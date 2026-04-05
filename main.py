# =========================
# main_android_optimized.py  —  Form Alchemist v1.6.0
# Android-optimized build entry point
#
# Patches applied over main.py:
#   P01  Bug fix: csv_col_spinner -> column_spinner
#   P02  Window.softinput_mode = "below_target" (soft keyboard keeps inputs visible)
#   P03  cv2 defensive import with CV2_AVAILABLE flag
#   P04  set_texture_from_bgr: explicit buffer release + cv2 fallback
#   P05  android_render_pdf_page: PIL fallback when cv2.imread unavailable
#   P06  PdfPageSource.render_page_bgr: cv2.cvtColor guard
#   P07  tempfile in android_render_pdf_page uses app-private dir (user_data_dir)
#   P08  get_app_output_dir always prefers user_data_dir (no storage permissions needed)
#   P09  set_status: Clock.schedule_once for thread-safe UI label updates
#   P10  Initial status chip shows plain-language 3-step hint
#   P11  PDF load errors shown as plain-language messages (not raw exceptions)
#   P12  Mobile make_button: min dp(44) tap target height (Android HIG)
#   P13  "?" help button added to appbar
#   P14  on_start, _request_android_permissions, first-run onboarding popup,
#        _show_help_popup, _show_user_error_popup injected into FormAlchemistApp
# =========================

# =========================
# PART 1 / 4
# Imports + Compatibility + Helpers + Engine Base
# =========================

import os
import io
import json
import math
import base64
import traceback
import re
import ssl
import urllib.request
import tempfile
import time
import sys
import hashlib
import uuid
import zipfile
from urllib.parse import urlparse, parse_qs
from types import SimpleNamespace

# -------------------------------------------------------------------
# Python 3.10+ compatibility patch for older reportlab builds
# Fixes:
# ImportError: cannot import name 'decodestring' from 'base64'
# -------------------------------------------------------------------
if not hasattr(base64, "decodestring"):
    base64.decodestring = base64.decodebytes
if not hasattr(base64, "encodestring"):
    base64.encodestring = base64.encodebytes

try:
    import cv2
    CV2_AVAILABLE = True
except Exception:
    cv2 = None
    CV2_AVAILABLE = False
from PIL import Image as PILImage
from PIL import ImageDraw, ImageFont
import importlib

FITZ_IMPORT_ERROR = None
try:
    fitz = importlib.import_module("fitz")
except Exception as _fitz_exc:
    fitz = None
    FITZ_IMPORT_ERROR = repr(_fitz_exc)

import numpy as np

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except Exception:
    PANDAS_AVAILABLE = False

    class _MiniColumns(list):
        def tolist(self):
            return list(self)

    class _MiniValueCounts(dict):
        def to_dict(self):
            return dict(self)

    class _MiniSeries:
        def __init__(self, data=None, index=None, dtype=None):
            self._data = list(data or [])
            self.index = list(index) if index is not None else list(range(len(self._data)))
            self.dtype = dtype

        def __iter__(self):
            return iter(self._data)

        def __len__(self):
            return len(self._data)

        def __getitem__(self, idx):
            return self._data[idx]

        def __eq__(self, other):
            return [str(v) == str(other) for v in self._data]

        def tolist(self):
            return list(self._data)

        def dropna(self):
            return _MiniSeries([v for v in self._data if not _mini_isna(v)])

        def astype(self, dtype):
            if dtype in (str, "str", "string"):
                return _MiniSeries(["" if _mini_isna(v) else str(v) for v in self._data], index=self.index, dtype=dtype)
            if dtype in (int, "int", "int64"):
                out = []
                for v in self._data:
                    if _mini_isna(v) or str(v).strip() == "":
                        out.append(0)
                    else:
                        out.append(int(float(v)))
                return _MiniSeries(out, index=self.index, dtype=dtype)
            return _MiniSeries(list(self._data), index=self.index, dtype=dtype)

        def unique(self):
            seen = []
            for v in self._data:
                if v not in seen:
                    seen.append(v)
            return seen

        def value_counts(self):
            counts = {}
            for v in self._data:
                counts[v] = counts.get(v, 0) + 1
            return _MiniValueCounts(counts)

    class _MiniRow(dict):
        pass

    class _MiniILoc:
        def __init__(self, df):
            self._df = df

        def __getitem__(self, idx):
            return _MiniRow(dict(self._df._rows[idx]))

    class _MiniDataFrame:
        def __init__(self, rows=None):
            self._rows = [dict(r) for r in (rows or [])]
            cols = []
            for row in self._rows:
                for key in row.keys():
                    if key not in cols:
                        cols.append(key)
            self._columns = _MiniColumns(cols)

        def __len__(self):
            return len(self._rows)

        @property
        def empty(self):
            return len(self._rows) == 0

        @property
        def columns(self):
            return self._columns

        @property
        def index(self):
            return list(range(len(self._rows)))

        @property
        def iloc(self):
            return _MiniILoc(self)

        def iterrows(self):
            for idx, row in enumerate(self._rows):
                yield idx, _MiniRow(dict(row))

        def fillna(self, value=""):
            for row in self._rows:
                for key, val in list(row.items()):
                    if _mini_isna(val):
                        row[key] = value
            return self

        def __getitem__(self, key):
            if isinstance(key, str):
                return _MiniSeries([row.get(key, "") for row in self._rows], index=self.index)
            if isinstance(key, (list, tuple)):
                if all(isinstance(v, bool) for v in key):
                    filtered = [row for row, keep in zip(self._rows, key) if keep]
                    return _MiniDataFrame(filtered)
                return _MiniDataFrame([{k: row.get(k, "") for k in key} for row in self._rows])
            raise KeyError(key)

        def __setitem__(self, key, value):
            if isinstance(value, _MiniSeries):
                values = value.tolist()
            elif isinstance(value, list):
                values = list(value)
            else:
                values = [value] * len(self._rows)
            if len(self._rows) == 0 and values:
                self._rows = [{} for _ in range(len(values))]
            if len(values) < len(self._rows):
                values.extend([""] * (len(self._rows) - len(values)))
            for idx, row in enumerate(self._rows):
                row[key] = values[idx] if idx < len(values) else ""
            if key not in self._columns:
                self._columns.append(key)

    def _mini_isna(value):
        if value is None:
            return True
        try:
            import math
            if isinstance(value, float) and math.isnan(value):
                return True
        except Exception:
            pass
        return False

    def _mini_read_csv(src, dtype=str):
        import csv
        import io as _io
        if hasattr(src, "read"):
            raw = src.read()
            if isinstance(raw, bytes):
                text = raw.decode("utf-8-sig", errors="replace")
            else:
                text = str(raw)
        else:
            with open(src, "r", encoding="utf-8-sig", newline="") as fh:
                text = fh.read()
        reader = csv.DictReader(_io.StringIO(text))
        rows = []
        for row in reader:
            rows.append({str(k): "" if _mini_isna(v) else str(v) for k, v in row.items()})
        return _MiniDataFrame(rows)

    def _mini_read_excel(src, dtype=str):
        from openpyxl import load_workbook
        wb = load_workbook(src, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            wb.close()
            return _MiniDataFrame([])
        headers = ["" if h is None else str(h) for h in headers]
        rows = []
        for row in rows_iter:
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                val = row[idx] if idx < len(row) else ""
                item[header] = "" if _mini_isna(val) else str(val)
            rows.append(item)
        wb.close()
        return _MiniDataFrame(rows)

    class _MiniPandasModule:
        DataFrame = _MiniDataFrame
        Series = _MiniSeries

        @staticmethod
        def read_csv(src, dtype=str):
            return _mini_read_csv(src, dtype=dtype)

        @staticmethod
        def read_excel(src, dtype=str):
            return _mini_read_excel(src, dtype=dtype)

        @staticmethod
        def isna(value):
            return _mini_isna(value)

    pd = _MiniPandasModule()

from pypdf import PdfReader, PdfWriter

try:
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.pdfbase.pdfmetrics import stringWidth as rl_string_width
    REPORTLAB_AVAILABLE = True

    # Android-safe patch: force ReportLab to use the pure-Python PDF escaping path.
    # Some compiled accelerator builds can crash at drawString()/canvas._escape with:
    # "PY_SSIZE_T_CLEAN macro must be defined for '#' formats".
    try:
        from reportlab.lib import rl_accel as _rl_accel
        _py_escape_pdf = getattr(_rl_accel, "_py_escapePDF", None)
        if callable(_py_escape_pdf):
            try:
                _rl_accel.escapePDF = _py_escape_pdf
            except Exception:
                pass
            try:
                rl_canvas.escapePDF = _py_escape_pdf
            except Exception:
                pass
    except Exception:
        pass
except Exception:
    rl_canvas = None
    REPORTLAB_AVAILABLE = False

    def rl_string_width(text, font_name, font_size):
        return len(str(text or "")) * float(font_size) * 0.6

class RectCompat:
    """Minimal rectangle compatibility layer used when PyMuPDF is unavailable."""
    __slots__ = ("x0", "y0", "x1", "y1")

    def __init__(self, x0=0.0, y0=0.0, x1=0.0, y1=0.0):
        rect_like = None
        if hasattr(x0, "x0") and hasattr(x0, "y0") and hasattr(x0, "x1") and hasattr(x0, "y1"):
            rect_like = x0
        elif y0 == 0.0 and x1 == 0.0 and y1 == 0.0 and not isinstance(x0, (int, float)):
            try:
                vals = list(x0)
                if len(vals) == 4:
                    rect_like = vals
            except Exception:
                rect_like = None

        if rect_like is not None:
            if hasattr(rect_like, "x0"):
                self.x0 = float(rect_like.x0)
                self.y0 = float(rect_like.y0)
                self.x1 = float(rect_like.x1)
                self.y1 = float(rect_like.y1)
            else:
                self.x0 = float(rect_like[0])
                self.y0 = float(rect_like[1])
                self.x1 = float(rect_like[2])
                self.y1 = float(rect_like[3])
            return

        self.x0 = float(x0)
        self.y0 = float(y0)
        self.x1 = float(x1)
        self.y1 = float(y1)

    @property
    def width(self):
        return self.x1 - self.x0

    @property
    def height(self):
        return self.y1 - self.y0

    def __iter__(self):
        yield self.x0
        yield self.y0
        yield self.x1
        yield self.y1

    def __repr__(self):
        return f"RectCompat({self.x0}, {self.y0}, {self.x1}, {self.y1})"

if fitz is None:
    fitz = SimpleNamespace(Rect=RectCompat)
elif not hasattr(fitz, "Rect"):
    fitz.Rect = RectCompat

from kivy.app import App
from kivy.animation import Animation
from kivy.clock import Clock
from kivy.core.text import Label as CoreLabel
from kivy.core.window import Window
from kivy.graphics import Color, Line, Rectangle, RoundedRectangle, Ellipse
from kivy.graphics.texture import Texture
from kivy.metrics import dp
from kivy.properties import BooleanProperty, ListProperty, ObjectProperty, StringProperty
from kivy.uix.anchorlayout import AnchorLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.checkbox import CheckBox
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.image import Image
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.slider import Slider
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput
from kivy.uix.widget import Widget
from kivy.utils import platform


RECORD_SELECT_TEXT = "Choose a patient / record"
COLUMN_SELECT_TEXT = "Choose which column fills this field"


class SearchableSelectField(Button):
    values = ListProperty([])
    placeholder = StringProperty("Select")
    picker_title = StringProperty("")
    search_hint = StringProperty("Type to filter")
    no_match_text = StringProperty("No matches found")
    allow_clear = BooleanProperty(True)
    _picker_popup = ObjectProperty(None, allownone=True)

    def __init__(self, **kwargs):
        placeholder = kwargs.pop("placeholder", kwargs.get("text", "Select"))
        picker_title = kwargs.pop("picker_title", "")
        search_hint = kwargs.pop("search_hint", "Type to filter")
        no_match_text = kwargs.pop("no_match_text", "No matches found")
        allow_clear = kwargs.pop("allow_clear", True)
        super().__init__(**kwargs)
        self.placeholder = str(placeholder or self.text or "Select")
        self.picker_title = str(picker_title or self.placeholder)
        self.search_hint = str(search_hint or "Type to filter")
        self.no_match_text = str(no_match_text or "No matches found")
        self.allow_clear = bool(allow_clear)
        if not str(self.text or "").strip():
            self.text = self.placeholder
        try:
            self.halign = "left"
            self.valign = "middle"
            self.shorten = True
            self.shorten_from = "right"
            self.padding = [dp(12), dp(10), dp(12), dp(10)]
            self.bind(size=self._sync_text_size)
        except Exception:
            pass
        self.bind(on_release=self._open_picker)

    def _sync_text_size(self, *_):
        try:
            self.text_size = (max(self.width - dp(24), 0), None)
        except Exception:
            pass

    def _normalized_values(self):
        out = []
        for value in list(self.values or []):
            value = str(value).strip()
            if value:
                out.append(value)
        return out

    def _open_picker(self, *_):
        app = App.get_running_app()
        palette = getattr(app, "theme_palette", {}) or {}
        is_mobile = bool(getattr(app, "ui_mobile", False))

        def _c(name, fallback):
            return palette.get(name, fallback)

        outer = BoxLayout(orientation="vertical", spacing=dp(8), padding=[dp(12), dp(12), dp(12), dp(12)])
        title = Label(
            text=self.picker_title or self.placeholder,
            color=_c("text", (0.93, 0.96, 1.0, 1)),
            size_hint_y=None,
            height=dp(24),
            halign="left",
            valign="middle",
            font_size=dp(15 if is_mobile else 16),
            bold=True,
        )
        title.bind(size=lambda inst, value: setattr(inst, "text_size", value))
        outer.add_widget(title)

        helper = Label(
            text="Type to filter the list. Press Enter to choose the top match.",
            color=_c("muted", (0.60, 0.68, 0.80, 1)),
            size_hint_y=None,
            height=dp(18),
            halign="left",
            valign="middle",
            font_size=dp(10.5),
        )
        helper.bind(size=lambda inst, value: setattr(inst, "text_size", value))
        outer.add_widget(helper)

        search_input = TextInput(
            text="",
            hint_text=self.search_hint,
            multiline=False,
            size_hint_y=None,
            height=dp(44),
            background_normal="",
            background_active="",
            background_color=_c("surface_alt", (0.11, 0.135, 0.185, 1)),
            foreground_color=_c("text", (0.93, 0.96, 1.0, 1)),
            cursor_color=_c("primary", (0.10, 0.78, 0.63, 1)),
            hint_text_color=_c("muted", (0.60, 0.68, 0.80, 1)),
            padding=[dp(12), dp(12), dp(12), dp(12)],
        )
        outer.add_widget(search_input)

        count_lbl = Label(
            text="",
            color=_c("muted", (0.60, 0.68, 0.80, 1)),
            size_hint_y=None,
            height=dp(18),
            halign="left",
            valign="middle",
            font_size=dp(10.5),
        )
        count_lbl.bind(size=lambda inst, value: setattr(inst, "text_size", value))
        outer.add_widget(count_lbl)

        quick_pick_btn = Button(
            text="Pick Top Match",
            size_hint_y=None,
            height=dp(44),
            background_normal="",
            background_color=_c("surface_alt", (0.11, 0.135, 0.185, 1)),
            color=_c("text", (0.93, 0.96, 1.0, 1)),
            disabled=True,
        )
        outer.add_widget(quick_pick_btn)

        results_scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False, bar_width=dp(6))
        results_grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        results_grid.bind(minimum_height=results_grid.setter("height"))
        results_scroll.add_widget(results_grid)
        outer.add_widget(results_scroll)

        button_row = GridLayout(cols=(2 if self.allow_clear else 1), size_hint_y=None, height=dp(44), spacing=dp(8))
        if self.allow_clear:
            clear_btn = Button(
                text="Clear",
                background_normal="",
                background_color=_c("surface_alt", (0.11, 0.135, 0.185, 1)),
                color=_c("text", (0.93, 0.96, 1.0, 1)),
            )
            button_row.add_widget(clear_btn)
        close_btn = Button(
            text="Close",
            background_normal="",
            background_color=_c("surface_alt", (0.11, 0.135, 0.185, 1)),
            color=_c("text", (0.93, 0.96, 1.0, 1)),
        )
        button_row.add_widget(close_btn)
        outer.add_widget(button_row)

        popup = Popup(
            title="",
            separator_height=0,
            background="",
            background_color=(0, 0, 0, 0.78),
            content=outer,
            size_hint=((0.94 if is_mobile else 0.68), (0.86 if is_mobile else 0.74)),
            auto_dismiss=True,
        )
        self._picker_popup = popup

        values = self._normalized_values()
        current_value = str(self.text or "").strip()
        max_rows = 250
        state = {"matches": list(values)}

        def _select_value(value):
            self.text = str(value)
            popup.dismiss()

        def _pick_top(*_):
            matches = list(state.get("matches") or [])
            if not matches:
                return
            _select_value(matches[0])

        def _refresh_results(*_):
            needle = str(search_input.text or "").strip().lower()
            matches = [value for value in values if needle in value.lower()] if needle else list(values)
            state["matches"] = matches
            results_grid.clear_widgets()
            shown = matches[:max_rows]
            quick_pick_btn.disabled = not bool(matches)
            quick_pick_btn.text = (f"Pick Top Match: {matches[0]}" if matches else "Pick Top Match")
            if not shown:
                empty_lbl = Label(
                    text=self.no_match_text,
                    color=_c("muted", (0.60, 0.68, 0.80, 1)),
                    size_hint_y=None,
                    height=dp(38),
                    halign="left",
                    valign="middle",
                    font_size=dp(11),
                )
                empty_lbl.bind(size=lambda inst, value: setattr(inst, "text_size", value))
                results_grid.add_widget(empty_lbl)
            else:
                for value in shown:
                    btn = Button(
                        text=value,
                        size_hint_y=None,
                        height=dp(44),
                        halign="left",
                        valign="middle",
                        shorten=True,
                        shorten_from="right",
                        background_normal="",
                        background_color=(_c("surface_alt", (0.11, 0.135, 0.185, 1)) if value != current_value else _c("primary", (0.10, 0.78, 0.63, 1))),
                        color=(_c("text", (0.93, 0.96, 1.0, 1)) if value != current_value else (1, 1, 1, 1)),
                    )
                    btn.bind(size=lambda inst, value: setattr(inst, "text_size", (max(inst.width - dp(18), 0), None)))
                    btn.bind(on_release=lambda inst, picked=value: _select_value(picked))
                    results_grid.add_widget(btn)
            if len(matches) > max_rows:
                count_lbl.text = f"Showing {max_rows} of {len(matches)} matches"
            else:
                count_lbl.text = f"{len(matches)} option{'s' if len(matches) != 1 else ''}"

        def _handle_text_validate(*_):
            needle = str(search_input.text or "").strip()
            if not needle:
                _pick_top()
                return
            matches = list(state.get("matches") or [])
            exact = next((value for value in matches if value.lower() == needle.lower()), None)
            if exact is not None:
                _select_value(exact)
            else:
                _pick_top()

        search_input.bind(text=_refresh_results)
        search_input.bind(on_text_validate=_handle_text_validate)
        quick_pick_btn.bind(on_release=_pick_top)
        if self.allow_clear:
            clear_btn.bind(on_release=lambda *_: (_select_value(self.placeholder)))
        close_btn.bind(on_release=lambda *_: popup.dismiss())
        _refresh_results()
        popup.open()
        Clock.schedule_once(lambda dt: setattr(search_input, "focus", True), 0.05)

def _utc_now_iso():
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def _safe_slug(value):
    value = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower())
    return value.strip("_") or "untitled"


def _sha1_json(value):
    raw = json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def get_form_alchemist_private_storage_root(app_name="form_alchemist"):
    try:
        running_app = App.get_running_app()
    except Exception:
        running_app = None

    user_data_dir = str(getattr(running_app, "user_data_dir", "") or "").strip()
    if user_data_dir:
        try:
            os.makedirs(user_data_dir, exist_ok=True)
            return user_data_dir
        except Exception:
            pass

    if platform == "android":
        try:
            from android.storage import app_storage_path
            base = str(app_storage_path() or "").strip()
            if base:
                os.makedirs(base, exist_ok=True)
                return base
        except Exception:
            pass

    fallback = os.path.join(os.path.expanduser("~"), f".{_safe_slug(app_name)}")
    try:
        os.makedirs(fallback, exist_ok=True)
        return fallback
    except Exception:
        final_fallback = os.path.join(tempfile.gettempdir(), _safe_slug(app_name))
        os.makedirs(final_fallback, exist_ok=True)
        return final_fallback


class LearningStorage:
    """App-private storage for learned profiles, revisions, and session state."""

    def __init__(self, base_dir=None, app_name="form_alchemist"):
        self.base_dir = str(base_dir or get_form_alchemist_private_storage_root(app_name))
        self.learning_dir = os.path.join(self.base_dir, "learning")
        self.revisions_dir = os.path.join(self.base_dir, "revisions")
        self.projects_dir = os.path.join(self.base_dir, "projects")
        self.cache_dir = os.path.join(self.base_dir, "cache")
        self._ensure_tree()

    def _ensure_tree(self):
        for folder in [self.base_dir, self.learning_dir, self.revisions_dir, self.projects_dir, self.cache_dir]:
            os.makedirs(folder, exist_ok=True)

    def _safe_load_json(self, path, default):
        try:
            if not os.path.exists(path):
                return default
            with open(path, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception:
            return default

    def _safe_write_json(self, path, payload):
        self._ensure_tree()
        parent = os.path.dirname(path)
        os.makedirs(parent, exist_ok=True)
        tmp_path = path + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, ensure_ascii=False)
        os.replace(tmp_path, path)

    def _rel(self, *parts):
        return os.path.join(self.base_dir, *parts)

    def load_profile_memory(self):
        return self._safe_load_json(self._rel("learning", "profile_memory.json"), {
            "version": 1,
            "profiles": {},
            "updated_at": None,
        })

    def save_profile_memory(self, payload):
        payload = dict(payload or {})
        payload["updated_at"] = _utc_now_iso()
        self._safe_write_json(self._rel("learning", "profile_memory.json"), payload)

    def load_template_stats(self):
        return self._safe_load_json(self._rel("learning", "template_stats.json"), {
            "version": 1,
            "families": {},
            "updated_at": None,
        })

    def save_template_stats(self, payload):
        payload = dict(payload or {})
        payload["updated_at"] = _utc_now_iso()
        self._safe_write_json(self._rel("learning", "template_stats.json"), payload)

    def load_correction_log(self):
        return self._safe_load_json(self._rel("learning", "correction_log.json"), {
            "version": 1,
            "events": [],
            "updated_at": None,
        })

    def save_correction_log(self, payload):
        payload = dict(payload or {})
        payload["updated_at"] = _utc_now_iso()
        self._safe_write_json(self._rel("learning", "correction_log.json"), payload)

    def save_revision(self, revision):
        revision = dict(revision or {})
        revision_id = str(revision.get("revision_id") or f"rev_{time.strftime('%Y%m%d_%H%M%S', time.gmtime())}_{uuid.uuid4().hex[:8]}")
        revision["revision_id"] = revision_id
        revision.setdefault("saved_at", _utc_now_iso())
        self._safe_write_json(self._rel("revisions", f"{revision_id}.json"), revision)

        index_path = self._rel("revisions", "index.json")
        index = self._safe_load_json(index_path, {"version": 1, "items": []})
        items = [item for item in index.get("items", []) if str(item.get("revision_id")) != revision_id]
        items.append({
            "revision_id": revision_id,
            "saved_at": revision.get("saved_at"),
            "reason": revision.get("reason"),
            "page_idx": revision.get("page_idx"),
            "pdf_path": revision.get("pdf_path"),
            "approved": bool(revision.get("approved", False)),
            "learning_weight": float(revision.get("learning_weight", 0.0) or 0.0),
        })
        index["items"] = sorted(items, key=lambda item: str(item.get("saved_at", "")))
        self._safe_write_json(index_path, index)
        return revision_id

    def load_revision(self, revision_id):
        revision_id = str(revision_id or "").strip()
        if not revision_id:
            return None
        return self._safe_load_json(self._rel("revisions", f"{revision_id}.json"), None)

    def load_revision_index(self):
        return self._safe_load_json(self._rel("revisions", "index.json"), {"version": 1, "items": []})

    def save_session(self, payload, name="current_session"):
        self._safe_write_json(self._rel("projects", f"{_safe_slug(name)}.json"), payload or {})

    def load_session(self, name="current_session"):
        return self._safe_load_json(self._rel("projects", f"{_safe_slug(name)}.json"), {})


try:
    from kivymd.app import MDApp
    from kivymd.uix.button import MDRaisedButton, MDFlatButton, MDRectangleFlatButton
    from kivymd.uix.textfield import MDTextField
    from kivymd.uix.label import MDLabel
    from kivymd.uix.selectioncontrol import MDSwitch
    KIVYMD_AVAILABLE = True
except Exception:
    MDApp = App
    MDRaisedButton = None
    MDFlatButton = None
    MDRectangleFlatButton = None
    MDTextField = None
    MDLabel = None
    MDSwitch = None
    KIVYMD_AVAILABLE = False


class NeoMobileButton(Button):
    """Rounded, layered button used to give the mobile UI a richer look."""

    def __init__(self, tone="primary", theme=None, **kwargs):
        self.tone = tone
        self.theme = theme or {}
        self.compact_mode = bool(kwargs.pop("compact_mode", False))
        self.corner_radius = kwargs.pop("corner_radius", dp(20))
        requested_font_size = kwargs.pop("font_size", dp(10.4))
        requested_padding = kwargs.pop("padding", [dp(7), dp(4), dp(7), dp(4)])
        super().__init__(font_size=requested_font_size, **kwargs)
        self.background_normal = ""
        self.background_down = ""
        self.background_color = (0, 0, 0, 0)
        self.border = (0, 0, 0, 0)
        self.color = (0.95, 0.97, 1.0, 1)
        self.base_font_size = float(requested_font_size)
        self.min_font_size = dp(7.2)
        self.font_size = self.base_font_size
        self.bold = True
        self.padding = list(requested_padding)
        self.halign = "center"
        self.valign = "middle"
        self.text_size = self.size
        self.shorten = False
        self.shorten_from = "right"
        self.max_lines = 2

        with self.canvas.before:
            self._shadow_color = Color(0, 0, 0, 0.32)
            self._shadow_rect = RoundedRectangle(radius=[self.corner_radius] * 4)
            self._outer_glow_color = Color(1, 1, 1, 0.06)
            self._outer_glow_rect = RoundedRectangle(radius=[self.corner_radius + dp(4)] * 4)
            self._base_color = Color(0.2, 0.32, 0.58, 1)
            self._base_rect = RoundedRectangle(radius=[self.corner_radius] * 4)
            self._nebula_color = Color(1, 1, 1, 0.08)
            self._nebula_rect = RoundedRectangle(radius=[max(dp(10), self.corner_radius - dp(2))] * 4)
            self._core_color = Color(1, 1, 1, 0.08)
            self._core_rect = RoundedRectangle(radius=[max(dp(8), self.corner_radius - dp(6))] * 4)
            self._shine_color = Color(1, 1, 1, 0.07)
            self._shine_rect = RoundedRectangle(radius=[max(dp(8), self.corner_radius - dp(4))] * 4)
            self._accent_color = Color(1, 1, 1, 0.18)
            self._accent_rect = RoundedRectangle(radius=[dp(10)] * 4)
            self._comet_color = Color(1, 1, 1, 0.14)
            self._comet_rect = RoundedRectangle(radius=[dp(8)] * 4)
            self._star1_color = Color(1, 1, 1, 0.18)
            self._star1 = Ellipse()
            self._star2_color = Color(1, 1, 1, 0.14)
            self._star2 = Ellipse()
        with self.canvas.after:
            self._border_color = Color(1, 1, 1, 0.12)
            self._border_line = Line(width=1.1)
            self._inner_border_color = Color(1, 1, 1, 0.04)
            self._inner_border_line = Line(width=1.0)

        self.bind(pos=self._update_graphics, size=self._update_graphics, state=self._update_graphics, disabled=self._update_graphics)
        self.bind(size=self._sync_text_size, text=self._sync_text_size)
        self._sync_text_size()
        self._update_graphics()

    @staticmethod
    def _mix(a, b, t):
        return tuple((a[i] * (1.0 - t)) + (b[i] * t) for i in range(len(a)))

    def _resolve_palette(self):
        theme = self.theme or {}
        return {
            "primary": theme.get("primary", (0.24, 0.48, 0.96, 1)),
            "soft": theme.get("primary_soft", (0.16, 0.24, 0.38, 1)),
            "accent": theme.get("accent", (0.10, 0.78, 0.63, 1)),
            "secondary": theme.get("secondary", (0.82, 0.52, 1.0, 1)),
            "aurora": theme.get("aurora", (0.28, 0.90, 1.0, 1)),
            "starlight": theme.get("starlight", (1.0, 0.94, 0.78, 1)),
            "plain": theme.get("surface_soft", (0.135, 0.16, 0.215, 1)),
            "ghost": theme.get("ghost", (0.055, 0.082, 0.150, 0.88)),
            "mini": theme.get("mini", (0.048, 0.072, 0.128, 0.96)),
            "danger": theme.get("danger", (0.87, 0.33, 0.41, 1)),
            "text": theme.get("text", (0.93, 0.96, 1.0, 1)),
            "muted": theme.get("muted", (0.60, 0.68, 0.80, 1)),
            "border": theme.get("border", (0.18, 0.24, 0.34, 1)),
        }

    def _sync_text_size(self, *_):
        try:
            pad_x = 0
            pad_y = 0
            if isinstance(getattr(self, "padding", None), (list, tuple)) and len(self.padding) >= 2:
                if len(self.padding) >= 4:
                    pad_x = float(self.padding[0]) + float(self.padding[2])
                    pad_y = float(self.padding[1]) + float(self.padding[3])
                else:
                    pad_x = float(self.padding[0]) * 2.0
                    pad_y = float(self.padding[1]) * 2.0
            avail_w = max(dp(1), self.width - pad_x)
            avail_h = max(dp(1), self.height - pad_y)
            self.text_size = (avail_w, avail_h)
            txt = str(getattr(self, "text", "") or "").replace("\n", " ").strip()
            if not txt:
                self.font_size = self.base_font_size
                return
            longest = max((len(chunk) for chunk in txt.split()), default=len(txt))
            weighted = max(longest, len(txt) * 0.62)
            if weighted <= 0:
                fitted = self.base_font_size
            else:
                fitted = min(self.base_font_size, (avail_w / max(1.0, weighted * 0.60)))
            if len(txt) >= 12:
                fitted = min(fitted, self.base_font_size * 0.96)
            if len(txt) >= 18:
                fitted = min(fitted, self.base_font_size * 0.88)
            if len(txt) >= 26:
                fitted = min(fitted, self.base_font_size * 0.80)
            if avail_h <= dp(28):
                fitted = min(fitted, self.base_font_size * 0.92)
            self.font_size = max(self.min_font_size, fitted)
        except Exception:
            self.text_size = self.size

    def _update_graphics(self, *_):
        pal = self._resolve_palette()
        base = pal.get(self.tone, pal["primary"])
        if self.tone == "accent":
            nebula = self._mix(base, pal["starlight"], 0.35)
            glow = self._mix(base, pal["aurora"], 0.20)
            comet = pal["starlight"]
        elif self.tone == "danger":
            nebula = self._mix(base, pal["secondary"], 0.22)
            glow = self._mix(base, pal["starlight"], 0.08)
            comet = self._mix(pal["starlight"], base, 0.35)
        elif self.tone == "ghost":
            nebula = self._mix(base, pal["aurora"], 0.08)
            glow = self._mix(base, pal["aurora"], 0.03)
            comet = self._mix(pal["starlight"], pal["aurora"], 0.18)
        elif self.tone == "mini":
            nebula = self._mix(base, pal["secondary"], 0.10)
            glow = self._mix(base, pal["aurora"], 0.02)
            comet = self._mix(pal["starlight"], base, 0.16)
        elif self.tone == "plain":
            nebula = self._mix(base, pal["secondary"], 0.18)
            glow = self._mix(base, pal["aurora"], 0.08)
            comet = self._mix(pal["starlight"], pal["aurora"], 0.40)
        elif self.tone == "soft":
            nebula = self._mix(base, pal["secondary"], 0.22)
            glow = self._mix(base, pal["aurora"], 0.11)
            comet = self._mix(pal["starlight"], pal["secondary"], 0.24)
        elif self.tone == "secondary":
            nebula = self._mix(base, pal["starlight"], 0.14)
            glow = self._mix(base, pal["aurora"], 0.08)
            comet = self._mix(pal["starlight"], pal["secondary"], 0.22)
        else:
            nebula = self._mix(base, pal["secondary"], 0.24)
            glow = self._mix(base, pal["aurora"], 0.18)
            comet = self._mix(pal["starlight"], pal["aurora"], 0.24)

        border = self._mix(glow, pal["border"], 0.34)
        if self.tone == "ghost":
            border = self._mix(border, pal["text"], 0.08)
        elif self.tone == "mini":
            border = self._mix(border, pal["text"], 0.04)
        gloss = (1, 1, 1, 0.08)
        accent = (1, 1, 1, 0.20)
        outer_glow_alpha = 0.035
        shadow_alpha = 0.22
        star_alpha_1 = 0.0
        star_alpha_2 = 0.0

        if self.compact_mode:
            gloss = (1, 1, 1, 0.032)
            accent = (1, 1, 1, 0.085)
            outer_glow_alpha *= 0.28
            shadow_alpha *= 0.50
            if self.tone == "ghost":
                gloss = (1, 1, 1, 0.018)
                accent = (1, 1, 1, 0.048)
                outer_glow_alpha *= 0.55
                shadow_alpha *= 0.62
            elif self.tone == "mini":
                gloss = (1, 1, 1, 0.014)
                accent = (1, 1, 1, 0.034)
                outer_glow_alpha *= 0.34
                shadow_alpha *= 0.48
            elif self.tone == "plain":
                gloss = (1, 1, 1, 0.024)
                accent = (1, 1, 1, 0.060)
            elif self.tone == "soft":
                gloss = (1, 1, 1, 0.028)
                accent = (1, 1, 1, 0.070)
            elif self.tone == "secondary":
                gloss = (1, 1, 1, 0.024)
                accent = (1, 1, 1, 0.064)
            elif self.tone == "accent":
                gloss = (1, 1, 1, 0.036)
                accent = (1, 1, 1, 0.095)
                outer_glow_alpha *= 1.16

        if self.disabled:
            base = self._mix(base, (0.12, 0.14, 0.19, 1), 0.60)
            nebula = self._mix(nebula, (0.10, 0.11, 0.15, 1), 0.58)
            glow = self._mix(glow, (0.12, 0.13, 0.16, 1), 0.60)
            border = self._mix(border, (0.20, 0.22, 0.27, 1), 0.48)
            gloss = (1, 1, 1, 0.02)
            accent = (1, 1, 1, 0.04)
            outer_glow_alpha = 0.02
            shadow_alpha = 0.08
            star_alpha_1 = 0.04
            star_alpha_2 = 0.03
            self.color = self._mix(pal["text"], pal["muted"], 0.58)
        elif self.state == "down":
            base = self._mix(base, (0.03, 0.05, 0.08, 1), 0.18)
            nebula = self._mix(nebula, (0.02, 0.03, 0.05, 1), 0.16)
            glow = self._mix(glow, pal["starlight"], 0.05)
            border = self._mix(border, pal["starlight"], 0.10)
            gloss = (1, 1, 1, 0.03)
            accent = (1, 1, 1, 0.10)
            outer_glow_alpha = 0.06
            shadow_alpha = 0.22
            self.color = (1, 1, 1, 1) if self.tone in ("primary", "accent", "danger", "secondary") else pal["text"]
        else:
            self.color = (1, 1, 1, 1) if self.tone in ("primary", "accent", "danger", "secondary") else pal["text"]

        x, y = self.pos
        w, h = self.size
        shadow_offset = dp(3 if self.state != "down" else 1)
        inset = dp(1.5)

        self._shadow_color.rgba = (0, 0, 0, shadow_alpha)
        self._shadow_rect.pos = (x, y - shadow_offset)
        self._shadow_rect.size = (w, h + dp(2))
        self._shadow_rect.radius = [self.corner_radius] * 4

        self._outer_glow_color.rgba = (*glow[:3], outer_glow_alpha if not self.disabled else 0.0)
        self._outer_glow_rect.pos = (x - dp(1.5), y - dp(1.5))
        self._outer_glow_rect.size = (w + dp(3), h + dp(3))
        self._outer_glow_rect.radius = [self.corner_radius + dp(4)] * 4

        self._base_color.rgba = base
        self._base_rect.pos = (x, y)
        self._base_rect.size = (w, h)
        self._base_rect.radius = [self.corner_radius] * 4

        nebula_alpha = 0.18 if not self.disabled else 0.07
        if self.compact_mode and not self.disabled:
            nebula_alpha *= 0.62
            if self.tone == "ghost":
                nebula_alpha *= 0.42
            elif self.tone == "mini":
                nebula_alpha *= 0.26
        self._nebula_color.rgba = (*nebula[:3], nebula_alpha)
        self._nebula_rect.pos = (x + inset, y + dp(1))
        self._nebula_rect.size = (max(0, w - inset * 2), max(0, h * 0.62))
        self._nebula_rect.radius = [max(dp(10), self.corner_radius - dp(2))] * 4

        shine_h = max(dp(12), h * 0.42)
        self._shine_color.rgba = gloss
        self._shine_rect.pos = (x + inset, y + h - shine_h - inset)
        self._shine_rect.size = (max(0, w - inset * 2), max(0, shine_h))
        self._shine_rect.radius = [max(dp(8), self.corner_radius - dp(4))] * 4

        core_alpha = 0.10 if not self.disabled else 0.03
        if self.compact_mode and not self.disabled:
            core_alpha *= 0.55
            if self.tone == "ghost":
                core_alpha *= 0.42
            elif self.tone == "mini":
                core_alpha *= 0.22
        self._core_color.rgba = (*comet[:3], core_alpha)
        self._core_rect.pos = (x + dp(10), y + dp(7))
        self._core_rect.size = (max(dp(18), w * 0.46), dp(5))
        self._core_rect.radius = [dp(8)] * 4

        accent_w = min(dp(28), w * 0.18)
        accent_h = dp(4)
        self._accent_color.rgba = accent
        self._accent_rect.pos = (x + dp(14), y + h - accent_h - dp(10))
        self._accent_rect.size = (accent_w, accent_h)
        self._accent_rect.radius = [dp(8)] * 4

        comet_w = min(dp(40), max(dp(20), w * 0.22))
        comet_alpha = 0.08 if not self.disabled else 0.02
        if self.compact_mode and not self.disabled:
            comet_alpha *= 0.45
            if self.tone == "ghost":
                comet_alpha *= 0.34
            elif self.tone == "mini":
                comet_alpha *= 0.18
        self._comet_color.rgba = (*comet[:3], comet_alpha)
        self._comet_rect.pos = (x + w - comet_w - dp(14), y + h - dp(16))
        self._comet_rect.size = (comet_w, dp(3.2))
        self._comet_rect.radius = [dp(8)] * 4

        star_size_1 = dp(0)
        star_size_2 = dp(0)
        self._star1_color.rgba = (*pal["starlight"][:3], star_alpha_1)
        self._star1.pos = (x + w - dp(18), y + h - dp(18))
        self._star1.size = (star_size_1, star_size_1)
        self._star2_color.rgba = (*pal["aurora"][:3], star_alpha_2)
        self._star2.pos = (x + w - dp(30), y + h - dp(13))
        self._star2.size = (star_size_2, star_size_2)

        self._border_color.rgba = border
        self._border_line.rounded_rectangle = [x, y, w, h, self.corner_radius]
        inner_border_alpha = 0.02 if not self.disabled else 0.0
        if self.compact_mode and self.tone in ("ghost", "mini"):
            inner_border_alpha = 0.012 if self.tone == "ghost" else 0.008
        self._inner_border_color.rgba = (1, 1, 1, inner_border_alpha)
        self._inner_border_line.rounded_rectangle = [x + dp(1), y + dp(1), max(0, w - dp(2)), max(0, h - dp(2)), max(dp(8), self.corner_radius - dp(2))]


FITZ_AVAILABLE = bool(fitz is not None and hasattr(fitz, "open") and hasattr(fitz, "Rect"))

if platform == "android":
    try:
        from jnius import autoclass
        from android import activity
        AndroidPythonActivity = autoclass("org.kivy.android.PythonActivity")
        AndroidIntent = autoclass("android.content.Intent")
        AndroidBitmap = autoclass("android.graphics.Bitmap")
        AndroidBitmapConfig = autoclass("android.graphics.Bitmap$Config")
        AndroidCompressFormat = autoclass("android.graphics.Bitmap$CompressFormat")
        AndroidPdfRenderer = autoclass("android.graphics.pdf.PdfRenderer")
        AndroidParcelFileDescriptor = autoclass("android.os.ParcelFileDescriptor")
        AndroidUri = autoclass("android.net.Uri")
        AndroidByteArrayOutputStream = autoclass("java.io.ByteArrayOutputStream")
        try:
            AndroidUriCopyHelper = autoclass("org.formalchemist.formalchemist.UriCopyHelper")
        except Exception:
            AndroidUriCopyHelper = None
        ANDROID_JAVA_AVAILABLE = True
    except Exception:
        AndroidPythonActivity = None
        AndroidIntent = None
        AndroidBitmap = None
        AndroidPdfRenderer = None
        AndroidParcelFileDescriptor = None
        AndroidUri = None
        AndroidByteArrayOutputStream = None
        AndroidBitmapConfig = None
        AndroidCompressFormat = None
        AndroidUriCopyHelper = None
        ANDROID_JAVA_AVAILABLE = False
    try:
        from androidssystemfilechooser import uri_to_stream, uri_to_filename, uri_to_extension
        ANDROID_SYSTEM_FILE_CHOOSER_AVAILABLE = True
    except Exception:
        uri_to_stream = None
        uri_to_filename = None
        uri_to_extension = None
        ANDROID_SYSTEM_FILE_CHOOSER_AVAILABLE = False
else:
    ANDROID_JAVA_AVAILABLE = False
    AndroidUriCopyHelper = None
    uri_to_stream = None
    uri_to_filename = None
    uri_to_extension = None
    ANDROID_SYSTEM_FILE_CHOOSER_AVAILABLE = False


def android_render_pdf_page(path, page_idx=0, preview_zoom=1.5):
    """Render a PDF page on Android using the native PdfRenderHelper Java bridge."""
    if platform != "android" or not ANDROID_JAVA_AVAILABLE:
        raise RuntimeError("Android PdfRenderer is unavailable.")

    tmp_path = None
    try:
        AndroidPdfRenderHelper = autoclass("org.formalchemist.formalchemist.PdfRenderHelper")
        # Use app-private dir for temp files to avoid storage permission issues
        _tmp_dir = None
        try:
            _running = App.get_running_app()
            _tmp_dir = str(getattr(_running, "user_data_dir", "") or "").strip() or None
        except Exception:
            pass
        fd, tmp_path = tempfile.mkstemp(
            prefix="form_alchemist_preview_",
            suffix=".png",
            dir=_tmp_dir,
        )
        os.close(fd)
        out_path = AndroidPdfRenderHelper.renderPageToPng(path, int(page_idx), float(preview_zoom), tmp_path)
        out_path = str(out_path or tmp_path)
        if cv2 is not None:
            img = cv2.imread(out_path, cv2.IMREAD_COLOR)
        else:
            from PIL import Image as _PILtmp
            import numpy as _nptmp
            _pil = _PILtmp.open(out_path).convert("RGB")
            img = _nptmp.array(_pil)[:, :, ::-1].copy()
            _pil.close()
        if img is None:
            raise ValueError("Android PdfRenderHelper produced an unreadable preview image.")
        return img
    finally:
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


# ============================================================
# Defaults
# ============================================================
APP_TITLE = "Form Alchemist"
CONFIG_FILENAME = "form_alchemist_config.json"
if platform == "android":
    ZOOM = 4.0
    PREVIEW_SCALE = 2.2
    # Keep the main content above the soft keyboard on Android
    Window.softinput_mode = "below_target"
else:
    ZOOM = 4.0
    PREVIEW_SCALE = 2.2

DEFAULTS = {
    "F_Area": 500,
    "F_MinW": 15,
    "F_MinH": 35,
    "F_Close": 1,
    "Line_MinW": 100,
    "Line_MaxW": 1680,
    "C_Strict": 40,
    "C_Size": (14, 65),
    "C_Border": 0.10,
    "C_Inner": 0.40,
    "ROI_Max": 200,
    "C_Open": 1,
    "C_Close": 0,
    "C_BandPct": 0.18,
    "C_AspectTol": 0.10,
    "Ext_Low": 0.04,
    "Ext_High": 0.60,
    "C_FillMin": 0.45,
    "C_Eps": 0.04,
    "Use_Extent": False,
    "Is_Grid": False,
    "Grid_N": 1,
}

TEXT_TUNING_DEFAULTS = {
    "uppercase": True,
    "overall_scale": 0.92,
    "box_scale": 1.00,
    "line_scale": 0.94,
    "grid_scale": 0.92,
    "content_inset_x": 0.07,
    "content_inset_y": 0.14,
    "box_vshift": 0.00,
    "line_vshift": 0.00,
    "grid_vshift": 0.00,
    "line_expand_h_mul": 2.10,
    "line_min_effective_h_px": 28,
}

CANDIDATE_TUNING_DEFAULTS = {
    "min_candidate_score": 0.34,
    "moderate_score": 0.58,
    "strong_score": 0.76,
    "auto_check_add_score": 0.58,
    "auto_check_remove_score": 0.76,
}


DETECTION_UI_META = {
    "f_area": {"label": "Minimum Box Area", "helper": "Ignore very tiny shapes. Raise this if the app finds too many tiny false boxes."},
    "f_minw": {"label": "Minimum Box Width", "helper": "Ignore narrow boxes that are too small to be useful."},
    "f_minh": {"label": "Minimum Box Height", "helper": "Ignore short boxes that are too small to be useful."},
    "f_close": {"label": "Box Cleanup Strength", "helper": "Smooth and join nearby box edges before the app decides what is a fill area."},
    "line_minw": {"label": "Minimum Line Length", "helper": "Ignore short marks and only keep longer answer lines."},
    "line_maxw": {"label": "Maximum Line Length", "helper": "Ignore extra-long lines that are probably table borders or page decorations."},
    "c_strict": {"label": "Checkbox Strictness", "helper": "Higher values make checkbox matching stricter. Lower values make it more forgiving."},
    "c_size_min": {"label": "Minimum Checkbox Size", "helper": "Ignore checkbox shapes smaller than this."},
    "c_size_max": {"label": "Maximum Checkbox Size", "helper": "Ignore checkbox shapes larger than this."},
    "c_border": {"label": "Border Visibility", "helper": "How much of the checkbox border should be visible for a good match."},
    "c_inner": {"label": "Inner Empty Space", "helper": "How empty the inside of the checkbox should look."},
    "roi_max": {"label": "Maximum Search Window", "helper": "The biggest area to inspect around a possible checkbox."},
    "c_open": {"label": "Noise Cleanup Before Check", "helper": "Removes tiny specks before checking if a shape is a checkbox."},
    "c_close": {"label": "Gap Closing After Check", "helper": "Closes small gaps so broken checkbox borders are easier to catch."},
    "c_band": {"label": "Edge Band Check", "helper": "Checks whether the top and bottom border bands look like a checkbox."},
    "c_aspect": {"label": "Square Shape Tolerance", "helper": "How close to a square the checkbox should be allowed to look."},
    "ext_low": {"label": "Minimum Shape Density", "helper": "Lower bound for how solid the checkbox-like shape should be."},
    "ext_high": {"label": "Maximum Shape Density", "helper": "Upper bound for how solid the checkbox-like shape should be."},
    "c_fill": {"label": "Minimum Filled Area", "helper": "Ignore shapes that do not fill enough of their own outline."},
    "c_eps": {"label": "Shape Simplify Amount", "helper": "How much the shape outline is simplified before checking corners."},
    "use_extent": {"label": "Use Extra Shape Filter", "helper": "Adds another shape test for checkbox matching. Helpful for noisy forms."},
}

DETECTION_GROUP_LABELS = {
    "Field": "Fill Areas / Boxes",
    "Lines": "Answer Lines",
    "Checkbox": "Checkboxes",
    "Extent": "Extra Shape Filters",
}

def detection_ui_label(key, fallback=None):
    meta = DETECTION_UI_META.get(str(key), {})
    return meta.get("label", fallback or str(key))

def detection_ui_helper(key, fallback=""):
    meta = DETECTION_UI_META.get(str(key), {})
    return meta.get("helper", fallback or "")


SEMANTIC_TARGET_KEYS = [
    "record_identity.first_name",
    "record_identity.middle_name",
    "record_identity.last_name",
    "record_identity.suffix",
    "record_identity.full_name",
    "record_identity.birth_month",
    "record_identity.birth_day",
    "record_identity.birth_year",
    "record_identity.membership_number",
]

NAME_TARGET_KEYS = [
    "record_identity.first_name",
    "record_identity.middle_name",
    "record_identity.last_name",
    "record_identity.suffix",
]

DOB_TARGET_KEYS = [
    "record_identity.birth_month",
    "record_identity.birth_day",
    "record_identity.birth_year",
]

LEGACY_GEOM_TO_SEMANTIC = {
    "names": NAME_TARGET_KEYS,
    "dob": DOB_TARGET_KEYS,
    "phil": ["record_identity.membership_number"],
}

SEMANTIC_TARGET_LABELS = {
    "record_identity.first_name": "Record First Name",
    "record_identity.middle_name": "Record Middle Name",
    "record_identity.last_name": "Record Last Name",
    "record_identity.suffix": "Record Suffix",
    "record_identity.full_name": "Record Full Name",
    "record_identity.birth_month": "Birth Month",
    "record_identity.birth_day": "Birth Day",
    "record_identity.birth_year": "Birth Year",
    "record_identity.membership_number": "Membership Number",
}

SEMANTIC_TARGET_PLACEHOLDER = "No special form target"
SEMANTIC_TARGETS_ENABLED = False

def semantic_target_label(key, fallback=None):
    key = str(key or "").strip()
    if not key:
        return fallback or ""
    return SEMANTIC_TARGET_LABELS.get(key, fallback or key)


# ============================================================
# Small helpers
# ============================================================
def safe_name(text):
    text = str(text or "").strip()
    text = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in text)
    return text.strip() or "Unknown"

def _norm_str(v):
    return str(v or "").strip()

def _rect_area(r):
    return max(0.0, r.width) * max(0.0, r.height)

def _rect_intersection_area(a, b):
    ix0 = max(a.x0, b.x0)
    iy0 = max(a.y0, b.y0)
    ix1 = min(a.x1, b.x1)
    iy1 = min(a.y1, b.y1)
    if ix1 <= ix0 or iy1 <= iy0:
        return 0.0
    return (ix1 - ix0) * (iy1 - iy0)

def _rect_iou(a, b):
    inter = _rect_intersection_area(a, b)
    if inter <= 0:
        return 0.0
    union = (a.width * a.height) + (b.width * b.height) - inter
    return inter / union if union > 0 else 0.0

def _x_intersection(a, b):
    return max(0.0, min(a.x1, b.x1) - max(a.x0, b.x0))

def _x_overlap_ratio(a, b):
    inter_x = _x_intersection(a, b)
    return inter_x / max(min(a.width, b.width), 1e-6)

def _rect_union(rects):
    return fitz.Rect(
        min(r.x0 for r in rects),
        min(r.y0 for r in rects),
        max(r.x1 for r in rects),
        max(r.y1 for r in rects)
    )

def _rect_close(a, b, tol=0.20):
    return (
        abs(a.x0 - b.x0) <= tol and
        abs(a.y0 - b.y0) <= tol and
        abs(a.x1 - b.x1) <= tol and
        abs(a.y1 - b.y1) <= tol
    )

def _rect_list_close(rects_a, rects_b, tol=0.20):
    if len(rects_a) != len(rects_b):
        return False
    aa = sorted(rects_a, key=lambda r: (round(r.y0, 3), r.x0, r.x1, r.y1))
    bb = sorted(rects_b, key=lambda r: (round(r.y0, 3), r.x0, r.x1, r.y1))
    return all(_rect_close(a, b, tol=tol) for a, b in zip(aa, bb))

def find_col_safe(df, *keywords):
    keys = [k.lower() for k in keywords]
    for c in df.columns:
        if all(k in str(c).lower() for k in keys):
            return c
    return None

def extract_gsheet_id(url):
    """
    Extract Google Sheet ID from common Google Sheets URLs.
    """
    if not url:
        return None

    url = str(url).strip()

    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if m:
        return m.group(1)

    parsed = urlparse(url)
    qs = parse_qs(parsed.query)
    if "id" in qs and qs["id"]:
        return qs["id"][0]

    return None


def extract_gsheet_gid(url):
    """
    Extract gid from Google Sheet URL. Defaults to 0 if missing.
    """
    if not url:
        return "0"

    parsed = urlparse(str(url).strip())
    qs = parse_qs(parsed.query)

    if "gid" in qs and qs["gid"]:
        return str(qs["gid"][0])

    if parsed.fragment:
        frag_qs = parse_qs(parsed.fragment)
        if "gid" in frag_qs and frag_qs["gid"]:
            return str(frag_qs["gid"][0])

        m = re.search(r"gid=([0-9]+)", parsed.fragment)
        if m:
            return m.group(1)

    return "0"


def gsheet_url_to_csv_export(url):
    """
    Convert a Google Sheet URL into a direct CSV export URL.
    Works for sheets that are accessible without OAuth
    (public/shared appropriately).
    """
    sheet_id = extract_gsheet_id(url)
    if not sheet_id:
        raise ValueError("Could not extract Google Sheet ID from URL.")

    gid = extract_gsheet_gid(url)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def _looks_like_ssl_cert_failure(exc):
    msg = str(exc or "").lower()
    return (
        "certificate verify failed" in msg or
        "unable to get local issuer certificate" in msg or
        "self signed certificate" in msg or
        "hostname" in msg and "doesn't match" in msg
    )


def download_url_bytes(url, timeout=30):
    """Download bytes from a URL with a cautious Android SSL fallback."""
    headers = {
        "User-Agent": "Mozilla/5.0 FormAlchemist/1.0",
        "Accept": "text/csv,text/plain,application/octet-stream,*/*",
    }

    verified_contexts = []
    try:
        import certifi
        verified_contexts.append(ssl.create_default_context(cafile=certifi.where()))
    except Exception:
        pass
    try:
        verified_contexts.append(ssl.create_default_context())
    except Exception:
        pass

    if not verified_contexts:
        verified_contexts = [None]

    last_error = None
    req = urllib.request.Request(url, headers=headers)

    for ctx in verified_contexts:
        try:
            if ctx is None:
                with urllib.request.urlopen(req, timeout=timeout) as resp:
                    return resp.read()
            with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
                return resp.read()
        except Exception as e:
            last_error = e

    if platform == "android" and last_error is not None and _looks_like_ssl_cert_failure(last_error):
        try:
            insecure_ctx = ssl._create_unverified_context()
            with urllib.request.urlopen(req, timeout=timeout, context=insecure_ctx) as resp:
                return resp.read()
        except Exception as e:
            last_error = e

    if last_error is not None:
        raise last_error
    raise RuntimeError("Failed to download URL data.")


def load_google_sheet_dataframe(url):
    csv_url = gsheet_url_to_csv_export(url)
    try:
        raw = download_url_bytes(csv_url, timeout=30)
    except Exception as e:
        if _looks_like_ssl_cert_failure(e):
            raise ValueError(
                "Android could not verify Google Sheets SSL certificates for this connection. "
                "Try again on a different network, or download/export the sheet as CSV/XLSX and load the file directly. "
                f"Details: {e}"
            )
        raise
    if not raw:
        raise ValueError("Google Sheet export returned an empty response.")

    sniff = raw[:512].lower()
    if b"<html" in sniff or b"<!doctype html" in sniff:
        raise ValueError(
            "Google Sheet export returned HTML instead of CSV. "
            "Make sure the sheet and selected tab are shared/accessible."
        )

    try:
        return pd.read_csv(io.BytesIO(raw), dtype=str).fillna("")
    except Exception:
        try:
            return pd.read_csv(io.StringIO(raw.decode("utf-8-sig")), dtype=str).fillna("")
        except Exception as e:
            raise ValueError(f"Downloaded Google Sheet data could not be parsed as CSV. Details: {e}")



class LocalImportStore:
    """Manage app-private imported files and generated outputs."""
    def __init__(self, base_dir_getter):
        self._base_dir_getter = base_dir_getter

    def get_output_dir(self):
        base_dir = self._base_dir_getter()
        os.makedirs(base_dir, exist_ok=True)
        return base_dir

    def get_import_dir(self):
        out_dir = os.path.join(self.get_output_dir(), "imports")
        os.makedirs(out_dir, exist_ok=True)
        return out_dir

    def build_local_path(self, display_name=None, default_name="selected_input", required_suffix=None):
        filename = safe_name(display_name or default_name)
        if required_suffix and not filename.lower().endswith(required_suffix.lower()):
            filename += required_suffix
        return os.path.join(self.get_import_dir(), filename)

    def save_bytes(self, payload, display_name=None, default_name="selected_input", required_suffix=None):
        if not payload:
            raise ValueError("Imported file is empty.")
        out_path = self.build_local_path(display_name=display_name, default_name=default_name, required_suffix=required_suffix)
        with open(out_path, "wb") as fh:
            fh.write(payload)
        if not os.path.exists(out_path) or os.path.getsize(out_path) <= 0:
            raise ValueError("Copied file is empty after saving to app storage.")
        return out_path


class AndroidDocumentPickerService:
    """Android Storage Access Framework helpers for import-style document picking."""
    def __init__(self, import_store, status_callback=None):
        self.import_store = import_store
        self.status_callback = status_callback

    def _status(self, message):
        if callable(self.status_callback):
            self.status_callback(message)


    def _coerce_binary_payload(self, payload):
        if payload is None:
            return b""
        if isinstance(payload, bytes):
            return payload
        if isinstance(payload, bytearray):
            return bytes(payload)
        if isinstance(payload, memoryview):
            return payload.tobytes()
        try:
            return bytes(payload)
        except Exception:
            pass
        try:
            return bytes(((int(b) + 256) % 256) for b in payload)
        except Exception:
            return b""

    def _looks_like_invalid_zero_payload(self, data):
        if not data:
            return True
        sample = data[: min(len(data), 64)]
        return bool(sample) and all(b == 0 for b in sample)

    def _read_uri_bytes_via_androidsystemfilechooser(self, uri, display_name=None):
        if not (ANDROID_SYSTEM_FILE_CHOOSER_AVAILABLE and callable(uri_to_stream)):
            return b"", display_name
        try:
            if not display_name and callable(uri_to_filename):
                display_name = uri_to_filename(uri)
        except Exception:
            pass

        try:
            stream = uri_to_stream(uri)
        except Exception:
            return b"", display_name

        try:
            payload = stream.read()
            data = self._coerce_binary_payload(payload)
            return data, display_name
        finally:
            try:
                stream.close()
            except Exception:
                pass

    def _read_uri_bytes_via_pfd(self, resolver, uri):
        pfd = None
        fd = None
        try:
            pfd = resolver.openFileDescriptor(uri, "r")
            if pfd is None:
                return b""
            fd = pfd.detachFd()
            chunks = []
            while True:
                chunk = os.read(fd, 65536)
                if not chunk:
                    break
                chunks.append(chunk)
            return b"".join(chunks)
        except Exception:
            return b""
        finally:
            try:
                if fd is not None:
                    os.close(fd)
            except Exception:
                pass
            try:
                if pfd is not None:
                    pfd.close()
            except Exception:
                pass

    def _read_uri_bytes_via_inputstream(self, resolver, uri):
        input_stream = None
        baos = None
        try:
            input_stream = resolver.openInputStream(uri)
            if input_stream is None:
                return b""
            baos = AndroidByteArrayOutputStream()
            ByteArray = autoclass("java.lang.reflect.Array")
            JavaByte = autoclass("java.lang.Byte").TYPE
            buffer = ByteArray.newInstance(JavaByte, 65536)
            while True:
                count = input_stream.read(buffer)
                if count == -1:
                    break
                if count > 0:
                    baos.write(buffer, 0, count)
            return self._coerce_binary_payload(baos.toByteArray())
        except Exception:
            return b""
        finally:
            try:
                if input_stream is not None:
                    input_stream.close()
            except Exception:
                pass
            try:
                if baos is not None:
                    baos.close()
            except Exception:
                pass


    def _resolve_display_name_native(self, context_obj, uri_string, default_name):
        if platform != "android" or AndroidUriCopyHelper is None:
            return None
        try:
            name = AndroidUriCopyHelper.resolveDisplayName(context_obj, uri_string, default_name or "")
            if name:
                return str(name)
        except Exception:
            pass
        return None

    def _copy_uri_to_local_file_native(self, context_obj, uri_string, out_path):
        if platform != "android" or AndroidUriCopyHelper is None:
            return False
        try:
            ok = AndroidUriCopyHelper.copyUriToPath(context_obj, uri_string, out_path)
            return bool(ok)
        except Exception:
            return False

    def write_bytes_to_uri(self, uri, payload, mode="w"):
        if platform != "android" or not ANDROID_JAVA_AVAILABLE:
            raise RuntimeError("Android document saver is unavailable.")
        if uri is None:
            raise ValueError("Android returned no destination URI.")

        data = self._coerce_binary_payload(payload)
        if isinstance(uri, str):
            uri_string = uri.strip()
        else:
            try:
                uri_string = str(uri.toString()).strip()
            except Exception:
                uri_string = ""

        if not uri_string:
            raise ValueError("Android returned an empty destination URI.")

        uri_obj = AndroidUri.parse(uri_string)
        activity_obj = AndroidPythonActivity.mActivity
        resolver = activity_obj.getContentResolver()

        pfd = None
        fd = None
        try:
            try:
                open_mode = "rwt" if "w" in str(mode or "w") and "t" not in str(mode or "w") else (mode or "w")
                pfd = resolver.openFileDescriptor(uri_obj, open_mode)
            except Exception:
                pfd = resolver.openFileDescriptor(uri_obj, mode or "w")
            if pfd is None:
                raise IOError("Android could not open the save destination.")

            fd = pfd.detachFd()
            try:
                os.ftruncate(fd, 0)
            except Exception:
                pass

            total = 0
            chunk_size = 65536
            while total < len(data):
                chunk = data[total:total + chunk_size]
                written = os.write(fd, chunk)
                if written is None:
                    written = 0
                if written <= 0:
                    raise IOError("Android write returned 0 bytes.")
                total += written

            try:
                os.fsync(fd)
            except Exception:
                pass
            return total
        finally:
            try:
                if fd is not None:
                    os.close(fd)
            except Exception:
                pass
            try:
                if pfd is not None:
                    pfd.close()
            except Exception:
                pass

    def copy_uri_to_local_file(self, uri, default_name="selected_input", required_suffix=None, allowed_suffixes=None):
        if platform != "android" or not ANDROID_JAVA_AVAILABLE:
            raise RuntimeError("Android document picker is unavailable.")
        if uri is None:
            raise ValueError("Android returned no document URI.")

        activity_obj = AndroidPythonActivity.mActivity
        resolver = activity_obj.getContentResolver()

        if isinstance(uri, str):
            uri_string = uri.strip()
        else:
            try:
                uri_string = str(uri.toString()).strip()
            except Exception:
                uri_string = ""

        if not uri_string:
            raise ValueError("Android returned an empty document URI.")
        if not (uri_string.startswith("content://") or uri_string.startswith("file://")):
            raise ValueError(f"Android returned an invalid document URI: {uri_string}")

        uri = AndroidUri.parse(uri_string)

        display_name = self._resolve_display_name_native(activity_obj, uri_string, default_name)

        if not display_name:
            cursor = None
            try:
                OpenableColumns = autoclass("android.provider.OpenableColumns")
                cursor = resolver.query(uri, None, None, None, None)
                if cursor is not None and cursor.moveToFirst():
                    idx = cursor.getColumnIndex(OpenableColumns.DISPLAY_NAME)
                    if idx >= 0:
                        display_name = cursor.getString(idx)
            except Exception:
                display_name = None
            finally:
                try:
                    if cursor is not None:
                        cursor.close()
                except Exception:
                    pass

        if (not display_name) and ANDROID_SYSTEM_FILE_CHOOSER_AVAILABLE and callable(uri_to_extension):
            try:
                inferred_ext = uri_to_extension(uri)
                if inferred_ext:
                    display_name = f"{safe_name(default_name)}.{str(inferred_ext).lstrip('.')}"
            except Exception:
                pass

        lower_name = str(display_name or default_name or "").lower()
        if allowed_suffixes:
            normalized = [str(s).lower() for s in allowed_suffixes if s]
            if normalized and not any(lower_name.endswith(s) for s in normalized):
                # SAF providers sometimes omit or rewrite the visible filename extension.
                # Do not reject the document purely by its display name; instead, keep the
                # required suffix on the local cached copy and let the downstream parser
                # validate the actual file contents.
                pass

        target_path = self.import_store.build_local_path(
            display_name=display_name,
            default_name=default_name,
            required_suffix=required_suffix,
        )

        native_ok = self._copy_uri_to_local_file_native(activity_obj, uri_string, target_path)
        read_backend = "nativehelper" if native_ok else "none"

        file_bytes = b""
        if native_ok and os.path.exists(target_path):
            try:
                with open(target_path, "rb") as fh:
                    file_bytes = fh.read()
            except Exception:
                file_bytes = b""

        if (not file_bytes) or self._looks_like_invalid_zero_payload(file_bytes):
            if os.path.exists(target_path):
                try:
                    os.remove(target_path)
                except Exception:
                    pass

            chooser_bytes, display_name = self._read_uri_bytes_via_androidsystemfilechooser(uri, display_name=display_name)
            if chooser_bytes and not self._looks_like_invalid_zero_payload(chooser_bytes):
                file_bytes = chooser_bytes
                read_backend = "androidssystemfilechooser"

            if (not file_bytes) or self._looks_like_invalid_zero_payload(file_bytes):
                fd_bytes = self._read_uri_bytes_via_pfd(resolver, uri)
                if fd_bytes and not self._looks_like_invalid_zero_payload(fd_bytes):
                    file_bytes = fd_bytes
                    read_backend = "pfd"

            if (not file_bytes) or self._looks_like_invalid_zero_payload(file_bytes):
                stream_bytes = self._read_uri_bytes_via_inputstream(resolver, uri)
                if stream_bytes and not self._looks_like_invalid_zero_payload(stream_bytes):
                    file_bytes = stream_bytes
                    read_backend = "inputstream"

            if (not file_bytes) or self._looks_like_invalid_zero_payload(file_bytes):
                sample = repr((file_bytes or b"")[:32])
                raise ValueError(
                    "Android returned an unreadable file payload. "
                    f"Backend={read_backend}; first bytes={sample}"
                )

            target_path = self.import_store.save_bytes(
                file_bytes,
                display_name=display_name,
                default_name=default_name,
                required_suffix=required_suffix,
            )
            try:
                with open(target_path, "rb") as fh:
                    file_bytes = fh.read()
            except Exception:
                file_bytes = b""

        if required_suffix and required_suffix.lower() == ".pdf":
            pdf_pos = file_bytes.find(b"%PDF-")
            if pdf_pos == -1 or pdf_pos > 1024:
                head = repr(file_bytes[:32])
                raise ValueError(
                    "Selected file was copied, but it does not contain a valid PDF header. "
                    f"Backend={read_backend}; first bytes: {head}"
                )
            if pdf_pos > 0:
                file_bytes = file_bytes[pdf_pos:]
                with open(target_path, "wb") as fh:
                    fh.write(file_bytes)

        if not os.path.exists(target_path) or os.path.getsize(target_path) <= 0:
            raise ValueError(f"Imported file is missing or empty after copy. Backend={read_backend}")

        return target_path


    def open_document_picker(self, request_code, mime_type="*/*", title="document", on_picked=None, cancel_message=None, required_suffix=None, allowed_suffixes=None, extra_mime_types=None):
        if platform != "android" or not ANDROID_JAVA_AVAILABLE:
            self._status("Android system document picker is unavailable.")
            return

        def _on_activity_result(request_code_result, result_code, intent):
            if request_code_result != request_code:
                return
            activity.unbind(on_activity_result=_on_activity_result)

            if result_code != -1 or intent is None:
                Clock.schedule_once(lambda dt, m=(cancel_message or f"{title.capitalize()} selection cancelled."): self._status(m), 0)
                return

            try:
                uri = intent.getData()
                if uri is None:
                    raise ValueError(f"No {title} URI was returned by Android.")

                try:
                    flags = intent.getFlags()
                    take_flags = flags & (AndroidIntent.FLAG_GRANT_READ_URI_PERMISSION | AndroidIntent.FLAG_GRANT_WRITE_URI_PERMISSION)
                    AndroidPythonActivity.mActivity.getContentResolver().takePersistableUriPermission(uri, take_flags)
                except Exception:
                    pass

                Clock.schedule_once(lambda dt, t=title: self._status(f"Importing {t}..."), 0)
                local_path = self.copy_uri_to_local_file(
                    uri,
                    default_name=safe_name(title),
                    required_suffix=required_suffix,
                    allowed_suffixes=allowed_suffixes,
                )

                if callable(on_picked):
                    def _safe_dispatch(dt, p=local_path):
                        try:
                            on_picked(p)
                        except Exception as e:
                            traceback.print_exc()
                            try:
                                self._status(f"{title.capitalize()} error: {e}")
                            except Exception:
                                pass
                    Clock.schedule_once(_safe_dispatch, 0)
            except Exception as e:
                traceback.print_exc()
                Clock.schedule_once(lambda dt, m=f"{title.capitalize()} error: {e}": self._status(m), 0)

        activity.bind(on_activity_result=_on_activity_result)
        intent = AndroidIntent(AndroidIntent.ACTION_OPEN_DOCUMENT)
        intent.addCategory(AndroidIntent.CATEGORY_OPENABLE)
        intent.addFlags(AndroidIntent.FLAG_GRANT_READ_URI_PERMISSION)
        intent.addFlags(AndroidIntent.FLAG_GRANT_PERSISTABLE_URI_PERMISSION)

        effective_mime = mime_type or "*/*"
        if extra_mime_types:
            try:
                mime_candidates = [str(m).strip() for m in extra_mime_types if str(m).strip()]
                if mime_candidates:
                    effective_mime = "*/*"
                    try:
                        intent.putExtra(AndroidIntent.EXTRA_MIME_TYPES, mime_candidates)
                    except Exception:
                        try:
                            intent.putExtra("android.intent.extra.MIME_TYPES", mime_candidates)
                        except Exception:
                            pass
            except Exception:
                pass

        intent.setType(effective_mime)
        AndroidPythonActivity.mActivity.startActivityForResult(intent, request_code)


class PdfPageSource:
    """Read page counts and rasterized page images using the best backend for the platform."""
    def page_count(self, path):
        if not path or not os.path.exists(path):
            raise FileNotFoundError(f"PDF file not found: {path}")

        try:
            with open(path, "rb") as fh:
                reader = PdfReader(fh)
                total = len(reader.pages)
            if total > 0:
                return total
        except Exception:
            pass

        if platform == "android" and ANDROID_JAVA_AVAILABLE:
            pfd = None
            renderer = None
            try:
                file_obj = autoclass("java.io.File")(path)
                pfd = AndroidParcelFileDescriptor.open(file_obj, AndroidParcelFileDescriptor.MODE_READ_ONLY)
                renderer = AndroidPdfRenderer(pfd)
                return renderer.getPageCount()
            finally:
                try:
                    if renderer is not None:
                        renderer.close()
                except Exception:
                    pass
                try:
                    if pfd is not None:
                        pfd.close()
                except Exception:
                    pass

        if FITZ_AVAILABLE:
            with fitz.open(path) as doc:
                return len(doc)

        raise RuntimeError("No PDF page-count backend is available.")

    def render_page_bgr(self, path, page_idx=0, preview_zoom=1.5):
        if not path or not os.path.exists(path):
            raise FileNotFoundError(f"PDF file not found: {path}")

        if platform == "android" and ANDROID_JAVA_AVAILABLE:
            return android_render_pdf_page(path, page_idx=page_idx, preview_zoom=preview_zoom)

        if FITZ_AVAILABLE:
            with fitz.open(path) as doc:
                total = len(doc)
                if total <= 0:
                    raise ValueError("PDF has no pages.")
                page_idx = max(0, min(int(page_idx), total - 1))
                page = doc.load_page(page_idx)
                mat = fitz.Matrix(float(preview_zoom), float(preview_zoom))
                pix = page.get_pixmap(matrix=mat, alpha=False)
                img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
                if cv2 is not None:
                    if pix.n == 4:
                        return cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)
                    return cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
                else:
                    # cv2 unavailable: convert without it
                    if pix.n == 4:
                        return img[:, :, :3][:, :, ::-1].copy()
                    return img[:, :, ::-1].copy()

        raise RuntimeError("No PDF rendering backend is available.")



# ============================================================
# Detection / mapping engine
# ============================================================
# ============================================================
# Detection / mapping engine
# ============================================================
class FormAlchemistEngine:
    def __init__(self):
        self.pdf_path = ""
        self.df = pd.DataFrame()
        self.patient_names = []

        self.first_col = None
        self.last_col = None
        self.mid_col = None
        self.suf_col = None
        self.dob_col = None
        self.phil_col = None
        self.record_label_source_column = ""

        self.geom = {"names": [], "dob": [], "phil": []}
        self.all_boxes = []
        self.box_types = []
        self.custom_mappings = {}
        self.selected_box_ids = []

        self.settings = dict(DEFAULTS)
        self.settings["C_Size"] = list(DEFAULTS["C_Size"])
        self.text_tuning = dict(TEXT_TUNING_DEFAULTS)
        self.candidate_tuning = dict(CANDIDATE_TUNING_DEFAULTS)
        self.pdf_source = PdfPageSource()
        self.detected_page_idx = None
        self.detected_settings_signature = None
        self.boxes_by_page = {}
        self.box_types_by_page = {}
        self.geom_by_page = {}
        self.boxes_settings_signature_by_page = {}
        self.candidate_boxes_by_page = {}
        self.candidate_types_by_page = {}
        self.candidate_scores_by_page = {}
        self.candidate_sources_by_page = {}
        self.candidate_review_state_by_page = {}
        self.candidate_meta_by_page = {}
        self.candidate_payloads_by_scope = {}
        self.area_templates_by_page = {}
        self.area_template_match_cache = {}
        self.area_template_match_review_state = {}
        self.area_template_index = {}
        self.manual_area_search_cache = {}
        self.repair_patches_by_page = {}
        self.repair_patch_index = {}
        self.sticky_detections_by_page = {}

        self.learning_storage = LearningStorage(app_name="form_alchemist")
        self.learning_enabled = True
        self.profile_memory = self.learning_storage.load_profile_memory()
        self.template_stats = self.learning_storage.load_template_stats()
        self.correction_log = self.learning_storage.load_correction_log()
        self.current_detection_context = {}
        self.current_revision_id = ""
        self.last_saved_revision_id = ""
        self.last_loaded_learning_meta = {}
        self.last_applied_profile_meta = {}
        self.semantic_targets = self._empty_semantic_targets()
        self.semantic_targets_by_page = {}
        self.semantic_target_overrides = {}

    def _normalized_text_tuning(self, tuning=None):
        base = dict(TEXT_TUNING_DEFAULTS)
        if isinstance(tuning, dict):
            for key, value in tuning.items():
                if key in base:
                    base[key] = value

        def _to_bool(value, default):
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)):
                return bool(value)
            txt = str(value or "").strip().lower()
            if txt in ("1", "true", "yes", "on"): return True
            if txt in ("0", "false", "no", "off"): return False
            return bool(default)

        def _to_float(value, default, lo=None, hi=None):
            try:
                out = float(value)
            except Exception:
                out = float(default)
            if lo is not None:
                out = max(float(lo), out)
            if hi is not None:
                out = min(float(hi), out)
            return float(out)

        def _to_int(value, default, lo=None, hi=None):
            try:
                out = int(round(float(value)))
            except Exception:
                out = int(default)
            if lo is not None:
                out = max(int(lo), out)
            if hi is not None:
                out = min(int(hi), out)
            return int(out)

        base["uppercase"] = _to_bool(base.get("uppercase", True), True)
        base["overall_scale"] = _to_float(base.get("overall_scale", 0.92), 0.92, 0.60, 1.50)
        base["box_scale"] = _to_float(base.get("box_scale", 1.00), 1.00, 0.60, 1.60)
        base["line_scale"] = _to_float(base.get("line_scale", 0.94), 0.94, 0.60, 1.60)
        base["grid_scale"] = _to_float(base.get("grid_scale", 0.92), 0.92, 0.60, 1.40)
        base["content_inset_x"] = _to_float(base.get("content_inset_x", 0.07), 0.07, 0.00, 0.22)
        base["content_inset_y"] = _to_float(base.get("content_inset_y", 0.14), 0.14, 0.00, 0.28)
        base["box_vshift"] = _to_float(base.get("box_vshift", -0.08), -0.08, -0.30, 0.20)
        base["line_vshift"] = _to_float(base.get("line_vshift", -0.08), -0.08, -0.30, 0.20)
        base["grid_vshift"] = _to_float(base.get("grid_vshift", -0.05), -0.05, -0.30, 0.20)
        base["line_expand_h_mul"] = _to_float(base.get("line_expand_h_mul", 2.10), 2.10, 1.00, 3.50)
        base["line_min_effective_h_px"] = _to_int(base.get("line_min_effective_h_px", 28), 28, 0, 80)
        return base

    def _normalized_candidate_tuning(self, tuning=None):
        base = dict(CANDIDATE_TUNING_DEFAULTS)
        if isinstance(tuning, dict):
            for key, value in tuning.items():
                if key in base:
                    base[key] = value

        def _to_float(value, default, lo=None, hi=None):
            try:
                out = float(value)
            except Exception:
                out = float(default)
            if lo is not None:
                out = max(float(lo), out)
            if hi is not None:
                out = min(float(hi), out)
            return float(out)

        base["min_candidate_score"] = _to_float(base.get("min_candidate_score", 0.34), 0.34, 0.0, 0.95)
        base["moderate_score"] = _to_float(base.get("moderate_score", 0.58), 0.58, 0.0, 0.98)
        base["strong_score"] = _to_float(base.get("strong_score", 0.76), 0.76, 0.0, 0.99)
        if base["strong_score"] < base["moderate_score"]:
            base["strong_score"] = base["moderate_score"]
        base["auto_check_add_score"] = _to_float(base.get("auto_check_add_score", 0.58), 0.58, 0.0, 0.99)
        base["auto_check_remove_score"] = _to_float(base.get("auto_check_remove_score", 0.76), 0.76, 0.0, 0.99)
        return base

    def get_candidate_tuning(self):
        tuning = getattr(self, "candidate_tuning", None)
        normalized = self._normalized_candidate_tuning(tuning)
        self.candidate_tuning = dict(normalized)
        return normalized

    def _serialize_area_templates(self):
        payload = {}
        for page_key, items in (getattr(self, 'area_templates_by_page', {}) or {}).items():
            serial = []
            for item in (items or []):
                if not isinstance(item, dict):
                    continue
                serial.append({
                    'template_id': str(item.get('template_id', '') or ''),
                    'page': int(item.get('page', page_key) or page_key),
                    'source_rect': self._serialize_rect(item.get('source_rect')) if item.get('source_rect') is not None else None,
                    'resolved_rects': self._serialize_rects(item.get('resolved_rects', [])),
                    'resolved_types': list(item.get('resolved_types', []) or []),
                    'needs_split': bool(item.get('needs_split', False)),
                    'template_kind': str(item.get('template_kind', '') or ''),
                    'split_count': int(item.get('split_count', 0) or 0),
                    'source_text_overlap': float(item.get('source_text_overlap', 0.0) or 0.0),
                    'features': dict(item.get('features', {}) or {}),
                    'shape_template': dict(item.get('shape_template', {}) or {}),
                    'created_at': str(item.get('created_at', '') or ''),
                })
            if serial:
                payload[str(int(page_key))] = serial
        return payload

    def _deserialize_area_templates(self, payload):
        self.area_templates_by_page = {}
        self.area_template_index = {}
        self.repair_patches_by_page = {}
        self.repair_patch_index = {}
        self.area_template_match_cache = {}
        self.area_template_match_review_state = {}
        self.manual_area_search_cache = {}
        if not isinstance(payload, dict):
            return
        for page_key, items in payload.items():
            try:
                page_idx = int(page_key)
            except Exception:
                continue
            page_items = []
            for item in (items or []):
                if not isinstance(item, dict):
                    continue
                source_rect = item.get('source_rect')
                try:
                    source_rect = fitz.Rect(source_rect) if source_rect is not None else None
                except Exception:
                    source_rect = None
                resolved_rects = []
                for rect in (item.get('resolved_rects', []) or []):
                    try:
                        resolved_rects.append(fitz.Rect(rect))
                    except Exception:
                        pass
                template = {
                    'template_id': str(item.get('template_id', '') or uuid.uuid4().hex[:12]),
                    'page': page_idx,
                    'source_rect': source_rect,
                    'resolved_rects': resolved_rects,
                    'resolved_types': [str(v) for v in (item.get('resolved_types', []) or [])],
                    'needs_split': bool(item.get('needs_split', False)),
                    'template_kind': str(item.get('template_kind', '') or ''),
                    'split_count': int(item.get('split_count', len(resolved_rects)) or len(resolved_rects)),
                    'source_text_overlap': float(item.get('source_text_overlap', 0.0) or 0.0),
                    'features': dict(item.get('features', {}) or {}),
                    'shape_template': dict(item.get('shape_template', {}) or {}),
                    'created_at': str(item.get('created_at', '') or _utc_now_iso()),
                }
                page_items.append(template)
                self.area_template_index[template['template_id']] = template
            if page_items:
                self.area_templates_by_page[page_idx] = page_items

    def analyze_manual_area_template(self, page_idx, source_rect):
        page_idx = int(page_idx or 0)
        source_rect = source_rect if isinstance(source_rect, fitz.Rect) else fitz.Rect(source_rect)
        if source_rect.width <= 2 or source_rect.height <= 2:
            return None
        img = self._render_pdf_page_bgr(self.pdf_path, page_idx=page_idx, preview_zoom=ZOOM)
        if img is None or getattr(img, 'size', 0) == 0:
            return None
        h_img, w_img = img.shape[:2]
        x0 = max(0, min(int(math.floor(source_rect.x0 * ZOOM)), w_img - 1))
        y0 = max(0, min(int(math.floor(source_rect.y0 * ZOOM)), h_img - 1))
        x1 = max(x0 + 1, min(int(math.ceil(source_rect.x1 * ZOOM)), w_img))
        y1 = max(y0 + 1, min(int(math.ceil(source_rect.y1 * ZOOM)), h_img))
        crop = img[y0:y1, x0:x1].copy()
        if crop.size == 0:
            return None
        text_mask, _ = self.build_text_like_mask(crop, header_band_pct=0.10)
        try:
            text_reject_mask = cv2.dilate(text_mask, np.ones((3, 3), np.uint8), iterations=1)
        except Exception:
            text_reject_mask = text_mask
        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
        base_bin_raw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        try:
            base_bin = cv2.bitwise_and(base_bin_raw, cv2.bitwise_not(text_reject_mask))
        except Exception:
            base_bin = base_bin_raw
        field_close_k = max(0, int(self.settings.get('F_Close', 1) or 0))
        bin_fields = cv2.morphologyEx(base_bin, cv2.MORPH_CLOSE, np.ones((field_close_k, field_close_k), np.uint8)) if field_close_k > 0 else base_bin
        bin_checks = base_bin.copy()
        red_close_k = max(0, int(self.settings.get('C_Close', 0) or 0))
        red_open_k = max(0, int(self.settings.get('C_Open', 1) or 0))
        if red_close_k > 0:
            bin_checks = cv2.morphologyEx(bin_checks, cv2.MORPH_CLOSE, np.ones((red_close_k, red_close_k), np.uint8))
        if red_open_k > 0:
            bin_checks = cv2.morphologyEx(bin_checks, cv2.MORPH_OPEN, np.ones((red_open_k, red_open_k), np.uint8))

        local_rects = []
        local_types = []
        # lines
        min_line_w = max(18, min(int(self.settings.get('Line_MinW', 100) or 100), int(crop.shape[1] * 0.85)))
        max_line_w = max(min_line_w + 8, int(crop.shape[1] * 0.98))
        for lr in self.find_answer_lines(crop, zoom_factor=1.0, min_line_w=min_line_w, max_line_w=max_line_w):
            local_rect = fitz.Rect(lr)
            rect = fitz.Rect(local_rect.x0 + x0, local_rect.y0 + y0, local_rect.x1 + x0, local_rect.y1 + y0)
            if not self._manual_capture_reject_text_artifact(local_rect, 'line', text_mask, crop.shape):
                local_rects.append(rect)
                local_types.append('line')
        # checks
        nf, _, stats, _ = cv2.connectedComponentsWithStats(bin_checks)
        checkbox_min_sz = max(7, int(self.settings.get('C_Size', [14,65])[0]) - 5)
        checkbox_max_sz = int(self.settings.get('C_Size', [14,65])[1])
        for i in range(1, nf):
            xx, yy, ww, hh, area = stats[i]
            if not (checkbox_min_sz <= ww <= checkbox_max_sz and checkbox_min_sz <= hh <= checkbox_max_sz):
                continue
            if self.looks_like_checkbox(bin_checks, xx, yy, ww, hh, area):
                local_rect = fitz.Rect(xx, yy, xx + ww, yy + hh)
                rect = fitz.Rect((xx + x0), (yy + y0), (xx + ww + x0), (yy + hh + y0))
                if not self._manual_capture_reject_text_artifact(local_rect, 'check', text_mask, crop.shape):
                    local_rects.append(rect)
                    local_types.append('check')
        # fields
        nf2, _, stats2, _ = cv2.connectedComponentsWithStats(bin_fields)
        field_area_thresh = max(30, int(self.settings.get('F_Area', 500) or 500) // 3)
        field_min_w = max(8, int(self.settings.get('F_MinW', 15) or 15))
        field_min_h = max(10, int(self.settings.get('F_MinH', 35) or 35) // 2)
        for i in range(1, nf2):
            xx, yy, ww, hh, area = stats2[i]
            if hh < field_min_h or ww < field_min_w or area < field_area_thresh:
                continue
            if ww >= crop.shape[1] * 0.96 or hh >= crop.shape[0] * 0.90:
                continue
            refined = self._refine_field_rect_from_mask(bin_fields, xx, yy, ww, hh, zoom_factor=1.0, row_frac_thresh=0.40, col_frac_thresh=0.14, min_inner_h=max(8, field_min_h // 2), pad_px=1)
            local_rect = fitz.Rect(refined)
            rect = fitz.Rect(local_rect.x0 + x0, local_rect.y0 + y0, local_rect.x1 + x0, local_rect.y1 + y0)
            if not self._manual_capture_reject_text_artifact(local_rect, 'field', text_mask, crop.shape):
                local_rects.append(rect)
                local_types.append('field')

        # dedupe/local cleanup
        dedup_rects = []
        dedup_types = []
        for rect, kind in sorted(zip(local_rects, local_types), key=lambda rt: (rt[0].y0, rt[0].x0, rt[0].width * rt[0].height)):
            duplicate = False
            for ex_rect, ex_kind in zip(dedup_rects, dedup_types):
                if ex_kind != kind:
                    continue
                if self._rect_iou(rect, ex_rect) >= 0.60 or self._rects_refer_to_same_target(rect, ex_rect, tol=0.60):
                    duplicate = True
                    break
            if not duplicate:
                dedup_rects.append(rect)
                dedup_types.append(kind)

        source_text_overlap = 0.0
        try:
            source_text_overlap = float(self.rect_text_overlap_ratio(fitz.Rect(0, 0, max(1, x1 - x0), max(1, y1 - y0)), text_mask, preview_zoom=1.0) or 0.0)
        except Exception:
            source_text_overlap = 0.0

        union_cover = 0.0
        if dedup_rects:
            union = fitz.Rect(min(r.x0 for r in dedup_rects), min(r.y0 for r in dedup_rects), max(r.x1 for r in dedup_rects), max(r.y1 for r in dedup_rects))
            union_cover = (union.width * union.height) / float(max(source_rect.width * ZOOM * source_rect.height * ZOOM, 1.0))
        prefers_single_field = False
        if len(dedup_rects) >= 2:
            try:
                prefers_single_field = bool(self._manual_capture_force_single_field_from_enclosing_box(source_rect, dedup_rects, dedup_types))
            except Exception:
                prefers_single_field = False
            if not prefers_single_field:
                try:
                    prefers_single_field = bool(self._manual_capture_prefers_single_field(crop, source_rect, dedup_rects, dedup_types, base_bin=base_bin, bin_fields=bin_fields))
                except Exception:
                    prefers_single_field = False
        needs_split = len(dedup_rects) >= 2 and not prefers_single_field
        if len(dedup_rects) <= 0:
            resolved_rects = [fitz.Rect(source_rect)]
            resolved_types = ['field']
        elif needs_split:
            resolved_rects = [fitz.Rect(r.x0 / ZOOM, r.y0 / ZOOM, r.x1 / ZOOM, r.y1 / ZOOM) for r in dedup_rects]
            resolved_types = list(dedup_types)
        else:
            resolved_rects = [fitz.Rect(source_rect)]
            resolved_types = ['field' if prefers_single_field else (dedup_types[0] if dedup_types else 'field')]
        type_counts = {}
        for kind in resolved_types:
            type_counts[kind] = int(type_counts.get(kind, 0) or 0) + 1
        template_kind = max(type_counts, key=type_counts.get) if type_counts else 'field'
        if needs_split:
            template_kind = f"split_{template_kind}"
        return {
            'page': page_idx,
            'source_rect': fitz.Rect(source_rect),
            'resolved_rects': resolved_rects,
            'resolved_types': resolved_types,
            'needs_split': bool(needs_split),
            'split_count': len(resolved_rects),
            'template_kind': template_kind,
            'source_text_overlap': float(round(source_text_overlap, 6)),
            'features': {
                'crop_width_px': int(crop.shape[1]),
                'crop_height_px': int(crop.shape[0]),
                'raw_internal_count': int(len(dedup_rects)),
                'union_cover_ratio': float(round(union_cover, 6)),
                'type_counts': type_counts,
                'prefers_single_field': bool(prefers_single_field),
            },
            'shape_template': {},
        }

    def _shape_template_from_contour(self, contour, roi_shape, source_rect, score=0.0, candidate_index=0):
        try:
            rr = source_rect if isinstance(source_rect, fitz.Rect) else fitz.Rect(source_rect)
        except Exception:
            return {}
        if contour is None or len(contour) < 3:
            return {}
        roi_h = int(max(1, roi_shape[0] if isinstance(roi_shape, (list, tuple)) else roi_shape[0]))
        roi_w = int(max(1, roi_shape[1] if isinstance(roi_shape, (list, tuple)) else roi_shape[1]))
        try:
            peri = float(cv2.arcLength(contour, True))
            area = float(cv2.contourArea(contour))
            x, y, w, h = cv2.boundingRect(contour)
            approx = cv2.approxPolyDP(contour, max(1.0, 0.02 * peri), True)
            moments = cv2.HuMoments(cv2.moments(contour)).flatten().tolist()
            contour_pts = contour.reshape(-1, 2)
        except Exception:
            return {}
        pdf_pts = []
        norm_pts = []
        for px, py in contour_pts.tolist():
            nx = float(px) / float(max(roi_w, 1))
            ny = float(py) / float(max(roi_h, 1))
            norm_pts.append([round(nx, 6), round(ny, 6)])
            pdf_pts.append([round(float(rr.x0 + (nx * rr.width)), 4), round(float(rr.y0 + (ny * rr.height)), 4)])
        bbox_rect = fitz.Rect(rr.x0 + (float(x) / max(roi_w, 1)) * rr.width, rr.y0 + (float(y) / max(roi_h, 1)) * rr.height, rr.x0 + (float(x + w) / max(roi_w, 1)) * rr.width, rr.y0 + (float(y + h) / max(roi_h, 1)) * rr.height)
        return {
            'candidate_index': int(candidate_index),
            'score': float(round(score, 6)),
            'roi_size_px': [int(roi_w), int(roi_h)],
            'norm_points': norm_pts,
            'pdf_points': pdf_pts,
            'hu_moments': [float(round(v, 10)) for v in moments],
            'bbox_rect': self._serialize_rect(bbox_rect),
            'bbox_aspect': float(round(float(w) / float(max(h, 1)), 6)),
            'area_ratio': float(round(area / float(max(roi_w * roi_h, 1)), 6)),
            'extent': float(round(area / float(max(w * h, 1)), 6)),
            'vertices': int(len(approx) if approx is not None else 0),
            'kind_hint': 'shape',
        }

    def extract_shape_candidates_from_roi(self, page_idx, source_rect, max_candidates=8):
        page_idx = int(page_idx or 0)
        try:
            rr = source_rect if isinstance(source_rect, fitz.Rect) else fitz.Rect(source_rect)
        except Exception:
            return []
        if rr.width <= 1 or rr.height <= 1:
            return []
        img = self._render_pdf_page_bgr(self.pdf_path, page_idx=page_idx, preview_zoom=ZOOM)
        if img is None or getattr(img, 'size', 0) == 0:
            return []
        h_img, w_img = img.shape[:2]
        x0 = max(0, min(int(math.floor(rr.x0 * ZOOM)), w_img - 1))
        y0 = max(0, min(int(math.floor(rr.y0 * ZOOM)), h_img - 1))
        x1 = max(x0 + 1, min(int(math.ceil(rr.x1 * ZOOM)), w_img))
        y1 = max(y0 + 1, min(int(math.ceil(rr.y1 * ZOOM)), h_img))
        crop = img[y0:y1, x0:x1].copy()
        if crop is None or getattr(crop, 'size', 0) == 0:
            return []
        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if len(crop.shape) == 3 else crop
        text_mask, _ = self.build_text_like_mask(crop, header_band_pct=0.10)
        try:
            light_text_mask = cv2.dilate(text_mask, np.ones((2, 2), np.uint8), iterations=1)
        except Exception:
            light_text_mask = text_mask
        work_masks = []
        try:
            otsu = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
            work_masks.append(('otsu', otsu))
        except Exception:
            pass
        try:
            adapt = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 21, 5)
            work_masks.append(('adaptive', adapt))
        except Exception:
            pass
        try:
            edges = cv2.Canny(gray, 60, 160)
            edges = cv2.dilate(edges, np.ones((2, 2), np.uint8), iterations=1)
            work_masks.append(('edge', edges))
        except Exception:
            pass
        seen = []
        candidates = []
        roi_h, roi_w = gray.shape[:2]
        roi_cx = roi_w * 0.5
        roi_cy = roi_h * 0.5
        for source_name, mask in work_masks:
            try:
                if light_text_mask is not None and getattr(light_text_mask, 'size', 0) != 0:
                    mask = cv2.bitwise_and(mask, cv2.bitwise_not(light_text_mask))
            except Exception:
                pass
            try:
                mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))
            except Exception:
                pass
            try:
                contours, _ = cv2.findContours(mask, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            except Exception:
                contours = []
            for c in contours:
                try:
                    area = float(cv2.contourArea(c))
                    peri = float(cv2.arcLength(c, True))
                    x, y, w, h = cv2.boundingRect(c)
                except Exception:
                    continue
                if area <= 8 or peri <= 0 or w < 4 or h < 4:
                    continue
                if w >= roi_w * 0.985 and h >= roi_h * 0.985:
                    continue
                border_touch = 0
                if x <= 1: border_touch += 1
                if y <= 1: border_touch += 1
                if (x + w) >= (roi_w - 1): border_touch += 1
                if (y + h) >= (roi_h - 1): border_touch += 1
                if border_touch >= 3:
                    continue
                extent = float(area / float(max(w * h, 1)))
                if extent <= 0.02:
                    continue
                approx = cv2.approxPolyDP(c, max(1.0, 0.03 * peri), True)
                vertices = int(len(approx) if approx is not None else 0)
                cx = x + (w * 0.5)
                cy = y + (h * 0.5)
                center_penalty = min(1.0, (abs(cx - roi_cx) / max(roi_w, 1)) + (abs(cy - roi_cy) / max(roi_h, 1)))
                area_ratio = float(area / float(max(roi_w * roi_h, 1)))
                compactness = float((4.0 * math.pi * area) / max(peri * peri, 1e-6))
                edge_hug_penalty = 0.08 * float(border_touch)
                score = (area_ratio * 0.85) + (extent * 0.30) + (min(max(compactness, 0.0), 1.0) * 0.16) - (center_penalty * 0.12) - edge_hug_penalty
                if vertices == 4:
                    score += 0.10
                if 0.55 <= (float(w) / float(max(h, 1))) <= 1.60:
                    score += 0.06
                norm_box = (round(x / max(roi_w, 1), 2), round(y / max(roi_h, 1), 2), round(w / max(roi_w, 1), 2), round(h / max(roi_h, 1), 2))
                if norm_box in seen:
                    continue
                seen.append(norm_box)
                sig = self._shape_template_from_contour(c, (roi_h, roi_w), rr, score=score, candidate_index=len(candidates))
                if not sig:
                    continue
                sig['source'] = str(source_name)
                sig['resolved_rects'] = [fitz.Rect(sig['bbox_rect'])] if sig.get('bbox_rect') else [fitz.Rect(rr)]
                sig['resolved_types'] = ['check' if int(sig.get('vertices', 0) or 0) == 4 and float(sig.get('bbox_aspect', 1.0) or 1.0) <= 1.6 else 'field']
                sig['split_count'] = 1
                sig['needs_split'] = False
                sig['template_kind'] = 'shape_template'
                candidates.append(sig)
        candidates.sort(key=lambda item: float(item.get('score', 0.0) or 0.0), reverse=True)
        final = []
        for cand in candidates:
            try:
                cand_rect = fitz.Rect(cand.get('bbox_rect')) if cand.get('bbox_rect') is not None else None
            except Exception:
                cand_rect = None
            duplicate = False
            for prev in final:
                try:
                    prev_rect = fitz.Rect(prev.get('bbox_rect')) if prev.get('bbox_rect') is not None else None
                except Exception:
                    prev_rect = None
                if cand_rect is not None and prev_rect is not None and self._rect_iou(cand_rect, prev_rect) >= 0.55:
                    duplicate = True
                    break
            if not duplicate:
                final.append(cand)
            if len(final) >= max(1, int(max_candidates or 3)):
                break
        return final

    def extract_shape_candidates_from_page(self, page_idx, template_shape=None, max_candidates=160):
        page_idx = int(page_idx or 0)
        img = self._render_pdf_page_bgr(self.pdf_path, page_idx=page_idx, preview_zoom=ZOOM)
        if img is None or getattr(img, 'size', 0) == 0:
            return []
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) == 3 else img
        text_mask, _ = self.build_text_like_mask(img, header_band_pct=0.10)
        try:
            light_text_mask = cv2.dilate(text_mask, np.ones((2, 2), np.uint8), iterations=1)
        except Exception:
            light_text_mask = text_mask
        work_masks = []
        try:
            otsu = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
            work_masks.append(('otsu', otsu))
        except Exception:
            pass
        try:
            adapt = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 25, 6)
            work_masks.append(('adaptive', adapt))
        except Exception:
            pass
        try:
            edges = cv2.Canny(gray, 60, 160)
            edges = cv2.dilate(edges, np.ones((2, 2), np.uint8), iterations=1)
            work_masks.append(('edge', edges))
        except Exception:
            pass
        h_img, w_img = gray.shape[:2]
        tpl_aspect = 0.0
        tpl_area_ratio = 0.0
        if isinstance(template_shape, dict):
            try:
                tpl_aspect = float(template_shape.get('bbox_aspect', 0.0) or 0.0)
            except Exception:
                tpl_aspect = 0.0
            try:
                tpl_area_ratio = float(template_shape.get('area_ratio', 0.0) or 0.0)
            except Exception:
                tpl_area_ratio = 0.0
        seen = []
        candidates = []
        for source_name, mask in work_masks:
            try:
                if light_text_mask is not None and getattr(light_text_mask, 'size', 0) != 0:
                    mask = cv2.bitwise_and(mask, cv2.bitwise_not(light_text_mask))
            except Exception:
                pass
            try:
                mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))
            except Exception:
                pass
            try:
                contours, _ = cv2.findContours(mask, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            except Exception:
                contours = []
            for c in contours:
                try:
                    area = float(cv2.contourArea(c))
                    peri = float(cv2.arcLength(c, True))
                    x, y, w, h = cv2.boundingRect(c)
                except Exception:
                    continue
                if area <= 10 or peri <= 0 or w < 4 or h < 4:
                    continue
                if w >= w_img * 0.985 and h >= h_img * 0.985:
                    continue
                border_touch = 0
                if x <= 1: border_touch += 1
                if y <= 1: border_touch += 1
                if (x + w) >= (w_img - 1): border_touch += 1
                if (y + h) >= (h_img - 1): border_touch += 1
                if border_touch >= 3:
                    continue
                area_ratio = float(area / float(max(w * h, 1)))
                if area_ratio <= 0.02:
                    continue
                bbox_aspect = float(w) / float(max(h, 1))
                norm_box = (round(x / max(w_img, 1), 3), round(y / max(h_img, 1), 3), round(w / max(w_img, 1), 3), round(h / max(h_img, 1), 3))
                if norm_box in seen:
                    continue
                seen.append(norm_box)
                pdf_rr = fitz.Rect(float(x) / ZOOM, float(y) / ZOOM, float(x + w) / ZOOM, float(y + h) / ZOOM)
                sig = self._shape_template_from_contour(c, (h_img, w_img), pdf_rr, score=0.0, candidate_index=len(candidates))
                if not sig:
                    continue
                compactness = float((4.0 * math.pi * area) / max(peri * peri, 1e-6))
                score = (min(float(sig.get('area_ratio', 0.0) or 0.0), 1.0) * 0.50) + (min(max(compactness, 0.0), 1.0) * 0.16) - (0.08 * float(border_touch))
                if tpl_aspect > 0:
                    score += max(0.0, 1.0 - min(abs(tpl_aspect - bbox_aspect) / 1.30, 1.0)) * 0.18
                if tpl_area_ratio > 0:
                    score += max(0.0, 1.0 - min(abs(tpl_area_ratio - float(sig.get('area_ratio', 0.0) or 0.0)) / 0.35, 1.0)) * 0.12
                if int(sig.get('vertices', 0) or 0) == 4:
                    score += 0.05
                sig['score'] = float(score)
                sig['source'] = str(source_name)
                sig['resolved_rects'] = [fitz.Rect(sig['bbox_rect'])] if sig.get('bbox_rect') else [fitz.Rect(pdf_rr)]
                sig['resolved_types'] = ['check' if int(sig.get('vertices', 0) or 0) == 4 and float(sig.get('bbox_aspect', 1.0) or 1.0) <= 1.6 else 'field']
                sig['split_count'] = 1
                sig['needs_split'] = False
                sig['template_kind'] = 'shape_template'
                candidates.append(sig)
        candidates.sort(key=lambda item: float(item.get('score', 0.0) or 0.0), reverse=True)
        final = []
        for cand in candidates:
            try:
                cand_rect = fitz.Rect(cand.get('bbox_rect')) if cand.get('bbox_rect') is not None else None
            except Exception:
                cand_rect = None
            duplicate = False
            for prev in final:
                try:
                    prev_rect = fitz.Rect(prev.get('bbox_rect')) if prev.get('bbox_rect') is not None else None
                except Exception:
                    prev_rect = None
                if cand_rect is not None and prev_rect is not None and self._rect_iou(cand_rect, prev_rect) >= 0.70:
                    duplicate = True
                    break
            if not duplicate:
                final.append(cand)
            if len(final) >= max(1, int(max_candidates or 160)):
                break
        return final

    def build_shape_template_analysis(self, page_idx, source_rect, candidate):
        try:
            rr = source_rect if isinstance(source_rect, fitz.Rect) else fitz.Rect(source_rect)
        except Exception:
            return None
        candidate = dict(candidate or {})
        resolved_rects = []
        for rect in (candidate.get('resolved_rects', []) or []):
            try:
                resolved_rects.append(rect if isinstance(rect, fitz.Rect) else fitz.Rect(rect))
            except Exception:
                pass
        if not resolved_rects:
            try:
                resolved_rects = [fitz.Rect(candidate.get('bbox_rect'))] if candidate.get('bbox_rect') is not None else [fitz.Rect(rr)]
            except Exception:
                resolved_rects = [fitz.Rect(rr)]
        resolved_types = [str(v) for v in (candidate.get('resolved_types', []) or [])]
        if not resolved_types:
            resolved_types = ['field']
        shape_template = dict(candidate or {})
        shape_template['source_rect'] = self._serialize_rect(rr)
        return {
            'page': int(page_idx or 0),
            'source_rect': fitz.Rect(rr),
            'resolved_rects': resolved_rects,
            'resolved_types': resolved_types,
            'needs_split': False,
            'split_count': len(resolved_rects),
            'template_kind': 'shape_template',
            'source_text_overlap': 0.0,
            'features': {
                'crop_width_px': int(candidate.get('roi_size_px', [0, 0])[0] or 0),
                'crop_height_px': int(candidate.get('roi_size_px', [0, 0])[1] or 0),
                'raw_internal_count': 1,
                'union_cover_ratio': float(candidate.get('area_ratio', 0.0) or 0.0),
                'type_counts': {str(resolved_types[0]): 1},
                'prefers_single_field': True,
                'shape_score': float(candidate.get('score', 0.0) or 0.0),
            },
            'shape_template': shape_template,
        }

    def _shape_similarity_score(self, template_shape, candidate_shape):
        try:
            tpl_hu = np.array(list(template_shape.get('hu_moments', []) or []), dtype=np.float64)
            cand_hu = np.array(list(candidate_shape.get('hu_moments', []) or []), dtype=np.float64)
        except Exception:
            return 0.0
        if tpl_hu.size != 7 or cand_hu.size != 7:
            return 0.0
        try:
            h1 = np.sign(tpl_hu) * np.log10(np.abs(tpl_hu) + 1e-12)
            h2 = np.sign(cand_hu) * np.log10(np.abs(cand_hu) + 1e-12)
            hu_dist = float(np.linalg.norm(h1 - h2))
        except Exception:
            hu_dist = 99.0
        hu_sim = 1.0 / (1.0 + hu_dist)
        aspect_tpl = float(template_shape.get('bbox_aspect', 1.0) or 1.0)
        aspect_cand = float(candidate_shape.get('bbox_aspect', 1.0) or 1.0)
        aspect_sim = max(0.0, 1.0 - min(abs(aspect_tpl - aspect_cand) / 1.25, 1.0))
        area_tpl = float(template_shape.get('area_ratio', 0.0) or 0.0)
        area_cand = float(candidate_shape.get('area_ratio', 0.0) or 0.0)
        size_sim = max(0.0, 1.0 - min(abs(area_tpl - area_cand) / 0.30, 1.0))
        try:
            tpl_cnt = np.array(template_shape.get('norm_points', []) or [], dtype=np.float32).reshape(-1, 1, 2)
            cand_cnt = np.array(candidate_shape.get('norm_points', []) or [], dtype=np.float32).reshape(-1, 1, 2)
            contour_diff = float(cv2.matchShapes(tpl_cnt, cand_cnt, cv2.CONTOURS_MATCH_I1, 0.0)) if tpl_cnt.size and cand_cnt.size else 99.0
        except Exception:
            contour_diff = 99.0
        contour_sim = 1.0 / (1.0 + max(contour_diff, 0.0))
        score = (contour_sim * 0.40) + (hu_sim * 0.25) + (aspect_sim * 0.20) + (size_sim * 0.15)
        return float(max(0.0, min(score, 0.99)))

    def _shape_template_preview_path(self, shape_template, preview_zoom=PREVIEW_SCALE):
        points = []
        for pt in (dict(shape_template or {}).get('pdf_points', []) or []):
            if not isinstance(pt, (list, tuple)) or len(pt) < 2:
                continue
            try:
                points.append([float(pt[0]) * float(preview_zoom or PREVIEW_SCALE), float(pt[1]) * float(preview_zoom or PREVIEW_SCALE)])
            except Exception:
                continue
        return points

    def _page_detection_cache_token(self, page_idx):
        try:
            page_idx = int(page_idx or 0)
        except Exception:
            page_idx = 0
        if int(getattr(self, 'detected_page_idx', -1) or -1) == page_idx and bool(getattr(self, 'all_boxes', []) or []):
            boxes = list(getattr(self, 'all_boxes', []) or [])
            types = list(getattr(self, 'box_types', []) or [])
        else:
            boxes = list((getattr(self, 'boxes_by_page', {}) or {}).get(page_idx, []) or [])
            types = list((getattr(self, 'box_types_by_page', {}) or {}).get(page_idx, []) or [])
        payload = {
            'page': int(page_idx),
            'boxes': [self._serialize_rect(r) for r in boxes],
            'types': [str(t) for t in types],
            'settings_signature': str(self._settings_signature() or ''),
        }
        return _sha1_json(payload)

    def _area_template_match_scope_key(self, page_idx, template_id):
        try:
            page_idx = int(page_idx or 0)
        except Exception:
            page_idx = 0
        sig = str(self._settings_signature() or '')
        detect_token = self._page_detection_cache_token(page_idx)
        return f"{page_idx}|{str(template_id or '').strip()}|{sig}|{detect_token}"

    def cache_area_template_matches(self, page_idx, template_id, matches):
        key = self._area_template_match_scope_key(page_idx, template_id)
        payload = []
        for item in (matches or []):
            if not isinstance(item, dict):
                continue
            payload.append({
                'match_id': str(item.get('match_id', '') or ''),
                'template_id': str(item.get('template_id', '') or ''),
                'page': int(item.get('page', page_idx) or page_idx),
                'rect': self._serialize_rect(item.get('rect')) if item.get('rect') is not None else None,
                'score': float(item.get('score', 0.0) or 0.0),
                'reason': str(item.get('reason', '') or ''),
                'resolved_rects': self._serialize_rects(item.get('resolved_rects', []) or []),
                'resolved_types': [str(v) for v in (item.get('resolved_types', []) or [])],
                'split_count': int(item.get('split_count', 0) or 0),
            })
        self.area_template_match_cache[key] = payload
        return key

    def get_cached_area_template_matches(self, page_idx, template_id):
        key = self._area_template_match_scope_key(page_idx, template_id)
        payload = list((self.area_template_match_cache or {}).get(key, []) or [])
        matches = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            matches.append({
                'match_id': str(item.get('match_id', '') or ''),
                'template_id': str(item.get('template_id', '') or ''),
                'page': int(item.get('page', page_idx) or page_idx),
                'rect': fitz.Rect(item.get('rect')) if item.get('rect') is not None else None,
                'score': float(item.get('score', 0.0) or 0.0),
                'reason': str(item.get('reason', '') or ''),
                'resolved_rects': [fitz.Rect(r) for r in (item.get('resolved_rects', []) or [])],
                'resolved_types': [str(v) for v in (item.get('resolved_types', []) or [])],
                'split_count': int(item.get('split_count', 0) or 0),
            })
        return matches

    def update_area_template_match_review_state(self, page_idx, template_id, review_state=None):
        key = self._area_template_match_scope_key(page_idx, template_id)
        state = {}
        for rk, rv in dict(review_state or {}).items():
            state[str(rk)] = bool(rv)
        self.area_template_match_review_state[key] = state
        return state

    def get_area_template_match_review_state(self, page_idx, template_id):
        key = self._area_template_match_scope_key(page_idx, template_id)
        return dict((self.area_template_match_review_state or {}).get(key, {}) or {})

    def invalidate_area_template_match_cache(self, page_idx=None, template_id=None):
        cache = dict(getattr(self, 'area_template_match_cache', {}) or {})
        review = dict(getattr(self, 'area_template_match_review_state', {}) or {})
        if page_idx is None and not str(template_id or '').strip():
            self.area_template_match_cache = {}
            self.area_template_match_review_state = {}
            return
        try:
            page_idx_norm = None if page_idx is None else int(page_idx)
        except Exception:
            page_idx_norm = None
        template_norm = str(template_id or '').strip()
        def _keep(scope_key):
            parts = str(scope_key or '').split('|')
            key_page_raw = parts[0] if parts else ''
            try:
                key_page = int(key_page_raw)
            except Exception:
                key_page = None
            key_template = parts[1] if len(parts) > 1 else ''
            if page_idx_norm is not None and key_page_raw != 'FORM' and key_page != page_idx_norm:
                return True
            if template_norm and key_template != template_norm:
                return True
            return False
        self.area_template_match_cache = {k: v for k, v in cache.items() if _keep(k)}
        self.area_template_match_review_state = {k: v for k, v in review.items() if _keep(k)}

    def _form_detection_cache_token(self):
        try:
            total_pages = int(self.total_pages() or 0)
        except Exception:
            total_pages = 0
        payload = {
            'pdf_path': str(getattr(self, 'pdf_path', '') or ''),
            'settings_signature': str(self._settings_signature() or ''),
            'page_tokens': [self._page_detection_cache_token(page_idx) for page_idx in range(max(total_pages, 0))],
        }
        return _sha1_json(payload)

    def _form_area_template_match_scope_key(self, template_id):
        template_norm = str(template_id or '').strip()
        return f"FORM|{template_norm}|{self._settings_signature()}|{self._form_detection_cache_token()}"

    def cache_form_area_template_matches(self, template_id, matches):
        key = self._form_area_template_match_scope_key(template_id)
        payload = []
        for item in (matches or []):
            if not isinstance(item, dict):
                continue
            payload.append({
                'match_id': str(item.get('match_id', '') or ''),
                'template_id': str(item.get('template_id', '') or ''),
                'page': int(item.get('page', 0) or 0),
                'rect': self._serialize_rect(item.get('rect')) if item.get('rect') is not None else None,
                'score': float(item.get('score', 0.0) or 0.0),
                'reason': str(item.get('reason', '') or ''),
                'resolved_rects': self._serialize_rects(item.get('resolved_rects', []) or []),
                'resolved_types': [str(v) for v in (item.get('resolved_types', []) or [])],
                'split_count': int(item.get('split_count', 0) or 0),
            })
        self.area_template_match_cache[key] = payload
        return payload

    def get_cached_form_area_template_matches(self, template_id):
        key = self._form_area_template_match_scope_key(template_id)
        payload = list((self.area_template_match_cache or {}).get(key, []) or [])
        restored = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            rect = item.get('rect')
            try:
                rect = fitz.Rect(rect) if rect is not None else None
            except Exception:
                rect = None
            restored.append({
                'match_id': str(item.get('match_id', '') or ''),
                'template_id': str(item.get('template_id', '') or ''),
                'page': int(item.get('page', 0) or 0),
                'rect': rect,
                'score': float(item.get('score', 0.0) or 0.0),
                'reason': str(item.get('reason', '') or ''),
                'resolved_rects': [fitz.Rect(r) for r in (item.get('resolved_rects', []) or [])],
                'resolved_types': [str(v) for v in (item.get('resolved_types', []) or [])],
                'split_count': int(item.get('split_count', 0) or 0),
            })
        return restored

    def update_form_area_template_match_review_state(self, template_id, review_state=None):
        key = self._form_area_template_match_scope_key(template_id)
        state = {}
        for rk, rv in dict(review_state or {}).items():
            state[str(rk)] = bool(rv)
        self.area_template_match_review_state[key] = state
        return state

    def get_form_area_template_match_review_state(self, template_id):
        key = self._form_area_template_match_scope_key(template_id)
        return dict((self.area_template_match_review_state or {}).get(key, {}) or {})

    def save_area_template(self, analysis):
        if not isinstance(analysis, dict):
            return None
        page_idx = int(analysis.get('page', 0) or 0)
        template = {
            'template_id': str(uuid.uuid4().hex[:12]),
            'page': page_idx,
            'source_rect': fitz.Rect(analysis.get('source_rect')),
            'resolved_rects': [fitz.Rect(r) for r in (analysis.get('resolved_rects', []) or [])],
            'resolved_types': [str(v) for v in (analysis.get('resolved_types', []) or [])],
            'needs_split': bool(analysis.get('needs_split', False)),
            'template_kind': str(analysis.get('template_kind', '') or ''),
            'split_count': int(analysis.get('split_count', 0) or 0),
            'source_text_overlap': float(analysis.get('source_text_overlap', 0.0) or 0.0),
            'features': dict(analysis.get('features', {}) or {}),
            'shape_template': dict(analysis.get('shape_template', {}) or {}),
            'created_at': _utc_now_iso(),
        }
        self.area_templates_by_page.setdefault(page_idx, []).append(template)
        self.area_template_index[template['template_id']] = template
        return template

    def merge_area_template_into_detections(self, page_idx, template):
        page_idx = int(page_idx or 0)
        if not isinstance(template, dict):
            return {'added_refs': [], 'selected_ids': []}
        resolved_rects = [fitz.Rect(r) for r in (template.get('resolved_rects', []) or [])]
        resolved_types = [str(v) for v in (template.get('resolved_types', []) or [])]
        if not resolved_rects:
            return {'added_refs': [], 'selected_ids': []}
        return self._merge_manual_rect_type_pairs(page_idx, list(zip(resolved_rects, resolved_types or ['field'] * len(resolved_rects))))

    def _merge_manual_rect_type_pairs(self, page_idx, rect_type_pairs):
        page_idx = int(page_idx or 0)
        added_refs = []
        for rect, kind in (rect_type_pairs or []):
            try:
                rect = fitz.Rect(rect)
            except Exception:
                continue
            kind = str(kind or 'field')
            kind = kind if kind in {'field', 'line', 'check'} else 'field'
            before = len(self.all_boxes)
            if kind == 'check':
                self._append_box_unique(rect, kind, iou_thresh=0.50)
            else:
                self._append_box_unique(rect, kind, iou_thresh=0.62)
            if len(self.all_boxes) > before:
                added_refs.append((fitz.Rect(rect), kind))
        selected_ids = []
        if added_refs:
            try:
                self._cleanup_field_fragments()
                self._cleanup_line_field_conflicts()
            except Exception:
                pass
            self.detected_page_idx = int(page_idx)
            self.detected_settings_signature = self._settings_signature()
            self._cache_detection_for_page(page_idx)
            used = set()
            for rect, kind in added_refs:
                best_idx = -1
                best_score = 0.0
                for idx_existing, existing in enumerate(self.all_boxes):
                    if idx_existing in used:
                        continue
                    existing_kind = str(self.box_types[idx_existing] if idx_existing < len(self.box_types) else '')
                    if kind and existing_kind != str(kind):
                        continue
                    score = self._mobile_repair_overlap_score(rect, existing)
                    if score > best_score:
                        best_idx = idx_existing
                        best_score = score
                if best_idx >= 0 and best_score >= 0.45:
                    used.add(best_idx)
                    selected_ids.append(best_idx)
        return {'added_refs': added_refs, 'selected_ids': sorted(set(int(x) for x in selected_ids if isinstance(x, int) or str(x).isdigit()))}


    def _manual_area_search_scope_key(self, page_idx):
        try:
            page_idx = int(page_idx or 0)
        except Exception:
            page_idx = 0
        return f"{page_idx}|{self._page_detection_cache_token(page_idx)}"

    def _get_manual_area_search_cache_entry(self, page_idx, page_img=None):
        try:
            page_idx = int(page_idx or 0)
        except Exception:
            page_idx = 0
        scope_key = self._manual_area_search_scope_key(page_idx)
        cache = dict(getattr(self, 'manual_area_search_cache', {}) or {})
        entry = cache.get(scope_key)
        if isinstance(entry, dict):
            if page_img is not None and entry.get('page_img') is None:
                entry['page_img'] = page_img
            return entry
        if page_img is None:
            page_img = self._render_pdf_page_bgr(self.pdf_path, page_idx=page_idx, preview_zoom=ZOOM)
        if page_img is None or getattr(page_img, 'size', 0) == 0:
            return None
        gray = cv2.cvtColor(page_img, cv2.COLOR_BGR2GRAY)
        base_bin = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        try:
            page_text_mask, _ = self.build_text_like_mask(page_img, header_band_pct=0.10)
            page_text_mask = cv2.dilate(page_text_mask, np.ones((3, 3), np.uint8), iterations=1)
            struct_bin = cv2.bitwise_and(base_bin, cv2.bitwise_not(page_text_mask))
        except Exception:
            page_text_mask = None
            struct_bin = base_bin
        n, _, stats, _ = cv2.connectedComponentsWithStats(struct_bin)
        entry = {
            'scope_key': scope_key,
            'page_img': page_img,
            'gray': gray,
            'struct_bin': struct_bin,
            'stats': stats,
            'component_count': int(n),
            'candidate_rects': None,
            'signature_cache': {},
        }
        cache = {k: v for k, v in cache.items() if str(k).endswith(str(self._page_detection_cache_token(page_idx))) is False or k == scope_key}
        cache[scope_key] = entry
        self.manual_area_search_cache = cache
        return entry

    def _get_cached_manual_area_signature(self, page_idx, rect, page_img=None):
        entry = self._get_manual_area_search_cache_entry(page_idx, page_img=page_img)
        if not isinstance(entry, dict):
            return self._build_manual_area_structural_signature(page_idx, rect, page_img=page_img)
        try:
            rect_obj = fitz.Rect(rect)
        except Exception:
            return None
        rect_key_payload = self._serialize_rect(rect_obj)
        if isinstance(rect_key_payload, list):
            rect_key = tuple(rect_key_payload)
        elif isinstance(rect_key_payload, tuple):
            rect_key = rect_key_payload
        else:
            rect_key = tuple(list(rect_key_payload)) if rect_key_payload is not None and not isinstance(rect_key_payload, (str, bytes, int, float, bool)) else rect_key_payload
        sig_cache = entry.setdefault('signature_cache', {})
        sig = sig_cache.get(rect_key)
        if sig is None:
            sig = self._build_manual_area_structural_signature(page_idx, rect_obj, page_img=entry.get('page_img'))
            sig_cache[rect_key] = sig
        return sig


    def _manual_capture_text_overlap_expanded(self, local_rect, text_mask, pad_px=4):
        try:
            rr = fitz.Rect(local_rect)
        except Exception:
            return 0.0
        if text_mask is None or getattr(text_mask, 'size', 0) == 0:
            return 0.0
        h, w = text_mask.shape[:2]
        x0 = max(0, int(math.floor(rr.x0)) - int(max(0, pad_px)))
        y0 = max(0, int(math.floor(rr.y0)) - int(max(0, pad_px)))
        x1 = min(w, int(math.ceil(rr.x1)) + int(max(0, pad_px)))
        y1 = min(h, int(math.ceil(rr.y1)) + int(max(0, pad_px)))
        if x1 <= x0 or y1 <= y0:
            return 0.0
        roi = text_mask[y0:y1, x0:x1]
        if roi.size == 0:
            return 0.0
        return float(np.count_nonzero(roi)) / float(max(roi.size, 1))


    def _manual_capture_force_single_field_from_enclosing_box(self, source_rect, dedup_rects, dedup_types):
        try:
            source_rect = source_rect if isinstance(source_rect, fitz.Rect) else fitz.Rect(source_rect)
        except Exception:
            return False
        rects = []
        types = []
        try:
            for rr, kind in zip((dedup_rects or []), (dedup_types or [])):
                rects.append(rr if isinstance(rr, fitz.Rect) else fitz.Rect(rr))
                types.append(str(kind or ''))
        except Exception:
            return False
        if len(rects) < 2:
            return False
        if any(t not in {'field', 'line'} for t in types):
            return False

        source_area = float(max(source_rect.width * source_rect.height, 1e-6))
        best_idx = -1
        best_area = 0.0
        for idx, (rr, kind) in enumerate(zip(rects, types)):
            if kind != 'field':
                continue
            area = float(max(rr.width, 0.0) * max(rr.height, 0.0))
            if area > best_area:
                best_idx = idx
                best_area = area
        if best_idx < 0:
            return False

        dominant = rects[best_idx]
        dominant_area = float(max(dominant.width, 0.0) * max(dominant.height, 0.0))
        dominant_cover = dominant_area / float(max(source_area * (ZOOM ** 2), 1e-6))
        if dominant_cover < 0.18:
            return False
        if float(max(dominant.width, 0.0)) / float(max(dominant.height, 1e-6)) < 1.35:
            return False

        enclosed_hits = 0
        inner_line_hits = 0
        other_hits = 0
        for idx, rr in enumerate(rects):
            if idx == best_idx:
                continue
            other_hits += 1
            inter = _rect_intersection_area(dominant, rr)
            rr_area = float(max(rr.width, 0.0) * max(rr.height, 0.0))
            if rr_area <= 0:
                continue
            enclosed_ratio = inter / rr_area
            if enclosed_ratio >= 0.88:
                enclosed_hits += 1
                if types[idx] == 'line':
                    inner_line_hits += 1
                continue
            # Accept tiny border fragments very close to the dominant box top edge.
            x_overlap = _x_intersection(dominant, rr) / float(max(rr.width, 1e-6))
            near_top = rr.y0 <= (dominant.y0 + max(3.0, dominant.height * 0.24))
            if x_overlap >= 0.75 and near_top and rr.height <= max(10.0, dominant.height * 0.22):
                enclosed_hits += 1
                if types[idx] == 'line':
                    inner_line_hits += 1

        if other_hits <= 0:
            return False
        if enclosed_hits < other_hits:
            return False
        if inner_line_hits >= 1:
            return True
        return dominant_cover >= 0.26

    def _manual_capture_prefers_single_field(self, crop, source_rect, dedup_rects, dedup_types, base_bin=None, bin_fields=None):
        try:
            source_rect = source_rect if isinstance(source_rect, fitz.Rect) else fitz.Rect(source_rect)
        except Exception:
            return False
        if crop is None or getattr(crop, 'size', 0) == 0:
            return False
        if source_rect.width <= 2 or source_rect.height <= 2:
            return False
        h_crop, w_crop = crop.shape[:2]
        if w_crop < 24 or h_crop < 16:
            return False
        type_set = {str(t or '') for t in (dedup_types or [])}
        if not type_set:
            return False
        if any(t not in {'line', 'field'} for t in type_set):
            return False
        try:
            source_aspect = float(source_rect.width) / float(max(source_rect.height, 1e-6))
        except Exception:
            source_aspect = 1.0
        if source_aspect < 1.4:
            return False
        if len(dedup_rects or []) > 4:
            return False

        px_rects = []
        try:
            for rr in (dedup_rects or []):
                r = rr if isinstance(rr, fitz.Rect) else fitz.Rect(rr)
                px_rects.append(fitz.Rect(r.x0 - (source_rect.x0 * ZOOM), r.y0 - (source_rect.y0 * ZOOM), r.x1 - (source_rect.x0 * ZOOM), r.y1 - (source_rect.y0 * ZOOM)))
        except Exception:
            px_rects = []
        if not px_rects:
            return False

        union = fitz.Rect(min(r.x0 for r in px_rects), min(r.y0 for r in px_rects), max(r.x1 for r in px_rects), max(r.y1 for r in px_rects))
        union_w_ratio = float(union.width) / float(max(w_crop, 1))
        union_h_ratio = float(union.height) / float(max(h_crop, 1))
        if union_w_ratio < 0.45:
            return False

        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if len(crop.shape) == 3 else crop
        work = base_bin if base_bin is not None else cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        contours, _ = cv2.findContours(work, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        best_cover = 0.0
        best_border_strength = 0.0
        best_extent = 0.0
        for c in contours:
            x, y, w, h = cv2.boundingRect(c)
            if w < max(18, int(w_crop * 0.35)) or h < max(12, int(h_crop * 0.18)):
                continue
            if w > int(w_crop * 0.985) and h > int(h_crop * 0.985):
                continue
            cand = fitz.Rect(x, y, x + w, y + h)
            inter = _rect_intersection_area(cand, union)
            cover = inter / float(max(union.width * union.height, 1e-6))
            if cover <= 0.0:
                continue
            roi = work[y:y+h, x:x+w]
            if roi.size == 0:
                continue
            band = max(1, min(int(min(w, h) * 0.10), 8))
            top = float(cv2.countNonZero(roi[:band, :])) / float(max(roi[:band, :].size, 1))
            bottom = float(cv2.countNonZero(roi[-band:, :])) / float(max(roi[-band:, :].size, 1))
            left = float(cv2.countNonZero(roi[:, :band])) / float(max(roi[:, :band].size, 1))
            right = float(cv2.countNonZero(roi[:, -band:])) / float(max(roi[:, -band:].size, 1))
            border_strength = min(top, bottom, left, right)
            extent = float(cv2.contourArea(c)) / float(max(w * h, 1))
            if cover > best_cover:
                best_cover = cover
                best_border_strength = border_strength
                best_extent = extent

        only_lines = type_set == {'line'}
        mixed_line_field = type_set.issubset({'line', 'field'})
        strong_outer_box = (best_cover >= 0.82 and best_border_strength >= 0.028)
        probable_box = (best_cover >= 0.72 and best_border_strength >= 0.020 and 0.02 <= best_extent <= 0.90)
        nested_line_pattern = (mixed_line_field and union_w_ratio >= 0.55 and union_h_ratio <= 0.60)

        dominant_rect = None
        dominant_kind = ''
        dominant_area = 0.0
        src_area = float(max(w_crop * h_crop, 1))
        for rr, kind in zip(px_rects, (dedup_types or [])):
            try:
                area = float(max(rr.width, 0.0) * max(rr.height, 0.0))
            except Exception:
                area = 0.0
            if area > dominant_area:
                dominant_area = area
                dominant_rect = rr
                dominant_kind = str(kind or '')
        dominant_cover = (dominant_area / src_area) if src_area > 0 else 0.0
        dominant_encloses_others = False
        if dominant_rect is not None and dominant_cover >= 0.34:
            enclosed_hits = 0
            other_hits = 0
            for rr in px_rects:
                if rr is dominant_rect:
                    continue
                other_hits += 1
                inter = _rect_intersection_area(dominant_rect, rr)
                rr_area = float(max(rr.width, 0.0) * max(rr.height, 0.0))
                if rr_area > 0 and (inter / rr_area) >= 0.84:
                    enclosed_hits += 1
            dominant_encloses_others = (other_hits > 0 and enclosed_hits >= other_hits)

        dominant_field_box = (
            dominant_rect is not None and
            dominant_kind == 'field' and
            dominant_cover >= 0.30 and
            float(max(dominant_rect.width, 0.0)) / float(max(dominant_rect.height, 1e-6)) >= 1.4
        )
        stacked_inner_pattern = False
        if dominant_rect is not None and len(px_rects) >= 2:
            inner_top_hits = 0
            for rr in px_rects:
                if rr is dominant_rect:
                    continue
                if rr.y0 <= (dominant_rect.y0 + max(8.0, dominant_rect.height * 0.34)):
                    inner_top_hits += 1
            stacked_inner_pattern = inner_top_hits >= 1

        if strong_outer_box and nested_line_pattern:
            return True
        if only_lines and probable_box and union_h_ratio <= 0.42:
            return True
        if mixed_line_field and probable_box and len(px_rects) <= 3 and source_aspect >= 2.0:
            return True
        if mixed_line_field and dominant_field_box and dominant_encloses_others:
            return True
        if mixed_line_field and dominant_field_box and stacked_inner_pattern and dominant_cover >= 0.38:
            return True
        return False

    def _manual_capture_reject_text_artifact(self, local_rect, kind, text_mask, crop_shape):
        try:
            rr = fitz.Rect(local_rect)
        except Exception:
            return True
        try:
            crop_h = int(crop_shape[0])
            crop_w = int(crop_shape[1])
        except Exception:
            crop_h, crop_w = 0, 0
        if rr.width <= 2 or rr.height <= 2:
            return True
        text_overlap = 0.0
        halo_overlap = 0.0
        try:
            text_overlap = float(self.rect_text_overlap_ratio(rr, text_mask, preview_zoom=1.0) or 0.0)
        except Exception:
            text_overlap = 0.0
        try:
            halo_overlap = float(self._manual_capture_text_overlap_expanded(rr, text_mask, pad_px=max(4, int(min(rr.width, rr.height) * 0.35))) or 0.0)
        except Exception:
            halo_overlap = text_overlap
        width = float(rr.width)
        height = float(rr.height)
        aspect = width / float(max(height, 1e-6))
        area = width * height
        near_top_band = bool(crop_h > 0 and rr.y0 <= max(10.0, crop_h * 0.10))
        # Strong text rejection for tiny boxes embedded in words or labels.
        if kind == 'check':
            if text_overlap >= 0.07 or halo_overlap >= 0.16:
                return True
            if area <= 520 and halo_overlap >= 0.10:
                return True
            if near_top_band and halo_overlap >= 0.06:
                return True
        elif kind == 'field':
            if text_overlap >= 0.10 or halo_overlap >= 0.22:
                return True
            if width <= 26 and height <= 26 and halo_overlap >= 0.08:
                return True
            if width <= 42 and height <= 22 and halo_overlap >= 0.09:
                return True
            if near_top_band and halo_overlap >= 0.07:
                return True
            if aspect <= 2.4 and area <= 900 and halo_overlap >= 0.08:
                return True
        elif kind == 'line':
            if text_overlap >= 0.12 or halo_overlap >= 0.24:
                return True
            if near_top_band and halo_overlap >= 0.10:
                return True
        if self.should_suppress_text_like_rect(rr, text_mask, preview_zoom=1.0, header_cut=int(max(crop_h, 1) * 0.08), overlap_thresh=(0.12 if kind != 'line' else 0.16)):
            return True
        return False

    def _build_manual_area_structural_signature(self, page_idx, rect, out_size=(40, 40), page_img=None):
        try:
            rect = fitz.Rect(rect)
        except Exception:
            return None
        if rect.width <= 1 or rect.height <= 1:
            return None
        img = page_img if page_img is not None else self._render_pdf_page_bgr(self.pdf_path, page_idx=page_idx, preview_zoom=ZOOM)
        if img is None or getattr(img, 'size', 0) == 0:
            return None
        h_img, w_img = img.shape[:2]
        x0 = max(0, min(int(math.floor(rect.x0 * ZOOM)), w_img - 1))
        y0 = max(0, min(int(math.floor(rect.y0 * ZOOM)), h_img - 1))
        x1 = max(x0 + 1, min(int(math.ceil(rect.x1 * ZOOM)), w_img))
        y1 = max(y0 + 1, min(int(math.ceil(rect.y1 * ZOOM)), h_img))
        crop = img[y0:y1, x0:x1].copy()
        if crop.size == 0:
            return None
        text_mask, _ = self.build_text_like_mask(crop, header_band_pct=0.10)
        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
        struct = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        try:
            text_dilate = cv2.dilate(text_mask, np.ones((3, 3), np.uint8), iterations=1)
            struct = cv2.bitwise_and(struct, cv2.bitwise_not(text_dilate))
        except Exception:
            pass
        struct = cv2.morphologyEx(struct, cv2.MORPH_CLOSE, np.ones((2, 2), np.uint8))
        small = cv2.resize(struct, out_size, interpolation=cv2.INTER_AREA)
        small = (small > 48).astype(np.uint8)
        return {
            'mask': small,
            'aspect': float(rect.width) / float(max(rect.height, 1e-6)),
            'density': float(np.mean(small)),
            'pixel_size': (int(crop.shape[1]), int(crop.shape[0])),
        }

    def _prune_area_template_match_redundancy(self, matches, max_matches=18):
        kept = []
        for match in list(matches or []):
            if not isinstance(match, dict):
                continue
            rect = match.get('rect')
            if rect is None:
                continue
            try:
                rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(rect)
            except Exception:
                continue
            resolved_rects = []
            for rr in (match.get('resolved_rects', []) or []):
                try:
                    resolved_rects.append(rr if isinstance(rr, fitz.Rect) else fitz.Rect(rr))
                except Exception:
                    continue
            duplicate = False
            for ex in kept:
                try:
                    ex_rect = ex.get('rect') if isinstance(ex.get('rect'), fitz.Rect) else fitz.Rect(ex.get('rect'))
                except Exception:
                    ex_rect = None
                if ex_rect is not None and self._rect_iou(rect, ex_rect) >= 0.62:
                    duplicate = True
                    break
                ex_resolved = []
                for rr in (ex.get('resolved_rects', []) or []):
                    try:
                        ex_resolved.append(rr if isinstance(rr, fitz.Rect) else fitz.Rect(rr))
                    except Exception:
                        continue
                if resolved_rects and ex_resolved:
                    overlap_hits = 0
                    for a in resolved_rects:
                        best = 0.0
                        for b in ex_resolved:
                            best = max(best, self._rect_iou(a, b))
                        if best >= 0.55:
                            overlap_hits += 1
                    overlap_ratio = overlap_hits / float(max(len(resolved_rects), len(ex_resolved), 1))
                    if overlap_ratio >= 0.60:
                        duplicate = True
                        break
            if not duplicate:
                kept.append(match)
            if len(kept) >= max(1, int(max_matches or 18)):
                break
        return kept


    def _area_template_match_family_key(self, match):
        if not isinstance(match, dict):
            return None
        try:
            rect = match.get('rect')
            rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(rect)
        except Exception:
            return None
        resolved_types = [str(v) for v in (match.get('resolved_types', []) or [])]
        type_counts = {}
        for kind in resolved_types:
            type_counts[kind] = int(type_counts.get(kind, 0) or 0) + 1
        dominant_kind = ''
        if type_counts:
            dominant_kind = sorted(type_counts.items(), key=lambda kv: (-int(kv[1]), str(kv[0])))[0][0]
        elif resolved_types:
            dominant_kind = str(resolved_types[0])
        else:
            dominant_kind = 'field'
        split_count = int(match.get('split_count', len(resolved_types) or 1) or 1)
        if split_count <= 0:
            split_count = max(1, len(resolved_types) or 1)
        try:
            w = float(rect.width)
            h = float(rect.height)
        except Exception:
            w, h = 0.0, 0.0
        aspect = w / float(max(h, 1e-6))
        area = max(1.0, w * h)
        if area < 40.0:
            area_bucket = 'xs'
        elif area < 140.0:
            area_bucket = 's'
        elif area < 320.0:
            area_bucket = 'm'
        elif area < 760.0:
            area_bucket = 'l'
        else:
            area_bucket = 'xl'
        if aspect < 0.85:
            aspect_bucket = 'tall'
        elif aspect <= 1.18:
            aspect_bucket = 'square'
        elif aspect <= 2.35:
            aspect_bucket = 'wide'
        else:
            aspect_bucket = 'long'
        return (str(dominant_kind), int(split_count), str(area_bucket), str(aspect_bucket))

    def _throttle_area_template_matches_by_family(self, matches, max_matches=18, family_soft_cap=2):
        family_soft_cap = max(1, int(family_soft_cap or 2))
        max_matches = max(1, int(max_matches or 18))
        family_counts = {}
        kept = []
        overflow = []
        for item in list(matches or []):
            fam = self._area_template_match_family_key(item)
            key = fam if fam is not None else ('unknown',)
            used = int(family_counts.get(key, 0) or 0)
            if used < family_soft_cap:
                kept.append(item)
                family_counts[key] = used + 1
            else:
                overflow.append(item)
        if len(kept) < max_matches and overflow:
            overflow.sort(key=lambda item: float((item or {}).get('score', 0.0) or 0.0), reverse=True)
            for item in overflow:
                kept.append(item)
                if len(kept) >= max_matches:
                    break
        return kept[:max_matches]

    def find_similar_area_templates_on_page(self, page_idx, template, max_matches=18):
        page_idx = int(page_idx or 0)
        if not isinstance(template, dict):
            return []
        try:
            source_rect = fitz.Rect(template.get('source_rect')) if template.get('source_rect') is not None else None
        except Exception:
            source_rect = None
        if source_rect is None or source_rect.width <= 1 or source_rect.height <= 1:
            return []
        page_cache = self._get_manual_area_search_cache_entry(page_idx)
        if not isinstance(page_cache, dict):
            return []
        page_img = page_cache.get('page_img')
        if page_img is None or getattr(page_img, 'size', 0) == 0:
            return []
        tpl_sig = self._get_cached_manual_area_signature(page_idx, source_rect, page_img=page_img)
        if not tpl_sig:
            return []
        template_shape = dict((template or {}).get('shape_template', {}) or {})
        tw = max(24.0, float(source_rect.width * ZOOM))
        th = max(24.0, float(source_rect.height * ZOOM))
        stats = page_cache.get('stats')
        n = int(page_cache.get('component_count', 0) or 0)
        candidate_rects = list(page_cache.get('candidate_rects') or [])
        shape_direct_candidates = []
        def _add_candidate(rr):
            rr = fitz.Rect(rr)
            if rr.width < max(source_rect.width * 0.55, 6.0) or rr.height < max(source_rect.height * 0.55, 6.0):
                return
            if self._rect_iou(rr, source_rect) >= 0.55:
                return
            for ex in candidate_rects:
                if self._rect_iou(rr, ex) >= 0.65:
                    return
            candidate_rects.append(rr)
        if template_shape:
            try:
                shape_direct_candidates = list(self.extract_shape_candidates_from_page(page_idx, template_shape=template_shape, max_candidates=max(120, int(max_matches or 18) * 12)) or [])
            except Exception:
                shape_direct_candidates = []
        if not candidate_rects:
            for i in range(1, n):
                xx, yy, ww, hh, area = stats[i]
                if area < 18 or ww < 4 or hh < 4:
                    continue
                cx = xx + (ww * 0.5)
                cy = yy + (hh * 0.5)
                rx0 = max(0.0, (cx - (tw * 0.5)) / ZOOM)
                ry0 = max(0.0, (cy - (th * 0.5)) / ZOOM)
                rx1 = min(float(page_img.shape[1]) / ZOOM, (cx + (tw * 0.5)) / ZOOM)
                ry1 = min(float(page_img.shape[0]) / ZOOM, (cy + (th * 0.5)) / ZOOM)
                _add_candidate(fitz.Rect(rx0, ry0, rx1, ry1))
            page_cache['candidate_rects'] = list(candidate_rects)
        matches = []
        existing_boxes = [fitz.Rect(r) for r in (self.all_boxes or [])] if int(getattr(self, 'detected_page_idx', -1) or -1) == page_idx else []
        existing_types = [str(t) for t in (self.box_types or [])] if existing_boxes else []
        template_type_counts = dict((template.get('features', {}) or {}).get('type_counts', {}) or {})
        ranked_candidates = []
        mask_a = tpl_sig.get('mask')
        # contour-first candidates when a shape template exists
        for shape_candidate in (shape_direct_candidates or []):
            try:
                cand_rect = fitz.Rect(shape_candidate.get('bbox_rect')) if shape_candidate.get('bbox_rect') is not None else None
            except Exception:
                cand_rect = None
            if cand_rect is None:
                continue
            cand_sig = self._get_cached_manual_area_signature(page_idx, cand_rect, page_img=page_img)
            if not cand_sig:
                continue
            mask_b = cand_sig.get('mask')
            inter = float(np.logical_and(mask_a > 0, mask_b > 0).sum())
            union = float(np.logical_or(mask_a > 0, mask_b > 0).sum())
            mask_iou = (inter / union) if union > 0 else 0.0
            aspect_sim = max(0.0, 1.0 - min(abs(float(tpl_sig.get('aspect', 1.0)) - float(cand_sig.get('aspect', 1.0))) / 1.1, 1.0))
            density_sim = max(0.0, 1.0 - min(abs(float(tpl_sig.get('density', 0.0)) - float(cand_sig.get('density', 0.0))) / 0.45, 1.0))
            quick_score = (float(shape_candidate.get('score', 0.0) or 0.0) * 0.42) + (mask_iou * 0.40) + (aspect_sim * 0.10) + (density_sim * 0.08)
            if quick_score < 0.34:
                continue
            ranked_candidates.append((quick_score, cand_rect, cand_sig, mask_iou, aspect_sim, density_sim, dict(shape_candidate or {})))
        for cand_rect in candidate_rects:
            cand_sig = self._get_cached_manual_area_signature(page_idx, cand_rect, page_img=page_img)
            if not cand_sig:
                continue
            mask_b = cand_sig.get('mask')
            inter = float(np.logical_and(mask_a > 0, mask_b > 0).sum())
            union = float(np.logical_or(mask_a > 0, mask_b > 0).sum())
            mask_iou = (inter / union) if union > 0 else 0.0
            aspect_sim = max(0.0, 1.0 - min(abs(float(tpl_sig.get('aspect', 1.0)) - float(cand_sig.get('aspect', 1.0))) / 1.1, 1.0))
            density_sim = max(0.0, 1.0 - min(abs(float(tpl_sig.get('density', 0.0)) - float(cand_sig.get('density', 0.0))) / 0.45, 1.0))
            quick_score = (mask_iou * 0.72) + (aspect_sim * 0.16) + (density_sim * 0.12)
            if quick_score < 0.40:
                continue
            ranked_candidates.append((quick_score, cand_rect, cand_sig, mask_iou, aspect_sim, density_sim, None))
        ranked_candidates.sort(key=lambda row: float(row[0]), reverse=True)
        for quick_score, cand_rect, cand_sig, mask_iou, aspect_sim, density_sim, direct_shape in ranked_candidates[:240]:
            analysis = None
            shape_match = 0.0
            if template_shape:
                candidate_shapes = [direct_shape] if isinstance(direct_shape, dict) and direct_shape else []
                if not candidate_shapes:
                    candidate_shapes = self.extract_shape_candidates_from_roi(page_idx, cand_rect, max_candidates=6)
                best_shape = None
                best_shape_score = 0.0
                for shape_candidate in (candidate_shapes or []):
                    sim = self._shape_similarity_score(template_shape, shape_candidate)
                    if sim > best_shape_score:
                        best_shape_score = sim
                        best_shape = shape_candidate
                if best_shape is not None and best_shape_score >= 0.42:
                    analysis = self.build_shape_template_analysis(page_idx, cand_rect, best_shape)
                    shape_match = float(best_shape_score)
            if analysis is None:
                analysis = self.analyze_manual_area_template(page_idx, cand_rect)
            if not analysis:
                continue
            cand_type_counts = dict((analysis.get('features', {}) or {}).get('type_counts', {}) or {})
            type_overlap = 0.0
            if template_type_counts and cand_type_counts:
                shared = 0
                total = max(sum(template_type_counts.values()), sum(cand_type_counts.values()), 1)
                for key, val in template_type_counts.items():
                    shared += min(int(val or 0), int(cand_type_counts.get(key, 0) or 0))
                type_overlap = float(shared) / float(total)
            existing_cover = 0.0
            if existing_boxes:
                cover_hits = 0
                for ex_rect, ex_kind in zip(existing_boxes, existing_types):
                    if self._rect_iou(cand_rect, ex_rect) >= 0.52:
                        cover_hits += 1
                existing_cover = min(1.0, cover_hits / float(max(len(analysis.get('resolved_rects', []) or []), 1)))
            if template_shape:
                score = (shape_match * 0.56) + (aspect_sim * 0.10) + (density_sim * 0.06) + (type_overlap * 0.12) + (mask_iou * 0.24) - (existing_cover * 0.08)
            else:
                score = (mask_iou * 0.52) + (aspect_sim * 0.14) + (density_sim * 0.12) + (type_overlap * 0.16) - (existing_cover * 0.10)
            score = max(0.0, min(0.98, score))
            if existing_cover >= 0.95:
                continue
            if score < (0.42 if template_shape else 0.46):
                continue
            reason = f'shape {shape_match:.2f} • structure {mask_iou:.2f} • type {type_overlap:.2f} • density {density_sim:.2f} • existing {existing_cover:.2f}' if template_shape else f'structure {mask_iou:.2f} • type {type_overlap:.2f} • density {density_sim:.2f} • existing {existing_cover:.2f}'
            matches.append({
                'match_id': str(uuid.uuid4().hex[:12]),
                'template_id': str(template.get('template_id', '') or ''),
                'page': page_idx,
                'rect': fitz.Rect(cand_rect),
                'score': float(round(score, 6)),
                'reason': reason,
                'resolved_rects': [fitz.Rect(r) for r in (analysis.get('resolved_rects', []) or [])],
                'resolved_types': [str(v) for v in (analysis.get('resolved_types', []) or [])],
                'split_count': int(analysis.get('split_count', 0) or 0),
            })
        matches.sort(key=lambda item: float(item.get('score', 0.0) or 0.0), reverse=True)
        kept = self._prune_area_template_match_redundancy(matches, max_matches=max_matches * 3)
        kept = self._throttle_area_template_matches_by_family(kept, max_matches=max_matches, family_soft_cap=2)
        return kept

    def find_similar_area_templates_in_form(self, template, max_total=36, per_page_max=10):
        if not isinstance(template, dict):
            return []
        try:
            total_pages = int(self.total_pages() or 0)
        except Exception:
            total_pages = 0
        if total_pages <= 0:
            return []
        all_matches = []
        for page_idx in range(total_pages):
            page_matches = self.find_similar_area_templates_on_page(page_idx, template, max_matches=per_page_max)
            for item in (page_matches or []):
                if not isinstance(item, dict):
                    continue
                payload = dict(item)
                payload['page'] = int(page_idx)
                reason = str(payload.get('reason', '') or '')
                payload['reason'] = f"page {page_idx} • {reason}" if reason else f"page {page_idx}"
                all_matches.append(payload)
        all_matches.sort(key=lambda item: float(item.get('score', 0.0) or 0.0), reverse=True)
        all_matches = self._throttle_area_template_matches_by_family(all_matches, max_matches=max_total * 2, family_soft_cap=3)
        kept = []
        per_page_seen = {}
        for item in all_matches:
            page_idx = int(item.get('page', 0) or 0)
            if per_page_seen.get(page_idx, 0) >= max(1, int(per_page_max or 10)):
                continue
            kept.append(item)
            per_page_seen[page_idx] = int(per_page_seen.get(page_idx, 0) or 0) + 1
            if len(kept) >= max(1, int(max_total or 36)):
                break
        return kept

    def merge_area_template_matches_into_detections(self, page_idx, matches):
        page_idx = int(page_idx or 0)
        rect_type_pairs = []
        for item in (matches or []):
            if not isinstance(item, dict):
                continue
            rects = [fitz.Rect(r) for r in (item.get('resolved_rects', []) or [])]
            types = [str(v) for v in (item.get('resolved_types', []) or [])]
            for rect, kind in zip(rects, types or ['field'] * len(rects)):
                rect_type_pairs.append((rect, kind))
        return self._merge_manual_rect_type_pairs(page_idx, rect_type_pairs)

    def auto_apply_saved_area_templates_after_detection(self, page_idx, max_matches_per_template=18, min_score=0.58):
        page_idx = int(page_idx or 0)
        templates = list((self.area_templates_by_page or {}).get(page_idx, []) or [])
        if not templates:
            return {'template_count': 0, 'match_count': 0, 'added_count': 0, 'selected_ids': []}
        has_live = bool(getattr(self, 'all_boxes', [])) and int(getattr(self, 'detected_page_idx', -1) or -1) == page_idx
        if not has_live:
            restored = bool(self._restore_detection_for_page(page_idx, required_signature=self._settings_signature()))
            if not restored:
                return {'template_count': len(templates), 'match_count': 0, 'added_count': 0, 'selected_ids': []}
        selected_ids = []
        total_matches = 0
        total_added = 0
        used_template_ids = set()
        for template in templates:
            if not isinstance(template, dict):
                continue
            template_id = str(template.get('template_id', '') or '')
            matches = self.find_similar_area_templates_on_page(page_idx, template, max_matches=max_matches_per_template)
            strong_matches = [m for m in (matches or []) if float((m or {}).get('score', 0.0) or 0.0) >= float(min_score)]
            if not strong_matches:
                continue
            total_matches += len(strong_matches)
            merge_result = self.merge_area_template_matches_into_detections(page_idx, strong_matches) or {}
            page_selected = list(merge_result.get('selected_ids', []) or [])
            if not page_selected:
                page_selected = []
                for rect_ref, kind in (merge_result.get('added_refs', []) or []):
                    best_idx = -1
                    best_score = 0.0
                    for idx_existing, existing in enumerate(self.all_boxes):
                        existing_kind = str(self.box_types[idx_existing] if idx_existing < len(self.box_types) else '')
                        if kind and existing_kind != str(kind):
                            continue
                        score = self._mobile_repair_overlap_score(rect_ref, existing)
                        if score > best_score:
                            best_idx = idx_existing
                            best_score = score
                    if best_idx >= 0 and best_score >= 0.45:
                        page_selected.append(best_idx)
            added_now = len(list(merge_result.get('added_refs', []) or []))
            total_added += int(added_now)
            selected_ids.extend(int(x) for x in page_selected if isinstance(x, int) or str(x).isdigit())
            if template_id:
                used_template_ids.add(template_id)
        if total_added > 0:
            self.detected_page_idx = int(page_idx)
            self.detected_settings_signature = self._settings_signature()
            self._cache_detection_for_page(page_idx)
        for template_id in used_template_ids:
            try:
                self.invalidate_area_template_match_cache(page_idx=page_idx, template_id=template_id)
            except Exception:
                pass
        return {
            'template_count': len(templates),
            'match_count': int(total_matches),
            'added_count': int(total_added),
            'selected_ids': sorted(set(int(x) for x in selected_ids if isinstance(x, int) or str(x).isdigit())),
        }


    def _serialize_repair_patches(self):
        payload = {}
        for page_key, items in (getattr(self, 'repair_patches_by_page', {}) or {}).items():
            serial = []
            for item in (items or []):
                if not isinstance(item, dict):
                    continue
                serial.append({
                    'patch_id': str(item.get('patch_id', '') or ''),
                    'page': int(item.get('page', page_key) or page_key),
                    'anchor_rects': self._serialize_rects(item.get('anchor_rects', []) or []),
                    'zone_rect': self._serialize_rect(item.get('zone_rect')) if item.get('zone_rect') is not None else None,
                    'add_rects': self._serialize_rects(item.get('add_rects', []) or []),
                    'add_types': [str(v) for v in (item.get('add_types', []) or [])],
                    'remove_rects': self._serialize_rects(item.get('remove_rects', []) or []),
                    'remove_types': [str(v) for v in (item.get('remove_types', []) or [])],
                    'created_at': str(item.get('created_at', '') or ''),
                })
            if serial:
                payload[str(int(page_key))] = serial
        return payload

    def _deserialize_repair_patches(self, payload):
        self.repair_patches_by_page = {}
        self.repair_patch_index = {}
        if not isinstance(payload, dict):
            return
        for page_key, items in payload.items():
            try:
                page_idx = int(page_key)
            except Exception:
                continue
            page_items = []
            for item in (items or []):
                if not isinstance(item, dict):
                    continue
                try:
                    zone_rect = fitz.Rect(item.get('zone_rect')) if item.get('zone_rect') is not None else None
                except Exception:
                    zone_rect = None
                anchor_rects, add_rects, remove_rects = [], [], []
                for rect in (item.get('anchor_rects', []) or []):
                    try: anchor_rects.append(fitz.Rect(rect))
                    except Exception: pass
                for rect in (item.get('add_rects', []) or []):
                    try: add_rects.append(fitz.Rect(rect))
                    except Exception: pass
                for rect in (item.get('remove_rects', []) or []):
                    try: remove_rects.append(fitz.Rect(rect))
                    except Exception: pass
                patch = {
                    'patch_id': str(item.get('patch_id', '') or uuid.uuid4().hex[:12]),
                    'page': page_idx,
                    'anchor_rects': anchor_rects,
                    'zone_rect': zone_rect,
                    'add_rects': add_rects,
                    'add_types': [str(v) for v in (item.get('add_types', []) or [])],
                    'remove_rects': remove_rects,
                    'remove_types': [str(v) for v in (item.get('remove_types', []) or [])],
                    'created_at': str(item.get('created_at', '') or _utc_now_iso()),
                }
                page_items.append(patch)
                self.repair_patch_index[patch['patch_id']] = patch
            if page_items:
                self.repair_patches_by_page[page_idx] = page_items

    def save_repair_patch(self, page_idx, anchor_rects=None, zone_rect=None, add_rects=None, add_types=None, remove_rects=None, remove_types=None):
        page_idx = int(page_idx or 0)
        patch = {
            'patch_id': str(uuid.uuid4().hex[:12]),
            'page': page_idx,
            'anchor_rects': [fitz.Rect(r) for r in (anchor_rects or [])],
            'zone_rect': (fitz.Rect(zone_rect) if zone_rect is not None else None),
            'add_rects': [fitz.Rect(r) for r in (add_rects or [])],
            'add_types': [str(v) for v in (add_types or [])],
            'remove_rects': [fitz.Rect(r) for r in (remove_rects or [])],
            'remove_types': [str(v) for v in (remove_types or [])],
            'created_at': _utc_now_iso(),
        }
        self.repair_patches_by_page.setdefault(page_idx, []).append(patch)
        self.repair_patch_index[patch['patch_id']] = patch
        return patch

    def auto_apply_saved_repair_patches_after_detection(self, page_idx, min_anchor_overlap=0.40):
        page_idx = int(page_idx or 0)
        patches = list((getattr(self, 'repair_patches_by_page', {}) or {}).get(page_idx, []) or [])
        if not patches:
            return {'patch_count': 0, 'applied_count': 0, 'added_count': 0, 'removed_count': 0}
        added_count = 0
        removed_count = 0
        applied_count = 0
        if not (bool(getattr(self, 'all_boxes', []) or []) and int(getattr(self, 'detected_page_idx', -1) or -1) == page_idx):
            restored = bool(self._restore_detection_for_page(page_idx, required_signature=self._settings_signature()))
            if not restored:
                return {'patch_count': len(patches), 'applied_count': 0, 'added_count': 0, 'removed_count': 0}
        for patch in patches:
            if not isinstance(patch, dict):
                continue
            anchor_rects = [fitz.Rect(r) for r in (patch.get('anchor_rects', []) or [])]
            if anchor_rects:
                hit_count = 0
                for ar in anchor_rects:
                    best = 0.0
                    for existing in (self.all_boxes or []):
                        try:
                            best = max(best, float(self._mobile_repair_overlap_score(ar, existing)))
                        except Exception:
                            best = max(best, float(self._rect_iou(ar, existing)))
                    if best >= float(min_anchor_overlap):
                        hit_count += 1
                if hit_count < max(1, int(math.ceil(len(anchor_rects) * 0.5))):
                    continue
            local_added = 0
            local_removed = 0
            add_rects = [fitz.Rect(r) for r in (patch.get('add_rects', []) or [])]
            add_types = [str(v) for v in (patch.get('add_types', []) or [])]
            merge_result = self._merge_manual_rect_type_pairs(page_idx, list(zip(add_rects, add_types or ['field'] * len(add_rects)))) if add_rects else {'added_refs': []}
            local_added += len(list((merge_result or {}).get('added_refs', []) or []))
            rem_rects = [fitz.Rect(r) for r in (patch.get('remove_rects', []) or [])]
            rem_types = [str(v) for v in (patch.get('remove_types', []) or [])]
            if rem_rects:
                keep_boxes = []
                keep_types = []
                for idx_existing, existing in enumerate(list(self.all_boxes or [])):
                    ex_kind = str(self.box_types[idx_existing] if idx_existing < len(self.box_types) else '')
                    drop = False
                    for rr, rk in zip(rem_rects, rem_types or ['field'] * len(rem_rects)):
                        if rk and ex_kind != str(rk):
                            continue
                        score = 0.0
                        try:
                            score = float(self._mobile_repair_overlap_score(rr, existing))
                        except Exception:
                            score = float(self._rect_iou(rr, existing))
                        if score >= 0.58:
                            drop = True
                            break
                    if drop:
                        local_removed += 1
                    else:
                        keep_boxes.append(existing)
                        keep_types.append(ex_kind)
                if local_removed > 0:
                    self.all_boxes = keep_boxes
                    self.box_types = keep_types
            if local_added > 0 or local_removed > 0:
                applied_count += 1
                added_count += int(local_added)
                removed_count += int(local_removed)
                self.detected_page_idx = int(page_idx)
                self.detected_settings_signature = self._settings_signature()
                self._cache_detection_for_page(page_idx)
        return {'patch_count': len(patches), 'applied_count': int(applied_count), 'added_count': int(added_count), 'removed_count': int(removed_count)}


    def remember_sticky_detections(self, page_idx, rect_type_pairs):
        page_idx = int(page_idx or 0)
        bucket = list((getattr(self, 'sticky_detections_by_page', {}) or {}).get(page_idx, []) or [])
        for rect, kind in list(rect_type_pairs or []):
            try:
                rr = rect if isinstance(rect, fitz.Rect) else fitz.Rect(rect)
            except Exception:
                continue
            kk = str(kind or 'field') or 'field'
            duplicate = False
            for ex_rect, ex_kind in bucket:
                try:
                    ex_rr = ex_rect if isinstance(ex_rect, fitz.Rect) else fitz.Rect(ex_rect)
                except Exception:
                    continue
                if str(ex_kind or '') != kk:
                    continue
                try:
                    if self._rect_iou(rr, ex_rr) >= 0.78 or self._rects_refer_to_same_target(rr, ex_rr, tol=0.35):
                        duplicate = True
                        break
                except Exception:
                    pass
            if not duplicate:
                bucket.append((fitz.Rect(rr), kk))
        self.sticky_detections_by_page[page_idx] = bucket
        return list(bucket)

    def auto_apply_sticky_detections_after_detection(self, page_idx, min_score=0.45):
        page_idx = int(page_idx or 0)
        sticky_pairs = list((getattr(self, 'sticky_detections_by_page', {}) or {}).get(page_idx, []) or [])
        if not sticky_pairs:
            return {'sticky_count': 0, 'added_count': 0, 'selected_ids': []}
        has_live = bool(getattr(self, 'all_boxes', []) or []) and int(getattr(self, 'detected_page_idx', -1) or -1) == page_idx
        if (not has_live) and (not bool(self._restore_detection_for_page(page_idx, required_signature=self._settings_signature()))):
            return {'sticky_count': len(sticky_pairs), 'added_count': 0, 'selected_ids': []}
        merge_result = self._merge_manual_rect_type_pairs(page_idx, sticky_pairs) or {}
        selected_ids = list(merge_result.get('selected_ids', []) or [])
        if not selected_ids:
            used = set()
            for rect_ref, kind in list(merge_result.get('added_refs', []) or []):
                best_idx = -1
                best_score = 0.0
                for idx_existing, existing in enumerate(list(self.all_boxes or [])):
                    if idx_existing in used:
                        continue
                    existing_kind = str(self.box_types[idx_existing] if idx_existing < len(self.box_types) else '')
                    if kind and existing_kind != str(kind):
                        continue
                    try:
                        score = float(self._mobile_repair_overlap_score(rect_ref, existing))
                    except Exception:
                        score = float(self._rect_iou(rect_ref, existing))
                    if score > best_score:
                        best_idx = idx_existing
                        best_score = score
                if best_idx >= 0 and best_score >= float(min_score):
                    used.add(best_idx)
                    selected_ids.append(best_idx)
        return {
            'sticky_count': len(sticky_pairs),
            'added_count': len(list(merge_result.get('added_refs', []) or [])),
            'selected_ids': sorted(set(int(x) for x in selected_ids if isinstance(x, int) or str(x).isdigit())),
        }

    def get_candidate_learning_priors(self, page_idx=0, max_events=600):
        try:
            page_idx = int(page_idx or 0)
        except Exception:
            page_idx = 0
        current_fp = ""
        try:
            current_fp = str(((self.current_detection_context or {}).get('fingerprint') or {}).get('fingerprint_key', '') or '')
        except Exception:
            current_fp = ""
        if not current_fp and self.pdf_path:
            try:
                current_fp = str((self.build_page_fingerprint(page_idx=page_idx) or {}).get('fingerprint_key', '') or '')
            except Exception:
                current_fp = ""

        events = list(((self.correction_log or {}).get('events', []) or []))[-max_events:]
        same_fp = {}
        global_stats = {}
        used = 0
        for event in events:
            if not isinstance(event, dict):
                continue
            if str(event.get('event_type', '') or '') != 'candidate_review_item':
                continue
            kind = str(event.get('kind', '') or '').strip().lower()
            source = str(event.get('source', '') or '').strip().lower()
            if not kind or not source:
                continue
            key = (kind, source)
            accepted = bool(event.get('accepted', False))
            bucket = global_stats.setdefault(key, {'accepted': 0, 'rejected': 0})
            bucket['accepted' if accepted else 'rejected'] += 1
            event_fp = str(event.get('fingerprint_key', '') or '')
            if current_fp and event_fp and event_fp == current_fp:
                bucket_fp = same_fp.setdefault(key, {'accepted': 0, 'rejected': 0})
                bucket_fp['accepted' if accepted else 'rejected'] += 1
                used += 1

        priors = {}
        for key in set(list(global_stats.keys()) + list(same_fp.keys())):
            g = global_stats.get(key, {})
            s = same_fp.get(key, {})
            g_total = int(g.get('accepted', 0) or 0) + int(g.get('rejected', 0) or 0)
            s_total = int(s.get('accepted', 0) or 0) + int(s.get('rejected', 0) or 0)
            g_balance = 0.0 if g_total <= 0 else (float(int(g.get('accepted', 0) or 0) - int(g.get('rejected', 0) or 0)) / float(g_total))
            s_balance = 0.0 if s_total <= 0 else (float(int(s.get('accepted', 0) or 0) - int(s.get('rejected', 0) or 0)) / float(s_total))
            adjustment = max(-0.18, min(0.18, (s_balance * 0.14) + (g_balance * 0.08)))
            priors[key] = {
                'adjustment': round(float(adjustment), 4),
                'same_fp_total': int(s_total),
                'global_total': int(g_total),
            }
        return {
            'fingerprint_key': current_fp,
            'pairs': priors,
            'same_fp_events': int(used),
        }

    def candidate_learning_adjustment(self, kind, source, page_idx=0, priors=None):
        kind = str(kind or '').strip().lower()
        source = str(source or '').strip().lower()
        if not kind or not source:
            return 0.0
        if not isinstance(priors, dict):
            priors = self.get_candidate_learning_priors(page_idx=page_idx)
        pair = dict((priors or {}).get('pairs', {}).get((kind, source), {}) or {})
        try:
            return float(pair.get('adjustment', 0.0) or 0.0)
        except Exception:
            return 0.0

    def _empty_geom(self):
        return {"names": [], "dob": [], "phil": []}

    def _empty_semantic_targets(self):
        return {key: [] for key in SEMANTIC_TARGET_KEYS}

    def _normalize_semantic_targets(self, targets=None):
        normalized = self._empty_semantic_targets()
        src = targets or {}
        for key in SEMANTIC_TARGET_KEYS:
            rects = []
            for rect in (src.get(key, []) or []):
                try:
                    rects.append(fitz.Rect(rect))
                except Exception:
                    try:
                        rects.append(fitz.Rect(*rect))
                    except Exception:
                        continue
            normalized[key] = rects
        return normalized

    def _semantic_targets_from_geom(self, geom=None):
        geom = geom or self._empty_geom()
        targets = self._empty_semantic_targets()
        for legacy_key, semantic_keys in LEGACY_GEOM_TO_SEMANTIC.items():
            rects = [fitz.Rect(r) for r in (geom.get(legacy_key, []) or [])]
            for semantic_key, rect in zip(semantic_keys, rects):
                targets[semantic_key].append(rect)
        return targets

    def _legacy_geom_from_semantic_targets(self, targets=None):
        targets = self._normalize_semantic_targets(targets)
        return {
            "names": [fitz.Rect(r) for key in NAME_TARGET_KEYS for r in targets.get(key, [])],
            "dob": [fitz.Rect(r) for key in DOB_TARGET_KEYS for r in targets.get(key, [])],
            "phil": [fitz.Rect(r) for r in targets.get("record_identity.membership_number", [])],
        }

    def _set_semantic_targets(self, targets=None):
        self.semantic_targets = self._normalize_semantic_targets(targets)
        self.geom = self._legacy_geom_from_semantic_targets(self.semantic_targets)

    def _normalize_semantic_target_overrides(self, overrides=None):
        normalized = {}
        if not isinstance(overrides, dict):
            return normalized
        for page_key, target_map in overrides.items():
            try:
                page_idx = int(page_key)
            except Exception:
                continue
            if not isinstance(target_map, dict):
                continue
            page_targets = self._empty_semantic_targets()
            for key in SEMANTIC_TARGET_KEYS:
                rects = []
                for rect in (target_map.get(key, []) or []):
                    try:
                        rects.append(fitz.Rect(rect))
                    except Exception:
                        continue
                page_targets[key] = rects
            if any(page_targets.values()):
                normalized[page_idx] = page_targets
        return normalized

    def _apply_semantic_target_overrides(self, page_idx, targets=None):
        merged = self._normalize_semantic_targets(targets)
        page_targets = self.semantic_target_overrides.get(int(page_idx), {})
        if isinstance(page_targets, dict):
            for key in SEMANTIC_TARGET_KEYS:
                rects = [fitz.Rect(r) for r in (page_targets.get(key, []) or [])]
                if rects:
                    merged[key] = rects
        return self._normalize_semantic_targets(merged)

    def _current_semantic_targets_for_page(self, page_idx):
        if not SEMANTIC_TARGETS_ENABLED:
            return self._empty_semantic_targets()
        page_idx = int(page_idx or 0)
        if self.detected_page_idx == page_idx and any(self.semantic_targets.values()):
            base = self.semantic_targets
        else:
            base = self.semantic_targets_by_page.get(page_idx, self._empty_semantic_targets())
        return self._apply_semantic_target_overrides(page_idx, base)

    def get_box_semantic_target_payload(self, box_id, page_idx):
        if (not SEMANTIC_TARGETS_ENABLED) or box_id < 0 or box_id >= len(self.all_boxes):
            return {"box_id": box_id, "page": page_idx, "status": "EMPTY", "target_key": ""}
        target_rect = self.all_boxes[box_id]
        best_key = ""
        best_score = -1.0
        semantic_src = self._current_semantic_targets_for_page(page_idx)
        for key in SEMANTIC_TARGET_KEYS:
            for rect in (semantic_src.get(key, []) or []):
                score = self._mapping_match_score(target_rect, rect)
                if score > best_score:
                    best_score = score
                    best_key = key
        if best_key and best_score >= 1.0:
            return {
                "box_id": box_id,
                "page": page_idx,
                "status": "TARGETED",
                "target_key": best_key,
                "target_label": semantic_target_label(best_key, best_key),
            }
        return {"box_id": box_id, "page": page_idx, "status": "EMPTY", "target_key": ""}

    def assign_semantic_target(self, box_ids, target_key, page_idx):
        if not SEMANTIC_TARGETS_ENABLED:
            return False
        target_key = str(target_key or "").strip()
        if target_key not in SEMANTIC_TARGET_KEYS:
            raise ValueError("Please choose a valid special form target.")
        page_idx = int(page_idx or 0)
        selected_rects = []
        for b_id in box_ids:
            if b_id < 0 or b_id >= len(self.all_boxes):
                raise ValueError(f"Box ID {b_id} is out of range.")
            selected_rects.append(self.all_boxes[b_id])
        selected_rects = sorted(selected_rects, key=lambda rr: (round(rr.y0, 3), rr.x0))
        page_targets = self._normalize_semantic_targets(self.semantic_target_overrides.get(page_idx, {}))
        for key in SEMANTIC_TARGET_KEYS:
            kept = []
            for existing_rect in (page_targets.get(key, []) or []):
                if any(self._rects_refer_to_same_target(existing_rect, sr, tol=0.20) for sr in selected_rects):
                    continue
                kept.append(existing_rect)
            page_targets[key] = kept
        page_targets[target_key] = [fitz.Rect(r) for r in selected_rects]
        self.semantic_target_overrides[page_idx] = page_targets
        merged_targets = self._apply_semantic_target_overrides(page_idx, self._current_semantic_targets_for_page(page_idx))
        self.semantic_targets_by_page[page_idx] = self._normalize_semantic_targets(merged_targets)
        self.geom_by_page[page_idx] = self._legacy_geom_from_semantic_targets(self.semantic_targets_by_page[page_idx])
        if self.detected_page_idx == page_idx:
            self._set_semantic_targets(merged_targets)

    def clear_semantic_target_for_box_ids(self, box_ids, page_idx, target_key=None):
        if not SEMANTIC_TARGETS_ENABLED:
            return False
        page_idx = int(page_idx or 0)
        selected_rects = []
        for b_id in box_ids:
            if b_id < 0 or b_id >= len(self.all_boxes):
                continue
            selected_rects.append(self.all_boxes[b_id])
        if not selected_rects:
            return
        page_targets = self._normalize_semantic_targets(self.semantic_target_overrides.get(page_idx, {}))
        keys = [str(target_key).strip()] if str(target_key or "").strip() in SEMANTIC_TARGET_KEYS else list(SEMANTIC_TARGET_KEYS)
        changed = False
        for key in keys:
            kept = []
            for existing_rect in (page_targets.get(key, []) or []):
                if any(self._rects_refer_to_same_target(existing_rect, sr, tol=0.20) for sr in selected_rects):
                    changed = True
                    continue
                kept.append(existing_rect)
            page_targets[key] = kept
        if any(page_targets.values()):
            self.semantic_target_overrides[page_idx] = page_targets
        else:
            self.semantic_target_overrides.pop(page_idx, None)
        merged_targets = self._apply_semantic_target_overrides(page_idx, self.semantic_targets_by_page.get(page_idx, self._empty_semantic_targets()))
        self.semantic_targets_by_page[page_idx] = self._normalize_semantic_targets(merged_targets)
        self.geom_by_page[page_idx] = self._legacy_geom_from_semantic_targets(self.semantic_targets_by_page[page_idx])
        if self.detected_page_idx == page_idx and changed:
            self._set_semantic_targets(merged_targets)

    def _semantic_rects(self, *keys):
        if not SEMANTIC_TARGETS_ENABLED:
            return []
        out = []
        for key in keys:
            out.extend([fitz.Rect(r) for r in self.semantic_targets.get(key, [])])
        return out

    # --------------------------------------------------------
    # Data loading
    # --------------------------------------------------------
    def load_dataframe(self, path_or_url):
        src = str(path_or_url).strip()
    
        if not src:
            raise ValueError("No data source provided.")
    
        # ---------------------------------------------
        # Google Sheet URL support (CSV export endpoint)
        # ---------------------------------------------
        if src.startswith("http://") or src.startswith("https://"):
            if "docs.google.com/spreadsheets" in src:
                try:
                    df = load_google_sheet_dataframe(src)
                except Exception as e:
                    raise ValueError(
                        "Failed to load Google Sheet from URL. "
                        "Make sure the sheet/tab is accessible. "
                        f"Details: {e}"
                    )
            else:
                raise ValueError("Only Google Sheets URLs are supported for URL loading.")
    
        # ---------------------------------------------
        # Local file support
        # ---------------------------------------------
        else:
            ext = os.path.splitext(src)[1].lower()
            if ext == ".csv":
                df = pd.read_csv(src, dtype=str).fillna("")
            elif ext in [".xlsx", ".xls"]:
                df = pd.read_excel(src, dtype=str).fillna("")
            else:
                raise ValueError("Unsupported file type. Use CSV, XLSX, XLS, or a Google Sheet URL.")
    
        self.df = df
        self.first_col = find_col_safe(df, "first", "name")
        self.last_col = find_col_safe(df, "surname") or find_col_safe(df, "last", "name")
        self.mid_col = find_col_safe(df, "middle", "name")
        self.suf_col = find_col_safe(df, "suffix")
        self.dob_col = find_col_safe(df, "birth", "date") or find_col_safe(df, "birthdate")
        self.phil_col = find_col_safe(df, "philhealth")
    
        self.rebuild_record_labels(source_column=self.record_label_source_column)

    def has_name_identity_columns(self):
        return bool(str(self.first_col or "").strip() or str(self.last_col or "").strip())

    def rebuild_record_labels(self, source_column=None):
        if self.df is None or self.df.empty:
            self.patient_names = []
            self.record_label_source_column = ""
            return

        df = self.df
        chosen_source = str(self.record_label_source_column if source_column is None else source_column or "").strip()
        if chosen_source and chosen_source not in df.columns:
            chosen_source = ""

        self.record_label_source_column = chosen_source
        f_col_safe = self.first_col if self.first_col else (df.columns[0] if len(df.columns) else None)
        l_col_safe = self.last_col if self.last_col else (df.columns[0] if len(df.columns) else None)

        def _clean_piece(value):
            value = str(value or "").strip()
            return re.sub(r"\s+", " ", value)

        base_labels = []
        dob_values = []
        for row_idx, (_, row) in enumerate(df.iterrows(), start=1):
            first_val = _clean_piece(row.get(f_col_safe, "")) if f_col_safe else ""
            last_val = _clean_piece(row.get(l_col_safe, "")) if l_col_safe else ""
            dob_val = _clean_piece(row.get(self.dob_col, "")) if self.dob_col else ""

            if chosen_source:
                base_label = _clean_piece(row.get(chosen_source, ""))
                if not base_label:
                    base_label = f"Record {row_idx}"
            else:
                if last_val and first_val:
                    base_label = f"{last_val}, {first_val}"
                else:
                    base_label = first_val or last_val or f"Record {row_idx}"

            base_labels.append(base_label)
            dob_values.append(dob_val)

        label_counts = pd.Series(base_labels, dtype=str).value_counts().to_dict() if base_labels else {}
        final_labels = []
        for row_idx, (base_label, dob_val) in enumerate(zip(base_labels, dob_values), start=1):
            extras = []
            if label_counts.get(base_label, 0) > 1 and dob_val:
                extras.append(dob_val)
            if label_counts.get(base_label, 0) > 1:
                extras.append(f"Row {row_idx}")
            final_labels.append(base_label + ((" — " + " • ".join(extras)) if extras else ""))

        self.df["_ROW_NUMBER"] = pd.Series(list(range(1, len(df) + 1)), index=df.index, dtype="int64")
        self.df["_RECORD_KEY"] = pd.Series([f"row:{i}" for i in range(1, len(df) + 1)], index=df.index, dtype=str)
        self.df["_RECORD_LABEL"] = pd.Series(final_labels, index=df.index, dtype=str)
        self.df["_DISPLAY_NAME"] = self.df["_RECORD_LABEL"]
        self.patient_names = sorted([
            str(x).strip()
            for x in self.df["_RECORD_LABEL"].dropna().tolist()
            if str(x).strip()
        ])

    def needs_record_label_source_fallback(self):
        return bool(self.df is not None and not self.df.empty and not self.has_name_identity_columns() and not str(self.record_label_source_column or "").strip())

    def _record_label_column(self):
        if self.df is None or self.df.empty:
            return "_DISPLAY_NAME"
        if "_RECORD_LABEL" in self.df.columns:
            return "_RECORD_LABEL"
        return "_DISPLAY_NAME"

    def _record_key_column(self):
        if self.df is None or self.df.empty:
            return ""
        return "_RECORD_KEY" if "_RECORD_KEY" in self.df.columns else ""

    def resolve_record_key(self, record_label=None):
        if self.df is None or self.df.empty:
            return ""
        label = str(record_label or "").strip()
        if not label:
            return ""
        label_col = self._record_label_column()
        matches = self.df[self.df[label_col].astype(str) == label]
        if matches.empty and "_DISPLAY_NAME" in self.df.columns and label_col != "_DISPLAY_NAME":
            matches = self.df[self.df["_DISPLAY_NAME"].astype(str) == label]
        if matches.empty:
            return ""
        key_col = self._record_key_column()
        if key_col and key_col in matches.columns:
            return str(matches.iloc[0].get(key_col, "") or "")
        return ""

    def resolve_record_label(self, record_key=None, fallback_label=""):
        if self.df is None or self.df.empty:
            return str(fallback_label or "").strip()
        key = str(record_key or "").strip()
        label_col = self._record_label_column()
        if key:
            key_col = self._record_key_column()
            if key_col and key_col in self.df.columns:
                matches = self.df[self.df[key_col].astype(str) == key]
                if not matches.empty:
                    return str(matches.iloc[0].get(label_col, matches.iloc[0].get("_DISPLAY_NAME", fallback_label)) or "").strip()
        fallback = str(fallback_label or "").strip()
        if fallback:
            matches = self.df[self.df[label_col].astype(str) == fallback]
            if matches.empty and "_DISPLAY_NAME" in self.df.columns and label_col != "_DISPLAY_NAME":
                matches = self.df[self.df["_DISPLAY_NAME"].astype(str) == fallback]
            if not matches.empty:
                return str(matches.iloc[0].get(label_col, fallback) or "").strip()
        return fallback

    def _get_record_row(self, record_ref=None, record_key=None):
        if self.df is None or self.df.empty:
            raise ValueError("DataFrame is empty.")

        key = str(record_key or "").strip()
        if key:
            key_col = self._record_key_column()
            if key_col and key_col in self.df.columns:
                matches = self.df[self.df[key_col].astype(str) == key]
                if not matches.empty:
                    if len(matches) > 1:
                        raise ValueError(f"Multiple records found for key '{key}'.")
                    return matches.iloc[0]

        record_label = str(record_ref or "").strip()
        if not record_label:
            raise ValueError("Record label is empty.")
        label_col = self._record_label_column()
        matches = self.df[self.df[label_col].astype(str) == record_label]
        if matches.empty and "_DISPLAY_NAME" in self.df.columns and label_col != "_DISPLAY_NAME":
            matches = self.df[self.df["_DISPLAY_NAME"].astype(str) == record_label]
        if matches.empty:
            raise ValueError(f"Record not found: {record_label}")
        if len(matches) > 1:
            raise ValueError(f"Multiple records found for '{record_label}'. Please refine the record label so each row stays unique.")
        return matches.iloc[0]

    def load_pdf(self, path):
        """Set PDF path, validate file, and return page count."""
        if not path:
            raise ValueError("No PDF path provided.")
        if not os.path.exists(path):
            raise FileNotFoundError(f"PDF file not found: {path}")

        try:
            total = self._pdf_page_count(path)
            if total <= 0:
                raise ValueError("PDF has no pages.")
        except Exception as e:
            raise ValueError(f"Failed to open PDF: {e}")

        self.pdf_path = path
        self.all_boxes = []
        self.box_types = []
        self.geom = self._empty_geom()
        self.semantic_targets = self._empty_semantic_targets()
        self.detected_page_idx = None
        self.detected_settings_signature = None
        self.boxes_by_page = {}
        self.box_types_by_page = {}
        self.geom_by_page = {}
        self.semantic_targets_by_page = {}
        self.boxes_settings_signature_by_page = {}
        self.candidate_boxes_by_page = {}
        self.candidate_types_by_page = {}
        self.candidate_scores_by_page = {}
        self.candidate_sources_by_page = {}
        self.candidate_review_state_by_page = {}
        self.candidate_meta_by_page = {}
        self.candidate_payloads_by_scope = {}
        self.area_templates_by_page = {}
        self.area_template_index = {}
        self.area_template_match_cache = {}
        self.area_template_match_review_state = {}
        self.manual_area_search_cache = {}
        self.repair_patches_by_page = {}
        self.repair_patch_index = {}
        self.sticky_detections_by_page = {}
        self.semantic_target_overrides = {}
        return total

    def total_pages(self):
        if not self.pdf_path or not os.path.exists(self.pdf_path):
            return 1
        try:
            return max(self._pdf_page_count(self.pdf_path), 1)
        except Exception:
            return 1

    def _pdf_page_count(self, path):
        return self.pdf_source.page_count(path)

    def _render_pdf_page_bgr(self, path, page_idx=0, preview_zoom=1.5):
        return self.pdf_source.render_page_bgr(path, page_idx=page_idx, preview_zoom=preview_zoom)


    def supports_detection_backend(self):
        """
        Return True when the app can rasterize PDF pages for detection/preview.
        Detection is image-first and can use Android PdfRenderer or PyMuPDF.
        """
        return bool(ANDROID_JAVA_AVAILABLE or FITZ_AVAILABLE)

    def supports_export_backend(self):
        """
        Return True when filled-PDF export is available.
        This variant exports by rasterizing pages and rebuilding the PDF as images,
        so it only requires a working PDF rendering backend.
        """
        return bool(ANDROID_JAVA_AVAILABLE or FITZ_AVAILABLE)

    def supports_processing_backend(self):
        """
        Backward-compatible alias for export-capable processing.
        """
        return self.supports_export_backend()

    def android_preview_only_mode(self):
        return platform == "android" and not self.supports_detection_backend()

    def _coerce_int_setting(self, key, default=None):
        default = DEFAULTS.get(key) if default is None else default
        value = self.settings.get(key, default)
        try:
            if isinstance(value, bool):
                return int(value)
            return int(float(value))
        except Exception:
            return int(float(default))

    def _coerce_float_setting(self, key, default=None):
        default = DEFAULTS.get(key) if default is None else default
        value = self.settings.get(key, default)
        try:
            if isinstance(value, (list, tuple)):
                value = value[0]
            return float(value)
        except Exception:
            return float(default)

    def _coerce_pair_setting(self, key, default=None):
        default = DEFAULTS.get(key) if default is None else default
        value = self.settings.get(key, default)
        if not isinstance(value, (list, tuple)) or len(value) < 2:
            value = default
        try:
            return (int(float(value[0])), int(float(value[1])))
        except Exception:
            return (int(float(default[0])), int(float(default[1])))

    def _coerce_bool_setting(self, key, default=None):
        default = DEFAULTS.get(key) if default is None else default
        value = self.settings.get(key, default)
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "yes", "on"}
        return bool(value)

    def _settings_signature(self):
        size_pair = self._coerce_pair_setting("C_Size")
        payload = {
            "F_Area": self._coerce_int_setting("F_Area"),
            "F_MinW": self._coerce_int_setting("F_MinW"),
            "F_MinH": self._coerce_int_setting("F_MinH"),
            "F_Close": self._coerce_int_setting("F_Close"),
            "Line_MinW": self._coerce_int_setting("Line_MinW"),
            "Line_MaxW": self._coerce_int_setting("Line_MaxW"),
            "C_Strict": self._coerce_int_setting("C_Strict"),
            "C_Size": [size_pair[0], size_pair[1]],
            "C_Border": round(self._coerce_float_setting("C_Border"), 6),
            "C_Inner": round(self._coerce_float_setting("C_Inner"), 6),
            "ROI_Max": self._coerce_int_setting("ROI_Max"),
            "C_Open": self._coerce_int_setting("C_Open"),
            "C_Close": self._coerce_int_setting("C_Close"),
            "C_BandPct": round(self._coerce_float_setting("C_BandPct"), 6),
            "C_AspectTol": round(self._coerce_float_setting("C_AspectTol"), 6),
            "Ext_Low": round(self._coerce_float_setting("Ext_Low"), 6),
            "Ext_High": round(self._coerce_float_setting("Ext_High"), 6),
            "C_FillMin": round(self._coerce_float_setting("C_FillMin"), 6),
            "C_Eps": round(self._coerce_float_setting("C_Eps"), 6),
            "Use_Extent": self._coerce_bool_setting("Use_Extent"),
        }
        return json.dumps(payload, sort_keys=True, separators=(",", ":"))

    def _cache_detection_for_page(self, page_idx):
        page_idx = int(page_idx)
        self.boxes_by_page[page_idx] = [fitz.Rect(r) for r in self.all_boxes]
        self.box_types_by_page[page_idx] = list(self.box_types)
        geom_copy = {}
        for key, rects in (self.geom or {}).items():
            geom_copy[key] = [fitz.Rect(r) for r in rects]
        self.geom_by_page[page_idx] = geom_copy
        self.semantic_targets_by_page[page_idx] = self._normalize_semantic_targets(self.semantic_targets)
        sig = self._settings_signature()
        self.boxes_settings_signature_by_page[page_idx] = sig
        self.detected_settings_signature = sig

    def invalidate_detection_cache(self, page_idx=None, clear_current=True):
        if page_idx is None:
            self.boxes_by_page = {}
            self.box_types_by_page = {}
            self.geom_by_page = {}
            self.semantic_targets_by_page = {}
            self.boxes_settings_signature_by_page = {}
            self.clear_candidate_cache()
            self.invalidate_area_template_match_cache()
            if clear_current:
                self.all_boxes = []
                self.box_types = []
                self.geom = self._empty_geom()
                self.semantic_targets = self._empty_semantic_targets()
                self.detected_page_idx = None
                self.detected_settings_signature = None
            return

        page_idx = int(page_idx)
        self.boxes_by_page.pop(page_idx, None)
        self.box_types_by_page.pop(page_idx, None)
        self.geom_by_page.pop(page_idx, None)
        self.semantic_targets_by_page.pop(page_idx, None)
        self.boxes_settings_signature_by_page.pop(page_idx, None)
        self.clear_candidate_cache(page_idx=page_idx)
        self.invalidate_area_template_match_cache(page_idx=page_idx)
        if clear_current and self.detected_page_idx == page_idx:
            self.all_boxes = []
            self.box_types = []
            self.geom = self._empty_geom()
            self.semantic_targets = self._empty_semantic_targets()
            self.detected_page_idx = None
            self.detected_settings_signature = None

    def _candidate_scope_key(self, page_idx, anchor_ids=None):
        page_idx = int(page_idx)
        norm_ids = []
        for value in list(anchor_ids or []):
            if isinstance(value, int) or str(value).isdigit():
                iv = int(value)
                if iv not in norm_ids:
                    norm_ids.append(iv)
        norm_ids = sorted(norm_ids)
        return f"{page_idx}:" + (",".join(str(v) for v in norm_ids) if norm_ids else "*")

    def clear_candidate_cache(self, page_idx=None):
        if page_idx is None:
            self.candidate_boxes_by_page = {}
            self.candidate_types_by_page = {}
            self.candidate_scores_by_page = {}
            self.candidate_sources_by_page = {}
            self.candidate_review_state_by_page = {}
            self.candidate_meta_by_page = {}
            self.candidate_payloads_by_scope = {}
            return
        page_idx = int(page_idx)
        self.candidate_boxes_by_page.pop(page_idx, None)
        self.candidate_types_by_page.pop(page_idx, None)
        self.candidate_scores_by_page.pop(page_idx, None)
        self.candidate_sources_by_page.pop(page_idx, None)
        self.candidate_review_state_by_page.pop(page_idx, None)
        self.candidate_meta_by_page.pop(page_idx, None)
        prefix = f"{page_idx}:"
        for scope_key in list((self.candidate_payloads_by_scope or {}).keys()):
            if str(scope_key).startswith(prefix):
                self.candidate_payloads_by_scope.pop(scope_key, None)

    def _normalize_cached_candidate_item(self, item):
        if not isinstance(item, dict) or item.get('rect') is None:
            return None
        try:
            rect = item.get('rect') if isinstance(item.get('rect'), fitz.Rect) else fitz.Rect(item.get('rect'))
        except Exception:
            return None
        source = str(item.get('source', 'scan') or 'scan')
        source_set = [str(s) for s in (item.get('source_set', []) or []) if str(s).strip()]
        if source and source not in source_set:
            source_set.append(source)
        try:
            support_count = int(item.get('support_count', len(source_set) if source_set else 1) or 1)
        except Exception:
            support_count = max(1, len(source_set) if source_set else 1)
        normalized = {
            'rect': rect,
            'kind': str(item.get('kind', 'field') or 'field'),
            'score': float(item.get('score', 0.0) or 0.0),
            'source': source,
            'source_set': list(source_set),
            'support_count': int(max(support_count, len(source_set) if source_set else 1)),
            'consensus_bonus': float(item.get('consensus_bonus', 0.0) or 0.0),
            'raw_score': float(item.get('raw_score', item.get('score', 0.0)) or 0.0),
            'base_score': float(item.get('base_score', item.get('score', 0.0)) or 0.0),
            'learning_bias': float(item.get('learning_bias', 0.0) or 0.0),
            'features': dict(item.get('features', {}) or {}),
        }
        if 'reason' in item:
            normalized['reason'] = str(item.get('reason', '') or '')
        return normalized

    def cache_candidates_for_page(self, page_idx, candidates=None, review_state=None, meta=None, anchor_ids=None):
        page_idx = int(page_idx)
        cand_list = []
        for item in list(candidates or []):
            norm = self._normalize_cached_candidate_item(item)
            if norm is not None:
                cand_list.append(norm)
        review_state = dict(review_state or {})
        meta_payload = dict(meta or {})
        norm_anchor_ids = []
        for value in list(anchor_ids or meta_payload.get('anchor_ids', []) or []):
            if isinstance(value, int) or str(value).isdigit():
                iv = int(value)
                if iv not in norm_anchor_ids:
                    norm_anchor_ids.append(iv)
        norm_anchor_ids = sorted(norm_anchor_ids)
        meta_payload['anchor_ids'] = list(norm_anchor_ids)

        boxes = [fitz.Rect(item.get('rect')) for item in cand_list]
        kinds = [str(item.get('kind', 'field') or 'field') for item in cand_list]
        scores = [float(item.get('score', 0.0) or 0.0) for item in cand_list]
        sources = [str(item.get('source', 'scan') or 'scan') for item in cand_list]

        self.candidate_boxes_by_page[page_idx] = list(boxes)
        self.candidate_types_by_page[page_idx] = list(kinds)
        self.candidate_scores_by_page[page_idx] = list(scores)
        self.candidate_sources_by_page[page_idx] = list(sources)
        self.candidate_review_state_by_page[page_idx] = dict(review_state)
        self.candidate_meta_by_page[page_idx] = dict(meta_payload)

        scope_key = self._candidate_scope_key(page_idx, norm_anchor_ids)
        self.candidate_payloads_by_scope[scope_key] = {
            'candidates': [self._normalize_cached_candidate_item(item) for item in cand_list],
            'review_state': dict(review_state),
            'meta': dict(meta_payload),
        }

    def get_candidates_for_page(self, page_idx, anchor_ids=None):
        page_idx = int(page_idx)
        scope_key = self._candidate_scope_key(page_idx, anchor_ids)
        scoped = dict((self.candidate_payloads_by_scope or {}).get(scope_key, {}) or {})
        if scoped:
            candidates = []
            for item in list(scoped.get('candidates', []) or []):
                norm = self._normalize_cached_candidate_item(item)
                if norm is not None:
                    candidates.append(norm)
            return {
                'candidates': candidates,
                'review_state': dict(scoped.get('review_state', {}) or {}),
                'meta': dict(scoped.get('meta', {}) or {}),
            }

        boxes = list(self.candidate_boxes_by_page.get(page_idx, []) or [])
        kinds = list(self.candidate_types_by_page.get(page_idx, []) or [])
        scores = list(self.candidate_scores_by_page.get(page_idx, []) or [])
        sources = list(self.candidate_sources_by_page.get(page_idx, []) or [])
        candidates = []
        for idx, rect in enumerate(boxes):
            norm = self._normalize_cached_candidate_item({
                'rect': fitz.Rect(rect),
                'kind': str(kinds[idx] if idx < len(kinds) else 'field'),
                'score': float(scores[idx] if idx < len(scores) else 0.0),
                'source': str(sources[idx] if idx < len(sources) else 'scan'),
            })
            if norm is not None:
                candidates.append(norm)
        return {
            'candidates': candidates,
            'review_state': dict(self.candidate_review_state_by_page.get(page_idx, {}) or {}),
            'meta': dict(self.candidate_meta_by_page.get(page_idx, {}) or {}),
        }

    def update_candidate_review_state_for_page(self, page_idx, review_state=None, anchor_ids=None):
        page_idx = int(page_idx)
        current = dict(self.candidate_review_state_by_page.get(page_idx, {}) or {})
        if isinstance(review_state, dict):
            current.update({str(k): bool(v) for k, v in review_state.items()})
        self.candidate_review_state_by_page[page_idx] = current
        scope_key = self._candidate_scope_key(page_idx, anchor_ids)
        scoped = dict((self.candidate_payloads_by_scope or {}).get(scope_key, {}) or {})
        if scoped:
            scoped_current = dict(scoped.get('review_state', {}) or {})
            scoped_current.update({str(k): bool(v) for k, v in (review_state or {}).items()})
            scoped['review_state'] = scoped_current
            self.candidate_payloads_by_scope[scope_key] = scoped
            return dict(scoped_current)
        return dict(current)

    def get_cached_mobile_repair_proposal(self, page_idx, anchor_ids=None):
        stored = self.get_candidates_for_page(page_idx, anchor_ids=anchor_ids)
        candidates = list(stored.get("candidates", []) or [])
        meta = dict(stored.get("meta", {}) or {})
        review_state = dict(stored.get("review_state", {}) or {})
        if not candidates and not meta.get("removal_candidates"):
            return None

        try:
            page_idx = int(page_idx or 0)
        except Exception:
            page_idx = 0

        cached_anchor_ids = [int(x) for x in (meta.get("anchor_ids", []) or []) if isinstance(x, int) or str(x).isdigit()]
        requested_anchor_ids = [int(x) for x in (anchor_ids or []) if isinstance(x, int) or str(x).isdigit()]
        if requested_anchor_ids and cached_anchor_ids and sorted(set(cached_anchor_ids)) != sorted(set(requested_anchor_ids)):
            return None

        expected_sig = str(meta.get("detection_signature", "") or "")
        current_sig = ""
        try:
            current_sig = str(self._settings_signature() or "")
        except Exception:
            current_sig = ""
        if expected_sig and current_sig and expected_sig != current_sig:
            return None

        expected_fp = str(meta.get("fingerprint_key", "") or "")
        current_fp = ""
        try:
            current_fp = str(((self.current_detection_context or {}).get('fingerprint') or {}).get('fingerprint_key', '') or '')
        except Exception:
            current_fp = ""
        if not current_fp and self.pdf_path:
            try:
                current_fp = str((self.build_page_fingerprint(page_idx=page_idx) or {}).get('fingerprint_key', '') or '')
            except Exception:
                current_fp = ""
        if expected_fp and current_fp and expected_fp != current_fp:
            return None

        detection_ready = False
        try:
            if int(getattr(self, "detected_page_idx", -1) or -1) == page_idx and bool(getattr(self, "all_boxes", []) or []):
                detection_ready = True
            else:
                detection_ready = bool(self._restore_detection_for_page(page_idx, required_signature=self._settings_signature()))
        except Exception:
            detection_ready = False
        if not detection_ready:
            return None

        try:
            zone_vals = list(meta.get("zone_rect", []) or [])
            zone_rect = fitz.Rect(*zone_vals) if len(zone_vals) == 4 else None
        except Exception:
            zone_rect = None

        all_boxes = list(getattr(self, 'all_boxes', []) or [])
        all_types = list(getattr(self, 'box_types', []) or [])

        def _resolve_existing_index(rect, kind, min_score=0.50):
            best_idx = -1
            best_score = 0.0
            try:
                rr = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
            except Exception:
                return -1
            kind = str(kind or '')
            for idx_existing, existing in enumerate(all_boxes):
                existing_kind = str(all_types[idx_existing] if idx_existing < len(all_types) else '')
                if kind and existing_kind != kind:
                    continue
                score = self._rect_iou(rr, existing)
                try:
                    if score <= 0.0:
                        score = self._mapping_match_score(rr, existing)
                except Exception:
                    pass
                if score > best_score:
                    best_idx = idx_existing
                    best_score = score
            return best_idx if best_score >= float(min_score) else -1

        removal_candidates = []
        for item in list(meta.get("removal_candidates", []) or []):
            if not isinstance(item, dict):
                continue
            try:
                rect_vals = list(item.get("rect", []) or [])
                rect = fitz.Rect(*rect_vals) if len(rect_vals) == 4 else None
            except Exception:
                rect = None
            if rect is None:
                continue
            kind = str(item.get("kind", "field") or "field")
            existing_index = _resolve_existing_index(rect, kind, min_score=0.45)
            if existing_index < 0:
                continue
            removal_candidates.append({
                "existing_index": int(existing_index),
                "rect": rect,
                "kind": kind,
                "reason": str(item.get("reason", "") or ""),
                "score": float(item.get("score", 0.0) or 0.0),
                "source": str(item.get("source", "cleanup") or "cleanup"),
            })
        return {
            "page_idx": int(page_idx),
            "anchor_ids": list(cached_anchor_ids or requested_anchor_ids),
            "anchor_rects": [],
            "zone_rect": zone_rect,
            "existing_local_ids": [int(x) for x in (meta.get("existing_local_ids", []) or []) if isinstance(x, int) or str(x).isdigit()],
            "existing_local_count": int(meta.get("existing_local_count", 0) or 0),
            "candidates": candidates,
            "candidate_type_counts": dict(meta.get("candidate_type_counts", {}) or {}),
            "source_counts": dict(meta.get("source_counts", {}) or {}),
            "inferred_structured_count": int(meta.get("inferred_structured_count", 0) or 0),
            "removal_candidates": removal_candidates,
            "review_defaults": dict(review_state),
            "cache_validated": True,
        }

    def current_detection_matches_settings(self, page_idx=0):
        page_idx = int(page_idx)
        if self.detected_page_idx != page_idx or not self.all_boxes:
            return False
        return str(self.detected_settings_signature or "") == self._settings_signature()

    def _restore_detection_for_page(self, page_idx, required_signature=None):
        page_idx = int(page_idx)
        if page_idx not in self.boxes_by_page:
            return False
        cached_sig = str(self.boxes_settings_signature_by_page.get(page_idx, "") or "")
        if required_signature is not None and cached_sig != str(required_signature or ""):
            return False
        self.all_boxes = [fitz.Rect(r) for r in self.boxes_by_page.get(page_idx, [])]
        self.box_types = list(self.box_types_by_page.get(page_idx, []))
        geom_src = self.geom_by_page.get(page_idx, {}) or {}
        semantic_src = self.semantic_targets_by_page.get(page_idx)
        if semantic_src:
            self.semantic_targets = self._normalize_semantic_targets(semantic_src)
            self.geom = self._legacy_geom_from_semantic_targets(self.semantic_targets)
        else:
            self.geom = {k: [fitz.Rect(r) for r in rects] for k, rects in geom_src.items()}
            self.semantic_targets = self._semantic_targets_from_geom(self.geom)
        self.detected_page_idx = page_idx
        self.detected_settings_signature = cached_sig or self._settings_signature()
        return True

    def _ensure_detection_for_page(self, page_idx=0):
        page_idx = int(page_idx)
        required_signature = self._settings_signature()
        if self.current_detection_matches_settings(page_idx):
            return
        if self._restore_detection_for_page(page_idx, required_signature=required_signature):
            return
        self.run_detection(page_idx=page_idx)

    def get_raw_preview_pixmap(self, page_idx=0, preview_zoom=1.5):
        """Render the raw loaded PDF page without filling or boxes."""
        if not self.pdf_path or not os.path.exists(self.pdf_path):
            raise FileNotFoundError("PDF path is missing or invalid.")
        return self._render_pdf_page_bgr(self.pdf_path, page_idx=page_idx, preview_zoom=preview_zoom)



    # --------------------------------------------------------
    # Box append helpers
    # --------------------------------------------------------
    def _append_box_unique(self, rect, box_type="check", iou_thresh=0.55):
        for j, existing in enumerate(self.all_boxes):
            if self.box_types[j] != box_type:
                continue
            if _rect_iou(rect, existing) >= iou_thresh:
                return
        self.all_boxes.append(rect)
        self.box_types.append(box_type)

    def _append_geom_fields(self):
        return

    # --------------------------------------------------------
    # Cleanup helpers
    # --------------------------------------------------------
    def _cleanup_field_fragments(self):
        if not self.all_boxes:
            return

        keep = [True] * len(self.all_boxes)
        field_idxs = [i for i, t in enumerate(self.box_types) if t == "field"]

        for si in field_idxs:
            if not keep[si]:
                continue

            sr = self.all_boxes[si]
            s_area = _rect_area(sr)
            if s_area <= 0:
                keep[si] = False
                continue

            for bi in field_idxs:
                if si == bi or not keep[bi]:
                    continue

                br = self.all_boxes[bi]
                if br.height < sr.height or _rect_area(br) <= s_area:
                    continue

                x_overlap = _x_overlap_ratio(sr, br)
                gap = br.y0 - sr.y1

                if (
                    x_overlap >= 0.80 and
                    0 <= gap <= 6.0 and
                    sr.height <= max(4.0, br.height * 0.45)
                ):
                    keep[si] = False
                    break

                inter = _rect_intersection_area(sr, br)
                if inter > 0:
                    cover_small = inter / max(s_area, 1e-6)
                    if cover_small >= 0.75 and sr.y1 <= (br.y0 + br.height * 0.45):
                        keep[si] = False
                        break

        self.all_boxes = [r for k, r in zip(keep, self.all_boxes) if k]
        self.box_types = [t for k, t in zip(keep, self.box_types) if k]

    def _cleanup_line_field_conflicts(self):
        if not self.all_boxes:
            return

        keep = [True] * len(self.all_boxes)
        field_idxs = [i for i, t in enumerate(self.box_types) if t == "field"]
        line_idxs = [i for i, t in enumerate(self.box_types) if t == "line"]

        for li in line_idxs:
            if not keep[li]:
                continue

            lr = self.all_boxes[li]
            la = _rect_area(lr)
            if la <= 0:
                keep[li] = False
                continue

            lmid_y = (lr.y0 + lr.y1) / 2.0

            for fi in field_idxs:
                fr = self.all_boxes[fi]

                inter = _rect_intersection_area(lr, fr)
                overlap_ratio = inter / max(la, 1e-6)
                x_overlap = _x_intersection(lr, fr) / max(lr.width, 1e-6)

                gap_above = fr.y0 - lr.y1
                gap_below = lr.y0 - fr.y1

                if overlap_ratio >= 0.35:
                    keep[li] = False
                    break

                if x_overlap >= 0.70 and (fr.y0 - 1.5) <= lmid_y <= (fr.y0 + fr.height * 0.60):
                    keep[li] = False
                    break

                if x_overlap >= 0.70 and (0 <= gap_above <= 4.0 or 0 <= gap_below <= 4.0):
                    keep[li] = False
                    break

                line_mid_x = (lr.x0 + lr.x1) / 2.0
                near_top_band = fr.y0 <= lr.y1 <= (fr.y0 + fr.height * 0.55)

                if (fr.x0 <= line_mid_x <= fr.x1) and near_top_band:
                    keep[li] = False
                    break

        self.all_boxes = [r for k, r in zip(keep, self.all_boxes) if k]
        self.box_types = [t for k, t in zip(keep, self.box_types) if k]

    # --------------------------------------------------------
    # Field refinement
    # --------------------------------------------------------
    def _refine_field_rect_from_mask(
        self,
        binv, x, y, w, h,
        zoom_factor=1.0,
        row_frac_thresh=0.45,
        col_frac_thresh=0.18,
        min_inner_h=18,
        pad_px=2
    ):
        roi = binv[y:y+h, x:x+w]
        if roi.size == 0:
            return fitz.Rect(x/zoom_factor, y/zoom_factor, (x+w)/zoom_factor, (y+h)/zoom_factor)

        row_frac = (roi > 0).sum(axis=1) / float(max(w, 1))
        col_frac = (roi > 0).sum(axis=0) / float(max(h, 1))

        strong_rows = np.where(row_frac >= row_frac_thresh)[0]
        strong_cols = np.where(col_frac >= col_frac_thresh)[0]

        if len(strong_rows) >= 2 and len(strong_cols) >= 2:
            ry0 = max(0, int(strong_rows[0]) - pad_px)
            ry1 = min(h, int(strong_rows[-1]) + pad_px + 1)
            rx0 = max(0, int(strong_cols[0]) - pad_px)
            rx1 = min(w, int(strong_cols[-1]) + pad_px + 1)

            if (ry1 - ry0) >= min_inner_h and (rx1 - rx0) >= max(20, int(w * 0.40)):
                return fitz.Rect(
                    (x + rx0) / zoom_factor,
                    (y + ry0) / zoom_factor,
                    (x + rx1) / zoom_factor,
                    (y + ry1) / zoom_factor
                )

        return fitz.Rect(x/zoom_factor, y/zoom_factor, (x+w)/zoom_factor, (y+h)/zoom_factor)

    # --------------------------------------------------------
    # Checkbox logic
    # --------------------------------------------------------
    def build_text_like_mask(self, img_bgr, header_band_pct=0.16):
        """Build a conservative mask of text-like components, with stronger support for headers/labels."""
        if img_bgr is None or getattr(img_bgr, "size", 0) == 0:
            return None, {"header_cut": 0, "component_count": 0}
        if len(img_bgr.shape) == 2:
            gray = img_bgr
        else:
            gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (3, 3), 0)
        _, binv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        h_img, w_img = binv.shape[:2]
        header_cut = int(max(8, min(h_img - 1, h_img * float(header_band_pct))))
        n, _, stats, _ = cv2.connectedComponentsWithStats(binv)
        mask = np.zeros_like(binv)
        comp_count = 0
        max_text_w = max(30, min(int(w_img * 0.20), 210))
        max_text_h = max(20, min(int(h_img * 0.11), 96))
        for i in range(1, n):
            x, y, w, h, area = [int(v) for v in stats[i]]
            if area < 5 or area > 2600 or w <= 0 or h <= 0:
                continue
            if w > max_text_w or h > max_text_h:
                continue
            roi = binv[y:y+h, x:x+w]
            if roi.size == 0:
                continue
            fill_ratio = float(area) / float(max(w * h, 1))
            if fill_ratio < 0.04 or fill_ratio > 0.90:
                continue
            aspect = float(w) / float(max(h, 1))
            if aspect < 0.08 or aspect > 20.0:
                continue
            row_cov = float(np.mean((roi > 0).sum(axis=1) / float(max(w, 1))))
            col_cov = float(np.mean((roi > 0).sum(axis=0) / float(max(h, 1))))
            if row_cov > 0.84 and col_cov > 0.84:
                continue
            # Compact glyphs and label fragments.
            compact_text_like = (
                (w <= 86 and h <= 44 and fill_ratio >= 0.08) or
                (w <= 56 and h <= 30 and fill_ratio >= 0.06)
            )
            # Header/title fragments can be slightly larger and wider.
            header_text_like = (
                y <= header_cut and
                w <= max(96, int(w_img * 0.14)) and
                h <= max(44, int(h_img * 0.06)) and
                fill_ratio >= 0.05
            )
            if not (compact_text_like or header_text_like):
                continue
            x0 = max(0, x - 1)
            y0 = max(0, y - 1)
            x1 = min(w_img, x + w + 1)
            y1 = min(h_img, y + h + 1)
            mask[y0:y1, x0:x1] = 255
            comp_count += 1
        if comp_count > 0:
            # Light horizontal support helps catch adjacent word fragments without spilling too far.
            mask = cv2.dilate(mask, np.ones((3, 3), np.uint8), iterations=1)
            mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, np.ones((5, 3), np.uint8), iterations=1)
        return mask, {"header_cut": int(header_cut), "component_count": int(comp_count)}

    def rect_text_overlap_ratio(self, rect, text_mask, preview_zoom=1.0):
        if text_mask is None:
            return 0.0
        try:
            rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
        except Exception:
            return 0.0
        h_img, w_img = text_mask.shape[:2]
        x0 = max(0, min(w_img, int(math.floor(rect.x0 * float(preview_zoom)))))
        y0 = max(0, min(h_img, int(math.floor(rect.y0 * float(preview_zoom)))))
        x1 = max(0, min(w_img, int(math.ceil(rect.x1 * float(preview_zoom)))))
        y1 = max(0, min(h_img, int(math.ceil(rect.y1 * float(preview_zoom)))))
        if x1 <= x0 or y1 <= y0:
            return 0.0
        roi = text_mask[y0:y1, x0:x1]
        if roi.size <= 0:
            return 0.0
        return float(cv2.countNonZero(roi)) / float(max(roi.size, 1))

    def _expanded_text_overlap_ratio(self, rect, text_mask, preview_zoom=1.0, pad_px=4):
        if text_mask is None:
            return 0.0
        try:
            rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
        except Exception:
            return 0.0
        h_img, w_img = text_mask.shape[:2]
        x0 = max(0, min(w_img, int(math.floor(rect.x0 * float(preview_zoom))) - int(max(0, pad_px))))
        y0 = max(0, min(h_img, int(math.floor(rect.y0 * float(preview_zoom))) - int(max(0, pad_px))))
        x1 = max(0, min(w_img, int(math.ceil(rect.x1 * float(preview_zoom))) + int(max(0, pad_px))))
        y1 = max(0, min(h_img, int(math.ceil(rect.y1 * float(preview_zoom))) + int(max(0, pad_px))))
        if x1 <= x0 or y1 <= y0:
            return 0.0
        roi = text_mask[y0:y1, x0:x1]
        if roi.size <= 0:
            return 0.0
        return float(cv2.countNonZero(roi)) / float(max(roi.size, 1))

    def should_suppress_text_like_rect(self, rect, text_mask=None, preview_zoom=1.0, header_cut=0, overlap_thresh=0.18):
        if text_mask is None:
            return False
        try:
            rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
        except Exception:
            return False
        overlap = self.rect_text_overlap_ratio(rect, text_mask, preview_zoom=preview_zoom)
        halo_overlap = self._expanded_text_overlap_ratio(
            rect,
            text_mask,
            preview_zoom=preview_zoom,
            pad_px=max(4, int(min(max(rect.width * float(preview_zoom), 1.0), max(rect.height * float(preview_zoom), 1.0)) * 0.25))
        )
        if overlap < float(overlap_thresh) and halo_overlap < max(0.20, float(overlap_thresh) + 0.06):
            return False
        w_px = float(max(rect.width, 0.0)) * float(preview_zoom)
        h_px = float(max(rect.height, 0.0)) * float(preview_zoom)
        y0_px = float(rect.y0) * float(preview_zoom)
        area_px = w_px * h_px
        smallish = (w_px <= 180.0 and h_px <= 88.0)
        tiny = (w_px <= 96.0 and h_px <= 62.0)
        glyphish = (w_px <= 56.0 and h_px <= 44.0)
        header_like = (y0_px <= float(max(header_cut, 0)) + 10.0)
        if header_like and tiny and (overlap >= max(0.10, float(overlap_thresh) * 0.80) or halo_overlap >= 0.22):
            return True
        if glyphish and (overlap >= max(0.14, float(overlap_thresh) * 0.88) or halo_overlap >= 0.28):
            return True
        if smallish and area_px <= 6400.0 and (overlap >= max(0.32, float(overlap_thresh) + 0.12) or halo_overlap >= 0.36):
            return True
        return False

    def _expanded_text_overlap_ratio(self, rect, text_mask=None, preview_zoom=1.0, pad_px=6):
        if text_mask is None:
            return 0.0
        try:
            rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
        except Exception:
            return 0.0
        try:
            h_img, w_img = text_mask.shape[:2]
        except Exception:
            return 0.0
        x0 = max(0, min(w_img, int(math.floor(rect.x0 * float(preview_zoom))) - int(max(0, pad_px))))
        y0 = max(0, min(h_img, int(math.floor(rect.y0 * float(preview_zoom))) - int(max(0, pad_px))))
        x1 = max(0, min(w_img, int(math.ceil(rect.x1 * float(preview_zoom))) + int(max(0, pad_px))))
        y1 = max(0, min(h_img, int(math.ceil(rect.y1 * float(preview_zoom))) + int(max(0, pad_px))))
        if x1 <= x0 or y1 <= y0:
            return 0.0
        roi = text_mask[y0:y1, x0:x1]
        if roi.size <= 0:
            return 0.0
        return float(cv2.countNonZero(roi)) / float(max(roi.size, 1))

    def _strict_title_cut_px(self, header_cut=0):
        try:
            header_cut = float(header_cut or 0.0)
        except Exception:
            header_cut = 0.0
        # The previous logic used the full header_cut from the text-mask pass,
        # which is ~16% of the page and can reach real controls on tall forms.
        # Keep this much shallower so only the actual title/title-subtitle band is affected.
        return float(max(26.0, min(58.0, (header_cut * 0.36) + 6.0)))

    def should_reject_header_word_field_artifact(self, rect, text_mask=None, preview_zoom=1.0, header_cut=0):
        if text_mask is None:
            return False
        try:
            rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
        except Exception:
            return False
        w_px = float(max(rect.width, 0.0)) * float(preview_zoom)
        h_px = float(max(rect.height, 0.0)) * float(preview_zoom)
        y0_px = float(rect.y0) * float(preview_zoom)
        title_cut_px = self._strict_title_cut_px(header_cut)
        # Ultra-narrow scope: only tiny field-like boxes in the actual title band.
        if y0_px > title_cut_px:
            return False
        if w_px > 54.0 or h_px > 54.0:
            return False
        overlap = self.rect_text_overlap_ratio(rect, text_mask, preview_zoom=preview_zoom)
        halo = self._expanded_text_overlap_ratio(rect, text_mask, preview_zoom=preview_zoom, pad_px=8)
        if overlap >= 0.22:
            return True
        if halo >= 0.34:
            return True
        if (w_px <= 38.0 and h_px <= 42.0) and (overlap >= 0.14 or halo >= 0.26):
            return True
        return False


    def _pdf_character_regions_for_page(self, page_idx, pad_x=0.35, pad_y=0.40):
        regions = []
        if not self.pdf_path or not os.path.exists(self.pdf_path):
            return regions
        if not FITZ_AVAILABLE:
            return regions
        try:
            with fitz.open(self.pdf_path) as doc:
                total = len(doc)
                if total <= 0:
                    return regions
                page_idx = max(0, min(int(page_idx or 0), total - 1))
                page = doc.load_page(page_idx)
                page_h = float(getattr(page.rect, "height", 0.0) or 0.0)

                # Preferred path: real character boxes from the PDF text layer.
                try:
                    raw = page.get_text("rawdict") or {}
                    char_regions = []
                    for block in list(raw.get("blocks", []) or []):
                        if int(block.get("type", 0) or 0) != 0:
                            continue
                        for line in list(block.get("lines", []) or []):
                            for span in list(line.get("spans", []) or []):
                                for ch in list(span.get("chars", []) or []):
                                    c = str(ch.get("c", "") or "")
                                    if not c or not c.strip():
                                        continue
                                    bbox = ch.get("bbox")
                                    if not bbox or len(bbox) < 4:
                                        continue
                                    x0, y0, x1, y1 = [float(v) for v in bbox[:4]]
                                    w = max(0.0, x1 - x0)
                                    h = max(0.0, y1 - y0)
                                    if w <= 0.08 or h <= 0.08:
                                        continue
                                    if w > 24.0 or h > 28.0:
                                        continue

                                    # Safer tweak: only enlarge header/title characters a bit more.
                                    is_header_char = bool(page_h > 0.0 and ((y0 + y1) * 0.5) <= (page_h * 0.18))
                                    use_pad_x = float(pad_x)
                                    use_pad_y = float(pad_y)
                                    if is_header_char:
                                        use_pad_x = max(use_pad_x, 0.70)
                                        use_pad_y = max(use_pad_y, 0.55)

                                    char_regions.append(
                                        fitz.Rect(
                                            max(0.0, x0 - use_pad_x),
                                            max(0.0, y0 - use_pad_y),
                                            x1 + use_pad_x,
                                            y1 + use_pad_y,
                                        )
                                    )
                    if char_regions:
                        return char_regions
                except Exception:
                    pass

                # Fallback: estimate per-character slots from word boxes.
                words = page.get_text("words") or []
        except Exception:
            return regions

        for word in words:
            try:
                x0, y0, x1, y1 = float(word[0]), float(word[1]), float(word[2]), float(word[3])
                txt = str(word[4] if len(word) > 4 else "")
            except Exception:
                continue
            txt = txt or ""
            stripped = txt.strip()
            if not stripped:
                continue

            chars = [ch for ch in stripped]
            n = len(chars)
            if n <= 0:
                continue

            w = max(0.0, x1 - x0)
            h = max(0.0, y1 - y0)
            if w <= 0.2 or h <= 0.2:
                continue
            if w > 220.0 or h > 40.0:
                continue

            char_w = w / float(max(n, 1))
            if char_w <= 0.10:
                continue

            # Slight extra padding only for fallback header words.
            is_header_word = bool(((y0 + y1) * 0.5) <= 32.0)
            use_pad_x = max(float(pad_x), 0.55 if is_header_word else float(pad_x))
            use_pad_y = max(float(pad_y), 0.50 if is_header_word else float(pad_y))

            for i, ch in enumerate(chars):
                if not str(ch).strip():
                    continue
                cx0 = x0 + (i * char_w)
                cx1 = x0 + ((i + 1) * char_w)
                regions.append(
                    fitz.Rect(
                        max(0.0, cx0 - use_pad_x),
                        max(0.0, y0 - use_pad_y),
                        cx1 + use_pad_x,
                        y1 + use_pad_y,
                    )
                )
        return regions

    def _remove_detections_inside_pdf_character_regions(self, page_idx, char_regions=None, text_mask=None, preview_zoom=1.0, header_cut=0):
        char_regions = [fitz.Rect(r) for r in (char_regions or [])]
        if not char_regions or not getattr(self, "all_boxes", None):
            return 0

        remove_ids = set()
        for idx, (rect, kind) in enumerate(zip(list(self.all_boxes or []), list(self.box_types or []))):
            try:
                rr = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
            except Exception:
                continue
            kind = str(kind or "")
            w = float(max(rr.width, 0.0))
            h = float(max(rr.height, 0.0))
            area = w * h

            # Still safely below real form controls, but a bit wider now that actual char boxes are preferred.
            if kind == "check":
                if w > 18.0 or h > 18.0 or area > 190.0:
                    continue
            elif kind == "field":
                if w > 20.0 or h > 15.0 or area > 220.0:
                    continue
            elif kind == "line":
                if w > 40.0 or h > 7.5:
                    continue
            else:
                continue

            cy_px = float((rr.y0 + rr.y1) * 0.5) * float(preview_zoom)
            title_cut_px = self._strict_title_cut_px(header_cut)
            # Safety: only suppress actual title/header artifacts, never body detections like real form checkboxes.
            if cy_px > title_cut_px:
                continue

            overlap = 0.0
            halo = 0.0
            if text_mask is not None:
                try:
                    overlap = float(self.rect_text_overlap_ratio(rr, text_mask, preview_zoom=preview_zoom) or 0.0)
                except Exception:
                    overlap = 0.0
                try:
                    halo = float(self._expanded_text_overlap_ratio(rr, text_mask, preview_zoom=preview_zoom, pad_px=5) or 0.0)
                except Exception:
                    halo = overlap

            # Require some actual rendered-text evidence.
            if overlap < 0.06 and halo < 0.16:
                continue

            rcx = (rr.x0 + rr.x1) * 0.5
            rcy = (rr.y0 + rr.y1) * 0.5
            r_area = max(area, 1e-6)

            best_ioa = 0.0
            center_in_char = False
            for cr in char_regions:
                inter = _rect_intersection_area(rr, cr)
                ioa = inter / r_area if r_area > 0 else 0.0
                if ioa > best_ioa:
                    best_ioa = ioa
                if cr.x0 <= rcx <= cr.x1 and cr.y0 <= rcy <= cr.y1:
                    center_in_char = True

            # Character-slot method: center inside a char slot is the main signal.
            micro_tiny = (
                (kind == "check" and w <= 10.5 and h <= 10.5 and area <= 82.0) or
                (kind == "field" and w <= 12.0 and h <= 9.0 and area <= 96.0) or
                (kind == "line" and w <= 20.0 and h <= 4.0)
            )

            if center_in_char:
                if kind == "line":
                    if best_ioa >= (0.10 if micro_tiny else 0.14):
                        remove_ids.add(int(idx))
                else:
                    if best_ioa >= (0.05 if micro_tiny else 0.09):
                        remove_ids.add(int(idx))
            elif best_ioa >= (0.68 if micro_tiny else 0.80):
                remove_ids.add(int(idx))

        if not remove_ids:
            return 0

        keep_boxes, keep_types = [], []
        removed = 0
        for idx, (rect, kind) in enumerate(zip(self.all_boxes, self.box_types)):
            if idx in remove_ids:
                removed += 1
                continue
            keep_boxes.append(rect)
            keep_types.append(kind)
        self.all_boxes = keep_boxes
        self.box_types = keep_types
        return int(removed)
    def should_reject_header_word_check_artifact(self, rect, text_mask=None, preview_zoom=1.0, header_cut=0):
        if text_mask is None:
            return False
        try:
            rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
        except Exception:
            return False
        w_px = float(max(rect.width, 0.0)) * float(preview_zoom)
        h_px = float(max(rect.height, 0.0)) * float(preview_zoom)
        y0_px = float(rect.y0) * float(preview_zoom)
        title_cut_px = self._strict_title_cut_px(header_cut)
        # Very narrow scope: only tiny checkbox-like detections in the actual title band.
        if y0_px > title_cut_px:
            return False
        if w_px > 30.0 or h_px > 30.0:
            return False
        overlap = self.rect_text_overlap_ratio(rect, text_mask, preview_zoom=preview_zoom)
        halo = self._expanded_text_overlap_ratio(rect, text_mask, preview_zoom=preview_zoom, pad_px=7)
        aspect = w_px / max(h_px, 1e-6)
        if not (0.60 <= aspect <= 1.55):
            return False
        if overlap >= 0.18:
            return True
        if halo >= 0.32:
            return True
        if (w_px <= 24.0 and h_px <= 24.0) and (overlap >= 0.10 or halo >= 0.22):
            return True
        return False

    def detect_faint_field_candidates(self, img_bgr, zoom_factor=1.0, zone_rect=None, text_mask=None, header_cut=0):
        """Repair-only candidate detector for faint or low-contrast rectangular fields."""
        if img_bgr is None or getattr(img_bgr, "size", 0) == 0:
            return []
        if len(img_bgr.shape) == 2:
            gray = img_bgr
        else:
            gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        try:
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            work = clahe.apply(gray)
        except Exception:
            work = gray
        work = cv2.GaussianBlur(work, (3, 3), 0)
        edges = cv2.Canny(work, 30, 110)
        edges = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8), iterations=1)
        cnts, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        h_img, w_img = gray.shape[:2]
        out = []
        for c in cnts:
            x, y, w, h = cv2.boundingRect(c)
            if w < 28 or h < 12:
                continue
            if w > int(w_img * 0.92) or h > int(h_img * 0.18):
                continue
            if w * h < 260 or w * h > int(w_img * h_img * 0.035):
                continue
            rect = fitz.Rect(x / zoom_factor, y / zoom_factor, (x + w) / zoom_factor, (y + h) / zoom_factor)
            if zone_rect is not None and _rect_intersection_area(rect, zone_rect) <= 0:
                continue
            if self.should_suppress_text_like_rect(rect, text_mask, preview_zoom=zoom_factor, header_cut=header_cut, overlap_thresh=0.14):
                continue
            aspect = float(w) / float(max(h, 1))
            if aspect < 1.4 or aspect > 18.0:
                continue
            peri = cv2.arcLength(c, True)
            area = max(float(cv2.contourArea(c)), 1.0)
            extent = area / float(max(w * h, 1))
            approx = cv2.approxPolyDP(c, 0.03 * peri, True) if peri > 0 else []
            if len(approx) < 4 or len(approx) > 10:
                continue
            roi_e = edges[y:y+h, x:x+w]
            if roi_e.size == 0:
                continue
            row_cov = float(np.mean((roi_e > 0).sum(axis=1) / float(max(w, 1))))
            col_cov = float(np.mean((roi_e > 0).sum(axis=0) / float(max(h, 1))))
            border_band = max(1, min(int(min(w, h) * 0.14), 10))
            top = roi_e[:border_band, :]
            bottom = roi_e[-border_band:, :]
            left = roi_e[:, :border_band]
            right = roi_e[:, -border_band:]
            border_strength = min(
                float(cv2.countNonZero(top)) / float(max(top.size, 1)),
                float(cv2.countNonZero(bottom)) / float(max(bottom.size, 1)),
                float(cv2.countNonZero(left)) / float(max(left.size, 1)),
                float(cv2.countNonZero(right)) / float(max(right.size, 1)),
            )
            side_hits = sum([
                1 if (float(cv2.countNonZero(top)) / float(max(top.size, 1))) >= 0.05 else 0,
                1 if (float(cv2.countNonZero(bottom)) / float(max(bottom.size, 1))) >= 0.05 else 0,
                1 if (float(cv2.countNonZero(left)) / float(max(left.size, 1))) >= 0.05 else 0,
                1 if (float(cv2.countNonZero(right)) / float(max(right.size, 1))) >= 0.05 else 0,
            ])
            if side_hits < 3:
                continue
            score = 0.42
            score += min(0.14, max(0.0, (border_strength - 0.04) * 1.6))
            score += min(0.10, max(0.0, (row_cov - 0.03) * 0.65))
            score += min(0.10, max(0.0, (col_cov - 0.03) * 0.65))
            score += 0.08 if side_hits >= 4 else 0.03
            if 0.04 <= extent <= 0.42:
                score += 0.08
            out.append({"rect": rect, "score": float(max(0.0, min(0.95, score))), "source": "faint_field"})
        return out

    def detect_partial_box_candidates(self, img_bgr, zoom_factor=1.0, zone_rect=None, text_mask=None, header_cut=0):
        """Repair-only detector for broken or incomplete rectangular field borders."""
        if img_bgr is None or getattr(img_bgr, "size", 0) == 0:
            return []
        if len(img_bgr.shape) == 2:
            gray = img_bgr
        else:
            gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (3, 3), 0)
        _, binv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        h_img, w_img = binv.shape[:2]
        cnts, _ = cv2.findContours(binv, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        out = []
        for c in cnts:
            x, y, w, h = cv2.boundingRect(c)
            if w < 28 or h < 12:
                continue
            if w > int(w_img * 0.92) or h > int(h_img * 0.18):
                continue
            if w * h < 240:
                continue
            rect = fitz.Rect(x / zoom_factor, y / zoom_factor, (x + w) / zoom_factor, (y + h) / zoom_factor)
            if zone_rect is not None and _rect_intersection_area(rect, zone_rect) <= 0:
                continue
            if self.should_suppress_text_like_rect(rect, text_mask, preview_zoom=zoom_factor, header_cut=header_cut, overlap_thresh=0.14):
                continue
            aspect = float(w) / float(max(h, 1))
            if aspect < 1.5 or aspect > 22.0:
                continue
            roi = binv[y:y+h, x:x+w]
            if roi.size == 0:
                continue
            band = max(1, min(int(min(w, h) * 0.12), 10))
            top = float(cv2.countNonZero(roi[:band, :])) / float(max(roi[:band, :].size, 1))
            bottom = float(cv2.countNonZero(roi[-band:, :])) / float(max(roi[-band:, :].size, 1))
            left = float(cv2.countNonZero(roi[:, :band])) / float(max(roi[:, :band].size, 1))
            right = float(cv2.countNonZero(roi[:, -band:])) / float(max(roi[:, -band:].size, 1))
            side_vals = [top, bottom, left, right]
            strong_sides = sum(1 for v in side_vals if v >= 0.05)
            weak_sides = sum(1 for v in side_vals if 0.015 <= v < 0.05)
            if strong_sides + weak_sides < 3:
                continue
            filled = float(cv2.countNonZero(roi)) / float(max(roi.size, 1))
            if filled > 0.42:
                continue
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.035 * peri, True) if peri > 0 else []
            if len(approx) < 2 or len(approx) > 12:
                continue
            score = 0.40
            score += min(0.18, strong_sides * 0.055)
            score += min(0.08, weak_sides * 0.025)
            score += 0.08 if 0.03 <= filled <= 0.24 else 0.0
            score += 0.06 if len(approx) >= 4 else 0.02
            out.append({"rect": rect, "score": float(max(0.0, min(0.92, score))), "source": "partial_box"})
        return out

    def detect_non_square_option_candidates(self, img_bgr, zoom_factor=1.0, zone_rect=None, text_mask=None, header_cut=0):
        """Repair-only detector for radio / diamond / weak non-square option-like candidates.
        These are intentionally softer than looks_like_checkbox() and feed the candidate layer only.
        """
        if img_bgr is None or getattr(img_bgr, "size", 0) == 0:
            return []
        if len(img_bgr.shape) == 2:
            img_bgr = cv2.cvtColor(img_bgr, cv2.COLOR_GRAY2BGR)

        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (3, 3), 0)
        _, binv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        binv = cv2.morphologyEx(binv, cv2.MORPH_OPEN, np.ones((1, 1), np.uint8))

        h_img, w_img = binv.shape[:2]
        if zone_rect is not None:
            try:
                zr = zone_rect if isinstance(zone_rect, fitz.Rect) else fitz.Rect(*zone_rect)
                zx0 = max(0, min(w_img, int(math.floor(float(zr.x0) * float(zoom_factor)))))
                zy0 = max(0, min(h_img, int(math.floor(float(zr.y0) * float(zoom_factor)))))
                zx1 = max(0, min(w_img, int(math.ceil(float(zr.x1) * float(zoom_factor)))))
                zy1 = max(0, min(h_img, int(math.ceil(float(zr.y1) * float(zoom_factor)))))
                if zx1 > zx0 and zy1 > zy0:
                    roi_mask = np.zeros_like(binv)
                    roi_mask[zy0:zy1, zx0:zx1] = binv[zy0:zy1, zx0:zx1]
                    work = roi_mask
                else:
                    work = binv
            except Exception:
                work = binv
        else:
            work = binv

        min_sz = max(8, int(self.settings.get("C_Size", (14, 65))[0]) - 6)
        max_sz = min(72, int(self.settings.get("C_Size", (14, 65))[1]) + 12)
        out = []
        seen = []

        cnts, _ = cv2.findContours(work, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        for c in cnts:
            try:
                x, y, w, h = cv2.boundingRect(c)
            except Exception:
                continue
            if w < min_sz or h < min_sz or w > max_sz or h > max_sz:
                continue
            if y < int(header_cut or 0):
                continue
            aspect = (w / float(max(h, 1)))
            if not (0.55 <= aspect <= 1.60):
                continue
            rect = fitz.Rect(x / zoom_factor, y / zoom_factor, (x + w) / zoom_factor, (y + h) / zoom_factor)
            if text_mask is not None:
                try:
                    if self.should_suppress_text_like_rect(rect, text_mask, preview_zoom=zoom_factor, header_cut=header_cut, overlap_thresh=0.16):
                        continue
                except Exception:
                    pass
            area = float(cv2.contourArea(c))
            if area <= 0:
                continue
            peri = float(cv2.arcLength(c, True))
            if peri <= 0:
                continue
            circularity = float((4.0 * math.pi * area) / max(peri * peri, 1e-6))
            extent = float(area / max(float(w * h), 1.0))
            approx = cv2.approxPolyDP(c, max(1.0, 0.05 * peri), True)
            vertices = int(len(approx) if approx is not None else 0)
            cx = x + (w * 0.5)
            cy = y + (h * 0.5)
            roi = work[max(0, y - 2):min(h_img, y + h + 2), max(0, x - 2):min(w_img, x + w + 2)]
            ink_ratio = float(cv2.countNonZero(roi) / float(max(roi.size, 1))) if roi.size > 0 else 0.0

            source = ''
            score = 0.0
            # Radio / circular options
            if 0.72 <= aspect <= 1.28 and circularity >= 0.45 and 0.18 <= extent <= 0.90:
                source = 'radio_like'
                score = 0.60 + min(0.20, max(0.0, circularity - 0.45))
                if vertices >= 6:
                    score += 0.05
            # Rotated diamonds / lozenge-like options
            elif 0.70 <= aspect <= 1.35 and vertices == 4 and 0.10 <= extent <= 0.46:
                source = 'diamond_like'
                score = 0.57
                if 0.16 <= extent <= 0.36:
                    score += 0.05
            # Broken / weak check-like options that are not strong enough for normal checkbox logic
            elif 0.70 <= aspect <= 1.40 and 0.10 <= extent <= 0.62 and 3 <= vertices <= 8 and ink_ratio <= 0.55:
                source = 'weak_option'
                score = 0.48
                if circularity >= 0.30:
                    score += 0.04
                if vertices == 4:
                    score += 0.03
            else:
                continue

            duplicate = False
            for prev in seen:
                if self._rect_iou(prev, rect) >= 0.55:
                    duplicate = True
                    break
            if duplicate:
                continue
            seen.append(rect)
            out.append({
                'rect': rect,
                'source': source,
                'score': round(float(min(max(score, 0.0), 0.95)), 4),
                'kind': 'check',
                'center': [round(float(cx / zoom_factor), 3), round(float(cy / zoom_factor), 3)],
                'circularity': round(float(circularity), 4),
                'extent': round(float(extent), 4),
                'vertices': int(vertices),
            })

        return out

    def looks_like_checkbox(self, binv, x, y, w, h, cc_area=None):
        strictness = self.settings["C_Strict"]
        border_override = self.settings["C_Border"]
        inner_override = self.settings["C_Inner"]
        band_pct = self.settings["C_BandPct"]
        aspect_tol = self.settings["C_AspectTol"]
        extent_low = self.settings["Ext_Low"]
        extent_high = self.settings["Ext_High"]
        fill_min = self.settings["C_FillMin"]
        eps = self.settings["C_Eps"]
        use_extent = self.settings["Use_Extent"]

        roi = binv[y:y+h, x:x+w]
        if roi.size == 0:
            return False

        s = max(10.0, min(float(strictness), 100.0))
        delta = (s - 40.0) / 60.0

        eff_border = border_override + (delta * 0.05)
        eff_inner  = inner_override - (delta * 0.06)
        eff_aspect = aspect_tol - (delta * 0.05)
        eff_fill   = fill_min + (delta * 0.05)

        eff_border = max(0.03, min(0.95, eff_border))
        eff_inner  = max(0.05, min(0.95, eff_inner))
        eff_aspect = max(0.08, min(1.00, eff_aspect))
        eff_fill   = max(0.18, min(0.95, eff_fill))

        aspect = w / float(h)
        if not (1 - eff_aspect <= aspect <= 1 + eff_aspect):
            return False

        if use_extent and cc_area is not None:
            extent = cc_area / float(w * h)
            if extent < extent_low or extent > extent_high:
                return False

        t = max(1, int(min(w, h) * band_pct))
        if w <= 2 * t or h <= 2 * t:
            return False

        top, bottom, left, right = roi[:t, :], roi[-t:, :], roi[:, :t], roi[:, -t:]
        inner = roi[t:-t, t:-t]

        def frac(a):
            return cv2.countNonZero(a) / float(a.size) if a.size > 0 else 0

        if min(frac(top), frac(bottom), frac(left), frac(right)) < eff_border:
            return False
        if frac(inner) > eff_inner:
            return False

        cnts, _ = cv2.findContours(roi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not cnts:
            return False

        c = max(cnts, key=cv2.contourArea)
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, eps * peri, True)

        if len(approx) != 4 or not cv2.isContourConvex(approx):
            return False

        bx, by, bw, bh = cv2.boundingRect(c)
        c_area = cv2.contourArea(c)
        rect_fill = c_area / float(max(bw * bh, 1))
        if rect_fill < 0.70:
            return False

        pad = max(2, int(min(w, h) * 0.35))
        y0 = max(0, y - pad)
        y1 = min(binv.shape[0], y + h + pad)
        x0 = max(0, x - pad)
        x1 = min(binv.shape[1], x + w + pad)
        outer = binv[y0:y1, x0:x1].copy()

        ix0 = x - x0
        iy0 = y - y0
        ix1 = ix0 + w
        iy1 = iy0 + h
        outer[iy0:iy1, ix0:ix1] = 0

        outer_frac = cv2.countNonZero(outer) / float(max(outer.size, 1))
        if outer_frac > 0.10:
            return False

        if (c_area / float(w * h)) < eff_fill:
            return False

        return True

# =========================
# PART 2 / 4
# Detection Engine Continuation
# Add this BELOW Part 1, inside the FormAlchemistEngine class
# =========================

    # --------------------------------------------------------
    # Line detection
    # --------------------------------------------------------
    def find_answer_lines(self, img, zoom_factor=3.0, min_line_w=100, max_line_w=900):
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (3, 3), 0)
        _, binv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(40, int(min_line_w) // 4), 2))
        detected_lines = cv2.morphologyEx(binv, cv2.MORPH_OPEN, kernel, iterations=1)
        cnts, _ = cv2.findContours(detected_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        line_rects = []
        img_h, img_w = img.shape[:2]

        for c in cnts:
            x, y, w, h = cv2.boundingRect(c)

            effective_max_line_w = int(max_line_w)

            if w > min_line_w and w < effective_max_line_w and 1 <= h <= 12:
                roi_tl = binv[max(0, y - 5):y, x:x + 4]
                roi_tr = binv[max(0, y - 5):y, x + w - 4:x + w]
                if cv2.countNonZero(roi_tl) > 4 and cv2.countNonZero(roi_tr) > 4:
                    continue

                roi_left_edge = binv[y:y + h, max(0, x - 5):x]
                roi_right_edge = binv[y:y + h, x + w:min(img_w, x + w + 5)]
                if cv2.countNonZero(roi_left_edge) > 8 and cv2.countNonZero(roi_right_edge) > 8:
                    continue

                roi_above = binv[max(0, y - 12):y, x:x + w]
                if roi_above.size > 0 and (cv2.countNonZero(roi_above) / float(roi_above.size)) > 0.35:
                    continue

                if y < (img_h * 0.05) or y > (img_h * 0.95):
                    continue

                line_rects.append(
                    fitz.Rect(
                        x / zoom_factor,
                        (y - 28) / zoom_factor,
                        (x + w) / zoom_factor,
                        y / zoom_factor
                    )
                )

        return line_rects

    # --------------------------------------------------------
    # Checkbox ROI scan
    # --------------------------------------------------------
    def find_checkbox_rects_in_roi(self, binv, x, y, w, h):
        roi = binv[y:y + h, x:x + w]
        if roi.size == 0:
            return []

        min_sz, max_sz = self.settings["C_Size"]
        cnts, _ = cv2.findContours(roi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        found = []

        for c in cnts:
            rx, ry, rw, rh = cv2.boundingRect(c)
            if not (min_sz <= rw <= max_sz and min_sz <= rh <= max_sz):
                continue

            cc_area = cv2.contourArea(c)
            if self.looks_like_checkbox(binv, x + rx, y + ry, rw, rh, cc_area=cc_area):
                found.append((x + rx, y + ry, rw, rh))

        return found

    # --------------------------------------------------------
    # Red checkbox color-assisted detection
    # --------------------------------------------------------
    def find_red_checkbox_candidates(
        self,
        img_bgr,
        zoom_factor=3.0,
        min_sz=8,
        max_sz=40,
        border_min=0.06,
        inner_max=0.65,
        band_pct=0.12,
        aspect_tol=0.45,
        fill_min=0.18,
        eps=0.06
    ):
        hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)

        lower1 = np.array([0, 70, 60], dtype=np.uint8)
        upper1 = np.array([12, 255, 255], dtype=np.uint8)
        lower2 = np.array([168, 70, 60], dtype=np.uint8)
        upper2 = np.array([180, 255, 255], dtype=np.uint8)

        mask1 = cv2.inRange(hsv, lower1, upper1)
        mask2 = cv2.inRange(hsv, lower2, upper2)
        red_mask = cv2.bitwise_or(mask1, mask2)

        red_mask = cv2.morphologyEx(red_mask, cv2.MORPH_CLOSE, np.ones((2, 2), np.uint8))
        red_mask = cv2.morphologyEx(red_mask, cv2.MORPH_OPEN, np.ones((1, 1), np.uint8))

        cnts, _ = cv2.findContours(red_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        out = []

        for c in cnts:
            x, y, w, h = cv2.boundingRect(c)

            if not (min_sz <= w <= max_sz and min_sz <= h <= max_sz):
                continue

            aspect = w / float(h) if h > 0 else 999
            if not (1 - aspect_tol <= aspect <= 1 + aspect_tol):
                continue

            roi = red_mask[y:y + h, x:x + w]
            if roi.size == 0:
                continue

            t = max(1, int(min(w, h) * band_pct))
            if w <= 2 * t or h <= 2 * t:
                continue

            top, bottom = roi[:t, :], roi[-t:, :]
            left, right = roi[:, :t], roi[:, -t:]
            inner = roi[t:-t, t:-t]

            def frac(a):
                return cv2.countNonZero(a) / float(a.size) if a.size > 0 else 0.0

            if min(frac(top), frac(bottom), frac(left), frac(right)) < border_min:
                continue
            if frac(inner) > inner_max:
                continue

            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, eps * peri, True)
            if len(approx) != 4:
                continue

            area = cv2.contourArea(c)
            if (area / float(w * h)) < fill_min:
                continue

            out.append(
                fitz.Rect(
                    x / zoom_factor,
                    y / zoom_factor,
                    (x + w) / zoom_factor,
                    (y + h) / zoom_factor
                )
            )

        return out

    def _infer_semantic_targets_from_dem_roi(self, img, field_area_thresh=500, field_min_h_val=20):
        if not SEMANTIC_TARGETS_ENABLED:
            return self._empty_semantic_targets()
        h_img, _ = img.shape[:2]
        y0_roi, y1_roi = int(0.22 * h_img), int(0.82 * h_img)
        gray_roi = cv2.cvtColor(img[y0_roi:y1_roi, :], cv2.COLOR_BGR2GRAY)
        _, bin_roi = cv2.threshold(gray_roi, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        _, _, stats, _ = cv2.connectedComponentsWithStats(bin_roi)

        min_h = max(18, int(field_min_h_val * 0.6))
        cands = []
        for x, y, ww, hh, area in stats[1:]:
            if hh < min_h or area <= field_area_thresh or ww < 24:
                continue
            cands.append((int(x), int(y), int(ww), int(hh)))

        targets = self._empty_semantic_targets()
        if not cands:
            return targets

        median_h = float(np.median([hh for _, _, _, hh in cands])) if cands else 24.0
        row_tol = max(14.0, median_h * 0.75)
        rows = []
        for cand in sorted(cands, key=lambda t: t[1] + t[3] / 2.0):
            _, y, _, hh = cand
            mid_y = y + hh / 2.0
            if not rows or abs(mid_y - rows[-1]["mid_y"]) > row_tol:
                rows.append({"mid_y": mid_y, "items": [cand]})
            else:
                rows[-1]["items"].append(cand)
                mids = [it[1] + it[3] / 2.0 for it in rows[-1]["items"]]
                rows[-1]["mid_y"] = sum(mids) / max(len(mids), 1)

        row_specs = []
        roi_h = float(max(gray_roi.shape[0], 1))
        for row in rows:
            items = sorted(row["items"], key=lambda t: t[0])
            row_specs.append({
                "items": items,
                "count": len(items),
                "mid_rel": float(row["mid_y"]) / roi_h,
                "total_width": sum(it[2] for it in items),
            })

        def choose_row(expected_rel, min_count, max_count, preferred_count=3):
            candidates = [r for r in row_specs if min_count <= r["count"] <= max_count]
            if not candidates:
                return None
            return min(
                candidates,
                key=lambda r: (abs(r["mid_rel"] - expected_rel), abs(r["count"] - preferred_count), -r["total_width"]),
            )

        record_row = choose_row(expected_rel=0.48, min_count=1, max_count=5, preferred_count=4)
        dob_row = choose_row(expected_rel=0.61, min_count=1, max_count=4, preferred_count=3)
        membership_row = choose_row(expected_rel=0.38, min_count=1, max_count=3, preferred_count=1)

        def to_rect(cand):
            x, y, ww, hh = cand
            return fitz.Rect(x / ZOOM, (y + y0_roi) / ZOOM, (x + ww) / ZOOM, (y + hh + y0_roi) / ZOOM)

        if record_row:
            rects = [to_rect(c) for c in record_row["items"][:4]]
            if len(rects) == 1:
                targets["record_identity.full_name"] = rects
            elif len(rects) == 2:
                targets["record_identity.first_name"] = [rects[0]]
                targets["record_identity.last_name"] = [rects[1]]
            elif len(rects) == 3:
                targets["record_identity.first_name"] = [rects[0]]
                targets["record_identity.middle_name"] = [rects[1]]
                targets["record_identity.last_name"] = [rects[2]]
            else:
                targets["record_identity.first_name"] = [rects[0]]
                targets["record_identity.middle_name"] = [rects[1]]
                targets["record_identity.last_name"] = [rects[2]]
                targets["record_identity.suffix"] = [rects[3]]

        if dob_row:
            rects = [to_rect(c) for c in dob_row["items"][:3]]
            for key, rect in zip(DOB_TARGET_KEYS, rects):
                targets[key] = [rect]

        if membership_row:
            rects = [to_rect(c) for c in membership_row["items"][:1]]
            if rects:
                targets["record_identity.membership_number"] = rects

        if not any(targets.values()):
            def get_band(l, h):
                band = [c for c in cands if l <= c[1] / gray_roi.shape[0] <= h]
                return [to_rect(c) for c in sorted(band, key=lambda t: t[0])]
            fallback_geom = {
                "names": get_band(0.46, 0.50),
                "dob": get_band(0.56, 0.66),
                "phil": get_band(0.36, 0.40),
            }
            targets = self._semantic_targets_from_geom(fallback_geom)

        return self._normalize_semantic_targets(targets)

    # --------------------------------------------------------
    # Main detection entry
    # --------------------------------------------------------
    def run_detection(self, page_idx=0):
        if not self.pdf_path or not os.path.exists(self.pdf_path):
            raise FileNotFoundError("PDF path is missing or invalid.")
        if not self.supports_detection_backend():
            raise RuntimeError("No PDF rasterization backend is available for detection in this build.")

        self.all_boxes = []
        self.box_types = []
        self.geom = self._empty_geom()
        self.semantic_targets = self._empty_semantic_targets()

        img = self._render_pdf_page_bgr(self.pdf_path, page_idx=page_idx, preview_zoom=ZOOM)
        if img is None or getattr(img, "size", 0) == 0:
            raise ValueError("Failed to render PDF page into image.")

        if len(img.shape) == 2:
            img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

        h_img, w_img = img.shape[:2]
        text_like_mask, text_like_meta = self.build_text_like_mask(img)
        header_cut = int((text_like_meta or {}).get("header_cut", max(8, h_img * 0.16)) or 0)

        field_area_thresh = int(self.settings["F_Area"])
        field_min_w_val = int(self.settings["F_MinW"])
        field_min_h_val = int(self.settings["F_MinH"])
        field_close_k = int(self.settings["F_Close"])
        line_min_w_val = int(self.settings["Line_MinW"])
        line_max_w_val = int(self.settings["Line_MaxW"])
        checkbox_min_sz = int(self.settings["C_Size"][0])
        checkbox_max_sz = int(self.settings["C_Size"][1])

        red_border_min = float(self.settings["C_Border"])
        red_inner_max = float(self.settings["C_Inner"])
        red_roi_max_val = int(self.settings["ROI_Max"])
        red_open_k = int(self.settings["C_Open"])
        red_close_k = int(self.settings["C_Close"])
        red_band_pct = float(self.settings["C_BandPct"])
        red_aspect_tol = float(self.settings["C_AspectTol"])
        red_extent_low_val = float(self.settings["Ext_Low"])
        red_extent_high_val = float(self.settings["Ext_High"])
        red_fill_min = float(self.settings["C_FillMin"])
        red_eps_val = float(self.settings["C_Eps"])
        red_use_extent_val = bool(self.settings["Use_Extent"])

        # --- Semantic auto-targeting disabled (manual mapping only)
        self.geom = self._empty_geom()
        self.semantic_targets = self._empty_semantic_targets()
        self.semantic_targets_by_page[page_idx] = self._empty_semantic_targets()

        # --- General detection
        full_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        base_bin = cv2.threshold(full_gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

        bin_checks = base_bin.copy()
        if red_close_k > 0:
            bin_checks = cv2.morphologyEx(
                bin_checks,
                cv2.MORPH_CLOSE,
                np.ones((red_close_k, red_close_k), np.uint8)
            )
        if red_open_k > 0:
            bin_checks = cv2.morphologyEx(
                bin_checks,
                cv2.MORPH_OPEN,
                np.ones((red_open_k, red_open_k), np.uint8)
            )

        if field_close_k > 0:
            bin_fields = cv2.morphologyEx(
                base_bin,
                cv2.MORPH_CLOSE,
                np.ones((field_close_k, field_close_k), np.uint8)
            )
        else:
            bin_fields = base_bin

        self.all_boxes = []
        self.box_types = []

        if page_idx == 0:
            self._append_geom_fields()

        # --- answer lines
        line_rects = self.find_answer_lines(
            img,
            ZOOM,
            min_line_w=line_min_w_val,
            max_line_w=line_max_w_val
        )
        for lr in line_rects:
            self.all_boxes.append(lr)
            self.box_types.append("line")

        # --- color-assisted red checkbox detection
        red_color_boxes = self.find_red_checkbox_candidates(
            img,
            zoom_factor=ZOOM,
            min_sz=max(7, checkbox_min_sz - 5),
            max_sz=min(int(checkbox_max_sz), 42),
            border_min=max(0.05, float(red_border_min) * 0.7),
            inner_max=min(0.75, float(red_inner_max) + 0.20),
            band_pct=max(0.10, min(float(red_band_pct), 0.16)),
            aspect_tol=max(0.30, float(red_aspect_tol)),
            fill_min=max(0.12, float(red_fill_min) * 0.45),
            eps=max(0.04, float(red_eps_val))
        )
        for rr in red_color_boxes:
            if self.should_reject_header_word_check_artifact(rr, text_like_mask, preview_zoom=ZOOM, header_cut=header_cut):
                continue
            self._append_box_unique(rr, "check", iou_thresh=0.45)

        # --- grayscale checkbox detection
        nf, _, sf, _ = cv2.connectedComponentsWithStats(bin_checks)
        small_min = max(7, int(checkbox_min_sz) - 5)

        for i in range(1, nf):
            x, y, w, h, area = sf[i]
            if w >= w_img * 0.9:
                continue

            aspect = w / float(h) if h > 0 else 999

            if (
                small_min <= w <= checkbox_max_sz and
                small_min <= h <= checkbox_max_sz and
                0.68 <= aspect <= 1.45
            ):
                if self.looks_like_checkbox(bin_checks, x, y, w, h, area):
                    candidate_rect = fitz.Rect(x / ZOOM, y / ZOOM, (x + w) / ZOOM, (y + h) / ZOOM)
                    if self.should_reject_header_word_check_artifact(candidate_rect, text_like_mask, preview_zoom=ZOOM, header_cut=header_cut):
                        continue
                    self._append_box_unique(
                        candidate_rect,
                        "check",
                        iou_thresh=0.45
                    )

            elif (
                max(w, h) <= int(red_roi_max_val) and
                min(w, h) >= small_min and
                0.55 <= aspect <= 1.75
            ):
                for fx, fy, fw, fh in self.find_checkbox_rects_in_roi(bin_checks, x, y, w, h):
                    candidate_rect = fitz.Rect(fx / ZOOM, fy / ZOOM, (fx + fw) / ZOOM, (fy + fh) / ZOOM)
                    if self.should_reject_header_word_check_artifact(candidate_rect, text_like_mask, preview_zoom=ZOOM, header_cut=header_cut):
                        continue
                    self._append_box_unique(
                        candidate_rect,
                            "check",
                            iou_thresh=0.45
                        )

            elif (
                max(w, h) <= int(red_roi_max_val) and
                min(w, h) >= small_min and
                0.60 <= aspect <= 1.60
            ):
                for fx, fy, fw, fh in self.find_checkbox_rects_in_roi(bin_checks, x, y, w, h):
                    candidate_rect = fitz.Rect(fx / ZOOM, fy / ZOOM, (fx + fw) / ZOOM, (fy + fh) / ZOOM)
                    if self.should_reject_header_word_check_artifact(candidate_rect, text_like_mask, preview_zoom=ZOOM, header_cut=header_cut):
                        continue
                    self._append_box_unique(
                        candidate_rect,
                            "check",
                            iou_thresh=0.45
                        )

            elif max(w, h) <= int(red_roi_max_val) and min(w, h) >= checkbox_min_sz:
                for fx, fy, fw, fh in self.find_checkbox_rects_in_roi(bin_checks, x, y, w, h):
                    candidate_rect = fitz.Rect(fx / ZOOM, fy / ZOOM, (fx + fw) / ZOOM, (fy + fh) / ZOOM)
                    if self.should_reject_header_word_check_artifact(candidate_rect, text_like_mask, preview_zoom=ZOOM, header_cut=header_cut):
                        continue
                    self._append_box_unique(
                        candidate_rect,
                            "check",
                            iou_thresh=0.45
                        )

        # --- field detection
        nf2, _, sf2, _ = cv2.connectedComponentsWithStats(bin_fields)
        for i in range(1, nf2):
            x, y, w, h, area = sf2[i]

            if w >= w_img * 0.90 or h >= h_img * 0.12 or area >= (w_img * h_img * 0.03):
                continue

            if (w > 1200 and h > 120) or (w > 1600 and h > 80):
                continue

            if h < field_min_h_val or w < field_min_w_val or area < field_area_thresh:
                continue

            refined_rect = self._refine_field_rect_from_mask(
                bin_fields,
                x, y, w, h,
                zoom_factor=ZOOM,
                row_frac_thresh=0.45,
                col_frac_thresh=0.18,
                min_inner_h=max(18, int(field_min_h_val * 0.45)),
                pad_px=2
            )

            if self.should_reject_header_word_field_artifact(refined_rect, text_like_mask, preview_zoom=ZOOM, header_cut=header_cut):
                continue
            self.all_boxes.append(refined_rect)
            self.box_types.append("field")

        self._cleanup_field_fragments()
        self._cleanup_line_field_conflicts()
        self._remove_detections_inside_pdf_character_regions(
            page_idx,
            char_regions=self._pdf_character_regions_for_page(page_idx),
            text_mask=text_like_mask,
            preview_zoom=ZOOM,
            header_cut=header_cut,
        )
        self.detected_page_idx = int(page_idx)
        self.detected_settings_signature = self._settings_signature()
        self._cache_detection_for_page(page_idx)

    # --------------------------------------------------------
    # Mapping helpers
    # --------------------------------------------------------
    def _mapping_rect_list(self, item):
        rects = item.get("rects")
        if rects:
            out = []
            for r in rects:
                out.append(r if isinstance(r, fitz.Rect) else fitz.Rect(*r))
            return sorted(out, key=lambda rr: (round(rr.y0, 3), rr.x0))

        r = item.get("rect")
        if r is None:
            return []
        return [r if isinstance(r, fitz.Rect) else fitz.Rect(*r)]

    def _rects_refer_to_same_target(self, a, b, tol=0.01):
        if (
            abs(a.x0 - b.x0) < tol and
            abs(a.y0 - b.y0) < tol and
            abs(a.x1 - b.x1) < tol and
            abs(a.y1 - b.y1) < tol
        ):
            return True

        acx = (a.x0 + a.x1) / 2.0
        acy = (a.y0 + a.y1) / 2.0
        bcx = (b.x0 + b.x1) / 2.0
        bcy = (b.y0 + b.y1) / 2.0

        if (b.x0 <= acx <= b.x1 and b.y0 <= acy <= b.y1):
            return True
        if (a.x0 <= bcx <= a.x1 and a.y0 <= bcy <= a.y1):
            return True

        inter = _rect_intersection_area(a, b)
        if inter <= 0:
            return False

        a_area = max(a.width * a.height, 1e-6)
        b_area = max(b.width * b.height, 1e-6)
        iou = inter / (a_area + b_area - inter)
        cover_small = inter / min(a_area, b_area)

        return iou >= 0.60 or cover_small >= 0.80

    def _mapping_match_score(self, target_rect, mapped_rect, tol=0.01):
        if (
            abs(target_rect.x0 - mapped_rect.x0) < tol and
            abs(target_rect.y0 - mapped_rect.y0) < tol and
            abs(target_rect.x1 - mapped_rect.x1) < tol and
            abs(target_rect.y1 - mapped_rect.y1) < tol
        ):
            return 9999.0

        inter = _rect_intersection_area(target_rect, mapped_rect)
        if inter <= 0:
            return -1.0

        t_area = max(target_rect.width * target_rect.height, 1e-6)
        iou = inter / (t_area + max(mapped_rect.width * mapped_rect.height, 1e-6) - inter)
        cover_target = inter / t_area

        cx = (target_rect.x0 + target_rect.x1) / 2.0
        cy = (target_rect.y0 + target_rect.y1) / 2.0
        center_inside = (mapped_rect.x0 <= cx <= mapped_rect.x1) and (mapped_rect.y0 <= cy <= mapped_rect.y1)

        score = 0.0
        if center_inside:
            score += 1.0
        score += iou * 10.0
        score += cover_target * 5.0
        return score


    def _mobile_repair_overlap_score(self, a, b):
        try:
            ra = a if isinstance(a, fitz.Rect) else fitz.Rect(*a)
            rb = b if isinstance(b, fitz.Rect) else fitz.Rect(*b)
            inter = _rect_intersection_area(ra, rb)
            if inter <= 0:
                return 0.0
            area_a = max(float(ra.width) * float(ra.height), 1e-6)
            area_b = max(float(rb.width) * float(rb.height), 1e-6)
            iou = inter / float(max(area_a + area_b - inter, 1e-6))
            cover = inter / float(max(min(area_a, area_b), 1e-6))
            return float(max(iou, cover))
        except Exception:
            return 0.0

    def describe_box_mapping(self, box_idx, page_idx):
        if box_idx < 0 or box_idx >= len(self.all_boxes):
            return "EMPTY"

        target_rect = self.all_boxes[box_idx]
        best_hit = None
        best_score = -1.0

        for map_key, configs in self.custom_mappings.items():
            for c in configs:
                if c.get("page", 0) != page_idx:
                    continue

                rects = self._mapping_rect_list(c)
                if not rects:
                    continue

                for r in rects:
                    score = self._mapping_match_score(target_rect, r)
                    if score > best_score:
                        best_score = score
                        best_hit = c

        parts = []
        if best_hit is not None and best_score >= 1.0:
            col = str(best_hit.get("column", "")).strip()
            trig = str(best_hit.get("trigger", "")).strip()
            is_grid = bool(best_hit.get("g", False))
            grid_n = int(best_hit.get("n", 1))

            parts.append(f"Mapped: {col}")
            if trig:
                parts.append(f"trigger={trig}")
            if is_grid:
                parts.append(f"grid={grid_n}")

        return " | ".join(parts) if parts else "EMPTY"

    def get_box_mapping_payload(self, box_id, page_idx):
        if box_id < 0 or box_id >= len(self.all_boxes):
            return {"box_id": box_id, "page": page_idx, "status": "EMPTY", "entries": []}

        target_rect = self.all_boxes[box_id]
        best_hit = None
        best_score = -1.0

        for map_key, configs in self.custom_mappings.items():
            for c in configs:
                if c.get("page", 0) != page_idx:
                    continue

                rects = self._mapping_rect_list(c)
                if not rects:
                    continue

                for r in rects:
                    score = self._mapping_match_score(target_rect, r)
                    if score > best_score:
                        best_score = score
                        best_hit = c

        if best_hit is not None and best_score >= 1.0:
            return {
                "box_id": box_id,
                "page": page_idx,
                "status": "MAPPED",
                "semantic_target": None,
                "entries": [
                    {
                        "column": str(best_hit.get("column", "")),
                        "trigger": str(best_hit.get("trigger", "")),
                        "g": bool(best_hit.get("g", False)),
                        "n": int(best_hit.get("n", 1)),
                    }
                ]
            }

        return {
            "box_id": box_id,
            "page": page_idx,
            "status": "EMPTY",
            "semantic_target": None,
            "entries": []
        }

    def assign_mapping(self, box_ids, column, trigger, is_grid, grid_n, page_idx):
        selected_rects = []
        for b_id in box_ids:
            if b_id < 0 or b_id >= len(self.all_boxes):
                raise ValueError(f"Box ID {b_id} is out of range.")
            selected_rects.append(self.all_boxes[b_id])

        selected_rects = sorted(selected_rects, key=lambda rr: (round(rr.y0, 3), rr.x0))
        map_key = f"{str(column).strip()}_{str(trigger).strip()}"

        def mapping_conflicts_with_selected(mapping_item, selected_rects_):
            existing_rects = self._mapping_rect_list(mapping_item)
            if not existing_rects:
                return False

            for er in existing_rects:
                for sr in selected_rects_:
                    if self._rects_refer_to_same_target(er, sr, tol=0.20):
                        return True
            return False

        for k in list(self.custom_mappings.keys()):
            kept = []
            for m in self.custom_mappings[k]:
                if m.get("page", 0) != page_idx:
                    kept.append(m)
                    continue

                if mapping_conflicts_with_selected(m, selected_rects):
                    continue
                kept.append(m)

            if kept:
                self.custom_mappings[k] = kept
            else:
                del self.custom_mappings[k]

        if map_key not in self.custom_mappings:
            self.custom_mappings[map_key] = []

        self.custom_mappings[map_key].append({
            "column": str(column).strip(),
            "trigger": str(trigger).strip(),
            "rects": selected_rects,
            "page": int(page_idx),
            "g": bool(is_grid),
            "n": int(grid_n),
        })

# =========================
# PART 3 / 4
# Fill + Config Engine
# Add this BELOW Part 2, inside the FormAlchemistEngine class
# =========================

    # --------------------------------------------------------
    # Drawing helpers
    # --------------------------------------------------------
    def _rect_union(self, rects):
        return fitz.Rect(
            min(r.x0 for r in rects),
            min(r.y0 for r in rects),
            max(r.x1 for r in rects),
            max(r.y1 for r in rects)
        )

    def _allocate_cells_by_width(self, rects, total_cells):
        total_cells = max(int(total_cells), 1)
        widths = [max(float(r.width), 1e-6) for r in rects]
        total_w = sum(widths)
        raw = [(w / total_w) * total_cells for w in widths]
        base = [int(np.floor(v)) for v in raw]
        remain = total_cells - sum(base)

        order = sorted(
            range(len(raw)),
            key=lambda i: (raw[i] - base[i]),
            reverse=True
        )
        for i in order[:remain]:
            base[i] += 1

        if sum(base) == 0 and base:
            base[-1] = total_cells

        return base

    def _draw_single_rect_text(self, page, val, rect, ox=0, oy=0, fs_scale=0.60):
        fs = rect.height * fs_scale
        est_w = len(val) * fs * 0.6
        p = fitz.Point(
            rect.x0 + max((rect.width - est_w) / 2, 1) + ox,
            rect.y1 - (rect.height * 0.2) + oy
        )
        page.insert_text(p, val, fontsize=fs, fontname="helv")

    def _draw_grid_rect_text(self, page, val, rect, grid_n=1, ox=0, oy=0, fs_scale=0.60):
        grid_n = max(int(grid_n), 1)
        fs = rect.height * fs_scale
        cell_w = rect.width / grid_n

        for i, ch in enumerate(val[:grid_n]):
            p = fitz.Point(
                rect.x0 + i * cell_w + (cell_w * 0.25) + ox,
                rect.y1 - (rect.height * 0.2) + oy
            )
            page.insert_text(p, ch, fontsize=fs, fontname="helv")

    def draw_logic(self, page, text, rect_or_rects, is_grid=False, grid_n=1, ox=0, oy=0, fs_scale=0.65):
        val = str(text).strip()
        if not val or val.lower() in ["nan", "none"]:
            return

        if val.endswith(".0"):
            val = val[:-2]
        val = val.upper()

        rects = rect_or_rects if isinstance(rect_or_rects, list) else [rect_or_rects]
        rects = [r if isinstance(r, fitz.Rect) else fitz.Rect(*r) for r in rects]
        rects = sorted(rects, key=lambda rr: (round(rr.y0, 3), rr.x0))
        if not rects:
            return

        if is_grid and len(rects) > 1:
            counts = self._allocate_cells_by_width(rects, grid_n)
            pos = 0
            for rect, n_cells in zip(rects, counts):
                if n_cells <= 0:
                    continue
                seg = val[pos:pos + n_cells]
                if not seg:
                    break
                self._draw_grid_rect_text(page, seg, rect, grid_n=n_cells, ox=ox, oy=oy, fs_scale=fs_scale)
                pos += n_cells
            return

        if is_grid:
            target = rects[0] if len(rects) == 1 else self._rect_union(rects)
            self._draw_grid_rect_text(page, val, target, grid_n=grid_n, ox=ox, oy=oy, fs_scale=fs_scale)
        else:
            target = rects[0] if len(rects) == 1 else self._rect_union(rects)
            self._draw_single_rect_text(page, val, target, ox=ox, oy=oy, fs_scale=fs_scale)

    # --------------------------------------------------------
    # Document processing / export
    # --------------------------------------------------------
    def _get_patient_row(self, patient_name, record_key=None):
        return self._get_record_row(record_ref=patient_name, record_key=record_key)

    def _coerce_rect(self, rect):
        if isinstance(rect, fitz.Rect):
            return rect
        if isinstance(rect, RectCompat):
            return rect
        return fitz.Rect(*rect)

    def _collect_overlay_ops(self, patient_name, page_idx=0, record_key=None):
        row = self._get_patient_row(patient_name, record_key=record_key)
        ops = []

        for configs in self.custom_mappings.values():
            for c in configs:
                if c.get("page", 0) != page_idx:
                    continue

                target_rects = [self._coerce_rect(r) for r in self._mapping_rect_list(c)]
                if not target_rects:
                    continue

                csv_val = str(row.get(c["column"], "")).strip()
                trigger = str(c.get("trigger", "")).strip()

                if trigger and csv_val.upper() == trigger.upper():
                    ops.append({
                        "kind": "check",
                        "rects": target_rects,
                    })
                elif not trigger and csv_val:
                    ops.append({
                        "kind": "text",
                        "text": csv_val,
                        "rects": target_rects,
                        "grid": bool(c.get("g", False)),
                        "grid_n": int(c.get("n", 1)),
                    })

        return ops

    def _collect_overlay_ops_by_page(self, patient_name, record_key=None):
        row = self._get_patient_row(patient_name, record_key=record_key)
        ops_by_page = {}

        for configs in self.custom_mappings.values():
            for c in configs:
                page_idx = int(c.get("page", 0) or 0)
                target_rects = [self._coerce_rect(r) for r in self._mapping_rect_list(c)]
                if not target_rects:
                    continue

                csv_val = str(row.get(c["column"], "")).strip()
                trigger = str(c.get("trigger", "")).strip()

                if trigger:
                    if csv_val.upper() != trigger.upper():
                        continue
                    ops_by_page.setdefault(page_idx, []).append({
                        "kind": "check",
                        "rects": target_rects,
                    })
                elif csv_val:
                    ops_by_page.setdefault(page_idx, []).append({
                        "kind": "text",
                        "text": csv_val,
                        "rects": target_rects,
                        "grid": bool(c.get("g", False)),
                        "grid_n": int(c.get("n", 1)),
                    })

        return ops_by_page

    def _reportlab_baseline_y(self, page_height, rect, oy=0):
        return float(page_height) - (float(rect.y1) - (float(rect.height) * 0.2) + float(oy))

    def _write_text_op_reportlab(self, c, page_height, text, rects, is_grid=False, grid_n=1, ox=0, oy=0, fs_scale=0.65):
        val = str(text or "").strip()
        if not val or val.lower() in ["nan", "none"]:
            return
        if val.endswith(".0"):
            val = val[:-2]
        val = val.upper()

        rects = sorted([self._coerce_rect(r) for r in rects], key=lambda rr: (round(rr.y0, 3), rr.x0))
        if not rects:
            return

        def draw_single(single_val, rect):
            fs = max(float(rect.height) * fs_scale, 6.0)
            text_w = rl_string_width(single_val, "Helvetica", fs)
            x = float(rect.x0) + max((float(rect.width) - float(text_w)) / 2.0, 1.0) + float(ox)
            y = self._reportlab_baseline_y(page_height, rect, oy=oy)
            c.setFont("Helvetica", fs)
            c.drawString(x, y, single_val)

        def draw_grid(grid_val, rect, cells):
            cells = max(int(cells), 1)
            fs = max(float(rect.height) * 0.60, 6.0)
            cell_w = float(rect.width) / cells
            y = self._reportlab_baseline_y(page_height, rect, oy=oy)
            c.setFont("Helvetica", fs)
            for i, ch in enumerate(grid_val[:cells]):
                x = float(rect.x0) + (i * cell_w) + (cell_w * 0.25) + float(ox)
                c.drawString(x, y, ch)

        if is_grid and len(rects) > 1:
            counts = self._allocate_cells_by_width(rects, grid_n)
            pos = 0
            for rect, n_cells in zip(rects, counts):
                seg = val[pos:pos + n_cells]
                if seg:
                    draw_grid(seg, rect, n_cells)
                pos += n_cells
            return

        target = rects[0] if len(rects) == 1 else self._rect_union(rects)
        if is_grid:
            draw_grid(val, target, grid_n)
        else:
            draw_single(val, target)

    def _write_check_op_reportlab(self, c, page_height, rects):
        rects = sorted([self._coerce_rect(r) for r in rects], key=lambda rr: (round(rr.y0, 3), rr.x0))
        if not rects:
            return
        target = rects[0] if len(rects) == 1 else self._rect_union(rects)
        fs = max(float(target.height) * 0.95, 8.0)
        x = float(target.x0) + (float(target.width) * 0.15)
        y = float(page_height) - (float(target.y1) - (float(target.height) * 0.15))
        c.setFont("Helvetica-Bold", fs)
        c.drawString(x, y, "X")

    def _build_filled_pdf_bytes_raster(self, patient_name, page_idx=0, record_key=None, preview_zoom=2.0, export_scope="document"):
        if not self.pdf_path or not os.path.exists(self.pdf_path):
            raise FileNotFoundError("PDF path is missing or invalid.")

        with open(self.pdf_path, "rb") as fh:
            reader = PdfReader(fh)
            total_pages = len(reader.pages)
        if total_pages <= 0:
            raise ValueError("PDF has no pages.")

        export_scope = str(export_scope or "document").strip().lower()
        target_idx = min(max(int(page_idx), 0), total_pages - 1)
        if export_scope == "page":
            ops_by_page = {target_idx: self._collect_overlay_ops(patient_name, page_idx=target_idx, record_key=record_key)}
        else:
            ops_by_page = self._collect_overlay_ops_by_page(patient_name, record_key=record_key)

        pil_pages = []

        for i in range(total_pages):
            img = self._render_pdf_page_bgr(self.pdf_path, page_idx=i, preview_zoom=preview_zoom)
            if img is None or getattr(img, "size", 0) == 0:
                raise ValueError(f"Failed to render page {i + 1} for raster export.")
            for op in ops_by_page.get(i, []):
                if op.get("kind") == "check":
                    self._draw_check_op_cv(img, op.get("rects", []), preview_zoom=preview_zoom)
                else:
                    self._draw_text_op_cv(
                        img,
                        op.get("text", ""),
                        op.get("rects", []),
                        preview_zoom=preview_zoom,
                        is_grid=bool(op.get("grid", False)),
                        grid_n=int(op.get("grid_n", 1)),
                    )
            rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            pil_pages.append(PILImage.fromarray(rgb))

        out_buf = io.BytesIO()
        first = pil_pages[0].convert("RGB")
        rest = [im.convert("RGB") for im in pil_pages[1:]]
        first.save(
            out_buf,
            format="PDF",
            save_all=True,
            append_images=rest,
            resolution=max(float(preview_zoom), 1.0) * 72.0,
        )
        return out_buf.getvalue()

    def _build_filled_pdf_bytes(self, patient_name, page_idx=0, record_key=None, export_scope="document"):
        if not self.supports_export_backend():
            raise RuntimeError("PDF export backend is unavailable in this build.")
        return self._build_filled_pdf_bytes_raster(patient_name, page_idx=page_idx, record_key=record_key, export_scope=export_scope)

    def export_filled_pdf(self, patient_name, out_path, page_idx=0, record_key=None, export_scope="document"):
        pdf_bytes = self._build_filled_pdf_bytes(patient_name, page_idx=page_idx, record_key=record_key, export_scope=export_scope)
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        with open(out_path, "wb") as fh:
            fh.write(pdf_bytes)

    def process_doc(self, patient_name, page_idx=0):
        """
        Backward-compatible helper. On desktops with PyMuPDF, returns an in-memory
        document object. On Android-safe paths, use export_filled_pdf() instead.
        """
        pdf_bytes = self._build_filled_pdf_bytes(patient_name, page_idx=page_idx)
        if FITZ_AVAILABLE:
            return fitz.open(stream=pdf_bytes, filetype="pdf")
        raise RuntimeError("process_doc() requires the optional PyMuPDF runtime. Use export_filled_pdf() for Android-safe export.")
    # --------------------------------------------------------
    # Config helpers
    # --------------------------------------------------------
    def _norm_str(self, v):
        return str(v or "").strip()

    def _rect_close(self, a, b, tol=0.20):
        return (
            abs(a.x0 - b.x0) <= tol and
            abs(a.y0 - b.y0) <= tol and
            abs(a.x1 - b.x1) <= tol and
            abs(a.y1 - b.y1) <= tol
        )

    def _rect_list_close(self, rects_a, rects_b, tol=0.20):
        if len(rects_a) != len(rects_b):
            return False

        aa = sorted(rects_a, key=lambda r: (round(r.y0, 3), r.x0, r.x1, r.y1))
        bb = sorted(rects_b, key=lambda r: (round(r.y0, 3), r.x0, r.x1, r.y1))
        return all(self._rect_close(a, b, tol=tol) for a, b in zip(aa, bb))

    def _mapping_entry_to_current(self, item, fallback_key=""):
        column = self._norm_str(item.get("column", ""))
        trigger = self._norm_str(item.get("trigger", ""))

        if not column and fallback_key:
            if "_" in fallback_key:
                parts = fallback_key.rsplit("_", 1)
                if len(parts) == 2:
                    column = self._norm_str(parts[0])
                    if not trigger:
                        trigger = self._norm_str(parts[1])
            else:
                column = self._norm_str(fallback_key)

        rects = []
        if item.get("rects"):
            for r in item.get("rects", []):
                if isinstance(r, fitz.Rect):
                    rects.append(r)
                elif isinstance(r, (list, tuple)) and len(r) == 4:
                    rects.append(fitz.Rect(*r))
        elif item.get("rect") is not None:
            r = item.get("rect")
            if isinstance(r, fitz.Rect):
                rects = [r]
            elif isinstance(r, (list, tuple)) and len(r) == 4:
                rects = [fitz.Rect(*r)]

        rects = sorted(rects, key=lambda rr: (round(rr.y0, 3), rr.x0, rr.x1, rr.y1))

        return {
            "column": column,
            "trigger": trigger,
            "rects": rects,
            "page": int(item.get("page", 0)),
            "g": bool(item.get("g", item.get("is_grid", False))),
            "n": int(item.get("n", item.get("grid_n", 1))),
        }

    def _extract_current_mapping_entries(self, mapping_dict):
        out = []
        for map_key, lst in mapping_dict.items():
            for item in lst:
                out.append(self._mapping_entry_to_current(item, fallback_key=map_key))
        return out

    def _extract_cfg_mapping_entries(self, cfg):
        out = []
        loaded = cfg.get("custom_mappings", {}) or {}
        for map_key, lst in loaded.items():
            if isinstance(lst, dict):
                lst = [lst]
            for item in lst:
                out.append(self._mapping_entry_to_current(item, fallback_key=map_key))
        return out

    def _same_mapping_identity(self, a, b, tol=0.20):
        return (
            int(a.get("page", 0)) == int(b.get("page", 0)) and
            self._norm_str(a.get("column", "")).upper() == self._norm_str(b.get("column", "")).upper() and
            self._norm_str(a.get("trigger", "")).upper() == self._norm_str(b.get("trigger", "")).upper() and
            self._rect_list_close(a.get("rects", []), b.get("rects", []), tol=tol)
        )

    def _same_target_rects(self, a, b, tol=0.20):
        return (
            int(a.get("page", 0)) == int(b.get("page", 0)) and
            self._rect_list_close(a.get("rects", []), b.get("rects", []), tol=tol)
        )

    def _entries_to_mapping_dict(self, entries):
        out = {}
        for e in entries:
            map_key = f"{self._norm_str(e.get('column', ''))}_{self._norm_str(e.get('trigger', ''))}"
            if map_key not in out:
                out[map_key] = []

            out[map_key].append({
                "column": self._norm_str(e.get("column", "")),
                "trigger": self._norm_str(e.get("trigger", "")),
                "rects": [r if isinstance(r, fitz.Rect) else fitz.Rect(*r) for r in e.get("rects", [])],
                "page": int(e.get("page", 0)),
                "g": bool(e.get("g", False)),
                "n": int(e.get("n", 1)),
            })
        return out

    def _learning_setting_keys(self):
        return (
            "F_Area", "F_MinW", "F_MinH", "F_Close",
            "Line_MinW", "Line_MaxW",
            "C_Strict", "C_Size", "C_Border", "C_Inner", "ROI_Max",
            "C_Open", "C_Close", "C_BandPct", "C_AspectTol",
            "Ext_Low", "Ext_High", "C_FillMin", "C_Eps", "Use_Extent",
            "Is_Grid", "Grid_N",
        )

    def _collect_learning_profile_settings(self):
        out = {}
        for key in self._learning_setting_keys():
            value = self.settings.get(key)
            if isinstance(value, tuple):
                value = list(value)
            out[key] = value
        return out

    def _apply_learning_profile_settings(self, settings_dict):
        if not isinstance(settings_dict, dict):
            return
        for key in self._learning_setting_keys():
            if key not in settings_dict:
                continue
            value = settings_dict.get(key)
            if value in {None, ""}:
                continue
            try:
                if key == "C_Size":
                    if isinstance(value, (list, tuple)) and len(value) == 2:
                        self.settings[key] = (int(float(value[0])), int(float(value[1])))
                    continue
                if key in {"F_Area", "F_MinW", "F_MinH", "F_Close", "Line_MinW", "Line_MaxW", "C_Strict", "ROI_Max", "C_Open", "C_Close", "Grid_N"}:
                    self.settings[key] = int(float(value))
                elif key in {"C_Border", "C_Inner", "C_BandPct", "C_AspectTol", "Ext_Low", "Ext_High", "C_FillMin", "C_Eps"}:
                    self.settings[key] = float(value)
                elif key in {"Use_Extent", "Is_Grid"}:
                    if isinstance(value, str):
                        self.settings[key] = value.strip().lower() in {"1", "true", "yes", "on"}
                    else:
                        self.settings[key] = bool(value)
                else:
                    self.settings[key] = value
            except Exception:
                continue

    def _serialize_rect(self, rect):
        rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
        return [float(rect.x0), float(rect.y0), float(rect.x1), float(rect.y1)]

    def _serialize_rects(self, rects):
        return [self._serialize_rect(r) for r in (rects or [])]

    def _serialize_geom_payload(self, geom=None):
        geom = geom or self.geom or {}
        return {
            "names": self._serialize_rects(geom.get("names", [])),
            "dob": self._serialize_rects(geom.get("dob", [])),
            "phil": self._serialize_rects(geom.get("phil", [])),
        }

    def _serialize_semantic_targets_payload(self, semantic_targets=None):
        semantic_targets = self._normalize_semantic_targets(semantic_targets or self.semantic_targets)
        return {key: self._serialize_rects(semantic_targets.get(key, [])) for key in SEMANTIC_TARGET_KEYS}

    def _serialize_detection_state(self, page_idx=None):
        return {
            "page_idx": int(self.detected_page_idx if page_idx is None else page_idx),
            "boxes": self._serialize_rects(self.all_boxes),
            "box_types": list(self.box_types or []),
            "geom": self._serialize_geom_payload(self.geom),
            "semantic_targets": self._serialize_semantic_targets_payload(self.semantic_targets),
            "selected_box_ids": list(self.selected_box_ids or []),
            "settings_signature": str(self.detected_settings_signature or ""),
        }

    def _page_family_hint(self):
        base = os.path.splitext(os.path.basename(str(self.pdf_path or "")))[0]
        base = re.sub(r"\d+", " ", base)
        base = re.sub(r"[_\-]+", " ", base)
        return _safe_slug(base)

    def build_page_fingerprint(self, page_idx=0, img_bgr=None):
        if img_bgr is None:
            img_bgr = self._render_pdf_page_bgr(self.pdf_path, page_idx=page_idx, preview_zoom=ZOOM)
        if img_bgr is None or getattr(img_bgr, "size", 0) == 0:
            raise ValueError("Unable to build page fingerprint from an empty image.")

        if len(img_bgr.shape) == 2:
            img_bgr = cv2.cvtColor(img_bgr, cv2.COLOR_GRAY2BGR)

        h, w = img_bgr.shape[:2]
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        mean_val = float(np.mean(gray))
        std_val = float(np.std(gray))
        _, binv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        ink_density = float(cv2.countNonZero(binv) / float(max(binv.size, 1)))

        h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(12, w // 30), 1))
        v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(12, h // 30)))
        horiz = cv2.morphologyEx(binv, cv2.MORPH_OPEN, h_kernel)
        vert = cv2.morphologyEx(binv, cv2.MORPH_OPEN, v_kernel)

        features = {
            "page_idx": int(page_idx),
            "width": int(w),
            "height": int(h),
            "aspect": round(float(w) / float(max(h, 1)), 6),
            "gray_mean": round(mean_val, 3),
            "gray_std": round(std_val, 3),
            "ink_density": round(ink_density, 6),
            "hline_density": round(float(cv2.countNonZero(horiz) / float(max(horiz.size, 1))), 6),
            "vline_density": round(float(cv2.countNonZero(vert) / float(max(vert.size, 1))), 6),
        }

        fingerprint = {
            "version": 1,
            "pdf_family_hint": self._page_family_hint(),
            "page_idx": int(page_idx),
            "page_count": int(self.total_pages() or 0) if self.pdf_path else 0,
            "features": features,
        }
        fingerprint["fingerprint_key"] = _sha1_json(fingerprint)
        return fingerprint

    def _score_profile_match(self, fingerprint, profile_entry):
        if not isinstance(profile_entry, dict):
            return 0.0
        score = 0.0
        if str(profile_entry.get("pdf_family_hint", "")) == str(fingerprint.get("pdf_family_hint", "")):
            score += 0.42
        if int(profile_entry.get("page_idx", -1)) == int(fingerprint.get("page_idx", -2)):
            score += 0.18

        pf = profile_entry.get("feature_snapshot", {}) or {}
        ff = fingerprint.get("features", {}) or {}
        for key, weight, scale in [
            ("aspect", 0.10, 0.20),
            ("ink_density", 0.12, 0.08),
            ("hline_density", 0.09, 0.06),
            ("vline_density", 0.09, 0.06),
            ("gray_std", 0.05, 45.0),
        ]:
            try:
                a = float(ff.get(key, 0.0))
                b = float(pf.get(key, 0.0))
                closeness = max(0.0, 1.0 - (abs(a - b) / float(scale)))
                score += weight * closeness
            except Exception:
                pass

        quality = float(profile_entry.get("avg_quality", 0.0) or 0.0)
        score += min(0.08, max(0.0, quality) * 0.08)
        if bool(profile_entry.get("approved", False)):
            score += 0.05
        return float(min(1.0, max(0.0, score)))

    def find_best_profile_match(self, fingerprint):
        profiles = (self.profile_memory or {}).get("profiles", {}) or {}
        fp_key = str((fingerprint or {}).get("fingerprint_key", "") or "")
        exact = profiles.get(fp_key)
        if isinstance(exact, dict):
            exact = dict(exact)
            exact["match_score"] = 1.0
            exact["match_kind"] = "exact"
            return exact

        best = None
        best_score = 0.0
        for entry in profiles.values():
            score = self._score_profile_match(fingerprint, entry)
            if score > best_score:
                best_score = score
                best = dict(entry)
        if best is not None:
            best["match_score"] = float(best_score)
            best["match_kind"] = "fuzzy"
        return best

    def prepare_learning_for_detection(self, page_idx=0, allow_profile_apply=True):
        if not self.learning_enabled or not self.pdf_path:
            self.current_detection_context = {
                "page_idx": int(page_idx),
                "profile_applied": False,
                "profile_apply_allowed": bool(allow_profile_apply),
            }
            return self.current_detection_context

        img = self._render_pdf_page_bgr(self.pdf_path, page_idx=page_idx, preview_zoom=ZOOM)
        fingerprint = self.build_page_fingerprint(page_idx=page_idx, img_bgr=img)
        profile = self.find_best_profile_match(fingerprint)
        applied = False
        if profile and isinstance(profile.get("settings"), dict):
            match_kind = str(profile.get("match_kind", ""))
            match_score = float(profile.get("match_score", 0.0) or 0.0)
            approved = bool(profile.get("approved", False))
            if allow_profile_apply and (match_kind == "exact" or (approved and match_score >= 0.92)):
                self._apply_learning_profile_settings(profile.get("settings", {}))
                applied = True
                self.last_applied_profile_meta = {
                    "profile_key": profile.get("fingerprint_key", ""),
                    "match_kind": match_kind,
                    "match_score": match_score,
                    "applied_at": _utc_now_iso(),
                }

        self.current_detection_context = {
            "started_at": _utc_now_iso(),
            "page_idx": int(page_idx),
            "fingerprint": fingerprint,
            "matched_profile": profile,
            "profile_applied": bool(applied),
            "profile_apply_allowed": bool(allow_profile_apply),
        }
        return self.current_detection_context

    def finalize_learning_after_detection(self, page_idx=0):
        ctx = dict(self.current_detection_context or {})
        ctx["completed_at"] = _utc_now_iso()
        ctx["page_idx"] = int(page_idx)
        ctx["raw_detection"] = self._serialize_detection_state(page_idx=page_idx)
        ctx["profile_used"] = self._collect_learning_profile_settings()
        self.current_detection_context = ctx
        self.persist_learning_session(current_page_idx=page_idx)
        return ctx

    def _rect_signature(self, rect, rect_type=""):
        rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
        return f"{rect_type}:{rect.x0:.3f}:{rect.y0:.3f}:{rect.x1:.3f}:{rect.y1:.3f}"

    def _normalize_learning_weight(self, reason="manual", approved=False, learning_weight=None):
        reason_key = str(reason or "manual").strip().lower()
        if approved:
            recommended = 1.0
            trust_class = "approved"
        elif reason_key.startswith("merge"):
            recommended = 0.25
            trust_class = "draft"
        elif reason_key.startswith("save") or reason_key in {"manual", "detect", "edit"}:
            recommended = 0.45 if reason_key == "manual" else 0.35
            trust_class = "saved"
        else:
            recommended = 0.30
            trust_class = "draft"

        try:
            requested = None if learning_weight is None else float(learning_weight)
        except Exception:
            requested = None

        if requested is None:
            effective = recommended
        elif approved:
            effective = max(requested, recommended)
        else:
            effective = max(0.0, min(requested, 0.95))
            if effective <= 0.0:
                effective = recommended
        return round(float(effective), 6), trust_class

    def _rect_iou(self, a, b):
        ra = a if isinstance(a, fitz.Rect) else fitz.Rect(*a)
        rb = b if isinstance(b, fitz.Rect) else fitz.Rect(*b)
        ix0 = max(ra.x0, rb.x0)
        iy0 = max(ra.y0, rb.y0)
        ix1 = min(ra.x1, rb.x1)
        iy1 = min(ra.y1, rb.y1)
        iw = max(0.0, ix1 - ix0)
        ih = max(0.0, iy1 - iy0)
        inter = iw * ih
        area_a = max(0.0, (ra.x1 - ra.x0) * (ra.y1 - ra.y0))
        area_b = max(0.0, (rb.x1 - rb.x0) * (rb.y1 - rb.y0))
        denom = area_a + area_b - inter
        return float(inter / denom) if denom > 0 else 0.0

    def _rect_change_metrics(self, raw_rect, final_rect):
        r0 = raw_rect if isinstance(raw_rect, fitz.Rect) else fitz.Rect(*raw_rect)
        r1 = final_rect if isinstance(final_rect, fitz.Rect) else fitz.Rect(*final_rect)
        raw_w = max(1.0, r0.width)
        raw_h = max(1.0, r0.height)
        final_w = max(1.0, r1.width)
        final_h = max(1.0, r1.height)
        center_shift = math.hypot(((r0.x0 + r0.x1) * 0.5) - ((r1.x0 + r1.x1) * 0.5), ((r0.y0 + r0.y1) * 0.5) - ((r1.y0 + r1.y1) * 0.5))
        size_shift = max(abs(raw_w - final_w), abs(raw_h - final_h))
        area_ratio = abs((raw_w * raw_h) - (final_w * final_h)) / float(max(raw_w * raw_h, final_w * final_h, 1.0))
        return {
            "center_shift": round(float(center_shift), 3),
            "size_shift": round(float(size_shift), 3),
            "area_ratio": round(float(area_ratio), 6),
            "iou": round(float(self._rect_iou(r0, r1)), 6),
        }

    def _build_delta_event(self, event_type, raw_rect=None, final_rect=None, raw_type="", final_type="", raw_idx=None, final_idx=None, note=""):
        payload = {
            "event_type": str(event_type or "change"),
            "raw_type": str(raw_type or ""),
            "final_type": str(final_type or ""),
            "raw_index": None if raw_idx is None else int(raw_idx),
            "final_index": None if final_idx is None else int(final_idx),
            "note": str(note or ""),
        }
        if raw_rect is not None:
            payload["raw_rect"] = self._serialize_rect(raw_rect)
            payload["raw_signature"] = self._rect_signature(raw_rect, raw_type)
        if final_rect is not None:
            payload["final_rect"] = self._serialize_rect(final_rect)
            payload["final_signature"] = self._rect_signature(final_rect, final_type)
        if raw_rect is not None and final_rect is not None:
            payload["metrics"] = self._rect_change_metrics(raw_rect, final_rect)
        return payload

    def _match_detection_pairs(self, raw_boxes, raw_types, final_boxes, final_types):
        exact_matches = []
        used_final = set()
        remaining_raw = []
        sig_to_final = {}
        for j, rect in enumerate(final_boxes):
            sig = self._rect_signature(rect, final_types[j] if j < len(final_types) else "")
            sig_to_final.setdefault(sig, []).append(j)

        for i, rect in enumerate(raw_boxes):
            sig = self._rect_signature(rect, raw_types[i] if i < len(raw_types) else "")
            bucket = sig_to_final.get(sig, [])
            picked = None
            while bucket:
                cand = bucket.pop(0)
                if cand not in used_final:
                    picked = cand
                    break
            if picked is not None:
                used_final.add(picked)
                exact_matches.append((i, picked))
            else:
                remaining_raw.append(i)

        remaining_final = [j for j in range(len(final_boxes)) if j not in used_final]
        fuzzy_matches = []
        score_pairs = []
        for i in remaining_raw:
            for j in remaining_final:
                iou = self._rect_iou(raw_boxes[i], final_boxes[j])
                if iou < 0.18:
                    continue
                type_bonus = 0.15 if ((raw_types[i] if i < len(raw_types) else "") == (final_types[j] if j < len(final_types) else "")) else 0.0
                score_pairs.append((iou + type_bonus, iou, i, j))
        score_pairs.sort(reverse=True)
        used_raw_fuzzy = set()
        used_final_fuzzy = set()
        for score, iou, i, j in score_pairs:
            if i in used_raw_fuzzy or j in used_final_fuzzy:
                continue
            metrics = self._rect_change_metrics(raw_boxes[i], final_boxes[j])
            center_shift = float(metrics.get("center_shift", 0.0) or 0.0)
            size_shift = float(metrics.get("size_shift", 0.0) or 0.0)
            if iou < 0.24 and center_shift > 18.0:
                continue
            if score < 0.24:
                continue
            used_raw_fuzzy.add(i)
            used_final_fuzzy.add(j)
            fuzzy_matches.append((i, j))

        unmatched_raw = [i for i in remaining_raw if i not in used_raw_fuzzy]
        unmatched_final = [j for j in remaining_final if j not in used_final_fuzzy]
        return exact_matches, fuzzy_matches, unmatched_raw, unmatched_final

    def _compute_learning_delta(self, raw_detection, final_detection):
        raw_boxes = raw_detection.get("boxes", []) if isinstance(raw_detection, dict) else []
        raw_types = raw_detection.get("box_types", []) if isinstance(raw_detection, dict) else []
        final_boxes = final_detection.get("boxes", []) if isinstance(final_detection, dict) else []
        final_types = final_detection.get("box_types", []) if isinstance(final_detection, dict) else []

        exact_matches, fuzzy_matches, unmatched_raw, unmatched_final = self._match_detection_pairs(raw_boxes, raw_types, final_boxes, final_types)
        added = []
        removed = []
        events = []
        stable_count = 0
        moved_count = 0
        resized_count = 0
        retyped_count = 0

        for i, j in exact_matches:
            stable_count += 1

        for i, j in fuzzy_matches:
            raw_type = raw_types[i] if i < len(raw_types) else ""
            final_type = final_types[j] if j < len(final_types) else ""
            metrics = self._rect_change_metrics(raw_boxes[i], final_boxes[j])
            center_shift = float(metrics.get("center_shift", 0.0) or 0.0)
            size_shift = float(metrics.get("size_shift", 0.0) or 0.0)
            area_ratio = float(metrics.get("area_ratio", 0.0) or 0.0)
            moved = center_shift >= 3.0
            resized = size_shift >= 3.0 or area_ratio >= 0.12
            retyped = raw_type != final_type
            if not (moved or resized or retyped):
                stable_count += 1
                continue
            if moved:
                moved_count += 1
            if resized:
                resized_count += 1
            if retyped:
                retyped_count += 1
            note_bits = []
            if retyped:
                note_bits.append(f"type {raw_type or '—'}→{final_type or '—'}")
            if moved:
                note_bits.append(f"move {center_shift:.1f}px")
            if resized:
                note_bits.append(f"resize {size_shift:.1f}px")
            event_type = "box_modified"
            if retyped and not (moved or resized):
                event_type = "box_retyped"
            elif resized and not moved and not retyped:
                event_type = "box_resized"
            elif moved and not resized and not retyped:
                event_type = "box_moved"
            events.append(self._build_delta_event(
                event_type,
                raw_rect=raw_boxes[i],
                final_rect=final_boxes[j],
                raw_type=raw_type,
                final_type=final_type,
                raw_idx=i,
                final_idx=j,
                note="; ".join(note_bits),
            ))

        for i in unmatched_raw:
            raw_type = raw_types[i] if i < len(raw_types) else ""
            sig = self._rect_signature(raw_boxes[i], raw_type)
            removed.append(sig)
            events.append(self._build_delta_event("box_removed", raw_rect=raw_boxes[i], raw_type=raw_type, raw_idx=i, note="Removed from final revision"))

        for j in unmatched_final:
            final_type = final_types[j] if j < len(final_types) else ""
            sig = self._rect_signature(final_boxes[j], final_type)
            added.append(sig)
            events.append(self._build_delta_event("box_added", final_rect=final_boxes[j], final_type=final_type, final_idx=j, note="Added in final revision"))

        total_units = max(1, max(len(raw_boxes), len(final_boxes), stable_count + len(events)))
        penalty = (len(added) + len(removed)) + (retyped_count * 0.8) + (resized_count * 0.45) + (moved_count * 0.35)
        quality = max(0.0, 1.0 - (float(penalty) / float(total_units + max(stable_count, 0))))

        event_counts = {}
        for event in events:
            key = str(event.get("event_type", "change") or "change")
            event_counts[key] = int(event_counts.get(key, 0) or 0) + 1

        return {
            "raw_count": int(len(raw_boxes)),
            "final_count": int(len(final_boxes)),
            "matched_count": int(len(exact_matches) + len(fuzzy_matches)),
            "stable_count": int(stable_count),
            "added_count": int(len(added)),
            "removed_count": int(len(removed)),
            "retyped_count": int(retyped_count),
            "resized_count": int(resized_count),
            "moved_count": int(moved_count),
            "added_signatures": added[:200],
            "removed_signatures": removed[:200],
            "event_counts": event_counts,
            "events": events[:200],
            "quality_score": round(float(quality), 6),
        }

    def _capture_revision_snapshot(self, reason="manual", approved=False, learning_weight=0.0, parent_revision_ids=None, source_path=""):
        page_idx = int((self.current_detection_context or {}).get("page_idx", self.detected_page_idx or 0) or 0)
        fingerprint = (self.current_detection_context or {}).get("fingerprint") or {
            "version": 1,
            "pdf_family_hint": self._page_family_hint(),
            "page_idx": page_idx,
            "fingerprint_key": "",
            "features": {},
        }
        raw_detection = (self.current_detection_context or {}).get("raw_detection") or self._serialize_detection_state(page_idx=page_idx)
        final_detection = self._serialize_detection_state(page_idx=page_idx)
        learning_delta = self._compute_learning_delta(raw_detection, final_detection)
        effective_weight, trust_class = self._normalize_learning_weight(reason=reason, approved=approved, learning_weight=learning_weight)
        revision = {
            "revision_id": "",
            "reason": str(reason or "manual"),
            "saved_at": _utc_now_iso(),
            "approved": bool(approved),
            "trust_class": str(trust_class),
            "learning_weight": float(effective_weight or 0.0),
            "page_idx": page_idx,
            "pdf_path": str(self.pdf_path or ""),
            "pdf_family_hint": str(fingerprint.get("pdf_family_hint", "")),
            "fingerprint": fingerprint,
            "profile_used": self._collect_learning_profile_settings(),
            "matched_profile": (self.current_detection_context or {}).get("matched_profile"),
            "profile_applied": bool((self.current_detection_context or {}).get("profile_applied", False)),
            "learning_meta": self.collect_learning_meta(),
            "raw_detection": raw_detection,
            "final_detection": final_detection,
            "learning_delta": learning_delta,
            "custom_mappings": self.collect_config().get("custom_mappings", {}),
            "ui_state": dict(getattr(self, "_config_ui_state", {}) or {}),
            "selected_record_label": str((getattr(self, "_config_ui_state", {}) or {}).get("selected_record_label", "") or ""),
            "selected_record_key": str((getattr(self, "_config_ui_state", {}) or {}).get("selected_record_key", "") or ""),
            "parent_revision_ids": list(parent_revision_ids or ([self.current_revision_id] if self.current_revision_id else [])),
            "source_path": str(source_path or ""),
        }
        return revision

    def _update_profile_memory_from_revision(self, revision, weight=None):
        if not isinstance(revision, dict):
            return None
        fingerprint = revision.get("fingerprint", {}) or {}
        fp_key = str(fingerprint.get("fingerprint_key", "") or "").strip()
        if not fp_key:
            return None

        profiles = self.profile_memory.setdefault("profiles", {})
        existing = profiles.get(fp_key, {}) if isinstance(profiles.get(fp_key), dict) else {}

        revision_weight = float(revision.get("learning_weight", 0.0) if weight is None else weight)
        trust_class = str(revision.get("trust_class", "approved" if revision.get("approved") else "draft") or "draft")
        trust_multiplier = 1.0 if revision.get("approved") else (0.65 if trust_class == "saved" else 0.35)
        effective_weight = max(0.05, revision_weight * trust_multiplier)
        save_count_prev = int(existing.get("save_count", 0) or 0)
        quality_prev = float(existing.get("avg_quality", 0.0) or 0.0)
        weighted_quality_sum_prev = float(existing.get("weighted_quality_sum", quality_prev * max(float(existing.get("learning_weight_total", 0.0) or 0.0), 1.0)) or 0.0)
        quality_new = float((revision.get("learning_delta", {}) or {}).get("quality_score", 0.0) or 0.0)
        combined_count = save_count_prev + 1
        total_weight = float(existing.get("learning_weight_total", 0.0) or 0.0) + effective_weight
        weighted_quality_sum = weighted_quality_sum_prev + (quality_new * effective_weight)
        avg_quality = ((quality_prev * save_count_prev) + quality_new) / float(max(combined_count, 1))
        avg_quality_weighted = weighted_quality_sum / float(max(total_weight, 0.05))

        updated = dict(existing)
        updated.update({
            "fingerprint_key": fp_key,
            "pdf_family_hint": str(fingerprint.get("pdf_family_hint", "")),
            "page_idx": int(fingerprint.get("page_idx", revision.get("page_idx", 0)) or 0),
            "feature_snapshot": dict((fingerprint.get("features") or {})),
            "settings": dict(revision.get("profile_used", {}) or {}),
            "save_count": combined_count,
            "approved_count": int(existing.get("approved_count", 0) or 0) + (1 if revision.get("approved") else 0),
            "approved": bool(existing.get("approved", False) or revision.get("approved", False)),
            "trust_class": trust_class,
            "learning_weight_total": round(float(total_weight), 6),
            "weighted_quality_sum": round(float(weighted_quality_sum), 6),
            "avg_quality": round(float(avg_quality), 6),
            "avg_quality_weighted": round(float(avg_quality_weighted), 6),
            "last_revision_id": str(revision.get("revision_id", "")),
            "last_used_at": _utc_now_iso(),
        })
        profiles[fp_key] = updated
        self.profile_memory["version"] = 1
        self.profile_memory["updated_at"] = _utc_now_iso()
        self.learning_storage.save_profile_memory(self.profile_memory)

        families = self.template_stats.setdefault("families", {})
        fam_key = str(updated.get("pdf_family_hint", "") or "unknown")
        fam = dict(families.get(fam_key, {}) or {})
        fam["last_used_at"] = _utc_now_iso()
        fam["page_hits"] = int(fam.get("page_hits", 0) or 0) + 1
        fam["last_fingerprint_key"] = fp_key
        fam["last_revision_id"] = str(revision.get("revision_id", ""))
        fam["avg_quality"] = updated.get("avg_quality_weighted", updated.get("avg_quality", 0.0))
        families[fam_key] = fam
        self.template_stats["version"] = 1
        self.learning_storage.save_template_stats(self.template_stats)
        return updated

    def _append_correction_event(self, event):
        payload = dict(self.correction_log or {})
        events = list(payload.get("events", []) or [])
        entry = dict(event or {})
        entry.setdefault("event_id", f"evt_{uuid.uuid4().hex[:10]}")
        entry.setdefault("recorded_at", _utc_now_iso())
        events.append(entry)
        max_keep = 600
        if len(events) > max_keep:
            events = events[-max_keep:]
        payload["version"] = 1
        payload["events"] = events
        self.correction_log = payload
        self.learning_storage.save_correction_log(payload)
        return entry

    def record_candidate_review_batch(self, page_idx=0, anchor_ids=None, zone_rect=None, reviewed_items=None, source='repair_review'):
        reviewed_items = list(reviewed_items or [])
        if not reviewed_items:
            return None
        try:
            page_idx = int(page_idx or 0)
        except Exception:
            page_idx = 0
        fingerprint = {}
        try:
            fingerprint = dict((self.current_detection_context or {}).get('fingerprint', {}) or {})
        except Exception:
            fingerprint = {}
        if not fingerprint and self.pdf_path:
            try:
                fingerprint = self.build_page_fingerprint(page_idx=page_idx)
            except Exception:
                fingerprint = {}
        summary_counts = {}
        summary = {
            'event_type': 'candidate_review_batch',
            'recorded_at': _utc_now_iso(),
            'page_idx': int(page_idx),
            'pdf_path': str(self.pdf_path or ''),
            'fingerprint_key': str((fingerprint or {}).get('fingerprint_key', '') or ''),
            'pdf_family_hint': str((fingerprint or {}).get('pdf_family_hint', '') or self._page_family_hint() or ''),
            'anchor_ids': [int(x) for x in (anchor_ids or []) if isinstance(x, int) or str(x).isdigit()],
            'zone_rect': self._serialize_rect(zone_rect) if zone_rect is not None else None,
            'source': str(source or 'repair_review'),
            'decision_count': int(len(reviewed_items)),
        }
        summary_entry = self._append_correction_event(summary)
        for item in reviewed_items:
            action = str(item.get('action', '') or '')
            accepted = bool(item.get('accepted', False))
            key = f"{action}:{'accepted' if accepted else 'rejected'}"
            summary_counts[key] = int(summary_counts.get(key, 0) or 0) + 1
            rect = item.get('rect')
            detail = {
                'event_type': 'candidate_review_item',
                'parent_event_id': str(summary_entry.get('event_id', '') or ''),
                'page_idx': int(page_idx),
                'action': action,
                'accepted': accepted,
                'kind': str(item.get('kind', '') or ''),
                'source': str(item.get('source', '') or ''),
                'score': float(item.get('score', 0.0) or 0.0),
                'reason': str(item.get('reason', '') or ''),
                'review_key': str(item.get('review_key', '') or ''),
                'fingerprint_key': str((fingerprint or {}).get('fingerprint_key', '') or ''),
            }
            try:
                if rect is not None:
                    detail['rect'] = self._serialize_rect(rect)
            except Exception:
                pass
            self._append_correction_event(detail)
        self._append_correction_event({
            'event_type': 'candidate_review_summary',
            'parent_event_id': str(summary_entry.get('event_id', '') or ''),
            'page_idx': int(page_idx),
            'decision_summary': summary_counts,
            'fingerprint_key': str((fingerprint or {}).get('fingerprint_key', '') or ''),
        })
        return summary_counts

    def _record_revision_learning_events(self, revision, event_type="revision_saved"):
        if not isinstance(revision, dict):
            return None
        delta = dict(revision.get("learning_delta", {}) or {})
        fingerprint = dict(revision.get("fingerprint", {}) or {})
        net_changes = sum(int(delta.get(k, 0) or 0) for k in ("added_count", "removed_count", "retyped_count", "resized_count", "moved_count"))
        if event_type == "revision_saved" and net_changes <= 0 and not bool(revision.get("approved", False)):
            return None
        profile_used = dict(revision.get("profile_used", {}) or {})
        matched = revision.get("matched_profile") if isinstance(revision.get("matched_profile"), dict) else {}
        base_event = {
            "event_type": str(event_type or "revision_saved"),
            "revision_id": str(revision.get("revision_id", "") or ""),
            "reason": str(revision.get("reason", "") or ""),
            "approved": bool(revision.get("approved", False)),
            "trust_class": str(revision.get("trust_class", "approved" if revision.get("approved") else "draft") or "draft"),
            "learning_weight": float(revision.get("learning_weight", 0.0) or 0.0),
            "pdf_path": str(revision.get("pdf_path", "") or ""),
            "page_idx": int(revision.get("page_idx", 0) or 0),
            "fingerprint_key": str(fingerprint.get("fingerprint_key", "") or ""),
            "pdf_family_hint": str(revision.get("pdf_family_hint", fingerprint.get("pdf_family_hint", "")) or ""),
            "profile_key": str(matched.get("fingerprint_key", "") or ""),
            "profile_applied": bool(revision.get("profile_applied", False)),
            "settings_snapshot": dict(profile_used),
            "delta": {
                "raw_count": int(delta.get("raw_count", 0) or 0),
                "final_count": int(delta.get("final_count", 0) or 0),
                "matched_count": int(delta.get("matched_count", 0) or 0),
                "stable_count": int(delta.get("stable_count", 0) or 0),
                "added_count": int(delta.get("added_count", 0) or 0),
                "removed_count": int(delta.get("removed_count", 0) or 0),
                "retyped_count": int(delta.get("retyped_count", 0) or 0),
                "resized_count": int(delta.get("resized_count", 0) or 0),
                "moved_count": int(delta.get("moved_count", 0) or 0),
                "quality_score": float(delta.get("quality_score", 0.0) or 0.0),
            },
            "event_counts": dict(delta.get("event_counts", {}) or {}),
        }
        summary_entry = self._append_correction_event(base_event)
        detail_events = list(delta.get("events", []) or [])
        for detail in detail_events[:80]:
            child = dict(base_event)
            child.update({
                "event_type": str(detail.get("event_type", "box_change") or "box_change"),
                "parent_event_id": str(summary_entry.get("event_id", "") or ""),
                "detail": dict(detail),
            })
            self._append_correction_event(child)
        return summary_entry

    def save_learning_revision(self, reason="manual", approved=False, learning_weight=0.0, parent_revision_ids=None, source_path=""):
        revision = self._capture_revision_snapshot(
            reason=reason,
            approved=approved,
            learning_weight=learning_weight,
            parent_revision_ids=parent_revision_ids,
            source_path=source_path,
        )
        revision_id = self.learning_storage.save_revision(revision)
        revision["revision_id"] = revision_id
        self.current_revision_id = revision_id
        self.last_saved_revision_id = revision_id
        self._update_profile_memory_from_revision(revision, weight=revision.get("learning_weight", learning_weight))
        self._record_revision_learning_events(revision, event_type="revision_saved")
        if approved:
            self._record_revision_learning_events(revision, event_type="revision_approved")
        self.persist_learning_session(current_page_idx=revision.get("page_idx", 0))
        return revision

    def approve_revision_for_learning(self, revision_id=None, learning_weight=1.0):
        revision_id = str(revision_id or self.current_revision_id or "").strip()
        if not revision_id:
            return None
        revision = self.learning_storage.load_revision(revision_id)
        if not isinstance(revision, dict):
            return None
        revision["approved"] = True
        effective_weight, trust_class = self._normalize_learning_weight(reason=revision.get("reason", "manual"), approved=True, learning_weight=learning_weight)
        revision["trust_class"] = str(trust_class)
        revision["learning_weight"] = float(effective_weight)
        self.learning_storage.save_revision(revision)
        self.current_revision_id = revision_id
        self.last_saved_revision_id = revision_id
        self._update_profile_memory_from_revision(revision, weight=learning_weight)
        self._record_revision_learning_events(revision, event_type="revision_approved")
        self.persist_learning_session(current_page_idx=revision.get("page_idx", self.detected_page_idx or 0))
        return revision

    def collect_learning_meta(self):
        ctx = self.current_detection_context or {}
        fingerprint = ctx.get("fingerprint", {}) if isinstance(ctx, dict) else {}
        return {
            "version": 1,
            "storage_root": str(self.learning_storage.base_dir),
            "current_revision_id": str(self.current_revision_id or ""),
            "last_saved_revision_id": str(self.last_saved_revision_id or ""),
            "fingerprint_key": str((fingerprint or {}).get("fingerprint_key", "") or ""),
            "page_idx": int(ctx.get("page_idx", self.detected_page_idx or 0) or 0),
            "profile_applied": bool(ctx.get("profile_applied", False)),
            "last_applied_profile_meta": dict(self.last_applied_profile_meta or {}),
            "updated_at": _utc_now_iso(),
        }

    def apply_learning_meta(self, learning_meta):
        meta = dict(learning_meta or {})
        self.last_loaded_learning_meta = meta
        self.current_revision_id = str(meta.get("current_revision_id", self.current_revision_id or "") or "")
        self.last_saved_revision_id = str(meta.get("last_saved_revision_id", self.last_saved_revision_id or "") or "")
        self.last_applied_profile_meta = dict(meta.get("last_applied_profile_meta", {}) or {})

    def persist_learning_session(self, current_page_idx=0):
        if not self.learning_enabled:
            return
        ui_state = dict(getattr(self, "_config_ui_state", {}) or {})
        session = {
            "version": 2,
            "saved_at": _utc_now_iso(),
            "pdf_path": str(self.pdf_path or ""),
            "page_idx": int(current_page_idx or 0),
            "detected_page_idx": int(self.detected_page_idx or 0),
            "current_revision_id": str(self.current_revision_id or ""),
            "learning_meta": self.collect_learning_meta(),
            "ui_state": ui_state,
            "selected_record_label": str(ui_state.get("selected_record_label", "") or ""),
            "selected_record_key": str(ui_state.get("selected_record_key", "") or ""),
            "current_detection_context": {
                "page_idx": int((self.current_detection_context or {}).get("page_idx", 0) or 0),
                "fingerprint": (self.current_detection_context or {}).get("fingerprint", {}),
                "profile_applied": bool((self.current_detection_context or {}).get("profile_applied", False)),
                "matched_profile_key": str((((self.current_detection_context or {}).get("matched_profile") or {}).get("fingerprint_key", "")) or ""),
            },
        }
        self.learning_storage.save_session(session, name="current_session")

    def collect_config(self):
        mappings_serial = {}

        for k, lst in self.custom_mappings.items():
            mappings_serial[k] = []
            for item in lst:
                rects = self._mapping_rect_list(item)
                mappings_serial[k].append({
                    "column": item.get("column", ""),
                    "trigger": item.get("trigger", ""),
                    "rects": [[r.x0, r.y0, r.x1, r.y1] for r in rects],
                    "page": int(item.get("page", 0)),
                    "g": bool(item.get("g", False)),
                    "n": int(item.get("n", 1)),
                })

        semantic_overrides_serial = {}

        ui_state = getattr(self, "_config_ui_state", {}) or {}
        if not isinstance(ui_state, dict):
            ui_state = {}

        return {
            "version": 7,
            "record_label_column": str(self._record_label_column()),
            "record_label_source_column": str(getattr(self, "record_label_source_column", "") or ""),
            "record_key_column": str(self._record_key_column()),
            "pdf_path": self.pdf_path,
            "zoom": float(ZOOM),
            "text_tuning": self._normalized_text_tuning(getattr(self, "text_tuning", None)),
            "candidate_tuning": self.get_candidate_tuning(),
            "ui_state": ui_state,
            "learning_meta": self.collect_learning_meta(),
            "current_detection_context": dict(self.current_detection_context or {}),
            "current_revision_id": str(self.current_revision_id or ""),
            "last_saved_revision_id": str(self.last_saved_revision_id or ""),

            "F_Area": int(self.settings["F_Area"]),
            "F_MinW": int(self.settings["F_MinW"]),
            "F_MinH": int(self.settings["F_MinH"]),
            "F_Close": int(self.settings["F_Close"]),

            "Line_MinW": int(self.settings["Line_MinW"]),
            "Line_MaxW": int(self.settings["Line_MaxW"]),

            "C_Strict": int(self.settings["C_Strict"]),
            "C_Size": [int(self.settings["C_Size"][0]), int(self.settings["C_Size"][1])],

            "C_Border": float(self.settings["C_Border"]),
            "C_Inner": float(self.settings["C_Inner"]),
            "ROI_Max": int(self.settings["ROI_Max"]),

            "C_Open": int(self.settings["C_Open"]),
            "C_Close": int(self.settings["C_Close"]),

            "C_BandPct": float(self.settings["C_BandPct"]),
            "C_AspectTol": float(self.settings["C_AspectTol"]),

            "Ext_Low": float(self.settings["Ext_Low"]),
            "Ext_High": float(self.settings["Ext_High"]),

            "C_FillMin": float(self.settings["C_FillMin"]),
            "C_Eps": float(self.settings["C_Eps"]),
            "Use_Extent": bool(self.settings["Use_Extent"]),

            "Is_Grid": bool(self.settings["Is_Grid"]),
            "Grid_N": int(self.settings["Grid_N"]),

            "custom_mappings": mappings_serial,
            "area_templates": self._serialize_area_templates(),
            "repair_patches": self._serialize_repair_patches(),
        }

    def apply_config(self, cfg):
        self.settings["F_Area"] = int(cfg.get("F_Area", self.settings["F_Area"]))
        self.settings["F_MinW"] = int(cfg.get("F_MinW", self.settings["F_MinW"]))
        self.settings["F_MinH"] = int(cfg.get("F_MinH", self.settings["F_MinH"]))
        self.settings["F_Close"] = int(cfg.get("F_Close", self.settings["F_Close"]))

        self.settings["Line_MinW"] = int(cfg.get("Line_MinW", self.settings["Line_MinW"]))
        self.settings["Line_MaxW"] = int(cfg.get("Line_MaxW", self.settings["Line_MaxW"]))

        self.settings["C_Strict"] = int(cfg.get("C_Strict", self.settings["C_Strict"]))

        c_size = cfg.get("C_Size", self.settings["C_Size"])
        if isinstance(c_size, (list, tuple)) and len(c_size) == 2:
            self.settings["C_Size"] = (int(c_size[0]), int(c_size[1]))

        self.settings["C_Border"] = float(cfg.get("C_Border", self.settings["C_Border"]))
        self.settings["C_Inner"] = float(cfg.get("C_Inner", self.settings["C_Inner"]))
        self.settings["ROI_Max"] = int(cfg.get("ROI_Max", self.settings["ROI_Max"]))
        self.settings["C_Open"] = int(cfg.get("C_Open", self.settings["C_Open"]))
        self.settings["C_Close"] = int(cfg.get("C_Close", self.settings["C_Close"]))
        self.settings["C_BandPct"] = float(cfg.get("C_BandPct", self.settings["C_BandPct"]))
        self.settings["C_AspectTol"] = float(cfg.get("C_AspectTol", self.settings["C_AspectTol"]))
        self.settings["Ext_Low"] = float(cfg.get("Ext_Low", self.settings["Ext_Low"]))
        self.settings["Ext_High"] = float(cfg.get("Ext_High", self.settings["Ext_High"]))
        self.settings["C_FillMin"] = float(cfg.get("C_FillMin", self.settings["C_FillMin"]))
        self.settings["C_Eps"] = float(cfg.get("C_Eps", self.settings["C_Eps"]))
        self.settings["Use_Extent"] = bool(cfg.get("Use_Extent", self.settings["Use_Extent"]))
        self.settings["Is_Grid"] = bool(cfg.get("Is_Grid", self.settings["Is_Grid"]))
        self.settings["Grid_N"] = int(cfg.get("Grid_N", self.settings["Grid_N"]))
        self.config_zoom = float(cfg.get("zoom", getattr(self, "config_zoom", ZOOM)))
        self.config_pdf_path = str(cfg.get("pdf_path", getattr(self, "config_pdf_path", "")) or "")
        self.text_tuning = self._normalized_text_tuning(cfg.get("text_tuning", getattr(self, "text_tuning", None)))
        self.candidate_tuning = self._normalized_candidate_tuning(cfg.get("candidate_tuning", getattr(self, "candidate_tuning", None)))
        self.record_label_source_column = str(cfg.get("record_label_source_column", getattr(self, "record_label_source_column", "")) or "").strip()
        self.apply_learning_meta(cfg.get("learning_meta", {}))
        restored_ctx = cfg.get("current_detection_context", {}) or {}
        self.current_detection_context = dict(restored_ctx) if isinstance(restored_ctx, dict) else {}
        self.current_revision_id = str(cfg.get("current_revision_id", self.current_revision_id or "") or self.current_revision_id or "")
        self.last_saved_revision_id = str(cfg.get("last_saved_revision_id", self.last_saved_revision_id or "") or self.last_saved_revision_id or "")

        self.custom_mappings.clear()
        self.semantic_target_overrides = {}
        self.area_templates_by_page = {}
        self.area_template_index = {}
        loaded_mappings = cfg.get("custom_mappings", {}) or {}
        for k, lst in loaded_mappings.items():
            if isinstance(lst, dict):
                lst = [lst]
            elif not isinstance(lst, list):
                continue

            self.custom_mappings[k] = []
            for item in lst:
                if not isinstance(item, dict):
                    continue

                rects = []
                if item.get("rects"):
                    for coords in item.get("rects", []):
                        if isinstance(coords, (list, tuple)) and len(coords) == 4:
                            rects.append(fitz.Rect(*coords))
                else:
                    rect_coords = item.get("rect", [0, 0, 0, 0])
                    if isinstance(rect_coords, (list, tuple)) and len(rect_coords) == 4:
                        rects = [fitz.Rect(*rect_coords)]

                self.custom_mappings[k].append({
                    "column": str(item.get("column", "")),
                    "trigger": str(item.get("trigger", "")),
                    "rects": rects,
                    "page": int(item.get("page", 0)),
                    "g": bool(item.get("g", item.get("is_grid", False))),
                    "n": int(item.get("n", item.get("grid_n", 1))),
                })

        self.all_boxes = []
        self.box_types = []
        self.geom = self._empty_geom()
        self.semantic_targets = self._empty_semantic_targets()
        self.selected_box_ids = []
        self.detected_page_idx = None
        self.boxes_by_page = {}
        self.box_types_by_page = {}
        self.geom_by_page = {}
        self.semantic_targets_by_page = {}
        self._deserialize_area_templates(cfg.get("area_templates", {}))
        self._deserialize_repair_patches(cfg.get("repair_patches", {}))

    def merge_config_into_current(self, incoming_cfg, keep_current_detection=True, prefer="incoming"):
        current_cfg = self.collect_config()

        if keep_current_detection:
            merged_cfg = dict(current_cfg)
            current_ui = dict(merged_cfg.get("ui_state", {}) or {})
            merged_cfg["ui_state"] = current_ui
            if isinstance(incoming_cfg, dict):
                for meta_key in ("record_label_column", "record_key_column"):
                    if meta_key in incoming_cfg:
                        merged_cfg[meta_key] = incoming_cfg.get(meta_key)
        else:
            merged_cfg = dict(current_cfg)
            for k, v in incoming_cfg.items():
                if k not in ("custom_mappings", "semantic_target_overrides", "semantic_target_schema"):
                    merged_cfg[k] = v

        current_entries = self._extract_current_mapping_entries(self.custom_mappings)
        incoming_entries = self._extract_cfg_mapping_entries(incoming_cfg)

        merged_entries = list(current_entries)

        for inc in incoming_entries:
            if not inc.get("column") and not inc.get("rects"):
                continue

            if any(self._same_mapping_identity(old, inc) for old in merged_entries):
                continue

            conflict_idx = None
            for i, old in enumerate(merged_entries):
                if self._same_target_rects(old, inc):
                    conflict_idx = i
                    break

            if conflict_idx is not None:
                if prefer == "incoming":
                    merged_entries[conflict_idx] = inc
            else:
                merged_entries.append(inc)

        merged_cfg["custom_mappings"] = {}
        merged_mapping_dict = self._entries_to_mapping_dict(merged_entries)

        for k, lst in merged_mapping_dict.items():
            merged_cfg["custom_mappings"][k] = []
            for item in lst:
                merged_cfg["custom_mappings"][k].append({
                    "column": item.get("column", ""),
                    "trigger": item.get("trigger", ""),
                    "rects": [[r.x0, r.y0, r.x1, r.y1] for r in item.get("rects", [])],
                    "page": int(item.get("page", 0)),
                    "g": bool(item.get("g", False)),
                    "n": int(item.get("n", 1)),
                })

        self.apply_config(merged_cfg)
        return merged_cfg

    def save_config(self, path):
        cfg = self.collect_config()
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False)
        self.persist_learning_session(current_page_idx=self.detected_page_idx or 0)

    def load_config(self, path):
        with open(path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        self.apply_config(cfg)
        self.persist_learning_session(current_page_idx=self.detected_page_idx or 0)
        return cfg



    def _draw_text_op_cv(self, img, text, rects, preview_zoom=1.5, is_grid=False, grid_n=1, ox=0, oy=0, fs_scale=0.65):
        import unicodedata

        tuning = self._normalized_text_tuning(getattr(self, "text_tuning", None))
        uppercase_enabled = bool(tuning.get("uppercase", True))

        def _apply_case(value):
            value = unicodedata.normalize("NFC", str(value or ""))
            return value.upper() if uppercase_enabled else value

        val = unicodedata.normalize("NFC", str(text or "").strip())
        if not val or val.lower() in ["nan", "none"]:
            return
        if val.endswith(".0"):
            val = val[:-2]
        val = _apply_case(val)

        rects = sorted([self._coerce_rect(r) for r in rects], key=lambda rr: (round(rr.y0, 3), rr.x0))
        if not rects:
            return

        rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        pil_img = PILImage.fromarray(rgb)
        draw = ImageDraw.Draw(pil_img)

        font_cache = {}
        sig_cache = {}
        font_path_cache = getattr(self, "_pil_font_path_cache", {})
        setattr(self, "_pil_font_path_cache", font_path_cache)

        def glyph_signature(font, ch):
            key = (id(font), ch)
            if key in sig_cache:
                return sig_cache[key]
            try:
                mask = font.getmask(ch, mode="L")
                sig = (mask.size, bytes(mask))
            except Exception:
                sig = None
            sig_cache[key] = sig
            return sig

        def classify_field(rect_list, force_grid=False):
            if force_grid:
                return "grid"
            target = rect_list[0] if len(rect_list) == 1 else self._rect_union(rect_list)
            w = max(float(target.width), 1.0)
            h = max(float(target.height), 1.0)
            aspect = w / h
            h_px = h * max(float(preview_zoom), 0.001)
            if len(rect_list) > 1:
                return "line" if (h_px <= 16 and aspect >= 3.4) else "box"
            if h_px <= 12 and aspect >= 2.8:
                return "line"
            if h_px <= 15 and aspect >= 5.4:
                return "line"
            if h_px <= 22 and aspect >= 11.0:
                return "line"
            return "box"

        FIELD_PROFILES = {
            "grid": {
                "font_bias": "uniform_bold",
                "start_mul": max(float(fs_scale), 0.86),
                "min_size": 8,
                "pad_x": 0.22,
                "pad_y": 0.26,
                "stroke_mul": 0.00,
                "stroke_min_h": 9999,
                "vshift": 0.00,
                "truncate_mul": 0.64,
                "min_effective_h_px": 0,
                "expand_h_mul": 1.00,
                "content_inset_x": 0.04,
                "content_inset_y": 0.08,
                "true_center": True,
            },
            "box": {
                "font_bias": "uniform_bold",
                "start_mul": max(float(fs_scale), 0.95),
                "min_size": 10,
                "pad_x": 0.16,
                "pad_y": 0.22,
                "stroke_mul": 0.00,
                "stroke_min_h": 9999,
                "vshift": 0.00,
                "truncate_mul": 0.72,
                "min_effective_h_px": 0,
                "expand_h_mul": 1.00,
                "content_inset_x": 0.05,
                "content_inset_y": 0.10,
                "true_center": True,
            },
            "line": {
                "font_bias": "uniform_bold_condensed",
                "start_mul": max(float(fs_scale), 0.90),
                "min_size": 10,
                "pad_x": 0.10,
                "pad_y": 0.22,
                "stroke_mul": 0.00,
                "stroke_min_h": 9999,
                "vshift": 0.00,
                "truncate_mul": 0.72,
                "min_effective_h_px": 24,
                "expand_h_mul": 1.75,
                "content_inset_x": 0.06,
                "content_inset_y": 0.10,
                "true_center": True,
            },
        }

        overall_scale = float(tuning.get("overall_scale", 0.92))
        profile_scale_map = {
            "grid": float(tuning.get("grid_scale", 0.92)),
            "box": float(tuning.get("box_scale", 1.00)),
            "line": float(tuning.get("line_scale", 0.94)),
        }
        inset_x = float(tuning.get("content_inset_x", 0.07))
        inset_y = float(tuning.get("content_inset_y", 0.14))
        vshift_map = {
            "grid": float(tuning.get("grid_vshift", 0.00)),
            "box": float(tuning.get("box_vshift", 0.00)),
            "line": float(tuning.get("line_vshift", 0.00)),
        }
        for _kind, _profile in FIELD_PROFILES.items():
            scale_mul = max(0.60, min(1.60, overall_scale * profile_scale_map.get(_kind, 1.0)))
            _profile["start_mul"] = max(0.50, float(_profile.get("start_mul", 1.0)) * scale_mul)
            _profile["min_size"] = max(8, int(round(float(_profile.get("min_size", 10)) * min(scale_mul, 1.10))))
            _profile["content_inset_x"] = inset_x
            _profile["content_inset_y"] = inset_y
            _profile["vshift"] = vshift_map.get(_kind, float(_profile.get("vshift", 0.0)))
        FIELD_PROFILES["line"]["min_effective_h_px"] = int(tuning.get("line_min_effective_h_px", 28))
        FIELD_PROFILES["line"]["expand_h_mul"] = float(tuning.get("line_expand_h_mul", 2.10))

        def build_font_candidates(style="box"):
            if style in ("line", "uniform_bold_condensed"):
                candidates = [
                    "/system/fonts/RobotoCondensed-Bold.ttf",
                    "/system/fonts/NotoSans-CondensedBold.ttf",
                    "/system/fonts/NotoSansDisplay-CondensedBold.ttf",
                    "/system/fonts/Roboto-Bold.ttf",
                    "/system/fonts/NotoSans-Bold.ttf",
                    "/system/fonts/RobotoCondensed-Regular.ttf",
                    "/system/fonts/NotoSans-Condensed.ttf",
                    "/system/fonts/NotoSansDisplay-Condensed.ttf",
                    "/system/fonts/Roboto-Regular.ttf",
                    "/system/fonts/NotoSans-Regular.ttf",
                ]
            else:
                candidates = [
                    "/system/fonts/NotoSans-Bold.ttf",
                    "/system/fonts/Roboto-Bold.ttf",
                    "/system/fonts/NotoSans-SemiBold.ttf",
                    "/system/fonts/Roboto-Medium.ttf",
                    "/system/fonts/NotoSansDisplay-Bold.ttf",
                    "/system/fonts/DroidSans-Bold.ttf",
                    "/system/fonts/NotoSans-Regular.ttf",
                    "/system/fonts/Roboto-Regular.ttf",
                ]
            candidates += [
                "/system/fonts/RobotoStatic-Bold.ttf",
                "/system/fonts/RobotoStatic-Regular.ttf",
                "/system/fonts/DroidSans.ttf",
                "/data/data/com.termux/files/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
                "/data/data/com.termux/files/usr/share/fonts/TTF/DejaVuSans.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            ]
            try:
                if os.path.isdir("/system/fonts"):
                    preferred = []
                    fallback = []
                    for name in sorted(os.listdir("/system/fonts")):
                        low = name.lower()
                        if not low.endswith((".ttf", ".otf", ".ttc")):
                            continue
                        full = os.path.join("/system/fonts", name)
                        if not any(tag in low for tag in ["noto", "roboto", "dejavu", "free", "sans", "droid"]):
                            fallback.append(full)
                            continue
                        if style in ("line", "uniform_bold_condensed"):
                            if "condensed" in low and any(tag in low for tag in ["bold", "semibold", "medium"]):
                                preferred.insert(0, full)
                            elif "condensed" in low:
                                preferred.append(full)
                            elif any(tag in low for tag in ["bold", "semibold", "medium"]):
                                preferred.append(full)
                            else:
                                fallback.append(full)
                        else:
                            if any(tag in low for tag in ["bold", "semibold", "medium"]):
                                preferred.insert(0, full)
                            elif "regular" in low:
                                preferred.append(full)
                            else:
                                fallback.append(full)
                    candidates.extend(preferred)
                    candidates.extend(fallback)
            except Exception:
                pass
            out = []
            seen = set()
            for path in candidates:
                if path and path not in seen and os.path.exists(path):
                    seen.add(path)
                    out.append(path)
            return out

        def best_font_path(sample_text, style="box"):
            sample_key = (style, tuple(sorted({ch for ch in str(sample_text or "") if ord(ch) > 127})))
            if sample_key in font_path_cache:
                return font_path_cache[sample_key]
            candidates = build_font_candidates(style=style)
            if not candidates:
                font_path_cache[sample_key] = None
                return None
            probe_size = 30 if style == "box" else 28
            best_path = None
            best_score = -1.0
            for path in candidates:
                try:
                    font = ImageFont.truetype(path, probe_size)
                except Exception:
                    continue
                qsig = glyph_signature(font, "?")
                score = 0.0
                for ch in sample_key[1]:
                    sig = glyph_signature(font, ch)
                    if sig and sig != qsig:
                        score += 1.0
                low = os.path.basename(path).lower()
                if style in ("line", "uniform_bold_condensed") and "condensed" in low:
                    score += 1.00
                elif style in ("line", "uniform_bold_condensed") and any(tag in low for tag in ["bold", "semibold", "medium"]):
                    score += 0.60
                elif any(tag in low for tag in ["bold", "semibold", "medium"]):
                    score += 0.75
                if score > best_score:
                    best_score = score
                    best_path = path
                if score >= len(sample_key[1]) + 0.55:
                    break
            font_path_cache[sample_key] = best_path
            return best_path

        def sanitize_for_font(value, font):
            value = _apply_case(value)
            qsig = glyph_signature(font, "?")
            pieces = []
            for ch in value:
                if ord(ch) <= 127 or ch.isspace():
                    pieces.append(ch)
                    continue
                sig = glyph_signature(font, ch)
                if sig and sig != qsig:
                    pieces.append(ch)
                    continue
                fallback = unicodedata.normalize("NFKD", ch).encode("ascii", "ignore").decode("ascii").upper()
                pieces.append(fallback if fallback else "?")
            return "".join(pieces)

        def pick_font(size_px, sample_text="", style="box"):
            size_px = max(int(size_px), 8)
            cache_key = (style, size_px, tuple(sorted({ch for ch in str(sample_text or "") if ord(ch) > 127})))
            if cache_key in font_cache:
                return font_cache[cache_key]
            best_path = best_font_path(sample_text, style=style)
            font = None
            if best_path:
                try:
                    font = ImageFont.truetype(best_path, size_px)
                except Exception:
                    font = None
            if font is None:
                try:
                    font = ImageFont.load_default()
                except Exception:
                    font = None
            font_cache[cache_key] = font
            return font

        def measure_bbox(value, font, stroke_width=0):
            bbox = draw.textbbox((0, 0), value, font=font, stroke_width=stroke_width)
            return bbox, max(1, bbox[2] - bbox[0]), max(1, bbox[3] - bbox[1])

        def fit_text(single_val, target_w, target_h, field_kind="box"):
            candidate = _apply_case(single_val)
            target_w = max(int(target_w), 1)
            target_h = max(int(target_h), 1)
            profile = FIELD_PROFILES.get(field_kind, FIELD_PROFILES["box"])
            effective_h = max(target_h, int(profile.get("min_effective_h_px", 0)))
            inner_w = max(int(target_w - max(4, target_w * profile["pad_x"])), 1)
            inner_h = max(int(effective_h - max(3, effective_h * profile["pad_y"])), 1)
            start_size = max(int(effective_h * profile["start_mul"]), profile["min_size"])
            min_size = int(profile["min_size"])
            base_stroke = 0
            for size in range(start_size, min_size - 1, -1):
                font = pick_font(size, candidate, style=profile["font_bias"])
                if font is None:
                    break
                safe_val = sanitize_for_font(candidate, font)
                for stroke in [base_stroke, 0]:
                    bbox, tw, th = measure_bbox(safe_val, font, stroke_width=stroke)
                    if tw <= inner_w and th <= inner_h:
                        return safe_val, font, bbox, tw, th, stroke, profile
            ellipsis = "..."
            trunc_start = max(min_size, int(target_h * profile["truncate_mul"]))
            for size in range(trunc_start, min_size - 1, -1):
                font = pick_font(size, candidate, style=profile["font_bias"])
                if font is None:
                    break
                safe_base = sanitize_for_font(candidate, font)
                for stroke in [base_stroke, 0]:
                    shortened = safe_base
                    while len(shortened) > 0:
                        probe = (shortened + ellipsis) if shortened else ellipsis
                        bbox, tw, th = measure_bbox(probe, font, stroke_width=stroke)
                        if tw <= inner_w and th <= inner_h:
                            return probe, font, bbox, tw, th, stroke, profile
                        shortened = shortened[:-1].rstrip()
            font = pick_font(min_size, candidate, style=profile["font_bias"])
            if font is None:
                return "", None, None, 0, 0, 0, profile
            bbox, tw, th = measure_bbox(ellipsis, font, stroke_width=0)
            return ellipsis, font, bbox, tw, th, 0, profile

        def _effective_rect(rect, profile):
            rect = self._coerce_rect(rect)
            expand_h_mul = max(float(profile.get("expand_h_mul", 1.0)), 1.0)
            min_effective_h_px = int(profile.get("min_effective_h_px", 0))
            actual_h_px = max(int(rect.height * preview_zoom), 1)
            desired_h_px = max(actual_h_px, min_effective_h_px, int(round(actual_h_px * expand_h_mul)))
            if desired_h_px <= actual_h_px:
                return rect
            extra_pdf_h = (desired_h_px - actual_h_px) / max(float(preview_zoom), 0.001)
            half_extra = extra_pdf_h / 2.0
            return fitz.Rect(rect.x0, rect.y0 - half_extra, rect.x1, rect.y1 + half_extra)

        def _target_rect(rect, profile, field_kind):
            rect = self._coerce_rect(rect)
            if field_kind == "line":
                return _effective_rect(rect, profile)
            return rect

        def _content_rect(rect, profile):
            rect = self._coerce_rect(rect)
            inset_x = max(0.0, float(rect.width) * float(profile.get("content_inset_x", 0.0)))
            inset_y = max(0.0, float(rect.height) * float(profile.get("content_inset_y", 0.0)))
            x0 = rect.x0 + inset_x
            x1 = rect.x1 - inset_x
            y0 = rect.y0 + inset_y
            y1 = rect.y1 - inset_y
            if x1 <= x0:
                x0, x1 = rect.x0, rect.x1
            if y1 <= y0:
                y0, y1 = rect.y0, rect.y1
            return fitz.Rect(x0, y0, x1, y1)

        def centered_anchor(rect, profile, field_kind):
            target_rect = _target_rect(rect, profile, field_kind)
            content_rect = _content_rect(target_rect, profile)
            left = (content_rect.x0 * preview_zoom) + (ox * preview_zoom)
            top = (content_rect.y0 * preview_zoom) + (oy * preview_zoom)
            target_w = max(float(content_rect.width * preview_zoom), 1.0)
            target_h = max(float(content_rect.height * preview_zoom), 1.0)
            cx = left + (target_w / 2.0)
            cy = top + (target_h / 2.0)
            cy += target_h * float(profile.get("vshift", 0.0))
            return cx, cy, content_rect

        def draw_centered_text(cx, cy, display_val, font, stroke, bbox):
            try:
                draw.text((cx, cy), display_val, font=font, fill=(0, 0, 0),
                          stroke_width=stroke, stroke_fill=(0, 0, 0), anchor="mm")
                return
            except Exception:
                pass
            bx0, by0, bx1, by1 = bbox
            x = cx - ((bx0 + bx1) / 2.0)
            y = cy - ((by0 + by1) / 2.0)
            draw.text((int(round(x)), int(round(y))), display_val, font=font, fill=(0, 0, 0),
                      stroke_width=stroke, stroke_fill=(0, 0, 0))

        field_kind = classify_field(rects, force_grid=bool(is_grid))

        def draw_single(single_val, rect):
            probe_profile = FIELD_PROFILES.get(field_kind, FIELD_PROFILES["box"])
            target_rect = _target_rect(rect, probe_profile, field_kind)
            target_w = max(int(target_rect.width * preview_zoom), 1)
            target_h = max(int(target_rect.height * preview_zoom), 1)
            display_val, font, bbox, tw, th, stroke, profile = fit_text(single_val, target_w, target_h, field_kind=field_kind)
            if not display_val or font is None or bbox is None:
                return
            cx, cy, _ = centered_anchor(rect, profile, field_kind)
            draw_centered_text(cx, cy, display_val, font, stroke, bbox)

        def draw_grid(grid_val, rect, cells):
            cells = max(int(cells), 1)
            target_rect = _target_rect(rect, FIELD_PROFILES["grid"], "grid")
            target_h = max(int(target_rect.height * preview_zoom), 1)
            cell_w = (target_rect.width * preview_zoom) / cells
            base_cell_w = rect.width / cells
            for i, ch in enumerate(_apply_case(grid_val)[:cells]):
                display_val, font, bbox, tw, th, stroke, profile = fit_text(str(ch), cell_w - 2, target_h, field_kind="grid")
                if not display_val or font is None or bbox is None:
                    continue
                cell_rect = fitz.Rect(
                    rect.x0 + (i * base_cell_w),
                    rect.y0,
                    rect.x0 + ((i + 1) * base_cell_w),
                    rect.y1,
                )
                cx, cy, _ = centered_anchor(cell_rect, profile, "grid")
                draw_centered_text(cx, cy, display_val, font, stroke, bbox)

        if is_grid and len(rects) > 1:
            counts = self._allocate_cells_by_width(rects, grid_n)
            pos = 0
            for rect, n_cells in zip(rects, counts):
                seg = val[pos:pos + n_cells]
                if seg:
                    draw_grid(seg, rect, n_cells)
                pos += n_cells
        else:
            target = rects[0] if len(rects) == 1 else self._rect_union(rects)
            if is_grid:
                draw_grid(val, target, grid_n)
            else:
                draw_single(val, target)

        img[:] = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)

    def _draw_check_op_cv(self, img, rects, preview_zoom=1.5):
        rects = sorted([self._coerce_rect(r) for r in rects], key=lambda rr: (round(rr.y0, 3), rr.x0))
        if not rects:
            return
        target = rects[0] if len(rects) == 1 else self._rect_union(rects)
        x0 = int(target.x0 * preview_zoom)
        y0 = int(target.y0 * preview_zoom)
        x1 = int(target.x1 * preview_zoom)
        y1 = int(target.y1 * preview_zoom)
        pad = max(1, int(min(x1 - x0, y1 - y0) * 0.18))
        thickness = 1 if (y1 - y0) < 36 else 2
        cv2.line(img, (x0 + pad, y0 + pad), (x1 - pad, y1 - pad), (0, 0, 0), thickness, cv2.LINE_AA)
        cv2.line(img, (x1 - pad, y0 + pad), (x0 + pad, y1 - pad), (0, 0, 0), thickness, cv2.LINE_AA)

    def get_processed_preview_pixmap(self, patient_name, page_idx=0, preview_zoom=1.5, record_key=None):
        img = self.get_raw_preview_pixmap(page_idx=page_idx, preview_zoom=preview_zoom)
        if img is None or getattr(img, "size", 0) == 0:
            raise ValueError("Failed to build preview image.")

        for op in self._collect_overlay_ops(patient_name, page_idx=page_idx, record_key=record_key):
            if op.get("kind") == "check":
                self._draw_check_op_cv(img, op.get("rects", []), preview_zoom=preview_zoom)
            else:
                self._draw_text_op_cv(
                    img,
                    op.get("text", ""),
                    op.get("rects", []),
                    preview_zoom=preview_zoom,
                    is_grid=bool(op.get("grid", False)),
                    grid_n=int(op.get("grid_n", 1)),
                )
        return img

    def get_preview_pixmap_with_boxes(self, patient_name, page_idx=0, preview_zoom=1.5, record_key=None):
        img = self.get_processed_preview_pixmap(patient_name, page_idx=page_idx, preview_zoom=preview_zoom, record_key=record_key)

        for i, r in enumerate(self.all_boxes):
            x0 = int(r.x0 * preview_zoom)
            y0 = int(r.y0 * preview_zoom)
            x1 = int(r.x1 * preview_zoom)
            y1 = int(r.y1 * preview_zoom)

            box_type = self.box_types[i]
            if box_type == "check":
                color = (0, 0, 255)
            elif box_type == "line":
                color = (0, 215, 255)
            else:
                color = (0, 255, 0)

            cv2.rectangle(img, (x0, y0), (x1, y1), color, 2)
            label = str(i)
            cv2.rectangle(img, (x0, max(0, y0 - 14)), (x0 + 20, y0), (0, 0, 0), -1)
            cv2.putText(img, label, (x0 + 2, y0 - 3), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (255, 255, 255), 1, cv2.LINE_AA)

        return img

# =========================
# PART 4 / 4
# Kivy App UI + Wiring
# Paste this BELOW Part 3
# =========================

class InteractivePreview(Image):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.size = (dp(1), dp(1))
        self.boxes_payload = []
        self.selected_ids = set()
        self.hovered_box_id = None
        self.preview_zoom = 1.0
        self.page_idx = 0
        self.label_mode = 'all'
        self.box_tap_callback = None
        self.box_double_tap_callback = None
        self.preview_tap_callback = None
        self.hover_callback = None
        self.zoom_request_callback = None
        self.display_scale = 1.0
        self.capture_rect_payload = None
        self.capture_split_rects_payload = []
        self.capture_path_payload = []
        self.capture_points_payload = []
        self._active_touches = {}
        self._touch_refs = {}
        self._pinch_start_dist = None
        self._pinch_start_scale = 1.0
        self._recent_taps = []
        self._mouse_bound = False
        self._tap_candidates = {}
        self._touch_start = {}
        self._touch_moved = set()
        self._tap_slop = float(dp(18))
        self._pinch_active = False
        self._pinch_last_midpoint = None
        self.bind(texture=self._redraw_overlay, size=self._redraw_overlay, pos=self._redraw_overlay)
        if platform != "android":
            Window.bind(mouse_pos=self._on_mouse_pos)
            self._mouse_bound = True

    def on_parent(self, *args):
        self._redraw_overlay()

    def set_texture_from_bgr(self, img_bgr):
        if img_bgr is None:
            self.texture = None
            self.canvas.after.clear()
            self.canvas.ask_update()
            return
        try:
            if cv2 is not None:
                rgba = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGBA)
            else:
                import numpy as _np_tex
                rgba = img_bgr[:, :, ::-1]
                if rgba.shape[2] == 3:
                    alpha = 255 * _np_tex.ones((*rgba.shape[:2], 1), dtype=_np_tex.uint8)
                    rgba = _np_tex.concatenate([rgba, alpha], axis=2)
        except Exception:
            import numpy as _np_tex2
            rgba = img_bgr
        texture = Texture.create(size=(rgba.shape[1], rgba.shape[0]), colorfmt="rgba")
        buf = bytes(rgba.data)       # explicit copy; releases numpy reference
        texture.blit_buffer(buf, colorfmt="rgba", bufferfmt="ubyte")
        texture.flip_vertical()
        del rgba, buf                # free large intermediate buffers immediately
        self.texture = texture
        self._redraw_overlay()
        self.canvas.ask_update()
        Clock.schedule_once(lambda dt: self._redraw_overlay(), 0)

    def set_boxes_payload(self, boxes_payload, selected_ids=None, preview_zoom=1.0, page_idx=0, label_mode='all'):
        self.boxes_payload = list(boxes_payload or [])
        self.preview_zoom = float(preview_zoom or 1.0)
        self.page_idx = int(page_idx or 0)
        self.selected_ids = set(selected_ids or [])
        self.label_mode = str(label_mode or 'all')
        self.hovered_box_id = None
        self._redraw_overlay()

    def set_selected_ids(self, selected_ids):
        self.selected_ids = set(selected_ids or [])
        self._redraw_overlay()

    def set_capture_guides(self, capture_rect=None, guide_rects=None, guide_paths=None, guide_points=None):
        self.capture_rect_payload = dict(capture_rect or {}) if isinstance(capture_rect, dict) else capture_rect
        self.capture_split_rects_payload = [dict(r) if isinstance(r, dict) else r for r in (guide_rects or [])]
        self.capture_path_payload = [list(p) for p in (guide_paths or []) if isinstance(p, (list, tuple))]
        self.capture_points_payload = list(guide_points or [])
        self._redraw_overlay()

    def clear_capture_guides(self):
        self.capture_rect_payload = None
        self.capture_split_rects_payload = []
        self.capture_path_payload = []
        self.capture_points_payload = []
        self._redraw_overlay()

    def clear_boxes(self):
        self.boxes_payload = []
        self.selected_ids = set()
        self.hovered_box_id = None
        self.clear_capture_guides()
        self._active_touches = {}
        self._touch_refs = {}
        self._pinch_start_dist = None
        self._pinch_active = False
        self._pinch_last_midpoint = None
        self._tap_candidates = {}
        self._touch_start = {}
        self._touch_moved = set()
        self._redraw_overlay()

    def _get_display_rect(self):
        tex = self.texture
        if not tex or tex.width <= 0 or tex.height <= 0:
            return None
        widget_w, widget_h = self.width, self.height
        tex_w, tex_h = tex.width, tex.height
        scale = min(widget_w / float(tex_w), widget_h / float(tex_h))
        draw_w = tex_w * scale
        draw_h = tex_h * scale
        x = self.x + (widget_w - draw_w) / 2.0
        y = self.y + (widget_h - draw_h) / 2.0
        return x, y, draw_w, draw_h

    def _box_color(self, box_type, selected=False, hovered=False):
        if selected:
            return (0.15, 0.85, 1.0, 1.0)
        if hovered:
            return (1.0, 0.6, 0.0, 1.0)
        if box_type == "check":
            return (1.0, 0.15, 0.15, 1.0)
        if box_type == "line":
            return (1.0, 0.85, 0.1, 1.0)
        return (0.1, 1.0, 0.3, 1.0)

    def _draw_label(self, x, y, text, anchor="top_left"):
        core = CoreLabel(text=str(text), font_size=11)
        core.refresh()
        tex = core.texture
        pad_x, pad_y = 4, 2
        bg_w = tex.width + pad_x * 2
        bg_h = tex.height + pad_y * 2
        draw_x = float(x)
        draw_y = float(y)
        if anchor == "left":
            draw_x = x - bg_w - 4
            draw_y = y + bg_h / 2.0
        elif anchor == "right":
            draw_x = x + 4
            draw_y = y + bg_h / 2.0
        elif anchor == "bottom":
            draw_x = x
            draw_y = y - 4
        else:
            draw_x = x
            draw_y = y + bg_h + 4
        Rectangle(pos=(draw_x, draw_y - bg_h), size=(bg_w, bg_h))
        Color(1, 1, 1, 1)
        Rectangle(texture=tex, pos=(draw_x + pad_x, draw_y - bg_h + pad_y), size=tex.size)

    def _label_anchor_for_box(self, box, x, y, w, h, disp):
        dx, dy, dw, dh = disp
        box_type = str(box.get("t", "field"))
        small_box = (w <= 28 or h <= 22)
        if box_type == "check" or small_box:
            if x - dx > 42:
                return "left"
            if (dx + dw) - (x + w) > 42:
                return "right"
            return "bottom" if y - dy > 18 else "top_left"
        if y + h + 22 <= dy + dh:
            return "top_left"
        return "bottom"

    def _should_draw_box_label(self, box):
        mode = str(getattr(self, 'label_mode', 'all') or 'all').strip().lower()
        try:
            box_id = int(box.get('id'))
        except Exception:
            box_id = box.get('id')
        if mode in {'none', 'off', 'hidden'}:
            return False
        if mode in {'selected', 'selected_only', 'focus', 'focus_only'}:
            return (box_id in self.selected_ids) or (box_id == self.hovered_box_id)
        if mode in {'hovered', 'hovered_only'}:
            return box_id == self.hovered_box_id
        return True

    def _resolve_box_image_rect(self, box):
        if not isinstance(box, dict):
            return 0.0, 0.0, 0.0, 0.0
        try:
            if all(k in box for k in ("pdf_x", "pdf_y", "pdf_w", "pdf_h")):
                zoom = float(getattr(self, "preview_zoom", 1.0) or 1.0)
                return (
                    float(box.get("pdf_x", 0.0)) * zoom,
                    float(box.get("pdf_y", 0.0)) * zoom,
                    float(box.get("pdf_w", 0.0)) * zoom,
                    float(box.get("pdf_h", 0.0)) * zoom,
                )
        except Exception:
            pass
        return (
            float(box.get("x", 0.0) or 0.0),
            float(box.get("y", 0.0) or 0.0),
            float(box.get("w", 0.0) or 0.0),
            float(box.get("h", 0.0) or 0.0),
        )

    def _resolve_box_display_rect(self, box):
        disp = self._get_display_rect()
        tex = self.texture
        if not disp or tex is None:
            return None
        dx, dy, dw, dh = disp
        bx, by, bw, bh = self._resolve_box_image_rect(box)
        sx = dw / float(max(tex.width, 1))
        sy = dh / float(max(tex.height, 1))
        x = dx + bx * sx
        y = dy + (tex.height - (by + bh)) * sy
        w = max(1.0, bw * sx)
        h = max(1.0, bh * sy)
        return (x, y, w, h, disp)

    def _resolve_label_display_rect(self, box):
        geom = self._resolve_box_display_rect(box)
        if not geom:
            return None
        x, y, w, h, disp = geom
        core = CoreLabel(text=str(box.get("id", "")), font_size=11)
        core.refresh()
        tex = core.texture
        pad_x, pad_y = 4, 2
        bg_w = tex.width + pad_x * 2
        bg_h = tex.height + pad_y * 2
        anchor = self._label_anchor_for_box(box, x, y, w, h, disp)
        draw_x = float(x)
        draw_y = float(y)
        if anchor == "left":
            draw_x = x - bg_w - 4
            draw_y = y + bg_h / 2.0
        elif anchor == "right":
            draw_x = x + 4
            draw_y = y + bg_h / 2.0
        elif anchor == "bottom":
            draw_x = x
            draw_y = y - 4
        else:
            draw_x = x
            draw_y = y + bg_h + 4
        return (draw_x, draw_y - bg_h, bg_w, bg_h)

    def _find_hit_box_for_touch(self, touch):
        pt = self._touch_to_image_point(touch)
        hit = self._find_hit_box(*pt) if pt is not None else None
        if hit is not None:
            return hit
        if not self.collide_point(*touch.pos):
            return None
        label_hits = []
        for box in self.boxes_payload:
            if not self._should_draw_box_label(box):
                continue
            rect = self._resolve_label_display_rect(box)
            if rect is None:
                continue
            lx, ly, lw, lh = rect
            if lx <= touch.x <= lx + lw and ly <= touch.y <= ly + lh:
                label_hits.append((box, lw * lh))
        if not label_hits:
            return None
        label_hits.sort(key=lambda item: (item[1], int(item[0].get("id", 0))))
        return label_hits[0][0]

    def _set_pointed_box(self, hit=None, notify=True):
        new_id = None
        if isinstance(hit, dict) and "id" in hit:
            try:
                new_id = int(hit.get("id"))
            except Exception:
                new_id = hit.get("id")
        changed = (new_id != self.hovered_box_id)
        self.hovered_box_id = new_id
        if changed:
            self._redraw_overlay()
        if notify and callable(self.hover_callback):
            try:
                self.hover_callback(hit)
            except Exception:
                pass

    def _redraw_overlay(self, *args):
        self.canvas.after.clear()
        disp = self._get_display_rect()
        if not disp:
            return
        dx, dy, dw, dh = disp
        tex = self.texture
        sx = dw / float(max(tex.width, 1))
        sy = dh / float(max(tex.height, 1))
        with self.canvas.after:
            for box in self.boxes_payload:
                bx, by, bw, bh = self._resolve_box_image_rect(box)
                x = dx + bx * sx
                y = dy + (tex.height - (by + bh)) * sy
                w = max(1.0, bw * sx)
                h = max(1.0, bh * sy)
                selected = box["id"] in self.selected_ids
                hovered = (box["id"] == self.hovered_box_id)
                Color(*self._box_color(box.get("t", "field"), selected=selected, hovered=hovered))
                Line(rectangle=(x, y, w, h), width=3.2 if selected else (2.4 if hovered else 1.35))
                if self._should_draw_box_label(box):
                    Color(0, 0, 0, 0.88)
                    self._draw_label(x, y + h, box["id"], anchor=self._label_anchor_for_box(box, x, y, w, h, disp))

            def _draw_pdf_rect(rect_payload, rgba, width=2.0):
                if rect_payload is None:
                    return
                try:
                    bx, by, bw, bh = self._resolve_box_image_rect(rect_payload)
                except Exception:
                    return
                x = dx + bx * sx
                y = dy + (tex.height - (by + bh)) * sy
                w = max(1.0, bw * sx)
                h = max(1.0, bh * sy)
                Color(*rgba)
                Line(rectangle=(x, y, w, h), width=width)

            capture_rect = getattr(self, 'capture_rect_payload', None)
            if capture_rect:
                _draw_pdf_rect(capture_rect, (0.20, 0.90, 1.00, 0.95), width=2.6)
            for rect_payload in (getattr(self, 'capture_split_rects_payload', []) or []):
                _draw_pdf_rect(rect_payload, (0.15, 1.00, 0.38, 0.92), width=2.2)
            for path_payload in (getattr(self, 'capture_path_payload', []) or []):
                pts = []
                for pt in (path_payload or []):
                    if not isinstance(pt, (list, tuple)) or len(pt) < 2:
                        continue
                    try:
                        img_x = float(pt[0]); img_y = float(pt[1])
                    except Exception:
                        continue
                    px = dx + img_x * sx
                    py = dy + (tex.height - img_y) * sy
                    pts.extend([px, py])
                if len(pts) >= 4:
                    Color(1.00, 0.58, 0.12, 0.98)
                    Line(points=pts, width=2.4, close=True)
            for point in (getattr(self, 'capture_points_payload', []) or []):
                if not isinstance(point, (list, tuple)) or len(point) < 2:
                    continue
                try:
                    img_x = float(point[0])
                    img_y = float(point[1])
                except Exception:
                    continue
                px = dx + img_x * sx
                py = dy + (tex.height - img_y) * sy
                rad = max(3.0, float(dp(4)))
                Color(1.0, 0.92, 0.20, 0.95)
                Ellipse(pos=(px - rad, py - rad), size=(rad * 2.0, rad * 2.0))

    def _touch_to_image_point(self, touch):
        disp = self._get_display_rect()
        tex = self.texture
        if not disp or not tex:
            return None
        dx, dy, dw, dh = disp
        if not (dx <= touch.x <= dx + dw and dy <= touch.y <= dy + dh):
            return None
        img_x = (touch.x - dx) * tex.width / float(max(dw, 1))
        img_y = tex.height - ((touch.y - dy) * tex.height / float(max(dh, 1)))
        return img_x, img_y

    def _android_hit_padding_img(self):
        try:
            scale = float(getattr(self, "display_scale", 1.0) or 1.0)
        except Exception:
            scale = 1.0
        finger_px = float(dp(18))
        return max(4.0, finger_px / max(scale, 0.35))

    def _find_hit_box(self, img_x, img_y):
        hits = []
        for box in self.boxes_payload:
            bx, by, bw, bh = self._resolve_box_image_rect(box)
            if bx <= img_x <= bx + bw and by <= img_y <= by + bh:
                hits.append((box, bw * bh, 0.0))
        if not hits and platform == "android":
            pad = self._android_hit_padding_img()
            near_hits = []
            for box in self.boxes_payload:
                bx, by, bw, bh = self._resolve_box_image_rect(box)
                ex0, ey0, ex1, ey1 = (bx - pad, by - pad, bx + bw + pad, by + bh + pad)
                if ex0 <= img_x <= ex1 and ey0 <= img_y <= ey1:
                    cx = bx + (bw / 2.0)
                    cy = by + (bh / 2.0)
                    near_hits.append((box, bw * bh, math.hypot(img_x - cx, img_y - cy)))
            hits = near_hits
        if not hits:
            return None
        type_priority = {"check": 0, "field": 1, "line": 2}
        hits.sort(key=lambda item: (item[2], item[1], type_priority.get(str(item[0].get("t", "field")), 9), item[0]["id"]))
        return hits[0][0]

    def _register_tap(self, x, y, window=0.55, distance=28.0):
        now = time.time()
        fresh = []
        for ts, tx, ty in getattr(self, "_recent_taps", []):
            if (now - float(ts)) <= float(window) and math.hypot(float(tx) - float(x), float(ty) - float(y)) <= float(distance):
                fresh.append((float(ts), float(tx), float(ty)))
        fresh.append((now, float(x), float(y)))
        self._recent_taps = fresh[-3:]
        return len(self._recent_taps)

    def _should_block_box_double_tap(self):
        try:
            app = App.get_running_app()
        except Exception:
            app = None
        if app is None:
            return False
        try:
            return bool(getattr(app, "ui_mobile", False)) and bool(getattr(app, "mobile_select_mode", False))
        except Exception:
            return False


    def _begin_android_pinch(self):
        if len(self._active_touches) < 2:
            return False
        pts = list(self._active_touches.values())[:2]
        self._pinch_start_dist = math.hypot(pts[1][0] - pts[0][0], pts[1][1] - pts[0][1])
        self._pinch_start_scale = float(getattr(self, "display_scale", 1.0) or 1.0)
        self._pinch_active = True
        self._pinch_last_midpoint = (
            (float(pts[0][0]) + float(pts[1][0])) / 2.0,
            (float(pts[0][1]) + float(pts[1][1])) / 2.0,
        )
        for ref in list(self._touch_refs.values())[:2]:
            try:
                if getattr(ref, "grab_current", None) is not self:
                    ref.grab(self)
            except Exception:
                pass
        return True

    def _release_android_tracked_grabs(self):
        for ref in list(getattr(self, "_touch_refs", {}).values()):
            try:
                if getattr(ref, "grab_current", None) is self:
                    ref.ungrab(self)
            except Exception:
                pass

    def on_touch_down(self, touch):
        if getattr(touch, "is_mouse_scrolling", False):
            if self.collide_point(*touch.pos):
                direction = "in" if getattr(touch, "button", "") == "scrollup" else "out"
                if callable(self.zoom_request_callback) and getattr(touch, "button", "") in ("scrollup", "scrolldown"):
                    self.zoom_request_callback("step", direction)
                    return True
            return super().on_touch_down(touch)

        if platform != "android":
            if not self.collide_point(*touch.pos):
                return super().on_touch_down(touch)

            pt = self._touch_to_image_point(touch)
            hit = self._find_hit_box_for_touch(touch)
            if pt is None and hit is None:
                return super().on_touch_down(touch)

            try:
                touch.grab(self)
            except Exception:
                pass

            self._active_touches[touch.uid] = (touch.x, touch.y)
            if len(self._active_touches) == 2:
                pts = list(self._active_touches.values())[:2]
                self._pinch_start_dist = math.hypot(pts[1][0] - pts[0][0], pts[1][1] - pts[0][1])
                self._pinch_start_scale = float(getattr(self, "display_scale", 1.0) or 1.0)
                return True

            if bool(getattr(self, 'capture_mode_active', False)) and callable(self.preview_tap_callback) and pt is not None:
                try:
                    handled = bool(self.preview_tap_callback(pt, hit))
                except Exception:
                    handled = False
                if handled:
                    try:
                        touch.ud['_capture_point_handled'] = True
                    except Exception:
                        pass
                    self._set_pointed_box(hit, notify=True)
                    return True

            tap_count = self._register_tap(touch.x, touch.y)

            if hit is not None and tap_count >= 2:
                if self._should_block_box_double_tap():
                    if callable(self.box_tap_callback):
                        self.box_tap_callback(hit)
                    self._recent_taps = []
                    return True
                if callable(self.box_double_tap_callback):
                    self.box_double_tap_callback(hit)
                    self._recent_taps = []
                    return True

            if hit is None and pt is not None and tap_count >= 3:
                if callable(self.zoom_request_callback):
                    self.zoom_request_callback("step", "in")
                    self._recent_taps = []
                    return True

            if hit is not None:
                if callable(self.box_tap_callback):
                    self.box_tap_callback(hit)
                return True
            return super().on_touch_down(touch)

        if not self.collide_point(*touch.pos):
            return super().on_touch_down(touch)

        pt = self._touch_to_image_point(touch)
        hit = self._find_hit_box_for_touch(touch)
        self._touch_refs[touch.uid] = touch
        self._set_pointed_box(hit, notify=True)
        self._touch_start[touch.uid] = (touch.x, touch.y)
        self._tap_candidates[touch.uid] = ({"hit": hit, "x": touch.x, "y": touch.y, "img_pt": pt} if pt is not None else None)
        self._active_touches[touch.uid] = (touch.x, touch.y)

        if len(self._active_touches) >= 2:
            self._begin_android_pinch()
            return True

        try:
            touch.grab(self)
        except Exception:
            pass
        return True

    def on_touch_move(self, touch):
        if platform == "android":
            start = self._touch_start.get(touch.uid)
            if start is not None:
                if math.hypot(float(touch.x) - float(start[0]), float(touch.y) - float(start[1])) > float(self._tap_slop):
                    self._touch_moved.add(touch.uid)
            if touch.uid in self._active_touches:
                prev_pt = self._active_touches.get(touch.uid, (touch.x, touch.y))
                self._active_touches[touch.uid] = (touch.x, touch.y)
                self._touch_refs[touch.uid] = touch
                if len(self._active_touches) >= 2:
                    if not self._pinch_active:
                        self._begin_android_pinch()
                    if callable(self.zoom_request_callback):
                        pts = list(self._active_touches.values())[:2]
                        cur_dist = math.hypot(pts[1][0] - pts[0][0], pts[1][1] - pts[0][1])
                        midpoint = (
                            (float(pts[0][0]) + float(pts[1][0])) / 2.0,
                            (float(pts[0][1]) + float(pts[1][1])) / 2.0,
                        )
                        last_mid = getattr(self, "_pinch_last_midpoint", midpoint) or midpoint
                        pan_delta = (
                            float(midpoint[0]) - float(last_mid[0]),
                            float(midpoint[1]) - float(last_mid[1]),
                        )
                        self._pinch_last_midpoint = midpoint
                        if self._pinch_start_dist and cur_dist > 0:
                            factor = cur_dist / float(max(self._pinch_start_dist, 1e-6))
                            self.zoom_request_callback("gesture", {
                                "scale": self._pinch_start_scale * factor,
                                "focus": midpoint,
                                "pan": pan_delta,
                            })
                    return True
                current_hit = self._find_hit_box_for_touch(touch)
                self._set_pointed_box(current_hit, notify=(current_hit is not None or self.hovered_box_id is not None))
                if getattr(touch, "grab_current", None) is self:
                    candidate = self._tap_candidates.get(touch.uid)
                    zoomed = float(getattr(self, "display_scale", 1.0) or 1.0) > 1.05
                    if candidate is not None and zoomed and callable(self.zoom_request_callback):
                        dx = float(touch.x) - float(prev_pt[0])
                        dy = float(touch.y) - float(prev_pt[1])
                        if abs(dx) > 0.01 or abs(dy) > 0.01:
                            self.zoom_request_callback("gesture", {
                                "scale": float(getattr(self, "display_scale", 1.0) or 1.0),
                                "focus": (float(touch.x), float(touch.y)),
                                "pan": (dx, dy),
                            })
                        return True
            if getattr(touch, "grab_current", None) is self:
                return True
            return super().on_touch_move(touch)

        if touch.uid in self._active_touches:
            prev_pt = self._active_touches.get(touch.uid, (touch.x, touch.y))
            self._active_touches[touch.uid] = (touch.x, touch.y)
            if len(self._active_touches) >= 2 and callable(self.zoom_request_callback):
                pts = list(self._active_touches.values())[:2]
                cur_dist = math.hypot(pts[1][0] - pts[0][0], pts[1][1] - pts[0][1])
                if self._pinch_start_dist and cur_dist > 0:
                    factor = cur_dist / float(max(self._pinch_start_dist, 1e-6))
                    midpoint = (
                        (float(pts[0][0]) + float(pts[1][0])) / 2.0,
                        (float(pts[0][1]) + float(pts[1][1])) / 2.0,
                    )
                    self.zoom_request_callback("gesture", {
                        "scale": self._pinch_start_scale * factor,
                        "focus": midpoint,
                        "pan": (float(touch.x) - float(prev_pt[0]), float(touch.y) - float(prev_pt[1])),
                    })
                    return True
        return super().on_touch_move(touch)

    def on_touch_up(self, touch):
        if platform == "android":
            was_grabbed = (getattr(touch, "grab_current", None) is self)
            if was_grabbed:
                try:
                    touch.ungrab(self)
                except Exception:
                    pass

            had_multi_touch = len(self._active_touches) >= 2 or bool(self._pinch_active)
            moved = (touch.uid in self._touch_moved)
            candidate = self._tap_candidates.get(touch.uid)

            if touch.uid in self._active_touches:
                self._active_touches.pop(touch.uid, None)
            self._touch_refs.pop(touch.uid, None)
            self._touch_start.pop(touch.uid, None)
            self._tap_candidates.pop(touch.uid, None)
            self._touch_moved.discard(touch.uid)

            if len(self._active_touches) < 2:
                self._pinch_start_dist = None
                self._pinch_active = False
                self._pinch_last_midpoint = None
                self._release_android_tracked_grabs()
                if had_multi_touch and callable(self.zoom_request_callback):
                    self.zoom_request_callback("gesture_end", None)

            if candidate and (not moved) and (not had_multi_touch):
                hit = candidate.get("hit")
                img_pt = candidate.get("img_pt")
                tap_count = self._register_tap(float(candidate.get("x", touch.x)), float(candidate.get("y", touch.y)))
                if callable(self.preview_tap_callback):
                    try:
                        handled = bool(self.preview_tap_callback(img_pt, hit))
                    except Exception:
                        handled = False
                    if handled:
                        self._set_pointed_box(hit, notify=True)
                        return True
                if hit is not None and tap_count >= 2:
                    if self._should_block_box_double_tap():
                        if callable(self.box_tap_callback):
                            self.box_tap_callback(hit)
                        self._recent_taps = []
                        self._set_pointed_box(hit, notify=True)
                        return True
                    if callable(self.box_double_tap_callback):
                        self.box_double_tap_callback(hit)
                        self._recent_taps = []
                        self._set_pointed_box(hit, notify=True)
                        return True
                if hit is not None:
                    if callable(self.box_tap_callback):
                        self.box_tap_callback(hit)
                        self._set_pointed_box(hit, notify=True)
                        return True
                if hit is None and tap_count >= 3:
                    if callable(self.zoom_request_callback):
                        self.zoom_request_callback("step", "in")
                        self._recent_taps = []
                        return True
            elif not had_multi_touch:
                self._set_pointed_box(None, notify=bool(self.hovered_box_id is not None))

            if was_grabbed or had_multi_touch:
                return True
            return super().on_touch_up(touch)

        was_grabbed = (getattr(touch, "grab_current", None) is self)
        if was_grabbed:
            try:
                touch.ungrab(self)
            except Exception:
                pass

        if bool(getattr(getattr(touch, 'ud', {}), 'get', lambda *a, **k: False)('_capture_point_handled', False)):
            if touch.uid in self._active_touches:
                self._active_touches.pop(touch.uid, None)
                if len(self._active_touches) < 2:
                    self._pinch_start_dist = None
            return True

        moved = False
        start = self._active_touches.get(touch.uid)
        if start is not None:
            try:
                moved = math.hypot(float(touch.x) - float(start[0]), float(touch.y) - float(start[1])) > float(self._tap_slop)
            except Exception:
                moved = False

        had_multi_touch = len(self._active_touches) >= 2
        if touch.uid in self._active_touches:
            self._active_touches.pop(touch.uid, None)
            if len(self._active_touches) < 2:
                self._pinch_start_dist = None
                if had_multi_touch and callable(self.zoom_request_callback):
                    self.zoom_request_callback("gesture_end", None)

        if self.collide_point(*touch.pos):
            pt = self._touch_to_image_point(touch)
            hit = self._find_hit_box_for_touch(touch)
            if (not moved) and (pt is not None or hit is not None):
                tap_count = self._register_tap(float(touch.x), float(touch.y))
                if callable(self.preview_tap_callback):
                    try:
                        handled = bool(self.preview_tap_callback(pt, hit))
                    except Exception:
                        handled = False
                    if handled:
                        self._set_pointed_box(hit, notify=True)
                        return True
                if hit is not None and tap_count >= 2:
                    if self._should_block_box_double_tap():
                        if callable(self.box_tap_callback):
                            self.box_tap_callback(hit)
                        self._recent_taps = []
                        self._set_pointed_box(hit, notify=True)
                        return True
                    if callable(self.box_double_tap_callback):
                        self.box_double_tap_callback(hit)
                        self._recent_taps = []
                        self._set_pointed_box(hit, notify=True)
                        return True
                if hit is not None:
                    if callable(self.box_tap_callback):
                        self.box_tap_callback(hit)
                        self._set_pointed_box(hit, notify=True)
                        return True
                if hit is None and tap_count >= 3:
                    if callable(self.zoom_request_callback):
                        self.zoom_request_callback("step", "in")
                        self._recent_taps = []
                        return True

        return super().on_touch_up(touch)

    def _on_mouse_pos(self, _window, pos):
        if platform == "android" or not self.get_root_window():
            return
        local = self.to_widget(*pos)
        class Dummy: pass
        d = Dummy(); d.x=local[0]; d.y=local[1]; d.pos=(local[0], local[1])
        hit = self._find_hit_box_for_touch(d)
        self._set_pointed_box(hit, notify=(hit is not None or self.hovered_box_id is not None))


# Main application layout

class FormAlchemistLayout(BoxLayout):
    pass




class FloatingPreviewHUD(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._drag_touch = None
        self._drag_offset = (0, 0)
        self._collapsed = False
        self._expanded_height = None
        self._body_widget = None
        self._toggle_button = None
        self._drag_handle_height = dp(40)
        self._interactive_widgets = []
        self._allow_body_passthrough = True
        self._drag_enabled = True
        self._collapse_callbacks = []

    def set_body_widget(self, widget, expanded_height=None):
        self._body_widget = widget
        if expanded_height is not None:
            self._expanded_height = float(expanded_height)
        elif widget is not None:
            self._expanded_height = max(
                float(getattr(self, 'height', 0) or 0),
                float(getattr(widget, 'height', 0) or 0) + self._drag_handle_height + dp(18)
            )

    def set_toggle_button(self, button):
        self._toggle_button = button
        if button is not None:
            button.bind(on_release=self.toggle_collapsed)
            self._refresh_collapse_state()

    def register_interactive_widgets(self, *widgets):
        collected = []
        for widget in widgets:
            if widget is None:
                continue
            if isinstance(widget, (list, tuple, set)):
                for sub in widget:
                    if sub is not None:
                        collected.append(sub)
            else:
                collected.append(widget)
        self._interactive_widgets = collected

    def set_drag_enabled(self, enabled=True):
        self._drag_enabled = bool(enabled)

    def set_allow_body_passthrough(self, enabled=True):
        self._allow_body_passthrough = bool(enabled)

    def bind_collapse_state(self, callback):
        if callable(callback):
            self._collapse_callbacks.append(callback)

    def _emit_collapse_state(self):
        for callback in list(self._collapse_callbacks):
            try:
                callback(bool(self._collapsed))
            except Exception:
                pass

    def toggle_collapsed(self, *_):
        self._collapsed = not bool(self._collapsed)
        self._refresh_collapse_state()

    def _refresh_collapse_state(self):
        if self._body_widget is not None:
            self._body_widget.opacity = 0 if self._collapsed else 1
            self._body_widget.disabled = bool(self._collapsed)
            self._body_widget.height = 0 if self._collapsed else max(
                dp(1),
                float(getattr(self._body_widget, 'minimum_height', 0) or getattr(self._body_widget, 'height', 0) or dp(110))
            )
            self._body_widget.size_hint_y = None
        if self._collapsed:
            self.height = self._drag_handle_height + dp(8)
        else:
            self.height = max(self._drag_handle_height + dp(8), float(self._expanded_height or self.height or dp(150)))
        if self._toggle_button is not None:
            collapsed_label = getattr(self._toggle_button, '_collapsed_label', 'Show')
            expanded_label = getattr(self._toggle_button, '_expanded_label', 'Hide')
            self._toggle_button.text = collapsed_label if self._collapsed else expanded_label
        self._emit_collapse_state()

    def _local_point_in_widget(self, widget, touch):
        try:
            lx, ly = widget.to_widget(*touch.pos, relative=True)
            return (0 <= lx <= widget.width) and (0 <= ly <= widget.height)
        except Exception:
            return False

    def _hud_contains(self, touch):
        return self._local_point_in_widget(self, touch)

    def _handle_contains(self, touch):
        if not self._hud_contains(touch):
            return False
        if self._toggle_button is not None and self._touch_hits_widget_tree(self._toggle_button, touch):
            return False
        if self._touch_hits_interactive(touch):
            return False
        return True

    def _find_deepest_touch_target(self, widget, touch):
        if widget is None:
            return None
        try:
            children = list(getattr(widget, 'children', []) or [])
        except Exception:
            children = []
        for child in children:
            target = self._find_deepest_touch_target(child, touch)
            if target is not None:
                return target
        if self._local_point_in_widget(widget, touch):
            return widget
        return None

    def _touch_hits_widget_tree(self, widget, touch):
        return self._find_deepest_touch_target(widget, touch) is not None

    def _find_interactive_target(self, touch):
        if self._toggle_button is not None:
            target = self._find_deepest_touch_target(self._toggle_button, touch)
            if target is not None:
                return target
        for widget in self._interactive_widgets:
            target = self._find_deepest_touch_target(widget, touch)
            if target is not None:
                return target
        return None

    def _touch_hits_interactive(self, touch):
        return self._find_interactive_target(touch) is not None

    def _dispatch_to_interactive_target(self, touch, method_name):
        target = None
        if method_name == 'on_touch_down':
            target = self._find_interactive_target(touch)
            if target is not None:
                touch.ud['floating_preview_hud_target'] = target
        else:
            target = touch.ud.get('floating_preview_hud_target')
            if target is None and self._hud_contains(touch):
                target = self._find_interactive_target(touch)
        if target is None:
            return None
        handler = getattr(target, method_name, None)
        if not callable(handler):
            return None
        try:
            return bool(handler(touch))
        except Exception:
            return None

    def on_touch_down(self, touch):
        if self._drag_enabled and self._handle_contains(touch):
            try:
                px, py = self.parent.to_widget(*touch.pos, relative=True)
            except Exception:
                px, py = touch.pos
            self._drag_touch = touch
            self._drag_offset = (px - self.x, py - self.y)
            touch.grab(self)
            self.pos_hint = {}
            return True
        if self._hud_contains(touch) and self._touch_hits_interactive(touch):
            dispatched = self._dispatch_to_interactive_target(touch, 'on_touch_down')
            if dispatched is not None:
                return dispatched
            return super().on_touch_down(touch)
        if self._allow_body_passthrough and self._hud_contains(touch):
            return False
        return super().on_touch_down(touch)

    def on_touch_move(self, touch):
        if touch.grab_current is self:
            parent = self.parent
            if parent is None:
                return True
            try:
                px, py = parent.to_widget(*touch.pos, relative=True)
            except Exception:
                px, py = touch.pos
            new_x = px - self._drag_offset[0]
            new_y = py - self._drag_offset[1]
            max_x = max(0.0, parent.width - self.width)
            max_y = max(0.0, parent.height - self.height)
            self.x = min(max(0.0, new_x), max_x)
            self.y = min(max(0.0, new_y), max_y)
            return True
        if self._hud_contains(touch) and (self._touch_hits_interactive(touch) or touch.ud.get('floating_preview_hud_target') is not None):
            dispatched = self._dispatch_to_interactive_target(touch, 'on_touch_move')
            if dispatched is not None:
                return dispatched
            return super().on_touch_move(touch)
        if self._allow_body_passthrough and self._hud_contains(touch):
            return False
        return super().on_touch_move(touch)

    def on_touch_up(self, touch):
        if touch.grab_current is self:
            touch.ungrab(self)
            self._drag_touch = None
            return True
        if self._hud_contains(touch) and (self._touch_hits_interactive(touch) or touch.ud.get('floating_preview_hud_target') is not None):
            dispatched = self._dispatch_to_interactive_target(touch, 'on_touch_up')
            touch.ud.pop('floating_preview_hud_target', None)
            if dispatched is not None:
                return dispatched
            return super().on_touch_up(touch)
        touch.ud.pop('floating_preview_hud_target', None)
        if self._allow_body_passthrough and self._hud_contains(touch):
            return False
        return super().on_touch_up(touch)


# Main application class

class FormAlchemistApp(MDApp):

    def build(self):
        self.title = APP_TITLE
        self.icon = "assets/icon.png"
        self.engine = FormAlchemistEngine()
        self.import_store = LocalImportStore(self.get_app_output_dir)
        self.android_picker = AndroidDocumentPickerService(self.import_store, status_callback=self.set_status)

        if KIVYMD_AVAILABLE and hasattr(self, "theme_cls"):
            try:
                self.theme_cls.theme_style = "Dark"
                self.theme_cls.primary_palette = "Blue"
                self.theme_cls.accent_palette = "Teal"
                if hasattr(self.theme_cls, "material_style"):
                    self.theme_cls.material_style = "M3"
            except Exception:
                pass

        self.force_mobile_layout = (os.environ.get("MEDIMAP_FORCE_MOBILE_UI", "0") == "1")
        desktop_phone_emulation = (platform != "android" and self.force_mobile_layout)
        is_mobile = (platform == "android") or desktop_phone_emulation
        self.ui_mobile = is_mobile
        self._is_mobile_ui = is_mobile
        self.desktop_phone_emulation = desktop_phone_emulation
        pad = dp(8) if is_mobile else dp(14)
        gap = dp(8) if is_mobile else dp(12)
        row_h = dp(52) if is_mobile else dp(46)
        input_h = dp(52) if is_mobile else dp(44)

        palette = {
            "bg": (0.006, 0.010, 0.030, 1),
            "surface": (0.028, 0.052, 0.110, 0.992),
            "surface_alt": (0.042, 0.068, 0.138, 0.995),
            "surface_soft": (0.064, 0.092, 0.172, 0.995),
            "primary": (0.255, 0.500, 0.980, 1),
            "primary_soft": (0.094, 0.144, 0.330, 1),
            "accent": (0.960, 0.710, 0.300, 1),
            "accent_soft": (0.280, 0.170, 0.058, 1),
            "secondary": (0.585, 0.515, 0.900, 1),
            "aurora": (0.255, 0.820, 0.960, 1),
            "starlight": (1.000, 0.950, 0.800, 1),
            "text": (0.970, 0.980, 1.000, 1),
            "muted": (0.710, 0.785, 0.930, 1),
            "border": (0.205, 0.275, 0.455, 0.54),
            "preview_bg": (0.004, 0.008, 0.022, 1),
            "chip": (0.055, 0.080, 0.152, 0.992),
            "ghost": (0.052, 0.078, 0.146, 0.86),
            "mini": (0.046, 0.070, 0.126, 0.96),
            "danger": (1.000, 0.360, 0.560, 1),
        }
        Window.clearcolor = palette["bg"]
        # Android: push layout above the soft keyboard so inputs stay visible
        if platform == "android":
            try:
                Window.softinput_mode = "below_target"
            except Exception:
                pass
        self.ui_palette = palette
        self._page_selected_box_ids = {}
        self._active_selection_page_idx = 0
        self.preview_zoom_mode = "fit_width" if is_mobile else "manual"
        self.preview_zoom_factor = 1.0
        self.desktop_zoom_mode = "manual"
        self.desktop_zoom_factor = 1.0
        self._mobile_zoom_bootstrap_pending = bool(is_mobile)
        self.mobile_select_mode = False
        self._mobile_quick_rail_open = False
        self._repair_anchor_box_ids = []
        self._repair_anchor_page_idx = -1
        self._repair_last_proposal = None
        self._capture_mode_active = False
        self._capture_page_idx = -1
        self._capture_points = []
        self._capture_preview_rect = None
        self._capture_split_preview = []
        self._capture_outline_preview = []
        self._capture_shape_candidates = []
        self._capture_shape_candidate_index = 0
        self._capture_active_template = None
        self.text_tuning_ui = self.engine._normalized_text_tuning(getattr(self.engine, "text_tuning", None))

        def style_card(widget, color=None, radius=dp(22)):
            color = color or palette["surface"]
            with widget.canvas.before:
                widget._card_shadow_color = Color(0, 0, 0, 0.18)
                widget._card_shadow = RoundedRectangle(radius=[radius + dp(2)] * 4)
                widget._card_bg_color = Color(*color)
                widget._card_bg = RoundedRectangle(radius=[radius] * 4)
                widget._card_shine_color = Color(1, 1, 1, 0.025)
                widget._card_shine = RoundedRectangle(radius=[max(dp(8), radius - dp(4))] * 4)
                widget._card_border_color = Color(*palette["border"])
                widget._card_border = Line(rounded_rectangle=[0, 0, 0, 0, radius], width=0.9)
            def _upd(*_):
                x, y = widget.pos
                w, h = widget.size
                widget._card_shadow.pos = (x, y - dp(1))
                widget._card_shadow.size = (w, h + dp(1.5))
                widget._card_bg.pos = (x, y)
                widget._card_bg.size = (w, h)
                widget._card_shine.pos = (x + dp(1), y + h - max(dp(14), h * 0.24) - dp(1))
                widget._card_shine.size = (max(0, w - dp(2)), max(dp(12), h * 0.24))
                widget._card_border.rounded_rectangle = [x, y, w, h, radius]
            widget.bind(pos=_upd, size=_upd)
            _upd()
            return widget

        def make_card(title, subtitle=None):
            card = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(12), size_hint_y=None)
            card.bind(minimum_height=card.setter("height"))
            style_card(card, palette["surface"])

            head = BoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(3))
            if is_mobile:
                head.bind(minimum_height=head.setter("height"))
            else:
                head.height = dp(44) if subtitle else dp(24)

            ttl = Label(
                text=title,
                markup=False,
                bold=True,
                color=palette["text"],
                size_hint_y=None,
                height=dp(24),
                halign="left",
                valign="middle",
                font_size=dp(16) if is_mobile else dp(15),
            )
            ttl.bind(size=self._sync_label_text_size)
            if is_mobile:
                self._bind_auto_height_label(ttl, min_height=dp(22), extra_pad=dp(1))
            head.add_widget(ttl)

            if subtitle:
                sub = Label(
                    text=subtitle,
                    color=palette["muted"],
                    size_hint_y=None,
                    height=dp(22),
                    halign="left",
                    valign="middle",
                    font_size=dp(11) if is_mobile else dp(11),
                )
                sub.bind(size=self._sync_label_text_size)
                if is_mobile:
                    self._bind_auto_height_label(sub, min_height=dp(16), extra_pad=dp(2))
                head.add_widget(sub)

            card.add_widget(head)

            body = GridLayout(cols=1, spacing=dp(10), size_hint_y=None)
            body.bind(minimum_height=body.setter("height"))
            card.add_widget(body)
            return card, body

        def make_button(text, tone="primary", icon=None):
            colors = {
                "primary": palette["primary"],
                "soft": palette["primary_soft"],
                "accent": palette["accent"],
                "secondary": palette.get("secondary", palette["accent"]),
                "plain": palette["surface_soft"],
                "danger": palette["danger"],
            }
            label = text if not icon else f"{icon}  {text}"

            if is_mobile:
                btn = NeoMobileButton(
                    text=label,
                    tone=tone,
                    theme=palette,
                    size_hint_y=None,
                    height=max(dp(44), row_h - dp(4)),  # Android HIG: min 44dp tap target
                    font_size=dp(10.6),
                    padding=[dp(7), dp(6), dp(7), dp(6)],
                )
                return btn

            if KIVYMD_AVAILABLE and MDRaisedButton is not None:
                if tone in ("primary", "accent", "danger"):
                    btn = MDRaisedButton(
                        text=label,
                        size_hint_y=None,
                        height=row_h,
                        md_bg_color=colors.get(tone, palette["primary"]),
                        text_color=(1, 1, 1, 1),
                        elevation=1,
                    )
                elif tone == "plain":
                    btn = MDRectangleFlatButton(
                        text=label,
                        size_hint_y=None,
                        height=row_h,
                        line_color=palette["border"],
                        text_color=palette["text"],
                    )
                else:
                    btn = MDFlatButton(
                        text=label,
                        size_hint_y=None,
                        height=row_h,
                        theme_text_color="Custom",
                        text_color=palette["text"],
                    )
                return btn

            txt_color = (1,1,1,1) if tone in ("primary","accent","danger") else palette["text"]
            btn = Button(
                text=label,
                size_hint_y=None, height=row_h,
                background_normal="", background_down="",
                background_color=colors.get(tone, palette["primary"]),
                color=txt_color, font_size=dp(14) if is_mobile else dp(13),
            )
            return btn

        def make_input(default="", hint="", input_filter=None):
            if (not is_mobile) and KIVYMD_AVAILABLE and MDTextField is not None:
                field = MDTextField(
                    text=str(default),
                    hint_text=hint,
                    mode="fill",
                    size_hint_y=None,
                    height=input_h + dp(8),
                    helper_text_mode="on_focus",
                )
                try:
                    field.fill_color_normal = palette["surface_alt"]
                    field.fill_color_focus = palette["surface_alt"]
                    field.line_color_normal = palette["border"]
                    field.line_color_focus = palette["primary"]
                    field.text_color_normal = palette["text"]
                    field.text_color_focus = palette["text"]
                    field.hint_text_color_normal = palette["muted"]
                    field.hint_text_color_focus = palette["primary"]
                except Exception:
                    pass
                if input_filter is not None:
                    try:
                        field.input_filter = input_filter
                    except Exception:
                        pass
                return field

            kwargs = dict(text=str(default), hint_text=hint, multiline=False,
                          size_hint_y=None, height=input_h, font_size=(dp(13) if is_mobile else dp(14)), write_tab=False,
                          background_normal="", background_active="",
                          background_color=palette["surface_alt"], foreground_color=palette["text"],
                          cursor_color=palette["primary"], padding=[dp(10), dp(10), dp(10), dp(10)])
            if input_filter is not None:
                kwargs["input_filter"] = input_filter
            return TextInput(**kwargs)

        def make_spinner(text, values=(), picker_title=None, search_hint=None):
            return SearchableSelectField(
                text=text,
                values=list(values),
                placeholder=text,
                picker_title=(picker_title or text),
                search_hint=(search_hint or f"Type to filter {str(text).lower()}"),
                size_hint_y=None,
                height=input_h,
                font_size=dp(14),
                background_normal="",
                background_color=palette["surface_alt"],
                color=palette["text"],
            )

        def labeled_field(label_text, widget, helper=None):
            box = BoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(4))
            box.bind(minimum_height=box.setter("height"))
            lbl = Label(text=label_text, color=palette["text"], size_hint_y=None,
                        height=dp(20), halign="left", valign="middle", font_size=dp(12))
            lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(lbl, min_height=dp(18), extra_pad=dp(4))
            box.add_widget(lbl)
            box.add_widget(widget)
            if helper:
                hlp = Label(text=helper, color=palette["muted"], size_hint_y=None,
                            height=dp(18), halign="left", valign="middle", font_size=dp(12) if is_mobile else dp(11))
                hlp.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(hlp, min_height=dp(14), extra_pad=dp(4))
                box.add_widget(hlp)
            return box

        def labeled_checkbox(label_text, checkbox, helper=None):
            slot_w = dp(56 if is_mobile else 64)
            row_h = dp(44 if is_mobile else 36)
            row = BoxLayout(orientation="horizontal", size_hint_y=None, height=row_h, spacing=dp(8))
            lbl_cls = MDLabel if KIVYMD_AVAILABLE and MDLabel is not None else Label
            lbl_kwargs = dict(text=label_text, halign="left", valign="middle")
            if lbl_cls is Label:
                lbl_kwargs.update(color=palette["text"], font_size=(dp(12) if is_mobile else dp(13)))
            else:
                lbl_kwargs.update(theme_text_color="Custom", text_color=palette["text"], font_style="Body1")
            lbl = lbl_cls(**lbl_kwargs)
            if lbl_cls is Label:
                lbl.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(lbl, min_height=(dp(22) if is_mobile else dp(18)), extra_pad=dp(4))
            row.add_widget(lbl)
            toggle_slot = AnchorLayout(
                anchor_x="right",
                anchor_y="center",
                size_hint=(None, None),
                width=slot_w,
                height=row_h,
                padding=(0, 0, dp(8 if is_mobile else 4), 0),
            )
            checkbox.size_hint = (None, None)
            checkbox.size = ((dp(32), dp(24)) if is_mobile else (dp(32), dp(22)))
            toggle_slot.add_widget(checkbox)
            row.add_widget(toggle_slot)
            if helper:
                wrap = BoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(3))
                wrap.bind(minimum_height=wrap.setter("height"))
                wrap.add_widget(row)
                sub = Label(text=helper, color=palette["muted"], size_hint_y=None,
                            height=(dp(22) if is_mobile else dp(16)), halign="left", valign="middle",
                            font_size=(dp(11) if is_mobile else dp(11)))
                sub.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(sub, min_height=(dp(18) if is_mobile else dp(14)), extra_pad=dp(4))
                wrap.add_widget(sub)
                return wrap
            return row

        def _toggle_phone_view(*_):
            if platform == "android":
                return
            next_state = not bool(getattr(self, "force_mobile_layout", False))
            os.environ["MEDIMAP_FORCE_MOBILE_UI"] = "1" if next_state else "0"
            self.set_status("Switching desktop view mode...")
            try:
                os.execv(sys.executable, [sys.executable] + sys.argv)
            except Exception as e:
                self.set_status(f"Could not restart for view toggle:\n{e}\nSet MEDIMAP_FORCE_MOBILE_UI={'1' if next_state else '0'} and relaunch.")
        self._toggle_phone_view = _toggle_phone_view

        def _make_preview_hud():
            hud = FloatingPreviewHUD(
                orientation="vertical",
                spacing=dp(8),
                size_hint=((1, None) if is_mobile else (None, None)),
                width=(dp(292) if is_mobile else dp(300)),
                height=(dp(168) if is_mobile else dp(178)),
                padding=[dp(12), dp(10), dp(12), dp(10)],
                pos_hint=({} if is_mobile else {"right": 0.985, "y": 0.02}),
            )
            self._style_popup_card(hud, palette.get("surface", (0.085, 0.105, 0.145, 0.96)), radius=dp(18))
            hud_header = BoxLayout(orientation="horizontal", spacing=dp(6), size_hint_y=None, height=dp(32))
            self.preview_hud_title_lbl = Label(
                text="Preview HUD",
                color=palette["text"],
                size_hint_y=None,
                height=(0 if is_mobile else dp(18)),
                halign="left",
                valign="middle",
                font_size=dp(11.5),
                bold=True,
            )
            self.preview_hud_title_lbl.bind(size=self._sync_label_text_size)
            hud_drag_hint = Label(
                text=("selected" if is_mobile else "drag"),
                color=palette["muted"],
                size_hint=(None, None),
                width=(dp(72) if is_mobile else dp(40)),
                height=dp(18),
                halign="right",
                valign="middle",
                font_size=dp(9.5),
            )
            hud_drag_hint.bind(size=self._sync_label_text_size)
            self.btn_hud_pin = self._make_compact_action_button("Hide", tone="accent")
            self.btn_hud_pin._expanded_label = "Hide"
            self.btn_hud_pin._collapsed_label = "Show"
            self.btn_hud_pin.size_hint = (None, None)
            self.btn_hud_pin.size = ((dp(58), dp(32)) if is_mobile else (dp(54), dp(26)))
            hud_header.add_widget(self.preview_hud_title_lbl)
            hud_header.add_widget(hud_drag_hint)
            hud_header.add_widget(self.btn_hud_pin)

            hud_body = BoxLayout(orientation="vertical", spacing=dp(5), size_hint_y=None)
            hud_body.bind(minimum_height=hud_body.setter("height"))
            self.preview_hud_detail_lbl = Label(
                text=("Tap or drag over a box to see its details here." if is_mobile else "Hover or tap a box to see its details here."),
                color=palette["text"],
                size_hint_y=None,
                height=dp(32),
                halign="left",
                valign="top",
                font_size=dp(11),
            )
            self.preview_hud_detail_lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(self.preview_hud_detail_lbl, min_height=dp(30), extra_pad=dp(6))
            self.preview_hud_mapping_lbl = Label(
                text="Mapping: —",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(36),
                halign="left",
                valign="top",
                font_size=dp(10.5),
            )
            self.preview_hud_mapping_lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(self.preview_hud_mapping_lbl, min_height=dp(32), extra_pad=dp(6))
            hud_btns = GridLayout(cols=3, spacing=dp(6), size_hint_y=None, height=row_h)
            self.btn_hud_map = make_button("Map", tone="primary")
            self.btn_hud_map.bind(on_release=self._open_mapping_from_hud)
            self.btn_hud_clear = make_button("Clear Field Links", tone="secondary")
            self.btn_hud_clear.bind(on_release=self.on_clear_selected_mapping)
            self.btn_hud_tune = make_button("Tune", tone="accent")
            self.btn_hud_tune.bind(on_release=self.open_detection_settings_popup)
            for w in [self.btn_hud_map, self.btn_hud_clear, self.btn_hud_tune]:
                hud_btns.add_widget(w)
            hud_body.add_widget(self.preview_hud_detail_lbl)
            hud_body.add_widget(self.preview_hud_mapping_lbl)
            hud_body.add_widget(hud_btns)
            hud.add_widget(hud_header)
            hud.add_widget(hud_body)
            hud.set_body_widget(hud_body, expanded_height=(dp(152) if is_mobile else dp(160)))
            hud.set_toggle_button(self.btn_hud_pin)
            hud.bind_collapse_state(lambda *_: self._refresh_mobile_hud_restore_button())
            hud.register_interactive_widgets(
                self.btn_hud_pin,
                self.btn_hud_map,
                self.btn_hud_clear,
                self.btn_hud_tune,
            )
            if is_mobile:
                hud.set_drag_enabled(False)
                hud.set_allow_body_passthrough(False)
                hud.bind(height=lambda *_: self._refresh_mobile_hud_restore_button())
            return hud

        root = BoxLayout(orientation="vertical", spacing=gap, padding=pad)

        appbar = BoxLayout(
            orientation="horizontal",
            spacing=dp(6) if is_mobile else dp(14),
            size_hint_y=None,
            height=dp(44) if is_mobile else dp(88),
            padding=[dp(8), dp(4), dp(8), dp(4)],
        )
        style_card(appbar, palette["surface"], radius=dp(18) if is_mobile else dp(26))

        brand_wrap = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None, size_hint_x=1)
        brand_wrap.bind(minimum_height=brand_wrap.setter("height"))
        hero_title = Label(
            text=(APP_TITLE if is_mobile else APP_TITLE),
            color=palette["text"],
            bold=True,
            size_hint_y=None,
            height=dp(24) if is_mobile else dp(30),
            halign="left",
            valign="middle",
            font_size=dp(14.5) if is_mobile else dp(20),
        )
        hero_title.bind(size=self._sync_label_text_size)
        if is_mobile:
            self._bind_auto_height_label(hero_title, min_height=dp(26), extra_pad=dp(2))
        brand_wrap.add_widget(hero_title)

        hero_sub = Label(
            text=("" if is_mobile else "Medical-tech PDF automation with guided preview, detection, mapping, and export."),
            color=palette["muted"],
            size_hint_y=None,
            height=(0 if is_mobile else dp(22)),
            halign="left",
            valign="middle",
            font_size=dp(10.5) if is_mobile else dp(11.5),
            opacity=(0 if is_mobile else 1),
        )
        hero_sub.bind(size=self._sync_label_text_size)
        if is_mobile:
            self._bind_auto_height_label(hero_sub, min_height=0, extra_pad=0)
        brand_wrap.add_widget(hero_sub)
        appbar.add_widget(brand_wrap)

        # "?" help button — opens the quick-reference popup
        _help_btn_appbar = self._make_compact_action_button("?", tone="ghost")
        _help_btn_appbar.size_hint = (None, None)
        _help_btn_appbar.size = (dp(44), dp(44))
        _help_btn_appbar.bind(on_release=self._show_help_popup)
        self._help_btn = _help_btn_appbar
        appbar.add_widget(_help_btn_appbar)

        self.mobile_meta_lbl = None
        self.android_status_file_lbl = None
        self.android_status_page_lbl = None
        self.android_status_boxes_lbl = None
        self.android_status_mode_lbl = None
        self.android_inspector_selected_lbl = None
        self.android_inspector_count_lbl = None
        self.android_inspector_box_lbl = None
        self.android_inspector_mapping_lbl = None
        self.android_inspector_resolution_lbl = None
        self.android_capture_summary_lbl = None
        self.android_capture_candidate_lbl = None
        self.android_capture_action_hint_lbl = None

        status_chip = Label(
            text=("Step 1: Tap Load PDF to open your form" if is_mobile else "Step 1: Open a PDF form  •  Step 2: Load Data  •  Step 3: Map & Export"),
            color=palette["text"],
            size_hint=(1, None) if is_mobile else (None, None),
            height=dp(44) if is_mobile else dp(48),
            halign="left" if is_mobile else "center",
            valign="middle",
            font_size=dp(12) if is_mobile else dp(11),
            padding=(dp(12), dp(0)),
        )
        status_chip.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(status_chip, min_height=dp(38), extra_pad=dp(10))
        style_card(status_chip, palette["chip"], radius=dp(18))
        self.status_lbl = status_chip

        phone_view_btn = None
        if platform != "android":
            phone_view_btn = make_button(
                "Desktop" if desktop_phone_emulation else "Phone",
                tone="soft" if not desktop_phone_emulation else "accent",
            )
            phone_view_btn.size_hint = (None, None)
            phone_view_btn.width = dp(104 if desktop_phone_emulation else 94)
            phone_view_btn.bind(on_release=_toggle_phone_view)

        if is_mobile:
            if phone_view_btn is not None:
                appbar.add_widget(phone_view_btn)
            root.add_widget(appbar)
        else:
            if phone_view_btn is not None:
                appbar.add_widget(phone_view_btn)
            status_chip.width = dp(420)
            appbar.add_widget(status_chip)
            root.add_widget(appbar)

        self.mobile_flow_lbl = None
        mobile_flow_card = None
        if is_mobile:
            mobile_flow_card = BoxLayout(
                orientation="vertical",
                spacing=dp(3),
                size_hint_y=None,
                height=dp(56),
                padding=[dp(8), dp(6), dp(8), dp(6)],
            )
            style_card(mobile_flow_card, palette["surface_alt"], radius=dp(16))

            flow_title = Label(
                text="Next",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(16),
                halign="left",
                valign="middle",
                font_size=dp(11),
            )
            flow_title.bind(size=self._sync_label_text_size)
            mobile_flow_card.add_widget(flow_title)

            self.mobile_flow_lbl = Label(
                text="Step 1: Tap '1 Open Files' and load your PDF form.",
                color=palette["text"],
                bold=True,
                size_hint_y=None,
                height=dp(22),
                halign="left",
                valign="middle",
                font_size=dp(11.5),
            )
            self.mobile_flow_lbl.bind(size=self._sync_label_text_size)
            mobile_flow_card.add_widget(self.mobile_flow_lbl)

        main = BoxLayout(orientation="vertical" if is_mobile else "horizontal", spacing=dp(10) if is_mobile else dp(16))

        self._mobile_section_buttons = {}
        self._mobile_section_cards = {}
        self._mobile_section_host = None
        self._mobile_active_section = None
        self._desktop_section_buttons = {}
        self._desktop_section_cards = {}
        self._desktop_section_host = None
        self._desktop_active_section = None

        def _style_mobile_section_button(btn, active=False):
            bg = palette["primary"] if active else palette["surface_soft"]
            fg = (1, 1, 1, 1) if active else palette["text"]
            if hasattr(btn, "tone"):
                try:
                    btn.tone = "primary" if active else "plain"
                except Exception:
                    pass
            if hasattr(btn, "md_bg_color"):
                try:
                    btn.md_bg_color = bg
                except Exception:
                    pass
                try:
                    btn.text_color = fg
                except Exception:
                    pass
            else:
                btn.background_normal = ""
                btn.background_down = ""
                btn.background_color = bg
                btn.color = fg

        def _show_mobile_section(section_key):
            if not is_mobile or self._mobile_section_host is None:
                return
            section_card = self._mobile_section_cards.get(section_key)
            if section_card is None:
                return
            self._mobile_section_host.clear_widgets()
            self._mobile_section_host.add_widget(section_card)
            self._mobile_active_section = section_key
            try:
                if platform == "android" and android_panel_card is not None and len(android_panel_card.children) >= 2:
                    _panel_title = android_panel_card.children[-1]
                    if hasattr(_panel_title, "text"):
                        _panel_title.text = f"Panel • {section_key}"
            except Exception:
                pass
            for key, btn in self._mobile_section_buttons.items():
                _style_mobile_section_button(btn, active=(key == section_key))
            try:
                if getattr(self, "android_status_mode_lbl", None) is not None:
                    self.android_status_mode_lbl.text = f"Mode: {section_key}"
            except Exception:
                pass
            try:
                self._refresh_android_capture_stage_ui()
            except Exception:
                pass
            try:
                self._refresh_android_selection_inspector()
            except Exception:
                pass
            try:
                self._refresh_android_capture_review_panel()
            except Exception:
                pass

        self._show_mobile_section = _show_mobile_section

        def _style_desktop_section_button(btn, active=False):
            bg = palette["primary"] if active else palette["surface_soft"]
            fg = (1, 1, 1, 1) if active else palette["text"]
            btn.background_normal = ""
            btn.background_down = ""
            btn.background_color = bg
            btn.color = fg

        def _show_desktop_section(section_key):
            if is_mobile or self._desktop_section_host is None:
                return
            section_card = self._desktop_section_cards.get(section_key)
            if section_card is None:
                return
            self._desktop_section_host.clear_widgets()
            self._desktop_section_host.add_widget(section_card)
            self._desktop_active_section = section_key
            for key, btn in self._desktop_section_buttons.items():
                _style_desktop_section_button(btn, active=(key == section_key))

        self._show_desktop_section = _show_desktop_section

        def _make_desktop_section_tabs(section_names):
            wrap = GridLayout(cols=len(section_names), spacing=dp(8), size_hint_y=None, height=row_h)
            for name in section_names:
                btn = Button(
                    text=name,
                    size_hint_y=None,
                    height=row_h,
                    background_normal="",
                    background_down="",
                    background_color=palette["surface_soft"],
                    color=palette["text"],
                    font_size=dp(12),
                )
                btn.bind(on_release=lambda inst, n=name: _show_desktop_section(n))
                self._desktop_section_buttons[name] = btn
                wrap.add_widget(btn)
            return wrap


        def _make_mobile_section_tabs(section_names):
            host = ScrollView(
                do_scroll_x=True,
                do_scroll_y=False,
                bar_width=0,
                scroll_type=["content"],
                size_hint_y=None,
                height=dp(44),
            )
            row = BoxLayout(orientation="horizontal", spacing=dp(8), size_hint=(None, None), height=dp(44))
            row.bind(minimum_width=row.setter("width"))
            display_map = {
                "Files": "1 Open Files",
                "Detection": "2 Find Fields",
                "Capture": "3 Draw Area",
                "Templates": "4 Templates",
                "Inspect": "5 Repair",
                "Mapping": "6 Map Fields",
                "Session": "7 Settings",
                "Export": "8 Export PDF",
                "Text": "Text",
                "Learning": "Memory",
            }
            for name in section_names:
                btn = self._make_compact_action_button(display_map.get(name, name), tone="ghost")
                btn.size_hint = (None, None)
                btn.size = (dp(90), dp(44))
                btn.bind(on_release=lambda inst, n=name: _show_mobile_section(n))
                self._mobile_section_buttons[name] = btn
                row.add_widget(btn)
            host.add_widget(row)
            return host

        controls_wrap = BoxLayout(orientation="vertical", size_hint=(1, 0.56) if is_mobile else (0.40, 1))
        controls_scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, bar_width=dp(6), scroll_type=["bars", "content"])
        controls = GridLayout(cols=1, spacing=gap, size_hint_y=None)
        controls.bind(minimum_height=controls.setter("height"))

        files_card, files_body = make_card("Workspace", "Open your PDF form and data file" if is_mobile else "Open your PDF form, data file (CSV/XLSX), and saved presets")
        file_grid = GridLayout(cols=1 if is_mobile else 3, spacing=dp(8), size_hint_y=None)
        file_buttons = [
            ("Load PDF", self.on_load_pdf, "primary"),
            ("Load Data File", self.on_load_csv, "soft"),
            ("Google Sheet", self.on_load_gsheet_url, "soft"),
            ("Open Saved Setup", self.on_load_config, "plain"),
            ("Merge Saved Setup", self.on_merge_config, "plain"),
            ("Save Setup", self.on_save_config, "accent"),
        ]
        rows = (len(file_buttons) + file_grid.cols - 1) // file_grid.cols
        file_grid.height = rows * row_h + max(rows-1,0) * dp(8)
        for txt, cb, tone in file_buttons:
            btn = make_button(txt, tone=tone)
            btn.bind(on_release=cb)
            file_grid.add_widget(btn)
        files_body.add_widget(file_grid)
        if is_mobile:
            self._mobile_section_cards["Files"] = files_card
        else:
            self._desktop_section_cards["Workspace"] = files_card

        nav_card, nav_body = make_card("Session", "Choose a record, page, and next action" if is_mobile else "Choose a record, move through pages, find fillable areas, and refresh the preview")
        self.patient_spinner = make_spinner(RECORD_SELECT_TEXT, picker_title="Choose a patient / record", search_hint="Type to search for a patient or record")
        self.patient_spinner.bind(text=self.on_patient_change)
        nav_body.add_widget(labeled_field("Record", self.patient_spinner, "Tap the field, type to filter the record list, then press Enter or choose the row to preview or export"))
        nav_grid = GridLayout(cols=3, spacing=dp(8), size_hint_y=None, height=row_h)
        self.page_input_main = make_input("0", "Page", input_filter="int")
        self.btn_prev = make_button("Prev", tone="plain")
        self.btn_next = make_button("Next", tone="plain")
        self.btn_prev.bind(on_release=self.on_prev_page)
        self.btn_next.bind(on_release=self.on_next_page)
        if not is_mobile:
            self.page_input = self.page_input_main
        self.page_input_main.bind(text=self._mirror_page_inputs)
        nav_grid.add_widget(self.page_input_main)
        nav_grid.add_widget(self.btn_prev)
        nav_grid.add_widget(self.btn_next)
        nav_body.add_widget(labeled_field("Page Number", nav_grid, "Page 0 is the first page"))
        action_grid = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=row_h)
        self.btn_detect = make_button("Find Fields", tone="accent")
        self.btn_detect.bind(on_release=self.on_run_detect)
        self.btn_preview = make_button("Refresh Form Preview", tone="primary")
        self.btn_preview.bind(on_release=self.on_preview)
        action_grid.add_widget(self.btn_detect)
        action_grid.add_widget(self.btn_preview)
        nav_body.add_widget(action_grid)
        self.backend_note_lbl = Label(
            text="",
            color=palette["muted"],
            size_hint_y=None,
            height=dp(40),
            halign="left",
            valign="middle",
            font_size=dp(11),
        )
        self.backend_note_lbl.bind(size=self._sync_label_text_size)
        nav_body.add_widget(self.backend_note_lbl)
        if is_mobile:
            self._mobile_section_cards["Session"] = nav_card
        else:
            self._desktop_section_cards["Session"] = nav_card

        detect_card, detect_body = make_card("Detection", "Find fillable areas on the form" if is_mobile else "Tune how the app finds fill areas, answer lines, and checkboxes")
        explain = Label(text="These controls decide how the app finds fill areas, answer lines, and checkboxes.",
                        color=palette["muted"], size_hint_y=None, height=dp(18), halign="left", valign="middle", font_size=dp(10.5 if is_mobile else 11))
        explain.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(explain, min_height=dp(18), extra_pad=dp(4))
        detect_body.add_widget(explain)

        self.f_area = make_input(DEFAULTS["F_Area"], "500", input_filter="int")
        self.f_minw = make_input(DEFAULTS["F_MinW"], "15", input_filter="int")
        self.f_minh = make_input(DEFAULTS["F_MinH"], "35", input_filter="int")
        self.f_close = make_input(DEFAULTS["F_Close"], "1", input_filter="int")
        self.line_minw = make_input(DEFAULTS["Line_MinW"], "100", input_filter="int")
        self.line_maxw = make_input(DEFAULTS["Line_MaxW"], "1680", input_filter="int")
        self.c_strict = make_input(DEFAULTS["C_Strict"], "40", input_filter="int")
        self.c_size_min = make_input(DEFAULTS["C_Size"][0], "14", input_filter="int")
        self.c_size_max = make_input(DEFAULTS["C_Size"][1], "65", input_filter="int")
        self.c_border = make_input(DEFAULTS["C_Border"], "0.10")
        self.c_inner = make_input(DEFAULTS["C_Inner"], "0.40")
        self.roi_max = make_input(DEFAULTS["ROI_Max"], "200", input_filter="int")
        self.c_open = make_input(DEFAULTS["C_Open"], "1", input_filter="int")
        self.c_close = make_input(DEFAULTS["C_Close"], "0", input_filter="int")
        self.c_band = make_input(DEFAULTS["C_BandPct"], "0.18")
        self.c_aspect = make_input(DEFAULTS["C_AspectTol"], "0.10")
        self.ext_low = make_input(DEFAULTS["Ext_Low"], "0.04")
        self.ext_high = make_input(DEFAULTS["Ext_High"], "0.60")
        self.c_fill = make_input(DEFAULTS["C_FillMin"], "0.45")
        self.c_eps = make_input(DEFAULTS["C_Eps"], "0.04")
        self.use_extent_chk = CheckBox(active=bool(DEFAULTS["Use_Extent"]))

        detection_fields = [
            labeled_field(detection_ui_label("f_area", "Field Area"), self.f_area, detection_ui_helper("f_area", "Minimum contour area kept as a field")),
            labeled_field(detection_ui_label("f_minw", "Field Min Width"), self.f_minw, detection_ui_helper("f_minw")),
            labeled_field(detection_ui_label("f_minh", "Field Min Height"), self.f_minh, detection_ui_helper("f_minh")),
            labeled_field(detection_ui_label("f_close", "Field Close"), self.f_close, detection_ui_helper("f_close", "Morph close kernel for field cleanup")),
            labeled_field(detection_ui_label("line_minw", "Line Min Width"), self.line_minw, detection_ui_helper("line_minw")),
            labeled_field(detection_ui_label("line_maxw", "Line Max Width"), self.line_maxw, detection_ui_helper("line_maxw")),
            labeled_field(detection_ui_label("c_strict", "Checkbox Strict"), self.c_strict, detection_ui_helper("c_strict", "Higher values tighten checkbox rules")),
            labeled_field(detection_ui_label("c_size_min", "Checkbox Min Size"), self.c_size_min, detection_ui_helper("c_size_min")),
            labeled_field(detection_ui_label("c_size_max", "Checkbox Max Size"), self.c_size_max, detection_ui_helper("c_size_max")),
            labeled_field(detection_ui_label("c_border", "Border Fill"), self.c_border, detection_ui_helper("c_border")),
            labeled_field(detection_ui_label("c_inner", "Inner Fill"), self.c_inner, detection_ui_helper("c_inner")),
            labeled_field(detection_ui_label("roi_max", "ROI Max"), self.roi_max, detection_ui_helper("roi_max")),
            labeled_field(detection_ui_label("c_open", "Checkbox Open"), self.c_open, detection_ui_helper("c_open")),
            labeled_field(detection_ui_label("c_close", "Checkbox Close"), self.c_close, detection_ui_helper("c_close")),
            labeled_field(detection_ui_label("c_band", "Band %"), self.c_band, detection_ui_helper("c_band")),
            labeled_field(detection_ui_label("c_aspect", "Aspect Tolerance"), self.c_aspect, detection_ui_helper("c_aspect")),
            labeled_field(detection_ui_label("ext_low", "Extent Low"), self.ext_low, detection_ui_helper("ext_low")),
            labeled_field(detection_ui_label("ext_high", "Extent High"), self.ext_high, detection_ui_helper("ext_high")),
            labeled_field(detection_ui_label("c_fill", "Fill Min"), self.c_fill, detection_ui_helper("c_fill")),
            labeled_field(detection_ui_label("c_eps", "Approx Eps"), self.c_eps, detection_ui_helper("c_eps")),
            labeled_checkbox(detection_ui_label("use_extent", "Use Extent"), self.use_extent_chk, detection_ui_helper("use_extent", "Use contour extent as an extra checkbox filter")),
        ]
        det_grid = GridLayout(cols=1 if is_mobile else 3, spacing=dp(8), size_hint_y=None)
        det_grid.bind(minimum_height=det_grid.setter("height"))
        for w in detection_fields:
            det_grid.add_widget(w)
        detect_body.add_widget(det_grid)
        det_action_row = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=row_h)
        self.btn_detection_dialog = make_button("Easy Tuning Sliders", tone="soft")
        self.btn_detection_dialog.bind(on_release=self.open_detection_settings_popup)
        self.btn_detection_run = make_button("Find Fields Now", tone="accent")
        self.btn_detection_run.bind(on_release=self.on_run_detect)
        det_action_row.add_widget(self.btn_detection_dialog)
        det_action_row.add_widget(self.btn_detection_run)
        detect_body.add_widget(det_action_row)
        if is_mobile:
            self._mobile_section_cards["Detection"] = detect_card
        else:
            self._desktop_section_cards["Detection"] = detect_card

        text_card, text_body = make_card("Text Rendering", "Adjust font fit, centering, and readability inside the mapped boxes")
        text_note = Label(text="These controls change how mapped text is fitted and centered inside the selected fields and exported PDF.",
                        color=palette["muted"], size_hint_y=None, height=dp(22), halign="left", valign="middle", font_size=dp(10.5 if is_mobile else 11))
        text_note.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(text_note, min_height=dp(20), extra_pad=dp(4))
        text_body.add_widget(text_note)
        self.text_tuning_summary_lbl = Label(text="", color=palette["text"], size_hint_y=None, height=dp(42), halign="left", valign="middle", font_size=dp(11))
        self.text_tuning_summary_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(self.text_tuning_summary_lbl, min_height=dp(38), extra_pad=dp(4))
        text_body.add_widget(labeled_field("Current text fit profile", self.text_tuning_summary_lbl, "Preview and export share these font-fitting rules."))
        text_action_row = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=row_h)
        self.btn_text_tuning_dialog = make_button("Font Tuning Console", tone="soft")
        self.btn_text_tuning_dialog.bind(on_release=self.open_text_tuning_popup)
        self.btn_text_preview_refresh = make_button("Preview Text Now", tone="primary")
        self.btn_text_preview_refresh.bind(on_release=self.on_preview)
        text_action_row.add_widget(self.btn_text_tuning_dialog)
        text_action_row.add_widget(self.btn_text_preview_refresh)
        text_body.add_widget(text_action_row)
        if is_mobile:
            self._mobile_section_cards["Text"] = text_card
        else:
            self._desktop_section_cards["Text"] = text_card
        self._update_text_tuning_summary()

        map_card, map_body = make_card("Mapping", "Link detected fields to your data columns" if is_mobile else "Link detected fields to your data columns directly from the preview")
        self.box_ids_input = make_input("", "0,1,2")
        self.column_spinner = make_spinner(COLUMN_SELECT_TEXT, picker_title="Choose which column fills this field", search_hint="Type to find a column name")
        self.trigger_input = make_input("", "Trigger")
        self.grid_flag_chk = CheckBox(active=False)
        self.grid_n_input = make_input("1", "1", input_filter="int")
        map_body.add_widget(labeled_field("Selected Boxes", self.box_ids_input, "Tap boxes in the preview to add or remove them" if is_mobile else "Click boxes in the preview to add or remove them"))
        map_body.add_widget(labeled_field("Data Column", self.column_spinner, "Tap the field, type to filter the column list, then press Enter or choose the column to fill the selected box or boxes"))
        map_body.add_widget(labeled_field("Checkbox Trigger Value", self.trigger_input, "Optional: only mark the checkbox when the selected value matches this text" if is_mobile else "Leave blank for normal text filling. Use a value only when this mapping should place an X in a checkbox."))
        map_opts = GridLayout(cols=1 if is_mobile else 2, spacing=dp(8), size_hint_y=None)
        map_opts.bind(minimum_height=map_opts.setter("height"))
        map_opts.add_widget(labeled_checkbox("Split Across Boxes", self.grid_flag_chk, "Spread characters across separate boxes or cells"))
        map_opts.add_widget(labeled_field("Number of Boxes", self.grid_n_input, "How many boxes or cells should receive the split text"))
        map_body.add_widget(map_opts)
        self.btn_assign = make_button("Save Field Link", tone="primary")
        self.btn_assign.height = dp(42) if is_mobile else self.btn_assign.height
        self.btn_assign.bind(on_release=self.on_assign_mapping)
        map_body.add_widget(self.btn_assign)
        if is_mobile:
            self._mobile_section_cards["Mapping"] = map_card
        else:
            self._desktop_section_cards["Mapping"] = map_card

        export_card, export_body = make_card("Export", "Generate filled PDFs from your data" if is_mobile else "Generate filled PDF files for one record or for all records at once")
        out_grid = GridLayout(cols=1 if is_mobile else 2, spacing=dp(8), size_hint_y=None, height=(2*row_h+dp(8)) if is_mobile else row_h)
        self.btn_generate_one = make_button("Export This Record", tone="accent")
        self.btn_generate_one.bind(on_release=self.on_generate_single)
        self.btn_generate_batch = make_button("Export All Records", tone="plain")
        self.btn_generate_batch.bind(on_release=self.on_generate_batch)
        out_grid.add_widget(self.btn_generate_one)
        out_grid.add_widget(self.btn_generate_batch)
        export_body.add_widget(out_grid)
        self.export_note_lbl = Label(
            text="",
            color=palette["muted"],
            size_hint_y=None,
            height=dp(42),
            halign="left",
            valign="middle",
            font_size=dp(11),
        )
        self.export_note_lbl.bind(size=self._sync_label_text_size)
        export_body.add_widget(self.export_note_lbl)
        if is_mobile:
            self._mobile_section_cards["Export"] = export_card
        else:
            self._desktop_section_cards["Export"] = export_card

        if is_mobile and platform == "android":
            def _make_android_panel_card(title, subtitle=""):
                card, body = make_card(title, subtitle)
                return card, body

            # Android-specific staged panels: these are lighter action-first cards
            # that sit in the visible Android panel host instead of the older dense
            # mobile sidebar cards.
            android_files_card, android_files_body = _make_android_panel_card(
                "1 Open Files",
                "Open your PDF form and data file to get started."
            )
            files_quick = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(192))
            btn_android_load_pdf = self._make_compact_action_button("Load PDF", tone="primary")
            btn_android_load_pdf.bind(on_release=self.on_load_pdf)
            btn_android_load_data = self._make_compact_action_button("Load Data", tone="plain")
            btn_android_load_data.bind(on_release=self.on_load_csv)
            btn_android_load_sheet = self._make_compact_action_button("Google Sheet", tone="plain")
            btn_android_load_sheet.bind(on_release=self.on_load_gsheet_url)
            btn_android_open_setup = self._make_compact_action_button("Open Setup", tone="ghost")
            btn_android_open_setup.bind(on_release=self.on_load_config)
            btn_android_merge_setup = self._make_compact_action_button("Merge Setup", tone="ghost")
            btn_android_merge_setup.bind(on_release=self.on_merge_config)
            btn_android_save_setup = self._make_compact_action_button("Save Setup", tone="ghost")
            btn_android_save_setup.bind(on_release=self.on_save_config)
            btn_android_templates = self._make_compact_action_button("Templates", tone="plain")
            btn_android_templates.bind(on_release=self._open_area_template_library_popup)
            btn_android_detect_jump = self._make_compact_action_button("Find Fields", tone="accent")
            btn_android_detect_jump.bind(on_release=lambda *_: self._show_mobile_section("Detection"))
            for _w in (
                btn_android_load_pdf,
                btn_android_load_data,
                btn_android_load_sheet,
                btn_android_open_setup,
                btn_android_merge_setup,
                btn_android_save_setup,
                btn_android_templates,
                btn_android_detect_jump,
            ):
                _w.size_hint = (1, None)
                _w.height = dp(42)
                files_quick.add_widget(_w)
            android_files_body.add_widget(files_quick)

            android_session_card, android_session_body = _make_android_panel_card(
                "7 Settings",
                "Navigate pages and refresh the form preview."
            )
            session_grid = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(92))
            btn_android_prev = self._make_compact_action_button("Prev Page", tone="plain")
            btn_android_prev.bind(on_release=self.on_prev_page)
            btn_android_next = self._make_compact_action_button("Next Page", tone="plain")
            btn_android_next.bind(on_release=self.on_next_page)
            btn_android_find = self._make_compact_action_button("Find", tone="accent")
            btn_android_find.bind(on_release=self.on_run_detect)
            btn_android_preview = self._make_compact_action_button("Refresh Preview", tone="ghost")
            btn_android_preview.bind(on_release=self.on_preview)
            for _w in (btn_android_prev, btn_android_next, btn_android_find, btn_android_preview):
                _w.size_hint = (1, None)
                _w.height = dp(42)
                session_grid.add_widget(_w)
            android_session_body.add_widget(session_grid)

            android_detection_card, android_detection_body = _make_android_panel_card(
                "2 Find Fields",
                "Find fillable areas on the form and adjust detection settings."
            )
            det_actions = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(92))
            btn_android_detect = self._make_compact_action_button("Find Fields", tone="accent")
            btn_android_detect.bind(on_release=self.on_run_detect)
            btn_android_tuning = self._make_compact_action_button("Open Tuning", tone="plain")
            btn_android_tuning.bind(on_release=self.open_detection_settings_popup)
            btn_android_capture_jump = self._make_compact_action_button("Capture", tone="accent")
            btn_android_capture_jump.bind(on_release=self._toggle_manual_area_capture_mode)
            btn_android_repair_jump = self._make_compact_action_button("Repair", tone="secondary")
            btn_android_repair_jump.bind(on_release=self._toggle_mobile_inspect_panel)
            for _w in (btn_android_detect, btn_android_tuning, btn_android_capture_jump, btn_android_repair_jump):
                _w.size_hint = (1, None)
                _w.height = dp(42)
                det_actions.add_widget(_w)
            android_detection_body.add_widget(det_actions)

            # NOTE: do not reuse detection TextInput widgets here because they are
            # already mounted in the main settings layout; Kivy widgets can have
            # only one parent at a time.
            det_compact_note = Label(
                text="Use [b]Open Tuning[/b] to edit detection thresholds.",
                markup=True,
                color=palette["muted"],
                size_hint_y=None,
                height=dp(28),
                halign="left",
                valign="middle",
                font_size=dp(12),
            )
            det_compact_note.bind(size=self._sync_label_text_size)
            android_detection_body.add_widget(det_compact_note)

            android_mapping_card, android_mapping_body = _make_android_panel_card(
                "6 Map Fields",
                "Link detected fields to your data columns."
            )
            map_action_grid = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(92))
            btn_android_map_save = self._make_compact_action_button("Save Field Link", tone="primary")
            btn_android_map_save.bind(on_release=self.on_assign_mapping)
            btn_android_map_clear = self._make_compact_action_button("Clear Field Links", tone="ghost")
            btn_android_map_clear.bind(on_release=self.on_clear_selected_mapping)
            btn_android_select_toggle = self._make_compact_action_button("Select Mode", tone="plain")
            btn_android_select_toggle.bind(on_release=self._toggle_mobile_selection_mode)
            btn_android_map_templates = self._make_compact_action_button("Templates", tone="plain")
            btn_android_map_templates.bind(on_release=self._open_area_template_library_popup)
            for _w in (btn_android_map_save, btn_android_map_clear, btn_android_select_toggle, btn_android_map_templates):
                _w.size_hint = (1, None)
                _w.height = dp(44)
                map_action_grid.add_widget(_w)
            android_mapping_body.add_widget(map_action_grid)
            # NOTE: these selectors are already mounted in the primary mapping UI.
            # Re-adding them here causes Kivy parent conflicts.
            map_compact_note = Label(
                text="Choose [b]Record[/b] and [b]Column[/b] in the main Mapping panel.",
                markup=True,
                color=palette["muted"],
                size_hint_y=None,
                height=dp(28),
                halign="left",
                valign="middle",
                font_size=dp(12),
            )
            map_compact_note.bind(size=self._sync_label_text_size)
            android_mapping_body.add_widget(map_compact_note)

            android_export_card, android_export_body = _make_android_panel_card(
                "8 Export PDF",
                "Generate filled PDFs from your data."
            )
            export_action_grid = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(92))
            btn_android_export_one = self._make_compact_action_button("Export This Record", tone="accent")
            btn_android_export_one.bind(on_release=self.on_generate_single)
            btn_android_export_batch = self._make_compact_action_button("Export All Records", tone="plain")
            btn_android_export_batch.bind(on_release=self.on_generate_batch)
            btn_android_save_cfg = self._make_compact_action_button("Save Config", tone="ghost")
            btn_android_save_cfg.bind(on_release=self.on_save_config)
            btn_android_load_cfg = self._make_compact_action_button("Load Config", tone="ghost")
            btn_android_load_cfg.bind(on_release=self.on_load_config)
            for _w in (btn_android_export_one, btn_android_export_batch, btn_android_save_cfg, btn_android_load_cfg):
                _w.size_hint = (1, None)
                _w.height = dp(44)
                export_action_grid.add_widget(_w)
            android_export_body.add_widget(export_action_grid)
            self.export_note_lbl_android = Label(
                text="",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(42),
                halign="left",
                valign="middle",
                font_size=dp(11),
            )
            self.export_note_lbl_android.bind(size=self._sync_label_text_size)
            android_export_body.add_widget(self.export_note_lbl_android)

            self._mobile_section_cards["Files"] = android_files_card
            # Keep the full Session card on Android so record selection/picker
            # (including imported Google Sheet rows) remains available.
            self._mobile_section_cards["Session"] = nav_card
            self._mobile_section_cards["Detection"] = android_detection_card
            self._mobile_section_cards["Mapping"] = android_mapping_card
            self._mobile_section_cards["Export"] = android_export_card

            android_capture_card, android_capture_body = _make_android_panel_card(
                "3 Draw Area",
                "Draw a region on the form to add or remove detected fields."
            )
            capture_stage_lbl = Label(
                text="Stage: Idle",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(18),
                halign="left",
                valign="middle",
                font_size=dp(10.5),
            )
            capture_stage_lbl.bind(size=self._sync_label_text_size)
            self.android_capture_stage_lbl = capture_stage_lbl
            android_capture_body.add_widget(capture_stage_lbl)

            capture_action_grid = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(92))
            btn_android_capture_start = self._make_compact_action_button("Start Capture", tone="accent")
            btn_android_capture_start.bind(on_release=self._toggle_manual_area_capture_mode)
            btn_android_capture_cancel = self._make_compact_action_button("Cancel Capture", tone="ghost")
            btn_android_capture_cancel.bind(on_release=lambda *_: self._clear_manual_capture_preview(preserve_mode=False))
            btn_android_capture_templates = self._make_compact_action_button("Saved Templates", tone="plain")
            btn_android_capture_templates.bind(on_release=self._open_area_template_library_popup)
            btn_android_capture_tuning = self._make_compact_action_button("Open Tuning", tone="plain")
            btn_android_capture_tuning.bind(on_release=self.open_detection_settings_popup)
            for _w in (btn_android_capture_start, btn_android_capture_cancel, btn_android_capture_templates, btn_android_capture_tuning):
                _w.size_hint = (1, None)
                _w.height = dp(42)
                capture_action_grid.add_widget(_w)
            android_capture_body.add_widget(capture_action_grid)

            capture_help = Label(
                text="1. Start Capture\n2. Tap top-left and bottom-right on the preview\n3. Tap an outline or use Prev/Next below\n4. Use Area / Delete / Find Similar from this panel",
                color=palette["text"],
                size_hint_y=None,
                height=dp(72),
                halign="left",
                valign="middle",
                font_size=dp(10.5),
            )
            capture_help.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(capture_help, min_height=dp(56), extra_pad=dp(6))
            android_capture_body.add_widget(capture_help)

            self.android_capture_candidate_lbl = Label(
                text="Candidate: —",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(18),
                halign="left",
                valign="middle",
                font_size=dp(10.5),
            )
            self.android_capture_candidate_lbl.bind(size=self._sync_label_text_size)
            android_capture_body.add_widget(self.android_capture_candidate_lbl)

            self.android_capture_summary_lbl = Label(
                text="No captured area yet.",
                color=palette["text"],
                size_hint_y=None,
                height=dp(56),
                halign="left",
                valign="middle",
                font_size=dp(10.5),
            )
            self.android_capture_summary_lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(self.android_capture_summary_lbl, min_height=dp(44), extra_pad=dp(6))
            android_capture_body.add_widget(self.android_capture_summary_lbl)

            candidate_switch_row = GridLayout(cols=3, spacing=dp(8), size_hint_y=None, height=dp(42))
            btn_android_capture_prev = self._make_compact_action_button("Prev Trace", tone="plain")
            btn_android_capture_prev.bind(on_release=lambda *_: self._select_capture_shape_candidate(-1))
            btn_android_capture_next = self._make_compact_action_button("Next Trace", tone="plain")
            btn_android_capture_next.bind(on_release=lambda *_: self._select_capture_shape_candidate(1))
            btn_android_capture_popup = self._make_compact_action_button("Open Popup", tone="ghost")
            btn_android_capture_popup.bind(on_release=lambda *_: self._open_manual_area_capture_confirm_popup())
            for _w in (btn_android_capture_prev, btn_android_capture_next, btn_android_capture_popup):
                _w.size_hint = (1, None)
                _w.height = dp(38)
                candidate_switch_row.add_widget(_w)
            android_capture_body.add_widget(candidate_switch_row)

            self.android_capture_action_hint_lbl = Label(
                text="Review stays here on Android. Popup is optional.",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(18),
                halign="left",
                valign="middle",
                font_size=dp(10),
            )
            self.android_capture_action_hint_lbl.bind(size=self._sync_label_text_size)
            android_capture_body.add_widget(self.android_capture_action_hint_lbl)

            capture_quick_row = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(92))
            btn_android_capture_find_page = self._make_compact_action_button("Find Similar Page", tone="plain")
            btn_android_capture_find_page.bind(on_release=self._find_similar_for_active_capture_template)
            btn_android_capture_find_form = self._make_compact_action_button("Find Similar Form", tone="plain")
            btn_android_capture_find_form.bind(on_release=self._find_similar_in_form_for_active_capture_template)
            btn_android_capture_apply = self._make_compact_action_button("Use Area", tone="primary")
            btn_android_capture_apply.bind(on_release=self._apply_confirmed_manual_area_capture)
            btn_android_capture_delete = self._make_compact_action_button("Delete In Area", tone="danger")
            if hasattr(self, '_delete_detected_items_for_active_capture'):
                btn_android_capture_delete.bind(on_release=lambda *_: self._delete_detected_items_for_active_capture(persist_as_repair_patch=True, include_mappings=True))
            for _w in (btn_android_capture_find_page, btn_android_capture_find_form, btn_android_capture_apply, btn_android_capture_delete):
                _w.size_hint = (1, None)
                _w.height = dp(42)
                capture_quick_row.add_widget(_w)
            android_capture_body.add_widget(capture_quick_row)

            self._mobile_section_cards["Capture"] = android_capture_card

            android_templates_card, android_templates_body = _make_android_panel_card(
                "4 Templates",
                "Reuse saved field layouts from previous forms."
            )

            templates_stage_lbl = Label(
                text="Saved Area Templates are reusable on Detect and Find Similar.",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(22),
                halign="left",
                valign="middle",
                font_size=dp(10.5),
            )
            templates_stage_lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(templates_stage_lbl, min_height=dp(20), extra_pad=dp(4))
            android_templates_body.add_widget(templates_stage_lbl)

            templates_action_grid = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(92))
            btn_android_templates_open = self._make_compact_action_button("Open Library", tone="accent")
            btn_android_templates_open.bind(on_release=self._open_area_template_library_popup)
            btn_android_templates_capture = self._make_compact_action_button("Go to Draw Area", tone="plain")
            btn_android_templates_capture.bind(on_release=lambda *_: self._show_mobile_section("Capture"))
            btn_android_templates_find_page = self._make_compact_action_button("Find Saved Page", tone="plain")
            btn_android_templates_find_page.bind(on_release=self._open_area_template_library_popup)
            btn_android_templates_find_form = self._make_compact_action_button("Find Saved Form", tone="plain")
            btn_android_templates_find_form.bind(on_release=self._open_area_template_library_popup)
            for _w in (
                btn_android_templates_open,
                btn_android_templates_capture,
                btn_android_templates_find_page,
                btn_android_templates_find_form,
            ):
                _w.size_hint = (1, None)
                _w.height = dp(42)
                templates_action_grid.add_widget(_w)
            android_templates_body.add_widget(templates_action_grid)

            templates_help = Label(
                text="Use Open Library to apply, find, edit, or delete saved templates.\nTemplates are the durable layer reused by Detect and Find Similar.",
                color=palette["text"],
                size_hint_y=None,
                height=dp(52),
                halign="left",
                valign="middle",
                font_size=dp(10.5),
            )
            templates_help.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(templates_help, min_height=dp(42), extra_pad=dp(6))
            android_templates_body.add_widget(templates_help)

            templates_manage_grid = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(92))
            btn_android_templates_detect = self._make_compact_action_button("Find Fields", tone="primary")
            btn_android_templates_detect.bind(on_release=self.on_run_detect)
            btn_android_templates_refresh = self._make_compact_action_button("Refresh Preview", tone="ghost")
            btn_android_templates_refresh.bind(on_release=self.on_preview)
            btn_android_templates_export = self._make_compact_action_button("Go to Export PDF", tone="plain")
            btn_android_templates_export.bind(on_release=lambda *_: self._show_mobile_section("Export"))
            btn_android_templates_mapping = self._make_compact_action_button("Go to Map Fields", tone="plain")
            btn_android_templates_mapping.bind(on_release=lambda *_: self._show_mobile_section("Mapping"))
            for _w in (
                btn_android_templates_detect,
                btn_android_templates_refresh,
                btn_android_templates_export,
                btn_android_templates_mapping,
            ):
                _w.size_hint = (1, None)
                _w.height = dp(42)
                templates_manage_grid.add_widget(_w)
            android_templates_body.add_widget(templates_manage_grid)

            self._mobile_section_cards["Templates"] = android_templates_card

            android_inspect_card, android_inspect_body = _make_android_panel_card(
                "5 Repair",
                "Inspect selected fields and fix incorrect detections."
            )

            inspect_summary = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
            inspect_summary.bind(minimum_height=inspect_summary.setter("height"))

            self.android_inspector_selected_lbl = Label(
                text="None",
                color=palette["text"],
                size_hint_y=None,
                height=dp(20),
                halign="left",
                valign="middle",
                font_size=dp(11),
            )
            self.android_inspector_selected_lbl.bind(size=self._sync_label_text_size)
            inspect_summary.add_widget(labeled_field("Selected IDs", self.android_inspector_selected_lbl))

            self.android_inspector_count_lbl = Label(
                text="0",
                color=palette["text"],
                size_hint_y=None,
                height=dp(20),
                halign="left",
                valign="middle",
                font_size=dp(11),
            )
            self.android_inspector_count_lbl.bind(size=self._sync_label_text_size)
            inspect_summary.add_widget(labeled_field("Selected Count", self.android_inspector_count_lbl))

            self.android_inspector_box_lbl = Label(
                text="No selection",
                color=palette["text"],
                size_hint_y=None,
                height=dp(42),
                halign="left",
                valign="middle",
                font_size=dp(11),
            )
            self.android_inspector_box_lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(self.android_inspector_box_lbl, min_height=dp(36), extra_pad=dp(6))
            inspect_summary.add_widget(labeled_field("Primary Box", self.android_inspector_box_lbl))

            self.android_inspector_mapping_lbl = Label(
                text="Choose a box in the preview to inspect it.",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(42),
                halign="left",
                valign="middle",
                font_size=dp(10.5),
            )
            self.android_inspector_mapping_lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(self.android_inspector_mapping_lbl, min_height=dp(36), extra_pad=dp(6))
            inspect_summary.add_widget(labeled_field("Current Link", self.android_inspector_mapping_lbl))

            self.android_inspector_resolution_lbl = Label(
                text="Resolved preview value: —",
                color=palette["text"],
                size_hint_y=None,
                height=dp(42),
                halign="left",
                valign="middle",
                font_size=dp(10.5),
            )
            self.android_inspector_resolution_lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(self.android_inspector_resolution_lbl, min_height=dp(36), extra_pad=dp(6))
            inspect_summary.add_widget(labeled_field("Resolved Value", self.android_inspector_resolution_lbl))

            android_inspect_body.add_widget(inspect_summary)

            inspect_actions = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(92))
            btn_android_inspect_select = self._make_compact_action_button("Select Mode", tone="plain")
            btn_android_inspect_select.bind(on_release=self._toggle_mobile_selection_mode)
            btn_android_inspect_clear = self._make_compact_action_button("Clear Field Links", tone="ghost")
            btn_android_inspect_clear.bind(on_release=self.on_clear_selected_mapping)
            btn_android_inspect_map = self._make_compact_action_button("Go to Map Fields", tone="primary")
            btn_android_inspect_map.bind(on_release=lambda *_: self._show_mobile_section("Mapping"))
            btn_android_inspect_repair = self._make_compact_action_button("Repair Tools", tone="secondary")
            btn_android_inspect_repair.bind(on_release=self._toggle_mobile_inspect_panel)
            for _w in (btn_android_inspect_select, btn_android_inspect_clear, btn_android_inspect_map, btn_android_inspect_repair):
                _w.size_hint = (1, None)
                _w.height = dp(42)
                inspect_actions.add_widget(_w)
            android_inspect_body.add_widget(inspect_actions)

            self._mobile_section_cards["Inspect"] = android_inspect_card

        learning_card, learning_body = make_card("Smart Memory", "Review saved profiles, revisions, and auto-adjustment history")
        learning_note = Label(
            text="The app can remember good settings for similar forms. Use this panel to review matches, save a good version, and reuse the best learned setup.",
            color=palette["muted"],
            size_hint_y=None,
            height=dp(34),
            halign="left",
            valign="middle",
            font_size=dp(11),
        )
        learning_note.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(learning_note, min_height=dp(30), extra_pad=dp(4))
        learning_body.add_widget(learning_note)

        def _make_learning_info_label(initial_text, color=None, min_height=dp(32)):
            lbl = Label(
                text=initial_text,
                color=color or palette["text"],
                size_hint_y=None,
                height=min_height,
                halign="left",
                valign="middle",
                font_size=dp(11),
            )
            lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(lbl, min_height=min_height, extra_pad=dp(4))
            return lbl

        self.learning_summary_lbl = _make_learning_info_label("Profiles: 0 • Revisions: 0 • Corrections: 0")
        self.learning_match_lbl = _make_learning_info_label("Match: none yet")
        self.learning_revision_lbl = _make_learning_info_label("Current revision: none")
        self.learning_recent_lbl = _make_learning_info_label("Recent revisions: none", min_height=dp(48))
        self.learning_storage_lbl = _make_learning_info_label("Storage: —", color=palette["muted"], min_height=dp(40))

        learning_body.add_widget(labeled_field("Summary", self.learning_summary_lbl, "Counts and current page learning state"))
        learning_body.add_widget(labeled_field("Best Match", self.learning_match_lbl, "The closest saved setup the app found for this page"))
        learning_body.add_widget(labeled_field("Saved Versions", self.learning_revision_lbl, "The current saved or approved setup version"))
        learning_body.add_widget(labeled_field("Recent History", self.learning_recent_lbl, "Recently saved setup versions for quick review"))
        learning_body.add_widget(labeled_field("Storage", self.learning_storage_lbl, "Where the app keeps its saved memory files"))

        learn_grid = GridLayout(cols=2 if is_mobile else 3, spacing=dp(8), size_hint_y=None)
        learning_buttons = [
            ("Refresh State", self.on_learning_refresh, "soft"),
            ("Save Revision", self.on_learning_save_revision, "accent"),
            ("Approve Current", self.on_learning_approve_current, "primary"),
            ("Auto-Apply Best Profile", self.on_learning_apply_best_profile, "plain"),
            ("View Current", self.on_learning_view_current_revision, "plain"),
            ("Browse Revisions", self.on_learning_browse_revisions, "plain"),
            ("Reload Memory", self.on_learning_reload_memory, "plain"),
        ]
        learn_rows = (len(learning_buttons) + learn_grid.cols - 1) // learn_grid.cols
        learn_grid.height = learn_rows * row_h + max(learn_rows - 1, 0) * dp(8)
        for txt, cb, tone in learning_buttons:
            btn = make_button(txt, tone=tone)
            btn.bind(on_release=cb)
            if txt == "Refresh State":
                self.btn_learning_refresh = btn
            elif txt == "Save Revision":
                self.btn_learning_save_revision = btn
            elif txt == "Approve Current":
                self.btn_learning_approve_current = btn
            elif txt == "Auto-Apply Best Profile":
                self.btn_learning_apply_best = btn
            elif txt == "View Current":
                self.btn_learning_view_current = btn
            elif txt == "Browse Revisions":
                self.btn_learning_browse = btn
            elif txt == "Reload Memory":
                self.btn_learning_reload = btn
            learn_grid.add_widget(btn)
        learning_body.add_widget(learn_grid)

        if is_mobile:
            self._mobile_section_cards["Learning"] = learning_card
        else:
            self._desktop_section_cards["Learning"] = learning_card

        selection_card, selection_body = make_card("Selection", "Review the currently selected box and its saved link")

        selection_summary = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(112))
        self.inspector_selected_lbl = Label(text="None", color=palette["text"], size_hint_y=None, height=dp(20), halign="left", valign="middle", font_size=dp(12))
        self.inspector_selected_lbl.bind(size=self._sync_label_text_size)
        self.inspector_count_lbl = Label(text="0", color=palette["text"], size_hint_y=None, height=dp(20), halign="left", valign="middle", font_size=dp(12))
        self.inspector_count_lbl.bind(size=self._sync_label_text_size)
        selection_summary.add_widget(labeled_field("Selected Box Numbers", self.inspector_selected_lbl, "Tap preview boxes to select them"))
        selection_summary.add_widget(labeled_field("Boxes Selected", self.inspector_count_lbl))

        selection_box_body = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        selection_box_body.bind(minimum_height=selection_box_body.setter("height"))
        selection_box_body.add_widget(selection_summary)
        self.inspector_box_lbl = Label(text="No box selected", color=palette["text"], size_hint_y=None, height=dp(40), halign="left", valign="middle", font_size=dp(12))
        self.inspector_box_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(self.inspector_box_lbl, min_height=dp(36), extra_pad=dp(6))
        selection_box_body.add_widget(labeled_field("Selected Box", self.inspector_box_lbl))

        mapping_section_body = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        mapping_section_body.bind(minimum_height=mapping_section_body.setter("height"))
        self.inspector_mapping_lbl = Label(text="Choose a box in the preview to select or map it.", color=palette["muted"], size_hint_y=None, height=dp(40), halign="left", valign="middle", font_size=dp(11))
        self.inspector_mapping_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(self.inspector_mapping_lbl, min_height=dp(36), extra_pad=dp(6))
        mapping_section_body.add_widget(labeled_field("Current Link", self.inspector_mapping_lbl))
        self.inspector_resolution_lbl = Label(text="Resolved preview value: —", color=palette["text"], size_hint_y=None, height=dp(40), halign="left", valign="middle", font_size=dp(11))
        self.inspector_resolution_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(self.inspector_resolution_lbl, min_height=dp(36), extra_pad=dp(6))
        mapping_section_body.add_widget(labeled_field("Value That Will Be Written", self.inspector_resolution_lbl))
        clear_row = GridLayout(cols=1, spacing=dp(8), size_hint_y=None, height=row_h)
        self.btn_clear_mapping = make_button("Clear Field Links", tone="plain")
        self.btn_clear_mapping.bind(on_release=self.on_clear_selected_mapping)
        clear_row.add_widget(self.btn_clear_mapping)
        mapping_section_body.add_widget(clear_row)

        hover_section_body = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        hover_section_body.bind(minimum_height=hover_section_body.setter("height"))
        self.inspector_hover_lbl = Label(text=("Tap a box to select it." if is_mobile else "Hover a box to review it."), color=palette["muted"], size_hint_y=None, height=dp(40), halign="left", valign="middle", font_size=dp(11))
        self.inspector_hover_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(self.inspector_hover_lbl, min_height=dp(36), extra_pad=dp(6))
        pointer_label = "Selected Box"
        hover_section_body.add_widget(labeled_field(pointer_label, self.inspector_hover_lbl))

        selection_body.add_widget(self._build_collapsible_panel("Selection Summary", selection_box_body, start_open=True))
        selection_body.add_widget(self._build_collapsible_panel("Mapping Details", mapping_section_body, start_open=True))
        selection_body.add_widget(self._build_collapsible_panel(pointer_label, hover_section_body, start_open=False))
        self._desktop_right_section_cards = {}
        self._desktop_right_section_buttons = {}
        self._desktop_right_section_host = None
        self._desktop_right_active_section = None

        def _style_desktop_right_button(btn, active=False):
            bg = palette["primary"] if active else palette["surface_soft"]
            fg = (1, 1, 1, 1) if active else palette["text"]
            btn.background_normal = ""
            btn.background_down = ""
            btn.background_color = bg
            btn.color = fg

        def _show_desktop_right_section(section_key):
            if is_mobile or self._desktop_right_section_host is None:
                return
            section_card = self._desktop_right_section_cards.get(section_key)
            if section_card is None:
                return
            self._desktop_right_section_host.clear_widgets()
            self._desktop_right_section_host.add_widget(section_card)
            self._desktop_right_active_section = section_key
            for key, btn in self._desktop_right_section_buttons.items():
                _style_desktop_right_button(btn, active=(key == section_key))

        self._show_desktop_right_section = _show_desktop_right_section

        def _make_desktop_right_tabs(section_names):
            wrap = GridLayout(cols=len(section_names), spacing=dp(6), size_hint_y=None, height=dp(40))
            for name in section_names:
                btn = Button(text=name, size_hint_y=None, height=dp(40), background_normal="", background_down="", background_color=palette["surface_soft"], color=palette["text"], font_size=dp(11.5))
                btn.bind(on_release=lambda inst, n=name: _show_desktop_right_section(n))
                self._desktop_right_section_buttons[name] = btn
                wrap.add_widget(btn)
            return wrap

        android_top_modes_card = None
        android_quick_actions_card = None
        android_bottom_status_card = None
        android_panel_card = None
        android_panel_scroll = None
        android_panel_host = None

        if is_mobile and platform == "android":
            android_top_modes_card = BoxLayout(
                orientation="vertical",
                spacing=dp(6),
                size_hint_y=None,
                height=dp(98),
                padding=[dp(8), dp(8), dp(8), dp(8)],
            )
            style_card(android_top_modes_card, palette["surface_alt"], radius=dp(16))

            mode_title = Label(
                text="Workspace",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(16),
                halign="left",
                valign="middle",
                font_size=dp(11),
            )
            mode_title.bind(size=self._sync_label_text_size)
            android_top_modes_card.add_widget(mode_title)

            top_mode_tabs = _make_mobile_section_tabs(["Files", "Detection", "Capture", "Templates", "Inspect", "Mapping", "Session", "Export"])  # keys unchanged; display_map applies friendly labels
            android_top_modes_card.add_widget(top_mode_tabs)

            android_quick_actions_card = BoxLayout(
                orientation="horizontal",
                spacing=dp(8),
                size_hint_y=None,
                height=dp(44),
            )
            btn_top_capture = self._make_compact_action_button("3 Draw Area", tone="accent")
            btn_top_capture.bind(on_release=lambda *_: self._show_mobile_section("Capture"))
            btn_top_templates = self._make_compact_action_button("4 Templates", tone="plain")
            btn_top_templates.bind(on_release=lambda *_: self._show_mobile_section("Templates"))
            btn_top_repair = self._make_compact_action_button("5 Repair", tone="secondary")
            btn_top_repair.bind(on_release=lambda *_: self._show_mobile_section("Inspect"))
            btn_top_help = self._make_compact_action_button("? Help", tone="plain")
            btn_top_help.bind(on_release=lambda *_: self._show_help_popup())
            for _w in (btn_top_capture, btn_top_templates, btn_top_repair, btn_top_help):
                _w.size_hint = (1, None)
                _w.height = dp(44)
                android_quick_actions_card.add_widget(_w)
            android_top_modes_card.add_widget(android_quick_actions_card)

            android_panel_card = BoxLayout(
                orientation="vertical",
                spacing=dp(6),
                size_hint_y=None,
                height=dp(210),
                padding=[dp(8), dp(8), dp(8), dp(8)],
            )
            style_card(android_panel_card, palette["surface_alt"], radius=dp(16))
            panel_title = Label(
                text="Panel",
                color=palette["muted"],
                size_hint_y=None,
                height=dp(16),
                halign="left",
                valign="middle",
                font_size=dp(11),
            )
            panel_title.bind(size=self._sync_label_text_size)
            android_panel_card.add_widget(panel_title)
            android_panel_scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                bar_width=dp(4),
                scroll_type=["bars", "content"],
            )
            android_panel_host = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
            android_panel_host.bind(minimum_height=android_panel_host.setter("height"))
            android_panel_scroll.add_widget(android_panel_host)
            android_panel_card.add_widget(android_panel_scroll)

            android_bottom_status_card = FloatingPreviewHUD(
                orientation="vertical",
                spacing=dp(4),
                size_hint=(None, None),
                width=dp(300),
                height=dp(136),
                padding=[dp(10), dp(8), dp(10), dp(8)],
                pos_hint={},
            )
            self._style_popup_card(android_bottom_status_card, palette.get("surface_alt", (0.042, 0.068, 0.138, 0.995)), radius=dp(16))

            helper_header = BoxLayout(orientation="horizontal", spacing=dp(6), size_hint_y=None, height=dp(28))
            helper_title = Label(text="Preview Helper", color=palette["text"], halign="left", valign="middle", font_size=dp(11.5), bold=True)
            helper_title.bind(size=self._sync_label_text_size)
            helper_drag_lbl = Label(text="drag", color=palette["muted"], size_hint=(None, None), width=dp(44), height=dp(18), halign="right", valign="middle", font_size=dp(9.5))
            helper_drag_lbl.bind(size=self._sync_label_text_size)
            self.btn_android_helper_toggle = self._make_compact_action_button("Hide", tone="ghost")
            self.btn_android_helper_toggle._expanded_label = "Hide"
            self.btn_android_helper_toggle._collapsed_label = "Show"
            self.btn_android_helper_toggle.size_hint = (None, None)
            self.btn_android_helper_toggle.size = (dp(50), dp(24))
            self.btn_android_helper_toggle = self._make_compact_action_button("—", tone="ghost")
            self.btn_android_helper_toggle._expanded_label = "—"
            self.btn_android_helper_toggle._collapsed_label = "✦"
            self.btn_android_helper_toggle.size_hint = (None, None)
            self.btn_android_helper_toggle.size = (dp(34), dp(24))
            helper_header.add_widget(helper_title)
            helper_header.add_widget(helper_drag_lbl)
            helper_header.add_widget(self.btn_android_helper_toggle)

            helper_body = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
            helper_body.bind(minimum_height=helper_body.setter("height"))
            status_row = GridLayout(cols=2, rows=2, spacing=dp(4), size_hint=(1, None), height=dp(42))
            self.android_status_file_lbl = Label(text="File: No PDF", color=palette["muted"], halign="left", valign="middle", font_size=dp(10.5))
            self.android_status_page_lbl = Label(text="Page: 1/1", color=palette["muted"], halign="left", valign="middle", font_size=dp(10.5))
            self.android_status_boxes_lbl = Label(text="Boxes: 0 • Sel: 0", color=palette["muted"], halign="left", valign="middle", font_size=dp(10.5))
            self.android_status_mode_lbl = Label(text="Mode: Files", color=palette["text"], halign="left", valign="middle", font_size=dp(11.5), bold=True)
            for _lbl in (self.android_status_file_lbl, self.android_status_page_lbl, self.android_status_boxes_lbl, self.android_status_mode_lbl):
                _lbl.bind(size=self._sync_label_text_size)
                status_row.add_widget(_lbl)
            helper_body.add_widget(status_row)

            legend_row = BoxLayout(orientation="horizontal", spacing=dp(6), size_hint_y=None, height=dp(22))
            legend_lbl = Label(
                text="Legend: Field / Check / ROI / Trace",
                color=palette["muted"],
                halign="left",
                valign="middle",
                font_size=dp(10.5),
            )
            legend_lbl.bind(size=self._sync_label_text_size)
            legend_row.add_widget(legend_lbl)
            for _txt, _col in [
                ("Field", (0.25, 0.90, 0.42, 1)),
                ("Check", (1.00, 0.88, 0.25, 1)),
                ("ROI", (0.35, 0.70, 1.00, 1)),
                ("Trace", (1.00, 0.60, 0.18, 1)),
            ]:
                chip = Label(text=str(_txt), color=_col, halign="left", valign="middle", font_size=dp(10.5))
                chip = Label(text=f"[color=#{int(_col[0]*255):02x}{int(_col[1]*255):02x}{int(_col[2]*255):02x}]●[/color] {_txt}", markup=True, color=palette["muted"], halign="left", valign="middle", font_size=dp(10.5))
                chip.bind(size=self._sync_label_text_size)
                legend_row.add_widget(chip)
            helper_body.add_widget(legend_row)

            android_bottom_status_card.add_widget(helper_header)
            android_bottom_status_card.add_widget(helper_body)
            android_bottom_status_card.set_body_widget(helper_body, expanded_height=dp(132))
            android_bottom_status_card.set_toggle_button(self.btn_android_helper_toggle)
            android_bottom_status_card.register_interactive_widgets(self.btn_android_helper_toggle)
            android_bottom_status_card.set_drag_enabled(True)
            android_bottom_status_card.set_allow_body_passthrough(False)
            android_bottom_status_card.set_allow_body_passthrough(True)

        if is_mobile:
            main.spacing = 0
            workspace = FloatLayout()
            self.mobile_workspace = workspace

            try:
                appbar.remove_widget(status_chip)
            except Exception:
                pass

            self.btn_sidebar_toggle = self._make_compact_action_button("Menu", tone="ghost")
            self.btn_sidebar_toggle.size_hint = (None, None)
            self.btn_sidebar_toggle.size = (dp(84), dp(34))
            if platform == "android":
                self.btn_sidebar_toggle.opacity = 0
                self.btn_sidebar_toggle.disabled = True
                self.btn_sidebar_toggle.size = (0, 0)
            appbar.add_widget(self.btn_sidebar_toggle)

            stage_column = BoxLayout(
                orientation="vertical",
                spacing=dp(6),
                padding=[dp(4), dp(0), dp(4), dp(4)],
                size_hint=(1, 1),
            )
            workspace.add_widget(stage_column)

            if platform == "android":
                if android_top_modes_card is not None:
                    stage_column.add_widget(android_top_modes_card)
                if android_panel_card is not None:
                    stage_column.add_widget(android_panel_card)
            else:
                mobile_workspace_bar = Widget(size_hint_y=None, height=0, opacity=0)
                stage_column.add_widget(mobile_workspace_bar)

            self.preview_shell = BoxLayout(
                orientation="vertical",
                spacing=0,
                padding=[dp(0), dp(0), dp(0), dp(0)],
                size_hint=(1, 1),
            )
            style_card(self.preview_shell, palette["preview_bg"], radius=dp(14))
            preview_stage = FloatLayout(size_hint=(1, 1))
            self.preview_stage = preview_stage
            preview_wrap = ScrollView(
                do_scroll_x=True,
                do_scroll_y=True,
                bar_width=dp(4),
                scroll_type=["bars", "content"],
                size_hint=(1, 1),
                pos_hint={"x": 0, "y": 0},
            )
            self.preview_wrap = preview_wrap
            preview_stack = BoxLayout(orientation="vertical", spacing=dp(4), size_hint=(None, None))
            self.preview_stack = preview_stack
            preview_stack.bind(minimum_height=preview_stack.setter("height"))
            preview_stack.bind(minimum_width=preview_stack.setter("width"))
            self.preview = InteractivePreview(size_hint=(None, None))
            self.preview.bind(texture=self._update_preview_size)
            self.preview.box_tap_callback = self.on_preview_box_tap
            self.preview.box_double_tap_callback = self.on_preview_box_double_tap
            self.preview.preview_tap_callback = self.on_preview_point_tap
            self.preview.hover_callback = self.on_preview_box_hover
            self.preview.zoom_request_callback = self.on_preview_zoom_request
            preview_wrap.bind(size=self._on_preview_viewport_change, pos=self._on_preview_viewport_change)
            self.preview_shell.bind(size=self._on_preview_viewport_change, pos=self._on_preview_viewport_change)
            self.preview.bind(size=self._sync_preview_stack_size)
            self.preview_info = Label(
                text="",
                color=palette["muted"],
                size_hint_y=None,
                height=0,
                halign="left",
                valign="middle",
                font_size=dp(8.8),
                opacity=0,
                disabled=True,
            )
            self.preview_info.bind(size=self._sync_label_text_size)
            preview_stack.add_widget(self.preview_info)
            self.preview_empty_hint = self._make_preview_empty_hint(palette, is_mobile=True)
            preview_stack.add_widget(self.preview_empty_hint)
            preview_stack.add_widget(self.preview)
            preview_wrap.add_widget(preview_stack)
            preview_stage.add_widget(preview_wrap)

            self.preview_hud = _make_preview_hud()
            self.preview_hud.size_hint = (None, None)
            self.preview_hud.width = dp(228) if is_mobile else dp(258)
            self.preview_hud.height = dp(126) if is_mobile else dp(142)
            self.preview_hud.pos_hint = {"right": 0.985, "y": 0.10}
            self.preview_hud.set_drag_enabled(True)
            self.preview_hud.set_allow_body_passthrough(False)
            self.preview_hud.set_allow_body_passthrough(True)
            self.preview_hud.opacity = 0
            self.preview_hud.disabled = True
            preview_stage.add_widget(self.preview_hud)
            self._bind_overlay_panel_bounds(self.preview_hud)

            self.btn_show_hud = None

            self.preview_shell.add_widget(preview_stage)
            stage_column.add_widget(self.preview_shell)
            if platform == "android" and android_bottom_status_card is not None:
                android_bottom_status_card.pos = (dp(10), dp(10))
                preview_stage.add_widget(android_bottom_status_card)
                self._bind_overlay_panel_bounds(android_bottom_status_card)

            # Focus Mode: floating toggle button anchored to top-left of the preview stage.
            # Tapping it hides the control panels so the preview fills the screen.
            self._focus_mode_active = False
            self.android_top_modes_card = android_top_modes_card
            self.android_panel_card = android_panel_card
            _focus_btn_size = dp(48)
            self.btn_focus_mode = self._make_compact_action_button("Focus", tone="ghost")
            self.btn_focus_mode.size_hint = (None, None)
            self.btn_focus_mode.size = (dp(72), dp(42))
            self.btn_focus_mode.pos_hint = {"x": 0.01, "top": 0.99}
            self.btn_focus_mode.opacity = 0.72
            self.btn_focus_mode.font_size = dp(11)
            self.btn_focus_mode.bind(on_release=self._toggle_focus_mode)
            preview_stage.add_widget(self.btn_focus_mode)

            quick_rail = BoxLayout(
                orientation="vertical",
                spacing=dp(6),
                size_hint=(None, None),
                width=dp(88),
                height=dp(156),
                padding=[dp(7), dp(7), dp(7), dp(7)],
                pos_hint={"right": 0.992, "center_y": 0.56},
            )
            style_card(quick_rail, palette["surface_alt"], radius=dp(16))
            self.btn_mobile_rail_fit_page = self._make_compact_action_button("Fit Pg", tone="ghost")
            self.btn_mobile_rail_fit_page.bind(on_release=self._fit_preview_page)
            self.btn_mobile_rail_fit_width = self._make_compact_action_button("Fit W", tone="ghost")
            self.btn_mobile_rail_fit_width.bind(on_release=self._fit_preview_width)
            self.btn_mobile_rail_zoom_reset = self._make_compact_action_button("100%", tone="ghost")
            self.btn_mobile_rail_zoom_reset.bind(on_release=lambda *_: self._zoom_preview("reset"))
            self.btn_mobile_map = None
            self.btn_mobile_focus = None
            for w in [
                self.btn_mobile_rail_fit_width,
                self.btn_mobile_rail_fit_page,
                self.btn_mobile_rail_zoom_reset,
            ]:
                w.size_hint = (None, None)
                w.size = (dp(74), dp(44))
                quick_rail.add_widget(w)
            self.mobile_quick_rail = quick_rail
            self._set_widget_enabled(quick_rail, False)
            quick_rail.opacity = 0
            quick_rail.disabled = True
            preview_stage.add_widget(quick_rail)

            bottom_dock = FloatLayout(size_hint=(1, 1))
            bottom_panel = BoxLayout(
                orientation="vertical",
                spacing=dp(6),
                size_hint=(None, None),
                width=min(Window.width - dp(24), dp(500)),
                height=dp(170),
                padding=[dp(8), dp(8), dp(8), dp(8)],
                pos=(Window.width - min(Window.width - dp(24), dp(500)) - dp(12), dp(64)),
                opacity=0,
                disabled=True,
            )
            style_card(bottom_panel, palette["surface_alt"], radius=dp(16))
            primary_row = GridLayout(
                cols=4,
                rows=1,
                spacing=dp(6),
                size_hint=(1, None),
                height=dp(44),
            )
            view_row = GridLayout(
                cols=4,
                rows=1,
                spacing=dp(6),
                size_hint=(1, None),
                height=dp(44),
            )
            context_row = GridLayout(
                cols=6,
                rows=1,
                spacing=dp(4),
                size_hint=(1, None),
                height=dp(40),
            )
            self.btn_mobile_prev = self._make_compact_action_button("Prev", tone="plain")
            self.btn_mobile_prev.bind(on_release=self.on_prev_page)
            self.page_input_mobile = make_input("0", "Pg", input_filter="int")
            self.page_input_mobile.size_hint = (None, None)
            self.page_input_mobile.size = (dp(58), dp(36))
            self.page_input_mobile.bind(text=self._mirror_page_inputs)
            self.btn_mobile_next = self._make_compact_action_button("Next", tone="plain")
            self.btn_mobile_next.bind(on_release=self.on_next_page)
            self.btn_mobile_fit_width = self._make_compact_action_button("Fit W", tone="ghost")
            self.btn_mobile_fit_width.bind(on_release=self._fit_preview_width)
            self.btn_mobile_fit_page = self._make_compact_action_button("Fit Pg", tone="ghost")
            self.btn_mobile_fit_page.bind(on_release=self._fit_preview_page)
            self.btn_mobile_zoom_out = self._make_compact_action_button("Zoom -", tone="mini")
            self.btn_mobile_zoom_out.bind(on_release=lambda *_: self._zoom_preview("out"))
            self.btn_mobile_zoom_in = self._make_compact_action_button("Zoom +", tone="mini")
            self.btn_mobile_zoom_in.bind(on_release=lambda *_: self._zoom_preview("in"))
            self.btn_mobile_zoom_reset = self._make_compact_action_button("100%", tone="ghost")
            self.btn_mobile_zoom_reset.bind(on_release=lambda *_: self._zoom_preview("reset"))
            self.btn_mobile_focus = None
            self.btn_mobile_map = None
            self.btn_mobile_refresh = self._make_compact_action_button("Refresh", tone="soft")
            self.btn_mobile_refresh.bind(on_release=self.on_preview)
            self.btn_mobile_detect_bar = self._make_compact_action_button("Find", tone="accent")
            self.btn_mobile_detect_bar.bind(on_release=self.on_run_detect)
            self.btn_mobile_select_mode = self._make_compact_action_button("Select Off", tone="plain")
            self.btn_mobile_select_mode.bind(on_release=self._toggle_mobile_selection_mode)
            self.btn_mobile_sidebar = self._make_compact_action_button("Menu", tone="ghost")
            if platform == "android":
                self.btn_mobile_sidebar.opacity = 0
                self.btn_mobile_sidebar.disabled = True
                self.btn_mobile_sidebar.size_hint = (None, None)
                self.btn_mobile_sidebar.size = (0, 0)
            else:
                self.btn_mobile_sidebar.bind(on_release=lambda *_: self._toggle_mobile_sidebar())
            self.btn_mobile_more = self._make_compact_action_button("Repair", tone="secondary")
            self.btn_mobile_more.bind(on_release=(lambda *_: self._show_mobile_section("Inspect")) if platform == "android" else self._toggle_mobile_inspect_panel)
            self.btn_mobile_capture = self._make_compact_action_button("Capture", tone="accent")
            self.btn_mobile_capture.bind(on_release=(lambda *_: self._show_mobile_section("Capture")) if platform == "android" else self._toggle_manual_area_capture_mode)
            self.btn_mobile_templates = self._make_compact_action_button("Tpls", tone="plain")
            self.btn_mobile_templates.bind(on_release=(lambda *_: self._show_mobile_section("Templates")) if platform == "android" else self._open_area_template_library_popup)
            self.btn_mobile_tools_fab = self._make_compact_action_button("More", tone="primary")
            self.btn_mobile_tools_fab.size_hint = (None, None)
            self.btn_mobile_tools_fab.size = (dp(88), dp(48))
            self.btn_mobile_tools_fab.pos = (Window.width - dp(86), dp(12))
            self.btn_mobile_tools_fab.bind(on_release=self._toggle_mobile_tools_tray)
            self.page_input = self.page_input_mobile
            self.btn_mobile_prev.text = "Prev"
            self.btn_mobile_next.text = "Next"
            self.btn_mobile_fit_width.text = "Fit W"
            self.btn_mobile_fit_page.text = "Fit Pg"
            self.btn_mobile_zoom_in.text = "Zoom +"
            self.btn_mobile_zoom_out.text = "Zoom -"
            self.btn_mobile_zoom_reset.text = "100%"
            self.btn_mobile_detect_bar.text = "Find"
            self.btn_mobile_select_mode.text = "Select Off"
            self.btn_mobile_sidebar.text = "Menu"
            self.btn_mobile_more.text = "Repair"
            self.btn_mobile_capture.text = "Capture"
            self.btn_mobile_templates.text = "Tpls"
            for widget in [
                self.btn_mobile_prev,
                self.page_input_mobile,
                self.btn_mobile_next,
                self.btn_mobile_detect_bar,
            ]:
                widget.size_hint = (1, None)
                widget.height = dp(38)
                primary_row.add_widget(widget)
            for widget in [
                self.btn_mobile_fit_width,
                self.btn_mobile_fit_page,
                self.btn_mobile_zoom_out,
                self.btn_mobile_zoom_in,
            ]:
                widget.size_hint = (1, None)
                widget.height = dp(38)
                view_row.add_widget(widget)
            for widget in [
                self.btn_mobile_zoom_reset,
                self.btn_mobile_select_mode,
                self.btn_mobile_more,
                self.btn_mobile_capture,
                self.btn_mobile_templates,
                self.btn_mobile_sidebar,
            ]:
                widget.size_hint = (1, None)
                widget.height = dp(38)
                context_row.add_widget(widget)
            bottom_panel.add_widget(primary_row)
            bottom_panel.add_widget(view_row)
            bottom_panel.add_widget(context_row)
            bottom_dock.add_widget(bottom_panel)
            bottom_dock.add_widget(self.btn_mobile_tools_fab)
            self.mobile_bottom_dock = bottom_dock
            self.mobile_bottom_panel = bottom_panel
            self.mobile_bottom_bar = primary_row
            self.mobile_tools_tray_open = False
            if platform == "android":
                try:
                    bottom_panel.opacity = 0
                    bottom_panel.disabled = True
                    bottom_panel.size_hint = (None, None)
                    bottom_panel.size = (0, 0)
                    bottom_panel.pos = (-10000, -10000)
                except Exception:
                    pass
                try:
                    bottom_dock.opacity = 0
                    bottom_dock.disabled = True
                    bottom_dock.size_hint = (None, None)
                    bottom_dock.size = (0, 0)
                    bottom_dock.pos = (-10000, -10000)
                except Exception:
                    pass
                try:
                    self.btn_mobile_tools_fab.opacity = 0
                    self.btn_mobile_tools_fab.disabled = True
                    self.btn_mobile_tools_fab.size_hint = (None, None)
                    self.btn_mobile_tools_fab.size = (0, 0)
                    self.btn_mobile_tools_fab.pos = (-10000, -10000)
                except Exception:
                    pass
            workspace.add_widget(bottom_dock)

            self._sync_mobile_tools_layout()
            self._update_capture_button_state()
            Window.bind(size=self._sync_mobile_tools_layout)

            self.mobile_sidebar_scrim = Button(
                text="",
                background_normal="",
                background_down="",
                background_color=(0.0, 0.0, 0.0, 0.36),
                size_hint=(None, None),
                size=(0, 0),
                pos=(-10000, -10000),
                opacity=0,
                disabled=True,
            )
            workspace.add_widget(self.mobile_sidebar_scrim)

            sidebar_w = min(dp(340), Window.width * 0.84)
            self.mobile_sidebar = BoxLayout(
                orientation="vertical",
                spacing=dp(6),
                size_hint=(None, 1),
                width=sidebar_w,
                padding=[dp(8), dp(8), dp(8), dp(8)],
                pos=(-sidebar_w - dp(8), 0),
            )
            style_card(self.mobile_sidebar, palette["surface"], radius=dp(18))
            sidebar_header = BoxLayout(orientation="horizontal", spacing=dp(6), size_hint_y=None, height=dp(34))
            sidebar_title = Label(text="Tools", color=palette["text"], bold=True, halign="left", valign="middle", font_size=dp(12.5))
            sidebar_title.bind(size=self._sync_label_text_size)
            self.btn_sidebar_close = self._make_compact_action_button("Close", tone="plain")
            self.btn_sidebar_close.size_hint = (None, None)
            self.btn_sidebar_close.size = (dp(104), dp(32))
            try:
                self.btn_sidebar_close.text_size = (None, None)
            except Exception:
                pass
            sidebar_header.add_widget(sidebar_title)
            sidebar_header.add_widget(self.btn_sidebar_close)
            self.mobile_sidebar.add_widget(sidebar_header)

            if status_chip.parent is not None:
                try:
                    status_chip.parent.remove_widget(status_chip)
                except Exception:
                    pass
            status_chip.size_hint = (1, None)
            status_chip.height = dp(30)
            self.mobile_sidebar.add_widget(status_chip)

            if mobile_flow_card is not None:
                if mobile_flow_card.parent is not None:
                    try:
                        mobile_flow_card.parent.remove_widget(mobile_flow_card)
                    except Exception:
                        pass
                self.mobile_sidebar.add_widget(mobile_flow_card)

            if platform != "android":
                sidebar_tabs = _make_mobile_section_tabs(["Files", "Detection", "Text", "Learning", "Mapping", "Session", "Export"])
                self.mobile_sidebar.add_widget(sidebar_tabs)

                sidebar_scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, bar_width=dp(4), bar_margin=dp(10), scroll_type=["bars", "content"])
                self._mobile_section_host = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
                self._mobile_section_host.bind(minimum_height=self._mobile_section_host.setter("height"))
                sidebar_scroll.add_widget(self._mobile_section_host)
                self.mobile_sidebar.add_widget(sidebar_scroll)
                workspace.add_widget(self.mobile_sidebar)

                self.btn_sidebar_toggle.bind(on_release=self._toggle_mobile_sidebar)
                self.btn_sidebar_close.bind(on_release=self._toggle_mobile_sidebar)
                self.mobile_sidebar_scrim.bind(on_release=self._toggle_mobile_sidebar)
                Clock.schedule_once(lambda dt: _show_mobile_section("Files"), 0)
                Clock.schedule_once(lambda dt: self._set_mobile_sidebar_state(False), 0)
            else:
                # Android cleanup: keep the new shell as the only navigation source.
                # Route the section tabs into the visible Android panel host instead
                # of the legacy slide-out sidebar.
                self._mobile_section_host = android_panel_host
                try:
                    self.mobile_sidebar.opacity = 0
                    self.mobile_sidebar.disabled = True
                    self.mobile_sidebar.size_hint = (None, None)
                    self.mobile_sidebar.size = (0, 0)
                    self.mobile_sidebar.pos = (-10000, -10000)
                except Exception:
                    pass
                try:
                    self.mobile_sidebar_scrim.opacity = 0
                    self.mobile_sidebar_scrim.disabled = True
                    self.mobile_sidebar_scrim.size = (0, 0)
                    self.mobile_sidebar_scrim.pos = (-10000, -10000)
                except Exception:
                    pass
                try:
                    self.btn_sidebar_close.opacity = 0
                    self.btn_sidebar_close.disabled = True
                    self.btn_sidebar_close.size_hint = (None, None)
                    self.btn_sidebar_close.size = (0, 0)
                except Exception:
                    pass
                self._mobile_sidebar_open = False
                if android_panel_host is not None:
                    Clock.schedule_once(lambda dt: _show_mobile_section("Detection"), 0)
                    try:
                        if getattr(self, "android_status_mode_lbl", None) is not None:
                            self.android_status_mode_lbl.text = "Mode: Detection"
                    except Exception:
                        pass

            Clock.schedule_once(lambda dt: self._set_mobile_quick_rail_visible(False), 0)
            Clock.schedule_once(lambda dt: self._refresh_mobile_select_mode_ui(), 0)
            Clock.schedule_once(lambda dt: self._refresh_mobile_hud_restore_button(), 0)

            main.add_widget(workspace)
        else:
            left_wrap = BoxLayout(orientation="vertical", size_hint=(0.23, 1), spacing=dp(10))
            left_scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, bar_width=dp(5), scroll_type=["bars", "content"])
            left_body = GridLayout(cols=1, spacing=dp(10), size_hint_y=None)
            left_body.bind(minimum_height=left_body.setter("height"))
            desktop_controls_card, desktop_controls_body = make_card("Document & Detection", "Compact desktop controls")
            tabs = _make_desktop_section_tabs(["Workspace", "Session", "Detection", "Text", "Learning"])
            desktop_controls_body.add_widget(tabs)
            self._desktop_section_host = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
            self._desktop_section_host.bind(minimum_height=self._desktop_section_host.setter("height"))
            desktop_controls_body.add_widget(self._desktop_section_host)
            left_body.add_widget(desktop_controls_card)
            self._desktop_section_cards["Workspace"] = files_card
            self._desktop_section_cards["Session"] = nav_card
            self._desktop_section_cards["Detection"] = detect_card
            left_scroll.add_widget(left_body)
            left_wrap.add_widget(left_scroll)
            Clock.schedule_once(lambda dt: _show_desktop_section("Workspace"), 0)

            center_wrap = BoxLayout(orientation="vertical", size_hint=(0.57, 1), spacing=dp(10))
            toolbar = BoxLayout(orientation="horizontal", size_hint_y=None, height=dp(44), spacing=dp(8))
            self.btn_open_pdf_toolbar = make_button("Open PDF", tone="primary")
            self.btn_open_pdf_toolbar.bind(on_release=self.on_load_pdf)
            self.btn_open_data_toolbar = make_button("Open Data File", tone="soft")
            self.btn_open_data_toolbar.bind(on_release=self.on_load_csv)
            self.btn_toolbar_detect = make_button("Find Fields", tone="accent")
            self.btn_toolbar_detect.bind(on_release=self.on_run_detect)
            self.btn_toolbar_refresh = make_button("Refresh Preview", tone="primary")
            self.btn_toolbar_refresh.bind(on_release=self.on_preview)
            self.btn_toolbar_prev = make_button("Prev", tone="plain")
            self.btn_toolbar_prev.bind(on_release=self.on_prev_page)
            self.btn_toolbar_next = make_button("Next", tone="plain")
            self.btn_toolbar_next.bind(on_release=self.on_next_page)
            self.btn_zoom_out = make_button("−", tone="plain")
            self.btn_zoom_out.bind(on_release=lambda *_: self._zoom_preview("out"))
            zoom_chip = BoxLayout(orientation="horizontal", size_hint=(None, None), size=(dp(84), dp(44)), padding=[dp(12), dp(10), dp(12), dp(10)])
            style_card(zoom_chip, palette["chip"], radius=dp(16))
            self.zoom_chip_lbl = Label(text="100%", color=palette["text"], halign="center", valign="middle", font_size=dp(12))
            self.zoom_chip_lbl.bind(size=self._sync_label_text_size)
            zoom_chip.add_widget(self.zoom_chip_lbl)
            self.btn_zoom_in = make_button("+", tone="plain")
            self.btn_zoom_in.bind(on_release=lambda *_: self._zoom_preview("in"))
            self.btn_zoom_reset = make_button("100%", tone="plain")
            self.btn_zoom_reset.bind(on_release=lambda *_: self._zoom_preview("reset"))
            self.btn_fit_width = make_button("Fit Width", tone="plain")
            self.btn_fit_width.bind(on_release=self._fit_preview_width)
            self.btn_fit_page = make_button("Fit Page", tone="plain")
            self.btn_fit_page.bind(on_release=self._fit_preview_page)
            self.btn_toolbar_export = make_button("Export PDF", tone="accent")
            self.btn_toolbar_export.bind(on_release=self.on_generate_single)
            self.btn_toolbar_focus = make_button("Focus", tone="ghost")
            self.btn_toolbar_focus.size_hint = (None, 1)
            self.btn_toolbar_focus.width = dp(70)
            self.btn_toolbar_focus.bind(on_release=self._toggle_focus_mode)
            self._focus_mode_active = False
            self._desktop_left_wrap = None
            self._desktop_right_wrap = None
            self.btn_focus_mode = self.btn_toolbar_focus
            for w in [self.btn_open_pdf_toolbar, self.btn_open_data_toolbar, self.btn_toolbar_detect, self.btn_toolbar_refresh, self.btn_toolbar_prev, self.btn_toolbar_next, self.btn_zoom_out, zoom_chip, self.btn_zoom_in, self.btn_zoom_reset, self.btn_fit_width, self.btn_fit_page, self.btn_toolbar_export, self.btn_toolbar_focus]:
                toolbar.add_widget(w)
            center_wrap.add_widget(toolbar)

            preview_card, preview_body = make_card("Preview Canvas", "Hover boxes, zoom, map, and export from the live document")
            preview_card.size_hint_y = 1
            self.preview_shell = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(8), size_hint_y=None)
            self.preview_shell.height = max(dp(620), Window.height - dp(220))
            style_card(self.preview_shell, palette["preview_bg"], radius=dp(24))
            preview_stage = FloatLayout(size_hint=(1, 1))
            self.preview_stage = preview_stage
            preview_wrap = ScrollView(do_scroll_x=True, do_scroll_y=True, bar_width=dp(6), scroll_type=["bars", "content"], size_hint=(1, 1), pos_hint={"x": 0, "y": 0})
            self.preview_wrap = preview_wrap
            preview_stack = BoxLayout(orientation="vertical", spacing=dp(6), size_hint=(None, None))
            self.preview_stack = preview_stack
            preview_stack.bind(minimum_height=preview_stack.setter("height"))
            preview_stack.bind(minimum_width=preview_stack.setter("width"))
            self.preview = InteractivePreview(size_hint=(None, None))
            self.preview.bind(texture=self._update_preview_size)
            self.preview.box_tap_callback = self.on_preview_box_tap
            self.preview.box_double_tap_callback = self.on_preview_box_double_tap
            self.preview.preview_tap_callback = self.on_preview_point_tap
            self.preview.hover_callback = self.on_preview_box_hover
            self.preview.zoom_request_callback = self.on_preview_zoom_request
            preview_wrap.bind(size=self._on_preview_viewport_change, pos=self._on_preview_viewport_change)
            self.preview_shell.bind(size=self._on_preview_viewport_change, pos=self._on_preview_viewport_change)
            self.preview.bind(size=self._sync_preview_stack_size)
            self.preview_info = Label(text="", color=palette["text"], size_hint_y=None, height=0, halign="left", valign="middle", font_size=dp(11), opacity=0, disabled=True)
            self.preview_info.bind(size=self._sync_label_text_size)
            preview_stack.add_widget(self.preview_info)
            self.preview_empty_hint = self._make_preview_empty_hint(palette, is_mobile=False)
            preview_stack.add_widget(self.preview_empty_hint)
            preview_stack.add_widget(self.preview)
            preview_wrap.add_widget(preview_stack)
            preview_stage.add_widget(preview_wrap)
            self.preview_hud = _make_preview_hud()
            preview_stage.add_widget(self.preview_hud)
            Clock.schedule_once(lambda dt: setattr(self.preview_hud, "pos", (max(0, preview_stage.width - self.preview_hud.width - dp(10)), dp(12))), 0)
            self.preview_shell.add_widget(preview_stage)
            preview_body.add_widget(self.preview_shell)
            center_wrap.add_widget(preview_card)

            right_wrap = BoxLayout(orientation="vertical", size_hint=(0.20, 1), spacing=dp(10))
            right_card, right_body = make_card("Inspector", "Selection, mapping, and export")
            right_tabs = _make_desktop_right_tabs(["Selection", "Mapping", "Export"])
            right_body.add_widget(right_tabs)
            self._desktop_right_section_host = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
            self._desktop_right_section_host.bind(minimum_height=self._desktop_right_section_host.setter("height"))
            right_body.add_widget(self._desktop_right_section_host)
            right_wrap.add_widget(right_card)
            self._desktop_right_section_cards["Selection"] = selection_card
            self._desktop_right_section_cards["Mapping"] = map_card
            self._desktop_right_section_cards["Export"] = export_card
            Clock.schedule_once(lambda dt: _show_desktop_right_section("Selection"), 0)

            self._desktop_left_wrap = left_wrap
            self._desktop_center_wrap = center_wrap
            self._desktop_right_wrap = right_wrap
            main.add_widget(left_wrap)
            main.add_widget(center_wrap)
            main.add_widget(right_wrap)

            bottom_status = BoxLayout(orientation="horizontal", spacing=dp(10), size_hint_y=None, height=dp(42), padding=[dp(12), dp(8), dp(12), dp(8)])
            style_card(bottom_status, palette["surface_alt"], radius=dp(18))
            def _status_piece(width_hint=0.2):
                lbl = Label(text="", color=palette["muted"], halign="left", valign="middle", size_hint=(width_hint, 1), font_size=dp(11))
                lbl.bind(size=self._sync_label_text_size)
                return lbl
            self.statusbar_file_lbl = _status_piece(0.26)
            self.statusbar_page_lbl = _status_piece(0.13)
            self.statusbar_patient_lbl = _status_piece(0.24)
            self.statusbar_boxes_lbl = _status_piece(0.17)
            self.statusbar_ready_lbl = _status_piece(0.20)
            for lbl in [self.statusbar_file_lbl, self.statusbar_page_lbl, self.statusbar_patient_lbl, self.statusbar_boxes_lbl, self.statusbar_ready_lbl]:
                bottom_status.add_widget(lbl)
            self.desktop_status_detail_lbl = self.status_lbl
            self.status_lbl.size_hint = (None, None)
            self.status_lbl.width = dp(420)
            root.add_widget(main)
            root.add_widget(bottom_status)
            Clock.schedule_once(lambda dt: self._bind_desktop_shortcuts(), 0)

        if is_mobile:
            root.add_widget(main)

        app_shell = FloatLayout()
        root.size_hint = (1, 1)
        root.pos_hint = {"x": 0, "y": 0}
        app_shell.add_widget(root)
        self.root_container = root
        self._app_shell = app_shell
        # Route B startup: no Python/Kivy startup splash. Startup visuals are owned by the custom native Android activity/theme.
        self._startup_presplash = None

        Clock.schedule_once(lambda dt: self.refresh_backend_capabilities_ui(), 0)
        Clock.schedule_once(lambda dt: self._refresh_preview_empty_hint(), 0)
        Clock.schedule_once(lambda dt: self.refresh_learning_ui(), 0)
        Clock.schedule_once(lambda dt: self._maybe_restore_learning_session(), 0.15)
        Clock.schedule_once(self._hide_android_loading_screen, 0)
        Clock.schedule_once(self._hide_android_loading_screen, 0.12)
        Clock.schedule_once(self._hide_android_loading_screen, 0.45)
        return app_shell

    def _make_preview_empty_hint(self, palette, is_mobile=False):
        outer = BoxLayout(
            orientation="vertical",
            spacing=dp(10 if is_mobile else 12),
            padding=[dp(12 if is_mobile else 16), dp(12 if is_mobile else 16), dp(12 if is_mobile else 16), dp(12 if is_mobile else 16)],
            size_hint_y=None,
        )
        outer.bind(minimum_height=outer.setter("height"))
        self._style_popup_card(outer, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(18 if is_mobile else 20))

        hero = BoxLayout(
            orientation="vertical",
            spacing=dp(6),
            padding=[dp(14 if is_mobile else 16), dp(12 if is_mobile else 14), dp(14 if is_mobile else 16), dp(12 if is_mobile else 14)],
            size_hint_y=None,
        )
        hero.bind(minimum_height=hero.setter("height"))
        self._style_popup_card(hero, palette.get("surface", (0.08, 0.11, 0.16, 1)), radius=dp(16 if is_mobile else 18))

        title = Label(
            text="Welcome! Open a PDF form to begin.",
            color=palette.get("text", (0.93, 0.96, 1.0, 1)),
            size_hint_y=None,
            halign="left",
            valign="middle",
            font_size=dp(17 if is_mobile else 19),
            bold=True,
        )
        title.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(title, min_height=dp(26), extra_pad=dp(4))

        subtitle = Label(
            text=(
                "Your filled form preview will appear here once you open a PDF."
                if is_mobile else
                "Your live form preview, detected fields, and mapping guides will appear here."
            ),
            color=palette.get("muted", (0.60, 0.68, 0.80, 1)),
            size_hint_y=None,
            halign="left",
            valign="top",
            font_size=dp(10.8 if is_mobile else 11.5),
        )
        subtitle.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(subtitle, min_height=dp(34 if is_mobile else 28), extra_pad=dp(6))

        hero.add_widget(title)
        hero.add_widget(subtitle)

        steps = Label(
            text=(
                "How to get started:\n"
                "1. Open a PDF form  (1 Open Files tab)\n"
                "2. Load your CSV or Excel data file\n"
                "3. Find fields on the form  (2 Find Fields tab)\n"
                "4. Link fields to your data columns  (6 Map Fields tab)\n"
                "5. Export the filled PDFs  (8 Export PDF tab)"
            ),
            color=palette.get("text", (0.93, 0.96, 1.0, 1)),
            size_hint_y=None,
            halign="left",
            valign="top",
            font_size=dp(11.0 if is_mobile else 11.6),
        )
        steps.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(steps, min_height=dp(128 if is_mobile else 116), extra_pad=dp(8))

        helper = Label(
            text=(
                "Open a PDF to get started. This guide will disappear once a form is loaded."
                if is_mobile else
                "Load a PDF form to replace this guide with the live page preview."
            ),
            color=palette.get("accent", (0.96, 0.71, 0.30, 1)),
            size_hint_y=None,
            halign="left",
            valign="middle",
            font_size=dp(10.0 if is_mobile else 10.6),
        )
        helper.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(helper, min_height=dp(20), extra_pad=dp(4))

        outer.add_widget(hero)
        outer.add_widget(steps)
        outer.add_widget(helper)
        outer._expanded_height = max(outer.height, dp(168 if is_mobile else 156))
        return outer

    def _refresh_preview_empty_hint(self, *_):
        hint = getattr(self, "preview_empty_hint", None)
        if hint is None:
            return
        has_pdf = bool(getattr(self.engine, "pdf_path", ""))
        show = not has_pdf
        if show:
            hint.disabled = False
            hint.opacity = 1
            expanded = getattr(hint, "_expanded_height", 0) or hint.minimum_height or hint.height or dp(170)
            hint.height = expanded
        else:
            hint.opacity = 0
            hint.disabled = True
            hint.height = 0
        info = getattr(self, "preview_info", None)
        if info is not None:
            info.text = ""
            info.opacity = 0
            info.disabled = True
            info.height = 0

    def _build_startup_presplash(self):
        # Route B startup: disable the in-app splash completely so the app never falls back
        # to assets/presplash.* or assets/icon.png during startup.
        return None

    def _start_presplash_fade(self, *_):
        self._startup_presplash = None
        return

    def _finish_presplash_fade(self, *_):
        self._startup_presplash = None
        return

    def _hide_android_loading_screen(self, *_):
        if platform != "android":
            return
        try:
            from android import loadingscreen
        except Exception:
            return
        try:
            loadingscreen.hide_loading_screen()
        except Exception:
            pass
        try:
            Window.canvas.ask_update()
        except Exception:
            pass

    # --------------------------------------------------------
    # UI helpers
    # --------------------------------------------------------
    def _sync_label_text_size(self, instance, value):
        instance.text_size = value

    def _bind_auto_height_label(self, instance, min_height=0, extra_pad=0):
        def _upd(*_):
            try:
                instance.text_size = (max(1, instance.width), None)
            except Exception:
                return
            try:
                instance.texture_update()
                tex_h = instance.texture_size[1]
            except Exception:
                tex_h = 0
            try:
                instance.height = max(min_height, tex_h + extra_pad)
            except Exception:
                pass
        try:
            instance.bind(width=_upd, text=_upd)
        except Exception:
            pass
        Clock.schedule_once(lambda dt: _upd(), 0)
        return instance

    def _sync_preview_stack_size(self, *_):
        stack = getattr(self, "preview_stack", None)
        preview = getattr(self, "preview", None)
        wrap = getattr(self, "preview_wrap", None)
        if stack is None or preview is None or wrap is None:
            return
        info = getattr(self, "preview_info", None)
        info_h = 0.0
        info_spacing = float(getattr(stack, "spacing", 0) or 0)
        if info is not None and float(getattr(info, "opacity", 1) or 0) > 0 and float(getattr(info, "height", 0) or 0) > 0:
            info_h = float(getattr(info, "height", 0) or 0)
        wrap_w = max(float(getattr(wrap, "width", 0) or 0), 1.0)
        preview_w = float(getattr(preview, "width", 0) or 0)
        preview_h = float(getattr(preview, "height", 0) or 0)
        stack.size_hint_x = None
        stack.size_hint_y = None
        stack.width = max(dp(1), wrap_w, preview_w)
        stack.height = max(dp(1), preview_h + info_h + (info_spacing if info_h > 0 else 0.0))

    def _clamp_overlay_panel_to_preview_stage(self, panel, *_):
        stage = getattr(self, "preview_stage", None)
        if panel is None or stage is None:
            return
        try:
            max_x = max(0.0, float(stage.width) - float(panel.width))
            max_y = max(0.0, float(stage.height) - float(panel.height))
            clamped_x = min(max(0.0, float(panel.x)), max_x)
            clamped_y = min(max(0.0, float(panel.y)), max_y)
            if abs(float(panel.x) - clamped_x) > 0.01:
                panel.x = clamped_x
            if abs(float(panel.y) - clamped_y) > 0.01:
                panel.y = clamped_y
            panel.x = min(max(0.0, float(panel.x)), max_x)
            panel.y = min(max(0.0, float(panel.y)), max_y)
        except Exception:
            pass

    def _bind_overlay_panel_bounds(self, panel):
        stage = getattr(self, "preview_stage", None)
        if panel is None or stage is None:
            return
        try:
            panel.bind(size=lambda *_: self._clamp_overlay_panel_to_preview_stage(panel))
            panel.bind(pos=lambda *_: self._clamp_overlay_panel_to_preview_stage(panel))
            stage.bind(size=lambda *_: self._clamp_overlay_panel_to_preview_stage(panel))
            Clock.schedule_once(lambda dt: self._clamp_overlay_panel_to_preview_stage(panel), 0)
        except Exception:
            pass

    def _on_preview_viewport_change(self, *_):
        Clock.unschedule(self._refresh_preview_for_viewport)
        Clock.schedule_once(self._refresh_preview_for_viewport, 0)

    def _refresh_preview_for_viewport(self, *_):
        preview = getattr(self, "preview", None)
        if preview is None:
            return
        tex = getattr(preview, "texture", None)
        if tex is not None:
            self._update_preview_size(preview, tex)
            self._refresh_preview_boxes_for_current_view()
        self._sync_preview_stack_size()
        if getattr(self, "ui_mobile", False):
            wrap = getattr(self, "preview_wrap", None)
            mode = str(getattr(self, "preview_zoom_mode", "fit_width") or "fit_width")
            if tex is not None and wrap is not None and bool(getattr(self, "_mobile_zoom_bootstrap_pending", False)):
                wrap_w = float(getattr(wrap, "width", 0) or 0)
                preview_w = float(getattr(preview, "width", 0) or 0)
                bootstrap_fit = (mode != "manual")
                if wrap_w > dp(220) and preview_w > 0 and preview_w < (wrap_w * 0.94):
                    bootstrap_fit = True
                if bootstrap_fit:
                    self.preview_zoom_mode = "fit_width"
                    self._update_preview_size(preview, tex)
                    self._sync_preview_stack_size()
                self._mobile_zoom_bootstrap_pending = False
            if wrap is not None and str(getattr(self, "preview_zoom_mode", "fit_width") or "fit_width") in ("fit_width", "fit_page"):
                wrap.scroll_x = 0.0
                wrap.scroll_y = 1.0

    def _update_preview_size(self, instance, texture):
        if not texture:
            return

        shell = getattr(self, "preview_shell", None)
        wrap = getattr(self, "preview_wrap", None)
        is_mobile = bool(getattr(self, "ui_mobile", False))
        if is_mobile:
            mode = getattr(self, "preview_zoom_mode", "fit_width")
            zoom = float(getattr(self, "preview_zoom_factor", 1.0) or 1.0)
            base_w = float(getattr(wrap, "width", 0) or 0) if wrap is not None else 0.0
            base_h = float(getattr(wrap, "height", 0) or 0) if wrap is not None else 0.0
            if base_w <= 1:
                base_w = float(getattr(shell, "width", 0) or 0) if shell is not None else float(Window.width)
            if base_h <= 1:
                base_h = float(getattr(shell, "height", 0) or 0) if shell is not None else float(Window.height * 0.76)
            avail_w = max(dp(280), base_w)
            avail_h = max(dp(420), base_h)
        else:
            mode = getattr(self, "desktop_zoom_mode", "manual")
            zoom = float(getattr(self, "desktop_zoom_factor", 1.0) or 1.0)
            base_w = float(getattr(wrap, "width", 0) or 0) if wrap is not None else 0.0
            base_h = float(getattr(wrap, "height", 0) or 0) if wrap is not None else 0.0
            if base_w <= 1:
                base_w = float(getattr(shell, "width", 0) or 0) if shell is not None else float(Window.width * 0.58)
            if base_h <= 1:
                base_h = float(getattr(shell, "height", 0) or 0) if shell is not None else float(Window.height - dp(240))
            avail_w = max(dp(260), base_w - dp(16))
            avail_h = max(dp(260), base_h - dp(22))

        if mode == "fit_width":
            scale = avail_w / float(max(texture.width, 1))
        elif mode == "fit_page":
            scale = min(avail_w / float(max(texture.width, 1)), avail_h / float(max(texture.height, 1)))
        else:
            scale = zoom

        max_scale = 8.0 if is_mobile else 6.0
        min_scale = 0.35 if is_mobile else 0.2
        scale = max(min_scale, min(float(scale), max_scale))
        self.preview.display_scale = scale
        self.preview.size = (max(dp(1), texture.width * scale), max(dp(1), texture.height * scale))

        if is_mobile:
            self.preview_zoom_factor = scale
            if hasattr(self, "mobile_zoom_chip_lbl") and self.mobile_zoom_chip_lbl is not None:
                self.mobile_zoom_chip_lbl.text = ("Fit" if mode != "manual" else f"{int(round(scale * 100))}%")
        else:
            self.desktop_zoom_factor = scale
            if hasattr(self, "zoom_chip_lbl") and self.zoom_chip_lbl is not None:
                self.zoom_chip_lbl.text = f"{int(round(scale * 100))}%"
        self._sync_preview_stack_size()

    def _apply_preview_zoom(self, mode=None, scale=None, anchor=None, anchor_window_point=None, schedule_resync=True):
        target_mode_attr = "preview_zoom_mode" if getattr(self, "ui_mobile", False) else "desktop_zoom_mode"
        target_scale_attr = "preview_zoom_factor" if getattr(self, "ui_mobile", False) else "desktop_zoom_factor"
        if getattr(self, "ui_mobile", False):
            self._mobile_zoom_bootstrap_pending = False
        if anchor_window_point is not None:
            try:
                anchor = self._capture_preview_anchor_at_window_point(float(anchor_window_point[0]), float(anchor_window_point[1]))
            except Exception:
                anchor = self._capture_preview_anchor()
        elif anchor is None:
            anchor = self._capture_preview_anchor()
        if mode is not None:
            setattr(self, target_mode_attr, mode)
        if scale is not None:
            max_scale = 8.0 if getattr(self, "ui_mobile", False) else 6.0
            min_scale = 0.35 if getattr(self, "ui_mobile", False) else 0.2
            setattr(self, target_scale_attr, max(min_scale, min(float(scale), max_scale)))
        tex = getattr(getattr(self, "preview", None), "texture", None)
        if tex is not None:
            self._update_preview_size(self.preview, tex)
            Clock.schedule_once(lambda dt: self._post_preview_refresh(reset_scroll=False), 0)
            if anchor_window_point is not None:
                Clock.schedule_once(lambda dt: self._restore_preview_anchor_at_window_point(anchor, float(anchor_window_point[0]), float(anchor_window_point[1])), 0)
            else:
                Clock.schedule_once(lambda dt: self._restore_preview_anchor(anchor), 0)
            if schedule_resync:
                self._schedule_preview_resync()

    def _zoom_preview(self, direction):
        current_attr = "preview_zoom_factor" if getattr(self, "ui_mobile", False) else "desktop_zoom_factor"
        current = float(getattr(self, current_attr, 1.0) or 1.0)
        step = 1.18 if getattr(self, "ui_mobile", False) else 1.15
        if direction == "in":
            current *= step
        elif direction == "out":
            current /= step
        else:
            current = 1.0
        self._apply_preview_zoom(mode="manual", scale=current)

    def _fit_preview_width(self, *_):
        self._apply_preview_zoom(mode="fit_width")

    def _fit_preview_page(self, *_):
        self._apply_preview_zoom(mode="fit_page")

    def _update_selection_inspector(self):
        ids = sorted(set(int(x) for x in getattr(self.engine, "selected_box_ids", []) if isinstance(x, int) or str(x).isdigit()))
        if hasattr(self, "inspector_selected_lbl") and self.inspector_selected_lbl is not None:
            self.inspector_selected_lbl.text = (", ".join(str(i) for i in ids) if ids else "None")
        if hasattr(self, "inspector_count_lbl") and self.inspector_count_lbl is not None:
            self.inspector_count_lbl.text = str(len(ids))

        if not ids:
            box_text = "No selection"
            mapping_text = "Choose a box in the preview to select or map it."
        else:
            first = ids[0]
            box_type = self.engine.box_types[first] if first < len(self.engine.box_types) else "field"
            rect = self.engine.all_boxes[first] if first < len(self.engine.all_boxes) else None
            rect_text = ""
            if rect is not None:
                rect_text = f" | ({int(rect.x0)}, {int(rect.y0)}) → ({int(rect.x1)}, {int(rect.y1)})"
            box_text = f"Box {first} • {box_type}{rect_text}"
            mapping = self.engine.describe_box_mapping(first, self.current_page_idx())
            mapping_text = mapping if mapping and mapping != "EMPTY" else "Unmapped"

        if hasattr(self, "inspector_box_lbl") and self.inspector_box_lbl is not None:
            self.inspector_box_lbl.text = box_text
        if hasattr(self, "inspector_mapping_lbl") and self.inspector_mapping_lbl is not None:
            self.inspector_mapping_lbl.text = mapping_text
        if hasattr(self, "inspector_resolution_lbl") and self.inspector_resolution_lbl is not None:
            if ids:
                first = ids[0]
                mapping = self._find_mapping_for_box(first, self.current_page_idx()) or {}
                self.inspector_resolution_lbl.text = "Resolved preview value: " + self._get_selected_column_preview(
                    mapping.get("column", self.column_spinner.text if hasattr(self, "column_spinner") else ""),
                    trigger=mapping.get("trigger", self.trigger_input.text if hasattr(self, "trigger_input") else ""),
                    is_grid=bool(mapping.get("g", getattr(self.grid_flag_chk, "active", False) if hasattr(self, "grid_flag_chk") else False)),
                    grid_n=int(mapping.get("n", self.grid_n_input.text if hasattr(self, "grid_n_input") else "1") or 1),
                )
            else:
                self.inspector_resolution_lbl.text = "Resolved preview value: —"
        self._update_preview_hud()

    def _update_bottom_statusbar(self):
        pdf_name = os.path.basename(getattr(self.engine, "pdf_path", "") or "") or "No PDF"
        try:
            page_idx = self.current_page_idx()
        except Exception:
            page_idx = 0
        total_pages = max(int(self.engine.total_pages()), 0)
        patient = self.selected_patient() or "No record"
        box_count = len(getattr(self.engine, "all_boxes", []) or [])
        sel_count = len(getattr(self.engine, "selected_box_ids", []) or [])
        ready = []
        if getattr(self.engine, "pdf_path", ""):
            ready.append("PDF")
        if getattr(self.engine, "df", None) is not None and not self.engine.df.empty:
            ready.append("DATA")
        if self.engine.supports_detection_backend():
            ready.append("DETECT")
        if self.engine.supports_export_backend():
            ready.append("EXPORT")
        readiness = " • ".join(ready) if ready else "Idle"
        if hasattr(self, "statusbar_file_lbl") and self.statusbar_file_lbl is not None:
            self.statusbar_file_lbl.text = f"File: {pdf_name}"
        if hasattr(self, "statusbar_page_lbl") and self.statusbar_page_lbl is not None:
            self.statusbar_page_lbl.text = f"Page: {page_idx + 1}/{max(total_pages, 1)}"
        if hasattr(self, "statusbar_patient_lbl") and self.statusbar_patient_lbl is not None:
            self.statusbar_patient_lbl.text = f"Record: {patient}"
        if hasattr(self, "statusbar_boxes_lbl") and self.statusbar_boxes_lbl is not None:
            self.statusbar_boxes_lbl.text = f"Boxes: {box_count} • Selected: {sel_count}"
        if hasattr(self, "statusbar_ready_lbl") and self.statusbar_ready_lbl is not None:
            self.statusbar_ready_lbl.text = f"Ready: {readiness}"
        if hasattr(self, "mobile_meta_lbl") and self.mobile_meta_lbl is not None:
            self.mobile_meta_lbl.text = ""
        if getattr(self, "android_status_file_lbl", None) is not None:
            self.android_status_file_lbl.text = f"File: {pdf_name}"
        if getattr(self, "android_status_page_lbl", None) is not None:
            self.android_status_page_lbl.text = f"Page: {page_idx + 1}/{max(total_pages, 1)}"
        if getattr(self, "android_status_boxes_lbl", None) is not None:
            self.android_status_boxes_lbl.text = f"Boxes: {box_count} • Sel: {sel_count}"
        if getattr(self, "android_status_mode_lbl", None) is not None:
            mode_txt = str(getattr(self, "_mobile_active_section", "Files") or "Files")
            self.android_status_mode_lbl.text = f"Mode: {mode_txt}"

    def _bind_desktop_shortcuts(self):
        if platform == "android":
            return
        try:
            Window.unbind(on_key_down=self._on_window_key_down)
        except Exception:
            pass
        try:
            Window.bind(on_key_down=self._on_window_key_down)
        except Exception:
            pass

    def _on_window_key_down(self, _window, key, _scancode, _codepoint, modifiers):
        if platform == "android":
            return False
        modifiers = modifiers or []
        ctrl = "ctrl" in modifiers or "meta" in modifiers
        try:
            if ctrl and key in (111, ord('o'), ord('O')):
                self.on_load_pdf(None)
                return True
            if ctrl and key in (115, ord('s'), ord('S')):
                self.on_save_config(None)
                return True
            if key in (276, 80):
                self.on_prev_page(None)
                return True
            if key in (275, 79):
                self.on_next_page(None)
                return True
            if key in (13, 271):
                self.on_run_detect(None)
                return True
            if key in (127, 8):
                self.on_clear_selected_mapping()
                return True
            if key in (45, 269):
                self._zoom_preview("out")
                return True
            if key in (61, 43, 270):
                self._zoom_preview("in")
                return True
            digit_map = {
                ord('1'): 'Workspace', ord('2'): 'Session', ord('3'): 'Detection',
                ord('4'): 'Selection', ord('5'): 'Mapping', ord('6'): 'Export'
            }
            if key in digit_map:
                target = digit_map[key]
                if target in getattr(self, '_desktop_section_cards', {}):
                    self._show_desktop_section(target)
                elif target in getattr(self, '_desktop_right_section_cards', {}):
                    self._show_desktop_right_section(target)
                return True
        except Exception:
            return False
        return False

    def _post_preview_refresh(self, *args, reset_scroll=True):
        try:
            tex = getattr(self.preview, "texture", None)
            if tex is not None:
                self._update_preview_size(self.preview, tex)
                self._refresh_preview_boxes_for_current_view()
            self._sync_preview_stack_size()
            self.preview._redraw_overlay()
            self.preview.canvas.ask_update()
            if reset_scroll and getattr(self, "preview_wrap", None) is not None:
                self.preview_wrap.scroll_x = 0
                self.preview_wrap.scroll_y = 1
            elif getattr(self, "ui_mobile", False) and getattr(self, "preview_wrap", None) is not None:
                mode = str(getattr(self, "preview_zoom_mode", "fit_width") or "fit_width")
                if mode in ("fit_width", "fit_page"):
                    self.preview_wrap.scroll_x = 0
            self._update_preview_hud()
            if getattr(self, "ui_mobile", False):
                Clock.schedule_once(lambda dt: self._refresh_preview_for_viewport(), 0.05)
                Clock.schedule_once(lambda dt: self._refresh_preview_for_viewport(), 0.18)
        except Exception:
            pass

    def _short_learning_path(self, path, max_len=82):
        txt = str(path or "").strip()
        if not txt:
            return "—"
        if len(txt) <= max_len:
            return txt
        keep = max(18, (max_len - 3) // 2)
        return f"{txt[:keep]}...{txt[-keep:]}"

    def refresh_learning_ui(self, *_):
        if not hasattr(self, "engine") or self.engine is None:
            return
        if not hasattr(self, "learning_summary_lbl"):
            return
        try:
            profiles = ((self.engine.profile_memory or {}).get("profiles", {}) or {})
            revision_index = self.engine.learning_storage.load_revision_index() if getattr(self.engine, "learning_storage", None) else {"items": []}
            rev_items = revision_index.get("items", []) if isinstance(revision_index, dict) else []
            corrections = ((self.engine.correction_log or {}).get("events", []) or [])
            ctx = dict(getattr(self.engine, "current_detection_context", {}) or {})
            fp = dict(ctx.get("fingerprint", {}) or {})
            match = dict(ctx.get("matched_profile", {}) or {})
            page_idx = int(ctx.get("page_idx", getattr(self.engine, "detected_page_idx", 0) or 0) or 0)
            box_count = len(getattr(self.engine, "all_boxes", []) or [])
            enabled = bool(getattr(self.engine, "learning_enabled", False))
            self.learning_summary_lbl.text = (
                f"Profiles: {len(profiles)} • Revisions: {len(rev_items)} • Corrections: {len(corrections)}\n"
                f"Page: {page_idx} • Boxes: {box_count} • Learning: {'On' if enabled else 'Off'}"
            )

            fp_key = str(fp.get("fingerprint_key", "") or "")
            fp_short = (fp_key[:12] + "…") if fp_key and len(fp_key) > 13 else (fp_key or "—")
            if match:
                kind = str(match.get("match_kind", "fuzzy") or "fuzzy").title()
                score = float(match.get("match_score", 0.0) or 0.0)
                applied = bool(ctx.get("profile_applied", False))
                approved = bool(match.get("approved", False))
                self.learning_match_lbl.text = (
                    f"Match: {kind} • Score: {score:.2f} • Applied: {'Yes' if applied else 'No'}\n"
                    f"Approved profile: {'Yes' if approved else 'No'} • Fingerprint: {fp_short}"
                )
            else:
                self.learning_match_lbl.text = f"Match: none yet\nFingerprint: {fp_short}"

            current_rev = str(getattr(self.engine, "current_revision_id", "") or "")
            last_saved = str(getattr(self.engine, "last_saved_revision_id", "") or "")
            approved_rev = None
            if current_rev:
                approved_rev = self.engine.learning_storage.load_revision(current_rev)
            if not approved_rev and last_saved:
                approved_rev = self.engine.learning_storage.load_revision(last_saved)
            approved_flag = bool((approved_rev or {}).get("approved", False)) if isinstance(approved_rev, dict) else False
            self.learning_revision_lbl.text = (
                f"Current: {current_rev or 'none'}\n"
                f"Last saved: {last_saved or 'none'} • Approved: {'Yes' if approved_flag else 'No'}"
            )

            recent_bits = []
            for item in list(reversed(rev_items))[:3]:
                rid = str(item.get("revision_id", "") or "")
                rid_short = (rid[:18] + "…") if len(rid) > 19 else rid
                reason = str(item.get("reason", "manual") or "manual")
                page = int(item.get("page_idx", 0) or 0)
                approved_txt = " ✓" if bool(item.get("approved", False)) else ""
                recent_bits.append(f"{rid_short} • p{page} • {reason}{approved_txt}")
            self.learning_recent_lbl.text = "\n".join(recent_bits) if recent_bits else "Recent revisions: none"

            self.learning_storage_lbl.text = (
                f"Store: {self._short_learning_path(getattr(self.engine.learning_storage, 'base_dir', ''))}\n"
                f"PDF: {os.path.basename(getattr(self.engine, 'pdf_path', '') or '') or 'None loaded'}"
            )
        except Exception:
            pass

    def on_learning_refresh(self, instance):
        self.refresh_learning_ui()
        profiles = len(((self.engine.profile_memory or {}).get("profiles", {}) or {})) if hasattr(self, "engine") else 0
        self.set_status(f"Learning state refreshed.\nProfiles: {profiles}")

    def on_learning_reload_memory(self, instance):
        try:
            self.engine.profile_memory = self.engine.learning_storage.load_profile_memory()
            self.engine.template_stats = self.engine.learning_storage.load_template_stats()
            self.engine.correction_log = self.engine.learning_storage.load_correction_log()
            self.refresh_learning_ui()
            self.set_status(
                f"Learning memory reloaded.\n"
                f"Profiles: {len((self.engine.profile_memory or {}).get('profiles', {}) or {})}"
            )
        except Exception as e:
            self.set_status(f"Reload learning error:\n{e}")

    def on_learning_save_revision(self, instance):
        try:
            if not getattr(self.engine, "pdf_path", ""):
                self.set_status("Load PDF first.", kind="warning", hold_seconds=1.5, force=True)
                return
            self.apply_ui_settings_to_engine()
            self.engine._config_ui_state = self._collect_runtime_config_state()
            revision = self.engine.save_learning_revision(
                reason="learning_ui_save",
                approved=False,
                learning_weight=0.45,
            ) or {}
            self.refresh_learning_ui()
            self.set_status(
                f"Learning revision saved.\n"
                f"Revision: {revision.get('revision_id', 'n/a')}"
            )
        except Exception as e:
            self.set_status(f"Save revision error:\n{e}")

    def on_learning_approve_current(self, instance):
        try:
            if not getattr(self.engine, "pdf_path", ""):
                self.set_status("Load PDF first.")
                return
            revision_id = str(getattr(self.engine, "current_revision_id", "") or getattr(self.engine, "last_saved_revision_id", "") or "").strip()
            revision = None
            if revision_id:
                revision = self.engine.approve_revision_for_learning(revision_id=revision_id, learning_weight=1.0)
            else:
                self.apply_ui_settings_to_engine()
                self.engine._config_ui_state = self._collect_runtime_config_state()
                revision = self.engine.save_learning_revision(
                    reason="learning_ui_approve",
                    approved=True,
                    learning_weight=1.0,
                )
            if not isinstance(revision, dict):
                self.set_status("No revision available to approve yet.")
                return
            self.engine.current_revision_id = str(revision.get("revision_id", self.engine.current_revision_id or "") or "")
            self.engine.last_saved_revision_id = str(revision.get("revision_id", self.engine.last_saved_revision_id or "") or "")
            self.engine.persist_learning_session(current_page_idx=self.current_page_idx())
            self.refresh_learning_ui()
            self.set_status(
                f"Revision approved for learning.\n"
                f"Revision: {revision.get('revision_id', 'n/a')}"
            )
        except Exception as e:
            self.set_status(f"Approve revision error:\n{e}")

    def on_learning_apply_best_profile(self, instance):
        try:
            if not getattr(self.engine, "pdf_path", ""):
                self.set_status("Load PDF first.")
                return
            page_idx = int(self.current_page_idx())
            img = self.engine._render_pdf_page_bgr(self.engine.pdf_path, page_idx=page_idx, preview_zoom=ZOOM)
            fingerprint = self.engine.build_page_fingerprint(page_idx=page_idx, img_bgr=img)
            profile = self.engine.find_best_profile_match(fingerprint)
            if not isinstance(profile, dict) or not isinstance(profile.get("settings"), dict):
                self.set_status("No learned profile available for this page yet.")
                return
            match_score = float(profile.get("match_score", 0.0) or 0.0)
            match_kind = str(profile.get("match_kind", "fuzzy") or "fuzzy")
            if match_kind != "exact" and match_score < 0.35:
                self.set_status(f"Best learned profile is too weak to apply.\nScore: {match_score:.2f}")
                return
            self.engine._apply_learning_profile_settings(profile.get("settings", {}))
            self.engine.last_applied_profile_meta = {
                "profile_key": profile.get("fingerprint_key", ""),
                "match_kind": match_kind,
                "match_score": match_score,
                "applied_at": _utc_now_iso(),
                "explicit_ui_apply": True,
            }
            ctx = dict(getattr(self.engine, "current_detection_context", {}) or {})
            ctx.update({
                "page_idx": page_idx,
                "fingerprint": fingerprint,
                "matched_profile": dict(profile),
                "profile_applied": True,
            })
            self.engine.current_detection_context = ctx
            self.engine.persist_learning_session(current_page_idx=page_idx)
            self.push_engine_settings_to_ui()
            self.refresh_learning_ui()
            self.set_status(
                f"Applied learned profile.\n"
                f"{match_kind.title()} match • Score {match_score:.2f}"
            )
        except Exception as e:
            self.set_status(f"Apply learned profile error:\n{e}")

    def _revision_popup_body(self, revision):
        palette = getattr(self, "ui_palette", {}) or {}
        data = dict(revision or {})
        delta = dict(data.get("learning_delta", {}) or {})
        fp = dict(data.get("fingerprint", {}) or {})
        body = BoxLayout(orientation="vertical", spacing=dp(8))
        scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, bar_width=dp(5), scroll_type=["bars", "content"])
        content = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        sections = [
            ("Revision", [
                f"ID: {data.get('revision_id', '—')}",
                f"Reason: {data.get('reason', 'manual')}",
                f"Saved: {data.get('saved_at', '—')}",
                f"Approved: {'Yes' if data.get('approved', False) else 'No'}",
                f"Trust: {data.get('trust_class', 'draft')}",
                f"Learning weight: {float(data.get('learning_weight', 0.0) or 0.0):.2f}",
            ]),
            ("Page / file", [
                f"PDF: {os.path.basename(str(data.get('pdf_path', '') or '')) or '—'}",
                f"Page: {int(data.get('page_idx', 0) or 0)}",
                f"Family: {data.get('pdf_family_hint', '—')}",
                f"Fingerprint: {fp.get('fingerprint_key', '—') or '—'}",
            ]),
            ("Delta", [
                f"Raw boxes: {int(delta.get('raw_count', 0) or 0)}",
                f"Final boxes: {int(delta.get('final_count', 0) or 0)}",
                f"Matched stable: {int(delta.get('stable_count', 0) or 0)} / {int(delta.get('matched_count', 0) or 0)}",
                f"Added: {int(delta.get('added_count', 0) or 0)}",
                f"Removed: {int(delta.get('removed_count', 0) or 0)}",
                f"Retyped: {int(delta.get('retyped_count', 0) or 0)}",
                f"Resized: {int(delta.get('resized_count', 0) or 0)}",
                f"Moved: {int(delta.get('moved_count', 0) or 0)}",
                f"Quality score: {float(delta.get('quality_score', 0.0) or 0.0):.3f}",
            ]),
        ]
        matched = data.get("matched_profile") if isinstance(data.get("matched_profile"), dict) else {}
        settings_keys = sorted(list((data.get("profile_used", {}) or {}).keys()))
        sections.append(("Profile", [
            f"Applied: {'Yes' if data.get('profile_applied', False) else 'No'}",
            f"Matched profile key: {matched.get('fingerprint_key', '—') or '—'}",
            f"Settings keys: {', '.join(settings_keys[:8]) or '—'}",
        ]))
        recent_events = list((delta.get('events', []) or []))[:8]
        if recent_events:
            sections.append(("Detected changes", [
                f"{str(evt.get('event_type', 'change')).replace('_', ' ').title()}: {evt.get('note', '—') or '—'}"
                for evt in recent_events
            ]))
        for title, lines in sections:
            card = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(10), dp(10), dp(10), dp(10)], size_hint_y=None)
            card.bind(minimum_height=card.setter("height"))
            self._style_popup_card(card, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(16))
            ttl = Label(text=title, color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(20), halign="left", valign="middle", font_size=dp(12.5), bold=True)
            ttl.bind(size=self._sync_label_text_size)
            card.add_widget(ttl)
            for line in lines:
                lbl = Label(text=str(line), color=palette.get("muted", (0.60, 0.68, 0.80, 1)), size_hint_y=None, height=dp(20), halign="left", valign="middle", font_size=dp(10.8))
                lbl.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(lbl, min_height=dp(18), extra_pad=dp(2))
                card.add_widget(lbl)
            content.add_widget(card)
        scroll.add_widget(content)
        body.add_widget(scroll)
        return body

    def _build_revision_compare_payload(self, base_revision, other_revision):
        base = dict(base_revision or {})
        other = dict(other_revision or {})
        base_final = dict(base.get("final_detection", {}) or {})
        other_final = dict(other.get("final_detection", {}) or {})
        delta = self.engine._compute_learning_delta(base_final, other_final)
        return {
            "base_revision": base,
            "other_revision": other,
            "delta": delta,
        }

    def _revision_compare_popup_body(self, base_revision, other_revision):
        payload = self._build_revision_compare_payload(base_revision, other_revision)
        base = payload.get("base_revision", {})
        other = payload.get("other_revision", {})
        delta = payload.get("delta", {})
        lines = [
            f"Base: {base.get('revision_id', '—')} • Page {int(base.get('page_idx', 0) or 0)}",
            f"Other: {other.get('revision_id', '—')} • Page {int(other.get('page_idx', 0) or 0)}",
            f"Base final boxes: {int(((base.get('final_detection', {}) or {}).get('boxes', []) or []).__len__())}",
            f"Other final boxes: {int(((other.get('final_detection', {}) or {}).get('boxes', []) or []).__len__())}",
            f"Stable: {int(delta.get('stable_count', 0) or 0)} / {int(delta.get('matched_count', 0) or 0)}",
            f"Added vs base: {int(delta.get('added_count', 0) or 0)}",
            f"Removed vs base: {int(delta.get('removed_count', 0) or 0)}",
            f"Retyped: {int(delta.get('retyped_count', 0) or 0)}",
            f"Resized: {int(delta.get('resized_count', 0) or 0)}",
            f"Moved: {int(delta.get('moved_count', 0) or 0)}",
            f"Similarity quality: {float(delta.get('quality_score', 0.0) or 0.0):.3f}",
        ]
        body = BoxLayout(orientation="vertical", spacing=dp(8))
        scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, bar_width=dp(5), scroll_type=["bars", "content"])
        content = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        palette = getattr(self, "ui_palette", {}) or {}
        card = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(10), dp(10), dp(10), dp(10)], size_hint_y=None)
        card.bind(minimum_height=card.setter("height"))
        self._style_popup_card(card, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(16))
        for line in lines:
            lbl = Label(text=str(line), color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(20), halign="left", valign="middle", font_size=dp(10.8))
            lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(lbl, min_height=dp(18), extra_pad=dp(2))
            card.add_widget(lbl)
        content.add_widget(card)
        event_lines = [f"{str(evt.get('event_type', 'change')).replace('_', ' ').title()}: {evt.get('note', '—') or '—'}" for evt in list(delta.get('events', []) or [])[:10]]
        if event_lines:
            card2 = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(10), dp(10), dp(10), dp(10)], size_hint_y=None)
            card2.bind(minimum_height=card2.setter("height"))
            self._style_popup_card(card2, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(16))
            ttl = Label(text="Top changes", color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(20), halign="left", valign="middle", font_size=dp(12.5), bold=True)
            ttl.bind(size=self._sync_label_text_size)
            card2.add_widget(ttl)
            for line in event_lines:
                lbl = Label(text=str(line), color=palette.get("muted", (0.60, 0.68, 0.80, 1)), size_hint_y=None, height=dp(20), halign="left", valign="middle", font_size=dp(10.8))
                lbl.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(lbl, min_height=dp(18), extra_pad=dp(2))
                card2.add_widget(lbl)
            content.add_widget(card2)
        scroll.add_widget(content)
        body.add_widget(scroll)
        return body

    def _open_revision_compare_popup(self, other_revision_id, browser_popup=None):
        current_id = str(getattr(self.engine, "current_revision_id", "") or getattr(self.engine, "last_saved_revision_id", "") or "").strip()
        if not current_id:
            self.set_status("Set or save a current revision first before comparing.")
            return
        base_revision = self.engine.learning_storage.load_revision(current_id)
        other_revision = self.engine.learning_storage.load_revision(other_revision_id)
        if not isinstance(base_revision, dict) or not isinstance(other_revision, dict):
            self.set_status("Revision compare data could not be loaded.")
            return
        popup, _outer = self._make_styled_popup(
            "Revision Diff",
            self._revision_compare_popup_body(base_revision, other_revision),
            subtitle="Compare another revision against the current revision",
            size_hint=(0.95 if getattr(self, "ui_mobile", False) else 0.72, 0.90 if getattr(self, "ui_mobile", False) else 0.84),
        )
        popup.open()

    def on_learning_view_current_revision(self, instance):
        revision_id = str(getattr(self.engine, "current_revision_id", "") or getattr(self.engine, "last_saved_revision_id", "") or "").strip()
        if not revision_id:
            self.set_status("No revision available to inspect yet.")
            return
        revision = self.engine.learning_storage.load_revision(revision_id)
        if not isinstance(revision, dict):
            self.set_status("Revision details could not be loaded.")
            return
        popup, _outer = self._make_styled_popup(
            "Current Revision",
            self._revision_popup_body(revision),
            subtitle="Saved learning snapshot details",
            size_hint=(0.94 if getattr(self, "ui_mobile", False) else 0.66, 0.86 if getattr(self, "ui_mobile", False) else 0.80),
        )
        popup.open()

    def on_learning_browse_revisions(self, instance):
        revision_index = self.engine.learning_storage.load_revision_index() if getattr(self.engine, "learning_storage", None) else {"items": []}
        items = list(revision_index.get("items", []) or [])
        if not items:
            self.set_status("No saved revisions yet.")
            return
        palette = getattr(self, "ui_palette", {}) or {}
        body = BoxLayout(orientation="vertical", spacing=dp(8))
        scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, bar_width=dp(5), scroll_type=["bars", "content"])
        content = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        popup_holder = {"popup": None}
        for item in list(reversed(items))[:24]:
            rid = str(item.get("revision_id", "") or "")
            reason = str(item.get("reason", "manual") or "manual")
            page = int(item.get("page_idx", 0) or 0)
            approved = bool(item.get("approved", False))
            saved_at = str(item.get("saved_at", "") or "")
            row = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(10), dp(10), dp(10), dp(10)], size_hint_y=None)
            row.bind(minimum_height=row.setter("height"))
            self._style_popup_card(row, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(16))
            summary = Label(
                text=f"{rid}\nPage {page} • {reason} • {'Approved' if approved else 'Draft'} • {str(item.get('trust_class', 'draft'))}\n{saved_at}",
                color=palette.get("text", (0.93, 0.96, 1.0, 1)),
                size_hint_y=None,
                height=dp(52),
                halign="left",
                valign="middle",
                font_size=dp(10.8),
            )
            summary.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(summary, min_height=dp(42), extra_pad=dp(2))
            btn_row = GridLayout(cols=3, spacing=dp(8), size_hint_y=None, height=dp(44))
            btn_view = self._make_compact_action_button("View", tone="plain")
            btn_compare = self._make_compact_action_button("Compare", tone="soft")
            btn_use = self._make_compact_action_button("Use as Current", tone="soft")
            btn_row.add_widget(btn_view)
            btn_row.add_widget(btn_compare)
            btn_row.add_widget(btn_use)
            def _view_factory(rev_id):
                return lambda *_: self._open_revision_from_browser(rev_id, popup_holder.get("popup"))
            def _compare_factory(rev_id):
                return lambda *_: self._open_revision_compare_popup(rev_id, popup_holder.get("popup"))
            def _use_factory(rev_id):
                return lambda *_: self._select_current_revision_from_browser(rev_id, popup_holder.get("popup"))
            btn_view.bind(on_release=_view_factory(rid))
            btn_compare.bind(on_release=_compare_factory(rid))
            btn_use.bind(on_release=_use_factory(rid))
            current_id = str(getattr(self.engine, "current_revision_id", "") or getattr(self.engine, "last_saved_revision_id", "") or "").strip()
            if not current_id or current_id == rid:
                btn_compare.disabled = True
            row.add_widget(summary)
            row.add_widget(btn_row)
            content.add_widget(row)
        scroll.add_widget(content)
        body.add_widget(scroll)
        popup, _outer = self._make_styled_popup(
            "Revision Browser",
            body,
            subtitle="Recent saved learning revisions • compare against current",
            size_hint=(0.95 if getattr(self, "ui_mobile", False) else 0.72, 0.90 if getattr(self, "ui_mobile", False) else 0.84),
        )
        popup_holder["popup"] = popup
        popup.open()

    def _open_revision_from_browser(self, revision_id, browser_popup=None):
        if browser_popup is not None:
            try:
                browser_popup.dismiss()
            except Exception:
                pass
        revision = self.engine.learning_storage.load_revision(revision_id)
        if not isinstance(revision, dict):
            self.set_status("Revision details could not be loaded.")
            return
        popup, _outer = self._make_styled_popup(
            "Revision Details",
            self._revision_popup_body(revision),
            subtitle="Saved learning snapshot details",
            size_hint=(0.94 if getattr(self, "ui_mobile", False) else 0.66, 0.86 if getattr(self, "ui_mobile", False) else 0.80),
        )
        popup.open()

    def _select_current_revision_from_browser(self, revision_id, browser_popup=None):
        revision = self.engine.learning_storage.load_revision(revision_id)
        if not isinstance(revision, dict):
            self.set_status("Revision could not be loaded.")
            return
        self.engine.current_revision_id = str(revision.get("revision_id", self.engine.current_revision_id or "") or "")
        self.engine.last_saved_revision_id = str(revision.get("revision_id", self.engine.last_saved_revision_id or "") or "")
        self.engine.apply_learning_meta(revision.get("learning_meta", {}) or self.engine.collect_learning_meta())
        self.engine.persist_learning_session(current_page_idx=self.current_page_idx())
        self.refresh_learning_ui()
        if browser_popup is not None:
            try:
                browser_popup.dismiss()
            except Exception:
                pass
        self.set_status(f"Current revision set to:\n{revision_id}")

    def _maybe_restore_learning_session(self, *_):
        try:
            if getattr(self, "_learning_session_restore_attempted", False):
                return
            self._learning_session_restore_attempted = True
            if not hasattr(self, "engine") or self.engine is None:
                return
            if getattr(self.engine, "pdf_path", ""):
                return
            session = self.engine.learning_storage.load_session("current_session") if getattr(self.engine, "learning_storage", None) else {}
            if not isinstance(session, dict) or not session:
                self.refresh_learning_ui()
                return
            pdf_path = str(session.get("pdf_path", "") or "").strip()
            page_idx = int(session.get("page_idx", 0) or 0)
            learning_meta = dict(session.get("learning_meta", {}) or {})
            current_ctx = dict(session.get("current_detection_context", {}) or {})
            pending_ui_state = dict(session.get("ui_state", {}) or {})
            self._pending_restored_record_label = str(session.get("selected_record_label", pending_ui_state.get("selected_record_label", "")) or "").strip()
            self._pending_restored_record_key = str(session.get("selected_record_key", pending_ui_state.get("selected_record_key", "")) or "").strip()
            if learning_meta:
                self.engine.apply_learning_meta(learning_meta)
            if current_ctx:
                self.engine.current_detection_context = current_ctx
            if not pdf_path or not os.path.exists(pdf_path):
                self.refresh_learning_ui()
                return
            if platform == "android":
                self._set_page_inputs_text(max(0, page_idx))
                self.engine.detected_page_idx = int(session.get("detected_page_idx", page_idx) or page_idx)
                self.set_status(
                    "Last session was restored safely.\n"
                    "Reopen your PDF form manually to continue."
                )
                Clock.schedule_once(lambda dt: self.refresh_learning_ui(), 0.05)
                return
            total = self.engine.load_pdf(pdf_path)
            max_idx = max(int(total) - 1, 0)
            page_idx = max(0, min(page_idx, max_idx))
            self._set_page_inputs_text(page_idx)
            self.engine.detected_page_idx = int(session.get("detected_page_idx", page_idx) or page_idx)
            self.set_status(f"Restoring last session…\n{os.path.basename(pdf_path)}")
            Clock.schedule_once(lambda dt, p=pdf_path, t=total: self._finish_pdf_load_preview(p, t), 0.05)
            Clock.schedule_once(lambda dt: self.refresh_learning_ui(), 0.08)
        except Exception as e:
            self.set_status(f"Session restore skipped:\n{e}")

    def _status_priority(self, kind):
        kind = str(kind or "info").strip().lower()
        priorities = {
            "preview": 10,
            "session": 20,
            "info": 30,
            "action": 40,
            "detect": 60,
            "warning": 65,
            "error": 80,
        }
        return priorities.get(kind, 30)

    def set_status(self, text, kind="info", hold_seconds=None, force=False):
        now = time.time()
        kind = str(kind or "info").strip().lower()
        current_kind = str(getattr(self, "_status_kind", "info") or "info").strip().lower()
        current_until = float(getattr(self, "_status_hold_until", 0.0) or 0.0)
        current_priority = self._status_priority(current_kind)
        incoming_priority = self._status_priority(kind)
        if not force and current_until > now and incoming_priority < current_priority:
            return False

        text = str(text or "")
        self._status_kind = kind
        self._status_text = text
        self._status_hold_until = now + max(0.0, float(hold_seconds or 0.0)) if hold_seconds else 0.0

        # Map kind to a visible text color so users can distinguish severity at a glance
        _palette = getattr(self, "ui_palette", {}) or {}
        _kind_color = {
            "error":   _palette.get("danger",  (1.0, 0.36, 0.56, 1)),
            "warning": _palette.get("accent",  (0.96, 0.71, 0.30, 1)),
            "action":  _palette.get("primary", (0.255, 0.50, 0.98, 1)),
            "detect":  _palette.get("aurora",  (0.255, 0.82, 0.96, 1)),
        }.get(kind, _palette.get("text", (0.97, 0.98, 1.0, 1)))

        # Thread-safety: schedule UI update on the Kivy main thread
        def _apply_status_text(dt):
            try:
                self.status_lbl.text = text
                self.status_lbl.color = _kind_color
            except Exception:
                pass
        Clock.schedule_once(_apply_status_text, 0)
        # Best-effort immediate update (safe if already on main thread)
        try:
            self.status_lbl.text = text
            self.status_lbl.color = _kind_color
        except Exception:
            pass
        if hasattr(self, "desktop_status_detail_lbl") and self.desktop_status_detail_lbl is not None:
            self.desktop_status_detail_lbl.text = text.replace("\n", " • ")[:220]
        try:
            self.refresh_backend_capabilities_ui()
            self.refresh_learning_ui()
            self._update_bottom_statusbar()
        except Exception:
            pass
        return True

    def _set_widget_enabled(self, widget, enabled=True):
        if widget is None:
            return
        try:
            widget.disabled = not bool(enabled)
        except Exception:
            pass
        try:
            widget.opacity = 1.0 if enabled else 0.45
        except Exception:
            pass

    def refresh_backend_capabilities_ui(self):
        detection_ok = bool(self.engine.supports_detection_backend())
        export_ok = bool(self.engine.supports_export_backend())
        has_pdf = bool(getattr(self.engine, "pdf_path", ""))
        has_data = bool(getattr(self.engine, "df", None) is not None and not self.engine.df.empty)
        android_mode = (platform == "android")

        detect_ready = detection_ok and has_pdf
        preview_ready = detection_ok and has_pdf
        assign_ready = detection_ok and has_pdf and has_data and bool(getattr(self.engine, "selected_box_ids", []))
        export_ready = export_ok and has_pdf and has_data
        page_ready = has_pdf

        for attr, enabled in [
            ("btn_detect", detect_ready),
            ("btn_preview", preview_ready),
            ("btn_assign", assign_ready),
            ("btn_generate_one", export_ready),
            ("btn_generate_batch", export_ready),
            ("btn_prev", page_ready),
            ("btn_next", page_ready),
            ("btn_toolbar_prev", page_ready),
            ("btn_toolbar_next", page_ready),
            ("btn_toolbar_detect", detect_ready),
            ("btn_toolbar_refresh", preview_ready),
            ("btn_zoom_out", has_pdf),
            ("btn_zoom_in", has_pdf),
            ("btn_fit_width", has_pdf),
            ("btn_fit_page", has_pdf),
            ("btn_zoom_reset", has_pdf),
            ("btn_mobile_rail_zoom_reset", has_pdf),
            ("btn_mobile_rail_fit_page", has_pdf),
            ("btn_mobile_rail_fit_width", has_pdf),
            ("btn_mobile_prev", page_ready),
            ("btn_mobile_next", page_ready),
            ("btn_mobile_fit_width", has_pdf),
            ("btn_mobile_fit_page", has_pdf),
            ("btn_mobile_zoom_out", has_pdf),
            ("btn_mobile_zoom_in", has_pdf),
            ("btn_mobile_zoom_reset", has_pdf),
            ("btn_mobile_refresh", preview_ready),
            ("btn_mobile_detect_bar", detect_ready),
            ("btn_mobile_select_mode", True),
            ("btn_mobile_more", True),
            ("btn_mobile_sidebar", True),
            ("btn_mobile_tools_fab", True),
            ("btn_clear_mapping", bool(getattr(self.engine, "selected_box_ids", []))),
            ("btn_toolbar_export", export_ready),
            ("btn_learning_refresh", True),
            ("btn_learning_reload", True),
            ("btn_learning_save_revision", has_pdf),
            ("btn_learning_apply_best", has_pdf),
            ("btn_learning_view_current", bool(getattr(self.engine, "current_revision_id", "") or getattr(self.engine, "last_saved_revision_id", ""))),
            ("btn_learning_browse", bool((self.engine.learning_storage.load_revision_index() or {}).get("items", [])) if getattr(self.engine, "learning_storage", None) else False),
            ("btn_learning_approve_current", bool(has_pdf and (getattr(self.engine, "current_revision_id", "") or getattr(self.engine, "last_saved_revision_id", "") or getattr(self.engine, "all_boxes", [])))),
        ]:
            self._set_widget_enabled(getattr(self, attr, None), enabled)

        if hasattr(self, "backend_note_lbl") and self.backend_note_lbl is not None:
            mode = "Android mode" if android_mode else "Desktop mode"
            record = self.selected_record() or "No record selected"
            page_txt = str(self.current_page_idx()) if has_pdf else "-"
            preview_mode = self._session_preview_mode() if has_pdf else "no_pdf"
            mode_label = {
                "no_pdf": "No preview yet",
                "raw": "Raw page preview",
                "android_preview": "Android detection preview",
                "processed": "Filled record preview",
            }.get(preview_mode, "Preview")
            next_step = self._session_next_action_hint()
            self.backend_note_lbl.text = (
                f"{mode} • Session mode: {mode_label} • Page: {page_txt}\n"
                f"Record: {record}\n"
                f"Find fields: {'ready' if detection_ok else 'not available'} • "
                f"Create PDFs: {'ready' if export_ok else 'not available'} • Next step: {next_step}"
            )

        export_note_text = ""
        if not export_ok:
            export_note_text = "PDF export is not available in this build."
        elif not has_pdf and not has_data:
            export_note_text = "Open a PDF form and a data file to enable export."
        elif not has_pdf:
            export_note_text = "Open a PDF form to enable export."
        elif not has_data:
            export_note_text = "Load a data file to enable export."
        else:
            export_note_text = "Export is ready. Create one PDF or the full batch."
        for _note in (getattr(self, "export_note_lbl", None), getattr(self, "export_note_lbl_android", None)):
            if _note is not None:
                _note.text = export_note_text

        if getattr(self, "mobile_flow_lbl", None) is not None:
            if not has_pdf:
                self.mobile_flow_lbl.text = "Step 1: Tap '1 Open Files' and load your PDF form."
            elif not has_data:
                self.mobile_flow_lbl.text = "PDF loaded! Now tap '1 Open Files' and load your CSV or Excel data file."
            elif not detection_ok:
                self.mobile_flow_lbl.text = "Data ready. Preview works, but field detection is not available in this build."
            elif not self.engine.all_boxes:
                self.mobile_flow_lbl.text = "Data loaded. Tap '2 Find Fields' to detect fillable areas on this page."
            elif not self.engine.custom_mappings:
                self.mobile_flow_lbl.text = "Fields found! Tap '6 Map Fields' and link each field to a data column."
            elif not export_ok:
                self.mobile_flow_lbl.text = "Field links saved. PDF export is not available in this build."
            else:
                self.mobile_flow_lbl.text = "All set! Tap '8 Export PDF' to create your filled PDF."

        self._refresh_preview_empty_hint()
        self._update_selection_inspector()
        self._update_bottom_statusbar()


    def _iter_page_inputs(self):
        seen = set()
        for name in ("page_input", "page_input_main", "page_input_mobile"):
            widget = getattr(self, name, None)
            if widget is None:
                continue
            wid = id(widget)
            if wid in seen:
                continue
            seen.add(wid)
            yield widget

    def _mirror_page_inputs(self, source, value):
        for widget in self._iter_page_inputs():
            if widget is source:
                continue
            try:
                if widget.text != value:
                    widget.text = value
            except Exception:
                pass

    def _read_page_input_text(self):
        preferred = []
        if getattr(self, "ui_mobile", False):
            preferred.extend([getattr(self, "page_input_mobile", None), getattr(self, "page_input_main", None)])
        else:
            preferred.extend([getattr(self, "page_input_main", None), getattr(self, "page_input_mobile", None)])
        preferred.append(getattr(self, "page_input", None))
        seen = set()
        for widget in preferred:
            if widget is None or id(widget) in seen:
                continue
            seen.add(id(widget))
            try:
                value = str(widget.text).strip()
                if value != "":
                    return value
            except Exception:
                pass
        return "0"

    def _set_page_inputs_text(self, value):
        text = str(value)
        for widget in self._iter_page_inputs():
            try:
                if widget.text != text:
                    widget.text = text
            except Exception:
                pass

    def current_page_idx(self):
        try:
            idx = max(0, int(self._read_page_input_text()))
        except Exception:
            idx = 0

        total = self.engine.total_pages()
        if total <= 0:
            return 0
        return min(idx, total - 1)

    def _box_type_counts(self):
        counts = {"check": 0, "line": 0, "field": 0}
        for kind in list(getattr(self.engine, "box_types", []) or []):
            counts[str(kind)] = counts.get(str(kind), 0) + 1
        return counts

    def _preview_boxes_payload_if_current(self, page_idx, preview_zoom):
        page_idx = int(page_idx)
        if self.engine.current_detection_matches_settings(page_idx):
            return self._build_preview_boxes_payload(page_idx=page_idx, preview_zoom=preview_zoom)
        if self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()):
            return self._build_preview_boxes_payload(page_idx=page_idx, preview_zoom=preview_zoom)
        return []

    def apply_ui_settings_to_engine(self):
        try:
            self.engine.settings["F_Area"] = int(self.f_area.text)
            self.engine.settings["F_MinW"] = int(self.f_minw.text)
            self.engine.settings["F_MinH"] = int(self.f_minh.text)
            self.engine.settings["F_Close"] = int(self.f_close.text)

            self.engine.settings["Line_MinW"] = int(self.line_minw.text)
            self.engine.settings["Line_MaxW"] = int(self.line_maxw.text)

            self.engine.settings["C_Strict"] = int(self.c_strict.text)
            self.engine.settings["C_Size"] = (
                int(self.c_size_min.text),
                int(self.c_size_max.text)
            )

            self.engine.settings["C_Border"] = float(self.c_border.text)
            self.engine.settings["C_Inner"] = float(self.c_inner.text)
            self.engine.settings["ROI_Max"] = int(self.roi_max.text)

            self.engine.settings["C_Open"] = int(self.c_open.text)
            self.engine.settings["C_Close"] = int(self.c_close.text)
            self.engine.settings["C_BandPct"] = float(self.c_band.text)
            self.engine.settings["C_AspectTol"] = float(self.c_aspect.text)
            self.engine.settings["Ext_Low"] = float(self.ext_low.text)
            self.engine.settings["Ext_High"] = float(self.ext_high.text)
            self.engine.settings["C_FillMin"] = float(self.c_fill.text)
            self.engine.settings["C_Eps"] = float(self.c_eps.text)
            self.engine.settings["Use_Extent"] = bool(getattr(self, "use_extent_chk", None).active if hasattr(self, "use_extent_chk") else int(self.use_extent.text.strip() or "0"))
            self.engine.text_tuning = self.engine._normalized_text_tuning(getattr(self, "text_tuning_ui", None))
        except Exception as e:
            raise ValueError(f"Invalid settings input: {e}")

    def get_app_output_dir(self):
        """Return a writable app-private directory for generated files/configs.
        Always prefers app.user_data_dir so no storage permissions are needed.
        """
        base_dir = str(getattr(self, "user_data_dir", "") or "").strip()
        if not base_dir:
            try:
                base_dir = get_form_alchemist_private_storage_root("form_alchemist")
            except Exception:
                pass
        if not base_dir:
            base_dir = os.path.join(os.path.expanduser("~"), "FormAlchemist")
        out_dir = os.path.join(base_dir, "output")
        try:
            os.makedirs(out_dir, exist_ok=True)
        except Exception:
            out_dir = base_dir
            os.makedirs(out_dir, exist_ok=True)
        return out_dir

    def get_default_file_path(self):
        if platform == "android":
            return self.get_app_output_dir()
        return os.path.expanduser("~")
        
    def push_engine_settings_to_ui(self):
        s = self.engine.settings
        self.f_area.text = str(s["F_Area"])
        self.f_minw.text = str(s["F_MinW"])
        self.f_minh.text = str(s["F_MinH"])
        self.f_close.text = str(s["F_Close"])
        self.line_minw.text = str(s["Line_MinW"])
        self.line_maxw.text = str(s["Line_MaxW"])

        self.c_strict.text = str(s["C_Strict"])
        self.c_size_min.text = str(s["C_Size"][0])
        self.c_size_max.text = str(s["C_Size"][1])
        self.c_border.text = str(s["C_Border"])
        self.c_inner.text = str(s["C_Inner"])
        self.roi_max.text = str(s["ROI_Max"])

        self.c_open.text = str(s["C_Open"])
        self.c_close.text = str(s["C_Close"])
        self.c_band.text = str(s["C_BandPct"])
        self.c_aspect.text = str(s["C_AspectTol"])
        self.ext_low.text = str(s["Ext_Low"])
        self.ext_high.text = str(s["Ext_High"])
        self.c_fill.text = str(s["C_FillMin"])
        self.c_eps.text = str(s["C_Eps"])
        
        if hasattr(self, "use_extent_chk"):
            self.use_extent_chk.active = bool(s["Use_Extent"])
        elif hasattr(self, "use_extent"):
            self.use_extent.text = "1" if s["Use_Extent"] else "0"

        self.text_tuning_ui = self.engine._normalized_text_tuning(getattr(self.engine, "text_tuning", None))
        self._update_text_tuning_summary()

    def _apply_record_label_source_column(self, column_name, popup=None):
        column_name = str(column_name or "").strip()
        if self.engine.df is None or self.engine.df.empty:
            if popup is not None:
                try:
                    popup.dismiss()
                except Exception:
                    pass
            return
        if column_name and column_name not in list(self.engine.df.columns):
            self.set_status(f"Record label column not found:\n{column_name}")
            return
        try:
            self.engine.rebuild_record_labels(source_column=column_name)
            self.refresh_patient_and_column_lists()
            self._persist_runtime_session_state(current_page_idx=self.current_page_idx())
            if popup is not None:
                try:
                    popup.dismiss()
                except Exception:
                    pass
            if column_name:
                self.set_status(f"Session fallback applied.\nRecord labels now use column:\n{column_name}")
            else:
                self.set_status("Session fallback cleared. Using generated row labels.")
            if getattr(self.engine, "pdf_path", ""):
                Clock.schedule_once(lambda dt: self.on_preview(None), 0.05)
        except Exception as e:
            traceback.print_exc()
            self.set_status(f"Record label fallback error:\n{e}")

    def _prompt_record_label_fallback_if_needed(self):
        if not getattr(self, "engine", None):
            return
        if not self.engine.needs_record_label_source_fallback():
            return
        if getattr(self, "_record_label_fallback_popup", None):
            try:
                if self._record_label_fallback_popup.parent is not None:
                    return
            except Exception:
                return

        cols = [str(c) for c in list(self.engine.df.columns) if str(c).strip() and not str(c).startswith("_")]
        if not cols:
            return

        palette = getattr(self, "theme_palette", {}) or {}
        def _c(name, fallback):
            return palette.get(name, fallback)

        outer = BoxLayout(orientation="vertical", spacing=dp(10), padding=[dp(14), dp(14), dp(14), dp(14)])
        title = Label(
            text="No name columns found",
            color=_c("text", (0.93, 0.96, 1.0, 1)),
            size_hint_y=None,
            height=dp(26),
            halign="left",
            valign="middle",
            font_size=dp(16),
            bold=True,
        )
        title.bind(size=lambda inst, value: setattr(inst, "text_size", value))
        outer.add_widget(title)

        helper = Label(
            text="Session mode needs something to identify each record. Choose a target column to use as the record label.",
            color=_c("muted", (0.60, 0.68, 0.80, 1)),
            size_hint_y=None,
            height=dp(52),
            halign="left",
            valign="middle",
            font_size=dp(11),
        )
        helper.bind(size=lambda inst, value: setattr(inst, "text_size", value))
        outer.add_widget(helper)

        picker = SearchableSelectField(
            text=(cols[0] if cols else COLUMN_SELECT_TEXT),
            values=cols,
            placeholder=COLUMN_SELECT_TEXT,
            picker_title="Choose record label column",
            search_hint="Type to find a target column",
            size_hint_y=None,
            height=dp(46),
            font_size=dp(14),
            background_normal="",
            background_color=_c("surface_alt", (0.11, 0.135, 0.185, 1)),
            color=_c("text", (0.93, 0.96, 1.0, 1)),
        )
        outer.add_widget(picker)

        btn_row = GridLayout(cols=3, size_hint_y=None, height=dp(44), spacing=dp(8))
        btn_rows = Button(text="Use Rows", background_normal="", background_color=_c("surface_alt", (0.11, 0.135, 0.185, 1)), color=_c("text", (0.93, 0.96, 1.0, 1)))
        btn_apply = Button(text="Apply", background_normal="", background_color=_c("primary", (0.10, 0.78, 0.63, 1)), color=_c("button_text", (1, 1, 1, 1)))
        btn_close = Button(text="Close", background_normal="", background_color=_c("surface_alt", (0.11, 0.135, 0.185, 1)), color=_c("text", (0.93, 0.96, 1.0, 1)))
        btn_row.add_widget(btn_rows)
        btn_row.add_widget(btn_apply)
        btn_row.add_widget(btn_close)
        outer.add_widget(btn_row)

        popup = Popup(
            title="",
            separator_height=0,
            background="",
            background_color=(0, 0, 0, 0.78),
            content=outer,
            size_hint=((0.94 if getattr(self, "ui_mobile", False) else 0.66), (0.46 if getattr(self, "ui_mobile", False) else 0.40)),
            auto_dismiss=False,
        )
        self._record_label_fallback_popup = popup

        btn_apply.bind(on_release=lambda *_: self._apply_record_label_source_column(picker.text, popup=popup))
        btn_rows.bind(on_release=lambda *_: self._apply_record_label_source_column("", popup=popup))
        btn_close.bind(on_release=lambda *_: popup.dismiss())
        popup.bind(on_dismiss=lambda *_: setattr(self, "_record_label_fallback_popup", None))
        popup.open()

    def refresh_patient_and_column_lists(self):
        if self.engine.df is None or self.engine.df.empty:
            self.patient_spinner.values = []
            self.column_spinner.values = []
            self.patient_spinner.text = RECORD_SELECT_TEXT
            self.column_spinner.text = COLUMN_SELECT_TEXT
            return

        label_col = "_RECORD_LABEL" if "_RECORD_LABEL" in self.engine.df.columns else "_DISPLAY_NAME"
        patient_names = sorted([
            str(x).strip()
            for x in self.engine.df[label_col].dropna().tolist()
            if str(x).strip()
        ])
        previous_label = self.selected_record() if hasattr(self, "patient_spinner") else None
        pending_key = str(getattr(self, "_pending_restored_record_key", "") or "").strip()
        pending_label = str(getattr(self, "_pending_restored_record_label", "") or "").strip()
        resolved_label = self.engine.resolve_record_label(record_key=pending_key, fallback_label=pending_label or previous_label or "")

        self.patient_spinner.values = patient_names
        if resolved_label and resolved_label in patient_names:
            self.patient_spinner.text = resolved_label
        else:
            self.patient_spinner.text = patient_names[0] if patient_names else RECORD_SELECT_TEXT

        self._pending_restored_record_key = ""
        self._pending_restored_record_label = ""

        cols = sorted([str(c) for c in self.engine.df.columns.tolist()])
        previous_column = str(getattr(self, "column_spinner", None).text or "").strip() if hasattr(self, "column_spinner") else ""
        self.column_spinner.values = cols
        if previous_column and previous_column in cols:
            self.column_spinner.text = previous_column
        else:
            self.column_spinner.text = cols[0] if cols else COLUMN_SELECT_TEXT
        self.refresh_backend_capabilities_ui()

    def selected_record(self):
        val = self.patient_spinner.text.strip()
        if not val or val in {RECORD_SELECT_TEXT, "Select Patient"}:
            return None
        return val

    def selected_record_key(self):
        label = self.selected_record()
        if not label or self.engine.df is None or self.engine.df.empty:
            return None
        try:
            key = str(self.engine.resolve_record_key(label) or "").strip()
            return key or None
        except Exception:
            return None

    def selected_patient(self):
        return self.selected_record()

    def _has_loaded_record_data(self):
        return bool(getattr(self.engine, "df", None) is not None and not self.engine.df.empty)

    def _session_preview_mode(self, page_idx=None):
        if not getattr(self.engine, "pdf_path", ""):
            return "no_pdf"
        if not self._has_loaded_record_data() or not self.selected_record():
            return "raw"
        if platform == "android" and not self.engine.supports_export_backend():
            return "android_preview"
        return "processed"

    def _session_next_action_hint(self):
        mode = self._session_preview_mode()
        if mode == "no_pdf":
            return "Open a PDF form"
        if not self._has_loaded_record_data():
            return "Load a data file"
        if not self.selected_record():
            return "Choose a record"
        if not getattr(self.engine, "all_boxes", []):
            return "Find fields on this page"
        if not getattr(self.engine, "custom_mappings", []):
            return "Link detected boxes to data columns"
        if self.engine.supports_export_backend():
            return "Preview or export this record"
        return "Preview this record"

    def _render_session_page(self, page_idx=None, reason="Preview"):
        if not self.engine.pdf_path:
            self.set_status("Load PDF first.")
            return False

        page_idx = self.current_page_idx() if page_idx is None else int(page_idx)
        self._set_page_inputs_text(page_idx)
        record = self.selected_record()
        self.apply_ui_settings_to_engine()
        self._prepare_page_context(page_idx)
        mode = self._session_preview_mode(page_idx)

        if mode == "android_preview":
            raw_img = self.engine.get_raw_preview_pixmap(page_idx=page_idx, preview_zoom=PREVIEW_SCALE)
            boxes_payload = self._preview_boxes_payload_if_current(page_idx=page_idx, preview_zoom=PREVIEW_SCALE)
            self.render_preview_image(raw_img, boxes_payload=boxes_payload, page_idx=page_idx, preview_zoom=PREVIEW_SCALE)
            self._sync_box_selection_ui()
            self.refresh_backend_capabilities_ui()
            self.set_status(
                f"{reason}.\n"
                f"Mode: Android preview\n"
                f"Record: {record or 'None selected'}\n"
                f"Page: {page_idx}\n"
                f"Boxes shown: {len(boxes_payload)}\n"
                f"PDF export backend is unavailable in this build."
            )
            self._persist_runtime_session_state(current_page_idx=page_idx)
            return True

        if mode == "raw":
            raw_img = self.engine.get_raw_preview_pixmap(page_idx=page_idx, preview_zoom=PREVIEW_SCALE)
            boxes_payload = self._preview_boxes_payload_if_current(page_idx=page_idx, preview_zoom=PREVIEW_SCALE)
            self.render_preview_image(raw_img, boxes_payload=boxes_payload, page_idx=page_idx, preview_zoom=PREVIEW_SCALE)
            self._sync_box_selection_ui()
            self.refresh_backend_capabilities_ui()
            reason_tail = "Choose a record to merge data into the preview." if self._has_loaded_record_data() else "Load a data file to preview filled values."
            self.set_status(
                f"{reason}.\n"
                f"Mode: Raw page preview\n"
                f"Page: {page_idx}\n"
                f"Boxes shown: {len(boxes_payload)}\n"
                f"{reason_tail}"
            )
            self._persist_runtime_session_state(current_page_idx=page_idx)
            return True

        img = self.engine.get_processed_preview_pixmap(
            patient_name=record,
            page_idx=page_idx,
            preview_zoom=PREVIEW_SCALE,
            record_key=self.selected_record_key(),
        )
        boxes_payload = self._build_preview_boxes_payload(page_idx=page_idx, preview_zoom=PREVIEW_SCALE)
        self.render_preview_image(img, boxes_payload=boxes_payload, page_idx=page_idx, preview_zoom=PREVIEW_SCALE)
        self._sync_box_selection_ui()
        self.refresh_backend_capabilities_ui()
        self.set_status(
            f"{reason}.\n"
            f"Mode: Filled record preview\n"
            f"Record: {record}\n"
            f"Page: {page_idx}\n"
            f"Boxes: {len(self.engine.all_boxes)}"
        )
        self._persist_runtime_session_state(current_page_idx=page_idx)
        return True

    def _go_to_page(self, target_idx, reason="Page changed"):
        self._prepare_mobile_page_transition()
        current_idx = self.current_page_idx()
        total = max(int(self.engine.total_pages() or 0), 0)
        if total > 0:
            target_idx = max(0, min(int(target_idx), total - 1))
        else:
            target_idx = max(0, int(target_idx))
        self._set_page_inputs_text(target_idx)
        if int(current_idx) != int(target_idx):
            self._clear_mobile_repair_state(page_idx=target_idx)
            self._clear_manual_capture_preview(preserve_mode=False)
        if self.engine.pdf_path:
            return self._render_session_page(page_idx=target_idx, reason=reason)
        self.refresh_backend_capabilities_ui()
        self.set_status(f"{reason}.\nPage: {target_idx}\nOpen a PDF form to render this page.")
        return False

    def _sanitize_box_id_list(self, ids):
        out = []
        for x in (ids or []):
            try:
                out.append(int(x))
            except Exception:
                continue
        return sorted(set(out))

    def _resolve_live_box_ids_for_rect_refs(self, rect_refs, min_score=0.45, exclude_ids=None):
        engine = getattr(self, 'engine', None)
        if engine is None:
            return []
        all_boxes = list(getattr(engine, 'all_boxes', []) or [])
        all_types = list(getattr(engine, 'box_types', []) or [])
        used = set(self._sanitize_box_id_list(exclude_ids or []))
        resolved = []
        for ref in (rect_refs or []):
            if not isinstance(ref, (list, tuple)) or len(ref) < 2:
                continue
            rect, kind = ref[0], str(ref[1] or '')
            try:
                rect = fitz.Rect(rect)
            except Exception:
                continue
            best_idx = -1
            best_score = 0.0
            for idx_existing, existing in enumerate(all_boxes):
                if idx_existing in used:
                    continue
                existing_kind = str(all_types[idx_existing] if idx_existing < len(all_types) else '')
                if kind and existing_kind != kind:
                    continue
                score = self._mobile_repair_overlap_score(rect, existing)
                if score > best_score:
                    best_idx = idx_existing
                    best_score = score
            if best_idx >= 0 and best_score >= float(min_score):
                used.add(best_idx)
                resolved.append(best_idx)
        return self._sanitize_box_id_list(resolved)

    def _stash_page_selection(self, page_idx=None):
        if not hasattr(self, "_page_selected_box_ids"):
            self._page_selected_box_ids = {}
        page_idx = self.current_page_idx() if page_idx is None else int(page_idx)
        self._page_selected_box_ids[page_idx] = self._sanitize_box_id_list(getattr(self.engine, "selected_box_ids", []))
        self._active_selection_page_idx = page_idx

    def _restore_page_selection(self, page_idx=None, clear_if_missing=True):
        if not hasattr(self, "_page_selected_box_ids"):
            self._page_selected_box_ids = {}
        page_idx = self.current_page_idx() if page_idx is None else int(page_idx)
        ids = list(self._page_selected_box_ids.get(page_idx, []))
        self.engine.selected_box_ids = self._sanitize_box_id_list(ids if ids else ([] if clear_if_missing else getattr(self.engine, "selected_box_ids", [])))
        self._active_selection_page_idx = page_idx
        return list(self.engine.selected_box_ids)

    def _prepare_page_context(self, page_idx, clear_selection_if_missing=True):
        page_idx = int(page_idx)
        prev_page = int(getattr(self, "_active_selection_page_idx", page_idx))
        if prev_page != page_idx:
            self._stash_page_selection(prev_page)

        restored = False
        if int(getattr(self.engine, "detected_page_idx", -1) or -1) == page_idx and bool(getattr(self.engine, "all_boxes", [])):
            restored = True
        else:
            restored = bool(self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()))

        if not restored and int(getattr(self.engine, "detected_page_idx", -1) or -1) != page_idx:
            self.engine.all_boxes = []
            self.engine.box_types = []
            self.engine.geom = self.engine._empty_geom()
            self.engine.semantic_targets = self.engine._empty_semantic_targets()
            self.engine.detected_page_idx = None
            self.engine.detected_settings_signature = None

        self._restore_page_selection(page_idx, clear_if_missing=clear_selection_if_missing)

    def render_preview_image(self, img_bgr, boxes_payload=None, page_idx=0, preview_zoom=PREVIEW_SCALE):
        if img_bgr is None:
            return
        self._last_preview_page_idx = int(page_idx or 0)
        self._last_preview_render_zoom = float(preview_zoom or PREVIEW_SCALE)
        self._last_preview_has_boxes = bool(boxes_payload)
        if getattr(self, "ui_mobile", False):
            mode = str(getattr(self, "preview_zoom_mode", "fit_width") or "fit_width")
            if mode in ("fit_width", "fit_page"):
                self._mobile_zoom_bootstrap_pending = True
        if hasattr(self.preview, "set_texture_from_bgr"):
            self.preview.set_texture_from_bgr(img_bgr)
            if boxes_payload is None:
                self.preview.clear_boxes()
            else:
                self.preview.set_boxes_payload(
                    boxes_payload,
                    selected_ids=self.engine.selected_box_ids,
                    preview_zoom=preview_zoom,
                    page_idx=page_idx,
                    label_mode=('selected_only' if getattr(self, 'ui_mobile', False) else 'all'),
                )
        else:
            rgba = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGBA)
            buf = rgba.tobytes()
            texture = Texture.create(size=(rgba.shape[1], rgba.shape[0]), colorfmt="rgba")
            texture.blit_buffer(buf, colorfmt="rgba", bufferfmt="ubyte")
            texture.flip_vertical()
            self.preview.texture = texture
        if bool(getattr(self, '_capture_mode_active', False)):
            self._refresh_manual_capture_overlay()
        else:
            try:
                self.preview.clear_capture_guides()
            except Exception:
                pass
        self._refresh_preview_empty_hint()
        Clock.schedule_once(self._post_preview_refresh, 0)
        Clock.schedule_once(self._post_preview_refresh, 0.05)

    def _build_preview_boxes_payload(self, page_idx, preview_zoom):
        payload = []
        for i, r in enumerate(self.engine.all_boxes):
            payload.append({
                "id": i,
                "pdf_x": float(r.x0),
                "pdf_y": float(r.y0),
                "pdf_w": float(r.width),
                "pdf_h": float(r.height),
                "x": float(r.x0 * preview_zoom),
                "y": float(r.y0 * preview_zoom),
                "w": float(r.width * preview_zoom),
                "h": float(r.height * preview_zoom),
                "t": self.engine.box_types[i] if i < len(self.engine.box_types) else "field",
                "mapping": self.engine.describe_box_mapping(i, page_idx),
            })
        return payload


    def _get_preview_content_widget(self):
        return getattr(self, "preview_stack", None) or getattr(self, "preview", None)

    def _capture_preview_anchor(self):
        wrap = getattr(self, "preview_wrap", None)
        content = self._get_preview_content_widget()
        if wrap is None or content is None:
            return None
        viewport_w = max(float(getattr(wrap, "width", 0) or 0), 1.0)
        viewport_h = max(float(getattr(wrap, "height", 0) or 0), 1.0)
        content_w = max(float(getattr(content, "width", 0) or 0), viewport_w)
        content_h = max(float(getattr(content, "height", 0) or 0), viewport_h)
        max_x = max(content_w - viewport_w, 0.0)
        max_y = max(content_h - viewport_h, 0.0)
        offset_x = float(getattr(wrap, "scroll_x", 0.0) or 0.0) * max_x
        offset_y = (1.0 - float(getattr(wrap, "scroll_y", 1.0) or 1.0)) * max_y
        center_x = (offset_x + viewport_w / 2.0) / max(content_w, 1.0)
        center_y = (offset_y + viewport_h / 2.0) / max(content_h, 1.0)
        return (
            max(0.0, min(1.0, center_x)),
            max(0.0, min(1.0, center_y)),
        )

    def _restore_preview_anchor(self, anchor):
        wrap = getattr(self, "preview_wrap", None)
        content = self._get_preview_content_widget()
        if wrap is None or content is None or not anchor:
            return
        viewport_w = max(float(getattr(wrap, "width", 0) or 0), 1.0)
        viewport_h = max(float(getattr(wrap, "height", 0) or 0), 1.0)
        content_w = max(float(getattr(content, "width", 0) or 0), viewport_w)
        content_h = max(float(getattr(content, "height", 0) or 0), viewport_h)
        max_x = max(content_w - viewport_w, 0.0)
        max_y = max(content_h - viewport_h, 0.0)
        target_x = max(0.0, min(max_x, anchor[0] * content_w - viewport_w / 2.0))
        target_y = max(0.0, min(max_y, anchor[1] * content_h - viewport_h / 2.0))
        wrap.scroll_x = 0.0 if max_x <= 1e-6 else (target_x / max_x)
        wrap.scroll_y = 1.0 if max_y <= 1e-6 else (1.0 - (target_y / max_y))

    def _capture_preview_anchor_at_window_point(self, win_x, win_y):
        wrap = getattr(self, "preview_wrap", None)
        content = self._get_preview_content_widget()
        if wrap is None or content is None:
            return None
        viewport_w = max(float(getattr(wrap, "width", 0) or 0), 1.0)
        viewport_h = max(float(getattr(wrap, "height", 0) or 0), 1.0)
        content_w = max(float(getattr(content, "width", 0) or 0), viewport_w)
        content_h = max(float(getattr(content, "height", 0) or 0), viewport_h)
        max_x = max(content_w - viewport_w, 0.0)
        max_y = max(content_h - viewport_h, 0.0)
        offset_x = float(getattr(wrap, "scroll_x", 0.0) or 0.0) * max_x
        offset_y = (1.0 - float(getattr(wrap, "scroll_y", 1.0) or 1.0)) * max_y
        local_x = max(0.0, min(viewport_w, float(win_x) - float(getattr(wrap, "x", 0.0) or 0.0)))
        local_y = max(0.0, min(viewport_h, float(win_y) - float(getattr(wrap, "y", 0.0) or 0.0)))
        content_x = offset_x + local_x
        content_y = offset_y + (viewport_h - local_y)
        return (
            max(0.0, min(1.0, content_x / max(content_w, 1.0))),
            max(0.0, min(1.0, content_y / max(content_h, 1.0))),
        )

    def _restore_preview_anchor_at_window_point(self, anchor, win_x, win_y):
        wrap = getattr(self, "preview_wrap", None)
        content = self._get_preview_content_widget()
        if wrap is None or content is None or not anchor:
            return
        viewport_w = max(float(getattr(wrap, "width", 0) or 0), 1.0)
        viewport_h = max(float(getattr(wrap, "height", 0) or 0), 1.0)
        content_w = max(float(getattr(content, "width", 0) or 0), viewport_w)
        content_h = max(float(getattr(content, "height", 0) or 0), viewport_h)
        max_x = max(content_w - viewport_w, 0.0)
        max_y = max(content_h - viewport_h, 0.0)
        local_x = max(0.0, min(viewport_w, float(win_x) - float(getattr(wrap, "x", 0.0) or 0.0)))
        local_y = max(0.0, min(viewport_h, float(win_y) - float(getattr(wrap, "y", 0.0) or 0.0)))
        target_x = max(0.0, min(max_x, anchor[0] * content_w - local_x))
        target_y = max(0.0, min(max_y, anchor[1] * content_h - (viewport_h - local_y)))
        wrap.scroll_x = 0.0 if max_x <= 1e-6 else (target_x / max_x)
        wrap.scroll_y = 1.0 if max_y <= 1e-6 else (1.0 - (target_y / max_y))

    def _scroll_preview_by_pixels(self, dx, dy):
        wrap = getattr(self, "preview_wrap", None)
        content = self._get_preview_content_widget()
        if wrap is None or content is None:
            return
        viewport_w = max(float(getattr(wrap, "width", 0) or 0), 1.0)
        viewport_h = max(float(getattr(wrap, "height", 0) or 0), 1.0)
        content_w = max(float(getattr(content, "width", 0) or 0), viewport_w)
        content_h = max(float(getattr(content, "height", 0) or 0), viewport_h)
        max_x = max(content_w - viewport_w, 0.0)
        max_y = max(content_h - viewport_h, 0.0)
        if max_x > 1e-6:
            offset_x = float(getattr(wrap, "scroll_x", 0.0) or 0.0) * max_x
            target_x = max(0.0, min(max_x, offset_x - float(dx)))
            wrap.scroll_x = target_x / max_x
        if max_y > 1e-6:
            offset_y = (1.0 - float(getattr(wrap, "scroll_y", 1.0) or 1.0)) * max_y
            target_y = max(0.0, min(max_y, offset_y - float(dy)))
            wrap.scroll_y = 1.0 - (target_y / max_y)

    def _focus_preview_box(self, box_id, *_args, **_kwargs):
        wrap = getattr(self, "preview_wrap", None)
        preview = getattr(self, "preview", None)
        content = self._get_preview_content_widget()
        if wrap is None or preview is None or content is None:
            return
        try:
            box_id = int(box_id)
        except Exception:
            return

        hit = None
        for box in getattr(preview, "boxes_payload", []) or []:
            try:
                if int(box.get("id", -1)) == box_id:
                    hit = box
                    break
            except Exception:
                continue

        tex = getattr(preview, "texture", None)
        disp = preview._get_display_rect() if hasattr(preview, "_get_display_rect") else None
        if hit is None or tex is None or disp is None:
            Clock.schedule_once(lambda dt: self._post_preview_refresh(reset_scroll=False), 0)
            return

        bx, by, bw, bh = preview._resolve_box_image_rect(hit) if hasattr(preview, "_resolve_box_image_rect") else (
            float(hit.get("x", 0.0) or 0.0),
            float(hit.get("y", 0.0) or 0.0),
            float(hit.get("w", 0.0) or 0.0),
            float(hit.get("h", 0.0) or 0.0),
        )
        dx, dy, dw, dh = disp
        sx = dw / float(max(getattr(tex, "width", 1), 1))
        sy = dh / float(max(getattr(tex, "height", 1), 1))
        box_cx = dx + (bx + bw / 2.0) * sx
        box_cy = dy + (float(getattr(tex, "height", 0)) - (by + bh / 2.0)) * sy

        viewport_w = max(float(getattr(wrap, "width", 0) or 0), 1.0)
        viewport_h = max(float(getattr(wrap, "height", 0) or 0), 1.0)
        content_w = max(float(getattr(content, "width", 0) or 0), viewport_w)
        content_h = max(float(getattr(content, "height", 0) or 0), viewport_h)
        max_x = max(content_w - viewport_w, 0.0)
        max_y = max(content_h - viewport_h, 0.0)

        target_x = max(0.0, min(max_x, box_cx - viewport_w / 2.0))
        target_y = max(0.0, min(max_y, box_cy - viewport_h / 2.0))

        wrap.scroll_x = 0.0 if max_x <= 1e-6 else (target_x / max_x)
        wrap.scroll_y = 1.0 if max_y <= 1e-6 else (1.0 - (target_y / max_y))

    def _refresh_preview_boxes_for_current_view(self):
        preview = getattr(self, "preview", None)
        if preview is None or getattr(preview, "texture", None) is None:
            return
        page_idx = int(getattr(self, "_last_preview_page_idx", self.current_page_idx()) or 0)
        render_zoom = float(getattr(self, "_last_preview_render_zoom", PREVIEW_SCALE) or PREVIEW_SCALE)
        self._prepare_page_context(page_idx, clear_selection_if_missing=False)
        boxes_payload = self._preview_boxes_payload_if_current(page_idx=page_idx, preview_zoom=render_zoom)
        preview.set_boxes_payload(
            boxes_payload,
            selected_ids=getattr(self.engine, "selected_box_ids", []),
            preview_zoom=render_zoom,
            page_idx=page_idx,
            label_mode=('selected_only' if getattr(self, 'ui_mobile', False) else 'all'),
        )

    def _desired_preview_render_zoom(self):
        mode_attr = "preview_zoom_mode" if getattr(self, "ui_mobile", False) else "desktop_zoom_mode"
        scale_attr = "preview_zoom_factor" if getattr(self, "ui_mobile", False) else "desktop_zoom_factor"
        mode = getattr(self, mode_attr, "fit_page")
        current_scale = float(getattr(self, scale_attr, 1.0) or 1.0)
        base_zoom = float(getattr(self, "_last_preview_render_zoom", PREVIEW_SCALE) or PREVIEW_SCALE)
        if mode != "manual":
            return max(PREVIEW_SCALE, min(base_zoom, 4.0 if getattr(self, "ui_mobile", False) else 5.0))
        boost = max(1.0, min(current_scale, 2.0 if getattr(self, "ui_mobile", False) else 2.6))
        target = PREVIEW_SCALE * boost
        cap = 4.0 if getattr(self, "ui_mobile", False) else 5.0
        return max(PREVIEW_SCALE, min(target, cap))

    def _schedule_preview_resync(self, *_):
        Clock.unschedule(self._perform_preview_resync)
        Clock.schedule_once(self._perform_preview_resync, 0.08 if getattr(self, "ui_mobile", False) else 0.04)

    def _perform_preview_resync(self, *_):
        if not getattr(self.engine, "pdf_path", ""):
            return
        preview = getattr(self, "preview", None)
        if preview is None or getattr(preview, "texture", None) is None:
            return

        anchor = self._capture_preview_anchor()
        desired_zoom = self._desired_preview_render_zoom()
        last_zoom = float(getattr(self, "_last_preview_render_zoom", PREVIEW_SCALE) or PREVIEW_SCALE)
        page_idx = self.current_page_idx()
        patient = self.selected_patient()
        has_data = bool(getattr(self.engine, "df", None) is not None and not self.engine.df.empty)
        has_boxes = bool(getattr(self.engine, "all_boxes", None))

        if abs(desired_zoom - last_zoom) / max(last_zoom, 1e-6) < 0.035:
            self._refresh_preview_boxes_for_current_view()
            Clock.schedule_once(lambda dt: self._restore_preview_anchor(anchor), 0)
            return

        try:
            if platform == "android" and not self.engine.supports_export_backend():
                img = self.engine.get_raw_preview_pixmap(page_idx=page_idx, preview_zoom=desired_zoom)
                boxes_payload = self._preview_boxes_payload_if_current(page_idx=page_idx, preview_zoom=desired_zoom)
            elif not patient or not has_data:
                img = self.engine.get_raw_preview_pixmap(page_idx=page_idx, preview_zoom=desired_zoom)
                boxes_payload = self._preview_boxes_payload_if_current(page_idx=page_idx, preview_zoom=desired_zoom)
            else:
                img = self.engine.get_processed_preview_pixmap(patient_name=patient, page_idx=page_idx, preview_zoom=desired_zoom)
                boxes_payload = self._build_preview_boxes_payload(page_idx=page_idx, preview_zoom=desired_zoom)
            self.render_preview_image(
                img,
                boxes_payload=boxes_payload,
                page_idx=page_idx,
                preview_zoom=desired_zoom,
            )
            self._sync_box_selection_ui()
            Clock.schedule_once(lambda dt: self._restore_preview_anchor(anchor), 0)
        except Exception:
            self._refresh_preview_boxes_for_current_view()
            Clock.schedule_once(lambda dt: self._restore_preview_anchor(anchor), 0)

    def _focus_selected_preview_box(self, *_):
        ids = sorted(set(int(x) for x in getattr(self.engine, "selected_box_ids", []) if isinstance(x, int) or str(x).isdigit()))
        if not ids:
            self.set_status("Select a preview box first.")
            return
        self._focus_preview_box(ids[0])

    def _find_mapping_for_box(self, box_id, page_idx):
        if box_id < 0 or box_id >= len(self.engine.all_boxes):
            return None
        target_rect = self.engine.all_boxes[box_id]
        best_hit = None
        best_score = -1.0
        for configs in self.engine.custom_mappings.values():
            for cfg in configs:
                if cfg.get("page", 0) != page_idx:
                    continue
                for r in self.engine._mapping_rect_list(cfg):
                    score = self.engine._mapping_match_score(target_rect, r)
                    if score > best_score:
                        best_score = score
                        best_hit = cfg
        if best_hit is not None and best_score >= 1.0:
            return best_hit
        return None

    def _sync_box_selection_ui(self):
        ids = sorted(set(int(x) for x in self.engine.selected_box_ids if isinstance(x, int) or str(x).isdigit()))
        self.engine.selected_box_ids = ids
        if hasattr(self, "_page_selected_box_ids"):
            self._page_selected_box_ids[int(self.current_page_idx())] = list(ids)
            self._active_selection_page_idx = int(self.current_page_idx())
        if hasattr(self, "box_ids_input") and self.box_ids_input is not None:
            self.box_ids_input.text = ",".join(str(i) for i in ids)
        preview = getattr(self, "preview", None)
        if preview is not None and hasattr(preview, "set_selected_ids"):
            preview.set_selected_ids(ids)

        if not ids:
            self._reset_mapping_editor_state(clear_box_ids=False)
            if hasattr(self, "preview_info") and self.preview_info is not None:
                if getattr(self, "ui_mobile", False):
                    mode_txt = "Select ON" if bool(getattr(self, "mobile_select_mode", False)) else "Select OFF"
                    self.preview_info.text = f"Preview ready. {mode_txt} • Tap a box to select • Map if correct • Repair only if needed."
                else:
                    self.preview_info.text = "Preview ready. Tap a box to select • Map if correct • Repair only if needed."
            self._update_selection_inspector()
            self._update_bottom_statusbar()
            return

        first = ids[0]
        self._load_mapping_into_editor(first)
        mapping = self.engine.describe_box_mapping(first, self.current_page_idx())
        if hasattr(self, "preview_info") and self.preview_info is not None:
            self.preview_info.text = f"Selected: {ids} • {mapping}"
        self._update_selection_inspector()
        self._refresh_android_selection_inspector()
        self._update_bottom_statusbar()
        try:
            if hasattr(self, '_show_desktop_right_section') and platform != 'android':
                self._show_desktop_right_section('Selection')
        except Exception:
            pass

    def _reset_mapping_editor_state(self, clear_box_ids=False):
        try:
            if hasattr(self, "column_spinner") and self.column_spinner is not None:
                self.column_spinner.text = COLUMN_SELECT_TEXT
        except Exception:
            pass
        try:
            if hasattr(self, "trigger_input") and self.trigger_input is not None:
                self.trigger_input.text = ""
        except Exception:
            pass
        try:
            if hasattr(self, "grid_flag_chk") and self.grid_flag_chk is not None:
                self.grid_flag_chk.active = False
            elif hasattr(self, "grid_flag_input") and self.grid_flag_input is not None:
                self.grid_flag_input.text = "0"
        except Exception:
            pass
        try:
            if hasattr(self, "grid_n_input") and self.grid_n_input is not None:
                self.grid_n_input.text = "1"
        except Exception:
            pass
        if clear_box_ids:
            try:
                if hasattr(self, "box_ids_input") and self.box_ids_input is not None:
                    self.box_ids_input.text = ""
            except Exception:
                pass

    def _load_mapping_into_editor(self, box_id):
        mapping = self._find_mapping_for_box(box_id, self.current_page_idx())
        if mapping:
            col = str(mapping.get("column", "")).strip()
            trig = str(mapping.get("trigger", "")).strip()
            if col:
                self.column_spinner.text = col
            else:
                self.column_spinner.text = COLUMN_SELECT_TEXT
            self.trigger_input.text = trig
            if hasattr(self, "grid_flag_chk"):
                self.grid_flag_chk.active = bool(mapping.get("g", False))
            else:
                self.grid_flag_input.text = "1" if bool(mapping.get("g", False)) else "0"
            self.grid_n_input.text = str(int(mapping.get("n", 1)))
        else:
            self._reset_mapping_editor_state(clear_box_ids=False)
        return mapping


    def _sync_mobile_tools_layout(self, *_):
        if not getattr(self, "ui_mobile", False):
            return
        if platform == "android":
            fab = getattr(self, "btn_mobile_tools_fab", None)
            panel = getattr(self, "mobile_bottom_panel", None)
            dock = getattr(self, "mobile_bottom_dock", None)
            for widget in (fab, panel, dock):
                if widget is not None:
                    try:
                        widget.opacity = 0
                        widget.disabled = True
                        widget.size_hint = (None, None)
                        widget.size = (0, 0)
                        widget.pos = (-10000, -10000)
                    except Exception:
                        pass
            return
        fab = getattr(self, "btn_mobile_tools_fab", None)
        panel = getattr(self, "mobile_bottom_panel", None)
        tray_open = bool(getattr(self, "mobile_tools_tray_open", False))
        if fab is not None:
            try:
                fab.pos = (max(dp(8), Window.width - fab.width - dp(12)), dp(12))
            except Exception:
                pass
        if panel is not None:
            try:
                panel.width = min(Window.width - dp(24), dp(500))
                panel.x = max(dp(12), Window.width - panel.width - dp(12))
                if tray_open:
                    panel.height = dp(170)
                    panel.y = dp(72)
                else:
                    panel.height = 0
                    panel.y = -10000
            except Exception:
                pass

    def _set_mobile_tools_tray_visible(self, open_state=True):
        if platform == "android":
            self.mobile_tools_tray_open = False
            panel = getattr(self, "mobile_bottom_panel", None)
            fab = getattr(self, "btn_mobile_tools_fab", None)
            dock = getattr(self, "mobile_bottom_dock", None)
            for widget in (panel, fab, dock):
                if widget is not None:
                    try:
                        widget.opacity = 0
                        widget.disabled = True
                        widget.size_hint = (None, None)
                        widget.size = (0, 0)
                        widget.pos = (-10000, -10000)
                    except Exception:
                        pass
            return
        self.mobile_tools_tray_open = bool(open_state)
        if self.mobile_tools_tray_open and bool(getattr(self, "_mobile_sidebar_open", False)):
            self._set_mobile_sidebar_state(False)
        self._set_mobile_quick_rail_visible(False)
        panel = getattr(self, "mobile_bottom_panel", None)
        fab = getattr(self, "btn_mobile_tools_fab", None)
        if panel is not None:
            panel.opacity = 1 if self.mobile_tools_tray_open else 0
            panel.disabled = not self.mobile_tools_tray_open
            try:
                panel.size_hint = (None, None)
                if self.mobile_tools_tray_open:
                    panel.height = dp(170)
                else:
                    panel.height = 0
                    panel.pos = (-10000, -10000)
            except Exception:
                pass
        if fab is not None:
            try:
                fab.text = "Close" if self.mobile_tools_tray_open else "More"
            except Exception:
                pass
        self._sync_mobile_tools_layout()
        self._refresh_mobile_hud_restore_button()

    def _toggle_mobile_tools_tray(self, *_):
        if platform == "android":
            return
        self._set_mobile_tools_tray_visible(not bool(getattr(self, "mobile_tools_tray_open", False)))

    def _refresh_mobile_select_mode_ui(self):
        if not getattr(self, "ui_mobile", False):
            return
        btn = getattr(self, "btn_mobile_select_mode", None)
        if btn is not None:
            try:
                active = bool(getattr(self, "mobile_select_mode", False))
                btn.text = "Select ON" if active else "Select OFF"
                btn.tone = ("accent" if active else "plain") if hasattr(btn, "tone") else getattr(btn, "tone", "plain")
                if hasattr(btn, "_update_graphics"):
                    btn._update_graphics()
            except Exception:
                pass

    def _toggle_mobile_selection_mode(self, *_):
        self.mobile_select_mode = not bool(getattr(self, "mobile_select_mode", False))
        if not bool(getattr(self, "mobile_select_mode", False)) and len(getattr(self.engine, "selected_box_ids", []) or []) > 1:
            self.engine.selected_box_ids = [int(self.engine.selected_box_ids[0])]
        self._refresh_mobile_select_mode_ui()
        self._sync_box_selection_ui()
        self._update_preview_hud()

    def _set_mobile_quick_rail_visible(self, open_state=True):
        self._mobile_quick_rail_open = bool(open_state)
        if self._mobile_quick_rail_open and bool(getattr(self, "mobile_tools_tray_open", False)):
            self._set_mobile_tools_tray_visible(False)
        if self._mobile_quick_rail_open and bool(getattr(self, "_mobile_sidebar_open", False)):
            self._set_mobile_sidebar_state(False)
        rail = getattr(self, "mobile_quick_rail", None)
        if rail is not None:
            rail.opacity = 1 if self._mobile_quick_rail_open else 0
            rail.disabled = not self._mobile_quick_rail_open
            try:
                rail.size_hint = (None, None)
                rail.size = (dp(84), dp(134)) if self._mobile_quick_rail_open else (0, 0)
                if not self._mobile_quick_rail_open:
                    rail.pos = (-10000, -10000)
            except Exception:
                pass
        self._refresh_mobile_hud_restore_button()

    def _toggle_mobile_quick_rail(self, *_):
        self._set_mobile_quick_rail_visible(not bool(getattr(self, "_mobile_quick_rail_open", False)))

    def _clear_mobile_repair_state(self, page_idx=None, clear_selection=False, clear_saved_selection=False):
        try:
            page_hint = int(page_idx) if page_idx is not None else -1
        except Exception:
            page_hint = -1
        self._repair_anchor_box_ids = []
        self._repair_anchor_page_idx = page_hint
        self._repair_last_proposal = None
        try:
            if hasattr(self, 'engine') and getattr(self, 'engine', None) is not None:
                if page_hint >= 0:
                    self.engine.clear_candidate_cache(page_idx=page_hint)
                else:
                    self.engine.clear_candidate_cache()
        except Exception:
            pass
        if clear_selection and hasattr(self, 'engine') and getattr(self, 'engine', None) is not None:
            try:
                self.engine.selected_box_ids = []
            except Exception:
                pass
        if clear_saved_selection and page_hint >= 0 and hasattr(self, '_page_selected_box_ids'):
            try:
                self._page_selected_box_ids.pop(page_hint, None)
            except Exception:
                pass

    def _mobile_repair_overlap_score(self, a, b):
        try:
            ra = a if isinstance(a, fitz.Rect) else fitz.Rect(*a)
            rb = b if isinstance(b, fitz.Rect) else fitz.Rect(*b)
            inter = _rect_intersection_area(ra, rb)
            if inter <= 0:
                return 0.0
            area_a = max(float(ra.width) * float(ra.height), 1e-6)
            area_b = max(float(rb.width) * float(rb.height), 1e-6)
            iou = inter / float(max(area_a + area_b - inter, 1e-6))
            cover = inter / float(max(min(area_a, area_b), 1e-6))
            return float(max(iou, cover))
        except Exception:
            return 0.0

    def _extract_mobile_candidate_features(self, rect, kind, source, zone, local_existing, anchor_rects):
        try:
            rr = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
        except Exception:
            return {}
        kind = str(kind or 'field')
        source = str(source or 'scan')
        zone = zone if isinstance(zone, fitz.Rect) else fitz.Rect(*zone)
        width = max(float(rr.width), 1.0)
        height = max(float(rr.height), 1.0)
        aspect = width / max(height, 1.0)
        anchor_overlap = 0.0
        for ar in list(anchor_rects or []):
            try:
                anchor_overlap = max(anchor_overlap, float(self._mobile_repair_overlap_score(rr, ar)))
            except Exception:
                pass
        same_kind = []
        for item in list(local_existing or []):
            try:
                ex_rect = item[0] if isinstance(item, (list, tuple)) and len(item) >= 1 else None
                ex_kind = str(item[1] if isinstance(item, (list, tuple)) and len(item) >= 2 else '')
                ex_rect = ex_rect if isinstance(ex_rect, fitz.Rect) else fitz.Rect(*ex_rect)
            except Exception:
                continue
            if ex_kind == kind:
                same_kind.append(ex_rect)
        size_similarity = 0.0
        if same_kind:
            med_w = sorted([max(float(r.width), 1.0) for r in same_kind])[len(same_kind)//2]
            med_h = sorted([max(float(r.height), 1.0) for r in same_kind])[len(same_kind)//2]
            dw = min(abs(width - med_w) / max(med_w, 1.0), 1.0)
            dh = min(abs(height - med_h) / max(med_h, 1.0), 1.0)
            size_similarity = max(0.0, 1.0 - ((dw + dh) * 0.5))
        structural_support = 0.0
        support_hits = 0
        for ex_rect in same_kind:
            try:
                ex_cx = (ex_rect.x0 + ex_rect.x1) * 0.5
                ex_cy = (ex_rect.y0 + ex_rect.y1) * 0.5
            except Exception:
                continue
            cx = (rr.x0 + rr.x1) * 0.5
            cy = (rr.y0 + rr.y1) * 0.5
            x_close = abs(ex_cx - cx) <= max(width, ex_rect.width) * 0.55
            y_close = abs(ex_cy - cy) <= max(height, ex_rect.height) * 0.55
            row_like = abs(ex_cy - cy) <= max(height, ex_rect.height) * 0.45
            col_like = abs(ex_cx - cx) <= max(width, ex_rect.width) * 0.45
            if kind == 'line' and row_like:
                support_hits += 1
            elif kind == 'field' and (row_like or col_like or (x_close and y_close)):
                support_hits += 1
            elif kind == 'check' and (row_like or col_like):
                support_hits += 1
        if support_hits > 0:
            structural_support = min(1.0, support_hits / 3.0)
        source_bonus = 0.0
        if source in {'grid_align', 'grid_edge'}:
            source_bonus += 0.04
        if source in {'faint_field', 'partial_box'}:
            source_bonus -= 0.01
        if source in {'red_scan', 'check_scan', 'field_scan', 'line_scan'}:
            source_bonus += 0.02
        aspect_penalty = 0.0
        if kind == 'check':
            if aspect < 0.55 or aspect > 1.85:
                aspect_penalty += 0.08
        elif kind == 'line':
            if aspect < 1.8:
                aspect_penalty += 0.10
        elif kind == 'field':
            if width < 10.0 or height < 5.0:
                aspect_penalty += 0.10
        edge_distance = min(abs(rr.x0 - zone.x0), abs(zone.x1 - rr.x1), abs(rr.y0 - zone.y0), abs(zone.y1 - rr.y1))
        edge_penalty = 0.04 if edge_distance <= 2.0 else 0.0
        return {
            'width': round(width, 3),
            'height': round(height, 3),
            'aspect': round(aspect, 4),
            'anchor_overlap': round(float(anchor_overlap), 6),
            'size_similarity': round(float(size_similarity), 6),
            'structural_support': round(float(structural_support), 6),
            'source_bonus': round(float(source_bonus), 6),
            'aspect_penalty': round(float(aspect_penalty), 6),
            'edge_penalty': round(float(edge_penalty), 6),
        }

    def _compute_mobile_candidate_penalty(self, features):
        try:
            aspect_penalty = float((features or {}).get('aspect_penalty', 0.0) or 0.0)
            edge_penalty = float((features or {}).get('edge_penalty', 0.0) or 0.0)
            anchor_overlap = float((features or {}).get('anchor_overlap', 0.0) or 0.0)
            size_similarity = float((features or {}).get('size_similarity', 0.0) or 0.0)
        except Exception:
            return 0.0
        penalty = aspect_penalty + edge_penalty
        if anchor_overlap < 0.05:
            penalty += 0.03
        if size_similarity < 0.22:
            penalty += 0.03
        return round(float(max(0.0, min(0.22, penalty))), 6)

    def _compute_mobile_candidate_confidence(self, base_score, features):
        try:
            score = float(base_score or 0.0)
        except Exception:
            score = 0.0
        try:
            source_bonus = float((features or {}).get('source_bonus', 0.0) or 0.0)
            anchor_overlap = float((features or {}).get('anchor_overlap', 0.0) or 0.0)
            size_similarity = float((features or {}).get('size_similarity', 0.0) or 0.0)
            structural_support = float((features or {}).get('structural_support', 0.0) or 0.0)
        except Exception:
            source_bonus = 0.0
            anchor_overlap = 0.0
            size_similarity = 0.0
            structural_support = 0.0
        score += source_bonus
        score += min(0.06, anchor_overlap * 0.10)
        score += min(0.05, size_similarity * 0.05)
        score += min(0.05, structural_support * 0.05)
        score -= float(self._compute_mobile_candidate_penalty(features))
        return max(0.0, min(0.99, round(float(score), 6)))


    def _prune_mobile_candidate_redundancy(self, candidates):
        pruned = []
        kind_rank = {'field': 0, 'check': 1, 'line': 2}
        for item in sorted(list(candidates or []), key=lambda entry: (-float(entry.get('score', 0.0) or 0.0), kind_rank.get(str(entry.get('kind', '')), 9), (entry.get('rect').y0 if entry.get('rect') is not None else 0.0), (entry.get('rect').x0 if entry.get('rect') is not None else 0.0))):
            rect = item.get('rect')
            kind = str(item.get('kind', '') or '')
            if rect is None or not kind:
                continue
            try:
                rr = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
            except Exception:
                continue
            keep = True
            for kept in pruned:
                try:
                    kr = kept.get('rect') if isinstance(kept.get('rect'), fitz.Rect) else fitz.Rect(kept.get('rect'))
                except Exception:
                    continue
                kk = str(kept.get('kind', '') or '')
                overlap = float(self._mobile_repair_overlap_score(rr, kr))
                if kk == kind and overlap >= 0.72:
                    keep = False
                    break
                if {kk, kind} == {'field', 'line'} and overlap >= 0.78:
                    kept_score = float(kept.get('score', 0.0) or 0.0)
                    item_score = float(item.get('score', 0.0) or 0.0)
                    # Prefer the stronger candidate unless scores are almost tied and the line is much wider.
                    if kept_score >= item_score + 0.03:
                        keep = False
                        break
                    if item_score <= kept_score + 0.02:
                        kept_w = max(float(kr.width), 1.0)
                        kept_h = max(float(kr.height), 1.0)
                        item_w = max(float(rr.width), 1.0)
                        item_h = max(float(rr.height), 1.0)
                        kept_aspect = kept_w / kept_h
                        item_aspect = item_w / item_h
                        if kind == 'line' and item_aspect >= kept_aspect * 1.35:
                            continue
                        keep = False
                        break
            if keep:
                enriched = dict(item)
                enriched['rect'] = rr
                pruned.append(enriched)
        return pruned

    def _generate_mobile_repair_candidates(self, page_idx, zone, local_existing, local_existing_ids, anchor_ids):
        engine = getattr(self, 'engine', None)
        if engine is None:
            return None
        try:
            img = engine._render_pdf_page_bgr(engine.pdf_path, page_idx=int(page_idx), preview_zoom=ZOOM)
        except Exception:
            img = None
        if img is None or getattr(img, 'size', 0) == 0:
            return None
        if len(img.shape) == 2:
            img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
        h_img, w_img = img.shape[:2]
        full_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        base_bin = cv2.threshold(full_gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        bin_checks = base_bin.copy()
        red_close_k = int(engine.settings.get('C_Close', 0) or 0)
        red_open_k = int(engine.settings.get('C_Open', 0) or 0)
        if red_close_k > 0:
            bin_checks = cv2.morphologyEx(bin_checks, cv2.MORPH_CLOSE, np.ones((red_close_k, red_close_k), np.uint8))
        if red_open_k > 0:
            bin_checks = cv2.morphologyEx(bin_checks, cv2.MORPH_OPEN, np.ones((red_open_k, red_open_k), np.uint8))
        field_close_k = int(engine.settings.get('F_Close', 1) or 1)
        if field_close_k > 0:
            bin_fields = cv2.morphologyEx(base_bin, cv2.MORPH_CLOSE, np.ones((field_close_k, field_close_k), np.uint8))
        else:
            bin_fields = base_bin
        text_like_mask, text_like_meta = engine.build_text_like_mask(img)
        header_cut = int((text_like_meta or {}).get('header_cut', 0) or 0)

        candidates = []
        candidate_type_counts = {'field': 0, 'check': 0, 'line': 0}
        local_existing_rects = []
        try:
            for item in list(local_existing or []):
                if isinstance(item, (list, tuple)) and len(item) >= 1:
                    ex_rect = item[0]
                else:
                    ex_rect = item
                ex_rect = ex_rect if isinstance(ex_rect, fitz.Rect) else fitz.Rect(*ex_rect)
                local_existing_rects.append(ex_rect)
        except Exception:
            local_existing_rects = []
        anchor_rects = []
        try:
            for box_id in list(anchor_ids or []):
                if 0 <= int(box_id) < len(getattr(engine, 'all_boxes', []) or []):
                    anchor_rects.append(engine.all_boxes[int(box_id)])
        except Exception:
            anchor_rects = []
        source_counts = {}
        tuning = {}
        try:
            tuning = getattr(engine, 'get_candidate_tuning', lambda: dict(CANDIDATE_TUNING_DEFAULTS))()
        except Exception:
            tuning = dict(CANDIDATE_TUNING_DEFAULTS)
        min_candidate_score = float((tuning or {}).get('min_candidate_score', 0.34) or 0.34)
        priors = {}
        try:
            priors = getattr(engine, 'get_candidate_learning_priors', lambda page_idx=0: {'pairs': {}, 'same_fp_events': 0})(page_idx=page_idx)
        except Exception:
            priors = {'pairs': {}, 'same_fp_events': 0}

        def _inside_zone(rect):
            return self._mobile_repair_overlap_score(zone, rect) >= 0.10

        def _append_candidate(rect, kind, source, score):
            try:
                rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
            except Exception:
                return False
            if not _inside_zone(rect):
                return False
            kind = str(kind or 'field')
            source = str(source or 'scan')
            try:
                base_score = float(score)
            except Exception:
                base_score = 0.0
            bias = 0.0
            try:
                bias = float(getattr(engine, 'candidate_learning_adjustment', lambda *a, **k: 0.0)(kind, source, page_idx=page_idx, priors=priors) or 0.0)
            except Exception:
                bias = 0.0
            raw_score = max(0.0, min(0.99, base_score + bias))
            features = self._extract_mobile_candidate_features(rect, kind, source, zone, local_existing, anchor_rects)
            score = self._compute_mobile_candidate_confidence(raw_score, features)
            if score < float(min_candidate_score):
                return False
            if engine.should_suppress_text_like_rect(rect, text_like_mask, preview_zoom=ZOOM, header_cut=header_cut, overlap_thresh=0.16):
                return False
            for existing in local_existing_rects:
                if self._mobile_repair_overlap_score(rect, existing) >= 0.50:
                    return False
            for existing in candidates:
                if str(existing.get('kind', '')) != str(kind):
                    continue
                if self._mobile_repair_overlap_score(rect, existing.get('rect')) >= 0.52:
                    merged_sources = list(existing.get('source_set', []) or [str(existing.get('source', '') or '')])
                    if str(source) not in merged_sources:
                        merged_sources.append(str(source))
                    consensus_bonus = self._mobile_candidate_consensus_bonus(merged_sources)
                    updated_score = max(float(score), float(existing.get('score', 0.0) or 0.0))
                    updated_score = min(0.98, updated_score + consensus_bonus)
                    if float(score) > float(existing.get('score', 0.0) or 0.0):
                        existing['rect'] = rect
                        existing['source'] = str(source)
                        existing['raw_score'] = float(raw_score)
                        existing['base_score'] = float(base_score)
                        existing['learning_bias'] = float(bias)
                        existing['features'] = dict(features or {})
                    existing['source_set'] = merged_sources
                    existing['support_count'] = int(len(merged_sources))
                    existing['consensus_bonus'] = float(consensus_bonus)
                    existing['score'] = float(updated_score)
                    return True
            source_list = [str(source)] if str(source) else []
            consensus_bonus = self._mobile_candidate_consensus_bonus(source_list)
            candidates.append({'rect': rect, 'kind': str(kind), 'source': str(source), 'source_set': source_list, 'support_count': int(len(source_list)), 'consensus_bonus': float(consensus_bonus), 'score': float(min(0.98, float(score) + consensus_bonus)), 'raw_score': float(raw_score), 'base_score': float(base_score), 'learning_bias': float(bias), 'features': dict(features or {})})
            candidate_type_counts[str(kind)] = int(candidate_type_counts.get(str(kind), 0) or 0) + 1
            source_counts[str(source)] = int(source_counts.get(str(source), 0) or 0) + 1
            return True

        try:
            line_rects = engine.find_answer_lines(img, ZOOM, min_line_w=int(engine.settings.get('Line_MinW', 100) or 100), max_line_w=int(engine.settings.get('Line_MaxW', 1680) or 1680))
            for lr in line_rects:
                _append_candidate(lr, 'line', 'line_scan', 0.68)
        except Exception:
            pass

        try:
            red_candidates = engine.find_red_checkbox_candidates(
                img,
                zoom_factor=ZOOM,
                min_sz=max(7, int(engine.settings.get('C_Size', (14, 65))[0]) - 5),
                max_sz=min(int(engine.settings.get('C_Size', (14, 65))[1]), 42),
                border_min=max(0.05, float(engine.settings.get('C_Border', 0.10)) * 0.7),
                inner_max=min(0.75, float(engine.settings.get('C_Inner', 0.40)) + 0.20),
                band_pct=max(0.10, min(float(engine.settings.get('C_BandPct', 0.18)), 0.16)),
                aspect_tol=max(0.30, float(engine.settings.get('C_AspectTol', 0.10))),
                fill_min=max(0.12, float(engine.settings.get('C_FillMin', 0.45)) * 0.45),
                eps=max(0.04, float(engine.settings.get('C_Eps', 0.04))),
            )
            for rr in red_candidates:
                _append_candidate(rr, 'check', 'red_scan', 0.79)
        except Exception:
            pass

        try:
            checkbox_min_sz = int(engine.settings.get('C_Size', (14, 65))[0])
            checkbox_max_sz = int(engine.settings.get('C_Size', (14, 65))[1])
            roi_max = int(engine.settings.get('ROI_Max', 200) or 200)
            nf, _, sf, _ = cv2.connectedComponentsWithStats(bin_checks)
            small_min = max(7, checkbox_min_sz - 5)
            for i in range(1, nf):
                x, y, w, h, area = sf[i]
                if w >= w_img * 0.9:
                    continue
                rect = fitz.Rect(x / ZOOM, y / ZOOM, (x + w) / ZOOM, (y + h) / ZOOM)
                if not _inside_zone(rect):
                    continue
                aspect = w / float(h) if h > 0 else 999.0
                if small_min <= w <= checkbox_max_sz and small_min <= h <= checkbox_max_sz and 0.68 <= aspect <= 1.45:
                    if engine.looks_like_checkbox(bin_checks, x, y, w, h, area):
                        _append_candidate(rect, 'check', 'check_scan', 0.71)
                elif max(w, h) <= roi_max and min(w, h) >= small_min and 0.55 <= aspect <= 1.75:
                    for fx, fy, fw, fh in engine.find_checkbox_rects_in_roi(bin_checks, x, y, w, h):
                        _append_candidate(fitz.Rect(fx / ZOOM, fy / ZOOM, (fx + fw) / ZOOM, (fy + fh) / ZOOM), 'check', 'check_roi', 0.66)
        except Exception:
            pass

        try:
            nf2, _, sf2, _ = cv2.connectedComponentsWithStats(bin_fields)
            field_area_thresh = int(engine.settings.get('F_Area', 500) or 500)
            field_min_w_val = int(engine.settings.get('F_MinW', 15) or 15)
            field_min_h_val = int(engine.settings.get('F_MinH', 35) or 35)
            for i in range(1, nf2):
                x, y, w, h, area = sf2[i]
                if w >= w_img * 0.90 or h >= h_img * 0.12 or area >= (w_img * h_img * 0.03):
                    continue
                if (w > 1200 and h > 120) or (w > 1600 and h > 80):
                    continue
                if h < field_min_h_val or w < field_min_w_val or area < field_area_thresh:
                    continue
                refined = engine._refine_field_rect_from_mask(bin_fields, x, y, w, h, zoom_factor=ZOOM, row_frac_thresh=0.45, col_frac_thresh=0.18, min_inner_h=max(18, int(field_min_h_val * 0.45)), pad_px=2)
                _append_candidate(refined, 'field', 'field_scan', 0.74)
        except Exception:
            pass

        try:
            faint_field_candidates = engine.detect_faint_field_candidates(
                img,
                zoom_factor=ZOOM,
                zone_rect=zone,
                text_mask=text_like_mask,
                header_cut=header_cut,
            )
            for item in faint_field_candidates:
                _append_candidate(item.get('rect'), 'field', item.get('source', 'faint_field'), float(item.get('score', 0.56) or 0.56))
        except Exception:
            pass

        try:
            partial_box_candidates = engine.detect_partial_box_candidates(
                img,
                zoom_factor=ZOOM,
                zone_rect=zone,
                text_mask=text_like_mask,
                header_cut=header_cut,
            )
            for item in partial_box_candidates:
                _append_candidate(item.get('rect'), 'field', item.get('source', 'partial_box'), float(item.get('score', 0.52) or 0.52))
        except Exception:
            pass

        try:
            option_candidates = engine.detect_non_square_option_candidates(
                img,
                zoom_factor=ZOOM,
                zone_rect=zone,
                text_mask=text_like_mask,
                header_cut=header_cut,
            )
            for item in option_candidates:
                _append_candidate(item.get('rect'), 'check', item.get('source', 'radio_like'), float(item.get('score', 0.54) or 0.54))
        except Exception:
            pass

        candidates = self._prune_mobile_candidate_redundancy(candidates)
        candidates.sort(key=lambda item: (-float(item.get('score', 0.0) or 0.0), {'field': 0, 'check': 1, 'line': 2}.get(str(item.get('kind', '')), 9), item['rect'].y0, item['rect'].x0))
        local_seed_pairs = []
        for existing in list(local_existing or []):
            if isinstance(existing, (list, tuple)) and len(existing) >= 2:
                try:
                    ex_rect = existing[0] if isinstance(existing[0], fitz.Rect) else fitz.Rect(*existing[0])
                except Exception:
                    continue
                ex_kind = str(existing[1] or '')
                local_seed_pairs.append((ex_rect, ex_kind))
        inferred_structured_count = self._infer_mobile_grid_like_candidates(zone, local_seed_pairs, candidates, _append_candidate)
        candidates = self._prune_mobile_candidate_redundancy(candidates)
        candidates.sort(key=lambda item: (-float(item.get('score', 0.0) or 0.0), {'field': 0, 'check': 1, 'line': 2}.get(str(item.get('kind', '')), 9), item['rect'].y0, item['rect'].x0))
        return {
            'candidates': candidates,
            'candidate_type_counts': dict(candidate_type_counts),
            'source_counts': dict(source_counts),
            'inferred_structured_count': int(inferred_structured_count or 0),
            'zone_rect': zone,
            'anchor_ids': list(anchor_ids or []),
            'learning_prior_pairs': int(len(dict((priors or {}).get('pairs', {}) or {}))),
            'learning_prior_same_fp_events': int((priors or {}).get('same_fp_events', 0) or 0),
            'candidate_min_score': float(min_candidate_score),
        }

    def _build_mobile_repair_proposal(self, anchor_ids, page_idx=0):
        page_idx = int(page_idx or 0)
        anchor_ids = sorted(set(int(x) for x in (anchor_ids or []) if isinstance(x, int) or str(x).isdigit()))
        if not anchor_ids:
            return None
        engine = getattr(self, 'engine', None)
        if engine is None:
            return None
        if int(getattr(engine, 'detected_page_idx', -1) or -1) != page_idx:
            engine._restore_detection_for_page(page_idx, required_signature=engine._settings_signature())
        anchor_rects = []
        for box_id in anchor_ids:
            if 0 <= int(box_id) < len(getattr(engine, 'all_boxes', []) or []):
                try:
                    anchor_rects.append(engine.all_boxes[int(box_id)])
                except Exception:
                    pass
        if not anchor_rects:
            return None

        zone = fitz.Rect(anchor_rects[0])
        for rect in anchor_rects[1:]:
            zone = fitz.Rect(min(zone.x0, rect.x0), min(zone.y0, rect.y0), max(zone.x1, rect.x1), max(zone.y1, rect.y1))
        pad_x = max(18.0, float(zone.width) * 0.65)
        pad_y = max(18.0, float(zone.height) * 1.00)
        zone = fitz.Rect(zone.x0 - pad_x, zone.y0 - pad_y, zone.x1 + pad_x, zone.y1 + pad_y)

        local_existing = []
        local_existing_ids = []
        local_type_counts = {'field': 0, 'check': 0, 'line': 0}
        all_boxes = list(getattr(engine, 'all_boxes', []) or [])
        all_types = list(getattr(engine, 'box_types', []) or [])
        for idx_box, rect in enumerate(all_boxes):
            try:
                if self._mobile_repair_overlap_score(zone, rect) >= 0.08:
                    t = str(all_types[idx_box] if idx_box < len(all_types) else '')
                    local_existing.append((rect, t))
                    local_existing_ids.append(idx_box)
                    if t in local_type_counts:
                        local_type_counts[t] += 1
            except Exception:
                pass

        candidate_payload = self._generate_mobile_repair_candidates(
            page_idx=page_idx,
            zone=zone,
            local_existing=local_existing,
            local_existing_ids=local_existing_ids,
            anchor_ids=anchor_ids,
        )
        if candidate_payload is None:
            return None
        candidates = list(candidate_payload.get('candidates', []) or [])
        candidate_type_counts = dict(candidate_payload.get('candidate_type_counts', {}) or {})
        source_counts = dict(candidate_payload.get('source_counts', {}) or {})
        inferred_structured_count = int(candidate_payload.get('inferred_structured_count', 0) or 0)
        review_defaults = {}
        for idx_item, item in enumerate(candidates):
            bucket = self._mobile_repair_confidence_bucket(float(item.get('score', 0.0) or 0.0))
            review_defaults[f'add:{idx_item}'] = bool(bucket != 'weak')

        removal_candidates = []
        removal_type_counts = {'field': 0, 'check': 0, 'line': 0}
        removal_source_counts = {}

        def _append_removal(existing_idx, rect, kind, reason, score):
            key = f"{int(existing_idx)}:{str(kind)}"
            for existing in removal_candidates:
                if str(existing.get('review_key', '')) == key:
                    if float(score) > float(existing.get('score', 0.0) or 0.0):
                        existing['reason'] = str(reason)
                        existing['score'] = float(score)
                    return
            removal_candidates.append({
                'review_key': key,
                'existing_index': int(existing_idx),
                'rect': rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect),
                'kind': str(kind),
                'source': 'cleanup_local',
                'reason': str(reason),
                'score': float(score),
            })
            removal_type_counts[str(kind)] = int(removal_type_counts.get(str(kind), 0) or 0) + 1
            removal_source_counts['cleanup_local'] = int(removal_source_counts.get('cleanup_local', 0) or 0) + 1

        local_entries = []
        for idx_existing, existing in zip(local_existing_ids, local_existing):
            if isinstance(existing, (list, tuple)) and len(existing) >= 1:
                rect = existing[0]
                inferred_kind = str(existing[1] if len(existing) >= 2 else '')
            else:
                rect = existing
                inferred_kind = ''
            try:
                rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
            except Exception:
                continue
            kind = str(inferred_kind or (all_types[idx_existing] if idx_existing < len(all_types) else ''))
            local_entries.append({'idx': int(idx_existing), 'rect': rect, 'kind': kind})

        anchor_set = set(int(x) for x in anchor_ids)
        # Conservative fragment/line-conflict removals, mirroring existing cleanup logic
        for entry in local_entries:
            idx_existing = int(entry['idx'])
            if idx_existing in anchor_set:
                continue
            rect = entry['rect']
            kind = str(entry['kind'] or '')
            if kind == 'field':
                for other in local_entries:
                    if other['idx'] == idx_existing or str(other['kind'] or '') != 'field':
                        continue
                    sr = rect
                    br = other['rect']
                    s_area = max(float(sr.width) * float(sr.height), 1e-6)
                    b_area = max(float(br.width) * float(br.height), 1e-6)
                    if b_area <= s_area or br.height < sr.height:
                        continue
                    x_overlap = _x_overlap_ratio(sr, br)
                    gap = br.y0 - sr.y1
                    inter = _rect_intersection_area(sr, br)
                    cover_small = inter / max(s_area, 1e-6)
                    if x_overlap >= 0.80 and 0 <= gap <= 6.0 and sr.height <= max(4.0, br.height * 0.45):
                        _append_removal(idx_existing, sr, kind, 'small fragment under larger field', 0.78)
                        break
                    if inter > 0 and cover_small >= 0.75 and sr.y1 <= (br.y0 + br.height * 0.45):
                        _append_removal(idx_existing, sr, kind, 'field fragment covered by larger field', 0.76)
                        break
            elif kind == 'line':
                la = max(float(rect.width) * float(rect.height), 1e-6)
                lmid_y = (rect.y0 + rect.y1) / 2.0
                line_mid_x = (rect.x0 + rect.x1) / 2.0
                for other in local_entries:
                    if str(other['kind'] or '') != 'field':
                        continue
                    fr = other['rect']
                    inter = _rect_intersection_area(rect, fr)
                    overlap_ratio = inter / max(la, 1e-6)
                    x_overlap = _x_intersection(rect, fr) / max(rect.width, 1e-6)
                    gap_above = fr.y0 - rect.y1
                    gap_below = rect.y0 - fr.y1
                    near_top_band = fr.y0 <= rect.y1 <= (fr.y0 + fr.height * 0.55)
                    if overlap_ratio >= 0.35:
                        _append_removal(idx_existing, rect, kind, 'line conflicts with nearby field', 0.74)
                        break
                    if x_overlap >= 0.70 and (fr.y0 - 1.5) <= lmid_y <= (fr.y0 + fr.height * 0.60):
                        _append_removal(idx_existing, rect, kind, 'line sits inside field top band', 0.71)
                        break
                    if x_overlap >= 0.70 and (0 <= gap_above <= 4.0 or 0 <= gap_below <= 4.0):
                        _append_removal(idx_existing, rect, kind, 'line hugs nearby field edge', 0.69)
                        break
                    if (fr.x0 <= line_mid_x <= fr.x1) and near_top_band:
                        _append_removal(idx_existing, rect, kind, 'line center falls inside field', 0.67)
                        break

        removal_candidates.sort(key=lambda item: (-float(item.get('score', 0.0) or 0.0), {'field': 0, 'line': 1, 'check': 2}.get(str(item.get('kind', '')), 9), item['rect'].y0, item['rect'].x0))
        for idx_item, item in enumerate(removal_candidates):
            bucket = self._mobile_repair_confidence_bucket(float(item.get('score', 0.0) or 0.0))
            review_defaults[f'remove:{idx_item}'] = bool(bucket == 'strong')
        try:
            engine.cache_candidates_for_page(
                page_idx,
                candidates=candidates,
                review_state=review_defaults,
                anchor_ids=anchor_ids,
                meta={
                    'anchor_ids': list(anchor_ids),
                    'existing_local_ids': list(local_existing_ids),
                    'existing_local_count': int(len(local_existing)),
                    'zone_rect': [float(zone.x0), float(zone.y0), float(zone.x1), float(zone.y1)],
                    'inferred_structured_count': int(inferred_structured_count or 0),
                    'candidate_type_counts': dict(candidate_type_counts),
                    'source_counts': dict(source_counts),
                    'detection_signature': str(getattr(engine, 'detected_settings_signature', '') or getattr(engine, '_settings_signature', lambda: '')() or ''),
                    'fingerprint_key': str((((getattr(engine, 'current_detection_context', {}) or {}).get('fingerprint') or {}).get('fingerprint_key', '') or '')),
                    'removal_review_keys': [f'remove:{idx}' for idx in range(len(removal_candidates))],
                    'removal_candidates': [{
                        'existing_index': int(item.get('existing_index', -1) or -1),
                        'rect': [float(item['rect'].x0), float(item['rect'].y0), float(item['rect'].x1), float(item['rect'].y1)] if item.get('rect') is not None else [],
                        'kind': str(item.get('kind', 'field') or 'field'),
                        'reason': str(item.get('reason', '') or ''),
                        'score': float(item.get('score', 0.0) or 0.0),
                        'source': str(item.get('source', 'cleanup') or 'cleanup'),
                    } for item in removal_candidates],
                },
            )
        except Exception:
            pass
        return {
            'page_idx': int(page_idx),
            'anchor_ids': list(anchor_ids),
            'anchor_rects': list(anchor_rects),
            'zone_rect': zone,
            'existing_local_ids': list(local_existing_ids),
            'existing_local_count': int(len(local_existing_ids)),
            'existing_local_type_counts': dict(local_type_counts),
            'candidates': candidates,
            'candidate_type_counts': dict(candidate_type_counts),
            'source_counts': dict(source_counts),
            'inferred_structured_count': int(inferred_structured_count or 0),
            'removal_candidates': removal_candidates,
            'removal_type_counts': dict(removal_type_counts),
            'removal_source_counts': dict(removal_source_counts),
        }

    def cluster_grid_like_alignment(self, items):
        rows = []
        try:
            ordered = sorted(list(items or []), key=lambda s: ((((s.get("rect").y0 + s.get("rect").y1) * 0.5) if isinstance(s.get("rect"), fitz.Rect) else 0.0), (s.get("rect").x0 if isinstance(s.get("rect"), fitz.Rect) else 0.0)))
        except Exception:
            ordered = list(items or [])
        for item in ordered:
            rect = item.get('rect')
            try:
                rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
            except Exception:
                continue
            cy = (rect.y0 + rect.y1) * 0.5
            placed = False
            for row in rows:
                tol = max(8.0, float(row.get('med_h', max(rect.height, 1.0))) * 0.75)
                if abs(cy - float(row.get('cy', cy))) <= tol:
                    row['items'].append(dict(item, rect=rect))
                    vals = [((it['rect'].y0 + it['rect'].y1) * 0.5) for it in row['items']]
                    hs = [max(float(it['rect'].height), 1.0) for it in row['items']]
                    ws = [max(float(it['rect'].width), 1.0) for it in row['items']]
                    row['cy'] = sum(vals) / max(len(vals), 1)
                    row['med_h'] = sorted(hs)[len(hs)//2]
                    row['med_w'] = sorted(ws)[len(ws)//2]
                    placed = True
                    break
            if not placed:
                row_h = max(float(rect.height), 1.0)
                row_w = max(float(rect.width), 1.0)
                rows.append({'cy': cy, 'med_h': row_h, 'med_w': row_w, 'items': [dict(item, rect=rect)]})
        for row in rows:
            row['items'] = sorted(row.get('items', []), key=lambda s: s['rect'].x0)
            if row['items']:
                widths = [max(float(it['rect'].width), 1.0) for it in row['items']]
                heights = [max(float(it['rect'].height), 1.0) for it in row['items']]
                row['med_w'] = sorted(widths)[len(widths)//2]
                row['med_h'] = sorted(heights)[len(heights)//2]
        return rows

    def score_grid_slot_candidate(self, inferred_rect, kind='field', med_w=0.0, med_h=0.0, gap=0.0, edge=False):
        try:
            rect = inferred_rect if isinstance(inferred_rect, fitz.Rect) else fitz.Rect(*inferred_rect)
        except Exception:
            return 0.0
        width = max(float(rect.width), 1.0)
        height = max(float(rect.height), 1.0)
        med_w = max(float(med_w or width), 1.0)
        med_h = max(float(med_h or height), 1.0)
        ratio_penalty = min(abs(width - med_w) / med_w, 1.0) * 0.08
        height_penalty = min(abs(height - med_h) / med_h, 1.0) * 0.06
        gap_penalty = 0.0
        if gap:
            gap_penalty = min(abs(float(gap) - med_w) / max(med_w, 1.0), 1.0) * 0.07
        base = 0.66 if str(kind or 'field') == 'field' else 0.61
        if edge:
            base -= 0.05
        score = base - ratio_penalty - height_penalty - gap_penalty
        return max(0.42, min(0.88, float(score)))

    def infer_missing_slots_from_alignment(self, rows, kind, zone, append_candidate):
        inferred = 0
        try:
            zone = zone if isinstance(zone, fitz.Rect) else fitz.Rect(*zone)
        except Exception:
            return 0
        for row in list(rows or []):
            row_items = sorted(list(row.get('items', []) or []), key=lambda s: s['rect'].x0)
            if len(row_items) < 2:
                continue
            med_w = max(float(row.get('med_w', 0.0) or 0.0), 1.0)
            med_h = max(float(row.get('med_h', 0.0) or 0.0), 1.0)
            if med_w <= 6 or med_h <= 4:
                continue
            for left, right in zip(row_items, row_items[1:]):
                lr = left['rect']; rr = right['rect']
                gap = float(rr.x0 - lr.x1)
                if gap < med_w * 0.35 or gap > med_w * 2.40:
                    continue
                x0 = lr.x1 + max(gap * 0.12, med_w * 0.08)
                x1 = rr.x0 - max(gap * 0.12, med_w * 0.08)
                if x1 <= x0:
                    continue
                inferred_rect = fitz.Rect(x0, min(lr.y0, rr.y0), x1, max(lr.y1, rr.y1))
                if inferred_rect.width < med_w * 0.52:
                    continue
                score = self.score_grid_slot_candidate(inferred_rect, kind=kind, med_w=med_w, med_h=med_h, gap=gap, edge=False)
                if append_candidate(inferred_rect, kind, 'grid_align', score):
                    inferred += 1
            if len(row_items) >= 3:
                centers = [((it['rect'].x0 + it['rect'].x1) * 0.5) for it in row_items]
                steps = [centers[i + 1] - centers[i] for i in range(len(centers) - 1)]
                steps = [s for s in steps if med_w * 0.9 <= s <= med_w * 3.1]
                if len(steps) >= 2:
                    step = sorted(steps)[len(steps)//2]
                    left_rect = row_items[0]['rect']
                    right_rect = row_items[-1]['rect']
                    for edge_rect, direction in ((left_rect, -1.0), (right_rect, 1.0)):
                        cx = ((edge_rect.x0 + edge_rect.x1) * 0.5) + (step * direction)
                        inferred_rect = fitz.Rect(cx - (med_w * 0.5), edge_rect.y0, cx + (med_w * 0.5), edge_rect.y1)
                        if inferred_rect.x0 < zone.x0 or inferred_rect.x1 > zone.x1:
                            continue
                        score = self.score_grid_slot_candidate(inferred_rect, kind=kind, med_w=med_w, med_h=med_h, gap=step, edge=True)
                        if append_candidate(inferred_rect, kind, 'grid_edge', score):
                            inferred += 1
        return int(inferred)

    def _infer_mobile_grid_like_candidates(self, zone, local_existing, existing_candidates, append_candidate):
        try:
            seeds = []
            for rect, kind in list(local_existing or []):
                kind = str(kind or '')
                if kind not in {'field', 'line'}:
                    continue
                try:
                    rr = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
                except Exception:
                    continue
                seeds.append({'rect': rr, 'kind': kind, 'seed': 'existing'})
            for cand in list(existing_candidates or []):
                if not isinstance(cand, dict):
                    continue
                kind = str(cand.get('kind', '') or '')
                if kind not in {'field', 'line'}:
                    continue
                rect = cand.get('rect')
                try:
                    rr = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
                except Exception:
                    continue
                seeds.append({'rect': rr, 'kind': kind, 'seed': 'candidate'})
            if len(seeds) < 2:
                return 0

            inferred = 0
            zone = zone if isinstance(zone, fitz.Rect) else fitz.Rect(*zone)
            for kind in ('field', 'line'):
                items = [s for s in seeds if s.get('kind') == kind]
                if len(items) < 2:
                    continue
                rows = self.cluster_grid_like_alignment(items)
                inferred += int(self.infer_missing_slots_from_alignment(rows, kind, zone, append_candidate) or 0)
            return int(inferred)
        except Exception:
            return 0

    def _mobile_repair_confidence_bucket(self, score):
        try:
            score = float(score)
        except Exception:
            score = 0.0
        tuning = {}
        try:
            tuning = getattr(self.engine, 'get_candidate_tuning', lambda: dict(CANDIDATE_TUNING_DEFAULTS))()
        except Exception:
            tuning = dict(CANDIDATE_TUNING_DEFAULTS)
        strong_score = float((tuning or {}).get('strong_score', 0.76) or 0.76)
        moderate_score = float((tuning or {}).get('moderate_score', 0.58) or 0.58)
        if score >= strong_score:
            return 'strong'
        if score >= moderate_score:
            return 'moderate'
        return 'weak'


    def _mobile_candidate_consensus_bonus(self, sources):
        try:
            uniq = []
            for src in list(sources or []):
                s = str(src or '').strip()
                if s and s not in uniq:
                    uniq.append(s)
        except Exception:
            uniq = []
        extra = max(0, len(uniq) - 1)
        return round(float(min(0.09, extra * 0.03)), 6)

    def _build_mobile_repair_review_payload(self, proposal):
        if not isinstance(proposal, dict):
            return None
        review_items = []
        add_counts = {'strong': 0, 'moderate': 0, 'weak': 0}
        remove_counts = {'strong': 0, 'moderate': 0, 'weak': 0}
        candidates = list(proposal.get('candidates', []) or [])
        removal_candidates = list(proposal.get('removal_candidates', []) or [])
        stored_review_state = {}
        try:
            page_idx = int(proposal.get('page_idx', 0) or 0)
            proposal_anchor_ids = [int(x) for x in (proposal.get('anchor_ids', []) or []) if isinstance(x, int) or str(x).isdigit()]
            stored_review_state = dict(getattr(self.engine, 'get_candidates_for_page')(page_idx, anchor_ids=proposal_anchor_ids).get('review_state', {}) or {})
        except Exception:
            stored_review_state = {}
        tuning = {}
        try:
            tuning = getattr(self.engine, 'get_candidate_tuning', lambda: dict(CANDIDATE_TUNING_DEFAULTS))()
        except Exception:
            tuning = dict(CANDIDATE_TUNING_DEFAULTS)
        auto_check_add_score = float((tuning or {}).get('auto_check_add_score', 0.58) or 0.58)
        auto_check_remove_score = float((tuning or {}).get('auto_check_remove_score', 0.76) or 0.76)
        for idx_item, item in enumerate(candidates):
            kind = str(item.get('kind', '') or 'field')
            source = str(item.get('source', '') or 'scan')
            score = float(item.get('score', 0.0) or 0.0)
            bucket = self._mobile_repair_confidence_bucket(score)
            add_counts[bucket] = int(add_counts.get(bucket, 0) or 0) + 1
            default_checked = bool(stored_review_state.get(f'add:{idx_item}', score >= auto_check_add_score))
            source_set = [str(s) for s in (item.get('source_set', []) or []) if str(s).strip()]
            support_count = int(item.get('support_count', len(source_set) if source_set else 1) or 1)
            label = f"Add {kind} from {source.replace('_', ' ')}"
            if source in {'grid_infer', 'grid_align', 'grid_edge'}:
                label = f"Add {kind} from grid-like structure"
            elif support_count > 1 or len(source_set) > 1:
                merged_names = ', '.join(s.replace('_', ' ') for s in source_set[:3])
                label = f"Add {kind} from combined evidence"
                if merged_names:
                    label += f" • {merged_names}"
            review_items.append({
                'review_key': f'add:{idx_item}',
                'action': 'add',
                'candidate_index': int(idx_item),
                'kind': kind,
                'source': source,
                'score': score,
                'bucket': bucket,
                'label': label,
                'support_count': int(support_count),
                'source_set': list(source_set),
                'checked': bool(default_checked),
            })
        for idx_item, item in enumerate(removal_candidates):
            kind = str(item.get('kind', '') or 'field')
            source = str(item.get('source', '') or 'cleanup_local')
            score = float(item.get('score', 0.0) or 0.0)
            bucket = self._mobile_repair_confidence_bucket(score)
            remove_counts[bucket] = int(remove_counts.get(bucket, 0) or 0) + 1
            default_checked = bool(stored_review_state.get(f'remove:{idx_item}', score >= auto_check_remove_score))
            label = f"Remove {kind} • {str(item.get('reason', '') or 'local interference')}"
            review_items.append({
                'review_key': f'remove:{idx_item}',
                'action': 'remove',
                'candidate_index': int(idx_item),
                'existing_index': int(item.get('existing_index', -1) or -1),
                'kind': kind,
                'source': source,
                'score': score,
                'bucket': bucket,
                'label': label,
                'checked': bool(default_checked),
            })
        out = dict(proposal)
        review_items.sort(key=lambda item: (0 if str(item.get('action')) == 'add' else 1, {'strong': 0, 'moderate': 1, 'weak': 2}.get(str(item.get('bucket', 'weak')), 9), -float(item.get('score', 0.0) or 0.0), str(item.get('kind', ''))))
        out['review_items'] = review_items
        out['review_count'] = len(review_items)
        out['review_add_bucket_counts'] = add_counts
        out['review_remove_bucket_counts'] = remove_counts
        return out

    def _open_mobile_repair_review_popup(self, proposal):
        proposal = self._build_mobile_repair_review_payload(proposal)
        if not isinstance(proposal, dict):
            self.set_status('No repair review proposal available.', kind='info', hold_seconds=2.0, force=True)
            return
        items = list(proposal.get('review_items', []) or [])
        if not items:
            self.set_status('Repair analysis found no new local changes to review.', kind='info', hold_seconds=2.2, force=True)
            return

        palette = getattr(self, 'ui_palette', {}) or {}
        outer = BoxLayout(orientation='vertical', spacing=dp(10), padding=[dp(14), dp(14), dp(14), dp(14)])
        self._style_popup_card(outer, palette.get('surface', (0.085, 0.105, 0.145, 1)), radius=dp(22))

        title = Label(text='Review Repair Changes', color=palette.get('text', (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(28), halign='left', valign='middle', font_size=dp(18), bold=True)
        title.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(title, min_height=dp(28), extra_pad=dp(4))
        outer.add_widget(title)

        add_counts = dict(proposal.get('review_add_bucket_counts', {}) or {})
        remove_counts = dict(proposal.get('review_remove_bucket_counts', {}) or {})
        sub = Label(
            text=(
                f"Additions — strong {int(add_counts.get('strong', 0) or 0)} • moderate {int(add_counts.get('moderate', 0) or 0)} • weak {int(add_counts.get('weak', 0) or 0)}\n"
                f"Removals — strong {int(remove_counts.get('strong', 0) or 0)} • moderate {int(remove_counts.get('moderate', 0) or 0)} • weak {int(remove_counts.get('weak', 0) or 0)}"
            ),
            color=palette.get('muted', (0.60, 0.68, 0.80, 1)),
            size_hint_y=None,
            halign='left',
            valign='middle',
            font_size=dp(11)
        )
        sub.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(sub, min_height=dp(34), extra_pad=dp(6))
        outer.add_widget(sub)

        state = {}
        counter_lbl = Label(text='', color=palette.get('muted', (0.60, 0.68, 0.80, 1)), size_hint_y=None, halign='left', valign='middle', font_size=dp(10.5))
        counter_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(counter_lbl, min_height=dp(18), extra_pad=dp(4))
        outer.add_widget(counter_lbl)

        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False, bar_width=dp(6))
        grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        grid.bind(minimum_height=grid.setter('height'))
        for item in items[:160]:
            row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint_y=None, padding=[dp(10), dp(8), dp(10), dp(8)])
            row.bind(minimum_height=row.setter('height'))
            self._style_popup_card(row, palette.get('surface_alt', (0.11, 0.135, 0.185, 1)), radius=dp(16))
            chk = CheckBox(active=bool(item.get('checked', False)), size_hint=(None, None), size=(dp(24), dp(24)))
            state[str(item.get('review_key'))] = chk
            row.add_widget(chk)
            body = BoxLayout(orientation='vertical', spacing=dp(2), size_hint_y=None)
            body.bind(minimum_height=body.setter('height'))
            bucket = str(item.get('bucket', 'weak') or 'weak')
            score = float(item.get('score', 0.0) or 0.0)
            action = str(item.get('action', 'add') or 'add')
            bucket_text = bucket.capitalize()
            line1 = Label(
                text=f"{item.get('label', 'Change')} • {bucket_text} {score:.2f}",
                color=palette.get('text', (0.93, 0.96, 1.0, 1)),
                size_hint_y=None,
                halign='left',
                valign='middle',
                font_size=dp(11.2),
                bold=True
            )
            line1.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(line1, min_height=dp(18), extra_pad=dp(4))
            body.add_widget(line1)
            detail = str(item.get('source', '')).replace('_', ' ')
            if action == 'remove':
                detail = f"Proposed local cleanup • {detail}"
            else:
                detail = f"Proposed addition • {detail}"
            line2 = Label(text=detail, color=palette.get('muted', (0.60, 0.68, 0.80, 1)), size_hint_y=None, halign='left', valign='middle', font_size=dp(10.2))
            line2.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(line2, min_height=dp(16), extra_pad=dp(4))
            body.add_widget(line2)
            row.add_widget(body)
            grid.add_widget(row)
        scroll.add_widget(grid)
        outer.add_widget(scroll)

        btn_bar = GridLayout(cols=4, spacing=dp(8), size_hint_y=None, height=dp(44))
        btn_all = self._make_compact_action_button('Select All', tone='ghost')
        btn_none = self._make_compact_action_button('Select None', tone='ghost')
        btn_close = self._make_compact_action_button('Close', tone='plain')
        btn_apply = self._make_compact_action_button('Apply Checked', tone='primary')
        btn_bar.add_widget(btn_all)
        btn_bar.add_widget(btn_none)
        btn_bar.add_widget(btn_close)
        btn_bar.add_widget(btn_apply)
        outer.add_widget(btn_bar)

        popup = Popup(title='', separator_height=0, background='', background_color=(0, 0, 0, 0.80), content=outer, size_hint=(0.96 if getattr(self, 'ui_mobile', False) else 0.82, 0.88 if getattr(self, 'ui_mobile', False) else 0.74), auto_dismiss=True)
        self._bind_popup_background_softening(popup)

        def _persist_review_state(*_):
            try:
                page_idx_local = int(proposal.get('page_idx', self.current_page_idx()) or 0)
                review_snapshot = {str(key): bool(getattr(chk, 'active', False)) for key, chk in state.items()}
                self.engine.update_candidate_review_state_for_page(page_idx_local, review_snapshot, anchor_ids=list(proposal.get('anchor_ids', []) or []))
            except Exception:
                pass

        def _refresh_counter(*_):
            active = sum(1 for chk in state.values() if bool(getattr(chk, 'active', False)))
            inactive = max(0, len(state) - active)
            counter_lbl.text = f"Checked changes: {active} • Unchecked: {inactive}"
            _persist_review_state()
        for chk in state.values():
            chk.bind(active=lambda *_: _refresh_counter())
        _refresh_counter()

        def _set_all(active):
            for chk in state.values():
                chk.active = bool(active)
            _refresh_counter()

        btn_all.bind(on_release=lambda *_: _set_all(True))
        btn_none.bind(on_release=lambda *_: _set_all(False))
        btn_close.bind(on_release=lambda *_: popup.dismiss())
        btn_apply.bind(on_release=lambda *_: self._apply_mobile_repair_review(proposal, state, popup=popup))
        popup.open()

    def _apply_mobile_repair_review(self, proposal, review_state, popup=None):
        if not isinstance(proposal, dict):
            self.set_status('No repair proposal to apply.', kind='warning', hold_seconds=2.0, force=True)
            return
        engine = getattr(self, 'engine', None)
        if engine is None:
            self.set_status('Repair apply failed: engine unavailable.', kind='warning', hold_seconds=2.0, force=True)
            return
        page_idx = int(proposal.get('page_idx', self.current_page_idx()) or 0)
        detection_ready = False
        try:
            if int(getattr(engine, 'detected_page_idx', -1) or -1) == page_idx and bool(getattr(engine, 'all_boxes', []) or []):
                detection_ready = True
            else:
                detection_ready = bool(engine._restore_detection_for_page(page_idx, required_signature=engine._settings_signature()))
        except Exception:
            detection_ready = False
        if not detection_ready:
            self.set_status('Repair apply requires a detected page. Run Detect again first.', kind='warning', hold_seconds=2.6, force=True)
            return
        candidates = list(proposal.get('candidates', []) or [])
        removal_candidates = list(proposal.get('removal_candidates', []) or [])
        review_items = list(proposal.get('review_items', []) or [])
        selected_candidate_indices = []
        selected_remove_indices = []
        reviewed_decisions = []
        for item in review_items:
            review_key = str(item.get('review_key') or '')
            chk = review_state.get(review_key)
            is_active = bool(getattr(chk, 'active', False)) if chk is not None else bool(item.get('default_checked', False))
            action = str(item.get('action') or '')
            try:
                idx = int(item.get('candidate_index'))
            except Exception:
                idx = -1
            reviewed_decisions.append({
                'review_key': review_key,
                'action': action,
                'accepted': bool(is_active),
                'kind': str(item.get('kind', '') or ''),
                'source': str(item.get('source', '') or ''),
                'score': float(item.get('score', 0.0) or 0.0),
                'reason': str(item.get('reason', '') or ''),
                'rect': item.get('rect'),
            })
            if not is_active:
                continue
            if idx < 0:
                continue
            if action == 'add':
                selected_candidate_indices.append(idx)
            elif action == 'remove':
                selected_remove_indices.append(idx)
        selected_candidate_indices = sorted(set(selected_candidate_indices))
        selected_remove_indices = sorted(set(selected_remove_indices))
        if not selected_candidate_indices and not selected_remove_indices:
            self.set_status('No repair changes were checked.', kind='info', hold_seconds=2.0, force=True)
            return

        all_boxes = list(getattr(engine, 'all_boxes', []) or [])
        all_types = list(getattr(engine, 'box_types', []) or [])
        added_rects = []
        removed_ids = []
        added_count = 0

        anchor_ids = [int(x) for x in (proposal.get('anchor_ids', []) or []) if isinstance(x, int) or str(x).isdigit()]
        anchor_refs = []
        for idx in anchor_ids:
            if 0 <= idx < len(all_boxes):
                try:
                    anchor_refs.append((fitz.Rect(all_boxes[idx]), str(all_types[idx] if idx < len(all_types) else 'field')))
                except Exception:
                    continue

        def _exists_close(rect, kind):
            for idx_existing, existing in enumerate(all_boxes):
                existing_kind = str(all_types[idx_existing] if idx_existing < len(all_types) else '')
                if existing_kind != str(kind):
                    continue
                if self._mobile_repair_overlap_score(rect, existing) >= 0.50:
                    return True
            return False

        for cand_idx in selected_candidate_indices:
            if not (0 <= cand_idx < len(candidates)):
                continue
            cand = candidates[cand_idx]
            rect = cand.get('rect') if isinstance(cand, dict) else None
            kind = str(cand.get('kind', 'field') if isinstance(cand, dict) else 'field')
            if rect is None:
                continue
            try:
                rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
            except Exception:
                continue
            if _exists_close(rect, kind):
                continue
            all_boxes.append(rect)
            all_types.append(kind)
            added_rects.append((fitz.Rect(rect), kind))
            added_count += 1

        def _resolve_live_existing_index(item, min_score=0.45):
            rect = item.get('rect') if isinstance(item, dict) else None
            kind = str(item.get('kind', '') if isinstance(item, dict) else '')
            try:
                target_rect = rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect)
            except Exception:
                target_rect = None
            best_idx = -1
            best_score = 0.0
            if target_rect is not None:
                for idx_existing, existing in enumerate(all_boxes):
                    existing_kind = str(all_types[idx_existing] if idx_existing < len(all_types) else '')
                    if kind and existing_kind != kind:
                        continue
                    score = self._mobile_repair_overlap_score(target_rect, existing)
                    if score > best_score:
                        best_idx = idx_existing
                        best_score = score
                if best_idx >= 0 and best_score >= float(min_score):
                    return int(best_idx)
            try:
                existing_index = int(item.get('existing_index', -1) or -1)
            except Exception:
                existing_index = -1
            if 0 <= existing_index < len(all_boxes):
                return existing_index
            return -1

        remove_existing_indices = []
        for rem_idx in selected_remove_indices:
            if not (0 <= rem_idx < len(removal_candidates)):
                continue
            item = removal_candidates[rem_idx]
            existing_index = _resolve_live_existing_index(item, min_score=0.45)
            if 0 <= existing_index < len(all_boxes):
                remove_existing_indices.append(existing_index)

        for existing_index in sorted(set(remove_existing_indices), reverse=True):
            try:
                all_boxes.pop(existing_index)
                all_types.pop(existing_index)
                removed_ids.append(existing_index)
            except Exception:
                pass

        if not added_rects and not removed_ids:
            self.set_status('Checked repair changes did not produce a local update.', kind='info', hold_seconds=2.4, force=True)
            return

        def _resolve_indices(refs, exclude=None, min_score=0.50):
            resolved = []
            used = set(int(x) for x in (exclude or []) if isinstance(x, int) or str(x).isdigit())
            for rect, kind in refs:
                best_idx = -1
                best_score = 0.0
                for idx_existing, existing in enumerate(all_boxes):
                    if idx_existing in used:
                        continue
                    existing_kind = str(all_types[idx_existing] if idx_existing < len(all_types) else '')
                    if kind and existing_kind != str(kind):
                        continue
                    score = self._mobile_repair_overlap_score(rect, existing)
                    if score > best_score:
                        best_idx = idx_existing
                        best_score = score
                if best_idx >= 0 and best_score >= float(min_score):
                    used.add(best_idx)
                    resolved.append(best_idx)
            return resolved

        resolved_anchor_ids = _resolve_indices(anchor_refs)
        resolved_added_ids = _resolve_indices(added_rects, exclude=resolved_anchor_ids, min_score=0.45)

        engine.all_boxes = all_boxes
        engine.box_types = all_types
        engine.detected_page_idx = int(page_idx)
        engine.detected_settings_signature = engine._settings_signature()
        engine._cache_detection_for_page(page_idx)
        engine.selected_box_ids = sorted(set([idx for idx in resolved_anchor_ids if 0 <= idx < len(all_boxes)] + [idx for idx in resolved_added_ids if 0 <= idx < len(all_boxes)]))
        self._stash_page_selection(page_idx)
        self._repair_anchor_box_ids = list(resolved_anchor_ids)
        self._repair_anchor_page_idx = int(page_idx)
        self._repair_last_proposal = None
        try:
            engine.clear_candidate_cache(page_idx=page_idx)
        except Exception:
            pass
        try:
            remove_rects = []
            remove_types = []
            for rem_idx in selected_remove_indices:
                if 0 <= rem_idx < len(removal_candidates):
                    item = removal_candidates[rem_idx]
                    rect = item.get('rect') if isinstance(item, dict) else None
                    kind = str(item.get('kind', 'field') if isinstance(item, dict) else 'field')
                    if rect is None:
                        continue
                    try:
                        remove_rects.append(rect if isinstance(rect, fitz.Rect) else fitz.Rect(*rect))
                        remove_types.append(kind)
                    except Exception:
                        continue
            engine.save_repair_patch(
                page_idx=page_idx,
                anchor_rects=[r for r, _ in anchor_refs],
                zone_rect=proposal.get('zone_rect'),
                add_rects=[r for r, _ in added_rects],
                add_types=[k for _, k in added_rects],
                remove_rects=remove_rects,
                remove_types=remove_types,
            )
        except Exception:
            pass
        try:
            engine.record_candidate_review_batch(page_idx=page_idx, anchor_ids=anchor_ids, zone_rect=proposal.get('zone_rect'), reviewed_items=reviewed_decisions, source='repair_apply')
        except Exception:
            pass
        if popup is not None:
            try:
                popup.dismiss()
            except Exception:
                pass
        self._render_session_page(page_idx=page_idx, reason=f"Repair applied ({added_count} adds / {len(removed_ids)} removals)")
        self.set_status(f"Repair applied: {added_count} addition{'s' if added_count != 1 else ''}, {len(removed_ids)} removal{'s' if len(removed_ids) != 1 else ''}. Candidate decisions were logged for learning and saved to carry into the next Find.", kind='action', hold_seconds=3.0, force=True)

    def _open_mobile_repair_analysis_popup(self, proposal):
        if not isinstance(proposal, dict):
            self.set_status('No repair proposal available.', kind='info', hold_seconds=2.0, force=True)
            return
        palette = getattr(self, 'ui_palette', {}) or {}
        counts = dict(proposal.get('candidate_type_counts', {}) or {})
        zone = proposal.get('zone_rect')
        zone_text = ''
        if zone is not None:
            try:
                zone_text = f"Repair zone: x {int(zone.x0)}–{int(zone.x1)} • y {int(zone.y0)}–{int(zone.y1)}"
            except Exception:
                zone_text = ''
        outer = BoxLayout(orientation='vertical', spacing=dp(10), padding=[dp(14), dp(14), dp(14), dp(14)])
        self._style_popup_card(outer, palette.get('surface', (0.085, 0.105, 0.145, 1)), radius=dp(22))
        title = Label(text='Repair Analysis', color=palette.get('text', (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(28), halign='left', valign='middle', font_size=dp(18), bold=True)
        title.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(title, min_height=dp(28), extra_pad=dp(4))
        outer.add_widget(title)
        tuning = {}
        try:
            tuning = getattr(self.engine, 'get_candidate_tuning', lambda: dict(CANDIDATE_TUNING_DEFAULTS))()
        except Exception:
            tuning = dict(CANDIDATE_TUNING_DEFAULTS)
        info_lines = [
            f"Example boxes: {', '.join(str(i) for i in proposal.get('anchor_ids', [])[:12])}",
            f"Local existing detections: {int(proposal.get('existing_local_count', 0) or 0)}",
            f"Candidate tuning: min {float((tuning or {}).get('min_candidate_score', 0.34) or 0.34):.2f} • add≥{float((tuning or {}).get('auto_check_add_score', 0.58) or 0.58):.2f} • remove≥{float((tuning or {}).get('auto_check_remove_score', 0.76) or 0.76):.2f}",
            f"Learning priors: {int(proposal.get('learning_prior_pairs', 0) or 0)} pair(s) • same-form events: {int(proposal.get('learning_prior_same_fp_events', 0) or 0)}",
            f"Suggested additions — fields: {int(counts.get('field', 0) or 0)} • checks: {int(counts.get('check', 0) or 0)} • lines: {int(counts.get('line', 0) or 0)}",
        ]
        if int(proposal.get('inferred_structured_count', 0) or 0) > 0:
            info_lines.append(f"Grid-like inferred slots: {int(proposal.get('inferred_structured_count', 0) or 0)}")
        if zone_text:
            info_lines.append(zone_text)
        src_counts = proposal.get('source_counts', {}) or {}
        if src_counts:
            info_lines.append('Sources: ' + ' • '.join(f"{k.replace('_', ' ')} {v}" for k, v in sorted(src_counts.items())))
        for line in info_lines:
            lbl = Label(text=str(line), color=palette.get('text', (0.93, 0.96, 1.0, 1)), size_hint_y=None, halign='left', valign='middle', font_size=dp(11))
            lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(lbl, min_height=dp(18), extra_pad=dp(4))
            outer.add_widget(lbl)
        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False, bar_width=dp(6))
        grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        grid.bind(minimum_height=grid.setter('height'))
        candidates = list(proposal.get('candidates', []) or [])
        if not candidates:
            empty = Label(text='No stronger local repair candidates were found around the selected examples.', color=palette.get('muted', (0.60, 0.68, 0.80, 1)), size_hint_y=None, halign='left', valign='middle', font_size=dp(11))
            empty.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(empty, min_height=dp(22), extra_pad=dp(6))
            grid.add_widget(empty)
        else:
            for idx_item, item in enumerate(candidates[:12], start=1):
                card = BoxLayout(orientation='vertical', spacing=dp(4), size_hint_y=None, padding=[dp(10), dp(8), dp(10), dp(8)])
                card.bind(minimum_height=card.setter('height'))
                self._style_popup_card(card, palette.get('surface_alt', (0.11, 0.135, 0.185, 1)), radius=dp(16))
                rect = item.get('rect')
                rect_text = ''
                if rect is not None:
                    try:
                        rect_text = f"x {int(rect.x0)}–{int(rect.x1)} • y {int(rect.y0)}–{int(rect.y1)}"
                    except Exception:
                        rect_text = ''
                line1 = Label(text=f"{idx_item}. {str(item.get('kind', '')).upper()} • {str(item.get('source', '')).replace('_', ' ')} • score {float(item.get('score', 0.0) or 0.0):.2f}", color=palette.get('text', (0.93, 0.96, 1.0, 1)), size_hint_y=None, halign='left', valign='middle', font_size=dp(11.5), bold=True)
                line1.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(line1, min_height=dp(20), extra_pad=dp(4))
                card.add_widget(line1)
                if rect_text:
                    line2 = Label(text=rect_text, color=palette.get('muted', (0.60, 0.68, 0.80, 1)), size_hint_y=None, halign='left', valign='middle', font_size=dp(10.5))
                    line2.bind(size=self._sync_label_text_size)
                    self._bind_auto_height_label(line2, min_height=dp(18), extra_pad=dp(4))
                    card.add_widget(line2)
                grid.add_widget(card)
        scroll.add_widget(grid)
        outer.add_widget(scroll)
        btn_row = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(44))
        btn_close = self._make_compact_action_button('Close', tone='plain')
        btn_review = self._make_compact_action_button('Review Changes', tone='primary')
        btn_row.add_widget(btn_close)
        btn_row.add_widget(btn_review)
        outer.add_widget(btn_row)
        popup = Popup(title='', separator_height=0, background='', background_color=(0, 0, 0, 0.80), content=outer, size_hint=(0.96 if getattr(self, 'ui_mobile', False) else 0.82, 0.86 if getattr(self, 'ui_mobile', False) else 0.72), auto_dismiss=True)
        self._bind_popup_background_softening(popup)
        def _review_changes(*_):
            self._repair_anchor_box_ids = list(proposal.get('anchor_ids', []) or [])
            self._repair_anchor_page_idx = int(proposal.get('page_idx', 0) or 0)
            self._repair_last_proposal = proposal
            self.set_status(f"Repair analysis ready: {len(candidates)} local candidate{'s' if len(candidates) != 1 else ''} found. Review the checked changes before applying.", kind='action', hold_seconds=2.6, force=True)
            try:
                popup.dismiss()
            except Exception:
                pass
            self._open_mobile_repair_review_popup(proposal)
        btn_close.bind(on_release=lambda *_: popup.dismiss())
        btn_review.bind(on_release=_review_changes)
        popup.open()

    def _run_mobile_repair_analysis(self, anchor_ids, page_idx=0, shell_popup=None):
        proposal = self._build_mobile_repair_proposal(anchor_ids, page_idx=page_idx)
        if proposal is None:
            self.set_status('Repair analysis could not build a local proposal.', kind='warning', hold_seconds=2.4, force=True)
            return
        self._repair_anchor_box_ids = list(proposal.get('anchor_ids', []) or [])
        self._repair_anchor_page_idx = int(proposal.get('page_idx', 0) or 0)
        self._repair_last_proposal = proposal
        if shell_popup is not None:
            try:
                shell_popup.dismiss()
            except Exception:
                pass
        self._open_mobile_repair_analysis_popup(proposal)

    def _open_mobile_repair_shell_popup(self, *_):
        sel_ids = sorted(set(int(x) for x in getattr(self.engine, "selected_box_ids", []) if isinstance(x, int) or str(x).isdigit())) if hasattr(self, "engine") else []
        if not sel_ids:
            self.set_status("Select example boxes first, then tap Repair.", kind="info", hold_seconds=2.2, force=True)
            return

        palette = getattr(self, "ui_palette", {}) or {}
        page_idx = int(self.current_page_idx()) if hasattr(self, 'current_page_idx') else 0
        box_types = list(getattr(self.engine, 'box_types', []) or [])
        counts = {
            'field': sum(1 for t in box_types if str(t) == 'field'),
            'check': sum(1 for t in box_types if str(t) == 'check'),
            'line': sum(1 for t in box_types if str(t) == 'line'),
        }

        outer = BoxLayout(orientation="vertical", spacing=dp(10), padding=[dp(14), dp(14), dp(14), dp(14)])
        self._style_popup_card(outer, palette.get("surface", (0.085, 0.105, 0.145, 1)), radius=dp(22))

        head = BoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None)
        head.bind(minimum_height=head.setter("height"))
        title_lbl = Label(text="Repair Detection", color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(28), halign="left", valign="middle", font_size=dp(18), bold=True)
        title_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(title_lbl, min_height=dp(28), extra_pad=dp(4))
        sub_lbl = Label(text="Optional fallback when the original detection misses fields or includes interference.", color=palette.get("muted", (0.60, 0.68, 0.80, 1)), size_hint_y=None, halign="left", valign="middle", font_size=dp(11))
        sub_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(sub_lbl, min_height=dp(20), extra_pad=dp(6))
        head.add_widget(title_lbl)
        head.add_widget(sub_lbl)
        outer.add_widget(head)

        chip_row = GridLayout(cols=3, spacing=dp(8), size_hint_y=None, height=dp(38))
        chip_row.add_widget(self._make_popup_chip(f"Page {page_idx}", tone="primary"))
        chip_row.add_widget(self._make_popup_chip(f"Examples: {len(sel_ids)}", tone="accent"))
        chip_row.add_widget(self._make_popup_chip("Base detect kept", tone="ghost"))
        outer.add_widget(chip_row)

        body = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
        body.bind(minimum_height=body.setter("height"))
        for text_line in [
            f"Selected example boxes: {', '.join(str(i) for i in sel_ids[:12])}{' …' if len(sel_ids) > 12 else ''}",
            f"Current page detections — fields: {counts['field']} • checks: {counts['check']} • lines: {counts['line']}",
            "This step analyzes only the local area around your current examples. The original detection stays unchanged until a later confirm step.",
        ]:
            lbl = Label(text=text_line, color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, halign="left", valign="middle", font_size=dp(11))
            lbl.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(lbl, min_height=dp(18), extra_pad=dp(4))
            body.add_widget(lbl)
        outer.add_widget(body)

        cached_proposal = None
        try:
            cached_proposal = self.engine.get_cached_mobile_repair_proposal(page_idx, anchor_ids=sel_ids)
        except Exception:
            cached_proposal = None
        if cached_proposal:
            cached_counts = dict(cached_proposal.get('candidate_type_counts', {}) or {})
            cached_lines = [
                f"Cached candidates ready — fields: {int(cached_counts.get('field', 0) or 0)} • checks: {int(cached_counts.get('check', 0) or 0)} • lines: {int(cached_counts.get('line', 0) or 0)}",
                "You can resume the cached repair review for this same example selection, or run a fresh analysis.",
            ]
            for text_line in cached_lines:
                lbl = Label(text=text_line, color=palette.get("muted", (0.60, 0.68, 0.80, 1)), size_hint_y=None, halign="left", valign="middle", font_size=dp(10.6))
                lbl.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(lbl, min_height=dp(18), extra_pad=dp(4))
                body.add_widget(lbl)

        btn_row = GridLayout(cols=4 if cached_proposal else 3, spacing=dp(8), size_hint_y=None, height=dp(44))
        btn_details = self._make_compact_action_button("Link Details", tone="ghost")
        btn_arm = self._make_compact_action_button("Analyze", tone="primary")
        btn_close = self._make_compact_action_button("Close", tone="plain")
        if cached_proposal:
            btn_resume = self._make_compact_action_button("Resume", tone="accent")
            btn_row.add_widget(btn_resume)
        btn_row.add_widget(btn_details)
        btn_row.add_widget(btn_arm)
        btn_row.add_widget(btn_close)
        outer.add_widget(btn_row)

        popup = Popup(title="", separator_height=0, background="", background_color=(0, 0, 0, 0.80), content=outer, size_hint=(0.96 if getattr(self, 'ui_mobile', False) else 0.78, 0.64 if getattr(self, 'ui_mobile', False) else 0.52), auto_dismiss=True)
        self._bind_popup_background_softening(popup)
        def _resume_cached(*_):
            if not cached_proposal:
                return
            self._repair_anchor_box_ids = list(cached_proposal.get('anchor_ids', []) or [])
            self._repair_anchor_page_idx = int(cached_proposal.get('page_idx', page_idx) or page_idx)
            self._repair_last_proposal = cached_proposal
            try:
                popup.dismiss()
            except Exception:
                pass
            self._open_mobile_repair_review_popup(cached_proposal)
        btn_details.bind(on_release=lambda *_: (popup.dismiss(), self.open_mapping_popup_for_selection(primary_box_id=sel_ids[0])))
        if cached_proposal:
            btn_resume.bind(on_release=_resume_cached)
        btn_arm.bind(on_release=lambda *_: self._run_mobile_repair_analysis(sel_ids, page_idx=page_idx, shell_popup=popup))
        btn_close.bind(on_release=lambda *_: popup.dismiss())
        popup.open()

    
    def _toggle_mobile_inspect_panel(self, *_):
        sel_ids = sorted(set(int(x) for x in getattr(self.engine, "selected_box_ids", []) if isinstance(x, int) or str(x).isdigit())) if hasattr(self, "engine") else []
        if not sel_ids:
            self.set_status("Select example boxes first, then tap Repair.", kind="info", hold_seconds=2.2, force=True)
            self._set_mobile_quick_rail_visible(False)
            return
        self._set_mobile_quick_rail_visible(False)
        self._open_mobile_repair_shell_popup()
        self._refresh_mobile_hud_restore_button()

    def _refresh_mobile_action_visibility(self):
        if not getattr(self, "ui_mobile", False):
            return
        selected = bool(getattr(self.engine, "selected_box_ids", []))
        hud_btn = getattr(self, "btn_show_hud", None)
        if hud_btn is not None:
            hud_btn.opacity = 0
            hud_btn.disabled = True
            try:
                hud_btn.size = (0, 0)
                hud_btn.pos = (-10000, -10000)
            except Exception:
                pass
        self._refresh_mobile_select_mode_ui()
        rail = getattr(self, "mobile_quick_rail", None)
        if rail is not None:
            rail_open = bool(getattr(self, "_mobile_quick_rail_open", False))
            rail.opacity = 1 if rail_open else 0
            rail.disabled = not rail_open
            try:
                rail.size_hint = (None, None)
                rail.size = (dp(84), dp(134)) if rail_open else (0, 0)
                if not rail_open:
                    rail.pos = (-10000, -10000)
            except Exception:
                pass
        for attr in (
            "btn_mobile_rail_fit_page",
            "btn_mobile_rail_fit_width",
            "btn_mobile_rail_zoom_reset",
        ):
            btn = getattr(self, attr, None)
            if btn is not None:
                btn.opacity = 1 if bool(getattr(self, "_mobile_quick_rail_open", False)) else 0

    def _set_mobile_sidebar_state(self, open_state, *_):
        if platform == "android":
            self._mobile_sidebar_open = False
            return
        self._mobile_sidebar_open = bool(open_state)
        if self._mobile_sidebar_open and bool(getattr(self, "mobile_tools_tray_open", False)):
            self._set_mobile_tools_tray_visible(False)
        if self._mobile_sidebar_open and bool(getattr(self, "_mobile_quick_rail_open", False)):
            self._set_mobile_quick_rail_visible(False)
        sidebar = getattr(self, "mobile_sidebar", None)
        scrim = getattr(self, "mobile_sidebar_scrim", None)
        if sidebar is not None:
            try:
                sidebar_x = 0 if self._mobile_sidebar_open else -float(sidebar.width) - dp(8)
                sidebar.x = sidebar_x
            except Exception:
                pass
        if scrim is not None:
            try:
                scrim.opacity = 1 if self._mobile_sidebar_open else 0
                scrim.disabled = not self._mobile_sidebar_open
                scrim.size_hint = (1, 1) if self._mobile_sidebar_open else (None, None)
                if self._mobile_sidebar_open:
                    scrim.size = getattr(self, 'mobile_workspace', scrim).size if getattr(self, 'mobile_workspace', None) is not None else Window.size
                    scrim.pos = (0, 0)
                else:
                    scrim.size = (0, 0)
                    scrim.pos = (-10000, -10000)
            except Exception:
                pass
        for attr in ("btn_sidebar_toggle", "btn_mobile_sidebar"):
            btn = getattr(self, attr, None)
            if btn is not None:
                try:
                    btn.text = "Close" if (self._mobile_sidebar_open and attr == "btn_sidebar_toggle") else ("Menu" if attr == "btn_mobile_sidebar" else ("Menu" if not self._mobile_sidebar_open else "Close"))
                except Exception:
                    pass



    def _refresh_android_selection_inspector(self):
        ids = sorted(set(int(x) for x in getattr(self.engine, "selected_box_ids", []) if isinstance(x, int) or str(x).isdigit()))
        lbl = getattr(self, "android_inspector_selected_lbl", None)
        if lbl is not None:
            lbl.text = ", ".join(str(i) for i in ids) if ids else "None"
        lbl = getattr(self, "android_inspector_count_lbl", None)
        if lbl is not None:
            lbl.text = str(len(ids))

        box_text = "No selection"
        mapping_text = "Choose a box in the preview to inspect it."
        resolved_text = "Resolved preview value: —"
        if ids:
            first = ids[0]
            try:
                box_type = self.engine.box_types[first] if first < len(self.engine.box_types) else "field"
            except Exception:
                box_type = "field"
            rect = self.engine.all_boxes[first] if first < len(self.engine.all_boxes) else None
            rect_text = ""
            if rect is not None:
                try:
                    rect_text = f" | ({int(rect.x0)}, {int(rect.y0)}) → ({int(rect.x1)}, {int(rect.y1)})"
                except Exception:
                    rect_text = ""
            box_text = f"Box {first} • {box_type}{rect_text}"
            try:
                mapping = self.engine.describe_box_mapping(first, self.current_page_idx())
            except Exception:
                mapping = ""
            mapping_text = mapping if mapping and mapping != "EMPTY" else "Unmapped"
            try:
                mapping_payload = self._find_mapping_for_box(first, self.current_page_idx()) or {}
            except Exception:
                mapping_payload = {}
            try:
                resolved_text = "Resolved preview value: " + self._get_selected_column_preview(
                    mapping_payload.get("column", self.column_spinner.text if hasattr(self, "column_spinner") else ""),
                    trigger=mapping_payload.get("trigger", self.trigger_input.text if hasattr(self, "trigger_input") else ""),
                    is_grid=bool(mapping_payload.get("g", getattr(self.grid_flag_chk, "active", False) if hasattr(self, "grid_flag_chk") else False)),
                    grid_n=int(mapping_payload.get("n", self.grid_n_input.text if hasattr(self, "grid_n_input") else "1") or 1),
                )
            except Exception:
                resolved_text = "Resolved preview value: —"

        for attr, value in [
            ("android_inspector_box_lbl", box_text),
            ("android_inspector_mapping_lbl", mapping_text),
            ("android_inspector_resolution_lbl", resolved_text),
        ]:
            lbl = getattr(self, attr, None)
            if lbl is not None:
                lbl.text = value

    def _refresh_android_capture_review_panel(self):
        summary_lbl = getattr(self, "android_capture_summary_lbl", None)
        cand_lbl = getattr(self, "android_capture_candidate_lbl", None)
        hint_lbl = getattr(self, "android_capture_action_hint_lbl", None)
        analysis = dict(getattr(self, "_capture_active_template", {}) or {})
        candidates = list(getattr(self, "_capture_shape_candidates", []) or [])
        idx = int(getattr(self, "_capture_shape_candidate_index", 0) or 0)
        if cand_lbl is not None:
            cand_lbl.text = (f"Candidate: {idx + 1} / {len(candidates)}" if candidates else "Candidate: ROI logic")
        if not analysis:
            if summary_lbl is not None:
                summary_lbl.text = "No captured area yet."
            if hint_lbl is not None:
                hint_lbl.text = "Review stays here on Android. Popup is optional."
            return
        try:
            rect = analysis.get("source_rect")
            rr = rect if isinstance(rect, fitz.Rect) else fitz.Rect(rect)
            rect_txt = f"ROI: ({int(rr.x0)}, {int(rr.y0)}) → ({int(rr.x1)}, {int(rr.y1)})"
        except Exception:
            rect_txt = "ROI: —"
        template_kind = str(analysis.get("template_kind", "field") or "field")
        split_count = int(analysis.get("split_count", 0) or 0)
        resolved_types = [str(v) for v in (analysis.get("resolved_types", []) or [])]
        kinds_txt = ", ".join(resolved_types[:4]) if resolved_types else "field"
        if summary_lbl is not None:
            summary_lbl.text = f"{rect_txt}\nType: {template_kind} • Targets: {split_count}\nResolved: {kinds_txt}"
        if hint_lbl is not None:
            hint_lbl.text = "Use panel actions after capture. Popup is optional." if bool(getattr(self, "_capture_mode_active", False)) else "Template ready. Use panel actions or open popup."

    def _refresh_android_capture_stage_ui(self):
        lbl = getattr(self, "android_capture_stage_lbl", None)
        if lbl is None:
            return
        if bool(getattr(self, "_capture_mode_active", False)):
            point_count = len(list(getattr(self, "_capture_points", []) or []))
            if point_count <= 0:
                lbl.text = "Stage: Capture active • tap top-left"
            elif point_count == 1:
                lbl.text = "Stage: Capture active • tap bottom-right"
            elif bool(getattr(self, "_capture_shape_candidates", []) or []):
                lbl.text = f"Stage: {len(list(getattr(self, '_capture_shape_candidates', []) or []))} candidate(s) ready"
            else:
                lbl.text = "Stage: Capture active"
        else:
            if getattr(self, "_capture_active_template", None):
                lbl.text = "Stage: Template ready"
            else:
                lbl.text = "Stage: Idle"

    def _toggle_mobile_sidebar(self, *_):
        self._set_mobile_sidebar_state(not getattr(self, "_mobile_sidebar_open", False))

    # ------------------------------------------------------------------
    # Focus Mode — hide control panels so the preview fills the screen
    # ------------------------------------------------------------------

    def _toggle_focus_mode(self, *_):
        self._set_focus_mode(not bool(getattr(self, "_focus_mode_active", False)))

    def _set_focus_mode(self, active, *_):
        """Enter or exit Focus Mode.

        Mobile (Android): collapses android_top_modes_card and
        android_panel_card by animating their height and opacity to 0.

        Mobile (emulation): hides the bottom dock and sidebar overlay.

        Desktop: collapses left_wrap and right_wrap so the preview
        center column expands to fill the full window width.
        """
        self._focus_mode_active = bool(active)
        focus_btn = getattr(self, "btn_focus_mode", None)

        # ---- Desktop path ----------------------------------------
        if not getattr(self, "_is_mobile_ui", getattr(self, "ui_mobile", False)):
            left_wrap = getattr(self, "_desktop_left_wrap", None)
            right_wrap = getattr(self, "_desktop_right_wrap", None)
            center_wrap = getattr(self, "_desktop_center_wrap", None)
            if self._focus_mode_active:
                # Collapse left and right panels
                if left_wrap is not None:
                    anim = Animation(opacity=0, width=0, duration=0.18, t="out_quad")
                    anim.start(left_wrap)
                    try:
                        left_wrap.size_hint_x = None
                    except Exception:
                        pass
                if right_wrap is not None:
                    anim = Animation(opacity=0, width=0, duration=0.18, t="out_quad")
                    anim.start(right_wrap)
                    try:
                        right_wrap.size_hint_x = None
                    except Exception:
                        pass
                if center_wrap is not None:
                    try:
                        center_wrap.size_hint_x = 1
                    except Exception:
                        pass
                if focus_btn is not None:
                    try:
                        focus_btn.text = "Panels"
                    except Exception:
                        pass
            else:
                # Restore panels
                if left_wrap is not None:
                    try:
                        left_wrap.size_hint_x = None
                        left_wrap.width = 0
                        left_wrap.opacity = 0
                    except Exception:
                        pass
                    anim = Animation(opacity=1, width=dp(260), duration=0.18, t="out_quad")
                    anim.start(left_wrap)
                if right_wrap is not None:
                    try:
                        right_wrap.size_hint_x = None
                        right_wrap.width = 0
                        right_wrap.opacity = 0
                    except Exception:
                        pass
                    anim = Animation(opacity=1, width=dp(220), duration=0.18, t="out_quad")
                    anim.start(right_wrap)
                if center_wrap is not None:
                    try:
                        center_wrap.size_hint_x = 0.57
                    except Exception:
                        pass
                if focus_btn is not None:
                    try:
                        focus_btn.text = "Focus"
                    except Exception:
                        pass
            return

        # ---- Android path ----------------------------------------
        if platform == "android":
            top_card = getattr(self, "android_top_modes_card", None)
            panel_card = getattr(self, "android_panel_card", None)
            if self._focus_mode_active:
                for card in (top_card, panel_card):
                    if card is not None:
                        try:
                            anim = Animation(opacity=0, height=0, duration=0.20, t="out_quad")
                            anim.start(card)
                            card.disabled = True
                        except Exception:
                            pass
                if focus_btn is not None:
                    try:
                        focus_btn.text = "Panels"
                        focus_btn.opacity = 0.9
                    except Exception:
                        pass
            else:
                if top_card is not None:
                    try:
                        anim = Animation(opacity=1, height=dp(98), duration=0.20, t="out_quad")
                        anim.start(top_card)
                        top_card.disabled = False
                    except Exception:
                        pass
                if panel_card is not None:
                    try:
                        anim = Animation(opacity=1, height=dp(210), duration=0.20, t="out_quad")
                        anim.start(panel_card)
                        panel_card.disabled = False
                    except Exception:
                        pass
                if focus_btn is not None:
                    try:
                        focus_btn.text = "Focus"
                        focus_btn.opacity = 0.72
                    except Exception:
                        pass
            return

        # ---- Non-Android mobile (emulation) path -----------------
        if self._focus_mode_active:
            # Close sidebar and bottom dock to maximise preview area
            self._set_mobile_sidebar_state(False)
            self._set_mobile_tools_tray_visible(False)
            dock = getattr(self, "mobile_bottom_dock", None)
            if dock is not None:
                try:
                    anim = Animation(opacity=0, duration=0.20, t="out_quad")
                    anim.bind(on_complete=lambda *_: setattr(dock, "disabled", True))
                    anim.start(dock)
                except Exception:
                    pass
            if focus_btn is not None:
                try:
                    focus_btn.text = "Panels"
                    focus_btn.opacity = 0.9
                except Exception:
                    pass
        else:
            dock = getattr(self, "mobile_bottom_dock", None)
            if dock is not None:
                try:
                    dock.disabled = False
                    anim = Animation(opacity=1, duration=0.20, t="out_quad")
                    anim.start(dock)
                except Exception:
                    pass
            if focus_btn is not None:
                try:
                    focus_btn.text = "Focus"
                    focus_btn.opacity = 0.72
                except Exception:
                    pass

    def _update_capture_button_state(self):
        btn = getattr(self, 'btn_mobile_capture', None)
        if btn is None:
            self._refresh_android_capture_stage_ui()
            return
        try:
            btn.text = 'Cancel Cap' if bool(getattr(self, '_capture_mode_active', False)) else 'Capture'
        except Exception:
            pass
        self._refresh_android_capture_stage_ui()

    def _clear_manual_capture_preview(self, preserve_mode=False):
        self._capture_points = []
        self._capture_preview_rect = None
        self._capture_split_preview = []
        self._capture_outline_preview = []
        self._capture_shape_candidates = []
        self._capture_shape_candidate_index = 0
        self._capture_active_template = None
        if not preserve_mode:
            self._capture_mode_active = False
            self._capture_page_idx = -1
        preview = getattr(self, 'preview', None)
        if preview is not None:
            try:
                preview.capture_mode_active = bool(self._capture_mode_active)
            except Exception:
                pass
            if hasattr(preview, 'clear_capture_guides'):
                preview.clear_capture_guides()
        self._update_capture_button_state()

    def _refresh_manual_capture_overlay(self):
        preview = getattr(self, 'preview', None)
        if preview is None or not hasattr(preview, 'set_capture_guides'):
            return
        try:
            preview.capture_mode_active = bool(getattr(self, '_capture_mode_active', False))
        except Exception:
            pass
        capture_rect = None
        if isinstance(getattr(self, '_capture_preview_rect', None), fitz.Rect):
            rect = self._capture_preview_rect
            capture_rect = {'pdf_x': float(rect.x0), 'pdf_y': float(rect.y0), 'pdf_w': float(rect.width), 'pdf_h': float(rect.height)}
        guide_rects = []
        for rect in (getattr(self, '_capture_split_preview', []) or []):
            try:
                rect = fitz.Rect(rect)
            except Exception:
                continue
            guide_rects.append({'pdf_x': float(rect.x0), 'pdf_y': float(rect.y0), 'pdf_w': float(rect.width), 'pdf_h': float(rect.height)})
        preview.set_capture_guides(capture_rect=capture_rect, guide_rects=guide_rects, guide_paths=list(getattr(self, '_capture_outline_preview', []) or []), guide_points=list(getattr(self, '_capture_points', []) or []))

    def _toggle_manual_area_capture_mode(self, *_):
        if bool(getattr(self, '_capture_mode_active', False)):
            self.set_status('Capture cancelled.')
            self._clear_manual_capture_preview(preserve_mode=False)
            return
        page_idx = self.current_page_idx()
        self._prepare_page_context(page_idx, clear_selection_if_missing=False)
        has_live = bool(getattr(self.engine, 'all_boxes', [])) and int(getattr(self.engine, 'detected_page_idx', -1) or -1) == int(page_idx)
        if (not has_live) and (not bool(self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()))):
            self.set_status('Run Detect first, then use Capture Missed Area.')
            return
        self._capture_mode_active = True
        self._capture_page_idx = int(page_idx)
        self._capture_points = []
        self._capture_preview_rect = None
        self._capture_split_preview = []
        self._capture_outline_preview = []
        self._capture_shape_candidates = []
        self._capture_shape_candidate_index = 0
        self._capture_active_template = None
        self._refresh_manual_capture_overlay()
        self._update_capture_button_state()
        self.set_status(f'Capture Missed Area active on page {page_idx}. Tap the top-left corner.')

    def _build_manual_area_rect_from_points(self, p1, p2):
        try:
            x0 = min(float(p1[0]), float(p2[0]))
            y0 = min(float(p1[1]), float(p2[1]))
            x1 = max(float(p1[0]), float(p2[0]))
            y1 = max(float(p1[1]), float(p2[1]))
        except Exception:
            return None
        tex = getattr(getattr(self, 'preview', None), 'texture', None)
        if tex is not None:
            x0 = max(0.0, min(x0, float(tex.width)))
            y0 = max(0.0, min(y0, float(tex.height)))
            x1 = max(0.0, min(x1, float(tex.width)))
            y1 = max(0.0, min(y1, float(tex.height)))
        if (x1 - x0) < 8 or (y1 - y0) < 8:
            return None
        zoom = float(getattr(self, '_last_preview_render_zoom', PREVIEW_SCALE) or PREVIEW_SCALE)
        return fitz.Rect(x0 / zoom, y0 / zoom, x1 / zoom, y1 / zoom)

    def on_preview_point_tap(self, img_pt, hit=None):
        if not bool(getattr(self, '_capture_mode_active', False)):
            return False
        if img_pt is None:
            return True
        if len(getattr(self, '_capture_points', []) or []) >= 2 and bool(getattr(self, '_capture_shape_candidates', []) or []):
            try:
                if self._pick_capture_shape_candidate_at_point(img_pt):
                    try:
                        self.set_status('Trace selected. Review actions in the Capture panel.')
                    except Exception:
                        pass
                    if platform == "android":
                        try:
                            self._show_mobile_section("Capture")
                        except Exception:
                            pass
                        self._refresh_android_capture_review_panel()
                    else:
                        Clock.schedule_once(lambda dt: self._open_manual_area_capture_confirm_popup(), 0)
                    return True
            except Exception:
                pass
            try:
                preview_rect = getattr(self, '_capture_preview_rect', None)
                if preview_rect is not None:
                    rr = preview_rect if isinstance(preview_rect, fitz.Rect) else fitz.Rect(preview_rect)
                    px = float(img_pt[0]); py = float(img_pt[1])
                    if rr.x0 <= px <= rr.x1 and rr.y0 <= py <= rr.y1:
                        try:
                            self.set_status('Using current trace selection. Review actions in the Capture panel.')
                        except Exception:
                            pass
                        if platform == "android":
                            try:
                                self._show_mobile_section("Capture")
                            except Exception:
                                pass
                            self._refresh_android_capture_review_panel()
                        else:
                            Clock.schedule_once(lambda dt: self._open_manual_area_capture_confirm_popup(), 0)
                        return True
            except Exception:
                pass
        page_idx = self.current_page_idx()
        if int(getattr(self, '_capture_page_idx', page_idx)) != int(page_idx):
            self._clear_manual_capture_preview(preserve_mode=False)
            return True
        try:
            point = (float(img_pt[0]), float(img_pt[1]))
        except Exception:
            return True
        self._capture_points.append(point)
        if len(self._capture_points) == 1:
            self._refresh_manual_capture_overlay()
            self._refresh_android_capture_stage_ui()
            self._refresh_android_capture_review_panel()
            self.set_status('Top-left captured. Tap the bottom-right corner.')
            return True
        p1 = self._capture_points[0]
        p2 = point
        rect = self._build_manual_area_rect_from_points(p1, p2)
        if rect is None:
            self.set_status('Captured area is too small. Tap the top-left corner again.')
            self._capture_points = []
            self._capture_preview_rect = None
            self._capture_split_preview = []
            self._refresh_manual_capture_overlay()
            return True
        self._capture_points = [p1, p2]
        self._capture_preview_rect = rect
        analysis = None
        shape_candidates = []
        try:
            shape_candidates = list(self.engine.extract_shape_candidates_from_roi(page_idx, rect, max_candidates=8) or [])
        except Exception:
            shape_candidates = []
        self._capture_shape_candidates = list(shape_candidates or [])
        self._capture_shape_candidate_index = 0
        if shape_candidates:
            try:
                analysis = self.engine.build_shape_template_analysis(page_idx, rect, shape_candidates[0])
            except Exception:
                analysis = None
        if analysis is None:
            try:
                analysis = self.engine.analyze_manual_area_template(page_idx, rect)
            except Exception:
                analysis = None
        if analysis is not None:
            self._capture_split_preview = [fitz.Rect(r) for r in (analysis.get('resolved_rects', []) or [])]
            self._capture_active_template = analysis
            shape_template = dict(analysis.get('shape_template', {}) or {})
            if shape_template:
                self._capture_outline_preview = [self.engine._shape_template_preview_path(shape_template, preview_zoom=getattr(self, '_last_preview_render_zoom', PREVIEW_SCALE))]
            else:
                self._capture_outline_preview = []
        else:
            self._capture_split_preview = []
            self._capture_outline_preview = []
            self._capture_active_template = None
        self._refresh_manual_capture_overlay()
        if shape_candidates:
            self.set_status('Captured area ready. Tap an orange outline to choose it, or tap inside the ROI to continue with the current trace.')
        else:
            self.set_status('Captured area ready. Review the template actions.')
            Clock.schedule_once(lambda dt: self._open_manual_area_capture_confirm_popup(), 0)
        return True

    def _capture_shape_candidate_hit_index(self, img_pt):
        candidates = list(getattr(self, '_capture_shape_candidates', []) or [])
        if not candidates or img_pt is None:
            return -1
        try:
            px = float(img_pt[0]); py = float(img_pt[1])
        except Exception:
            return -1
        best_idx = -1
        best_dist = 1e18
        for idx, cand in enumerate(candidates):
            pts = list((cand or {}).get('pdf_points', []) or [])
            if pts:
                poly = []
                for pt in pts:
                    if not isinstance(pt, (list, tuple)) or len(pt) < 2:
                        continue
                    try:
                        poly.append([float(pt[0]) * float(getattr(self, '_last_preview_render_zoom', PREVIEW_SCALE) or PREVIEW_SCALE), float(pt[1]) * float(getattr(self, '_last_preview_render_zoom', PREVIEW_SCALE) or PREVIEW_SCALE)])
                    except Exception:
                        continue
                if len(poly) >= 3:
                    try:
                        arr = np.array(poly, dtype=np.float32)
                        inside = cv2.pointPolygonTest(arr, (px, py), False)
                        if inside >= 0:
                            return int(idx)
                        dist = abs(float(cv2.pointPolygonTest(arr, (px, py), True)))
                    except Exception:
                        dist = 1e18
                else:
                    dist = 1e18
            else:
                try:
                    rr = fitz.Rect((cand or {}).get('bbox_rect'))
                    z = float(getattr(self, '_last_preview_render_zoom', PREVIEW_SCALE) or PREVIEW_SCALE)
                    x0 = rr.x0 * z; y0 = rr.y0 * z; x1 = rr.x1 * z; y1 = rr.y1 * z
                    if x0 <= px <= x1 and y0 <= py <= y1:
                        return int(idx)
                    cx = (x0 + x1) * 0.5; cy = (y0 + y1) * 0.5
                    dist = math.hypot(px - cx, py - cy)
                except Exception:
                    dist = 1e18
            if dist < best_dist:
                best_dist = dist
                best_idx = int(idx)
        if best_dist <= float(dp(28)):
            return best_idx
        return -1

    def _pick_capture_shape_candidate_at_point(self, img_pt):
        hit_idx = self._capture_shape_candidate_hit_index(img_pt)
        if hit_idx < 0:
            return False
        self._capture_shape_candidate_index = int(hit_idx)
        candidates = list(getattr(self, '_capture_shape_candidates', []) or [])
        rect = getattr(self, '_capture_preview_rect', None)
        page_idx = int(getattr(self, '_capture_page_idx', self.current_page_idx()) or self.current_page_idx())
        if hit_idx >= len(candidates):
            return False
        analysis = self.engine.build_shape_template_analysis(page_idx, rect, dict(candidates[hit_idx] or {})) or {}
        self._capture_active_template = dict(analysis or {})
        self._capture_split_preview = [fitz.Rect(r) for r in (analysis.get('resolved_rects', []) or [])]
        shape_template = dict(analysis.get('shape_template', {}) or {})
        self._capture_outline_preview = [self.engine._shape_template_preview_path(shape_template, preview_zoom=getattr(self, '_last_preview_render_zoom', PREVIEW_SCALE))] if shape_template else []
        self._refresh_manual_capture_overlay()
        cb = getattr(self, '_capture_popup_summary_updater', None)
        if callable(cb):
            try:
                cb()
            except Exception:
                pass
        try:
            self.set_status(f'Trace {hit_idx + 1} selected by tap. Now tap Use Area or Find Similar.')
        except Exception:
            pass
        return True

    def _select_capture_shape_candidate(self, offset=0):
        candidates = list(getattr(self, '_capture_shape_candidates', []) or [])
        if not candidates:
            self._refresh_android_capture_review_panel()
            return dict(getattr(self, '_capture_active_template', {}) or {})
        idx = int(getattr(self, '_capture_shape_candidate_index', 0) or 0)
        idx = (idx + int(offset or 0)) % max(len(candidates), 1)
        self._capture_shape_candidate_index = idx
        rect = getattr(self, '_capture_preview_rect', None)
        page_idx = int(getattr(self, '_capture_page_idx', self.current_page_idx()) or self.current_page_idx())
        candidate = dict(candidates[idx] or {})
        analysis = self.engine.build_shape_template_analysis(page_idx, rect, candidate) or {}
        self._capture_active_template = dict(analysis or {})
        self._capture_split_preview = [fitz.Rect(r) for r in (analysis.get('resolved_rects', []) or [])]
        shape_template = dict(analysis.get('shape_template', {}) or {})
        self._capture_outline_preview = [self.engine._shape_template_preview_path(shape_template, preview_zoom=getattr(self, '_last_preview_render_zoom', PREVIEW_SCALE))] if shape_template else []
        self._refresh_manual_capture_overlay()
        self._refresh_android_capture_review_panel()
        return analysis

    def _open_manual_area_capture_confirm_popup(self):
        analysis = dict(getattr(self, '_capture_active_template', {}) or {})
        rect = analysis.get('source_rect') if analysis else getattr(self, '_capture_preview_rect', None)
        if rect is None:
            return
        page_idx = int(analysis.get('page', self.current_page_idx()) or self.current_page_idx())
        popup_box = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(12))
        self._style_popup_card(popup_box, color=(self.ui_palette or {}).get('surface_alt', (0.11, 0.135, 0.185, 1)), radius=dp(18))
        title_lbl = Label(text='Captured Shape Template', color=(self.ui_palette or {}).get('text', (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(26), halign='left', valign='middle', font_size=dp(18), bold=True)
        title_lbl.bind(size=self._sync_label_text_size)
        popup_box.add_widget(title_lbl)
        helper_lbl = Label(text='Step 1: choose a contour candidate if needed. Step 2: Use Area to add it, Delete In Area to remove overlapping detections and links, or Find Similar to search from it. The current candidate is already selected.', color=(self.ui_palette or {}).get('muted', (0.70, 0.78, 0.90, 1)), size_hint_y=None, height=dp(44), halign='left', valign='top', font_size=dp(11))
        helper_lbl.bind(size=self._sync_label_text_size)
        popup_box.add_widget(helper_lbl)
        summary = Label(text='', color=(self.ui_palette or {}).get('muted', (0.70, 0.78, 0.90, 1)), size_hint_y=None, height=dp(96), halign='left', valign='top', font_size=dp(12))
        summary.bind(size=self._sync_label_text_size)
        popup_box.add_widget(summary)
        candidate_row = GridLayout(cols=4, spacing=dp(8), size_hint_y=None, height=dp(40))
        btn_prev = self._make_compact_action_button('Prev', tone='plain')
        btn_next = self._make_compact_action_button('Next', tone='plain')
        btn_pick = self._make_compact_action_button('Select This Trace', tone='accent')
        btn_fallback = self._make_compact_action_button('Use ROI Logic', tone='ghost')
        for btn in (btn_prev, btn_next, btn_pick, btn_fallback):
            candidate_row.add_widget(btn)
        popup_box.add_widget(candidate_row)
        actions = GridLayout(cols=3, spacing=dp(8), size_hint_y=None, height=dp(132))
        btn_use = self._make_compact_action_button('Use Area', tone='accent')
        btn_delete = self._make_compact_action_button('Delete In Area', tone='danger')
        btn_similar = self._make_compact_action_button('Find Similar (Page)', tone='plain')
        btn_similar_form = self._make_compact_action_button('Find Similar (Form)', tone='plain')
        btn_retake = self._make_compact_action_button('Retake', tone='plain')
        btn_cancel = self._make_compact_action_button('Cancel', tone='ghost')
        actions.add_widget(btn_use); actions.add_widget(btn_delete); actions.add_widget(btn_similar); actions.add_widget(btn_similar_form); actions.add_widget(btn_retake); actions.add_widget(btn_cancel)
        popup_box.add_widget(actions)
        popup = Popup(title='', separator_height=0, background='', background_color=(0,0,0,0.82), content=popup_box, size_hint=(0.92 if self.ui_mobile else 0.60, 0.52 if self.ui_mobile else 0.40), auto_dismiss=False)
        state = {'mode': 'shape' if bool(getattr(self, '_capture_shape_candidates', []) or []) else 'roi'}

        def _update_summary(*_):
            current = dict(getattr(self, '_capture_active_template', {}) or {})
            shape_template = dict(current.get('shape_template', {}) or {})
            split_count = int(current.get('split_count', 1) or 1)
            template_kind = str(current.get('template_kind', 'field') or 'field')
            overlap = float(current.get('source_text_overlap', 0.0) or 0.0)
            candidates = list(getattr(self, '_capture_shape_candidates', []) or [])
            idx = int(getattr(self, '_capture_shape_candidate_index', 0) or 0)
            if shape_template:
                summary.text = (
                    f'Page {page_idx} • shape template\n'
                    f'Area: {rect.width:.1f} × {rect.height:.1f} pdf units\n'
                    f'Candidate {idx + 1} of {max(len(candidates), 1)} selected • score {float(shape_template.get("score", 0.0) or 0.0):.2f}\n'
                    f'Vertices: {int(shape_template.get("vertices", 0) or 0)} • aspect {float(shape_template.get("bbox_aspect", 0.0) or 0.0):.2f} • next: Use Area or Find Similar'
                )
            else:
                summary.text = (
                    f'Page {page_idx} • {template_kind}\n'
                    f'Area: {rect.width:.1f} × {rect.height:.1f} pdf units\n'
                    f'Smart split detected: {split_count} target{"s" if split_count != 1 else ""}\n'
                    f'Text overlap ignored: {overlap:.2f}'
                )


        def _pick_shape(*_):
            state['mode'] = 'shape'
            self._select_capture_shape_candidate(0)
            _update_summary()
            try:
                self.set_status('Trace selected. Step 2: choose Use Area or Find Similar.')
            except Exception:
                pass

        def _fallback_roi(*_):
            state['mode'] = 'roi'
            try:
                analysis_roi = self.engine.analyze_manual_area_template(page_idx, rect) or {}
            except Exception:
                analysis_roi = {}
            self._capture_active_template = dict(analysis_roi or {})
            self._capture_split_preview = [fitz.Rect(r) for r in (analysis_roi.get('resolved_rects', []) or [])]
            self._capture_outline_preview = []
            self._refresh_manual_capture_overlay()
            _update_summary()
            try:
                self.set_status('ROI logic selected. Step 2: choose Use Area or Find Similar.')
            except Exception:
                pass

        btn_prev.bind(on_release=lambda *_: (self._select_capture_shape_candidate(-1), _update_summary()))
        btn_next.bind(on_release=lambda *_: (self._select_capture_shape_candidate(1), _update_summary()))
        btn_pick.bind(on_release=_pick_shape)
        btn_fallback.bind(on_release=_fallback_roi)

        def _apply(*_):
            self._capture_popup_summary_updater = None
            popup.dismiss()
            self._apply_confirmed_manual_area_capture()
        def _retake(*_):
            self._capture_popup_summary_updater = None
            popup.dismiss()
            self._capture_points = []
            self._capture_preview_rect = None
            self._capture_split_preview = []
            self._capture_outline_preview = []
            self._capture_shape_candidates = []
            self._capture_shape_candidate_index = 0
            self._capture_active_template = None
            self._refresh_manual_capture_overlay()
            self.set_status('Retake capture: tap the top-left corner again.')
        def _cancel(*_):
            self._capture_popup_summary_updater = None
            popup.dismiss()
            self._clear_manual_capture_preview(preserve_mode=False)
            self.set_status('Capture cancelled.')
        def _delete(*_):
            self._capture_popup_summary_updater = None
            popup.dismiss()
            self._freeze_active_manual_capture_template()
            self._delete_detected_items_for_active_capture(persist_as_repair_patch=True, include_mappings=True)
        btn_use.bind(on_release=_apply)
        btn_delete.bind(on_release=_delete)
        btn_similar.bind(on_release=lambda *_: (setattr(self, '_capture_popup_summary_updater', None), popup.dismiss(), self._freeze_active_manual_capture_template(), self._find_similar_for_active_capture_template()))
        btn_similar_form.bind(on_release=lambda *_: (setattr(self, '_capture_popup_summary_updater', None), popup.dismiss(), self._freeze_active_manual_capture_template(), self._find_similar_in_form_for_active_capture_template()))
        btn_retake.bind(on_release=_retake)
        btn_cancel.bind(on_release=_cancel)
        self._capture_popup_summary_updater = _update_summary
        _update_summary()
        try:
            candidate_count = len(list(getattr(self, '_capture_shape_candidates', []) or []))
        except Exception:
            candidate_count = 0
        if candidate_count <= 1:
            try:
                btn_prev.disabled = True
                btn_next.disabled = True
            except Exception:
                pass
        popup.open()

    def _delete_detected_items_for_active_capture(self, persist_as_repair_patch=True, include_mappings=True):
        analysis = dict(getattr(self, '_capture_active_template', {}) or {})
        page_idx = int(analysis.get('page', self.current_page_idx()) or self.current_page_idx())
        source_rect = analysis.get('source_rect') or getattr(self, '_capture_preview_rect', None)
        if source_rect is None:
            self.set_status('No captured area available for deletion.')
            return
        try:
            source_rect = source_rect if isinstance(source_rect, fitz.Rect) else fitz.Rect(source_rect)
        except Exception:
            self.set_status('Captured area is invalid for deletion.')
            return

        self._prepare_page_context(page_idx, clear_selection_if_missing=False)
        has_live = bool(getattr(self.engine, 'all_boxes', []) or []) and int(getattr(self.engine, 'detected_page_idx', -1) or -1) == int(page_idx)
        if (not has_live) and (not bool(self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()))):
            self.set_status('Run Detect again before deleting detections in the captured area.')
            return

        target_rects = []
        for rr in list(analysis.get('resolved_rects', []) or []):
            try:
                target_rects.append(rr if isinstance(rr, fitz.Rect) else fitz.Rect(rr))
            except Exception:
                pass
        if not target_rects:
            target_rects = [fitz.Rect(source_rect)]

        remove_ids = []
        removed_refs = []
        all_boxes = list(getattr(self.engine, 'all_boxes', []) or [])
        all_types = list(getattr(self.engine, 'box_types', []) or [])

        for idx_existing, existing in enumerate(all_boxes):
            try:
                ex_rect = existing if isinstance(existing, fitz.Rect) else fitz.Rect(existing)
            except Exception:
                continue
            ex_kind = str(all_types[idx_existing] if idx_existing < len(all_types) else 'field') or 'field'
            remove_hit = False

            # Remove anything strongly matching any resolved target rect.
            for rr in target_rects:
                try:
                    score = float(self.engine._mobile_repair_overlap_score(ex_rect, rr))
                except Exception:
                    score = float(self.engine._rect_iou(ex_rect, rr))
                if score >= 0.45:
                    remove_hit = True
                    break

            # Also remove detections whose center falls inside the full captured ROI,
            # so initial/base detections inside the user-selected area can be deleted too.
            if not remove_hit:
                try:
                    cx = (float(ex_rect.x0) + float(ex_rect.x1)) * 0.5
                    cy = (float(ex_rect.y0) + float(ex_rect.y1)) * 0.5
                    if (float(source_rect.x0) <= cx <= float(source_rect.x1)) and (float(source_rect.y0) <= cy <= float(source_rect.y1)):
                        remove_hit = True
                except Exception:
                    pass

            if remove_hit:
                remove_ids.append(int(idx_existing))
                removed_refs.append((fitz.Rect(ex_rect), ex_kind))

        remove_ids = self._sanitize_box_id_list(remove_ids)
        if not remove_ids:
            self.set_status('No detections overlapped the captured area.')
            return

        if include_mappings:
            try:
                self.clear_mapping_for_box_ids(remove_ids, page_idx)
            except Exception:
                pass

        keep_boxes = []
        keep_types = []
        for idx_existing, existing in enumerate(all_boxes):
            if idx_existing in set(remove_ids):
                continue
            keep_boxes.append(existing)
            keep_types.append(str(all_types[idx_existing] if idx_existing < len(all_types) else 'field'))

        self.engine.all_boxes = list(keep_boxes)
        self.engine.box_types = list(keep_types)
        self.engine.selected_box_ids = []

        try:
            self.engine._cache_detection_for_page(page_idx)
        except Exception:
            pass

        if persist_as_repair_patch and removed_refs:
            try:
                self.engine.save_repair_patch(
                    page_idx,
                    anchor_rects=[fitz.Rect(source_rect)],
                    zone_rect=fitz.Rect(source_rect),
                    add_rects=[],
                    add_types=[],
                    remove_rects=[fitz.Rect(r) for r, _k in removed_refs],
                    remove_types=[str(_k or 'field') for _r, _k in removed_refs],
                )
            except Exception:
                pass

        self._stash_page_selection(page_idx)
        self._clear_manual_capture_preview(preserve_mode=False)
        self._render_session_page(page_idx=page_idx, reason=f'Capture deletion applied ({len(remove_ids)} removed)')
        msg = f'Removed {len(remove_ids)} detection(s) from the captured area'
        if include_mappings:
            msg += ' and cleared linked mappings'
        if persist_as_repair_patch:
            msg += '. This removal will persist on later Detect runs.'
        self.set_status(msg)

    def _apply_confirmed_manual_area_capture(self):
        analysis = dict(getattr(self, '_capture_active_template', {}) or {})
        page_idx = int(analysis.get('page', self.current_page_idx()) or self.current_page_idx())
        rect = analysis.get('source_rect') or getattr(self, '_capture_preview_rect', None)
        if rect is None:
            self.set_status('No captured area to apply.')
            self._clear_manual_capture_preview(preserve_mode=False)
            return
        if not analysis:
            analysis = self.engine.analyze_manual_area_template(page_idx, rect) or {'page': page_idx, 'source_rect': fitz.Rect(rect), 'resolved_rects': [fitz.Rect(rect)], 'resolved_types': ['field'], 'needs_split': False, 'split_count': 1, 'template_kind': 'field', 'source_text_overlap': 0.0, 'features': {}}
        template = self.engine.save_area_template(analysis)
        self._prepare_page_context(page_idx, clear_selection_if_missing=False)
        has_live = bool(getattr(self.engine, 'all_boxes', [])) and int(getattr(self.engine, 'detected_page_idx', -1) or -1) == int(page_idx)
        if (not has_live) and (not bool(self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()))):
            self.set_status('Run Detect again before applying captured area.')
            self._clear_manual_capture_preview(preserve_mode=False)
            return
        merge_result = self.engine.merge_area_template_into_detections(page_idx, template) or {}
        try:
            self.engine.remember_sticky_detections(page_idx, list(merge_result.get('added_refs', []) or []))
        except Exception:
            pass
        added_ids = list(merge_result.get('selected_ids', []) or [])
        if not added_ids:
            added_ids = self._resolve_live_box_ids_for_rect_refs(merge_result.get('added_refs', []) or [], min_score=0.45)
        self._stash_page_selection(page_idx)
        merged_sel = list(self._sanitize_box_id_list(list(getattr(self.engine, 'selected_box_ids', []) or []) + list(added_ids or [])))
        self.engine.selected_box_ids = merged_sel
        self._stash_page_selection(page_idx)
        self._clear_manual_capture_preview(preserve_mode=False)
        self._render_session_page(page_idx=page_idx, reason=f'Captured area applied ({len(added_ids)} added)')
        split_note = 'split into smart green boxes' if bool(template.get('needs_split', False)) else 'used as one target'
        self.set_status(f'Captured area applied on page {page_idx}. Added {len(added_ids)} detections and saved a template ({split_note}).')

    def _ensure_saved_active_area_template(self, analysis=None):
        analysis = dict(analysis or getattr(self, '_capture_active_template', {}) or {})
        if not analysis:
            return None
        existing_id = str(analysis.get('template_id', '') or '')
        if existing_id and existing_id in getattr(self.engine, 'area_template_index', {}):
            return self.engine.area_template_index.get(existing_id)
        template = self.engine.save_area_template(analysis)
        if template is not None:
            analysis['template_id'] = str(template.get('template_id', '') or '')
            self._capture_active_template = analysis
        return template

    def _freeze_active_manual_capture_template(self):
        analysis = dict(getattr(self, '_capture_active_template', {}) or {})
        self._clear_manual_capture_preview(preserve_mode=False)
        if analysis:
            self._capture_active_template = analysis
        return analysis

    def _find_similar_for_active_capture_template(self, *_):
        if bool(getattr(self, '_capture_mode_active', False)):
            self._freeze_active_manual_capture_template()
        analysis = dict(getattr(self, '_capture_active_template', {}) or {})
        page_idx = int(analysis.get('page', self.current_page_idx()) or self.current_page_idx())
        rect = analysis.get('source_rect') or getattr(self, '_capture_preview_rect', None)
        if rect is None:
            self.set_status('No captured area available for similar-match search.')
            return
        if not analysis:
            analysis = self.engine.analyze_manual_area_template(page_idx, rect) or {}
            self._capture_active_template = dict(analysis or {})
        template = self._ensure_saved_active_area_template(analysis)
        if not template:
            self.set_status('Failed to save captured template. Find Similar only runs for templates saved into Saved Area Templates.')
            return
        self._prepare_page_context(page_idx, clear_selection_if_missing=False)
        has_live = bool(getattr(self.engine, 'all_boxes', [])) and int(getattr(self.engine, 'detected_page_idx', -1) or -1) == int(page_idx)
        if (not has_live) and (not bool(self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()))):
            self.set_status('Run Detect again before finding similar areas.')
            return
        template_id = str(template.get('template_id', '') or '')
        matches = self.engine.get_cached_area_template_matches(page_idx, template_id) if template_id else []
        if not matches:
            matches = self.engine.find_similar_area_templates_on_page(page_idx, template, max_matches=18)
            if template_id and matches:
                self.engine.cache_area_template_matches(page_idx, template_id, matches)
        if not matches:
            self.set_status(f'No similar captured areas found on page {page_idx}.')
            return
        self._open_area_template_match_review_popup(template, matches)

    def _find_similar_in_form_for_active_capture_template(self):
        analysis = dict(getattr(self, '_capture_active_template', {}) or {})
        page_idx = int(analysis.get('page', self.current_page_idx()) or self.current_page_idx())
        rect = analysis.get('source_rect') if analysis else None
        if rect is None:
            self.set_status('No captured area available for form-wide search.')
            return
        if not analysis:
            analysis = self.engine.analyze_manual_area_template(page_idx, rect) or {}
            self._capture_active_template = dict(analysis or {})
        template = self._ensure_saved_active_area_template(analysis)
        if not template:
            self.set_status('Failed to save captured template. Find Similar only runs for templates saved into Saved Area Templates.')
            return
        template_id = str(template.get('template_id', '') or '')
        matches = self.engine.get_cached_form_area_template_matches(template_id) if template_id else []
        if not matches:
            matches = self.engine.find_similar_area_templates_in_form(template, max_total=36, per_page_max=10)
            if template_id and matches:
                self.engine.cache_form_area_template_matches(template_id, matches)
        if not matches:
            self.set_status('No similar captured areas found in this form.')
            return
        self._open_area_template_form_match_review_popup(template, matches)

    def _open_area_template_match_review_popup(self, template, matches):
        page_idx = int(template.get('page', self.current_page_idx()) or self.current_page_idx())
        popup_box = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(12))
        self._style_popup_card(popup_box, color=(self.ui_palette or {}).get('surface_alt', (0.11, 0.135, 0.185, 1)), radius=dp(18))
        title_lbl = Label(text='Find Similar on This Page', color=(self.ui_palette or {}).get('text', (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(26), halign='left', valign='middle', font_size=dp(18), bold=True)
        title_lbl.bind(size=self._sync_label_text_size)
        popup_box.add_widget(title_lbl)
        summary = Label(text=(f"Page {page_idx} • {template.get('template_kind', 'field')}\n"
                              f"Template split targets: {int(template.get('split_count', 0) or 0)}\n"
                              f"Similar regions found: {len(matches)}"),
                        color=(self.ui_palette or {}).get('muted', (0.70, 0.78, 0.90, 1)),
                        size_hint_y=None, height=dp(70), halign='left', valign='top', font_size=dp(12))
        summary.bind(size=self._sync_label_text_size)
        popup_box.add_widget(summary)
        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False, bar_width=dp(5))
        rows = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        rows.bind(minimum_height=rows.setter('height'))
        scroll.add_widget(rows)
        popup_box.add_widget(scroll)
        template_id = str(template.get('template_id', '') or '')
        stored_review_state = self.engine.get_area_template_match_review_state(page_idx, template_id) if template_id else {}
        review_state = {}
        for idx, match in enumerate(matches):
            key = str(match.get('match_id', idx))
            checked = bool(stored_review_state.get(key, bool(float(match.get('score', 0.0) or 0.0) >= 0.58)))
            review_state[key] = checked
            row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint_y=None, height=dp(56))
            cb = CheckBox(active=checked, size_hint=(None, None), size=(dp(28), dp(28)))
            lbl = Label(text=(f"[{idx + 1}] score {float(match.get('score', 0.0) or 0.0):.2f} • "
                              f"{int(match.get('split_count', 0) or 0)} target{'s' if int(match.get('split_count', 0) or 0) != 1 else ''}\n"
                              f"{str(match.get('reason', '') or 'structure match')}"),
                        color=(self.ui_palette or {}).get('text', (0.93, 0.96, 1.0, 1)),
                        halign='left', valign='middle', font_size=dp(11))
            lbl.bind(size=self._sync_label_text_size)
            def _sync_state(inst, value, review_key=key):
                review_state[review_key] = bool(value)
            cb.bind(active=_sync_state)
            row.add_widget(cb)
            row.add_widget(lbl)
            row._match_index = idx
            rows.add_widget(row)
        actions = GridLayout(cols=4, spacing=dp(8), size_hint_y=None, height=dp(40))
        btn_all = self._make_compact_action_button('Select All', tone='plain')
        btn_none = self._make_compact_action_button('Select None', tone='ghost')
        btn_apply = self._make_compact_action_button('Apply Checked', tone='accent')
        btn_cancel = self._make_compact_action_button('Cancel', tone='ghost')
        for btn in (btn_all, btn_none, btn_apply, btn_cancel):
            actions.add_widget(btn)
        popup_box.add_widget(actions)
        popup = Popup(title='', separator_height=0, background='', background_color=(0,0,0,0.82), content=popup_box, size_hint=(0.92 if self.ui_mobile else 0.62, 0.78 if self.ui_mobile else 0.68), auto_dismiss=False)
        def _set_all(flag):
            for child in rows.children:
                for widget in child.children:
                    if isinstance(widget, CheckBox):
                        widget.active = bool(flag)
        def _apply(*_):
            chosen = []
            for idx, match in enumerate(matches):
                key = str(match.get('match_id', idx))
                if review_state.get(key, False):
                    chosen.append(match)
            if template_id:
                self.engine.update_area_template_match_review_state(page_idx, template_id, review_state)
            popup.dismiss()
            self._apply_area_template_match_candidates(template, chosen)
        btn_all.bind(on_release=lambda *_: _set_all(True))
        btn_none.bind(on_release=lambda *_: _set_all(False))
        btn_apply.bind(on_release=_apply)
        btn_cancel.bind(on_release=lambda *_: popup.dismiss())
        popup.open()

    def _apply_area_template_match_candidates(self, template, matches):
        page_idx = int(template.get('page', self.current_page_idx()) or self.current_page_idx())
        if not matches:
            self.set_status('No similar area matches selected.')
            return
        self._prepare_page_context(page_idx, clear_selection_if_missing=False)
        has_live = bool(getattr(self.engine, 'all_boxes', [])) and int(getattr(self.engine, 'detected_page_idx', -1) or -1) == int(page_idx)
        if (not has_live) and (not bool(self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()))):
            self.set_status('Run Detect again before applying similar area matches.')
            return
        merge_result = self.engine.merge_area_template_matches_into_detections(page_idx, matches) or {}
        try:
            self.engine.remember_sticky_detections(page_idx, list(merge_result.get('added_refs', []) or []))
        except Exception:
            pass
        self.engine.invalidate_area_template_match_cache(page_idx=page_idx, template_id=str(template.get('template_id', '') or ''))
        added_ids = list(merge_result.get('selected_ids', []) or [])
        if not added_ids:
            added_ids = self._resolve_live_box_ids_for_rect_refs(merge_result.get('added_refs', []) or [], min_score=0.45)
        self._stash_page_selection(page_idx)
        merged_sel = list(self._sanitize_box_id_list(list(getattr(self.engine, 'selected_box_ids', []) or []) + list(added_ids or [])))
        self.engine.selected_box_ids = merged_sel
        self._stash_page_selection(page_idx)
        self._render_session_page(page_idx=page_idx, reason=f'Similar areas applied ({len(added_ids)} added)')
        self._capture_active_template = None
        self.set_status(f'Applied {len(added_ids)} detections from similar captured areas on page {page_idx}.')


    def _open_area_template_form_match_review_popup(self, template, matches):
        popup_box = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(12))
        self._style_popup_card(popup_box, color=(self.ui_palette or {}).get('surface_alt', (0.11, 0.135, 0.185, 1)), radius=dp(18))
        title_lbl = Label(text='Find Similar in This Form', color=(self.ui_palette or {}).get('text', (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(26), halign='left', valign='middle', font_size=dp(18), bold=True)
        title_lbl.bind(size=self._sync_label_text_size)
        popup_box.add_widget(title_lbl)
        page_hits = sorted(set(int(item.get('page', 0) or 0) for item in (matches or [])))
        summary = Label(text=(f"Template: {str(template.get('template_kind', 'field') or 'field')}\n"
                              f"Pages with matches: {', '.join(str(p) for p in page_hits)}\n"
                              f"Similar regions found: {len(matches)}"),
                        color=(self.ui_palette or {}).get('muted', (0.70, 0.78, 0.90, 1)),
                        size_hint_y=None, height=dp(70), halign='left', valign='top', font_size=dp(12))
        summary.bind(size=self._sync_label_text_size)
        popup_box.add_widget(summary)
        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False, bar_width=dp(5))
        rows = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        rows.bind(minimum_height=rows.setter('height'))
        scroll.add_widget(rows)
        popup_box.add_widget(scroll)
        template_id = str(template.get('template_id', '') or '')
        saved_review = self.engine.get_form_area_template_match_review_state(template_id) if template_id else {}
        review_state = {}
        for idx, match in enumerate(matches):
            key = str(match.get('match_id', idx))
            checked = bool(saved_review.get(key, float(match.get('score', 0.0) or 0.0) >= 0.60))
            review_state[key] = checked
            row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint_y=None, height=dp(58))
            cb = CheckBox(active=checked, size_hint=(None, None), size=(dp(28), dp(28)))
            lbl = Label(text=(f"P{int(match.get('page', 0) or 0)} • score {float(match.get('score', 0.0) or 0.0):.2f} • "
                              f"{int(match.get('split_count', 0) or 0)} target{'s' if int(match.get('split_count', 0) or 0) != 1 else ''}\n"
                              f"{str(match.get('reason', '') or 'structure match')}"),
                        color=(self.ui_palette or {}).get('text', (0.93, 0.96, 1.0, 1)),
                        halign='left', valign='middle', font_size=dp(11))
            lbl.bind(size=self._sync_label_text_size)
            def _sync_state(inst, value, review_key=key):
                review_state[review_key] = bool(value)
            cb.bind(active=_sync_state)
            row.add_widget(cb)
            row.add_widget(lbl)
            rows.add_widget(row)
        actions = GridLayout(cols=6, spacing=dp(8), size_hint_y=None, height=dp(40))
        btn_all = self._make_compact_action_button('Select All', tone='plain')
        btn_none = self._make_compact_action_button('Select None', tone='ghost')
        btn_this = self._make_compact_action_button(f'Select P{int(self.current_page_idx())}', tone='plain')
        btn_clear_this = self._make_compact_action_button(f'Clear P{int(self.current_page_idx())}', tone='ghost')
        btn_apply = self._make_compact_action_button('Apply Checked', tone='accent')
        btn_cancel = self._make_compact_action_button('Cancel', tone='ghost')
        for btn in (btn_all, btn_none, btn_this, btn_clear_this, btn_apply, btn_cancel):
            actions.add_widget(btn)
        popup_box.add_widget(actions)
        popup = Popup(title='', separator_height=0, background='', background_color=(0,0,0,0.82), content=popup_box, size_hint=(0.96 if self.ui_mobile else 0.72, 0.82 if self.ui_mobile else 0.76), auto_dismiss=False)
        current_page = int(self.current_page_idx())
        def _set_all(flag):
            for child in rows.children:
                for widget in child.children:
                    if isinstance(widget, CheckBox):
                        widget.active = bool(flag)
        def _set_current_page(flag):
            for idx, match in enumerate(matches):
                if int(match.get('page', 0) or 0) != current_page:
                    continue
                key = str(match.get('match_id', idx))
                review_state[key] = bool(flag)
            for child in rows.children:
                try:
                    match_idx = getattr(child, '_match_index', None)
                    if match_idx is None:
                        continue
                    match = matches[int(match_idx)]
                    if int(match.get('page', 0) or 0) != current_page:
                        continue
                    for widget in child.children:
                        if isinstance(widget, CheckBox):
                            widget.active = bool(flag)
                except Exception:
                    continue
        def _apply(*_):
            chosen = []
            for idx, match in enumerate(matches):
                key = str(match.get('match_id', idx))
                if review_state.get(key, False):
                    chosen.append(match)
            if template_id:
                self.engine.update_form_area_template_match_review_state(template_id, review_state)
            popup.dismiss()
            self._apply_area_template_form_match_candidates(template, chosen)
        for idx, child in enumerate(rows.children):
            pass
        btn_all.bind(on_release=lambda *_: _set_all(True))
        btn_none.bind(on_release=lambda *_: _set_all(False))
        btn_this.bind(on_release=lambda *_: _set_current_page(True))
        btn_clear_this.bind(on_release=lambda *_: _set_current_page(False))
        btn_apply.bind(on_release=_apply)
        btn_cancel.bind(on_release=lambda *_: popup.dismiss())
        popup.open()

    def _apply_area_template_form_match_candidates(self, template, matches):
        if not matches:
            self.set_status('No form-wide similar area matches selected.')
            return
        by_page = {}
        for item in (matches or []):
            if not isinstance(item, dict):
                continue
            page_idx = int(item.get('page', 0) or 0)
            by_page.setdefault(page_idx, []).append(item)
        added_total = 0
        selected_ids = []
        current_page = int(self.current_page_idx())
        template_id = str((template or {}).get('template_id', '') or '')
        touched_pages = []
        for page_idx in sorted(by_page):
            has_live = bool(getattr(self.engine, 'all_boxes', [])) and int(getattr(self.engine, 'detected_page_idx', -1) or -1) == int(page_idx)
            if (not has_live) and (not bool(self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()))):
                continue
            merge_result = self.engine.merge_area_template_matches_into_detections(page_idx, by_page[page_idx]) or {}
            try:
                self.engine.remember_sticky_detections(page_idx, list(merge_result.get('added_refs', []) or []))
            except Exception:
                pass
            touched_pages.append(int(page_idx))
            try:
                if template_id:
                    self.engine.invalidate_area_template_match_cache(page_idx=page_idx, template_id=template_id)
                    self.engine.invalidate_area_template_match_cache(template_id=template_id)
            except Exception:
                pass
            page_added = list(merge_result.get('selected_ids', []) or [])
            if not page_added:
                page_added = self._resolve_live_box_ids_for_rect_refs(merge_result.get('added_refs', []) or [], min_score=0.45)
            added_total += len(page_added)
            if page_idx == current_page:
                selected_ids.extend(page_added)
        if selected_ids:
            self._stash_page_selection(current_page)
            merged_sel = list(self._sanitize_box_id_list(list(getattr(self.engine, 'selected_box_ids', []) or []) + list(selected_ids or [])))
            self.engine.selected_box_ids = merged_sel
            self._stash_page_selection(current_page)
        self._render_session_page(page_idx=current_page, reason=f'Form-wide similar areas applied ({added_total} added)')
        self._capture_active_template = None
        if touched_pages:
            pages_txt = ', '.join(str(p) for p in sorted(set(touched_pages)))
            self.set_status(f'Applied {added_total} detections from similar captured areas across this form (pages {pages_txt}).')
        else:
            self.set_status('No form-wide similar areas could be applied to the current detection state.')


    def _templates_for_page(self, page_idx=None):
        try:
            page_idx = int(self.current_page_idx() if page_idx is None else page_idx)
        except Exception:
            page_idx = int(self.current_page_idx())
        items = list(((getattr(self.engine, 'area_templates_by_page', {}) or {}).get(page_idx, []) or []))
        items.sort(key=lambda item: str(item.get('created_at', '') or ''), reverse=True)
        return items

    def _all_saved_area_templates(self, scope='page', page_idx=None, query='', sort_mode='recent'):
        try:
            page_idx = int(self.current_page_idx() if page_idx is None else page_idx)
        except Exception:
            page_idx = int(self.current_page_idx())
        area_map = dict(getattr(self.engine, 'area_templates_by_page', {}) or {})
        items = []
        if str(scope or 'page') == 'form':
            page_keys = sorted(int(k) for k in area_map.keys())
            if page_idx in page_keys:
                page_keys.remove(page_idx)
                page_keys = [page_idx] + page_keys
        else:
            page_keys = [page_idx]
        for p in page_keys:
            for item in list(area_map.get(p, []) or []):
                try:
                    entry = dict(item or {})
                except Exception:
                    continue
                entry['_page_idx'] = int(p)
                items.append(entry)
        q = str(query or '').strip().lower()
        if q:
            filtered = []
            for item in items:
                page_label = f"page {int(item.get('_page_idx', page_idx) or page_idx)}"
                bits = [
                    page_label,
                    str(item.get('template_kind', '') or ''),
                    str(item.get('created_at', '') or ''),
                    self._template_type_mix_text(item),
                    f"text {float(item.get('source_text_overlap', 0.0) or 0.0):.2f}",
                    f"targets {int(item.get('split_count', 0) or 0)}",
                ]
                hay = ' '.join(bits).lower()
                if q in hay:
                    filtered.append(item)
            items = filtered
        mode = str(sort_mode or 'recent').strip().lower()
        if mode == 'page':
            items.sort(key=lambda item: (int(item.get('_page_idx', 0) or 0), str(item.get('created_at', '') or '')), reverse=False)
        elif mode == 'kind':
            items.sort(key=lambda item: (str(item.get('template_kind', '') or ''), -int(item.get('_page_idx', 0) or 0), str(item.get('created_at', '') or '')))
        else:
            items.sort(key=lambda item: str(item.get('created_at', '') or ''), reverse=True)
        return items

    def _find_saved_area_template_page(self, template_id):
        template_id = str(template_id or '').strip()
        if not template_id:
            return None
        area_map = dict(getattr(self.engine, 'area_templates_by_page', {}) or {})
        for page_key, page_items in area_map.items():
            try:
                page_idx = int(page_key)
            except Exception:
                continue
            for item in list(page_items or []):
                if str(item.get('template_id', '') or '') == template_id:
                    return page_idx
        return None

    def _delete_saved_area_template(self, template_id, page_idx=None):
        template_id = str(template_id or '').strip()
        if not template_id:
            return False
        located_page = self._find_saved_area_template_page(template_id)
        if located_page is not None:
            page_idx = located_page
        try:
            page_idx = int(self.current_page_idx() if page_idx is None else page_idx)
        except Exception:
            page_idx = int(self.current_page_idx())
        page_items = list(((getattr(self.engine, 'area_templates_by_page', {}) or {}).get(page_idx, []) or []))
        kept = [item for item in page_items if str(item.get('template_id', '') or '') != template_id]
        changed = len(kept) != len(page_items)
        if changed:
            if kept:
                self.engine.area_templates_by_page[page_idx] = kept
            else:
                try:
                    del self.engine.area_templates_by_page[page_idx]
                except Exception:
                    pass
            try:
                self.engine.invalidate_area_template_match_cache(page_idx=page_idx, template_id=template_id)
            except Exception:
                pass
        try:
            if template_id in (getattr(self.engine, 'area_template_index', {}) or {}):
                del self.engine.area_template_index[template_id]
        except Exception:
            pass
        return bool(changed)

    def _compact_datetime_label(self, value):
        txt = str(value or '').strip()
        if not txt:
            return 'unknown time'
        txt = txt.replace('T', ' ').replace('Z', ' UTC')
        return txt[:19] if len(txt) >= 19 else txt

    def _template_type_mix_text(self, template):
        types = [str(v or '') for v in (dict(template or {}).get('resolved_types', []) or []) if str(v or '').strip()]
        if not types:
            return 'Types: n/a'
        counts = {}
        for kind in types:
            counts[kind] = int(counts.get(kind, 0) or 0) + 1
        bits = []
        for kind in ('field', 'check', 'line'):
            count = int(counts.get(kind, 0) or 0)
            if count > 0:
                bits.append(f"{kind}s {count}")
        return 'Types: ' + (', '.join(bits) if bits else ', '.join(f"{k} {v}" for k, v in counts.items()))

    def _build_rect_thumbnail_widget(self, page_idx, rect, width=100, height=72, fallback_text='Template'):
        palette = getattr(self, 'ui_palette', {}) or {}
        holder = BoxLayout(orientation='vertical', size_hint=(None, None), size=(dp(width), dp(height)))
        self._style_popup_card(holder, color=palette.get('surface', (0.09, 0.12, 0.18, 1)), radius=dp(10), border_color=palette.get('border', (0.18, 0.24, 0.34, 1)))
        label = Label(text=str(fallback_text or 'Template'), color=palette.get('text', (0.93, 0.96, 1.0, 1)), halign='center', valign='middle', font_size=dp(10.5))
        label.bind(size=self._sync_label_text_size)
        holder.add_widget(label)
        try:
            rect_obj = rect if isinstance(rect, fitz.Rect) else fitz.Rect(rect)
        except Exception:
            rect_obj = None
        if rect_obj is None:
            return holder
        try:
            img = self.engine._render_pdf_page_bgr(self.engine.pdf_path, page_idx=int(page_idx or 0), preview_zoom=1.6)
            if img is None or getattr(img, 'size', 0) == 0:
                return holder
            h_img, w_img = img.shape[:2]
            x0 = max(0, min(int(rect_obj.x0 * 1.6), max(w_img - 1, 0)))
            y0 = max(0, min(int(rect_obj.y0 * 1.6), max(h_img - 1, 0)))
            x1 = max(x0 + 1, min(int(rect_obj.x1 * 1.6), w_img))
            y1 = max(y0 + 1, min(int(rect_obj.y1 * 1.6), h_img))
            crop = img[y0:y1, x0:x1].copy()
            if crop.size == 0:
                return holder
            crop = cv2.cvtColor(crop, cv2.COLOR_BGR2RGBA)
            crop = cv2.resize(crop, (int(width), int(height)), interpolation=cv2.INTER_AREA)
            tex = Texture.create(size=(crop.shape[1], crop.shape[0]), colorfmt='rgba')
            tex.blit_buffer(crop.tobytes(), colorfmt='rgba', bufferfmt='ubyte')
            tex.flip_vertical()
            holder.clear_widgets()
            holder.add_widget(Image(texture=tex, allow_stretch=True, keep_ratio=False))
        except Exception:
            return holder
        return holder

    def _apply_saved_area_template_direct(self, template):
        if not isinstance(template, dict):
            self.set_status('Saved template is unavailable.')
            return
        page_idx = int(self.current_page_idx())
        self._prepare_page_context(page_idx, clear_selection_if_missing=False)
        has_live = bool(getattr(self.engine, 'all_boxes', [])) and int(getattr(self.engine, 'detected_page_idx', -1) or -1) == int(page_idx)
        if (not has_live) and (not bool(self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()))):
            self.set_status('Run Detect again before using a saved template.')
            return
        merge_result = self.engine.merge_area_template_into_detections(page_idx, template) or {}
        added_ids = list(merge_result.get('selected_ids', []) or [])
        if not added_ids:
            added_ids = self._resolve_live_box_ids_for_rect_refs(merge_result.get('added_refs', []) or [], min_score=0.45)
        self._stash_page_selection(page_idx)
        merged_sel = list(self._sanitize_box_id_list(list(getattr(self.engine, 'selected_box_ids', []) or []) + list(added_ids or [])))
        self.engine.selected_box_ids = merged_sel
        self._stash_page_selection(page_idx)
        self._render_session_page(page_idx=page_idx, reason=f'Saved template used ({len(added_ids)} added)')
        added_count = int(len(merge_result.get('added_refs', []) or []) or len(added_ids))
        if added_count > 0:
            self.set_status(f'Applied saved template on page {page_idx}. Added {added_count} detection(s).')
        else:
            self.set_status('Saved template matched existing detections, so nothing new was added.')

    def _run_saved_area_template_on_current_page(self, template):
        if not isinstance(template, dict):
            self.set_status('Saved template is unavailable.')
            return
        page_idx = int(self.current_page_idx())
        self._prepare_page_context(page_idx, clear_selection_if_missing=False)
        has_live = bool(getattr(self.engine, 'all_boxes', [])) and int(getattr(self.engine, 'detected_page_idx', -1) or -1) == int(page_idx)
        if (not has_live) and (not bool(self.engine._restore_detection_for_page(page_idx, required_signature=self.engine._settings_signature()))):
            self.set_status('Run Detect again before using a saved template.')
            return
        template_id = str(template.get('template_id', '') or '')
        matches = self.engine.find_similar_area_templates_on_page(page_idx, template, max_matches=18)
        if template_id:
            self.engine.cache_area_template_matches(page_idx, template_id, matches or [])
        if not matches:
            self.set_status(f'No similar regions found for template on page {page_idx}.')
            return
        self._capture_active_template = None
        self._open_area_template_match_review_popup(template, matches)

    def _find_similar_in_form_for_saved_template(self, template):
        if not isinstance(template, dict):
            self.set_status('Saved template is unavailable.')
            return
        template_id = str(template.get('template_id', '') or '')
        matches = self.engine.find_similar_area_templates_in_form(template, max_total=36, per_page_max=10)
        if template_id:
            self.engine.cache_form_area_template_matches(template_id, matches or [])
        if not matches:
            self.set_status('No similar regions found for this template in the current form.')
            return
        self._capture_active_template = None
        self._open_area_template_form_match_review_popup(template, matches)

    def _open_area_template_library_popup(self, *_):
        if bool(getattr(self, '_capture_mode_active', False)):
            self._freeze_active_manual_capture_template()
        page_idx = int(self.current_page_idx())
        palette = getattr(self, 'ui_palette', {}) or {}
        popup_box = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(12))
        self._style_popup_card(popup_box, color=palette.get('surface_alt', (0.11, 0.135, 0.185, 1)), radius=dp(18))
        title_lbl = Label(text='Saved Area Templates', color=palette.get('text', (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(26), halign='left', valign='middle', font_size=dp(18), bold=True)
        title_lbl.bind(size=self._sync_label_text_size)
        popup_box.add_widget(title_lbl)
        state = {'scope': 'page', 'sort': 'recent', 'query': '', 'mode': 'compact'}

        control_row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint_y=None, height=dp(38))
        btn_page = self._make_compact_action_button('Current Page', tone='accent')
        btn_form = self._make_compact_action_button('Whole Form', tone='plain')
        sort_spinner = Spinner(text='Recent', values=('Recent', 'By Page', 'By Kind'), size_hint=(None, 1), width=dp(118), background_normal='', background_color=palette.get('surface', (0.09, 0.12, 0.18, 1)), color=palette.get('text', (0.93, 0.96, 1.0, 1)))
        btn_mode = self._make_compact_action_button('Compact', tone='plain')
        control_row.add_widget(btn_page)
        control_row.add_widget(btn_form)
        control_row.add_widget(sort_spinner)
        control_row.add_widget(btn_mode)
        popup_box.add_widget(control_row)

        search_input = TextInput(text='', hint_text='Search kind, page, type, date, overlap', multiline=False, size_hint_y=None, height=dp(42), background_normal='', background_active='', background_color=palette.get('surface', (0.09, 0.12, 0.18, 1)), foreground_color=palette.get('text', (0.93, 0.96, 1.0, 1)), hint_text_color=palette.get('muted', (0.70, 0.78, 0.90, 1)), padding=[dp(12), dp(10), dp(12), dp(10)])
        popup_box.add_widget(search_input)

        summary = Label(text='', color=palette.get('muted', (0.70, 0.78, 0.90, 1)), size_hint_y=None, height=dp(44), halign='left', valign='middle', font_size=dp(12))
        summary.bind(size=self._sync_label_text_size)
        popup_box.add_widget(summary)

        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False, bar_width=dp(5))
        rows = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        rows.bind(minimum_height=rows.setter('height'))
        scroll.add_widget(rows)
        popup_box.add_widget(scroll)

        footer = GridLayout(cols=1, spacing=dp(8), size_hint_y=None, height=dp(40))
        btn_close = self._make_compact_action_button('Close', tone='ghost')
        footer.add_widget(btn_close)
        popup_box.add_widget(footer)
        popup = Popup(title='', separator_height=0, background='', background_color=(0, 0, 0, 0.82), content=popup_box, size_hint=(0.94 if self.ui_mobile else 0.78, 0.86 if self.ui_mobile else 0.82), auto_dismiss=False)

        def _scope_items():
            return self._all_saved_area_templates(scope=state['scope'], page_idx=page_idx, query=state['query'], sort_mode=state['sort'])

        def _set_scope(scope_name):
            state['scope'] = str(scope_name or 'page')
            btn_page.tone = 'accent' if state['scope'] == 'page' else 'plain'
            btn_form.tone = 'accent' if state['scope'] == 'form' else 'plain'
            try:
                btn_page._update_graphics(); btn_form._update_graphics()
            except Exception:
                pass
            _refresh_rows()

        def _render_template_row(template, display_index=None, target_rows=None):
            actual_page = int(template.get('_page_idx', page_idx) or page_idx)
            expanded = (state['mode'] == 'expanded')
            row = BoxLayout(orientation='horizontal' if expanded else 'vertical', spacing=dp(8), size_hint_y=None, padding=[dp(8), dp(8), dp(8), dp(8)])
            row.height = dp(134 if expanded else 92)
            self._style_popup_card(row, color=palette.get('surface_soft', (0.135, 0.16, 0.215, 1)), radius=dp(14))
            if expanded:
                thumb = self._build_rect_thumbnail_widget(actual_page, template.get('source_rect'), width=96 if self.ui_mobile else 108, height=74 if self.ui_mobile else 82, fallback_text=f'P{actual_page}')
                row.add_widget(thumb)
            body = BoxLayout(orientation='vertical', spacing=dp(6))
            split_count = int(template.get('split_count', 0) or 0)
            template_kind = str(template.get('template_kind', 'field') or 'field')
            prefix = f"[{display_index}] " if display_index is not None else ''
            top_text = f"{prefix}P{actual_page} • {template_kind} • {split_count} target{'s' if split_count != 1 else ''}"
            if state['scope'] == 'form':
                top_text = f"{prefix}Page {actual_page} • {template_kind} • {split_count} target{'s' if split_count != 1 else ''}"
            info_text = top_text + "\nSaved: " + self._compact_datetime_label(template.get('created_at', ''))
            info = Label(text=info_text, color=palette.get('text', (0.93, 0.96, 1.0, 1)), halign='left', valign='middle', font_size=dp(11.2))
            info.bind(size=self._sync_label_text_size)
            body.add_widget(info)
            meta_row = BoxLayout(orientation='horizontal', spacing=dp(6), size_hint_y=None, height=dp(28))
            meta_row.add_widget(self._make_popup_chip(self._template_type_mix_text(template), tone='muted'))
            try:
                overlap_val = float(template.get('source_text_overlap', 0.0) or 0.0)
            except Exception:
                overlap_val = 0.0
            meta_row.add_widget(self._make_popup_chip(f"Text {overlap_val:.2f}", tone=('danger' if overlap_val >= 0.24 else 'accent')))
            body.add_widget(meta_row)
            action_cols = 4 if expanded else 4
            actions = GridLayout(cols=action_cols, spacing=dp(8), size_hint_y=None, height=dp(44))
            btn_use = self._make_compact_action_button('Use Area', tone='accent')
            btn_run = self._make_compact_action_button('Find Page', tone='plain')
            btn_form2 = self._make_compact_action_button('Find Form', tone='plain')
            btn_delete = self._make_compact_action_button('Delete', tone='danger')
            btn_use.bind(on_release=lambda *_ , tpl=template: (popup.dismiss(), self._apply_saved_area_template_direct(tpl)))
            btn_run.bind(on_release=lambda *_ , tpl=template: (popup.dismiss(), self._run_saved_area_template_on_current_page(tpl)))
            btn_form2.bind(on_release=lambda *_ , tpl=template: (popup.dismiss(), self._find_similar_in_form_for_saved_template(tpl)))
            btn_delete.bind(on_release=lambda *_ , tpl=template, pg=actual_page: (self._delete_saved_area_template(str(tpl.get('template_id', '') or ''), page_idx=pg), _refresh_rows(), self.set_status('Saved template deleted.')))
            for btn in (btn_use, btn_run, btn_form2, btn_delete):
                actions.add_widget(btn)
            body.add_widget(actions)
            row.add_widget(body)
            (target_rows or rows).add_widget(row)

        def _refresh_rows(*_):
            rows.clear_widgets()
            items = _scope_items()
            total_templates = sum(len(v or []) for v in ((getattr(self.engine, 'area_templates_by_page', {}) or {}).values()))
            if state['scope'] == 'page':
                summary.text = f"Page {page_idx} • {len(items)} visible template{'s' if len(items) != 1 else ''} • whole form {total_templates}"
            else:
                page_set = sorted(set(int(item.get('_page_idx', page_idx) or page_idx) for item in items))
                summary.text = f"Whole form • {len(items)} visible template{'s' if len(items) != 1 else ''} across {len(page_set)} page{'s' if len(page_set) != 1 else ''}"
            if not items:
                empty_text = 'No saved templates matched the current view.' if state['query'] else ('No saved templates for this page yet. Switch to Whole Form to browse other pages.' if state['scope'] == 'page' else 'No saved templates in this form yet.')
                empty_lbl = Label(text=empty_text, color=palette.get('muted', (0.70, 0.78, 0.90, 1)), size_hint_y=None, height=dp(52), halign='left', valign='middle', font_size=dp(12))
                empty_lbl.bind(size=self._sync_label_text_size)
                rows.add_widget(empty_lbl)
                return
            if state['scope'] == 'form':
                grouped = {}
                for item in items:
                    grouped.setdefault(int(item.get('_page_idx', page_idx) or page_idx), []).append(item)
                order = list(grouped.keys())
                if page_idx in order:
                    order.remove(page_idx)
                    order = [page_idx] + sorted(order)
                else:
                    order = sorted(order)
                display_index = 1
                for grp_page in order:
                    section = BoxLayout(orientation='vertical', spacing=dp(8), size_hint_y=None, padding=[dp(8), dp(8), dp(8), dp(8)])
                    self._style_popup_card(section, color=palette.get('surface', (0.09, 0.12, 0.18, 1)), radius=dp(12))
                    header = Label(text=f"Page {grp_page}{' • current page' if grp_page == page_idx else ''} • {len(grouped.get(grp_page, []))} template{'s' if len(grouped.get(grp_page, [])) != 1 else ''}", color=palette.get('text', (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(24), halign='left', valign='middle', font_size=dp(12), bold=True)
                    header.bind(size=self._sync_label_text_size)
                    section.add_widget(header)
                    inner = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
                    inner.bind(minimum_height=inner.setter('height'))
                    section.add_widget(inner)
                    rows.add_widget(section)
                    for item in grouped.get(grp_page, []):
                        _render_template_row(item, display_index, target_rows=inner)
                        display_index += 1
                    section.height = header.height + inner.height + dp(24)
            else:
                for idx, item in enumerate(items, start=1):
                    _render_template_row(item, idx)

        btn_page.bind(on_release=lambda *_: _set_scope('page'))
        btn_form.bind(on_release=lambda *_: _set_scope('form'))
        sort_spinner.bind(text=lambda inst, value: (state.__setitem__('sort', {'Recent': 'recent', 'By Page': 'page', 'By Kind': 'kind'}.get(str(value), 'recent')), _refresh_rows()))
        search_input.bind(text=lambda inst, value: (state.__setitem__('query', str(value or '')), _refresh_rows()))
        def _toggle_mode(*_):
            state['mode'] = 'expanded' if state['mode'] == 'compact' else 'compact'
            btn_mode.text = 'Expanded' if state['mode'] == 'expanded' else 'Compact'
            _refresh_rows()
        btn_mode.bind(on_release=_toggle_mode)
        _set_scope('page')
        btn_close.bind(on_release=lambda *_: popup.dismiss())
        popup.open()

    def on_preview_box_hover(self, hit):
        is_mobile = bool(getattr(self, "ui_mobile", False))
        verb = "Selected" if is_mobile else "Hover"
        if not hit or "id" not in hit:
            if hasattr(self, "inspector_hover_lbl") and self.inspector_hover_lbl is not None:
                self.inspector_hover_lbl.text = ("Tap or drag over a box to review it." if is_mobile else "Hover a box to review it.")
            if hasattr(self, "desktop_status_detail_lbl") and self.desktop_status_detail_lbl is not None:
                self.desktop_status_detail_lbl.text = "Ready • No box targeted"
            self._update_preview_hud(hit=None)
            return
        try:
            hit_id = int(hit.get("id"))
        except Exception:
            self._update_preview_hud(hit=None)
            return
        mapping = hit.get("mapping") or "Unmapped"
        hover_text = f"{verb} Box {hit_id} | Type: {hit.get('t', 'field')} | {mapping}"
        sel_ids = sorted(set(int(x) for x in getattr(self.engine, 'selected_box_ids', []) if isinstance(x, int) or str(x).isdigit()))
        if hasattr(self, "preview_info") and self.preview_info is not None:
            if sel_ids:
                first = sel_ids[0]
                sel_mapping = self.engine.describe_box_mapping(first, self.current_page_idx())
                self.preview_info.text = f"Selected: {sel_ids} • {sel_mapping}"
            else:
                self.preview_info.text = hover_text
        if hasattr(self, "desktop_status_detail_lbl") and self.desktop_status_detail_lbl is not None:
            self.desktop_status_detail_lbl.text = f"{verb} • Box {hit_id} • {mapping}"
        if hasattr(self, "inspector_hover_lbl") and self.inspector_hover_lbl is not None:
            self.inspector_hover_lbl.text = hover_text
        self._update_preview_hud(hit=hit)

    def on_preview_box_tap(self, hit):
        if not hit or "id" not in hit:
            return
        try:
            box_id = int(hit["id"])
        except Exception:
            return
        existing = []
        for x in getattr(self.engine, "selected_box_ids", []) or []:
            if isinstance(x, int) or str(x).isdigit():
                v = int(x)
                if v not in existing:
                    existing.append(v)

        if getattr(self, "ui_mobile", False):
            if bool(getattr(self, "mobile_select_mode", False)):
                if box_id in existing:
                    existing = [v for v in existing if v != box_id]
                else:
                    existing.append(box_id)
            else:
                existing = [box_id]
        else:
            if box_id in existing:
                existing = [v for v in existing if v != box_id]
            else:
                existing.append(box_id)

        self.engine.selected_box_ids = existing
        if existing:
            self._load_mapping_into_editor(box_id)
        self._sync_box_selection_ui()
        self._update_preview_hud()

    def on_preview_box_double_tap(self, hit):
        if not hit or "id" not in hit:
            return
        try:
            box_id = int(hit["id"])
        except Exception:
            return

        # Guard mobile Select ON mode so accidental double-tap dispatches from a single tap
        # do not open the link popup. In Select ON, single taps should only manage selection.
        if getattr(self, "ui_mobile", False) and bool(getattr(self, "mobile_select_mode", False)):
            existing = []
            for x in getattr(self.engine, "selected_box_ids", []) or []:
                if isinstance(x, int) or str(x).isdigit():
                    v = int(x)
                    if v not in existing:
                        existing.append(v)
            if box_id not in existing:
                existing.append(box_id)
            self.engine.selected_box_ids = existing
            self._load_mapping_into_editor(box_id)
            self._sync_box_selection_ui()
            self._focus_preview_box(box_id)
            self._update_preview_hud()
            return

        existing = []
        for x in getattr(self.engine, "selected_box_ids", []) or []:
            if isinstance(x, int) or str(x).isdigit():
                v = int(x)
                if v not in existing:
                    existing.append(v)

        if box_id not in existing:
            existing.append(box_id)

        self.engine.selected_box_ids = existing
        self._load_mapping_into_editor(box_id)
        self._sync_box_selection_ui()
        self._focus_preview_box(box_id)
        self._update_preview_hud()
        self.open_mapping_popup_for_selection(primary_box_id=box_id)

    def _style_popup_card(self, widget, color=None, radius=None, border_color=None):
        palette = getattr(self, "ui_palette", {}) or {}
        color = color or palette.get("surface_alt", (0.11, 0.135, 0.185, 1))
        border_color = border_color or palette.get("border", (0.18, 0.24, 0.34, 1))
        radius = dp(18) if radius is None else radius
        with widget.canvas.before:
            widget._popup_shadow_color = Color(0, 0, 0, 0.24)
            widget._popup_shadow = RoundedRectangle(radius=[radius + dp(4)] * 4)
            widget._popup_glow_color = Color(*palette.get("aurora", (0.30, 0.92, 1.0, 1))[:3], 0.045)
            widget._popup_glow = RoundedRectangle(radius=[radius + dp(3)] * 4)
            widget._popup_bg_color = Color(*color)
            widget._popup_bg = RoundedRectangle(radius=[radius] * 4)
            widget._popup_nebula_color = Color(*palette.get("secondary", (0.69, 0.43, 1.0, 1))[:3], 0.085)
            widget._popup_nebula = RoundedRectangle(radius=[max(dp(10), radius - dp(2))] * 4)
            widget._popup_shine_color = Color(*palette.get("starlight", (1.0, 0.95, 0.8, 1))[:3], 0.050)
            widget._popup_shine = RoundedRectangle(radius=[max(dp(10), radius - dp(5))] * 4)
            widget._popup_border_color = Color(*border_color)
            widget._popup_border = Line(rounded_rectangle=[0, 0, 0, 0, radius], width=1.05)
        def _upd(*_):
            x, y = widget.pos
            w, h = widget.size
            widget._popup_shadow.pos = (x, y - dp(2))
            widget._popup_shadow.size = (w, h + dp(3))
            widget._popup_glow.pos = (x - dp(1), y - dp(1))
            widget._popup_glow.size = (w + dp(2), h + dp(2))
            widget._popup_bg.pos = (x, y)
            widget._popup_bg.size = (w, h)
            widget._popup_nebula.pos = (x + dp(1), y + dp(1))
            widget._popup_nebula.size = (max(0, w - dp(2)), max(0, h * 0.50))
            widget._popup_shine.pos = (x + dp(1.5), y + h - max(dp(16), h * 0.34) - dp(1.5))
            widget._popup_shine.size = (max(0, w - dp(3)), max(dp(16), h * 0.34))
            widget._popup_border.rounded_rectangle = [x, y, w, h, radius]
        widget.bind(pos=_upd, size=_upd)
        _upd()
        return widget

    def _make_popup_chip(self, text, tone="muted"):
        palette = getattr(self, "ui_palette", {}) or {}
        tone_map = {
            "primary": palette.get("primary_soft", (0.16, 0.24, 0.38, 1)),
            "accent": palette.get("accent_soft", (0.10, 0.28, 0.24, 1)),
            "muted": palette.get("chip", (0.13, 0.18, 0.27, 1)),
            "danger": (0.30, 0.12, 0.15, 1),
        }
        chip = BoxLayout(orientation="horizontal", size_hint_y=None, height=dp(34), padding=[dp(10), dp(6), dp(10), dp(6)])
        self._style_popup_card(chip, tone_map.get(tone, tone_map["muted"]), radius=dp(14))
        lbl = Label(text=str(text), color=palette.get("text", (0.93, 0.96, 1.0, 1)), halign="center", valign="middle", font_size=dp(11))
        lbl.bind(size=self._sync_label_text_size)
        chip.add_widget(lbl)
        return chip


    def _set_popup_background_softened(self, is_open, state_store):
        for attr in ("btn_mobile_tools_toggle", "mobile_bottom_tools", "preview_hud"):
            widget = getattr(self, attr, None)
            if widget is None:
                continue
            try:
                key = id(widget)
                if is_open:
                    if key not in state_store:
                        state_store[key] = {
                            "widget": widget,
                            "opacity": getattr(widget, "opacity", 1),
                            "disabled": bool(getattr(widget, "disabled", False)) if hasattr(widget, "disabled") else None,
                        }
                    widget.opacity = 0
                    if hasattr(widget, "disabled"):
                        widget.disabled = True
                else:
                    state = state_store.pop(key, None)
                    if state is None:
                        continue
                    widget.opacity = state.get("opacity", 1)
                    if hasattr(widget, "disabled") and state.get("disabled", None) is not None:
                        widget.disabled = bool(state.get("disabled", False))
            except Exception:
                pass

    def _bind_popup_background_softening(self, popup):
        popup_soften_state = {}
        popup.bind(
            on_open=lambda *_: self._set_popup_background_softened(True, popup_soften_state),
            on_dismiss=lambda *_: self._set_popup_background_softened(False, popup_soften_state),
        )
        return popup_soften_state

    def _make_compact_action_button(self, text, tone="plain"):
        palette = getattr(self, "ui_palette", {}) or {}
        if bool(getattr(self, "_is_mobile_ui", getattr(self, "ui_mobile", False))):
            actual_tone = tone if tone in {"primary", "accent", "secondary", "soft", "plain", "ghost", "mini", "danger"} else "plain"
            compact_height = dp(44 if tone not in ("mini", "ghost") else (32 if tone == "mini" else 40))
            compact_font = dp(11.5 if tone not in ("mini", "ghost") else (9.0 if tone == "mini" else 10.5))
            compact_radius = dp(10 if tone not in ("mini", "ghost") else (9 if tone == "mini" else 10))
            compact_padding = [dp(8), dp(4), dp(8), dp(4)] if tone != "mini" else [dp(7), dp(3), dp(7), dp(3)]
            btn = NeoMobileButton(
                text=str(text),
                tone=actual_tone,
                theme=palette,
                compact_mode=True,
                size_hint_y=None,
                height=compact_height,
                font_size=compact_font,
                corner_radius=compact_radius,
                padding=compact_padding,
            )
            return btn
        tone_map = {
            "primary": palette.get("primary", (0.24, 0.48, 0.96, 1)),
            "accent": palette.get("accent", (0.10, 0.78, 0.63, 1)),
            "secondary": palette.get("secondary", palette.get("accent", (0.10, 0.78, 0.63, 1))),
            "soft": palette.get("surface_soft", (0.135, 0.16, 0.215, 1)),
            "plain": palette.get("chip", (0.13, 0.18, 0.27, 1)),
            "ghost": palette.get("ghost", palette.get("chip", (0.13, 0.18, 0.27, 1))),
            "mini": palette.get("mini", palette.get("surface_soft", (0.135, 0.16, 0.215, 1))),
            "danger": palette.get("danger", (0.92, 0.32, 0.42, 1)),
        }
        return Button(
            text=str(text),
            background_normal="",
            background_down="",
            background_color=tone_map.get(tone, tone_map["plain"]),
            color=palette.get("text", (0.93, 0.96, 1.0, 1)),
            font_size=dp(11),
        )

    def _build_collapsible_panel(self, title, content, start_open=True):
        palette = getattr(self, "ui_palette", {}) or {}
        wrap = BoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None)
        header = self._make_compact_action_button(("▾ " if start_open else "▸ ") + str(title), tone="soft")
        header.size_hint_y = None
        header.height = dp(38)
        header.halign = "left"
        header.valign = "middle"
        try:
            header.text_size = header.size
            header.bind(size=self._sync_label_text_size)
        except Exception:
            pass
        body_wrap = BoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(0))
        body_wrap.add_widget(content)
        state = {"open": bool(start_open)}

        def _measure_content():
            h = 0
            try:
                h = max(h, float(getattr(content, "minimum_height", 0)))
            except Exception:
                pass
            try:
                h = max(h, float(content.height))
            except Exception:
                pass
            return max(h, dp(1))

        def _refresh(*_):
            state["open"] = bool(state.get("open", True))
            if state["open"]:
                body_wrap.opacity = 1
                body_wrap.disabled = False
                body_wrap.height = _measure_content()
                wrap.height = header.height + dp(6) + body_wrap.height
                header.text = "▾ " + str(title)
            else:
                body_wrap.opacity = 0
                body_wrap.disabled = True
                body_wrap.height = 0
                wrap.height = header.height
                header.text = "▸ " + str(title)
            wrap._panel_open = bool(state["open"])

        def _toggle(*_):
            state["open"] = not state["open"]
            _refresh()

        def _set_open(value=True):
            state["open"] = bool(value)
            _refresh()

        wrap._set_open = _set_open
        wrap._toggle_panel = _toggle
        header.bind(on_release=_toggle)
        content.bind(height=_refresh)
        if hasattr(content, "bind") and hasattr(content, "minimum_height"):
            content.bind(minimum_height=_refresh)
        wrap.add_widget(header)
        wrap.add_widget(body_wrap)
        Clock.schedule_once(_refresh, 0)
        return wrap

    def _semantic_target_key_from_label(self, label):
        label = str(label or "").strip()
        if not label or label == SEMANTIC_TARGET_PLACEHOLDER:
            return ""
        for key in SEMANTIC_TARGET_KEYS:
            if semantic_target_label(key, key) == label:
                return key
        if label in SEMANTIC_TARGET_KEYS:
            return label
        return ""

    def _get_selected_semantic_target_preview(self, target_key):
        target_key = str(target_key or "").strip()
        record_label = self.selected_patient()
        if not record_label or not target_key:
            return "—"
        try:
            row = self.engine._get_patient_row(record_label, record_key=self.selected_record_key())
        except Exception:
            return "—"

        def _clean(value):
            if pd.isna(value):
                return ""
            return str(value or "").strip()

        if target_key == "record_identity.first_name":
            return _clean(row.get(self.engine.first_col, "")) or "blank"
        if target_key == "record_identity.middle_name":
            return _clean(row.get(self.engine.mid_col, "")) or "blank"
        if target_key == "record_identity.last_name":
            return _clean(row.get(self.engine.last_col, "")) or "blank"
        if target_key == "record_identity.suffix":
            return _clean(row.get(self.engine.suf_col, "")) or "blank"
        if target_key == "record_identity.full_name":
            pieces = [
                _clean(row.get(self.engine.first_col, "")),
                _clean(row.get(self.engine.mid_col, "")),
                _clean(row.get(self.engine.last_col, "")),
                _clean(row.get(self.engine.suf_col, "")),
            ]
            pieces = [p for p in pieces if p]
            return " ".join(pieces) or "blank"
        if target_key in DOB_TARGET_KEYS:
            dob_value = _clean(row.get(self.engine.dob_col, ""))
            parts = [p for p in re.split(r"[^0-9A-Za-z]+", dob_value) if p]
            if len(parts) >= 3:
                mapping = {
                    "record_identity.birth_month": parts[0],
                    "record_identity.birth_day": parts[1],
                    "record_identity.birth_year": parts[2],
                }
                return mapping.get(target_key, dob_value) or "blank"
            return dob_value or "blank"
        if target_key == "record_identity.membership_number":
            return _clean(row.get(self.engine.phil_col, "")) or "blank"
        return "—"

    def _get_selected_column_preview(self, column_name, trigger="", is_grid=False, grid_n=1):
        column_name = str(column_name or "").strip()
        patient = self.selected_patient()
        if not patient or not column_name or column_name == COLUMN_SELECT_TEXT:
            return "—"
        try:
            row = self.engine._get_patient_row(patient)
            raw_val = row.get(column_name, "")
        except Exception:
            return "—"

        if pd.isna(raw_val):
            raw_val = ""
        csv_val = str(raw_val).strip()
        trigger = str(trigger or "").strip()

        if trigger:
            state = "checked" if csv_val.upper() == trigger.upper() else "unchecked"
            return f"{csv_val or 'blank'} → trigger '{trigger}' → {state}"
        if is_grid:
            return f"{csv_val or 'blank'} → grid x{int(grid_n or 1)}"
        return csv_val or "blank"

    def _refresh_mobile_hud_restore_button(self, *_):
        btn = getattr(self, "btn_show_hud", None)
        hud = getattr(self, "preview_hud", None)
        sel_ids = sorted(set(int(x) for x in getattr(self.engine, "selected_box_ids", []) if isinstance(x, int) or str(x).isdigit())) if hasattr(self, "engine") else []
        if btn is not None:
            btn.disabled = True
            btn.opacity = 0
            try:
                btn.size = (0, 0)
                btn.pos = (-10000, -10000)
            except Exception:
                pass
        inspect_btn = getattr(self, "btn_mobile_more", None)
        if inspect_btn is None:
            return
        has_selection = bool(sel_ids)
        try:
            inspect_btn.text = "Repair"
            inspect_btn.tone = "secondary" if has_selection else "ghost"
            inspect_btn.opacity = 1.0 if has_selection else 0.92
            if hasattr(inspect_btn, "_update_graphics"):
                inspect_btn._update_graphics()
            inspect_btn.disabled = False
        except Exception:
            pass

    def _restore_mobile_hud(self, *_):
        hud = getattr(self, "preview_hud", None)
        if hud is None:
            return
        if bool(getattr(hud, "_collapsed", False)):
            try:
                hud.toggle_collapsed()
            except Exception:
                hud._collapsed = False
                hud._refresh_collapse_state()
        self._refresh_mobile_hud_restore_button()

    def _update_preview_hud(self, hit=None):
        sel_ids = sorted(set(int(x) for x in getattr(self.engine, "selected_box_ids", []) if isinstance(x, int) or str(x).isdigit()))
        if hasattr(self, "btn_hud_map") and self.btn_hud_map is not None:
            self.btn_hud_map.disabled = not bool(sel_ids)
        if hasattr(self, "btn_hud_clear") and self.btn_hud_clear is not None:
            self.btn_hud_clear.disabled = not bool(sel_ids)

        hud = getattr(self, "preview_hud", None)
        if getattr(self, "ui_mobile", False) and hud is not None:
            if not sel_ids:
                hud.opacity = 0
                hud.disabled = True
                self._refresh_mobile_hud_restore_button()
                title = "Preview HUD"
                detail = ("Tap a box to show actions." if getattr(self, "ui_mobile", False) else "Select a box to show actions.")
                mapping = "Mapping: —"
                if hasattr(self, "preview_hud_title_lbl") and self.preview_hud_title_lbl is not None:
                    self.preview_hud_title_lbl.text = title
                if hasattr(self, "preview_hud_detail_lbl") and self.preview_hud_detail_lbl is not None:
                    self.preview_hud_detail_lbl.text = detail
                if hasattr(self, "preview_hud_mapping_lbl") and self.preview_hud_mapping_lbl is not None:
                    self.preview_hud_mapping_lbl.text = mapping
                self._refresh_mobile_action_visibility()
                self._refresh_mobile_hud_restore_button()
                return
            hud.opacity = 1
            hud.disabled = False

        if hit:
            box_id = int(hit.get("id", -1))
            rect_txt = f"{int(hit.get('x', 0))},{int(hit.get('y', 0))} • {int(hit.get('w', 0))}×{int(hit.get('h', 0))}"
            map_txt = hit.get("mapping") or "Unmapped"
            verb = "Selected" if getattr(self, "ui_mobile", False) else "Hover"
            title = f"{verb} • Box {box_id}"
            detail = f"Type: {hit.get('t', 'field')} • Rect: {rect_txt}"
            mapping = f"Mapping: {map_txt}"
        elif sel_ids:
            first = sel_ids[0]
            box_type = self.engine.box_types[first] if first < len(self.engine.box_types) else "field"
            rect = self.engine.all_boxes[first] if first < len(self.engine.all_boxes) else None
            rect_txt = "—"
            if rect is not None:
                rect_txt = f"{int(rect.x0)},{int(rect.y0)} → {int(rect.x1)},{int(rect.y1)}"
            map_txt = self.engine.describe_box_mapping(first, self.current_page_idx())
            title = f"Selected • {len(sel_ids)} box{'es' if len(sel_ids) != 1 else ''}"
            detail = f"Focus: {first} • {box_type} • {rect_txt}"
            mapping = f"Mapping: {map_txt if map_txt and map_txt != 'EMPTY' else 'Unmapped'}"
        else:
            title = "Preview HUD"
            detail = ("Tap or drag over a box to see its details here." if getattr(self, "ui_mobile", False) else "Hover or tap a box to see its details here.")
            mapping = "Mapping: —"

        if hasattr(self, "preview_hud_title_lbl") and self.preview_hud_title_lbl is not None:
            self.preview_hud_title_lbl.text = title
        if hasattr(self, "preview_hud_detail_lbl") and self.preview_hud_detail_lbl is not None:
            self.preview_hud_detail_lbl.text = detail
        if hasattr(self, "preview_hud_mapping_lbl") and self.preview_hud_mapping_lbl is not None:
            self.preview_hud_mapping_lbl.text = mapping
        self._refresh_mobile_action_visibility()
        self._refresh_mobile_hud_restore_button()

    def _open_mapping_from_hud(self, *_):
        ids = sorted(set(int(x) for x in getattr(self.engine, "selected_box_ids", []) if isinstance(x, int) or str(x).isdigit()))
        if not ids:
            self.set_status("Select a preview box before opening the mapping HUD action.")
            return
        self.open_mapping_popup_for_selection(primary_box_id=ids[0])

    def on_preview_zoom_request(self, action, value=None):
        if action == "step":
            self._zoom_preview("in" if str(value).lower() == "in" else "out")
        elif action == "scale":
            self._apply_preview_zoom(mode="manual", scale=float(value or 1.0))
        elif action == "gesture":
            payload = value or {}
            focus = payload.get("focus")
            scale = float(payload.get("scale", getattr(self, "preview_zoom_factor", 1.0) or 1.0))
            pan = payload.get("pan", (0.0, 0.0))
            self._apply_preview_zoom(mode="manual", scale=scale, anchor_window_point=focus, schedule_resync=False)
            try:
                pan_dx = float(pan[0])
                pan_dy = float(pan[1])
            except Exception:
                pan_dx = 0.0
                pan_dy = 0.0
            if abs(pan_dx) > 0.01 or abs(pan_dy) > 0.01:
                Clock.schedule_once(lambda dt, dx=pan_dx, dy=pan_dy: self._scroll_preview_by_pixels(dx, dy), 0)
        elif action == "gesture_end":
            self._schedule_preview_resync()

    def _normalized_text_tuning_ui(self, tuning=None):
        src = tuning if tuning is not None else getattr(self, "text_tuning_ui", None)
        return self.engine._normalized_text_tuning(src if isinstance(src, dict) else None)

    def _update_text_tuning_summary(self):
        tuning = self._normalized_text_tuning_ui()
        self.text_tuning_ui = dict(tuning)
        summary = (
            f"FIT INSIDE • CENTERED • {'UPPERCASE' if tuning.get('uppercase', True) else 'ORIGINAL CASE'}\n"
            f"Overall {tuning['overall_scale']:.2f} • Box {tuning['box_scale']:.2f} • Line {tuning['line_scale']:.2f} • Grid {tuning['grid_scale']:.2f}\n"
            f"Inset X {tuning['content_inset_x']:.2f} • Inset Y {tuning['content_inset_y']:.2f} • Line height {tuning['line_expand_h_mul']:.2f}x / min {int(tuning['line_min_effective_h_px'])} px"
        )
        lbl = getattr(self, "text_tuning_summary_lbl", None)
        if lbl is not None:
            lbl.text = summary

    def open_text_tuning_popup(self, *_):
        palette = getattr(self, "ui_palette", {}) or {}
        base = self._normalized_text_tuning_ui()
        outer = BoxLayout(orientation="vertical", spacing=dp(10), padding=[dp(14), dp(14), dp(14), dp(14)])
        self._style_popup_card(outer, palette.get("surface", (0.085, 0.105, 0.145, 1)), radius=dp(22))

        head = BoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None)
        head.bind(minimum_height=head.setter("height"))
        title_lbl = Label(text="Text Fit & Font Console", color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(28), halign="left", valign="middle", font_size=dp(18), bold=True)
        title_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(title_lbl, min_height=dp(28), extra_pad=dp(4))
        sub_lbl = Label(text="Adjust font fitting, centering, and line height. Apply now refreshes the current preview and also affects export.", color=palette.get("muted", (0.60, 0.68, 0.80, 1)), size_hint_y=None, height=dp(20), halign="left", valign="middle", font_size=dp(11))
        sub_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(sub_lbl, min_height=dp(20), extra_pad=dp(6))
        head.add_widget(title_lbl)
        head.add_widget(sub_lbl)
        outer.add_widget(head)

        scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, bar_width=dp(6), bar_margin=dp(10))
        content = GridLayout(cols=1, spacing=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        scroll.add_widget(content)

        def _slider_spec(key, label, min_v, max_v, step=0.01, decimals=2, helper="", widget_type="float"):
            return {"key": key, "label": label, "min": min_v, "max": max_v, "step": step, "decimals": decimals, "helper": helper, "widget_type": widget_type}

        groups = [
            ("Scale", [
                _slider_spec("overall_scale", "Overall Size", 0.60, 1.50, 0.01, 2, "Master size control used by preview and export."),
                _slider_spec("box_scale", "Box Size", 0.60, 1.60, 0.01, 2, "How large text appears inside normal green/box fields."),
                _slider_spec("line_scale", "Yellow Line Size", 0.60, 1.60, 0.01, 2, "How large text appears inside thin yellow answer lines."),
                _slider_spec("grid_scale", "Grid Cell Size", 0.60, 1.40, 0.01, 2, "How large one character appears inside split/grid cells."),
            ]),
            ("Centering", [
                _slider_spec("content_inset_x", "Horizontal Safe Padding", 0.00, 0.22, 0.01, 2, "Raises left/right padding so text stays inside the box."),
                _slider_spec("content_inset_y", "Vertical Safe Padding", 0.00, 0.28, 0.01, 2, "Raises top/bottom padding so text stays clear of the box edges."),
                _slider_spec("box_vshift", "Box Vertical Nudge", -0.30, 0.20, 0.01, 2, "Negative values move box text upward."),
                _slider_spec("line_vshift", "Yellow Line Vertical Nudge", -0.30, 0.20, 0.01, 2, "Negative values move yellow line text upward."),
                _slider_spec("grid_vshift", "Grid Vertical Nudge", -0.30, 0.20, 0.01, 2, "Negative values move grid-cell text upward."),
            ]),
            ("Yellow Line Readability", [
                _slider_spec("line_expand_h_mul", "Yellow Height Boost", 1.00, 3.50, 0.05, 2, "Creates a virtual taller area so yellow line text can stay readable."),
                _slider_spec("line_min_effective_h_px", "Yellow Minimum Height", 0, 80, 1, 0, "Minimum effective height in preview pixels for yellow line text.", "int"),
            ]),
        ]

        controls = {}

        def _format_slider_value(val, decimals=0, as_int=False):
            return str(int(round(float(val)))) if as_int else f"{float(val):.{int(decimals)}f}"

        def _make_slider_control(spec):
            is_mobile = bool(getattr(self, "ui_mobile", False))
            start_val = float(base.get(spec["key"], spec["min"]))
            row = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
            row.bind(minimum_height=row.setter("height"))
            top = BoxLayout(orientation="horizontal", spacing=dp(8), size_hint_y=None, height=dp(18 if is_mobile else 20))
            lbl = Label(text=spec["label"], color=palette.get("text", (0.93, 0.96, 1.0, 1)), halign="left", valign="middle", font_size=dp(11))
            lbl.bind(size=self._sync_label_text_size)
            top.add_widget(lbl)
            val_lbl = None
            if not is_mobile:
                val_lbl = Label(text=_format_slider_value(start_val, spec["decimals"], spec["widget_type"] == "int"), color=palette.get("accent", (0.10, 0.78, 0.63, 1)), halign="right", valign="middle", font_size=dp(11))
                val_lbl.bind(size=self._sync_label_text_size)
                top.add_widget(val_lbl)
            slider = Slider(min=spec["min"], max=spec["max"], value=start_val, step=spec["step"], size_hint_y=None, height=dp(28 if is_mobile else 30))
            input_widget = TextInput(
                text=_format_slider_value(start_val, spec["decimals"], spec["widget_type"] == "int"),
                multiline=False,
                size_hint_y=None,
                height=dp(28 if is_mobile else 28),
                input_filter=("int" if spec["widget_type"] == "int" else None),
                background_normal="",
                background_active="",
                background_color=palette.get("surface_soft", (0.135, 0.16, 0.215, 1)),
                foreground_color=palette.get("text", (0.93, 0.96, 1.0, 1)),
                cursor_color=palette.get("accent", (0.10, 0.78, 0.63, 1)),
                padding=[dp(8), dp(6), dp(8), dp(6)],
            )

            def _sync_from_slider(_instance, value):
                formatted = _format_slider_value(value, spec["decimals"], spec["widget_type"] == "int")
                if val_lbl is not None:
                    val_lbl.text = formatted
                if input_widget.text != formatted:
                    input_widget.text = formatted

            def _sync_from_input(_instance=None):
                try:
                    value = float(input_widget.text.strip())
                    value = max(float(spec["min"]), min(float(spec["max"]), value))
                    slider.value = value
                except Exception:
                    pass

            slider.bind(value=_sync_from_slider)
            input_widget.bind(on_text_validate=_sync_from_input, focus=lambda inst, focused: (None if focused else _sync_from_input(inst)))
            row.add_widget(top)
            if is_mobile:
                control_row = BoxLayout(orientation="horizontal", spacing=dp(8), size_hint_y=None, height=dp(34))
                slider.size_hint_x = 1
                input_widget.size_hint_x = None
                input_widget.width = dp(72)
                control_row.add_widget(slider)
                control_row.add_widget(input_widget)
                row.add_widget(control_row)
            else:
                row.add_widget(slider)
                row.add_widget(input_widget)
            if spec.get("helper"):
                helper = Label(text=spec["helper"], color=palette.get("muted", (0.60, 0.68, 0.80, 0.92)), size_hint_y=None, height=dp(18), halign="left", valign="middle", font_size=dp(11))
                helper.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(helper, min_height=dp(14), extra_pad=dp(2))
                row.add_widget(helper)
            controls[spec["key"]] = {"slider": slider, "input": input_widget, "spec": spec}
            return row

        for group_name, specs in groups:
            card = GridLayout(cols=1, spacing=dp(8), size_hint_y=None, padding=[dp(12), dp(10), dp(12), dp(12)])
            card.bind(minimum_height=card.setter("height"))
            title = Label(text=group_name, color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(18 if getattr(self, "ui_mobile", False) else 20), halign="left", valign="middle", font_size=dp(12 if getattr(self, "ui_mobile", False) else 13), bold=True)
            title.bind(size=self._sync_label_text_size)
            grid = GridLayout(cols=1 if getattr(self, "ui_mobile", False) else 2, spacing=dp(10), size_hint_y=None)
            grid.bind(minimum_height=grid.setter("height"))
            for spec in specs:
                grid.add_widget(_make_slider_control(spec))
            card.add_widget(title)
            card.add_widget(grid)
            self._style_popup_card(card, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(18))
            content.add_widget(card)

        uppercase_toggle = CheckBox(active=bool(base.get("uppercase", True)))
        toggle_row = BoxLayout(orientation="horizontal", spacing=dp(8), size_hint_y=None, height=dp(44))
        toggle_lbl = Label(text="Force uppercase", color=palette.get("text", (0.93, 0.96, 1.0, 1)), halign="left", valign="middle", font_size=dp(11))
        toggle_lbl.bind(size=self._sync_label_text_size)
        toggle_row.add_widget(toggle_lbl)
        toggle_slot = AnchorLayout(anchor_x="right", anchor_y="center", size_hint=(None, None), width=dp(56), height=dp(44), padding=(0, 0, dp(10), 0))
        uppercase_toggle.size_hint = (None, None)
        uppercase_toggle.size = (dp(32), dp(24))
        toggle_slot.add_widget(uppercase_toggle)
        toggle_row.add_widget(toggle_slot)
        toggle_wrap = BoxLayout(orientation="vertical", spacing=dp(3), padding=[dp(10), dp(8), dp(10), dp(8)], size_hint_y=None)
        toggle_wrap.bind(minimum_height=toggle_wrap.setter("height"))
        self._style_popup_card(toggle_wrap, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(16))
        toggle_wrap.add_widget(toggle_row)
        toggle_help = Label(text="Keep this on if you want the whole form to stay visually uniform.", color=palette.get("muted", (0.60, 0.68, 0.80, 0.96)), halign="left", valign="middle", font_size=dp(11), size_hint_y=None)
        toggle_help.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(toggle_help, min_height=dp(18), extra_pad=dp(4))
        toggle_wrap.add_widget(toggle_help)
        content.add_widget(toggle_wrap)

        outer.add_widget(scroll)

        if getattr(self, "ui_mobile", False):
            btn_row = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
            btn_row.bind(minimum_height=btn_row.setter("height"))
            top_actions = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(44))
            btn_defaults = self._make_compact_action_button("Reset", tone="ghost")
            btn_apply = self._make_compact_action_button("Apply", tone="primary")
            for btn in [btn_defaults, btn_apply]:
                top_actions.add_widget(btn)
            btn_close = self._make_compact_action_button("Close", tone="ghost")
            btn_close.size_hint_y = None
            btn_close.height = dp(42)
            btn_row.add_widget(top_actions)
            btn_row.add_widget(btn_close)
        else:
            btn_row = GridLayout(cols=3, spacing=dp(8), size_hint_y=None, height=dp(46))
            btn_defaults = self._make_compact_action_button("Defaults", tone="ghost")
            btn_apply = self._make_compact_action_button("Apply", tone="primary")
            btn_close = self._make_compact_action_button("Close", tone="plain")
            for btn in [btn_defaults, btn_apply, btn_close]:
                btn_row.add_widget(btn)
        outer.add_widget(btn_row)

        popup = Popup(
            title="",
            separator_height=0,
            background="",
            background_color=(0, 0, 0, 0.80),
            content=outer,
            size_hint=(0.96 if getattr(self, "ui_mobile", False) else 0.86, 0.92 if getattr(self, "ui_mobile", False) else 0.88),
            auto_dismiss=True,
        )
        self._bind_popup_background_softening(popup)

        def _apply_text_settings(*_):
            try:
                updated = dict(base)
                for key, payload in controls.items():
                    slider = payload["slider"]
                    spec = payload["spec"]
                    if spec["widget_type"] == "int":
                        updated[key] = int(round(float(slider.value)))
                    else:
                        updated[key] = float(slider.value)
                updated["uppercase"] = bool(uppercase_toggle.active)
                self.text_tuning_ui = self._normalized_text_tuning_ui(updated)
                self.apply_ui_settings_to_engine()
                self._update_text_tuning_summary()
                popup.dismiss()
                if not self.engine.pdf_path:
                    self.set_status("Text tuning applied. Load a PDF to see the updated fit and centering.", kind="action", hold_seconds=2.0, force=True)
                    return
                self.set_status("Text tuning applied. Refreshing preview with the new fit and centering rules...", kind="action", hold_seconds=2.5, force=True)
                Clock.schedule_once(lambda dt: self.on_preview(None), 0)
            except Exception as e:
                traceback.print_exc()
                self.set_status(f"Text tuning error:\n{e}", kind="error", force=True)

        def _restore_text_defaults(*_):
            defaults = self._normalized_text_tuning_ui(TEXT_TUNING_DEFAULTS)
            for key, payload in controls.items():
                if key in defaults:
                    payload["slider"].value = float(defaults[key])
            uppercase_toggle.active = bool(defaults.get("uppercase", True))

        btn_defaults.bind(on_release=_restore_text_defaults)
        btn_apply.bind(on_release=_apply_text_settings)
        btn_close.bind(on_release=lambda *_: popup.dismiss())
        popup.open()

    def open_detection_settings_popup(self, *_):
        palette = getattr(self, "ui_palette", {}) or {}
        outer = BoxLayout(orientation="vertical", spacing=dp(10), padding=[dp(14), dp(14), dp(14), dp(14)])
        self._style_popup_card(outer, palette.get("surface", (0.085, 0.105, 0.145, 1)), radius=dp(22))

        head = BoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None)
        head.bind(minimum_height=head.setter("height"))
        title_lbl = Label(text="Detection Tuning Console", color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(28), halign="left", valign="middle", font_size=dp(18), bold=True)
        title_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(title_lbl, min_height=dp(28), extra_pad=dp(4))
        sub_lbl = Label(text="Adjust how the app finds fill areas. Apply now reruns detection and refreshes the page immediately.", color=palette.get("muted", (0.60, 0.68, 0.80, 1)), size_hint_y=None, height=dp(20), halign="left", valign="middle", font_size=dp(11))
        sub_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(sub_lbl, min_height=dp(20), extra_pad=dp(6))
        head.add_widget(title_lbl)
        head.add_widget(sub_lbl)
        outer.add_widget(head)

        scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, bar_width=dp(6), bar_margin=dp(10))
        content = GridLayout(cols=1, spacing=dp(10), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        scroll.add_widget(content)

        def _popup_input_block(widget, helper_text=""):
            block = BoxLayout(orientation="vertical", spacing=dp(3), size_hint_y=None)
            block.bind(minimum_height=block.setter("height"))
            block.add_widget(widget)
            if helper_text:
                helper = Label(text=helper_text, color=palette.get("muted", (0.60, 0.68, 0.80, 0.92)), size_hint_y=None, height=(dp(18) if getattr(self, "ui_mobile", False) else dp(16)), halign="left", valign="middle", font_size=(dp(11) if getattr(self, "ui_mobile", False) else dp(11)))
                helper.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(helper, min_height=(dp(14) if getattr(self, "ui_mobile", False) else dp(14)), extra_pad=dp(2))
                block.add_widget(helper)
            return block

        def _slider_spec(key, label, min_v, max_v, step=1.0, decimals=0, helper="", widget_type="float"):
            return {"key": key, "label": label, "min": min_v, "max": max_v, "step": step, "decimals": decimals, "helper": helper, "widget_type": widget_type}

        groups = [
            ("Field", [
                _slider_spec("f_area", detection_ui_label("f_area", "Field Area"), 100, 5000, 10, 0, detection_ui_helper("f_area", "Minimum contour area kept as a field"), "int"),
                _slider_spec("f_minw", detection_ui_label("f_minw", "Field Min Width"), 5, 200, 1, 0, detection_ui_helper("f_minw"), "int"),
                _slider_spec("f_minh", detection_ui_label("f_minh", "Field Min Height"), 5, 120, 1, 0, detection_ui_helper("f_minh"), "int"),
                _slider_spec("f_close", detection_ui_label("f_close", "Field Close"), 0, 10, 1, 0, detection_ui_helper("f_close", "Morph close kernel for field cleanup"), "int"),
            ]),
            ("Lines", [
                _slider_spec("line_minw", detection_ui_label("line_minw", "Line Min Width"), 50, 800, 5, 0, detection_ui_helper("line_minw"), "int"),
                _slider_spec("line_maxw", detection_ui_label("line_maxw", "Line Max Width"), 150, 2600, 10, 0, detection_ui_helper("line_maxw"), "int"),
            ]),
            ("Checkbox", [
                _slider_spec("c_strict", detection_ui_label("c_strict", "Checkbox Strict"), 10, 100, 1, 0, detection_ui_helper("c_strict"), "int"),
                _slider_spec("c_size_min", detection_ui_label("c_size_min", "Checkbox Min Size"), 5, 120, 1, 0, detection_ui_helper("c_size_min"), "int"),
                _slider_spec("c_size_max", detection_ui_label("c_size_max", "Checkbox Max Size"), 10, 160, 1, 0, detection_ui_helper("c_size_max"), "int"),
                _slider_spec("c_border", detection_ui_label("c_border", "Border Fill"), 0.0, 1.0, 0.01, 2, detection_ui_helper("c_border"), "float"),
                _slider_spec("c_inner", detection_ui_label("c_inner", "Inner Fill"), 0.0, 1.0, 0.01, 2, detection_ui_helper("c_inner"), "float"),
                _slider_spec("roi_max", detection_ui_label("roi_max", "ROI Max"), 50, 1000, 5, 0, detection_ui_helper("roi_max"), "int"),
                _slider_spec("c_open", detection_ui_label("c_open", "Checkbox Open"), 0, 5, 1, 0, detection_ui_helper("c_open"), "int"),
                _slider_spec("c_close", detection_ui_label("c_close", "Checkbox Close"), 0, 5, 1, 0, detection_ui_helper("c_close"), "int"),
                _slider_spec("c_band", detection_ui_label("c_band", "Band %"), 0.01, 0.4, 0.01, 2, detection_ui_helper("c_band"), "float"),
                _slider_spec("c_aspect", detection_ui_label("c_aspect", "Aspect Tolerance"), 0.05, 1.0, 0.01, 2, detection_ui_helper("c_aspect"), "float"),
            ]),
            ("Extent", [
                _slider_spec("ext_low", detection_ui_label("ext_low", "Extent Low"), 0.01, 0.5, 0.01, 2, detection_ui_helper("ext_low"), "float"),
                _slider_spec("ext_high", detection_ui_label("ext_high", "Extent High"), 0.10, 1.0, 0.01, 2, detection_ui_helper("ext_high"), "float"),
                _slider_spec("c_fill", detection_ui_label("c_fill", "Fill Min"), 0.10, 1.0, 0.01, 2, detection_ui_helper("c_fill"), "float"),
                _slider_spec("c_eps", detection_ui_label("c_eps", "Approx Eps"), 0.01, 0.20, 0.01, 2, detection_ui_helper("c_eps"), "float"),
            ]),
        ]

        controls = {}

        def _format_slider_value(val, decimals=0, as_int=False):
            return str(int(round(float(val)))) if as_int else f"{float(val):.{int(decimals)}f}"

        def _make_slider_control(spec):
            src_widget = getattr(self, spec["key"])
            is_mobile = bool(getattr(self, "ui_mobile", False))
            try:
                start_val = float(src_widget.text)
            except Exception:
                start_val = float(spec["min"])
            row = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
            row.bind(minimum_height=row.setter("height"))
            top = BoxLayout(orientation="horizontal", spacing=dp(8), size_hint_y=None, height=dp(18 if is_mobile else 20))
            lbl = Label(text=spec["label"], color=palette.get("text", (0.93, 0.96, 1.0, 1)), halign="left", valign="middle", font_size=dp(11))
            lbl.bind(size=self._sync_label_text_size)
            top.add_widget(lbl)
            val_lbl = None
            if not is_mobile:
                val_lbl = Label(text=_format_slider_value(start_val, spec["decimals"], spec["widget_type"] == "int"), color=palette.get("accent", (0.10, 0.78, 0.63, 1)), halign="right", valign="middle", font_size=dp(11))
                val_lbl.bind(size=self._sync_label_text_size)
                top.add_widget(val_lbl)
            slider = Slider(min=spec["min"], max=spec["max"], value=start_val, step=spec["step"], size_hint_y=None, height=dp(28 if is_mobile else 30))
            input_widget = TextInput(
                text=_format_slider_value(start_val, spec["decimals"], spec["widget_type"] == "int"),
                multiline=False,
                size_hint_y=None,
                height=dp(28 if is_mobile else 28),
                input_filter=("int" if spec["widget_type"] == "int" else None),
                background_normal="",
                background_active="",
                background_color=palette.get("surface_soft", (0.135, 0.16, 0.215, 1)),
                foreground_color=palette.get("text", (0.93, 0.96, 1.0, 1)),
                cursor_color=palette.get("accent", (0.10, 0.78, 0.63, 1)),
                padding=[dp(8), dp(6), dp(8), dp(6)],
            )

            def _sync_from_slider(_instance, value):
                formatted = _format_slider_value(value, spec["decimals"], spec["widget_type"] == "int")
                if val_lbl is not None:
                    val_lbl.text = formatted
                if input_widget.text != formatted:
                    input_widget.text = formatted

            def _sync_from_input(_instance=None):
                try:
                    value = float(input_widget.text.strip())
                    value = max(float(spec["min"]), min(float(spec["max"]), value))
                    slider.value = value
                except Exception:
                    pass

            slider.bind(value=_sync_from_slider)
            input_widget.bind(on_text_validate=_sync_from_input, focus=lambda inst, focused: (None if focused else _sync_from_input(inst)))
            row.add_widget(top)
            if is_mobile:
                control_row = BoxLayout(orientation="horizontal", spacing=dp(8), size_hint_y=None, height=dp(34))
                slider.size_hint_x = 1
                input_widget.size_hint_x = None
                input_widget.width = dp(72)
                input_widget.height = dp(32)
                control_row.add_widget(slider)
                control_row.add_widget(input_widget)
                row.add_widget(control_row)
            else:
                row.add_widget(slider)
                row.add_widget(input_widget)
            controls[spec["key"]] = {"slider": slider, "input": input_widget, "spec": spec}
            return _popup_input_block(row, spec.get("helper", ""))

        for group_name, specs in groups:
            card = BoxLayout(orientation="vertical", spacing=dp(6 if getattr(self, "ui_mobile", False) else 8), padding=[dp(8), dp(8), dp(8), dp(8)] if getattr(self, "ui_mobile", False) else [dp(10), dp(10), dp(10), dp(10)], size_hint_y=None)
            card.bind(minimum_height=card.setter("height"))
            title = Label(text=DETECTION_GROUP_LABELS.get(group_name, group_name), color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(18 if getattr(self, "ui_mobile", False) else 20), halign="left", valign="middle", font_size=dp(12 if getattr(self, "ui_mobile", False) else 13), bold=True)
            title.bind(size=self._sync_label_text_size)
            grid = GridLayout(cols=1 if getattr(self, "ui_mobile", False) else 3, spacing=dp(10), size_hint_y=None)
            grid.bind(minimum_height=grid.setter("height"))
            for spec in specs:
                grid.add_widget(_make_slider_control(spec))
            card.add_widget(title)
            card.add_widget(grid)
            self._style_popup_card(card, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(18))
            content.add_widget(card)

        extent_toggle = CheckBox(active=bool(getattr(self.use_extent_chk, "active", False)))
        extent_row_h = dp(42)
        extent_row = BoxLayout(orientation="horizontal", spacing=dp(8), size_hint_y=None, height=extent_row_h)
        extent_lbl = Label(text=detection_ui_label("use_extent", "Use Extent"), color=palette.get("text", (0.93, 0.96, 1.0, 1)), halign="left", valign="middle", font_size=dp(11))
        extent_lbl.bind(size=self._sync_label_text_size)
        extent_row.add_widget(extent_lbl)
        extent_slot = AnchorLayout(anchor_x="right", anchor_y="center", size_hint=(None, None), width=dp(56), height=extent_row_h, padding=(0, 0, dp(10), 0))
        extent_toggle.size_hint = (None, None)
        extent_toggle.size = (dp(24), dp(18))
        extent_slot.add_widget(extent_toggle)
        extent_row.add_widget(extent_slot)
        extent_wrap = BoxLayout(orientation="vertical", spacing=dp(3), padding=[dp(10), dp(8), dp(10), dp(8)], size_hint_y=None)
        extent_wrap.bind(minimum_height=extent_wrap.setter("height"))
        self._style_popup_card(extent_wrap, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(16))
        extent_wrap.add_widget(extent_row)
        extent_help = Label(text=detection_ui_helper("use_extent", "Use contour extent as an extra checkbox filter"), color=palette.get("muted", (0.60, 0.68, 0.80, 0.96)), halign="left", valign="middle", font_size=dp(9.5), size_hint_y=None)
        extent_help.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(extent_help, min_height=dp(18), extra_pad=dp(4))
        extent_wrap.add_widget(extent_help)
        content.add_widget(extent_wrap)

        outer.add_widget(scroll)

        preset_wrap = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None, height=dp(56 if getattr(self, "ui_mobile", False) else 76))
        preset_lbl = Label(text=("" if getattr(self, "ui_mobile", False) else "Quick presets"), color=palette.get("muted", (0.60, 0.68, 0.80, 0.96)), size_hint_y=None, height=dp(0 if getattr(self, "ui_mobile", False) else 15), halign="left", valign="middle", font_size=dp(10.0))
        preset_lbl.bind(size=self._sync_label_text_size)
        preset_row = GridLayout(cols=3, spacing=dp(6), size_hint_y=None, height=dp(38 if getattr(self, "ui_mobile", False) else 44))
        btn_preset_balanced = self._make_compact_action_button("Balanced", tone="mini")
        btn_preset_sensitive = self._make_compact_action_button("Sensitive", tone="mini")
        btn_preset_strict = self._make_compact_action_button("Strict", tone="mini")
        for btn in [btn_preset_balanced, btn_preset_sensitive, btn_preset_strict]:
            preset_row.add_widget(btn)
        preset_wrap.add_widget(preset_lbl)
        preset_wrap.add_widget(preset_row)
        outer.add_widget(preset_wrap)

        if getattr(self, "ui_mobile", False):
            btn_row = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
            btn_row.bind(minimum_height=btn_row.setter("height"))
            top_actions = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(44))
            btn_defaults = self._make_compact_action_button("Reset", tone="ghost")
            btn_apply = self._make_compact_action_button("Apply", tone="primary")
            for btn in [btn_defaults, btn_apply]:
                top_actions.add_widget(btn)
            btn_close = self._make_compact_action_button("Close", tone="ghost")
            btn_close.size_hint_y = None
            btn_close.height = dp(42)
            btn_row.add_widget(top_actions)
            btn_row.add_widget(btn_close)
        else:
            btn_row = GridLayout(cols=3, spacing=dp(8), size_hint_y=None, height=dp(46))
            btn_defaults = self._make_compact_action_button("Defaults", tone="ghost")
            btn_apply = self._make_compact_action_button("Apply", tone="primary")
            btn_close = self._make_compact_action_button("Close", tone="plain")
            for btn in [btn_defaults, btn_apply, btn_close]:
                btn_row.add_widget(btn)
        outer.add_widget(btn_row)

        popup = Popup(
            title="",
            separator_height=0,
            background="",
            background_color=(0, 0, 0, 0.80),
            content=outer,
            size_hint=(0.96 if getattr(self, "ui_mobile", False) else 0.86, 0.92 if getattr(self, "ui_mobile", False) else 0.88),
            auto_dismiss=True,
        )

        self._bind_popup_background_softening(popup)

        def _apply_settings(*_):
            try:
                for key, payload in controls.items():
                    slider = payload["slider"]
                    spec = payload["spec"]
                    as_int = spec["widget_type"] == "int"
                    new_text = _format_slider_value(slider.value, spec["decimals"], as_int)
                    getattr(self, key).text = new_text
                self.use_extent_chk.active = bool(extent_toggle.active)
                self.apply_ui_settings_to_engine()
                popup.dismiss()
                if not self.engine.pdf_path:
                    self.set_status("Detection tuning applied. Load a PDF to see the updated field finding.", kind="action", hold_seconds=2.0, force=True)
                    return
                page_idx = self.current_page_idx()
                self.engine.invalidate_detection_cache(page_idx=page_idx, clear_current=True)
                self.set_status("Detection tuning applied. Refreshing field finding on the current page...", kind="action", hold_seconds=2.5, force=True)
                Clock.schedule_once(lambda dt: self.on_run_detect(None), 0)
            except Exception as e:
                self.set_status(f"Detection tuning error:\n{e}", kind="error", force=True)

        def _restore_defaults(*_):
            default_map = {
                "f_area": DEFAULTS["F_Area"],
                "f_minw": DEFAULTS["F_MinW"],
                "f_minh": DEFAULTS["F_MinH"],
                "f_close": DEFAULTS["F_Close"],
                "line_minw": DEFAULTS["Line_MinW"],
                "line_maxw": DEFAULTS["Line_MaxW"],
                "c_strict": DEFAULTS["C_Strict"],
                "c_size_min": DEFAULTS["C_Size"][0],
                "c_size_max": DEFAULTS["C_Size"][1],
                "c_border": DEFAULTS["C_Border"],
                "c_inner": DEFAULTS["C_Inner"],
                "roi_max": DEFAULTS["ROI_Max"],
                "c_open": DEFAULTS["C_Open"],
                "c_close": DEFAULTS["C_Close"],
                "c_band": DEFAULTS["C_BandPct"],
                "c_aspect": DEFAULTS["C_AspectTol"],
                "ext_low": DEFAULTS["Ext_Low"],
                "ext_high": DEFAULTS["Ext_High"],
                "c_fill": DEFAULTS["C_FillMin"],
                "c_eps": DEFAULTS["C_Eps"],
            }
            for key, val in default_map.items():
                if key in controls:
                    controls[key]["slider"].value = float(val)
            extent_toggle.active = bool(DEFAULTS["Use_Extent"])

        def _apply_preset(kind):
            preset = {
                "balanced": {
                    "f_area": DEFAULTS["F_Area"], "f_close": DEFAULTS["F_Close"],
                    "c_strict": DEFAULTS["C_Strict"], "c_open": DEFAULTS["C_Open"], "c_close": DEFAULTS["C_Close"],
                    "ext_low": DEFAULTS["Ext_Low"], "ext_high": DEFAULTS["Ext_High"], "c_fill": DEFAULTS["C_FillMin"]
                },
                "sensitive": {
                    "f_area": max(50, DEFAULTS["F_Area"] * 0.75), "f_close": max(0, DEFAULTS["F_Close"] - 1),
                    "c_strict": max(8, DEFAULTS["C_Strict"] - 8), "c_open": 0, "c_close": max(0, DEFAULTS["C_Close"] - 1),
                    "ext_low": max(0.01, DEFAULTS["Ext_Low"] * 0.65), "ext_high": min(1.0, DEFAULTS["Ext_High"] + 0.08), "c_fill": max(0.08, DEFAULTS["C_FillMin"] - 0.05)
                },
                "strict": {
                    "f_area": DEFAULTS["F_Area"] * 1.25, "f_close": DEFAULTS["F_Close"] + 1,
                    "c_strict": DEFAULTS["C_Strict"] + 10, "c_open": DEFAULTS["C_Open"] + 1, "c_close": DEFAULTS["C_Close"] + 1,
                    "ext_low": min(0.5, DEFAULTS["Ext_Low"] + 0.04), "ext_high": max(DEFAULTS["Ext_Low"] + 0.12, DEFAULTS["Ext_High"] - 0.04), "c_fill": min(1.0, DEFAULTS["C_FillMin"] + 0.05)
                },
            }.get(str(kind).lower(), {})
            for key, val in preset.items():
                if key in controls:
                    controls[key]["slider"].value = float(val)
            self.set_status(f"Detection preset applied: {str(kind).title()}")

        btn_defaults.bind(on_release=_restore_defaults)
        btn_preset_balanced.bind(on_release=lambda *_: _apply_preset("balanced"))
        btn_preset_sensitive.bind(on_release=lambda *_: _apply_preset("sensitive"))
        btn_preset_strict.bind(on_release=lambda *_: _apply_preset("strict"))
        btn_apply.bind(on_release=_apply_settings)
        btn_close.bind(on_release=lambda *_: popup.dismiss())
        popup.open()

    def open_mapping_popup_for_selection(self, primary_box_id=None):
        ids = sorted(set(int(x) for x in self.engine.selected_box_ids if isinstance(x, int) or str(x).isdigit()))
        if not ids:
            self.set_status("Tap at least one preview box first.")
            return

        page_idx = self.current_page_idx()
        focus_id = ids[0] if primary_box_id is None else int(primary_box_id)
        existing = self._find_mapping_for_box(focus_id, page_idx)
        palette = getattr(self, "ui_palette", {}) or {}

        current_col = self.column_spinner.text if self.column_spinner.text != COLUMN_SELECT_TEXT else ""
        current_trigger = self.trigger_input.text.strip()
        current_grid_flag = "1" if getattr(self, "grid_flag_chk", None).active else "0" if hasattr(self, "grid_flag_chk") else (self.grid_flag_input.text.strip() or "0")
        current_grid_n = self.grid_n_input.text.strip() or "1"

        if existing:
            current_col = str(existing.get("column", "")).strip() or current_col
            current_trigger = str(existing.get("trigger", "")).strip()
            current_grid_flag = "1" if bool(existing.get("g", False)) else "0"
            current_grid_n = str(int(existing.get("n", 1)))

        outer = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(10))
        self._style_popup_card(outer, palette.get("surface", (0.085, 0.105, 0.145, 1)), radius=dp(22))

        head = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
        head.bind(minimum_height=head.setter("height"))
        title_lbl = Label(
            text=("Link Selected Boxes" if getattr(self, "ui_mobile", False) else "Preview Link Inspector"),
            color=palette.get("text", (0.93, 0.96, 1.0, 1)),
            size_hint_y=None,
            height=dp(28),
            halign="left",
            valign="middle",
            font_size=dp(18),
            bold=True,
        )
        title_lbl.bind(size=self._sync_label_text_size)
        if getattr(self, "ui_mobile", False):
            subtitle_text = f"Box {focus_id} • {len(ids)} selected"
        else:
            subtitle_text = f"Page {page_idx} • Focus box {focus_id} • {len(ids)} selected"
        subtitle = Label(
            text=subtitle_text,
            color=palette.get("muted", (0.60, 0.68, 0.80, 1)),
            size_hint_y=None,
            height=dp(22),
            halign="left",
            valign="middle",
            font_size=dp(11.5),
        )
        subtitle.bind(size=self._sync_label_text_size)
        head.add_widget(title_lbl)
        head.add_widget(subtitle)
        if not getattr(self, "ui_mobile", False):
            chip_row = GridLayout(cols=3, spacing=dp(8), size_hint_y=None)
            chip_row.bind(minimum_height=chip_row.setter("height"))
            chip_row.add_widget(self._make_popup_chip(f"Boxes: {len(ids)}", tone="primary"))
            chip_row.add_widget(self._make_popup_chip(f"Mode: {'Grid' if str(current_grid_flag).strip() == '1' else 'Direct'}", tone="accent"))
            chip_row.add_widget(self._make_popup_chip("Mapped" if existing else "New Mapping", tone="muted" if existing else "danger"))
            head.add_widget(chip_row)
        outer.add_widget(head)

        if existing:
            parts = []
            if existing:
                parts.append(existing.get("column", "") or "—")
            summary_text = "Current: " + " • ".join(parts)
        else:
            summary_text = ("Choose a data column for these selected boxes." if getattr(self, "ui_mobile", False) else "Choose a data column for the selected boxes.")
        summary_lbl = Label(
            text=summary_text,
            color=palette.get("text", (0.93, 0.96, 1.0, 1)),
            size_hint_y=None,
            height=dp(42),
            halign="left",
            valign="middle",
            font_size=dp(11.5),
        )
        summary_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(summary_lbl, min_height=dp(44), extra_pad=dp(8))
        self._style_popup_card(summary_lbl, palette.get("chip", (0.13, 0.18, 0.27, 1)), radius=dp(16))
        body_scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, bar_width=dp(6), bar_margin=dp(10), size_hint=(1, 1))
        body = BoxLayout(orientation="vertical", spacing=dp(10), size_hint_y=None)
        body.bind(minimum_height=body.setter("height"))
        body.add_widget(summary_lbl)

        live_preview_lbl = Label(
            text="Sample Output: —",
            color=palette.get("text", (0.93, 0.96, 1.0, 1)),
            size_hint_y=None,
            height=dp(42),
            halign="left",
            valign="middle",
            font_size=dp(11.5),
        )
        live_preview_lbl.bind(size=self._sync_label_text_size)
        self._bind_auto_height_label(live_preview_lbl, min_height=dp(44), extra_pad=dp(8))
        self._style_popup_card(live_preview_lbl, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(16))
        body.add_widget(live_preview_lbl)

        def _popup_field_block(label_text, widget, helper_text=""):
            block = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
            block.bind(minimum_height=block.setter("height"))
            label = Label(text=label_text, color=palette.get("muted", (0.60, 0.68, 0.80, 1)), size_hint_y=None, height=dp(18), halign="left", valign="middle", font_size=dp(11.5))
            label.bind(size=self._sync_label_text_size)
            self._bind_auto_height_label(label, min_height=dp(18), extra_pad=dp(2))
            block.add_widget(label)
            block.add_widget(widget)
            if helper_text:
                helper_lbl = Label(text=helper_text, color=palette.get("muted", (0.60, 0.68, 0.80, 1)), size_hint_y=None, height=dp(14), halign="left", valign="middle", font_size=dp(10.5))
                helper_lbl.bind(size=self._sync_label_text_size)
                self._bind_auto_height_label(helper_lbl, min_height=dp(14), extra_pad=dp(4))
                block.add_widget(helper_lbl)
            return block

        form = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        form.bind(minimum_height=form.setter("height"))

        column_spinner = SearchableSelectField(
            text=current_col if current_col else COLUMN_SELECT_TEXT,
            values=self.column_spinner.values,
            placeholder=COLUMN_SELECT_TEXT,
            picker_title="Choose which column fills this field",
            search_hint="Type to find a column name",
            size_hint_y=None,
            height=dp(46),
            background_normal="",
            background_color=palette.get("surface_soft", (0.135, 0.16, 0.215, 1)),
            color=palette.get("text", (0.93, 0.96, 1.0, 1)),
        )
        form.add_widget(_popup_field_block("Data Column", column_spinner, "Type to filter the column list, then press Enter or choose the column for the selected box or boxes"))

        trigger_input = TextInput(
            text=current_trigger,
            hint_text="Checkbox value to match (optional)",
            multiline=False,
            size_hint_y=None,
            height=dp(46),
            background_normal="",
            background_active="",
            background_color=palette.get("surface_soft", (0.135, 0.16, 0.215, 1)),
            foreground_color=palette.get("text", (0.93, 0.96, 1.0, 1)),
            cursor_color=palette.get("accent", (0.10, 0.78, 0.63, 1)),
            hint_text_color=palette.get("muted", (0.60, 0.68, 0.80, 1)),
            padding=[dp(12), dp(12), dp(12), dp(12)],
        )
        form.add_widget(_popup_field_block("Checkbox Trigger Value", trigger_input, "Only needed when this link should place an X in a checkbox" if getattr(self, "ui_mobile", False) else "Only needed when this link should place an X in a checkbox"))

        grid_row = GridLayout(cols=1 if getattr(self, "ui_mobile", False) else 2, size_hint_y=None, spacing=dp(8))
        grid_row.bind(minimum_height=grid_row.setter("height"))
        popup_grid_chk = CheckBox(active=(str(current_grid_flag).strip() == "1"))
        grid_flag_wrap = BoxLayout(orientation="horizontal", spacing=dp(8), padding=[dp(10), dp(6), dp(10), dp(6)], size_hint_y=None, height=dp(46))
        self._style_popup_card(grid_flag_wrap, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(16))
        grid_flag_lbl = Label(text="Split Across Boxes", color=palette.get("text", (0.93, 0.96, 1.0, 1)), halign="left", valign="middle", font_size=dp(12))
        grid_flag_lbl.bind(size=self._sync_label_text_size)
        grid_flag_wrap.add_widget(grid_flag_lbl)
        popup_grid_chk.size_hint_x = None
        popup_grid_chk.width = dp(36)
        grid_flag_wrap.add_widget(popup_grid_chk)
        grid_n_input = TextInput(
            text=current_grid_n,
            hint_text="Number of boxes",
            multiline=False,
            input_filter="int",
            size_hint_y=None,
            height=dp(46),
            background_normal="",
            background_active="",
            background_color=palette.get("surface_soft", (0.135, 0.16, 0.215, 1)),
            foreground_color=palette.get("text", (0.93, 0.96, 1.0, 1)),
            cursor_color=palette.get("accent", (0.10, 0.78, 0.63, 1)),
            hint_text_color=palette.get("muted", (0.60, 0.68, 0.80, 1)),
            padding=[dp(12), dp(12), dp(12), dp(12)],
        )
        grid_row.add_widget(_popup_field_block("Split Across Boxes", grid_flag_wrap, "Turn this on when one value should be spread across several small boxes"))
        grid_row.add_widget(_popup_field_block("Number of Boxes", grid_n_input, "How many boxes or cells should receive the split text" if not getattr(self, "ui_mobile", False) else "How many boxes or cells should receive the split text"))
        form.add_widget(grid_row)
        body.add_widget(form)
        body_scroll.add_widget(body)
        outer.add_widget(body_scroll)

        def _refresh_live_preview(*_):
            column_preview = self._get_selected_column_preview(
                column_spinner.text,
                trigger=trigger_input.text.strip(),
                is_grid=bool(getattr(popup_grid_chk, "active", False)),
                grid_n=int(grid_n_input.text.strip() or "1"),
            )
            live_preview_lbl.text = "Sample Output: " + column_preview

        column_spinner.bind(text=_refresh_live_preview)
        trigger_input.bind(text=_refresh_live_preview)
        popup_grid_chk.bind(active=lambda *_: _refresh_live_preview())
        grid_n_input.bind(text=_refresh_live_preview)
        _refresh_live_preview()


        if getattr(self, "ui_mobile", False):
            btn_row = BoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(8))
            btn_row.bind(minimum_height=btn_row.setter("height"))
            primary_row = GridLayout(cols=2, size_hint_y=None, spacing=dp(8), height=dp(46))
            btn_assign = self._make_compact_action_button("Save Link", tone="primary")
            btn_clear = self._make_compact_action_button("Clear Link", tone="soft")
            primary_row.add_widget(btn_assign)
            primary_row.add_widget(btn_clear)
            btn_close = self._make_compact_action_button("Close", tone="plain")
            btn_close.size_hint_y = None
            btn_close.height = dp(44)
            btn_row.add_widget(primary_row)
            btn_row.add_widget(btn_close)
        else:
            btn_row = GridLayout(cols=3, size_hint_y=None, spacing=dp(8), height=dp(48))
            btn_assign = self._make_compact_action_button("Save Link", tone="primary")
            btn_clear = self._make_compact_action_button("Clear Link", tone="soft")
            btn_close = self._make_compact_action_button("Close", tone="plain")
            for btn in (btn_assign, btn_clear, btn_close):
                btn.size_hint_y = None
                btn.height = dp(42)
                btn_row.add_widget(btn)
        outer.add_widget(btn_row)

        popup = Popup(
            title="",
            separator_height=0,
            background="",
            background_color=(0, 0, 0, 0.80),
            content=outer,
            size_hint=(0.96 if getattr(self, "ui_mobile", False) else 0.62, 0.90 if getattr(self, "ui_mobile", False) else 0.64),
            auto_dismiss=True,
        )

        self._bind_popup_background_softening(popup)

        def _assign(*_):
            try:
                column = column_spinner.text.strip()
                if not column or column == COLUMN_SELECT_TEXT:
                    raise ValueError("Please choose a data column.")
                trigger = trigger_input.text.strip()
                is_grid = bool(getattr(popup_grid_chk, "active", False))
                grid_n = int(grid_n_input.text.strip() or "1")

                did_mapping = False
                if column and column != COLUMN_SELECT_TEXT:
                    self.engine.assign_mapping(ids, column, trigger, is_grid, grid_n, page_idx)
                    did_mapping = True

                self.column_spinner.text = column if (column and column != COLUMN_SELECT_TEXT) else COLUMN_SELECT_TEXT
                self.trigger_input.text = trigger if did_mapping else ""

                if hasattr(self, "grid_flag_chk"):
                    self.grid_flag_chk.active = bool(is_grid)
                else:
                    self.grid_flag_input.text = "1" if is_grid else "0"
                self.grid_n_input.text = str(grid_n)
                self.box_ids_input.text = ",".join(str(i) for i in ids)
                self.engine.selected_box_ids = ids
                self._sync_box_selection_ui()
                saved_parts = []
                if did_mapping:
                    saved_parts.append(f"Column: {column}")
                self.set_status(
                    f"Links saved from preview.\nPage: {page_idx}\n" + "\n".join(saved_parts) + f"\nBoxes: {ids}"
                )
                popup.dismiss()
                self.on_preview(None)
            except Exception as e:
                self.set_status(f"Preview mapping error:\n{e}")

        def _clear(*_):
            try:
                self.clear_mapping_for_box_ids(ids, page_idx)
                self.engine.selected_box_ids = ids
                self._sync_box_selection_ui()
                self.set_status(f"Cleared mapping for boxes: {ids} on page {page_idx}")
                popup.dismiss()
                self.on_preview(None)
            except Exception as e:
                self.set_status(f"Clear mapping error:\n{e}")

        btn_assign.bind(on_release=_assign)
        btn_clear.bind(on_release=_clear)
        btn_close.bind(on_release=lambda *_: popup.dismiss())
        popup.open()

    def clear_mapping_for_box_ids(self, box_ids, page_idx):
        box_ids = sorted(set(int(x) for x in box_ids))
        target_rects = []
        for b_id in box_ids:
            if b_id < 0 or b_id >= len(self.engine.all_boxes):
                continue
            target_rects.append(self.engine.all_boxes[b_id])

        def overlaps_any(mapping_item):
            existing_rects = self.engine._mapping_rect_list(mapping_item)
            for er in existing_rects:
                for sr in target_rects:
                    if self.engine._rects_refer_to_same_target(er, sr, tol=0.20):
                        return True
            return False

        for k in list(self.engine.custom_mappings.keys()):
            kept = []
            for m in self.engine.custom_mappings[k]:
                if m.get("page", 0) != page_idx:
                    kept.append(m)
                    continue
                if overlaps_any(m):
                    continue
                kept.append(m)
            if kept:
                self.engine.custom_mappings[k] = kept
            else:
                del self.engine.custom_mappings[k]

    def on_clear_selected_mapping(self, *_):
        ids = sorted(set(int(x) for x in getattr(self.engine, "selected_box_ids", []) if isinstance(x, int) or str(x).isdigit()))
        if not ids:
            self.set_status("Select at least one box first.")
            return
        self.clear_mapping_for_box_ids(ids, self.current_page_idx())
        if ids:
            self._load_mapping_into_editor(ids[0])
        else:
            self._reset_mapping_editor_state(clear_box_ids=False)
        self._sync_box_selection_ui()
        self.set_status(f"Cleared links for box(es): {', '.join(str(i) for i in ids)}")
        self.on_preview(None)

    def _make_styled_popup(self, title, body, subtitle="", size_hint=(0.9, 0.9), auto_dismiss=True):
        palette = getattr(self, "ui_palette", {}) or {}
        outer = BoxLayout(orientation="vertical", spacing=dp(10), padding=[dp(12), dp(12), dp(12), dp(12)])
        self._style_popup_card(outer, palette.get("surface", (0.085, 0.105, 0.145, 1)), radius=dp(22))
        head = BoxLayout(orientation="horizontal", spacing=dp(8), size_hint_y=None, height=dp(52))
        title_wrap = BoxLayout(orientation="vertical", spacing=dp(2))
        ttl = Label(text=str(title), color=palette.get("text", (0.93, 0.96, 1.0, 1)), size_hint_y=None, height=dp(24), halign="left", valign="middle", font_size=dp(17), bold=True)
        ttl.bind(size=self._sync_label_text_size)
        title_wrap.add_widget(ttl)
        if subtitle:
            sub = Label(text=str(subtitle), color=palette.get("muted", (0.60, 0.68, 0.80, 1)), size_hint_y=None, height=dp(18), halign="left", valign="middle", font_size=dp(10.5))
            sub.bind(size=self._sync_label_text_size)
            title_wrap.add_widget(sub)
        head.add_widget(title_wrap)
        popup = Popup(title="", separator_height=0, background="", background_color=(0, 0, 0, 0.72), content=outer, size_hint=size_hint, auto_dismiss=auto_dismiss)
        btn_close = self._make_compact_action_button("Close", tone="soft")
        btn_close.size_hint = (None, None)
        btn_close.size = (dp(86), dp(36))
        btn_close.bind(on_release=lambda *_: popup.dismiss())
        head.add_widget(btn_close)
        outer.add_widget(head)
        outer.add_widget(body)
        return popup, outer

    def _open_styled_filechooser_popup(self, title, filters, on_submit, subtitle="Double-click a file or use Open."):
        chooser = FileChooserListView(filters=list(filters or []), path=self.get_default_file_path())
        chooser_box = BoxLayout(orientation="vertical", spacing=dp(8))
        self._style_popup_card(chooser_box, getattr(self, "ui_palette", {}).get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(18))
        chooser_box.padding = [dp(8), dp(8), dp(8), dp(8)]
        chooser_box.add_widget(chooser)
        btn_row = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(44))
        btn_cancel = self._make_compact_action_button("Cancel", tone="plain")
        btn_open = self._make_compact_action_button("Open", tone="primary")
        btn_row.add_widget(btn_cancel)
        btn_row.add_widget(btn_open)
        chooser_box.add_widget(btn_row)
        popup, _outer = self._make_styled_popup(title, chooser_box, subtitle=subtitle, size_hint=(0.92, 0.92 if getattr(self, "ui_mobile", False) else 0.90))

        def _submit(selection=None):
            sel = list(selection if selection is not None else getattr(chooser, "selection", []) or [])
            if not sel:
                self.set_status(f"No file selected for {title}.")
                return
            on_submit(sel, popup)

        chooser.bind(on_submit=lambda _obj, sel, _touch: _submit(sel))
        btn_cancel.bind(on_release=lambda *_: popup.dismiss())
        btn_open.bind(on_release=lambda *_: _submit())
        popup.open()
        return popup


    # --------------------------------------------------------
    # File loading
    # --------------------------------------------------------
    def _load_dataframe_from_local_path(self, path):
        self.engine.load_dataframe(path)
        self.refresh_patient_and_column_lists()
        self._persist_runtime_session_state(current_page_idx=self.current_page_idx())
        self.set_status(
            f"Data loaded:\n{os.path.basename(path)}\nRows: {len(self.engine.df)}"
        )
        Clock.schedule_once(lambda dt: self._prompt_record_label_fallback_if_needed(), 0.05)
        if self.engine.pdf_path:
            Clock.schedule_once(lambda dt: self.on_preview(None), 0.1)

    def on_load_csv(self, instance):
        if platform == "android":
            return self._open_android_document_picker(
                request_code=42431,
                mime_type="*/*",
                title="data file",
                on_picked=self._load_dataframe_from_local_path,
                cancel_message="Data file selection cancelled.",
                allowed_suffixes=[".csv", ".xlsx", ".xls"],
                extra_mime_types=[
                    "text/csv",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel",
                    "text/plain",
                    "application/octet-stream",
                ],
            )

        self._open_styled_filechooser_popup(
            "Select CSV/XLSX File",
            ["*.csv", "*.xlsx", "*.xls"],
            self._handle_csv_selection,
            subtitle="Load the spreadsheet or CSV that will supply record values.",
        )

    def _handle_csv_selection(self, selection, popup):
        if selection:
            try:
                path = selection[0]
                self._load_dataframe_from_local_path(path)
            except Exception as e:
                traceback.print_exc()
                msg = str(e)
                if "codec" in msg.lower() or "encoding" in msg.lower() or "decode" in msg.lower():
                    friendly = "Could not read the file — try saving it as UTF-8 CSV and reload."
                elif "empty" in msg.lower() or "no columns" in msg.lower():
                    friendly = "The file appears to be empty or has no data columns."
                elif "sheet" in msg.lower():
                    friendly = "Could not read the spreadsheet. Make sure it is a valid XLSX or CSV file."
                else:
                    friendly = "Could not load the data file. Make sure it is a valid CSV or XLSX file."
                self.set_status(friendly, kind="error")
        popup.dismiss()

    def _load_config_from_local_path(self, path):
        if not path.lower().endswith(".json"):
            raise ValueError("Please select a JSON config file.")

        cfg = self.engine.load_config(path)
        self._reset_config_runtime_state()
        self.push_engine_settings_to_ui()

        saved_pdf_path = str(cfg.get("pdf_path", "") or "").strip()
        loaded_pdf = False
        loaded_total = 0
        if saved_pdf_path:
            try:
                if os.path.exists(saved_pdf_path):
                    loaded_total = int(self.engine.load_pdf(saved_pdf_path) or 0)
                    loaded_pdf = True
                else:
                    self.set_status(
                        f"Config loaded, but saved PDF was not found:\n{saved_pdf_path}"
                    )
            except Exception as e:
                traceback.print_exc()
                self.set_status(
                    f"Config loaded, but PDF restore failed:\n{os.path.basename(saved_pdf_path)}\n{e}"
                )

        self._apply_runtime_config_state(cfg)

        if loaded_pdf:
            try:
                page_idx = self.current_page_idx()
                total = max(int(self.engine.total_pages()), 1)
                page_idx = max(0, min(page_idx, total - 1))
                if hasattr(self, "page_input"):
                    self._set_page_inputs_text(page_idx)
                Clock.schedule_once(lambda dt: self.on_preview(None), 0.05)
                self.set_status(
                    f"Config loaded:\n{os.path.basename(path)}\n"
                    f"PDF restored: {os.path.basename(saved_pdf_path)}\n"
                    f"Page: {page_idx + 1}/{total}"
                )
            except Exception as e:
                traceback.print_exc()
                self.set_status(
                    f"Config loaded, but preview restore failed:\n{os.path.basename(path)}\n{e}"
                )
        else:
            self.set_status(f"Config loaded:\n{os.path.basename(path)}")

        self._persist_runtime_session_state(current_page_idx=self.current_page_idx())
        self.refresh_backend_capabilities_ui()

    def on_load_config(self, instance):
        if platform == "android":
            return self._open_android_document_picker(
                request_code=42432,
                mime_type="*/*",
                title="config file",
                on_picked=self._load_config_from_local_path,
                cancel_message="Config selection cancelled.",
                required_suffix=".json",
                allowed_suffixes=[".json"],
                extra_mime_types=["application/json", "text/plain", "application/octet-stream"],
            )

        self._open_styled_filechooser_popup(
            "Select Config File",
            ["*.json"],
            self._handle_load_config_selection,
            subtitle="Open a saved Form Alchemist JSON configuration.",
        )

    def _handle_load_config_selection(self, selection, popup):
        if selection:
            try:
                path = selection[0]
                self._load_config_from_local_path(path)
            except Exception as e:
                traceback.print_exc()
                self.set_status(f"Load config error:\n{e}")
        popup.dismiss()

    def _merge_config_from_local_path(self, path):
        if not path.lower().endswith(".json"):
            raise ValueError("Please select a JSON config file.")

        with open(path, "r", encoding="utf-8") as f:
            incoming = json.load(f)

        self.apply_ui_settings_to_engine()
        self.engine.merge_config_into_current(
            incoming_cfg=incoming,
            keep_current_detection=True,
            prefer="incoming"
        )
        self.engine.save_learning_revision(
            reason="merge_config",
            approved=False,
            learning_weight=0.25,
            source_path=path,
        )
        self.engine.all_boxes = []
        self.engine.box_types = []
        self.push_engine_settings_to_ui()
        self.set_status(f"Config merged:\n{os.path.basename(path)}")

    def on_merge_config(self, instance):
        if platform == "android":
            return self._open_android_document_picker(
                request_code=42433,
                mime_type="*/*",
                title="config file to merge",
                on_picked=self._merge_config_from_local_path,
                cancel_message="Merge config selection cancelled.",
                required_suffix=".json",
                allowed_suffixes=[".json"],
                extra_mime_types=["application/json", "text/plain", "application/octet-stream"],
            )

        self._open_styled_filechooser_popup(
            "Select Config File to Merge",
            ["*.json"],
            self._handle_merge_config_selection,
            subtitle="Merge another JSON config into the current session.",
        )

    def _handle_merge_config_selection(self, selection, popup):
        if selection:
            try:
                path = selection[0]
                self._merge_config_from_local_path(path)
            except Exception as e:
                traceback.print_exc()
                self.set_status(f"Merge config error:\n{e}")
        popup.dismiss()

    def _collect_runtime_config_state(self):
        page_idx = 0
        try:
            page_idx = int(self.current_page_idx())
        except Exception:
            page_idx = 0

        selected_ids = []
        try:
            selected_ids = list(self._sanitize_box_id_list(getattr(self.engine, "selected_box_ids", [])))
        except Exception:
            selected_ids = []

        ui_state = {
            "page_idx": page_idx,
            "preview_zoom_mode": str(getattr(self, "preview_zoom_mode", "fit_width")),
            "preview_zoom_factor": float(getattr(self, "preview_zoom_factor", 1.0)),
            "desktop_zoom_mode": str(getattr(self, "desktop_zoom_mode", "manual")),
            "desktop_zoom_factor": float(getattr(self, "desktop_zoom_factor", 1.0)),
            "selected_box_ids": selected_ids,
            "selected_record_label": str(self.selected_record() or ""),
            "selected_record_key": str(self.selected_record_key() or ""),
            "record_label_source_column": str(getattr(self.engine, "record_label_source_column", "") or ""),
            "session_preview_mode": str(self._session_preview_mode(page_idx)),
            "status_kind": str(getattr(self, "_status_kind", "info") or "info"),
            "mobile_sidebar_open": bool(getattr(self, "_mobile_sidebar_open", False)),
        }
        if hasattr(self, "_page_selected_box_ids"):
            try:
                ui_state["page_selected_box_ids"] = {
                    str(int(k)): list(self._sanitize_box_id_list(v))
                    for k, v in dict(self._page_selected_box_ids).items()
                }
            except Exception:
                ui_state["page_selected_box_ids"] = {}
        return ui_state

    def _persist_runtime_session_state(self, current_page_idx=None):
        try:
            self.engine._config_ui_state = self._collect_runtime_config_state()
            page_idx = self.current_page_idx() if current_page_idx is None else int(current_page_idx)
            self.engine.persist_learning_session(current_page_idx=page_idx)
        except Exception:
            pass

    def _reset_config_runtime_state(self):
        self.engine.all_boxes = []
        self.engine.box_types = []
        self.engine.geom = self.engine._empty_geom()
        self.engine.semantic_targets = self.engine._empty_semantic_targets()
        self.engine.selected_box_ids = []
        self.engine.detected_page_idx = None
        self.engine.detected_settings_signature = None
        self.engine.boxes_by_page = {}
        self.engine.box_types_by_page = {}
        self.engine.geom_by_page = {}
        self.engine.semantic_targets_by_page = {}
        self.engine.boxes_settings_signature_by_page = {}

        self._page_selected_box_ids = {}
        self._active_selection_page_idx = 0

        if hasattr(self, "preview"):
            try:
                self.preview.boxes_payload = []
            except Exception:
                pass
            try:
                self.preview.set_selected_ids([])
            except Exception:
                pass
        if hasattr(self, "box_ids_input"):
            self.box_ids_input.text = ""

    def _config_ready_preview_hint(self):
        if getattr(self, "ui_mobile", False):
            mode_txt = "Select ON" if bool(getattr(self, "mobile_select_mode", False)) else "Select OFF"
            return f"Preview ready. {mode_txt} • Tap a box to select • Map if correct • Repair only if needed."
        return "Preview ready. Tap a box to select • Map if correct • Repair only if needed."

    def _clear_config_import_selection_state(self, clear_page_cache=True, update_hint=True):
        self.engine.selected_box_ids = []
        if clear_page_cache or not hasattr(self, "_page_selected_box_ids"):
            self._page_selected_box_ids = {}
        else:
            try:
                self._page_selected_box_ids[int(self.current_page_idx())] = []
            except Exception:
                pass
        try:
            self._active_selection_page_idx = int(self.current_page_idx())
        except Exception:
            self._active_selection_page_idx = 0
        try:
            self._reset_mapping_editor_state(clear_box_ids=True)
        except Exception:
            pass
        preview = getattr(self, "preview", None)
        if preview is not None:
            try:
                preview.set_selected_ids([])
            except Exception:
                pass
            try:
                preview.hovered_box_id = None
                preview._redraw_overlay()
            except Exception:
                pass
        if update_hint and hasattr(self, "preview_info") and self.preview_info is not None:
            try:
                self.preview_info.text = self._config_ready_preview_hint()
            except Exception:
                pass
        try:
            self._update_preview_hud(hit=None)
        except Exception:
            pass
        try:
            self._update_selection_inspector()
        except Exception:
            pass
        try:
            self._update_bottom_statusbar()
        except Exception:
            pass

    def _refresh_after_config_import(self, page_idx=None, reason="Config updated", anchor=None):
        try:
            page_idx = self.current_page_idx() if page_idx is None else int(page_idx)
        except Exception:
            page_idx = 0
        self._clear_config_import_selection_state(clear_page_cache=False, update_hint=False)
        try:
            self._render_session_page(page_idx=page_idx, reason=reason)
        finally:
            if anchor is not None:
                Clock.schedule_once(lambda dt, a=anchor: self._restore_preview_anchor(a), 0)
                Clock.schedule_once(lambda dt, a=anchor: self._restore_preview_anchor(a), 0.06)

    def _apply_runtime_config_state(self, cfg, restore_selection=False):
        ui_state = cfg.get("ui_state", {}) if isinstance(cfg, dict) else {}
        if not isinstance(ui_state, dict):
            ui_state = {}

        def _safe_int(value, default=0):
            try:
                return int(value)
            except Exception:
                try:
                    return int(float(value))
                except Exception:
                    return int(default)

        def _safe_float(value, default=1.0):
            try:
                return float(value)
            except Exception:
                return float(default)

        def _loose_box_ids(values):
            out = []
            for item in (values or []):
                try:
                    out.append(int(item))
                except Exception:
                    continue
            return sorted(set(out))

        selected_record_label = str(ui_state.get("selected_record_label", cfg.get("selected_record_label", "") if isinstance(cfg, dict) else "") or "").strip()
        selected_record_key = str(ui_state.get("selected_record_key", cfg.get("selected_record_key", "") if isinstance(cfg, dict) else "") or "").strip()
        restored_record_label_source = str(ui_state.get("record_label_source_column", cfg.get("record_label_source_column", "") if isinstance(cfg, dict) else "") or "").strip()
        if restored_record_label_source and getattr(self.engine, "df", None) is not None and not self.engine.df.empty and restored_record_label_source in list(self.engine.df.columns):
            try:
                self.engine.rebuild_record_labels(source_column=restored_record_label_source)
                if hasattr(self, "patient_spinner") and hasattr(self, "column_spinner"):
                    self.refresh_patient_and_column_lists()
            except Exception:
                pass
        self._pending_restored_record_label = selected_record_label
        self._pending_restored_record_key = selected_record_key

        target_page = _safe_int(ui_state.get("page_idx", 0), 0)
        fallback_zoom = _safe_float(cfg.get("zoom", getattr(self.engine, "config_zoom", getattr(self, "preview_zoom_factor", 1.0))) or 1.0, 1.0)
        self.preview_zoom_mode = str(ui_state.get("preview_zoom_mode", getattr(self, "preview_zoom_mode", "fit_width")))
        self.preview_zoom_factor = max(0.35, min(_safe_float(ui_state.get("preview_zoom_factor", fallback_zoom), fallback_zoom), 8.0))
        self.desktop_zoom_mode = str(ui_state.get("desktop_zoom_mode", getattr(self, "desktop_zoom_mode", "manual")))
        self.desktop_zoom_factor = max(0.2, min(_safe_float(ui_state.get("desktop_zoom_factor", fallback_zoom), fallback_zoom), 6.0))
        sidebar_open = bool(ui_state.get("mobile_sidebar_open", False))
        self._restored_session_preview_mode = str(ui_state.get("session_preview_mode", "") or "")
        self._status_kind = str(ui_state.get("status_kind", getattr(self, "_status_kind", "info")) or "info")

        total = 0
        try:
            total = int(self.engine.total_pages())
        except Exception:
            total = 0
        if total > 0:
            target_page = max(0, min(target_page, total - 1))
        else:
            target_page = 0

        self._page_selected_box_ids = {}
        selected_ids = []
        if restore_selection:
            page_selected = ui_state.get("page_selected_box_ids", {}) or {}
            if isinstance(page_selected, dict):
                for k, v in page_selected.items():
                    try:
                        self._page_selected_box_ids[_safe_int(k, 0)] = list(_loose_box_ids(v))
                    except Exception:
                        continue

            selected_ids = list(_loose_box_ids(ui_state.get("selected_box_ids", [])))
            if selected_ids:
                self._page_selected_box_ids[target_page] = list(selected_ids)

        self._active_selection_page_idx = target_page
        self.engine.selected_box_ids = list(selected_ids)

        if hasattr(self, "page_input"):
            self._set_page_inputs_text(target_page)

        if self.engine.df is not None and not self.engine.df.empty and hasattr(self, "patient_spinner"):
            restored_label = self.engine.resolve_record_label(record_key=selected_record_key, fallback_label=selected_record_label)
            if restored_label and restored_label in list(getattr(self.patient_spinner, "values", []) or []):
                self.patient_spinner.text = restored_label
                self._pending_restored_record_label = ""
                self._pending_restored_record_key = ""

        self._restore_page_selection(target_page, clear_if_missing=False)
        self._sync_box_selection_ui()
        self._update_bottom_statusbar()
        if platform == "android":
            Clock.schedule_once(lambda dt, open_state=sidebar_open: self._set_mobile_sidebar_state(open_state), 0)
        try:
            if getattr(self.engine, "pdf_path", ""):
                self.preview_info.text = "Config loaded. Rendering the active page with the imported mappings."
            else:
                self.preview_info.text = "Config loaded. Load the saved PDF to resume preview on the active page."
        except Exception:
            pass

    def _build_config_json_payload(self):
        self.apply_ui_settings_to_engine()
        self.engine._config_ui_state = self._collect_runtime_config_state()
        cfg = self.engine.collect_config()
        payload = json.dumps(cfg, indent=2, ensure_ascii=False).encode("utf-8")
        if not payload:
            raise ValueError("Config payload is empty.")
        return cfg, payload

    def _save_android_config_via_picker(self, suggested_name=CONFIG_FILENAME, request_code=42443):
        cfg, payload = self._build_config_json_payload()

        def _handle_uri(uri):
            try:
                written = self.android_picker.write_bytes_to_uri(uri, payload)
                uri_text = str(uri) if uri is not None else ""
                self.engine.save_learning_revision(
                    reason="save_config",
                    approved=False,
                    learning_weight=0.35,
                    source_path=uri_text or str(suggested_name or CONFIG_FILENAME),
                )
                self.engine.persist_learning_session(current_page_idx=self.current_page_idx())
                lines = ["Config saved:", str(suggested_name or CONFIG_FILENAME), f"Bytes: {written}"]
                if uri_text:
                    lines.append(uri_text)
                self.set_status("\n".join(lines))
            except Exception as e:
                traceback.print_exc()
                self.set_status(f"Save config error:\n{e}")

        self.set_status(f"Choose save location:\n{suggested_name}")
        return self._open_android_create_document(
            request_code=request_code,
            suggested_name=str(suggested_name or CONFIG_FILENAME),
            mime_type="application/json",
            on_picked=_handle_uri,
            cancel_message="Config save cancelled.",
        )

    def on_save_config(self, instance):
        try:
            if platform == "android":
                return self._save_android_config_via_picker(suggested_name=CONFIG_FILENAME, request_code=42443)

            self.apply_ui_settings_to_engine()
            self.engine._config_ui_state = self._collect_runtime_config_state()
            out_path = os.path.join(self.get_app_output_dir(), CONFIG_FILENAME)
            self.engine.save_config(out_path)
            self.engine.save_learning_revision(
                reason="save_config",
                approved=False,
                learning_weight=0.35,
                source_path=out_path,
            )
            self.set_status(f"Config saved:\n{out_path}")
        except Exception as e:
            self.set_status(f"Save config error:\n{e}")

    # --------------------------------------------------------
    # Navigation
    # --------------------------------------------------------
    def _prepare_mobile_page_transition(self):
        if not getattr(self, "ui_mobile", False):
            return
        mode = str(getattr(self, "preview_zoom_mode", "fit_width") or "fit_width")
        if mode in ("fit_width", "fit_page"):
            self._mobile_zoom_bootstrap_pending = True
        self._set_mobile_quick_rail_visible(False)

    def on_prev_page(self, instance):
        try:
            return self._go_to_page(self.current_page_idx() - 1, reason="Previous page")
        except Exception as e:
            self.set_status(f"Prev page error:\n{e}", kind="error", force=True)

    def on_next_page(self, instance):
        try:
            return self._go_to_page(self.current_page_idx() + 1, reason="Next page")
        except Exception as e:
            self.set_status(f"Next page error:\n{e}", kind="error", force=True)

    def on_patient_change(self, spinner, text):
        page_idx = self.current_page_idx()
        self._persist_runtime_session_state(current_page_idx=page_idx)
        self.refresh_backend_capabilities_ui()
        if not self.engine.pdf_path:
            self.set_status(
                "Record selection updated.\nOpen a PDF form to preview this record.",
                kind="session",
                hold_seconds=1.5,
                force=True,
            )
            return
        reason = "Record cleared" if not text or text in {RECORD_SELECT_TEXT, "Select Patient"} else "Record changed"
        self.set_status(f"{reason}.\nRefreshing current page…", kind="session", hold_seconds=1.5, force=True)
        Clock.schedule_once(lambda dt, idx=page_idx, why=reason: self._render_session_page(page_idx=idx, reason=why), 0.05)

    # --------------------------------------------------------
    # Detection / preview    # --------------------------------------------------------
    # Detection / preview
    # --------------------------------------------------------
    def on_run_detect(self, instance):
        try:
            if not self.engine.pdf_path:
                self.set_status("Load PDF first.")
                return

            self.apply_ui_settings_to_engine()
            page_idx = self.current_page_idx()
            self._clear_mobile_repair_state(page_idx=page_idx)
            prev_count = len(getattr(self.engine, "all_boxes", []) or []) if getattr(self.engine, "detected_page_idx", None) == page_idx else 0
            ctx = self.engine.prepare_learning_for_detection(page_idx=page_idx, allow_profile_apply=False)
            profile_applied = bool((ctx or {}).get("profile_applied", False))
            profile_matched = bool((ctx or {}).get("matched_profile"))
            self.engine.invalidate_detection_cache(page_idx=page_idx, clear_current=True)
            self.engine.run_detection(page_idx=page_idx)

            # Detect should re-apply explicit saved Area Templates and saved repair
            # patches, but should NOT silently carry over transient sticky/manual
            # detections. This keeps template behavior durable while avoiding hidden
            # one-off carry-over state.
            auto_template_result = self.engine.auto_apply_saved_area_templates_after_detection(
                page_idx=page_idx,
                max_matches_per_template=18,
                min_score=0.58,
            ) or {}
            auto_repair_result = self.engine.auto_apply_saved_repair_patches_after_detection(
                page_idx=page_idx,
                min_anchor_overlap=0.40,
            ) or {}
            sticky_result = {'sticky_count': 0, 'added_count': 0, 'selected_ids': []}

            self.engine.finalize_learning_after_detection(page_idx=page_idx)
            # Do not force-select template hits after Detect; just keep them merged into
            # the live detections so mappings remain stable while startup selection stays calm.
            self.engine.selected_box_ids = []
            self._stash_page_selection(page_idx)
            counts = self._box_type_counts()
            if profile_applied:
                profile_msg = "\nLearned profile: applied"
            elif profile_matched:
                profile_msg = "\nLearned profile: matched but skipped (manual tuning kept)"
            else:
                profile_msg = "\nLearned profile: no match"

            auto_tpl_count = int(auto_template_result.get('template_count', 0) or 0)
            auto_tpl_added = int(auto_template_result.get('added_count', 0) or 0)
            auto_tpl_matches = int(auto_template_result.get('match_count', 0) or 0)
            auto_repair_count = int(auto_repair_result.get('patch_count', 0) or 0)
            auto_repair_applied = int(auto_repair_result.get('applied_count', 0) or 0)
            auto_repair_added = int(auto_repair_result.get('added_count', 0) or 0)
            auto_repair_removed = int(auto_repair_result.get('removed_count', 0) or 0)
            auto_tpl_msg = ''
            if auto_tpl_count > 0:
                auto_tpl_msg = (
                    f"\nTemplates: {auto_tpl_count} saved • "
                    f"{auto_tpl_matches} strong match{'es' if auto_tpl_matches != 1 else ''} • "
                    f"{auto_tpl_added} added"
                )
            auto_repair_msg = ''
            if auto_repair_count > 0:
                auto_repair_msg = (
                    f"\nRepairs: {auto_repair_count} saved • "
                    f"{auto_repair_applied} applied • "
                    f"+{auto_repair_added} / -{auto_repair_removed}"
                )
            sticky_count = int(sticky_result.get('sticky_count', 0) or 0)
            sticky_added = int(sticky_result.get('added_count', 0) or 0)
            sticky_msg = ''
            if sticky_count > 0:
                sticky_msg = f"\nSticky: {sticky_count} saved • {sticky_added} restored"

            summary = (
                f"Detection done.\n"
                f"Page: {page_idx}\n"
                f"Boxes: {prev_count} -> {len(self.engine.all_boxes)}"
                f"\nChecks: {counts.get('check', 0)}"
                f"\nLines: {counts.get('line', 0)}"
                f"\nFields: {counts.get('field', 0)}"
                f"\nCache: refreshed{profile_msg}{auto_tpl_msg}{auto_repair_msg}{sticky_msg}"
            )
            self._persist_runtime_session_state(current_page_idx=page_idx)
            Clock.schedule_once(lambda dt: self.on_preview(None), 0)
            Clock.schedule_once(lambda dt: self.on_preview(None), 0.05)
            self.set_status(summary, kind="detect", hold_seconds=3.5, force=True)
        except Exception as e:
            traceback.print_exc()
            msg = str(e)
            if "memory" in msg.lower() or "oom" in msg.lower():
                friendly = "Detection ran out of memory.\nTry reducing the zoom or closing other apps."
            elif "opencv" in msg.lower() or "cv2" in msg.lower():
                friendly = "Field detection is not available on this device.\nYou can still map fields manually."
            else:
                friendly = "Detection failed. Try refreshing the preview first, then detect again."
            self.set_status(friendly, kind="error", force=True)

    def on_preview(self, instance):
        try:
            return self._render_session_page(page_idx=self.current_page_idx(), reason="Preview refreshed")
        except Exception as e:
            traceback.print_exc()
            self.set_status(
                "Could not render the preview.\nTap Refresh to try again.",
                kind="error",
            )

    def on_assign_mapping(self, instance):
        try:
            column = self.column_spinner.text.strip()
            if not column or column == COLUMN_SELECT_TEXT:
                self.set_status("Please select a column.")
                return

            raw_ids = [x.strip() for x in self.box_ids_input.text.split(",") if x.strip()]
            box_ids = []
            for x in raw_ids:
                if not x.isdigit():
                    raise ValueError(f"Invalid Box ID: {x}")
                box_ids.append(int(x))

            box_ids = self._sanitize_box_id_list(box_ids)
            if not box_ids:
                raise ValueError("Please enter at least one Box ID.")

            trigger = self.trigger_input.text.strip()
            is_grid = bool(self.grid_flag_chk.active) if hasattr(self, "grid_flag_chk") else bool(int(self.grid_flag_input.text.strip() or "0"))
            grid_n = int(self.grid_n_input.text.strip() or "1")
            page_idx = self.current_page_idx()

            self.engine.assign_mapping(box_ids, column, trigger, is_grid, grid_n, page_idx)
            self.engine.selected_box_ids = sorted(set(box_ids))
            self._sync_box_selection_ui()

            self.set_status(
                f"Mapping saved.\nPage: {page_idx}\nColumn: {column}\nBoxes: {box_ids}"
            )
            self.on_preview(None)
        except Exception as e:
            self.set_status(f"Assign mapping error:\n{e}", kind="error", force=True)

    def _open_android_create_document(self, request_code, suggested_name, mime_type="application/pdf", on_picked=None, cancel_message="Save cancelled."):
        if platform != "android" or not ANDROID_JAVA_AVAILABLE:
            raise RuntimeError("Android save picker is unavailable.")

        result_ok = getattr(activity, "result_ok", -1)

        def _on_activity_result(request_code_result, result_code, intent):
            if request_code_result != request_code:
                return

            activity.unbind(on_activity_result=_on_activity_result)

            if result_code != result_ok or intent is None:
                Clock.schedule_once(lambda dt, m=cancel_message: self.set_status(m), 0)
                return

            try:
                uri = intent.getData()
            except Exception:
                uri = None

            if uri is None:
                Clock.schedule_once(lambda dt, m=cancel_message: self.set_status(m), 0)
                return

            if callable(on_picked):
                def _dispatch(dt, picked_uri=uri):
                    try:
                        on_picked(picked_uri)
                    except Exception as e:
                        traceback.print_exc()
                        self.set_status(f"Save error:\n{e}")
                Clock.schedule_once(_dispatch, 0)

        activity.bind(on_activity_result=_on_activity_result)
        try:
            intent = AndroidIntent(AndroidIntent.ACTION_CREATE_DOCUMENT)
            intent.addCategory(AndroidIntent.CATEGORY_OPENABLE)
            intent.addFlags(AndroidIntent.FLAG_GRANT_READ_URI_PERMISSION)
            intent.addFlags(AndroidIntent.FLAG_GRANT_WRITE_URI_PERMISSION)
            intent.setType(mime_type or "*/*")
            try:
                intent.putExtra(AndroidIntent.EXTRA_TITLE, str(suggested_name or "output"))
            except Exception:
                pass
            AndroidPythonActivity.mActivity.startActivityForResult(intent, request_code)
        except Exception:
            try:
                activity.unbind(on_activity_result=_on_activity_result)
            except Exception:
                pass
            raise
    def _save_android_bytes_via_picker(self, payload, suggested_name, mime_type="application/pdf", request_code=42441, success_message="Saved", extra_status_lines=None):
        data = bytes(payload or b"")
        if not data:
            raise ValueError("Generated file is empty.")

        def _handle_uri(uri):
            try:
                written = self.android_picker.write_bytes_to_uri(uri, data)
                lines = [f"{success_message}:", str(suggested_name)]
                if extra_status_lines:
                    lines.extend([str(line) for line in extra_status_lines if str(line or "").strip()])
                lines.append(f"Bytes: {written}")
                self.set_status("\n".join(lines))
            except Exception as e:
                traceback.print_exc()
                self.set_status(f"Save error:\n{e}")

        self.set_status(f"Choose save location:\n{suggested_name}")
        return self._open_android_create_document(
            request_code=request_code,
            suggested_name=suggested_name,
            mime_type=mime_type,
            on_picked=_handle_uri,
            cancel_message="Save cancelled.",
        )

    # --------------------------------------------------------
    # Output generation
    # --------------------------------------------------------
    def on_generate_single(self, instance):
        try:
            if not self.engine.supports_export_backend():
                self.set_status("PDF export backend is unavailable in this build.")
                return
            if not self.engine.pdf_path:
                self.set_status("Load PDF first.")
                return

            patient = self.selected_patient()
            record_key = self.selected_record_key()
            if not patient:
                self.set_status("Select a record first.")
                return

            page_idx = self.current_page_idx()
            filename = f"Filled_{safe_name(patient)}.pdf"

            if platform == "android":
                pdf_bytes = self.engine._build_filled_pdf_bytes(patient, page_idx=page_idx, record_key=record_key, export_scope="document")
                return self._save_android_bytes_via_picker(
                    pdf_bytes,
                    filename,
                    mime_type="application/pdf",
                    request_code=42441,
                    success_message="Export saved",
                    extra_status_lines=[f"Record: {patient}", "Scope: full document"],
                )

            out_dir = self.get_app_output_dir()
            out_path = os.path.join(out_dir, filename)
            self.engine.export_filled_pdf(patient, out_path, page_idx=page_idx, record_key=record_key, export_scope="document")
            self.set_status("Generated:\n" + out_path)
        except Exception as e:
            traceback.print_exc()
            self.set_status("Generate single error:\n" + str(e))

    def _copy_android_uri_to_local_file(self, uri, default_name="selected_input", required_suffix=None, allowed_suffixes=None):
        return self.android_picker.copy_uri_to_local_file(
            uri,
            default_name=default_name,
            required_suffix=required_suffix,
            allowed_suffixes=allowed_suffixes,
        )

    def _open_android_document_picker(self, request_code, mime_type="*/*", title="document", on_picked=None, cancel_message=None, required_suffix=None, allowed_suffixes=None, extra_mime_types=None):
        return self.android_picker.open_document_picker(
            request_code=request_code,
            mime_type=mime_type,
            title=title,
            on_picked=on_picked,
            cancel_message=cancel_message,
            required_suffix=required_suffix,
            allowed_suffixes=allowed_suffixes,
            extra_mime_types=extra_mime_types,
        )

    def _load_pdf_from_path(self, path):
        try:
            self.set_status(f"Loading PDF...\n{os.path.basename(path)}")
            total = self.engine.load_pdf(path)
            self._clear_mobile_repair_state(page_idx=0, clear_selection=True)
            self._persist_runtime_session_state(current_page_idx=0)

            cur_idx = self.current_page_idx()
            max_idx = max(total - 1, 0)
            if cur_idx > max_idx:
                self._set_page_inputs_text(0)

            self.set_status(
                f"PDF loaded: {os.path.basename(path)}\n"
                f"Pages: {total}\n"
                f"Rendering preview..."
            )
            Clock.schedule_once(lambda dt: self._finish_pdf_load_preview(path, total), 0.05)
        except Exception as e:
            traceback.print_exc()
            msg = str(e)
            if "password" in msg.lower() or "encrypt" in msg.lower():
                friendly = "This PDF is password-protected and cannot be opened."
            elif "permission" in msg.lower() or "access" in msg.lower():
                friendly = "Could not read the file. Check that it is accessible and try again."
            elif "corrupt" in msg.lower() or "invalid" in msg.lower() or "eof" in msg.lower():
                friendly = "The PDF file appears to be damaged or invalid."
            else:
                friendly = "Could not open the PDF. Make sure it is a valid PDF file."
            self.set_status(f"{friendly}", kind="error")

    def _copy_android_uri_to_local_pdf(self, uri):
        return self._copy_android_uri_to_local_file(
            uri,
            default_name="selected_pdf",
            required_suffix=".pdf",
            allowed_suffixes=[".pdf"],
        )

    def _open_android_pdf_picker(self):
        return self._open_android_document_picker(
            request_code=42421,
            mime_type="application/pdf",
            title="PDF",
            on_picked=self._load_pdf_from_path,
            cancel_message="PDF selection cancelled.",
            required_suffix=".pdf",
            allowed_suffixes=[".pdf"],
        )

    def _open_legacy_pdf_picker(self):
        self._open_styled_filechooser_popup(
            "Select PDF Template",
            ["*.pdf"],
            self._handle_pdf_selection,
            subtitle="Choose the PDF form template to preview and map.",
        )

    def on_load_pdf(self, instance):
        if platform == "android":
            return self._open_android_pdf_picker()
        return self._open_legacy_pdf_picker()
    
    def _handle_pdf_selection(self, selection, popup):
        popup.dismiss()

        if selection:
            try:
                path = selection[0]
                self.set_status(f"Loading PDF...\n{os.path.basename(path)}")

                total = self.engine.load_pdf(path)
                self._clear_mobile_repair_state(page_idx=0, clear_selection=True)
                self.engine.persist_learning_session(current_page_idx=0)

                cur_idx = self.current_page_idx()
                max_idx = max(total - 1, 0)
                if cur_idx > max_idx:
                    self._set_page_inputs_text(0)

                self.set_status(
                    f"PDF Loaded: {os.path.basename(path)}\n"
                    f"Pages: {total}\n"
                    f"Rendering preview..."
                )
                Clock.schedule_once(lambda dt: self._finish_pdf_load_preview(path, total), 0.05)
            except Exception as e:
                traceback.print_exc()
                _emsg = str(e or "")
                if "not found" in _emsg.lower() or "no such file" in _emsg.lower():
                    _emsg = "The PDF file could not be found. Please try selecting it again."
                elif "invalid" in _emsg.lower() or "header" in _emsg.lower() or "pdf" in _emsg.lower():
                    _emsg = "The selected file does not appear to be a valid PDF. Please choose a different file."
                elif "permission" in _emsg.lower():
                    _emsg = "Permission denied when accessing the PDF. Try selecting it through the file picker again."
                else:
                    _emsg = f"Could not open the PDF. Details: {e}"
                self.set_status(f"PDF Error: {_emsg}")

    def _finish_pdf_load_preview(self, path, total):
        try:
            page_idx = self.current_page_idx()
            raw_img = self.engine.get_raw_preview_pixmap(
                page_idx=page_idx,
                preview_zoom=PREVIEW_SCALE
            )
            self.render_preview_image(
                raw_img,
                boxes_payload=[],
                page_idx=page_idx,
                preview_zoom=PREVIEW_SCALE
            )
            self._sync_box_selection_ui()
            Clock.schedule_once(self._post_preview_refresh, 0.05)

            mode_note = "\nMode: export is limited in this build" if not self.engine.supports_export_backend() else ""
            self.set_status(
                f"PDF Loaded: {os.path.basename(path)}\n"
                f"Pages: {total}\n"
                f"Showing raw template page: {page_idx}"
                f"{mode_note}"
            )
        except Exception as e:
            traceback.print_exc()
            self.set_status(
                "Preview could not be rendered.\n"
                "Try tapping Refresh, or reload the PDF.",
                kind="error",
            )


    def open_text_input_popup(self, title, hint_text, on_submit_callback, default_text=""):
        palette = getattr(self, "ui_palette", {}) or {}
        wrap = BoxLayout(orientation="vertical", spacing=dp(10), padding=[dp(10), dp(10), dp(10), dp(10)])
        self._style_popup_card(wrap, palette.get("surface_alt", (0.11, 0.135, 0.185, 1)), radius=dp(18))

        txt = TextInput(
            text=default_text,
            hint_text=hint_text,
            multiline=False,
            size_hint_y=None,
            height=dp(48),
            background_normal="",
            background_active="",
            background_color=palette.get("surface_soft", (0.135, 0.16, 0.215, 1)),
            foreground_color=palette.get("text", (0.93, 0.96, 1.0, 1)),
            cursor_color=palette.get("accent", (0.10, 0.78, 0.63, 1)),
            hint_text_color=palette.get("muted", (0.60, 0.68, 0.80, 1)),
            padding=[dp(12), dp(12), dp(12), dp(12)],
        )
        wrap.add_widget(txt)

        btn_row = GridLayout(cols=2, size_hint_y=None, height=dp(46), spacing=dp(8))
        btn_cancel = self._make_compact_action_button("Cancel", tone="plain")
        btn_ok = self._make_compact_action_button("OK", tone="primary")
        btn_row.add_widget(btn_cancel)
        btn_row.add_widget(btn_ok)
        wrap.add_widget(btn_row)

        popup, _outer = self._make_styled_popup(title, wrap, subtitle=hint_text, size_hint=(0.92 if getattr(self, "ui_mobile", False) else 0.58, 0.34 if getattr(self, "ui_mobile", False) else 0.30))

        btn_cancel.bind(on_release=lambda *_: popup.dismiss())

        def _submit(*_):
            try:
                on_submit_callback(txt.text.strip())
            finally:
                popup.dismiss()

        btn_ok.bind(on_release=_submit)
        txt.bind(on_text_validate=lambda *_: _submit())

        popup.open()
    
    
    def on_load_gsheet_url(self, instance):
        self.open_text_input_popup(
            title="Load Google Sheet URL",
            hint_text="Paste Google Sheet URL here",
            on_submit_callback=self._handle_gsheet_url_submit
        )
    
    
    def _handle_gsheet_url_submit(self, url):
        try:
            if not url:
                self.set_status("No Google Sheet URL provided.")
                return
    
            self.engine.load_dataframe(url)
            self.refresh_patient_and_column_lists()
    
            self.set_status(
                f"Google Sheet loaded.\n"
                f"Rows: {len(self.engine.df)}\n"
                f"Records: {len(self.engine.patient_names)}"
            )
            Clock.schedule_once(lambda dt: self._prompt_record_label_fallback_if_needed(), 0.05)
    
            if self.engine.pdf_path:
                Clock.schedule_once(lambda dt: self.on_preview(None), 0.1)
    
        except Exception as e:
            traceback.print_exc()
            msg = str(e)
            if "certificate" in msg.lower() or "ssl" in msg.lower():
                msg += "\n\nTip: exporting/downloading the sheet as CSV/XLSX and loading the file directly is usually more reliable on Android."
            self.set_status(f"Google Sheet load error:\n{msg}")
    
    def on_generate_batch(self, instance):
        """Processes all rows in the data file and generates PDFs."""
        try:
            if not self.engine.supports_export_backend():
                self.set_status("PDF export backend is unavailable in this build.")
                return
            if self.engine.df is None or self.engine.df.empty:
                self.set_status("Load CSV/XLSX first.")
                return

            if not self.engine.pdf_path:
                self.set_status("Load PDF first.")
                return

            names = sorted(self.engine.df["_DISPLAY_NAME"].dropna().astype(str).unique())
            success = 0
            skipped = 0
            page_idx = self.current_page_idx()

            if platform == "android":
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    for patient_name in names:
                        try:
                            safe_p_name = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in patient_name).strip()
                            pdf_name = f"Filled_{safe_p_name or 'Unknown'}.pdf"
                            pdf_bytes = self.engine._build_filled_pdf_bytes(patient_name, page_idx=page_idx, export_scope="document")
                            zf.writestr(pdf_name, pdf_bytes)
                            success += 1
                        except Exception:
                            traceback.print_exc()
                            skipped += 1

                if success <= 0:
                    self.set_status(f"Batch error:\nNo PDFs were generated. Skipped: {skipped}")
                    return
                batch_name = "FormAlchemist_Batch.zip"
                return self._save_android_bytes_via_picker(
                    zip_buffer.getvalue(),
                    batch_name,
                    mime_type="application/zip",
                    request_code=42442,
                    success_message="Batch ZIP saved",
                    extra_status_lines=[f"Success: {success}", f"Skipped: {skipped}", "Scope: full document"],
                )

            out_dir = os.path.join(self.get_app_output_dir(), "batch_output")
            os.makedirs(out_dir, exist_ok=True)

            for patient_name in names:
                try:
                    safe_p_name = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in patient_name).strip()
                    out_path = os.path.join(out_dir, f"Filled_{safe_p_name or 'Unknown'}.pdf")
                    self.engine.export_filled_pdf(patient_name, out_path, page_idx=page_idx, export_scope="document")
                    success += 1
                except Exception:
                    traceback.print_exc()
                    skipped += 1

            self.set_status(f"Batch done.\nFolder: {out_dir}\nSuccess: {success} | Skipped: {skipped}")
        except Exception as e:
            traceback.print_exc()
            self.set_status("Batch Error: " + str(e))



    # ================================================================
    # Android-Optimized: Runtime permissions + First-run onboarding
    # ================================================================

    def on_start(self):
        """Called by Kivy after build() completes."""
        if platform == "android":
            Clock.schedule_once(self._request_android_permissions, 0.3)
        Clock.schedule_once(self._maybe_show_first_run_welcome, 0.8)

    def _request_android_permissions(self, *_):
        """Request READ/WRITE_EXTERNAL_STORAGE on Android 6+ (API 23+).
        SAF-based picking works without these, but they improve compatibility
        with older pre-SAF workflows.
        """
        try:
            from android.permissions import (
                request_permissions,
                Permission,
                check_permission,
            )
            needed = []
            for perm in [
                Permission.READ_EXTERNAL_STORAGE,
                Permission.WRITE_EXTERNAL_STORAGE,
            ]:
                try:
                    if not check_permission(perm):
                        needed.append(perm)
                except Exception:
                    needed.append(perm)
            if needed:
                def _on_permissions(permissions, grants):
                    denied = [p for p, g in zip(permissions, grants) if not g]
                    if denied:
                        Clock.schedule_once(lambda dt: self.set_status(
                            "Storage permission denied.\n"
                            "File picking still works via the system picker.\n"
                            "Grant permission in Settings > Apps > MediMapPro if needed.",
                            kind="warning",
                        ), 0)
                request_permissions(needed, _on_permissions)
        except Exception:
            pass  # android.permissions unavailable — SAF still works fine

    def _maybe_show_first_run_welcome(self, *_):
        """Show the onboarding popup only on the very first launch."""
        flag_path = os.path.join(
            get_form_alchemist_private_storage_root("form_alchemist"),
            "first_run_done.flag",
        )
        if os.path.exists(flag_path):
            return
        try:
            os.makedirs(os.path.dirname(flag_path), exist_ok=True)
            with open(flag_path, "w") as _fh:
                _fh.write("1")
        except Exception:
            pass
        Clock.schedule_once(lambda dt: self._show_first_run_welcome_popup(), 0)

    def _show_first_run_welcome_popup(self):
        """First-run welcome dialog explaining the 3-step workflow."""
        palette = getattr(self, "ui_palette", {}) or {}
        is_mobile = bool(getattr(self, "ui_mobile", False))

        outer = BoxLayout(
            orientation="vertical",
            spacing=dp(14),
            padding=[dp(16), dp(16), dp(16), dp(16)],
        )

        title_lbl = Label(
            text="Welcome to MediMapPro",
            color=palette.get("text", (0.97, 0.98, 1.0, 1)),
            bold=True,
            size_hint_y=None,
            height=dp(34),
            halign="left",
            valign="middle",
            font_size=dp(18) if is_mobile else dp(20),
        )
        title_lbl.bind(size=self._sync_label_text_size)
        outer.add_widget(title_lbl)

        steps_text = (
            "Here is how to get started in 3 steps:\n\n"
            "Step 1 — Load PDF\n"
            "  Tap Load PDF and select a blank PDF form from your device.\n\n"
            "Step 2 — Load Data\n"
            "  Tap Load Data File and select a CSV or Excel spreadsheet\n"
            "  (one row per patient / record), or paste a Google Sheet URL.\n\n"
            "Step 3 — Map and Export\n"
            "  Tap Find Fields to detect fill areas on the form.\n"
            "  Tap a box, choose a data column, press Save Link.\n"
            "  Then tap Export Current Record PDF to save a filled copy.\n\n"
            "Tap the ? button at the top right for this guide again at any time."
        )
        steps_lbl = Label(
            text=steps_text,
            color=palette.get("muted", (0.71, 0.79, 0.93, 1)),
            size_hint_y=None,
            height=dp(280) if is_mobile else dp(240),
            halign="left",
            valign="top",
            font_size=dp(12) if is_mobile else dp(13),
        )
        steps_lbl.bind(size=self._sync_label_text_size)
        outer.add_widget(steps_lbl)

        btn_ok = self._make_compact_action_button("Got it — Start Now", tone="primary")
        btn_ok.size_hint = (1, None)
        btn_ok.height = dp(48)
        outer.add_widget(btn_ok)

        popup, _w = self._make_styled_popup(
            "Welcome",
            outer,
            subtitle="Quick-start guide",
            size_hint=(0.94 if is_mobile else 0.62, 0.86 if is_mobile else 0.76),
        )
        btn_ok.bind(on_release=lambda *_a: popup.dismiss())
        popup.open()

    def _show_help_popup(self, *_):
        """Quick-reference help popup — accessible any time via the ? button."""
        palette = getattr(self, "ui_palette", {}) or {}
        is_mobile = bool(getattr(self, "ui_mobile", False))

        outer = BoxLayout(
            orientation="vertical",
            spacing=dp(12),
            padding=[dp(14), dp(14), dp(14), dp(14)],
        )

        help_text = (
            "MediMapPro — Quick Reference\n\n"
            "Load PDF\n"
            "  Opens your device file picker. Select a blank PDF form.\n"
            "  The form is shown in the preview panel on the right.\n\n"
            "Load Data File\n"
            "  Select a CSV or Excel spreadsheet with one row per record.\n"
            "  Or use Google Sheet to paste a public Google Sheets URL.\n\n"
            "Find Fields (Detect)\n"
            "  Scans the current page for fill areas, lines, and checkboxes.\n"
            "  Detected areas are shown as coloured overlays on the preview.\n\n"
            "Mapping\n"
            "  Tap a detected area in the preview to select it.\n"
            "  Choose the matching data column, then tap Save Link.\n"
            "  Repeat for each area you want to fill.\n\n"
            "Export\n"
            "  Export Current Record PDF — saves one filled PDF for the selected record.\n"
            "  Export All Record PDFs — saves a ZIP with one PDF per row."
        )
        help_lbl = Label(
            text=help_text,
            color=palette.get("text", (0.97, 0.98, 1.0, 1)),
            size_hint_y=None,
            height=dp(360) if is_mobile else dp(300),
            halign="left",
            valign="top",
            font_size=dp(11.5) if is_mobile else dp(12.5),
        )
        help_lbl.bind(size=self._sync_label_text_size)
        outer.add_widget(help_lbl)

        btn_close = self._make_compact_action_button("Close", tone="plain")
        btn_close.size_hint = (1, None)
        btn_close.height = dp(44)
        outer.add_widget(btn_close)

        popup, _w = self._make_styled_popup(
            "How to Use MediMapPro",
            outer,
            subtitle="Quick-start reference",
            size_hint=(0.94 if is_mobile else 0.62, 0.90 if is_mobile else 0.80),
        )
        btn_close.bind(on_release=lambda *_a: popup.dismiss())
        popup.open()

    def _show_user_error_popup(self, title, message):
        """Show a user-friendly error popup instead of a raw stack trace."""
        palette = getattr(self, "ui_palette", {}) or {}
        is_mobile = bool(getattr(self, "ui_mobile", False))

        outer = BoxLayout(
            orientation="vertical",
            spacing=dp(12),
            padding=[dp(14), dp(14), dp(14), dp(14)],
        )
        msg_lbl = Label(
            text=str(message or "An unexpected error occurred. Please try again."),
            color=palette.get("text", (0.97, 0.98, 1.0, 1)),
            size_hint_y=None,
            height=dp(180) if is_mobile else dp(130),
            halign="left",
            valign="top",
            font_size=dp(12) if is_mobile else dp(13),
        )
        msg_lbl.bind(size=self._sync_label_text_size)
        outer.add_widget(msg_lbl)

        btn_row = BoxLayout(orientation="horizontal", size_hint_y=None, height=dp(44), spacing=dp(8))
        btn_help = self._make_compact_action_button("? Help", tone="plain")
        btn_help.size_hint = (None, None)
        btn_help.width = dp(96)
        btn_help.height = dp(44)
        btn_ok = self._make_compact_action_button("OK", tone="primary")
        btn_ok.size_hint = (1, None)
        btn_ok.height = dp(44)
        btn_row.add_widget(btn_help)
        btn_row.add_widget(btn_ok)
        outer.add_widget(btn_row)

        popup, _w = self._make_styled_popup(
            title,
            outer,
            size_hint=(0.88 if is_mobile else 0.52, 0.50 if is_mobile else 0.42),
        )
        btn_ok.bind(on_release=lambda *_a: popup.dismiss())
        btn_help.bind(on_release=lambda *_a: (popup.dismiss(), self._show_help_popup()))
        popup.open()

# App entry point

if __name__ == "__main__":
    FormAlchemistApp().run()
